"""PR-concentration thesis experiment (faithful — calls real main()).

Thesis (GREENFIELD_COORDINATOR_LOCKS §Q-COORD.F): residual density
violations are bound by ProductionReport concentration of B47, NOT the
coordinator algorithm. Proof: spread B47's PR cohort across more tanks
and re-run the REAL pipeline; if B47 violations drop, thesis holds.

Method: copy Forecast.xlsm to a temp file, monkeypatch
`forecast.run.read_production_report` to spread B47's tanks, call the
real `forecast.run.main(temp)`, then read density violations from the
temp workbook's BatchLocations. The original workbook is untouched.

Run from Python/:
    python ../scripts/experiment_pr_deconcentrate.py
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # type: ignore

import forecast.run as run_mod  # type: ignore
from forecast.production_report import (  # type: ignore
    PRTankRecord,
    read_production_report as _real_read_pr,
)


def _spread_b47(og_records, facility, n_tanks):
    tank_sys = {t.tank_id: t.system_id for t in facility.tanks if t.type == "OG"}
    occupied = {r.tank_id for r in og_records}
    og12 = {"OG1N", "OG1S", "OG2N", "OG2S"}
    target = [r for r in og_records if r.batch_id == "B47"]
    others = [r for r in og_records if r.batch_id != "B47"]
    if not target:
        return list(og_records)
    total_count = sum(r.closing_count for r in target)
    total_bio = sum(r.closing_biomass_kg for r in target)
    avg_wt = (total_bio * 1000.0 / total_count) if total_count else 0.0
    own = sorted(r.tank_id for r in target)
    empty_og12 = sorted(t for t, s in tank_sys.items()
                        if s in og12 and t not in occupied)
    chosen = (own + empty_og12)[:n_tanks]
    per_c = total_count / len(chosen)
    per_b = total_bio / len(chosen)
    return others + [
        PRTankRecord(batch_id="B47", tank_id=t, closing_count=per_c,
                     closing_biomass_kg=per_b, closing_avg_wt_g=avg_wt)
        for t in chosen
    ]


def _count_violations(wb_path):
    wb = openpyxl.load_workbook(wb_path, keep_vba=True, data_only=True)
    ws = wb["BatchLocations"]
    viols = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row:
            continue
        d = row[8]
        if isinstance(d, (int, float)) and d > 95:
            viols.append((row[2], d))
    worst = max((d for _, d in viols), default=0.0)
    by_batch = Counter(b for b, _ in viols)
    return len(viols), worst, dict(by_batch)


def _run_with_spread(n_tanks, base_wb):
    """Copy workbook, patch PR reader to spread B47, run real main()."""
    tmp = base_wb.parent / f"_exp_b47_{n_tanks or 'baseline'}.xlsm"
    shutil.copy(base_wb, tmp)

    if n_tanks is None:
        patched = None
    else:
        # Need facility to map tanks; read it once via a throwaway wb.
        from forecast.excel_io import load_workbook, read_facility_config
        fac = read_facility_config(load_workbook(str(base_wb)))

        def patched(wb):
            cd, og, fw = _real_read_pr(wb)
            return cd, _spread_b47(og, fac, n_tanks), fw

    orig = run_mod.read_production_report
    try:
        if patched is not None:
            run_mod.read_production_report = patched
        run_mod.main(str(tmp))
    finally:
        run_mod.read_production_report = orig

    stats = _count_violations(tmp)
    tmp.unlink(missing_ok=True)
    return stats


def main() -> int:
    base_wb = ROOT / "Forecast.xlsm"
    print("=" * 72)
    print("PR-CONCENTRATION THESIS EXPERIMENT (faithful — real pipeline)")
    print("=" * 72)
    print(f"{'scenario':<22}{'violations':>11}{'worst':>9}{'B47':>6}{'B46':>6}")
    for label, n in [("baseline (as-is)", None), ("B47 -> 4 tanks", 4),
                     ("B47 -> 6 tanks", 6), ("B47 -> 8 tanks", 8)]:
        n_v, worst, bb = _run_with_spread(n, base_wb)
        print(f"{label:<22}{n_v:>11}{worst:>9.1f}"
              f"{bb.get('B47', 0):>6}{bb.get('B46', 0):>6}")
    print()
    print("Thesis holds if B47 violations fall as B47 spreads across more tanks.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
