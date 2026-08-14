"""Runner for the L1<->L3 feasibility loop (tankless GLOBAL planner POC).

METHOD: GLOBAL LOOP (L1 loading <-> L3 tank realizability)

Loads the repo config/ + scenario/, hydrates in-flight OG state from the
ProductionReport in Forecast.xlsm, then runs the L1<->L3 feasibility loop
(forecast.global_planner_loop_poc.run_loop): iterate L1 (per-week biomass
ceiling) -> L3 (lexicographic whole-tank placement) -> lower the ceiling of any
week whose whole-tank demand over-subscribes the mode-aware available OG tanks,
until fully tank-realizable or it converges.

Then runs the PRODUCTION pipeline (forecast.run.main) on the SAME inputs into a
TEMP workbook and reads back a head-to-head: total HOG, peak facility biomass,
per-system biomass/feed over-cap counts, transfers.

Usage:
    python -m tools.run_loop_poc
    python -m tools.run_loop_poc --max-iterations 12 --no-controller
    python -m tools.run_loop_poc --no-pr      # skip PR; incoming-only

Creates ONLY this runner + forecast/global_planner_loop_poc.py. Touches no
production file; not imported by the pipeline.
"""
from __future__ import annotations

import argparse
import shutil
import sys
import tempfile
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.config_io import load_config
from forecast.scenario_io import load_batches, load_limits
from forecast import global_planner_l2_poc as l2
from forecast import global_planner_l3_poc as l3
from forecast import global_planner_loop_poc as loop
from tools.run_global_poc import _hydrate_inflight_og
from tools.run_full_facility_poc import _hydrate_pr


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--scenario-dir", default=str(_ROOT / "scenario"))
    ap.add_argument("--workbook", default=str(_ROOT / "Forecast.xlsm"))
    ap.add_argument("--harvest-tank-density-pct", type=float, default=1.25)
    ap.add_argument("--max-iterations", type=int, default=10)
    ap.add_argument("--margin-frac", type=float, default=0.5)
    ap.add_argument("--slack-epsilon", type=float, default=1000.0)
    ap.add_argument("--mip-time-limit", type=float, default=180.0)
    ap.add_argument("--mip-rel-gap", type=float, default=0.01)
    ap.add_argument("--no-pr", action="store_true")
    ap.add_argument("--no-controller", action="store_true",
                    help="skip the production-pipeline head-to-head")
    ap.add_argument("--no-purge-hold", action="store_true",
                    help="model harvest as INSTANT removal (the old, mis-modeled "
                         "POC); default models the 6N off-feed purge-hold flow")
    args = ap.parse_args()
    model_purge_hold = not args.no_purge_hold

    print("=" * 72)
    print("  METHOD: GLOBAL LOOP (L1 loading <-> L3 tank realizability)")
    print("=" * 72)

    control, tables, facility = load_config(args.config_dir)
    batches = load_batches(args.scenario_dir)
    _facility_limits, system_limits = load_limits(args.scenario_dir, control)
    print(f"  Config:   {args.config_dir}")
    print(f"  Scenario: {len(batches)} batches from {args.scenario_dir}")

    inflight = {}
    fw_inflight = {}
    purge_inflight = {}
    if not args.no_pr:
        # Hydrate BOTH OG and FW in-flight units: model_full_facility (now the
        # default) needs fw_inflight or it under-counts the FW phase.
        inflight, fw_inflight, derived_start, purge_inflight = _hydrate_pr(
            Path(args.workbook), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
            print(f"  ForecastStart: derived {derived_start.date()} from PR closing +1d")
        print(f"  In-flight OG batches from PR: {len(inflight)}; "
              f"FW-in-flight batches: {len(fw_inflight)}")

    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    print(f"  forecast_start={fs_date}, horizon={control.horizon_weeks}w")
    print(f"  facility biomass cap: {control.max_biomass_kg:,.0f} kg")

    # Show the mode-aware available-tank counts (purge vs production).
    n_by_sys = l3.n_tanks_per_system(facility)
    prod_sys = [s for s in l2.og_systems_from_facility(facility)
                if s in (set(l2.NURSERY_SYSTEMS) | set(l2.GROWOUT_SYSTEMS))]
    sixn = sum(n_by_sys.get(s, 0) for s in l2.PURGE_SYSTEMS)
    feeding = sum(n_by_sys.get(s, 0) for s in prod_sys)
    psd = getattr(control, "sixn_production_start", None)
    print(f"  OG tank inventory: {feeding} feeding (11 NURSERY+GROWOUT systems) "
          f"+ {sixn} 6N -> {feeding + sixn} total")
    print(f"  sixn_production_start={psd.date() if hasattr(psd,'date') else psd}")
    print(f"  available BIOMASS tanks: purge={feeding + sixn}, "
          f"production={feeding + sixn}; FEED tanks: purge={feeding}, "
          f"production={feeding + sixn}")

    # ---- The loop.
    print("\n  RUNNING L1<->L3 FEASIBILITY LOOP")
    l3_kwargs = dict(
        slack_epsilon=args.slack_epsilon,
        mip_time_limit=args.mip_time_limit,
        mip_rel_gap=args.mip_rel_gap,
        verbose=False,
    )
    print(f"  6N flow model: "
          f"{'PURGE-HOLD (2-wk off-feed depuration through 6N pairs)' if model_purge_hold else 'INSTANT removal (mis-modeled; --no-purge-hold)'}")
    result = loop.run_loop(
        batches, tables, control, facility, system_limits,
        inflight_og=inflight,
        harvest_tank_density_pct=args.harvest_tank_density_pct,
        max_iterations=args.max_iterations,
        margin_frac=args.margin_frac,
        l3_kwargs=l3_kwargs,
        model_purge_hold=model_purge_hold,
        # The whole-facility model is the correct default; tie it to the same
        # toggle so --no-purge-hold recovers the OG-only instant-removal compare.
        model_full_facility=model_purge_hold,
        fw_inflight=fw_inflight,
        purge_inflight=purge_inflight,
        verbose=True,
    )

    print("\n  LOOP CONVERGENCE")
    print(f"    iterations run: {result.n_iterations}")
    print(f"    converged to FULLY tank-realizable (0 weeks over-sub): "
          f"{result.converged}")
    if not result.converged:
        print(f"    residual: {result.residual_over_weeks} weeks over-subscribed "
              f"({result.residual_total_tanks_over} tanks over)")
        for o in result.final_l3 and result.iterations[-1].over_weeks[:10] or []:
            print(f"      wk {o.week:>2} {o.week_label}: demand {o.tank_demand} "
                  f"vs {o.avail_biomass_tanks} tanks ({o.tanks_over} over), "
                  f"standing {o.standing_kg:,.0f} kg")

    print("\n  CONVERGED REALIZABLE ENVELOPE")
    print(f"    peak facility biomass: {result.peak_biomass_kg:,.0f} kg "
          f"({result.pct_of_cap_peak:.1f}% of {result.facility_cap_kg:,.0f} cap)")
    print(f"    mean facility biomass: {result.mean_biomass_kg:,.0f} kg "
          f"({result.pct_of_cap_mean:.1f}% of cap)")
    print(f"    total HOG harvested:   {result.total_hog_kg:,.0f} kg "
          f"(L1 round/live envelope)")
    fl3 = result.final_l3
    print(f"    L3 over-cap: {fl3.over_biomass_system_weeks} biomass / "
          f"{fl3.over_feed_system_weeks} feed system-weeks")
    print(f"    L3 transfers: {fl3.realized_transfers:.0f}; "
          f"worst bio/feed ratio {fl3.worst_biomass_ratio:.2f}x/"
          f"{fl3.worst_feed_ratio:.2f}x")

    # 6N purge-hold accounting (only when modeled).
    if model_purge_hold and result.final_l1.purge_trace:
        pt = result.final_l1.purge_trace
        peak_held = max((r.held_biomass_kg for r in pt), default=0.0)
        peak_tanks = max((r.sixn_tanks_used for r in pt), default=0)
        n6n = sum(l3.n_tanks_per_system(facility).get(s, 0)
                  for s in l2.PURGE_SYSTEMS)
        fin = result.iterations[-1]
        wks_held = sum(1 for r in pt if r.held_biomass_kg > 1e-6)
        modes = {}
        for r in pt:
            modes[r.mode] = modes.get(r.mode, 0) + 1
        print(f"    6N purge hold: peak {peak_held:,.0f} kg in {peak_tanks}/{n6n} "
              f"6N tanks; held some weeks={wks_held}/{len(pt)}; "
              f"6N over-subscribed weeks={fin.n_sixn_over_weeks}")
        print(f"    6N weeks by mode: {modes}; peak facility feed "
              f"{fin.peak_feed_kg_day:,.0f} kg/day "
              f"({100*fin.peak_feed_kg_day/control.max_feed_per_day_kg:.1f}% cap)")

    # ---- Conservation (per batch, from final L1).
    print("\n  CONSERVATION (final L1; per-batch seeded ~= harv+standing+mort+cull)")
    worst = max((abs(c["residual_pct"]) for c in result.final_l1.conservation.values()),
                default=0.0)
    print(f"    worst |residual| across batches: {worst:.4f}% "
          f"({'OK conserves' if worst < 0.01 else 'CHECK'})")

    # ---- Head-to-head vs the production controller.
    if not args.no_controller:
        ctrl = _run_controller(args)
        if ctrl is not None:
            _print_headtohead(result, ctrl)

    return 0


def _run_controller(args) -> dict | None:
    """Run forecast.run.main on the SAME inputs into a TEMP workbook, read back
    HOG / peak biomass / over-cap counts / transfers. Returns a metrics dict."""
    print("\n  RUNNING PRODUCTION CONTROLLER (forecast.run.main) for head-to-head")
    src = Path(args.workbook)
    if not src.exists():
        print(f"    (workbook {src} not found; skipping controller comparison)")
        return None
    try:
        from forecast.run import main as run_main
        from openpyxl import load_workbook as _load_xl
    except Exception as e:  # noqa: BLE001
        print(f"    (cannot import pipeline/openpyxl: {e}); skipping")
        return None

    tmpdir = Path(tempfile.mkdtemp(prefix="loop_poc_ctrl_"))
    try:
        in_copy = tmpdir / src.name
        shutil.copy2(src, in_copy)
        out_path = tmpdir / "controller_out.xlsx"
        run_main(input_path=str(in_copy), output_path=str(out_path),
                 config_dir=args.config_dir, scenario_dir=args.scenario_dir)
        wb = _load_xl(out_path, data_only=True)
        metrics = _read_controller_metrics(wb)
        wb.close()
        return metrics
    except Exception as e:  # noqa: BLE001
        import traceback
        print(f"    (controller run failed: {e})")
        traceback.print_exc()
        return None
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _read_controller_metrics(wb) -> dict:
    """Extract head-to-head metrics from the controller's output workbook."""
    m = {"hog_kg": None, "gross_kg": None, "peak_biomass_kg": None,
         "over_biomass_sw": None, "over_feed_sw": None, "transfers": None}

    # ---- YearlySummary: HOG (t) col3, Gross (t) col4, Peak (t) col7. Header
    # row 3, data from row 4. Sum HOG+Gross across years; peak = max.
    if "YearlySummary" in wb.sheetnames:
        ws = wb["YearlySummary"]
        hdr = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and row[0] == "Year":
                hdr = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
                hdr_row = ri
                break
        if hdr is not None:
            ci_hog = hdr.get("Harvest_HOG (t)")
            ci_gross = hdr.get("Harvest_Gross (t)")
            ci_peak = hdr.get("Peak_Biomass (t)")
            hog_t = gross_t = 0.0
            peak_t = 0.0
            for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
                if not row or row[0] is None:
                    continue
                if ci_hog is not None and row[ci_hog] is not None:
                    hog_t += float(row[ci_hog])
                if ci_gross is not None and row[ci_gross] is not None:
                    gross_t += float(row[ci_gross])
                if ci_peak is not None and row[ci_peak] is not None:
                    peak_t = max(peak_t, float(row[ci_peak]))
            m["hog_kg"] = hog_t * 1000.0
            m["gross_kg"] = gross_t * 1000.0
            m["peak_biomass_kg"] = peak_t * 1000.0

    # ---- SystemLimitsAudit: count non-empty Bio_flag / Feed_flag (cols 5,8).
    if "SystemLimitsAudit" in wb.sheetnames:
        ws = wb["SystemLimitsAudit"]
        nb = nf = 0
        seen_hdr = False
        for row in ws.iter_rows(values_only=True):
            if not seen_hdr:
                if row and row[0] == "Week":
                    seen_hdr = True
                continue
            if not row or row[0] is None:
                continue
            if len(row) > 4 and row[4]:
                nb += 1
            if len(row) > 7 and row[7]:
                nf += 1
        m["over_biomass_sw"] = nb
        m["over_feed_sw"] = nf

    # ---- TransferPlan: count data rows of Type == "Transfer" (intra-SW moves)
    # and the broader Transfer+TranOG total.
    if "TransferPlan" in wb.sheetnames:
        ws = wb["TransferPlan"]
        seen_hdr = False
        n_transfer = n_all = 0
        for row in ws.iter_rows(values_only=True):
            if not seen_hdr:
                if row and row[0] == "Week":
                    seen_hdr = True
                continue
            if not row or row[0] is None:
                continue
            typ = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            n_all += 1
            if typ == "Transfer":
                n_transfer += 1
        m["transfers"] = n_transfer
        m["transfer_rows_all"] = n_all

    return m


def _print_headtohead(result: "loop.LoopResult", ctrl: dict) -> None:
    fl3 = result.final_l3
    print("\n" + "=" * 72)
    print("  HEAD-TO-HEAD: GLOBAL LOOP vs PRODUCTION CONTROLLER")
    print("=" * 72)

    def _fmt(v, kind="kg"):
        if v is None:
            return "n/a"
        if kind == "kg":
            return f"{v:,.0f}"
        return f"{v}"

    rows = [
        ("Total HOG harvested (kg)",
         _fmt(result.total_hog_kg), _fmt(ctrl.get("hog_kg")),
         "L1 envelope is round/live; controller is HOG-yield"),
        ("Total round/live harvested (kg)",
         _fmt(result.total_hog_kg), _fmt(ctrl.get("gross_kg")),
         "compare round-to-round"),
        ("Peak facility biomass (kg)",
         _fmt(result.peak_biomass_kg), _fmt(ctrl.get("peak_biomass_kg")), ""),
        ("Peak % of facility cap",
         f"{result.pct_of_cap_peak:.1f}%",
         (f"{100*ctrl['peak_biomass_kg']/result.facility_cap_kg:.1f}%"
          if ctrl.get("peak_biomass_kg") else "n/a"), ""),
        ("Over-cap BIOMASS system-weeks",
         _fmt(fl3.over_biomass_system_weeks, "n"),
         _fmt(ctrl.get("over_biomass_sw"), "n"), ""),
        ("Over-cap FEED system-weeks",
         _fmt(fl3.over_feed_system_weeks, "n"),
         _fmt(ctrl.get("over_feed_sw"), "n"), ""),
        ("Transfers",
         _fmt(fl3.realized_transfers, "n"),
         _fmt(ctrl.get("transfers"), "n"),
         "L3 = tanks-worth moved INTO new system; controller = TransferPlan "
         "'Transfer' rows (different units)"),
    ]
    print(f"  {'metric':<34} {'GLOBAL LOOP':>16} {'CONTROLLER':>16}")
    print(f"  {'-'*34} {'-'*16} {'-'*16}")
    for name, g, c, note in rows:
        print(f"  {name:<34} {g:>16} {c:>16}")
        if note:
            print(f"      ({note})")


if __name__ == "__main__":
    raise SystemExit(main())
