"""Put a CONFIDENCE BAND on a forecast, from measured history.

    python tools/forecast_bands.py --forecast out.xlsm \
                                   --model backtest_fixed/error_model.json

A production plan is not really asked "what is the number", it is asked "how
much should I trust it". This answers that from the operator's OWN history:
21 monthly Production Reports, each replayed through the model and graded
against what actually happened (tools/backtest.py -> tools/error_model.py).

WHY THIS IS ANALYTIC AND NOT MONTE CARLO
----------------------------------------
The obvious build is to perturb growth, re-run the forecast a few hundred
times and take percentiles. That would be WRONG here, and expensively so: this
planner was measured swinging total HOG by 1.5% and its density-breach count by
35% on a 0.01% input change (a handling-mortality rate). Sampling through it
would mostly measure PLANNER INSTABILITY and report it as biological
uncertainty -- an authoritative-looking band around the wrong quantity.

So the measured distribution is propagated directly onto the plan instead. The
band answers one clean question: given that the model's average-weight error at
this horizon has historically fallen in this range, where does the tonnage land?

THE INVERSION, because it is easy to get backwards
--------------------------------------------------
The error is defined as (predicted - actual) / actual. So actual = predicted /
(1 + err). A HIGH error means the model ran HOT, which means the ACTUAL comes
in LOW. The p90 error therefore produces the LOW edge of the tonnage band and
p10 produces the HIGH edge -- the bounds cross over.

WHAT THE BAND COVERS, AND WHAT IT DOES NOT
------------------------------------------
COVERS: biological uncertainty in average weight -- the thing the backtest can
actually measure, and the only part of the forecast graded against reality.

DOES NOT COVER: execution. The harvest COUNT is a plan, not a prediction; if
the facility harvests a different number of fish than planned, tonnage moves
for reasons no growth model can foresee. The band is therefore conditional:
"if the harvest count plan executes, tonnage lands here". Stated on every row
rather than in a footnote, because a band read as covering everything is worse
than no band.

HORIZON LIMIT: only horizons the error model rates as sound are quoted. The
operator's ruling (2026-08-21) is that 1-3 months are trustworthy; beyond that
residual harvest-execution effects still contaminate the measurement, so those
months are printed as INDICATIVE and never as a band.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:                 # so it runs from any cwd
    sys.path.insert(0, str(_ROOT))

import openpyxl                                                  # noqa: E402

# Band arithmetic lives in forecast/error_bands.py — the SAME module the app
# renders from, so a number on screen and a number here cannot disagree. This
# file is the CLI around it, not a second implementation.
from forecast.error_bands import (QUOTABLE_MAX_MONTHS, apply_band,   # noqa: E402
                                  band_for_horizon, describe)


def _months_between(a: dt.date, b: dt.date) -> int:
    return (b.year - a.year) * 12 + (b.month - a.month)


def _week_to_date(label: str):
    """ISO week label ('2026-W23') -> that week's Monday."""
    try:
        y, w = str(label).split("-W")
        return dt.date.fromisocalendar(int(y), int(w), 1)
    except Exception:
        return None


def read_harvest_months(path: Path):
    """(month_date, count, hog_kg) per month, aggregated from HarvestPlan.

    NOT the 'Harvest Model' sheet, which looks like the obvious source and is
    not: it is a two-row summary (one row, whole horizon), so banding it
    produces exactly one line. HarvestPlan carries the per-(week, batch, tank)
    harvest rows the plan actually contains, which is what a month's tonnage is
    the sum of.

    A month is attributed by the harvest WEEK's Monday. That is coarser than
    the engine's own day-accurate month split (excel_io uses working days for a
    week straddling a boundary), but a band whose edges are tens of percent
    wide is not moved by a few days of attribution.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "HarvestPlan" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"{path} has no 'HarvestPlan' sheet to band.")
    agg: dict[dt.date, list] = {}
    for i, r in enumerate(wb["HarvestPlan"].iter_rows(max_col=9,
                                                      values_only=True), 1):
        if i <= 4:
            continue
        wl, _b, _t, cnt, _gw, _gb, _y, _hw, hog = (list(r) + [None] * 9)[:9]
        if not isinstance(cnt, (int, float)) or cnt <= 0:
            continue
        d = _week_to_date(wl)
        if d is None:
            continue
        key = dt.date(d.year, d.month, 1)
        a = agg.setdefault(key, [0.0, 0.0])
        a[0] += float(cnt)
        a[1] += float(hog) if isinstance(hog, (int, float)) else 0.0
    wb.close()
    return sorted((m, v[0], v[1]) for m, v in agg.items())


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--forecast", required=True)
    ap.add_argument("--model", required=True)
    ap.add_argument("--max-horizon", type=int, default=QUOTABLE_MAX_MONTHS)
    args = ap.parse_args()

    model = json.loads(Path(args.model).read_text(encoding="utf-8"))
    months = read_harvest_months(Path(args.forecast))
    if not months:
        print("no harvest months in that workbook")
        return 1

    start = months[0][0]
    print(f"forecast starts {start}. " + describe(model) + "\n")
    print(f"{'month':>10}{'horizon':>9}{'plan HOG t':>12}"
          f"{'low t':>10}{'high t':>10}{'band':>18}   basis")
    tot_p = tot_lo = tot_hi = 0.0
    for m, _cnt, hog in months:
        h = _months_between(start, m)
        t = hog / 1000.0
        b = band_for_horizon(model, h, args.max_horizon)
        if b is None:
            why = ("in-month" if h < 1 else
                   f"beyond {args.max_horizon}m — INDICATIVE, not a band")
            print(f"{m:%Y-%m}{h:>9}{t:>12,.0f}{'—':>10}{'—':>10}{'—':>18}   {why}")
            continue
        med, p10, p90, n = (b["median_pct"], b["p10_pct"], b["p90_pct"], b["n"])
        _ap = apply_band(t, b)      # the inversion lives in error_bands
        if _ap is None:
            continue
        lo, hi = _ap
        tot_p += t
        tot_lo += lo
        tot_hi += hi
        print(f"{m:%Y-%m}{h:>9}{t:>12,.0f}{lo:>10,.0f}{hi:>10,.0f}"
              f"{f'{-p90:+.1f}% / {-p10:+.1f}%':>18}   "
              f"n={n}, median {med:+.1f}%")
    if tot_p:
        print(f"\n{'TOTAL over the quotable window':>40}: "
              f"{tot_p:,.0f} t  (band {tot_lo:,.0f} – {tot_hi:,.0f} t)")
    print("\nBand covers BIOLOGY (average weight) only, and is conditional on "
          "the harvest COUNT plan executing.\nA different number of fish "
          "harvested moves tonnage for reasons no growth model can foresee.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
