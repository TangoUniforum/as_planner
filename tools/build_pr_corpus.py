"""Normalise the historical monthly Production Reports into a parseable corpus.

    python tools/build_pr_corpus.py --out ../pr_corpus

WHY
---
The operator's monthly Production Report workbooks carry the per-tank closing
state on a sheet called **"Month Batch Summery"**. The pipeline's reader
(forecast/production_report.py) opens `wb["ProductionReport"]` by name. The
LAYOUT is already identical — closing banner in col 1, `Site:` col 2,
`Fish group name:` col 3, `Unit:` col 4, closing count / biomass / avg weight
in cols 7 / 9 / 11 — so the only thing standing between two years of history
and a usable corpus is the sheet name.

This copies that one sheet, verbatim and values-only, into a small workbook
whose sheet IS called ProductionReport. Nothing is recomputed, reordered or
"cleaned": a corpus entry must be the operator's own numbers, or an error
measured against it means nothing.

Values-only matters: the reader opens workbooks `data_only=True`, so a formula
with no cached value reads as None. Copying the evaluated values makes the
corpus independent of whether Excel last recalculated the source.

The output is ~6 kB per month, so two years of history is committable and
survives a clean clone — unlike the 3.2 MB source workbooks.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

# The operator's name for the per-tank sheet. It has drifted over the years,
# so match on a normalised form rather than an exact string:
#   2025-2026  "Month Batch Summery"
#   2024       "Month Summary Batch"   (same words, different order)
SRC_SHEET = "Month Batch Summery"
_SRC_TOKENS = frozenset(("month", "batch", "summery", "summary"))
DST_SHEET = "ProductionReport"        # the name the reader requires


def _find_source_sheet(wb) -> str | None:
    """The per-tank sheet, under whichever name this vintage used.

    Requires 'month' + 'batch' + a summary word, so it cannot accidentally
    match the batch-less 'Month Summary' rollup (which carries no Unit rows
    and would hydrate an empty facility).
    """
    for name in wb.sheetnames:
        toks = {t for t in re.split(r"\W+", name.lower()) if t}
        if "month" in toks and "batch" in toks and toks & {"summery", "summary"}:
            return name
    return None

# Month folder -> the two-digit year is in the filename, e.g. "Jan'25".
_MONTH_ORDER = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def _closing_date(rows) -> str | None:
    """The `Closing Month: m/d/yyyy` banner, from ANY column.

    The reader requires it in column 1, but the 2024/early-2025 vintage nests
    it under Site (Site in col 1, Closing Month in col 2). `_normalise_rows`
    moves it; this just finds it.
    """
    for r in rows[:8]:
        for v in r:
            if isinstance(v, str) and "Closing Month" in v:
                m = re.search(r"(\d+/\d+/\d+)", v)
                if m:
                    return m.group(1)
    return None


def _normalise_rows(rows) -> tuple[list, bool]:
    """Put the `Closing Month:` banner in column 1, where the reader looks.

    Two hierarchy vintages exist. The current one is date-outermost:
        col1 Closing Month | col2 Site | col3 Fish group name | col4 Unit
    The 2024/early-2025 one is site-outermost, shifting the banner to col 2:
        col1 Site | col2 Closing Month | col3 Fish group name | col4 Unit
    Only the banner needs moving: production_report.py never reads the Site
    column at all, and the batch and unit columns are already in position.
    """
    shifted = False
    out = []
    for r in rows:
        r = list(r)
        c1 = r[0] if r else None
        if not (isinstance(c1, str) and "Closing Month" in c1):
            for j, v in enumerate(r[1:4], start=1):
                if isinstance(v, str) and "Closing Month" in v:
                    r[0], r[j] = v, None
                    shifted = True
                    break
        out.append(r)
    return out, shifted


def _sort_key(p: Path):
    """Chronological by the year folder and the month folder name."""
    parts = [q.name.lower() for q in p.parents]
    year = next((int(q) for q in parts if q.isdigit() and len(q) == 4), 0)
    mon = 0
    for q in parts:
        for k, v in _MONTH_ORDER.items():
            if q.startswith(k):
                mon = v
                break
        if mon:
            break
    return (year, mon, p.name)


def convert(src: Path, out_dir: Path) -> tuple[bool, str]:
    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except Exception as e:
        return False, f"load failed: {type(e).__name__}: {e}"
    src_name = _find_source_sheet(wb)
    if src_name is None:
        names = wb.sheetnames[:6]
        wb.close()
        return False, f"no per-tank batch sheet (has {names})"

    rows = [list(r) for r in wb[src_name].iter_rows(values_only=True)]
    wb.close()
    rows, shifted = _normalise_rows(rows)
    closing = _closing_date(rows)
    if closing is None:
        return False, "no parseable 'Closing Month: m/d/yyyy' banner"

    # The sheet NAME is not proof of content. Some 2024 vintages carry a
    # 'Month Summary Batch' sheet that is batch rollups only, with no `Unit:`
    # rows at all — it converts cleanly, parses cleanly, and hydrates an EMPTY
    # facility. A corpus entry that silently yields zero tanks is worse than a
    # missing one, so require the per-tank rows that carry tank continuity.
    units = sum(1 for r in rows for v in r
                if isinstance(v, str) and v.strip().startswith("Unit"))
    if units == 0:
        return False, (f"sheet {src_name!r} has NO 'Unit:' rows — batch rollups "
                       f"only, would hydrate an empty facility")

    out = openpyxl.Workbook()
    dst = out.active
    dst.title = DST_SHEET
    for r in rows:
        dst.append(list(r))

    mm, dd, yyyy = closing.split("/")
    name = f"{int(yyyy):04d}-{int(mm):02d}-{int(dd):02d}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    out.save(out_dir / name)
    tag = " [col-shifted]" if shifted else ""
    return True, f"{name}  ({len(rows)} rows, closing {closing}, sheet={src_name!r}){tag}"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--reports-root", default=r"C:\Users\julian.f\OneDrive - Atlantic Sapphire\Production\Reports")
    ap.add_argument("--out", default="pr_corpus")
    args = ap.parse_args()

    root = Path(args.reports_root)
    out_dir = Path(args.out)
    srcs = sorted(root.glob("*/*/*Production Report.xlsx"), key=_sort_key)
    if not srcs:
        print(f"no 'Production Report.xlsx' files under {root}")
        return 1

    ok = fail = 0
    for p in srcs:
        good, msg = convert(p, out_dir)
        tag = "OK  " if good else "FAIL"
        if good:
            ok += 1
        else:
            fail += 1
        print(f"  {tag} {p.parent.parent.name}/{p.parent.name:<6} {msg}")
    print(f"\n{ok} converted, {fail} failed -> {out_dir}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
