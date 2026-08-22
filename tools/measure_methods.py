"""RE-MEASURE the method blurbs' harvest claims, across the whole PR corpus.

    python tools/measure_methods.py --corpus pr_corpus \
        --registries pr_corpus/registries --methods controller,controller-hybrid

The blurbs in forecast/methods.py quote figures measured on 2026-08-03 across
"6 real July-2026 PRs" -- the numbers that tell an operator which method to
pick. Four changes since (handling mortality per deposit, grade_efficiency,
purge move-in Thursday, 6N one-batch-one-tank) all move the weekly harvest
series those figures describe, and the corpus is now 21 months rather than 6.
A stale number that decides a method choice is worse than no number.

WHAT IS MEASURED, and why these three
-------------------------------------
    zero-harvest weeks   weeks the plan harvests NOTHING. The steady-harvest
                         contract's hard failure: a packing line with no fish.
    weeks under floor    weeks below min_harvest_per_week (per-week overrides
                         from FacilityLimits honoured). Softer, but it is the
                         sales contract.
    worst week           the smallest NON-EMPTY harvest. "0 -> 16,148 fish" in
                         the current blurb is this metric.

Each PR runs through methods.run_method, so every method gets its own overrides
applied in an isolated temp config -- identical to how the Compare board runs
them. Same PR, same era registry; only the method differs.

READ THE PER-PR SPREAD, not just the mean. These plans are chaos-sensitive (a
0.01% input change was measured moving HOG 1.5%), so a mean across PRs hides
how unstable an arm is. Both are printed.
"""
from __future__ import annotations

import argparse
import collections
import contextlib
import datetime as dt
import io
import statistics
import sys
import tempfile
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl                                                  # noqa: E402
import yaml                                                      # noqa: E402

from forecast.methods import REGISTRY, run_method                # noqa: E402


def corpus_months(corpus: Path):
    out = []
    for p in sorted(corpus.glob("*.xlsx")):
        try:
            y, m, d = p.stem.split("-")
            out.append((dt.date(int(y), int(m), int(d)), p))
        except ValueError:
            continue
    return out


def floor_lookup(config_dir: Path, scenario_dir: Path):
    """(default_floor, {week_label: floor}) from control + FacilityLimits."""
    ctl = yaml.safe_load((config_dir / "control.yaml").read_text(encoding="utf-8"))
    default = float(ctl.get("min_harvest_per_week") or 0.0)
    ov = {}
    f = scenario_dir / "limits.yaml"
    if f.exists():
        def walk(o):
            if isinstance(o, dict):
                if o.get("metric") == "min_harvest_per_week":
                    w = o.get("week") or o.get("week_label")
                    v = o.get("value")
                    if w is not None and isinstance(v, (int, float)):
                        ov[str(w)] = float(v)
                for v in o.values():
                    walk(v)
            elif isinstance(o, list):
                for v in o:
                    walk(v)
        walk(yaml.safe_load(f.read_text(encoding="utf-8")))
    return default, ov


def metrics(path: Path, default_floor: float, ov: dict):
    """zero-harvest weeks, weeks under floor, worst non-empty week.

    Weeks come from BatchLocations as well as HarvestPlan: a week the plan
    harvests nothing has no HarvestPlan row at all, and counting only the rows
    that exist would report zero empty weeks for every method.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hv = collections.defaultdict(float)
    weeks = set()
    if "BatchLocations" in wb.sheetnames:
        for i, r in enumerate(wb["BatchLocations"].iter_rows(max_col=2,
                                                             values_only=True), 1):
            if i > 4 and r and r[0]:
                weeks.add(str(r[0]))
    if "HarvestPlan" in wb.sheetnames:
        for i, r in enumerate(wb["HarvestPlan"].iter_rows(max_col=4,
                                                          values_only=True), 1):
            if i <= 4:
                continue
            row = list(r) + [None] * 4
            wl, cnt = row[0], row[3]
            if isinstance(cnt, (int, float)) and cnt > 0 and wl:
                hv[str(wl)] += float(cnt)
                weeks.add(str(wl))
    wb.close()
    if not weeks:
        return None
    zero = sum(1 for w in weeks if hv.get(w, 0.0) <= 0.0)
    under = sum(1 for w in weeks if hv.get(w, 0.0) < ov.get(w, default_floor))
    nonzero = [v for v in hv.values() if v > 0]
    return {"weeks": len(weeks), "zero": zero, "under": under,
            "worst": min(nonzero) if nonzero else 0.0}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--corpus", required=True)
    ap.add_argument("--registries", required=True)
    ap.add_argument("--methods", default="controller,controller-hybrid")
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--limit", type=int, default=0, help="first N PRs only")
    ap.add_argument("--force", default="",
                    help="extra control overrides applied to EVERY method, as "
                         "k=v,k=v. Use to measure a feature that the shipped "
                         "config disables: controller-hybrid overrides "
                         "hybrid_follow but NOT hybrid_purge_lever / "
                         "hybrid_production_lever, and control.yaml ships both "
                         "false, so the guide is built and then steers nothing "
                         "-- the arm is byte-identical to the plain controller. "
                         "e.g. --force hybrid_production_lever=true")
    args = ap.parse_args()

    forced = {}
    for part in (args.force or "").split(","):
        if "=" in part:
            k, v = part.split("=", 1)
            v = v.strip()
            forced[k.strip()] = (True if v.lower() == "true" else
                                 False if v.lower() == "false" else
                                 float(v) if v.replace(".", "", 1).isdigit() else v)
    if forced:
        print("FORCED overrides on every method: " + str(forced))
    keys = [k.strip() for k in args.methods.split(",") if k.strip()]
    for k in keys:
        if k not in REGISTRY:
            print("unknown method " + repr(k) + "; have " + str(sorted(REGISTRY)))
            return 1

    months = corpus_months(Path(args.corpus))
    if args.limit:
        months = months[:args.limit]
    per = {k: [] for k in keys}
    print(str(len(months)) + " PRs x " + str(len(keys)) + " methods\n")
    print(f"{'PR':>12}{'method':>20}{'weeks':>7}{'zero':>6}{'under':>7}{'worst':>10}")
    for closing, pr in months:
        scn = Path(args.registries) / closing.isoformat()
        if not scn.is_dir():
            continue
        default_floor, ov = floor_lookup(Path(args.config_dir), scn)
        for k in keys:
            m = None
            with tempfile.TemporaryDirectory() as td:
                out = Path(td) / (k + ".xlsx")
                buf = io.StringIO()
                try:
                    _m = REGISTRY[k]
                    if forced:
                        import dataclasses as _dc
                        _m = _dc.replace(_m, overrides={**(_m.overrides or {}),
                                                        **forced})
                    with contextlib.redirect_stdout(buf):
                        run_method(_m, str(pr), str(out),
                                   args.config_dir, str(scn))
                except Exception as e:                       # noqa: BLE001
                    print(f"{str(closing):>12}{k:>20}   FAILED "
                          + type(e).__name__ + ": " + str(e)[:60])
                    continue
                prod = out if out.exists() else out.with_suffix(".xlsm")
                if not prod.exists():
                    print(f"{str(closing):>12}{k:>20}   no workbook produced")
                    continue
                m = metrics(prod, default_floor, ov)
            if not m:
                continue
            per[k].append(m)
            print(f"{str(closing):>12}{k:>20}{m['weeks']:>7}{m['zero']:>6}"
                  f"{m['under']:>7}{m['worst']:>10,.0f}")

    print("")
    print(f"{'method':>20}{'PRs':>5}{'zero wks/PR':>13}{'PRs w/ a zero wk':>18}"
          f"{'under/PR':>10}{'worst wk':>11}")
    for k in keys:
        v = per[k]
        if not v:
            print(f"{k:>20}    no successful runs")
            continue
        anyzero = sum(1 for m in v if m["zero"] > 0)
        share = str(anyzero) + "/" + str(len(v))
        print(f"{k:>20}{len(v):>5}"
              f"{statistics.mean(m['zero'] for m in v):>13.1f}"
              f"{share:>18}"
              f"{statistics.mean(m['under'] for m in v):>10.1f}"
              f"{statistics.median(m['worst'] for m in v):>11,.0f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
