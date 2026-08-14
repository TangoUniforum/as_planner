"""Runner for the tankless GLOBAL placement-configuration LP (Layer 3) POC.

METHOD: GLOBAL L3 (lexicographic placement LP)

Loads the repo config/ + scenario/, hydrates in-flight OG state from the
ProductionReport in Forecast.xlsm, runs L1 (forecast.global_planner_poc.plan
with record_standing=True) to get the per-(batch, week) standing population,
derives the Step-2 whole-tank demand, then runs L3
(forecast.global_planner_l3_poc.plan_l3) — a two-pass lexicographic LP that
assigns whole tanks to the 11 production OG systems: Pass A minimizes total
per-system cap violation, Pass B (slack fixed) minimizes transfers.

Prints + writes a small .xlsx:
  1. SystemLoadTrace — per (system, week): tanks/biomass/feed vs caps + OVER.
  2. Placement       — per (batch, system, week): integer tank count.
  3. Conservation    — per (batch, week): placed biomass vs L1 standing.

Usage:
    python -m tools.run_l3_poc
    python -m tools.run_l3_poc --workbook Forecast.xlsm --out l3_out.xlsx
    python -m tools.run_l3_poc --no-pr      # skip PR; incoming-only

Creates ONLY this runner + forecast/global_planner_l3_poc.py. Touches no
production file; not imported by the pipeline. If scipy is unavailable, falls
back to L2's greedy water-filler and says so.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.config_io import load_config
from forecast.scenario_io import load_batches, load_limits
from forecast import global_planner_poc as gpp
from forecast import global_planner_l2_poc as l2
from forecast import global_planner_l3_poc as l3
from tools.run_global_poc import _hydrate_inflight_og


# Greedy L2 baseline (reported for contrast).
_L2_BASELINE_BIO_OVER = 149
_L2_BASELINE_FEED_OVER = 170
_L2_BASELINE_WORST_BIO = 2.87
_L2_BASELINE_WORST_FEED = 2.46


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--scenario-dir", default=str(_ROOT / "scenario"))
    ap.add_argument("--workbook", default=str(_ROOT / "Forecast.xlsm"))
    ap.add_argument("--out", default=str(_ROOT / "global_l3_poc_out.xlsx"))
    ap.add_argument("--harvest-tank-density-pct", type=float, default=1.25)
    ap.add_argument("--slack-epsilon", type=float, default=1000.0)
    ap.add_argument("--mip-time-limit", type=float, default=180.0,
                    help="per-solve HiGHS time limit (s); returns best incumbent")
    ap.add_argument("--mip-rel-gap", type=float, default=0.01)
    ap.add_argument("--lp-relax", action="store_true",
                    help="solve the continuous LP relaxation + round (default is "
                         "a true MILP via HiGHS branch-and-bound)")
    ap.add_argument("--no-pr", action="store_true",
                    help="skip ProductionReport hydration (incoming batches only)")
    args = ap.parse_args()

    print("=" * 72)
    print("  METHOD: GLOBAL L3 (lexicographic placement LP)")
    print("=" * 72)

    control, tables, facility = load_config(args.config_dir)
    batches = load_batches(args.scenario_dir)
    _facility_limits, system_limits = load_limits(args.scenario_dir, control)
    print(f"  Config:   {args.config_dir}")
    print(f"  Scenario: {len(batches)} batches from {args.scenario_dir}")

    inflight = {}
    if not args.no_pr:
        inflight, derived_start = _hydrate_inflight_og(Path(args.workbook), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
            print(f"  ForecastStart: derived {derived_start.date()} from PR closing +1d")
        print(f"  In-flight OG batches from PR: {len(inflight)}")

    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    print(f"  forecast_start={fs_date}, horizon={control.horizon_weeks}w")

    # ---- L1: envelope + per-(batch, week) standing.
    # OG-ONLY instant-removal diagnostic: this runner's hardcoded greedy-L2
    # baseline contrast is calibrated to the OG-only model, so pin both
    # whole-facility flags False (the new defaults would otherwise shift it and
    # under-count FW without fw_inflight). The loop entry point is the
    # whole-facility path.
    l1 = gpp.plan(batches, tables, control, facility,
                  inflight_og=inflight,
                  harvest_tank_density_pct=args.harvest_tank_density_pct,
                  record_standing=True,
                  model_purge_hold=False, model_full_facility=False)
    peak_bio = max((r.standing_biomass_kg for r in l1.trace), default=0.0)
    print(f"  L1 peak standing biomass: {peak_bio:,.0f} kg "
          f"({100*peak_bio/control.max_biomass_kg:.1f}% of "
          f"{control.max_biomass_kg:,.0f} cap); L1 feasible={l1.feasible}")
    print(f"  L1 per-(batch,week) standing rows: {len(l1.batch_standing)}")

    # ---- Step 2: whole-tank demand.
    cap = l3.per_tank_capacity_kg(facility, control)
    demand = l3.build_tank_demand(l1, facility, control)
    total_tank_weeks = sum(d.tanks for d in demand)
    print(f"\n  STEP 2 (whole-tank demand)")
    print(f"    per_tank_capacity = {cap:,.0f} kg "
          f"(smallest OG tank x density_target_pct={control.density_target_pct})")
    print(f"    (batch,week) cells with demand: {len(demand)}; "
          f"total tank-weeks: {total_tank_weeks}")
    n_tanks = l3.n_tanks_per_system(facility)
    prod_sys = [s for s in l2.og_systems_from_facility(facility)
                if s in (set(l2.NURSERY_SYSTEMS) | set(l2.GROWOUT_SYSTEMS))]
    total_tank_capacity = sum(n_tanks.get(s, 0) for s in prod_sys)
    print(f"    production systems ({len(prod_sys)}): "
          f"{', '.join(prod_sys)}  ->  {total_tank_capacity} tanks total")

    # ---- L3: the LP (with greedy fallback).
    try:
        res = l3.plan_l3(l1, control, facility, system_limits,
                         slack_epsilon=args.slack_epsilon,
                         integer=not args.lp_relax,
                         mip_time_limit=args.mip_time_limit,
                         mip_rel_gap=args.mip_rel_gap, verbose=True)
        print(f"  MODE: {'LP-relaxation + rounding' if args.lp_relax else 'MILP (integer y)'}")
    except RuntimeError as e:
        print(f"\n  [L3] LP unavailable ({e}); FALLING BACK to L2 greedy water-fill.")
        return _run_greedy_fallback(l1, control, facility, system_limits,
                                    Path(args.out))

    _report(res)
    _write_outputs(res, control, Path(args.out))
    return 0


def _report(res: "l3.L3Result") -> None:
    print(f"\n  SOLVER: {res.solver}")
    print(f"    y-vars: {res.n_y_vars:,}; constraints (Pass A): "
          f"{res.n_constraints:,}")

    print("\n  PASS A — meet the limits (minimize total cap violation)")
    print(f"    LP optimum sum(slk_tank): {res.passA_tank_slack:.1f} tanks over "
          f"physical (prioritized — the harder limit)")
    print(f"    LP optimum sum(slk_bio+slk_feed): "
          f"{res.passA_slack_total:,.1f} kg (status {res.passA_status})")

    # Realized tank overflow from the rounded layout.
    over_tank_sw = sum(1 for r in res.loads if r.n_tanks > r.n_tanks_cap)
    over_tank_tanks = sum(max(0, r.n_tanks - r.n_tanks_cap) for r in res.loads)
    print("\n  REALIZED cap violations (after LP-relaxation rounding)")
    print(f"    system-weeks over TANK cap:    {over_tank_sw} "
          f"({over_tank_tanks} tanks over physical, cumulative)")
    print(f"    system-weeks over BIOMASS cap: {res.over_biomass_system_weeks}")
    print(f"    system-weeks over FEED cap:    {res.over_feed_system_weeks}")
    print(f"    total over-biomass: {res.total_over_biomass_kg:,.0f} kg")
    print(f"    total over-feed:    {res.total_over_feed_kg:,.0f} kg/day")
    print(f"    worst biomass fill ratio: {res.worst_biomass_ratio:.3f}")
    print(f"    worst feed fill ratio:    {res.worst_feed_ratio:.3f}")
    fully_legal = (res.over_biomass_system_weeks == 0
                   and res.over_feed_system_weeks == 0)
    print(f"    FULLY SYSTEM-LEGAL (0 over-cap): {fully_legal}")

    print("\n  CONTRAST vs greedy L2 baseline")
    print(f"    L2 greedy: {_L2_BASELINE_BIO_OVER} biomass / "
          f"{_L2_BASELINE_FEED_OVER} feed system-weeks over; "
          f"worst {_L2_BASELINE_WORST_BIO}x / {_L2_BASELINE_WORST_FEED}x")
    print(f"    L3 LP:     {res.over_biomass_system_weeks} biomass / "
          f"{res.over_feed_system_weeks} feed system-weeks over; "
          f"worst {res.worst_biomass_ratio:.2f}x / {res.worst_feed_ratio:.2f}x")

    print("\n  PASS B — minimize transfers (cap-slack fixed at Pass-A optimum)")
    print(f"    method: {res.passB_status}")
    print(f"    transfers (tanks-worth of a batch entering a NEW system "
          f"week-to-week): {res.realized_transfers:,.0f}")
    print(f"    avg systems per (batch, week): "
          f"{res.avg_systems_per_batch_week:.2f}")
    print(f"    avg distinct systems per batch (whole horizon): "
          f"{res.avg_systems_per_batch_horizon:.2f}")

    print("\n  CONSERVATION (exact)")
    print(f"    worst |placed biomass - L1 standing| per (batch,week): "
          f"{res.worst_biomass_residual_kg:.6f} kg "
          f"-- {'OK conserves' if res.worst_biomass_residual_kg < 1e-3 else 'CHECK'}")
    print(f"    worst |sum_s y - tanks[b,w]|: "
          f"{res.worst_tankcount_residual:.0f} tanks "
          f"-- {'OK (= tanks)' if res.worst_tankcount_residual < 1e-9 else 'CHECK'}")

    print("\n  LP-RELAXATION INTEGRALITY")
    print(f"    max |y_lp - round(y_lp)|: {res.integrality_gap:.4f}")
    print(f"    rounding fixups (tank-count repairs to restore = tanks): "
          f"{res.rounding_fixups}")


def _run_greedy_fallback(l1, control, facility, system_limits, out_path) -> int:
    res = l2.assign_split(l1, control, facility, system_limits)
    print("\n  [FALLBACK] L2 split water-filler used (no LP).")
    print(f"    over biomass system-weeks: {res.over_biomass_system_weeks}")
    print(f"    over feed system-weeks:    {res.over_feed_system_weeks}")
    print(f"    worst biomass ratio: {res.worst_biomass_ratio:.3f}; "
          f"worst feed ratio: {res.worst_feed_ratio:.3f}")
    print(f"    avg systems per (batch, week): "
          f"{res.avg_systems_per_batch_week:.2f}")
    return 0


def _write_outputs(res: "l3.L3Result", control, out_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        print(f"\n  (openpyxl unavailable: {e}); skipping xlsx")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "METHOD"
    ws["A1"] = "METHOD: GLOBAL L3 (lexicographic placement LP)"
    ws["A2"] = f"forecast_start={control.forecast_start}"
    ws["A3"] = f"horizon_weeks={control.horizon_weeks}"
    ws["A4"] = f"solver={res.solver}"
    ws["A5"] = f"systems={','.join(res.systems)}"
    ws["A6"] = f"passA_slack_total_kg={res.passA_slack_total:.1f}"
    ws["A7"] = f"over_biomass_system_weeks={res.over_biomass_system_weeks}"
    ws["A8"] = f"over_feed_system_weeks={res.over_feed_system_weeks}"
    ws["A9"] = f"passB_transfers={res.passB_transfers}"
    ws["A10"] = f"realized_transfers={res.realized_transfers:.0f}"
    ws["A11"] = f"avg_systems_per_batch_week={res.avg_systems_per_batch_week:.3f}"
    ws["A12"] = f"integrality_gap={res.integrality_gap:.4f}"
    ws["A13"] = f"rounding_fixups={res.rounding_fixups}"
    ws["A14"] = f"worst_biomass_residual_kg={res.worst_biomass_residual_kg:.6f}"

    wl = wb.create_sheet("SystemLoadTrace")
    wl.append(["week", "week_label", "system_id", "tier", "n_tanks",
               "n_tanks_cap", "biomass_kg", "biomass_cap", "feed_kg_day",
               "feed_cap", "over_biomass_kg", "over_feed_kg",
               "over_biomass", "over_feed"])
    for r in res.loads:
        wl.append([r.week, r.week_label, r.system_id, r.tier, r.n_tanks,
                   r.n_tanks_cap, round(r.biomass_kg, 1), r.biomass_cap,
                   round(r.feed_kg_day, 1), r.feed_cap,
                   round(r.over_biomass_kg, 1), round(r.over_feed_kg, 1),
                   r.over_biomass, r.over_feed])

    wp = wb.create_sheet("Placement")
    wp.append(["week", "week_label", "batch_id", "system_id", "tier", "tanks",
               "biomass_kg", "feed_kg_day"])
    for p in res.placements:
        wp.append([p.week, p.week_label, p.batch_id, p.system_id, p.tier,
                   p.tanks, round(p.biomass_kg, 1), round(p.feed_kg_day, 1)])

    try:
        wb.save(out_path)
        print(f"\n  Wrote {out_path}")
    except Exception as e:  # noqa: BLE001
        print(f"\n  (could not save xlsx: {e})")


if __name__ == "__main__":
    raise SystemExit(main())
