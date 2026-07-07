"""The compare harness: run several planning methods on ONE set of inputs and
write a single, legible RunComparison workbook.

Every method (see forecast.methods) runs from the SAME PR + config + scenario
(so the SAME manual override window and control rules apply — only the planning
method differs), each producing its own full forecast workbook. Each workbook is
then scored on the SAME axes via forecast.optimize.metrics_from_workbook (which
carries the hard conservation gate), plus the harvest-contract counts the
operator cares about (min→max weekly, weeks below min, zero-harvest weeks). The
results land side-by-side in the RunComparison sheet.

Usage:
    python -m tools.run_compare PATH_TO_PR.xlsm
    python -m tools.run_compare PR.xlsm --methods controller,global-lp --out cmp.xlsx

Nothing here touches production files: every method runs in an isolated temp
copy of config/scenario, and outputs go to a chosen directory (default: a
sibling '<pr-stem>_methods/' folder). Each method run is ~minutes, so a full
four-method compare is a batch job — progress is printed per method, and a
method that errors is recorded as a failed column rather than aborting the run.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from forecast import methods as _methods           # noqa: E402
from forecast import optimize as _opt              # noqa: E402
from forecast.excel_io import write_run_comparison  # noqa: E402


def _harvest_extras(out_path, min_harvest):
    """Harvest-contract counts from a produced workbook's HarvestPlan: total
    weeks, min/max weekly fish, weeks below the min-harvest floor, and zero-
    harvest weeks (the 'never an empty week' contract). Computed over ALL harvest
    weeks — the manual override weeks are identical across every method, so they
    offset every column equally and the comparison stays fair."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    try:
        fish = _opt._harvest_weekly_fish(wb)
    finally:
        wb.close()
    if not fish:
        return {"n_weeks": 0, "min_week": 0.0, "max_week": 0.0,
                "weeks_below_min": 0, "zero_weeks": 0, "min_harvest": min_harvest}
    return {
        "n_weeks": len(fish),
        "min_week": min(fish),
        "max_week": max(fish),
        "weeks_below_min": (sum(1 for x in fish if x < min_harvest)
                            if min_harvest else 0),
        "zero_weeks": sum(1 for x in fish if x < 1.0),
        "min_harvest": min_harvest,
    }


def _conservation_verdict(out_path):
    """Authoritative per-method conservation, using each method's OWN proof.

    Global emits a batch-level ReconciliationReport (seeded == harvested +
    standing + mortality + cull). That is the mass-conservation truth. Its
    InputConservationAudit 'never placed' flag OVER-reports on LP output — the
    LP holds late arrivals in the biomass envelope without a per-tank row — so
    for Global that flag means a REALIZED-PLACEMENT gap, not lost mass. The
    Controller places tank-by-tank, so its InputConservationAudit drops ARE
    real lost fish (use tuning._conservation).

    Returns dict: gate ('PASS'|'PARTIAL'|'FAIL'), dropped, overprod,
    unplaced_batches, unplaced_fish, residual_pct.
    """
    from forecast import tuning
    wb = openpyxl.load_workbook(out_path, data_only=True)
    sheets = set(wb.sheetnames)

    # Realized-placement gap headline (present on both engines' audit).
    unplaced_b = unplaced_f = 0
    if "InputConservationAudit" in sheets:
        for row in wb["InputConservationAudit"].iter_rows(values_only=True):
            line = " ".join(str(c) for c in row if c is not None)
            m = re.search(r"(\d[\d,]*)\s+batch\(es\)\s+DROPPED\s*[—-]+\s*"
                          r"([\d,]+)\s+stocked fish", line)
            if m:
                unplaced_b = int(m.group(1).replace(",", ""))
                unplaced_f = int(m.group(2).replace(",", ""))
                break

    # BOTH engines emit a sheet named 'ReconciliationReport', but of different
    # shape: Global's carries a FACILITY summary row (seeded == harvested +
    # standing + mortality + cull) whose last column is the mass residual %; the
    # Controller's is a per-(batch, week) balance table with NO FACILITY row.
    # Key the gate off the FACILITY row's PRESENCE, not the sheet name — else the
    # Controller wrongly takes the mass-gate branch, finds no residual, and FAILs
    # a run that conserves perfectly.
    facility_row = None
    if "ReconciliationReport" in sheets:
        for row in wb["ReconciliationReport"].iter_rows(values_only=True):
            if row and str(row[0]).strip().upper() == "FACILITY":
                facility_row = row
                break

    if facility_row is not None:                    # -> Global: mass gate
        try:
            residual_pct = float(facility_row[-1])
        except (TypeError, ValueError):
            residual_pct = None
        wb.close()
        mass_ok = residual_pct is not None and abs(residual_pct) < 0.01
        gate = ("FAIL" if not mass_ok else
                "PARTIAL" if unplaced_b > 0 else "PASS")
        return {"gate": gate, "dropped": 0 if mass_ok else unplaced_f,
                "overprod": 0, "unplaced_batches": unplaced_b,
                "unplaced_fish": unplaced_f, "residual_pct": residual_pct}

    wb.close()                                      # -> Controller: audit drops
    dropped, overprod = tuning._conservation(out_path)
    gate = "PASS" if dropped == 0 and overprod == 0 else "FAIL"
    return {"gate": gate, "dropped": dropped, "overprod": overprod,
            "unplaced_batches": unplaced_b if dropped else 0,
            "unplaced_fish": dropped, "residual_pct": None}


def run_compare(input_path, *, config_dir=None, scenario_dir=None,
                out_path=None, out_dir=None, method_keys=None,
                quiet=True, reuse=False) -> int:
    """Run the roster of methods on `input_path` and write the RunComparison
    workbook to `out_path`. Returns 0 on success (even if individual methods
    fail — those become failed columns). See module docstring.

    reuse=True skips the (multi-minute) engine run for any method whose output
    workbook already exists, re-scoring the cached workbook instead — so the
    comparison can be re-rendered after a scoring/gate change without paying for
    the engine runs again. Runtimes are read back from a small sidecar cache."""
    # Method labels/blurbs carry unicode (em dash, arrows); the Windows console
    # defaults to cp1252 and would crash a multi-minute run on the first print.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:                                            # noqa: BLE001
        pass
    input_path = Path(input_path)
    config_dir = Path(config_dir) if config_dir else ROOT / "config"
    scenario_dir = Path(scenario_dir) if scenario_dir else ROOT / "scenario"
    stem = input_path.stem
    out_dir = Path(out_dir) if out_dir else input_path.with_name(f"{stem}_methods")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = (Path(out_path) if out_path
                else input_path.with_name(f"{stem}_COMPARISON.xlsx"))

    roster = _methods.get_roster(method_keys)

    # Harvest rules from the SAME config every method reads.
    with open(config_dir / "control.yaml") as f:
        cfg = yaml.safe_load(f) or {}
    min_harvest = float(cfg.get("min_harvest_per_week", 0) or 0)
    harvest_cap = float(cfg.get("max_harvest_per_week", 55000) or 55000)

    print(f"RUN COMPARISON — {len(roster)} method(s) on {input_path.name}")
    print(f"  config={config_dir}  scenario={scenario_dir}")
    print(f"  per-method workbooks -> {out_dir}")
    print(f"  min_harvest_per_week={min_harvest:.0f}  "
          f"max_harvest_per_week={harvest_cap:.0f}"
          f"{'   [reuse: skip runs with an existing workbook]' if reuse else ''}\n")

    # Sidecar runtime cache so reuse can keep each method's wall time.
    et_path = out_dir / "_elapsed.json"
    try:
        elapsed_cache = json.loads(et_path.read_text())
    except Exception:                                            # noqa: BLE001
        elapsed_cache = {}

    def _score(rec, wb_path):
        metrics, _d, _o = _opt.metrics_from_workbook(str(wb_path), harvest_cap)
        verdict = _conservation_verdict(str(wb_path))
        rec["metrics"] = metrics
        rec["dropped"] = verdict["dropped"]
        rec["overprod"] = verdict["overprod"]
        rec["gate"] = verdict["gate"]
        rec["placement"] = {"unplaced_batches": verdict["unplaced_batches"],
                            "unplaced_fish": verdict["unplaced_fish"]}
        rec["harvest"] = _harvest_extras(str(wb_path), min_harvest)
        extra = (f" ({verdict['unplaced_batches']} batch(es) / "
                 f"{verdict['unplaced_fish']:,} fish unplaced)"
                 if verdict["gate"] == "PARTIAL" else "")
        print(f"    conservation {verdict['gate']}{extra}, "
              f"peak {metrics.overall_peak_biomass:,.0f} kg, "
              f"transfers/fish {metrics.transfers_per_fish:.3f}")

    def _written(wb_path):
        # run.main coerces the output extension to the workbook's content type
        # (.xlsm when it kept a VBA archive from an .xlsm template, else .xlsx),
        # so a controller run we asked to save '.xlsx' lands on disk as '.xlsm'.
        # Score whichever file actually exists.
        if wb_path.exists():
            return wb_path
        alt = wb_path.with_suffix(".xlsm")
        return alt if alt.exists() else wb_path

    records = []
    for i, m in enumerate(roster, 1):
        wb_path = out_dir / f"{stem}_{m.key}.xlsx"
        rec = {"key": m.key, "label": m.label, "family": m.family,
               "blurb": m.blurb, "workbook": wb_path.name, "failed": None,
               "elapsed": None, "dropped": None, "overprod": None,
               "gate": None, "metrics": None, "harvest": None,
               "placement": None}
        print(f"[{i}/{len(roster)}] {m.key} — {m.label} ...", flush=True)
        try:
            cached = _written(wb_path)
            if reuse and cached.exists() and cached.stat().st_size > 0:
                rec["elapsed"] = elapsed_cache.get(m.key)
                rec["workbook"] = cached.name
                print(f"    reusing cached workbook {cached.name}")
                _score(rec, cached)
            else:
                rc, elapsed = _methods.run_method(
                    m, input_path, wb_path, config_dir, scenario_dir, quiet=quiet)
                rec["elapsed"] = elapsed
                elapsed_cache[m.key] = elapsed
                et_path.write_text(json.dumps(elapsed_cache))
                if rc != 0:
                    rec["failed"] = f"engine returned rc={rc}"
                    print(f"    FAILED rc={rc} ({elapsed:.0f}s)")
                else:
                    print(f"    done {elapsed:.0f}s", flush=True)
                    written = _written(wb_path)
                    rec["workbook"] = written.name
                    _score(rec, written)
        except Exception as e:                                   # noqa: BLE001
            rec["failed"] = f"{type(e).__name__}: {e}"
            print(f"    FAILED: {rec['failed']}")
            traceback.print_exc()
        records.append(rec)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_run_comparison(wb, records, pr_name=input_path.name,
                         generated=datetime.now().isoformat(timespec="seconds"))
    wb.save(out_path)
    wb.close()

    n_pass = sum(1 for r in records if r.get("gate") == "PASS")
    n_part = sum(1 for r in records if r.get("gate") == "PARTIAL")
    n_fail = len(records) - n_pass - n_part
    print(f"\nWrote {out_path}")
    print(f"  gates: {n_pass} PASS · {n_part} PARTIAL · {n_fail} FAILED/errored "
          f"(of {len(records)}). Only PASS methods are directly comparable.")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", help="path to the PR workbook (.xlsm/.xlsx)")
    ap.add_argument("--config-dir", default=None)
    ap.add_argument("--scenario-dir", default=None)
    ap.add_argument("--out", default=None,
                    help="RunComparison workbook path "
                         "(default: <pr-stem>_COMPARISON.xlsx)")
    ap.add_argument("--out-dir", default=None,
                    help="dir for each method's full workbook "
                         "(default: <pr-stem>_methods/)")
    ap.add_argument("--methods", default=None,
                    help="comma-separated method keys (default: all). "
                         f"available: {', '.join(_methods.DEFAULT_ROSTER)}")
    ap.add_argument("--verbose", action="store_true",
                    help="stream each engine's own stdout (default: quiet)")
    ap.add_argument("--reuse", action="store_true",
                    help="skip the engine run for any method whose output "
                         "workbook already exists; re-score + re-render only")
    args = ap.parse_args()
    keys = ([k.strip() for k in args.methods.split(",") if k.strip()]
            if args.methods else None)
    return run_compare(args.workbook, config_dir=args.config_dir,
                       scenario_dir=args.scenario_dir, out_path=args.out,
                       out_dir=args.out_dir, method_keys=keys,
                       quiet=not args.verbose, reuse=args.reuse)


if __name__ == "__main__":
    raise SystemExit(main())
