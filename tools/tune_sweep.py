"""Controller tuning sweep — answer "what knobs reduce per-batch density over-cap?"

Runs the forecast repeatedly over a grid of Control knob values and reports, for
each variant, the PER-BATCH PEAK-DENSITY DISTRIBUTION (from the TransferTemplate
Section B "Plan" data) plus the conservation gates. This is the empirical recipe
for tuning a new scenario: sweep, read the distribution, pick the row that
minimises *severe* over-cap while holding conservation and utilisation.

WHY a distribution and not a single "OVER CAP" count:
  Running a facility near capacity means many batches PEAK right at the density
  cap. With ~10%/week growth and weekly rebalancing, a tank sitting at cap will
  cross it mid-week before the next check — a structural +3-10% touch, not a
  problem. So 1.00-1.10 peaks are "at cap / normal"; the rows that matter are the
  SEVERE ones (>1.3x), which are usually a CAPACITY collision (too much biomass
  wanting grow-out tanks at once) that NO controller knob can fix — only a
  stocking/cadence/tank-count change. The sweep makes that distinction visible.

Usage (from the repo root):
  python -m tools.tune_sweep                         # repo config/ + scenario/ yaml
  python -m tools.tune_sweep --config-template "C:\\path\\config_template (7).xlsx"
  python -m tools.tune_sweep --input "C:\\path\\Forecast.xlsm"

Edit GRID below to sweep different knobs/values. Each run is ~90s.
"""
from __future__ import annotations
import argparse
import contextlib
import io
import os
import shutil
import statistics
import sys
import tempfile

import openpyxl
import yaml

import forecast.run as run

# Knob grid: each entry is (label, {control-knob: value, ...}). The first row is
# conventionally the baseline (no overrides). Add/remove rows freely.
GRID = [
    ("baseline",            {}),
    ("density=0.90",        {"density_target_pct": 0.90}),
    ("density=0.85",        {"density_target_pct": 0.85}),
    ("varqty=20",           {"rebalance_varqty_budget": 20}),
    ("balance=60",          {"rebalance_balance_budget": 60}),
    ("setpoint=0.90",       {"harvest_setpoint_lookahead_weeks": 0.90}),
    ("setpoint=1.20",       {"harvest_setpoint_lookahead_weeks": 1.20}),
]


def _base_config(args, work):
    """Return (config_dir, scenario_dir) seeded either from a config_template
    workbook or from the repo's config/ + scenario/ yaml."""
    cdir = os.path.join(work, "config")
    sdir = os.path.join(work, "scenario")
    os.makedirs(cdir)
    os.makedirs(sdir)
    for src, dst in (("config", cdir), ("scenario", sdir)):
        for f in os.listdir(src):
            if f.endswith(".yaml"):
                shutil.copy(os.path.join(src, f), dst)
    if args.config_template:
        import forecast.config_template as ct
        wb = openpyxl.load_workbook(args.config_template, data_only=True)
        ct.import_config_template(wb, cdir, sdir)
    return cdir, sdir


def _peak_densities(out_path):
    """Pull each batch's peak density ratio from TransferTemplate Section B."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    hdr = None
    vals = []
    for row in wb["TransferTemplate"].iter_rows(values_only=True):
        if row and row[0] == "Batch":
            hdr = [str(c) for c in row if c is not None]
            continue
        if (hdr and row and isinstance(row[0], str) and row[0].startswith("B")
                and len(row[0]) > 1 and row[0][1].isdigit()):
            pc = next((c for c in hdr if c.startswith("Peak_Density")), None)
            d = {hdr[i]: row[i] for i in range(min(len(hdr), len(row)))}
            try:
                vals.append(float(d.get(pc) or 0))
            except (TypeError, ValueError):
                pass
    return vals


def _conservation(out_path):
    """Scan the audit sheets for dropped / over-produced fish (must be 0)."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    dropped = overprod = 0
    for sh in ("TankContinuityAudit", "InputConservationAudit"):
        if sh not in wb.sheetnames:
            continue
        for row in wb[sh].iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            line = " ".join(cells).upper()
            for c in cells:
                try:
                    v = int(float(c))
                except (TypeError, ValueError):
                    continue
                if v > 0 and "DROP" in line:
                    dropped = max(dropped, v)
                if v > 0 and "OVER-PRODUCED" in line:
                    overprod = max(overprod, v)
    return dropped, overprod


def _run_one(label, overrides, cdir0, sdir0, input_xlsm):
    work = tempfile.mkdtemp()
    cdir = os.path.join(work, "config")
    sdir = os.path.join(work, "scenario")
    shutil.copytree(cdir0, cdir)
    shutil.copytree(sdir0, sdir)
    cy = os.path.join(cdir, "control.yaml")
    with open(cy) as f:
        cfg = yaml.safe_load(f)
    cfg.update(overrides)
    with open(cy, "w") as f:
        yaml.safe_dump(cfg, f)
    inp = os.path.join(work, "in.xlsm")
    out = os.path.join(work, "out.xlsm")
    shutil.copy(input_xlsm, inp)
    with contextlib.redirect_stdout(io.StringIO()):
        run.main(inp, out, config_dir=cdir, scenario_dir=sdir)
    v = _peak_densities(out)
    dropped, overprod = _conservation(out)
    return v, dropped, overprod


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", default="Forecast.xlsm",
                    help="Input forecast workbook (PR source). Default: repo Forecast.xlsm")
    ap.add_argument("--config-template", default=None,
                    help="Optional config_template.xlsx to seed config/scenario from. "
                         "If omitted, the repo config/ + scenario/ yaml are used.")
    args = ap.parse_args(argv)

    if not os.path.exists(args.input):
        sys.exit(f"input workbook not found: {args.input}")

    seed = tempfile.mkdtemp()
    cdir0, sdir0 = _base_config(args, seed)

    print(f"input={args.input}  config={'template:'+args.config_template if args.config_template else 'repo yaml'}")
    print(f"{'variant':<18} {'OVER':>5} {'sev>1.3':>7} {'worst':>6} {'median':>6} | "
          f"{'<=1.0':>5} {'1.0-1.1':>7} {'1.1-1.3':>7} {'>1.3':>5} | cons")
    print("-" * 96)
    for label, overrides in GRID:
        try:
            v, dropped, overprod = _run_one(label, overrides, cdir0, sdir0, args.input)
        except Exception as e:  # noqa: BLE001 - sweep should survive one bad cell
            print(f"{label:<18} ERROR: {e}")
            continue
        n = len(v)
        b = lambda lo, hi: sum(1 for x in v if lo <= x < hi)
        over = sum(1 for x in v if x > 1.0001)
        sev = sum(1 for x in v if x >= 1.3)
        cons = "OK" if (dropped == 0 and overprod == 0) else f"FAIL d={dropped} o={overprod}"
        print(f"{label:<18} {over:>2}/{n:<2} {sev:>7} {max(v):>6.2f} {statistics.median(v):>6.2f} | "
              f"{b(0,1.0001):>5} {b(1.0001,1.1):>7} {b(1.1,1.3):>7} {sum(1 for x in v if x>=1.3):>5} | {cons}")
    print("\nRead: minimise sev>1.3 while cons==OK. If no row beats baseline, the "
          "severe peaks are a CAPACITY collision (stocking/cadence/tank-count), "
          "not a controller-tuning problem -- see docs/USER_GUIDE.md sec 7.")


if __name__ == "__main__":
    main()
