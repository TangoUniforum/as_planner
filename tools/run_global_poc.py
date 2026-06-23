"""Runner for the tankless GLOBAL harvest planner (Layer 1) proof-of-concept.

METHOD: GLOBAL (tankless L1 POC)

Loads the repo config/ + scenario/, hydrates in-flight OG state from the
ProductionReport in Forecast.xlsm (so running batches seed at their real PR
state), runs forecast.global_planner_poc.plan(), and writes a small .xlsx
(falling back to CSV + stdout) with:

  1. HarvestEnvelope  — per (batch, week): count, biomass, avg weight harvested.
  2. StandingTrace    — per week: facility biomass + feed/day vs the caps.
  3. Feasibility      — every week legal? which cap / over-stock kg if not.
  4. Conservation     — per batch: input ~= harvested + standing + mort + cull.

Usage:
    python -m tools.run_global_poc
    python -m tools.run_global_poc --workbook Forecast.xlsm --out out.xlsx
    python -m tools.run_global_poc --no-pr        # skip PR; incoming-only

Creates ONLY this runner + forecast/global_planner_poc.py. Touches no existing
file and is not imported by the production pipeline.
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

# Allow `python tools/run_global_poc.py` as well as `-m tools.run_global_poc`.
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.config_io import load_config
from forecast.scenario_io import load_batches
from forecast import global_planner_poc as gpp


def _hydrate_inflight_og(workbook_path: Path, batches):
    """Read PR -> per-batch in-flight OG (count, avg_wt_g, cv_pct) + derived start.

    Mirrors what run.py does: derive forecast_start = PR closing + 1 day, and
    aggregate OG (batch, tank) rows into one (count, avg_wt, cv) per batch.
    Returns (inflight_og dict, derived_forecast_start or None).
    """
    try:
        from forecast.excel_io import load_workbook
        from forecast.production_report import read_production_report
    except Exception as e:  # noqa: BLE001
        print(f"  (could not import PR reader: {e}); running incoming-only")
        return {}, None
    if not workbook_path.exists():
        print(f"  (workbook {workbook_path} not found; running incoming-only)")
        return {}, None
    wb = load_workbook(workbook_path)
    pr_closing, og_records, fw_records = read_production_report(wb)
    wb.close()
    derived_start = None
    if pr_closing is not None:
        derived_start = datetime(pr_closing.year, pr_closing.month,
                                 pr_closing.day) + timedelta(days=1)
    # Aggregate OG rows per batch.
    agg: dict[str, dict] = {}
    for r in og_records:
        e = agg.setdefault(r.batch_id, {"count": 0.0, "biomass_kg": 0.0})
        e["count"] += r.closing_count
        e["biomass_kg"] += r.closing_biomass_kg
    batch_cv = {b.batch_id: b.tran_og_cv for b in batches}
    inflight = {}
    for bid, e in agg.items():
        if e["count"] > 0:
            avg_wt = e["biomass_kg"] * 1000.0 / e["count"]
            inflight[bid] = (e["count"], avg_wt, batch_cv.get(bid, 16.0))
    return inflight, derived_start


def _fmt(n, w=12, dp=0):
    return f"{n:>{w},.{dp}f}"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--scenario-dir", default=str(_ROOT / "scenario"))
    ap.add_argument("--workbook", default=str(_ROOT / "Forecast.xlsm"))
    ap.add_argument("--out", default=str(_ROOT / "global_poc_out.xlsx"))
    ap.add_argument("--no-pr", action="store_true",
                    help="skip ProductionReport hydration (incoming batches only)")
    args = ap.parse_args()

    print("=" * 72)
    print("  METHOD: GLOBAL (tankless L1 POC)")
    print("=" * 72)

    control, tables, facility = load_config(args.config_dir)
    batches = load_batches(args.scenario_dir)
    print(f"  Config:   {args.config_dir}")
    print(f"  Scenario: {len(batches)} batches from {args.scenario_dir}")

    inflight = {}
    if not args.no_pr:
        inflight, derived_start = _hydrate_inflight_og(Path(args.workbook), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
            print(f"  ForecastStart: derived {derived_start.date()} from PR closing +1d")
        print(f"  In-flight OG batches from PR: {len(inflight)}")

    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    print(f"  forecast_start={fs_date}, horizon={control.horizon_weeks}w")
    print(f"  Caps: biomass<={control.max_biomass_kg:,.0f} kg, "
          f"feed<={control.max_feed_per_day_kg:,.0f} kg/day, "
          f"max_harvest={control.max_harvest_per_week:,.0f} fish/wk, "
          f"min_harvest_wt={control.min_harvest_weight_g:,.0f} g")

    # OG-ONLY instant-removal diagnostic: this runner intentionally exercises
    # the original L1 envelope (no 6N purge hold, OG-only cap). Pass both flags
    # explicitly False so the new whole-facility defaults do not silently change
    # this diagnostic (and avoid under-counting FW without fw_inflight).
    res = gpp.plan(batches, tables, control, facility, inflight_og=inflight,
                   model_purge_hold=False, model_full_facility=False)
    print(f"  OG-tank weekly draw ceiling (smallest OG tank): "
          f"{res.og_tank_ceiling_kg:,.0f} kg")
    print(f"  Seeds (batches entering OG in horizon): {len(res.seeds)}")

    # ---- Standing trace summary ----
    peak_bio = max((r.standing_biomass_kg for r in res.trace), default=0.0)
    peak_feed = max((r.feed_kg_day for r in res.trace), default=0.0)
    bio_cap = control.max_biomass_kg
    feed_cap = control.max_feed_per_day_kg
    print("\n  STANDING TRACE SUMMARY")
    print(f"    peak standing biomass: {peak_bio:,.0f} kg "
          f"({100*peak_bio/bio_cap:.1f}% of {bio_cap:,.0f} cap)")
    print(f"    peak feed/day:         {peak_feed:,.0f} kg "
          f"({100*peak_feed/feed_cap:.1f}% of {feed_cap:,.0f} cap)")
    total_harv_kg = sum(r.harvested_kg for r in res.trace)
    total_harv_ct = sum(r.harvested_count for r in res.trace)
    print(f"    total harvested:       {total_harv_kg:,.0f} kg, "
          f"{total_harv_ct:,.0f} fish over {control.horizon_weeks} weeks")

    # ---- Feasibility verdict ----
    print("\n  FEASIBILITY VERDICT")
    if res.feasible:
        print(f"    LEGAL every week: rides under both caps for all "
              f"{control.horizon_weeks} weeks.")
    else:
        print(f"    INFEASIBLE in {len(res.infeasible_weeks)} week(s) "
              f"(over-stocked beyond the weekly harvest ceiling):")
        for wk, label, cap, over in res.infeasible_weeks[:15]:
            print(f"      wk {wk:>2} {label}: {cap} cap exceeded by {over:,.0f} kg")
        if len(res.infeasible_weeks) > 15:
            print(f"      ... ({len(res.infeasible_weeks) - 15} more)")

    # ---- Conservation ----
    print("\n  CONSERVATION (per batch: seeded ~= harvested + standing + mort + cull)")
    print(f"    {'batch':<6} {'seeded':>11} {'harv':>11} {'standing':>11} "
          f"{'mort':>11} {'cull':>10} {'resid%':>8}")
    worst = 0.0
    for bid in sorted(res.conservation):
        c = res.conservation[bid]
        worst = max(worst, abs(c["residual_pct"]))
        print(f"    {bid:<6} {_fmt(c['seeded_count'],11)} "
              f"{_fmt(c['harvested_count'],11)} {_fmt(c['standing_count'],11)} "
              f"{_fmt(c['mortality_count'],11)} {_fmt(c['cull_count'],10)} "
              f"{c['residual_pct']:>7.3f}%")
    print(f"    worst |residual| across batches: {worst:.4f}% "
          f"({'OK — conserves' if worst < 0.01 else 'check'})")

    # ---- Write workbook (xlsx) or CSV fallback ----
    _write_outputs(res, control, Path(args.out))
    return 0


def _write_outputs(res, control, out_path: Path):
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        print(f"\n  (openpyxl unavailable: {e}); writing CSVs instead")
        _write_csvs(res, out_path.with_suffix(""))
        return
    wb = Workbook()
    # Banner sheet.
    ws = wb.active
    ws.title = "METHOD"
    ws["A1"] = "METHOD: GLOBAL (tankless L1 POC)"
    ws["A2"] = f"forecast_start={control.forecast_start}"
    ws["A3"] = f"horizon_weeks={control.horizon_weeks}"
    ws["A4"] = f"biomass_cap={control.max_biomass_kg}"
    ws["A5"] = f"feed_cap={control.max_feed_per_day_kg}"
    ws["A6"] = f"OG_tank_weekly_ceiling_kg={res.og_tank_ceiling_kg:.0f}"
    ws["A7"] = f"feasible={res.feasible}"

    we = wb.create_sheet("HarvestEnvelope")
    we.append(["week", "week_label", "batch_id", "count", "biomass_kg", "avg_wt_g"])
    for r in res.envelope:
        we.append([r.week, r.week_label, r.batch_id, round(r.count, 1),
                   round(r.biomass_kg, 1), round(r.avg_wt_g, 1)])

    wt = wb.create_sheet("StandingTrace")
    wt.append(["week", "week_label", "standing_biomass_kg", "biomass_cap",
               "feed_kg_day", "feed_cap", "harvested_kg", "harvested_count",
               "required_kg", "binding", "legal", "over_biomass_kg", "over_feed_kg"])
    for r in res.trace:
        wt.append([r.week, r.week_label, round(r.standing_biomass_kg, 1),
                   r.biomass_cap, round(r.feed_kg_day, 1), r.feed_cap,
                   round(r.harvested_kg, 1), round(r.harvested_count, 1),
                   round(r.required_kg, 1), r.binding, r.legal,
                   round(r.over_biomass_kg, 1), round(r.over_feed_kg, 1)])

    wf = wb.create_sheet("Feasibility")
    wf.append(["feasible_overall", res.feasible])
    wf.append(["week", "week_label", "cap_exceeded", "over_kg"])
    for wk, label, cap, over in res.infeasible_weeks:
        wf.append([wk, label, cap, round(over, 1)])

    wc = wb.create_sheet("Conservation")
    wc.append(["batch_id", "seeded_count", "harvested_count", "standing_count",
               "mortality_count", "cull_count", "accounted_count",
               "residual_count", "residual_pct"])
    for bid in sorted(res.conservation):
        c = res.conservation[bid]
        wc.append([bid, round(c["seeded_count"], 1),
                   round(c["harvested_count"], 1), round(c["standing_count"], 1),
                   round(c["mortality_count"], 1), round(c["cull_count"], 1),
                   round(c["accounted_count"], 1), round(c["residual_count"], 3),
                   round(c["residual_pct"], 4)])
    try:
        wb.save(out_path)
        print(f"\n  Wrote {out_path}")
    except Exception as e:  # noqa: BLE001
        print(f"\n  (could not save xlsx: {e}); writing CSVs instead")
        _write_csvs(res, out_path.with_suffix(""))


def _write_csvs(res, stem: Path):
    import csv
    stem.parent.mkdir(parents=True, exist_ok=True)
    with open(f"{stem}_envelope.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["week", "week_label", "batch_id", "count", "biomass_kg", "avg_wt_g"])
        for r in res.envelope:
            w.writerow([r.week, r.week_label, r.batch_id, r.count, r.biomass_kg, r.avg_wt_g])
    with open(f"{stem}_trace.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["week", "week_label", "standing_biomass_kg", "biomass_cap",
                    "feed_kg_day", "feed_cap", "harvested_kg", "required_kg",
                    "binding", "legal", "over_biomass_kg", "over_feed_kg"])
        for r in res.trace:
            w.writerow([r.week, r.week_label, r.standing_biomass_kg, r.biomass_cap,
                        r.feed_kg_day, r.feed_cap, r.harvested_kg, r.required_kg,
                        r.binding, r.legal, r.over_biomass_kg, r.over_feed_kg])
    print(f"  Wrote {stem}_envelope.csv and {stem}_trace.csv")


if __name__ == "__main__":
    raise SystemExit(main())
