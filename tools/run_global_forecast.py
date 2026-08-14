"""ENTRY POINT: produce the STANDARD forecast workbook from the GLOBAL method.

METHOD: GLOBAL (precalculated L1->L3)
====================================

This is a NEW entry point — it does NOT touch `forecast/run.py`. It hydrates the
PR exactly like the loop runners, runs the converged L1<->L3 global planner
(`forecast.global_planner_loop_poc.run_loop`) with the CORRECT whole-facility
models (6N purge hold + FW+OG+purge counted vs the cap), adapts the result into
the structures the SHARED `forecast.excel_io` writers consume
(`forecast.global_forecast.build_tables`), and emits a standard workbook with a
clear METHOD STAMP and a "_GLOBAL" filename suffix.

The production pipeline (`forecast.run.main`) stays byte-identical; this runner
imports the writers but adds no branch to run.py.

Sheets emitted
--------------
  RunConfig (method stamp), HarvestPlan, HarvestReport, Batch Plan,
  FeedForecastWeekly, FeedForecastMonthly, Advisory, WeeklyReport,
  MonthlyReport, ReconciliationReport (L1 conservation), StandingTrace (L1).
SPECIFIC-TANK PICK (step #2 — NOW REAL, `forecast.global_tank_pick_poc`):
  BatchLocations (real per-physical-tank occupancy, continuity-preserving),
  FacilityMap (real physical tank grid), TransferPlan (real tank-to-tank moves;
  6N sixn pair round-robin), and TankContinuityAudit (proves 0 TANK_DRIFT /
  0 BIO_DRIFT over the emitted locations + transfers + harvest events). The known
  1-week structural over-subscription is double-stacked + flagged, not dropped.

Usage:
    python -m tools.run_global_forecast
    python -m tools.run_global_forecast --workbook Forecast.xlsm --out out.xlsx
    python -m tools.run_global_forecast --no-pr        # incoming-only
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.config_io import load_config
from forecast.scenario_io import load_batches, load_limits
from forecast import global_forecast as gf
from forecast import global_planner_loop_poc as loop
from forecast import config_snapshot as _cs
from tools.run_full_facility_poc import _hydrate_pr


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--scenario-dir", default=str(_ROOT / "scenario"))
    ap.add_argument("--workbook", default=str(_ROOT / "Forecast.xlsm"))
    ap.add_argument("--out", default=None,
                    help="output .xlsx path; default = <workbook stem>_GLOBAL.xlsx")
    ap.add_argument("--max-iterations", type=int, default=10)
    ap.add_argument("--margin-frac", type=float, default=0.5)
    ap.add_argument("--slack-epsilon", type=float, default=1000.0)
    ap.add_argument("--mip-time-limit", type=float, default=180.0)
    ap.add_argument("--mip-rel-gap", type=float, default=0.01)
    ap.add_argument("--no-pr", action="store_true")
    args = ap.parse_args()

    print("=" * 72)
    print(f"  METHOD: GLOBAL (precalculated L1->L3) — STANDARD WORKBOOK EXPORT")
    print("=" * 72)

    control, tables, facility = load_config(args.config_dir)
    batches = load_batches(args.scenario_dir)
    _facility_limits, system_limits = load_limits(args.scenario_dir)
    print(f"  Config:   {args.config_dir}")
    print(f"  Scenario: {len(batches)} batches from {args.scenario_dir}")

    inflight_og, fw_inflight, purge_inflight = {}, {}, {}
    if not args.no_pr:
        # Hydrate BOTH OG and FW in-flight units — model_full_facility (the
        # correct default) REQUIRES fw_inflight or it under-counts the FW phase.
        # purge_inflight = fish already in 6N at hand-over (primes the purge
        # pipeline so L1 doesn't spin a fresh backlog -> startup overshoot).
        inflight_og, fw_inflight, derived_start, purge_inflight = _hydrate_pr(
            Path(args.workbook), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
            print(f"  ForecastStart: derived {derived_start.date()} from PR closing +1d")
        print(f"  In-flight OG batches: {len(inflight_og)}; "
              f"FW-in-flight batches: {len(fw_inflight)}")

    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    print(f"  forecast_start={fs_date}, horizon={control.horizon_weeks}w")
    print(f"  facility caps: biomass<={control.max_biomass_kg:,.0f} kg, "
          f"feed<={control.max_feed_per_day_kg:,.0f} kg/day")

    # ---- Run the converged L1<->L3 loop with the CORRECT whole-facility models.
    print("\n  RUNNING L1<->L3 LOOP (model_purge_hold=True, model_full_facility=True)")
    result = loop.run_loop(
        batches, tables, control, facility, system_limits,
        inflight_og=inflight_og,
        max_iterations=args.max_iterations,
        margin_frac=args.margin_frac,
        l3_kwargs=dict(slack_epsilon=args.slack_epsilon,
                       mip_time_limit=args.mip_time_limit,
                       mip_rel_gap=args.mip_rel_gap, verbose=False),
        model_purge_hold=True,
        model_full_facility=True,
        fw_inflight=fw_inflight,
        purge_inflight=purge_inflight,
        verbose=True,
    )
    print(f"  converged={result.converged}, iterations={result.n_iterations}, "
          f"peak {result.peak_biomass_kg:,.0f} kg ({result.pct_of_cap_peak:.1f}% cap), "
          f"HOG {result.total_hog_kg:,.0f} kg")

    # ---- Adapt to excel_io-ready structures.
    gft = gf.build_tables(result, batches, tables, control, facility,
                          fw_inflight=fw_inflight)
    cons = gf.conservation_summary(gft)
    print(f"\n  CONSERVATION (facility): seeded {cons['seeded']:,.0f} == "
          f"harvested {cons['harvested']:,.0f} + standing {cons['standing']:,.0f} "
          f"+ mort {cons['mortality']:,.0f} + cull {cons['cull']:,.0f}")
    print(f"    accounted {cons['accounted']:,.0f}; residual {cons['residual']:,.3f} "
          f"({cons['residual_pct']:.4f}%); worst per-batch "
          f"{cons['worst_batch_residual_pct']:.4f}% "
          f"({'OK — conserves' if cons['worst_batch_residual_pct'] < 0.01 else 'CHECK'})")
    print(f"  SPECIFIC-TANK PICK (step #2): {gft.n_transfers} tank-to-tank "
          f"transfers; {len(gft.batch_locations)} tank-week rows; "
          f"over-subscribed weeks: {len(gft.oversub_weeks)} "
          f"({gft.n_oversub_rows} double-stacked rows)"
          + (f" {gft.oversub_weeks}" if gft.oversub_weeks else ""))

    # ---- Emit the standard workbook via the SHARED writers.
    out_path = (Path(args.out) if args.out
                else Path(args.workbook).with_name(
                    Path(args.workbook).stem + "_GLOBAL.xlsx"))
    _emit_workbook(gft, result, batches, tables, control, facility,
                   _facility_limits, cons, out_path, system_limits=system_limits)
    print(f"\n  Wrote {out_path}")
    return 0


def _apply_manual_window(input_path, scenario_dir, control, tables, facility,
                         batches, inflight_og, fw_inflight, purge_inflight):
    """Execute the operator's manual override window (scenario/manual_events.yaml)
    BEFORE the global plan, then CONVERT the resulting facility state into the
    Global engine's aggregated seeds — so the precalculated method plans forward
    from the SAME starting point the controller uses (the manual transfers /
    harvests + their biology), not the raw PR.

    The global engine consumes aggregated per-batch dicts, not a tank-level state,
    so we: (1) hydrate a tank-level FacilityState from the PR, (2) run
    advance_facility_window (events + biology for the manual weeks), (3) aggregate
    the post-window tanks back into inflight_og (non-6N OG) + purge_inflight (6N),
    and drop any FW batch manually moved to OG from fw_inflight. Returns
    (inflight_og, fw_inflight, purge_inflight, new_start_datetime, new_horizon,
    warnings, window_weeks). No manual events -> inputs unchanged (window_weeks=0).
    """
    from datetime import datetime as _dt, timedelta as _td
    from collections import defaultdict
    from forecast.manual_events import load_manual_events
    # PR-specific events: closing = forecast_start - 1 day (same contract
    # as forecast.run).
    events = load_manual_events(
        str(scenario_dir), pr_closing=control.forecast_start - _td(days=1))
    if not events:
        return (inflight_og, fw_inflight, purge_inflight,
                control.forecast_start, control.horizon_weeks, [], 0, None)
    from forecast.excel_io import load_workbook
    from forecast.production_report import (
        read_production_report, hydrate_facility_state)
    from forecast.state import FacilityState
    from forecast.manual_window import advance_facility_window
    from forecast.sixn import SIXN_ALL_TANKS
    wb = load_workbook(str(input_path))
    pr_closing, og_records, fw_records = read_production_report(wb)
    wb.close()
    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    state = FacilityState.from_facility_config(facility, today=fs_date)
    hydrate_facility_state(state, og_records, batches)
    # Snapshot the PR-hydrated OPENING (pre-window) as a second, un-advanced state
    # so the TankContinuityAudit can open the first manual week from the real
    # in-flight tanks instead of from 0 (else every opening batch reads as a
    # +drift on the manual window's week 1 — the +2.07M boundary artifact).
    init_state = FacilityState.from_facility_config(facility, today=fs_date)
    hydrate_facility_state(init_state, og_records, batches)
    bbid = {b.batch_id: b for b in batches}
    window_n = max((e.week or 1) for e in events)
    win = advance_facility_window(
        state, bbid, tables, fs_date, window_n, events=events, control=control,
        pr_closing=pr_closing, fw_records=fw_records)
    # `state` is now the manual window's CLOSE (last manual week) — the tank
    # positions the global plan must CONTINUE from (not re-stock), so the W27
    # hand-off reconciles as transfers rather than vanish+restock.
    # Aggregate the post-window tanks into the Global engine's seed dicts.
    og_agg = defaultdict(lambda: [0.0, 0.0])
    purge_agg = defaultdict(lambda: [0.0, 0.0])
    for t in state.tanks_by_id.values():
        if t.is_empty:
            continue
        tgt = purge_agg if t.tank_id in SIXN_ALL_TANKS else og_agg
        tgt[t.batch_id][0] += t.count
        tgt[t.batch_id][1] += t.biomass_kg
    cv = {b.batch_id: b.tran_og_cv for b in batches}
    new_og = {bid: (c, b * 1000.0 / c, cv.get(bid, 16.0))
              for bid, (c, b) in og_agg.items() if c > 0}
    new_purge = {bid: (c, b * 1000.0 / c)
                 for bid, (c, b) in purge_agg.items() if c > 0}
    # WINDOW SEMANTICS — the 6N contents at the window close carry their release
    # timing into L1 explicitly (scripted stagings release _PURGE_HOLD_WEEKS
    # after their scripted week; untouched PR-start fish from the handoff), so
    # the purge hold is honored from the handoff. See
    # manual_window.sixn_release_schedule.
    from forecast.manual_window import sixn_release_schedule
    purge_schedule = sixn_release_schedule(
        state, win.get("transfer_events", []), fs_date, window_n)
    # FW batches manually crossed to OG are now in new_og -> drop from fw_inflight.
    transferred = set(win.get("transferred_fw_batches", set()))
    new_fw = {bid: v for bid, v in fw_inflight.items() if bid not in transferred}
    ns = win["new_start"]
    new_start = _dt(ns.year, ns.month, ns.day)
    # STITCH bundle: the manual weeks' per-tank rows + events, read-compatible with
    # the Global writers (placement.BatchLocationRow == TankLocRow read shape;
    # events.Transfer/Harvest/TranOGEntry == TankTransfer/Harvest/TranOG). Prepended
    # to gft so the output timeline OPENS with the manual work, then the global plan.
    stitch = {
        "batch_locations": win.get("batch_locations", []),
        "transfer_events": win.get("transfer_events", []),
        "harvest_events": win.get("harvest_events", []),
        "tranog_events": win.get("tranog_events", []),
        "realized_biology": win.get("realized_biology", {}),
        "opening_locations": win.get("opening_locations", []),  # manual-week ledger opens
        "initial_state": init_state,   # W23 audit opening (pre-window PR tanks)
        "window_close_state": state,   # W27 pick continuation (manual close tanks)
        # Release timing for the window-close 6N contents (window semantics).
        "purge_release_schedule": purge_schedule,
    }
    return (new_og, new_fw, new_purge, new_start,
            control.horizon_weeks - window_n, win.get("warnings", []), window_n, stitch)


def _build_manual_week_states(opening_locations, realized_biology):
    """Synthesize per-(batch, manual-week) BatchWeekState rows so the WeeklyReport/
    MonthlyReport ledger reconciles the manual override window (W23-W26).

    fw_states cover only FW/EGG weeks and the pick's mort_states cover only the
    GLOBAL weeks (W27+), so the manual weeks' IN-FLIGHT OG batches would open from
    their close (a manual harvest then reads as -Count_Check) and their OG-phase
    mortality would be uncredited. Each synth row carries the start-of-week PR
    opening (opening_locations, a start-of-week PRE-biology snapshot) AND the
    realized weekly mortality (realized_biology mort_count) in one object, so the
    ledger's open-lift and mortality-credit stay coupled (fixing one without the
    other would flip a currently-cancelling row). stage='SW' + week_from_input=1
    keep the ledger on its realized-OG path (no FW feed/fallback, not a stocking
    week). Global-only: the base run has no manual window and never calls this.
    """
    from collections import defaultdict
    from forecast.models import BatchWeekState
    op = defaultdict(lambda: {"count": 0.0, "bio": 0.0, "ws": None})
    for r in opening_locations:
        e = op[(r.batch_id, r.week_label)]
        e["count"] += r.count
        e["bio"] += r.biomass_kg
        e["ws"] = r.week_start
    mort = defaultdict(float)
    for (_tid, wl, bid), (_bio_delta, mort_ct) in (realized_biology or {}).items():
        mort[(bid, wl)] += mort_ct
    out = []
    for (bid, wl), e in op.items():
        c = e["count"]
        if c <= 0:
            continue
        wt = e["bio"] * 1000.0 / c
        mc = mort.get((bid, wl), 0.0)
        out.append(BatchWeekState(
            batch_id=bid, week_label=wl, week_start=e["ws"],
            days_since_input=0, week_from_input=1,
            count=c, avg_weight_g=wt, biomass_kg=e["bio"],
            feed_kg_day=0.0, feed_kg_week=0.0, sgr_pct_day=0.0, fcr=0.0,
            stage="SW", feed_type="",
            mortality_pct_weekly=(100.0 * mc / c),
            open_count=c, open_avg_weight_g=wt, open_biomass_kg=e["bio"],
            mort_count_week=mc))
    return out


def run_global(input_path, output_path, config_dir, scenario_dir, *,
               no_pr: bool = False, overstock: bool = False,
               max_iterations: int = 10, margin_frac: float = 0.5,
               slack_epsilon: float = 1000.0, mip_time_limit: float = 180.0,
               mip_rel_gap: float = 0.01,
               optimal: bool = False, cpsat_time: float = 300.0,
               cpsat_workers: int = 8, cpsat_det_time: float = 30.0) -> int:
    """Produce the standard GLOBAL-method workbook at `output_path` from the PR at
    `input_path` + the app's config/scenario. Callable mirror of `main()` for the
    UI (parallel to `forecast.run.main`). Returns 0 on success. Touches no
    production file.

    `overstock` (DEFAULT OFF since 2026-08-12, operator's call) is the SELECTIVE
    over-stock lever: batches averaging under 2.5 kg placed to 100% of the HARD
    density cap (95 kg/m3) instead of the operating target every other method
    plans to. It was default ON, which broke comparison parity — the controller
    family plans to `density_target_pct` (a knob the tournament SEARCHES, 0.9
    live / 0.95 in the tuned winner) while this hardcoded 100% for a whole
    biomass class, so Global entered every comparison with capacity its rivals
    did not take, and part of its density/utilisation lead was bought with it.
    The operator's rule is 95 hard / target 85 "as much as possible", and no
    weight-based density exemption exists in it: the 2.5 kg threshold and the
    "light fish are safe to concentrate" claim were engineering inventions.
    Leaving it available as an explicit, off-by-default lever so a placement
    study can still measure it (tools/run_placement_optimize.py sweeps it) —
    but a comparison run must not turn it on for one method only.
    """
    from forecast import global_planner_l3_poc as _l3
    control, tables, facility = load_config(str(config_dir))
    batches = load_batches(str(scenario_dir))
    _facility_limits, system_limits = load_limits(str(scenario_dir))
    inflight_og, fw_inflight, purge_inflight = {}, {}, {}
    _mw_stitch = None   # manual-window rows/events to prepend into the output
    _mw_weeks = 0       # manual-window length (0 = no window)
    _mw_warns = []      # manual-window lints (stays empty on the no_pr path)
    if not no_pr:
        inflight_og, fw_inflight, derived_start, purge_inflight = _hydrate_pr(
            Path(input_path), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
        # Manual override window: execute the operator's starting-state events
        # (manual transfers/harvests) + biology, then seed the global plan from the
        # resulting state — same starting point the controller honors. No events =>
        # unchanged. FW re-anchoring is approximate for now (fine when no fw_to_og).
        (inflight_og, fw_inflight, purge_inflight, _mw_start, _mw_horizon,
         _mw_warns, _mw_weeks, _mw_stitch) = _apply_manual_window(
            input_path, scenario_dir, control, tables, facility, batches,
            inflight_og, fw_inflight, purge_inflight)
        if _mw_weeks:
            control.forecast_start = _mw_start
            control.horizon_weeks = _mw_horizon
            print(f"  Manual override window: {_mw_weeks} week(s) executed before "
                  f"the global plan ({len(_mw_warns)} warning(s)); global now opens "
                  f"{_mw_start.date()}, horizon {_mw_horizon}w.")

    _prev = (_l3._OVERSTOCK_DENSITY_PCT, _l3._OVERSTOCK_MAX_WT_G)
    if overstock:
        _l3._OVERSTOCK_DENSITY_PCT, _l3._OVERSTOCK_MAX_WT_G = 1.0, 2500.0
        print("  overstock=True: selective over-stock active — <2.5 kg fish "
              "may be placed to the HARD density cap (100%). NOT COMPARABLE "
              "with the controller family, which plans to density_target_pct.")
    try:
        result = loop.run_loop(
            batches, tables, control, facility, system_limits,
            inflight_og=inflight_og, max_iterations=max_iterations,
            margin_frac=margin_frac,
            l3_kwargs=dict(slack_epsilon=slack_epsilon, mip_time_limit=mip_time_limit,
                           mip_rel_gap=mip_rel_gap, verbose=False),
            model_purge_hold=True, model_full_facility=True,
            fw_inflight=fw_inflight, purge_inflight=purge_inflight,
            # Window semantics: no implicit pre-start 6N staging after a manual
            # window; the window-close 6N contents carry explicit release timing.
            purge_release_schedule=(_mw_stitch or {}).get("purge_release_schedule"),
            manual_window_weeks=_mw_weeks,
            verbose=False)
        grow_q = None
        _engine_warns = []
        if optimal:
            grow_q, _cpsat_info = _solve_cpsat_q(
                result, facility, system_limits, control,
                cpsat_time, workers=cpsat_workers, det_time=cpsat_det_time,
                initial_tb=(_mw_stitch or {}).get("window_close_state"))
            _w = cpsat_degrade_warning(_cpsat_info)
            if _w:
                _engine_warns.append(_w)
                print(f"  !! {_w}")
        gft = gf.build_tables(result, batches, tables, control, facility,
                              fw_inflight=fw_inflight, grow_q_by_week=grow_q,
                              initial_tank_state=(_mw_stitch or {}).get(
                                  "window_close_state"))
        # Stitch the MANUAL WINDOW weeks into the output so the timeline OPENS with
        # the operator's starting-state work (transfers/harvests + biology), then the
        # global plan. The manual rows/events are read-compatible with the writers
        # (BatchLocationRow == TankLocRow; events.* == Tank* shapes), so BatchLocations
        # / HarvestPlan / TransferPlan all show the manual weeks first. Prepended (the
        # writers sort by week). This does NOT touch the ReconciliationReport — that
        # is L1's per-batch conservation of the GLOBAL plan (still residual ~0); the
        # manual window conserves in its own right (it is the controller's audited
        # advance). The per-tank TankContinuityAudit stays an advisory cross-check.
        if _mw_stitch:
            gft.batch_locations = (list(_mw_stitch["batch_locations"])
                                   + list(gft.batch_locations))
            gft.harvest_events = (list(_mw_stitch["harvest_events"])
                                  + list(gft.harvest_events))
            gft.transfer_events = (list(_mw_stitch["transfer_events"])
                                   + list(gft.transfer_events))
            gft.tranog_events = (list(_mw_stitch["tranog_events"])
                                 + list(gft.tranog_events))
            if isinstance(getattr(gft, "realized_biology", None), dict):
                gft.realized_biology.update(_mw_stitch["realized_biology"])
            print(f"  Stitched the manual override window into the output: "
                  f"{len(_mw_stitch['batch_locations'])} BatchLocations rows, "
                  f"{len(_mw_stitch['transfer_events'])} transfer + "
                  f"{len(_mw_stitch['harvest_events'])} harvest event(s).")
        # A batch the pick could not give ANY legal tank is ABSENT from the
        # plan even though L1's batch-level reconciliation still calls it
        # standing. Surface it as an ERROR row, never a silent gap.
        _engine_warns.extend(getattr(gft, "unplaced_warnings", []) or [])
        # Solver-health findings from L3 (non-reproducible solves, Pass B
        # fallbacks). A plan whose layout depended on machine load must say
        # so — otherwise every later A/B silently compares noise.
        _engine_warns.extend(list(_l3.SOLVER_WARNINGS))
        _engine_warns.extend(getattr(gft, "topology_warnings", []) or [])
        _engine_warns.extend(getattr(gft, "depuration_warnings", []) or [])
        _engine_warns.extend(placement_gap_warnings(gft))
        cons = gf.conservation_summary(gft)
        _mw_states = (_build_manual_week_states(
            _mw_stitch.get("opening_locations", []),
            _mw_stitch.get("realized_biology", {})) if _mw_stitch else [])
        _emit_workbook(gft, result, batches, tables, control, facility,
                       _facility_limits, cons, Path(output_path),
                       initial_state=(_mw_stitch or {}).get("initial_state"),
                       manual_week_states=_mw_states, system_limits=system_limits,
                       manual_warnings=list(_mw_warns) + _engine_warns)
    finally:
        _l3._OVERSTOCK_DENSITY_PCT, _l3._OVERSTOCK_MAX_WT_G = _prev
    return 0


def placement_gap_warnings(gft, tol_kg: float = 1000.0) -> list:
    """How much of L1's standing the tank pick could NOT put in a physical tank.

    THE CROSS-CHECK THAT WAS MISSING. Advisory / SystemLimitsAudit are both
    written FROM `batch_locations`, so comparing them to BatchLocations is
    circular: it proves the arithmetic of what WAS placed and says nothing about
    what was left out. `StandingTrace` is L1's own facility standing, computed
    before any tank exists, so L1(OG + 6N purge) - sum(BatchLocations) is the
    first non-circular measure of placement COMPLETENESS.

    It is not a rounding check. Measured on the operator's 7.29 PR, global-lp
    carried 18,900,608 kg-weeks of standing that never reached a tank — fish
    that L1 planned, that the batch-level ReconciliationReport reported as
    standing and conserving, and that the per-tank TankContinuityAudit could not
    see (a batch with no tank has nothing to reconcile). Every existing gate was
    blind to it. Reported per week and in total; `tol_kg` ignores per-week
    rounding dust.
    """
    trace = list(getattr(gft, "trace", None) or [])
    locs = list(getattr(gft, "batch_locations", None) or [])
    if not trace or not locs:
        return []
    placed: dict = {}
    for r in locs:
        wl = getattr(r, "week_label", None)
        placed[wl] = placed.get(wl, 0.0) + (getattr(r, "biomass_kg", 0.0) or 0.0)
    gaps = []
    total = 0.0
    for t in trace:
        wl = getattr(t, "week_label", None)
        want = ((getattr(t, "og_biomass_kg", 0.0) or 0.0)
                + (getattr(t, "purge_biomass_kg", 0.0) or 0.0))
        if want <= 0:
            continue
        gap = want - placed.get(wl, 0.0)
        if gap > tol_kg:
            gaps.append((wl, gap, want))
            total += gap
    if not gaps:
        return []
    worst = sorted(gaps, key=lambda g: -g[1])[:3]
    return [(
        f"PLACEMENT GAP - {len(gaps)} week(s) carry L1 standing that reached NO "
        f"physical tank: {total:,.0f} kg-weeks in total, worst "
        + ", ".join(f"{w} {g:,.0f} kg ({100.0 * g / v:.0f}% of that week)"
                    for w, g, v in worst)
        + ". These fish are in L1's plan and in the batch-level reconciliation "
        "but are absent from BatchLocations, so they load no tank, eat no feed "
        "and appear in no density or per-system metric. Advisory and "
        "SystemLimitsAudit are written FROM BatchLocations and therefore cannot "
        "show this."
    )]


def cpsat_degrade_warning(info) -> str:
    """The ValidationLog line for a CP-SAT placement that gave up on part of the
    horizon, or "" when every week solved.

    A week CP-SAT cannot place comes back EMPTY (`q_by_w[w] = {}`) and is then
    laid out by the tank pick's FALLBACK, which enforces no per-tank density
    cap. Until 2026-08-11 that degrade existed only as a stdout line, so a run
    whose placement failed on 103 of 127 weeks still reached the compare board
    labelled "Global - CP-SAT optimal" with a PASS gate and a peak density of
    689.9 kg/m3 against a 95 cap. It must be recorded where the graders and the
    operator look.

    The text deliberately carries NO ISO week label and neither "MANUAL EVENT"
    nor "MANUAL WINDOW", so `forecast.window_weeks.manual_window_weeks` can
    never mistake it for an operator-scripted window row — that would EXCLUDE
    planner weeks from the harvest-compliance gates, hiding breaches in exactly
    the run that degraded."""
    n_inf = int((info or {}).get("n_infeasible", 0) or 0)
    if n_inf <= 0:
        return ""
    n_wk = int((info or {}).get("n_weeks", 0) or 0)
    pct = (100.0 * n_inf / n_wk) if n_wk else 0.0
    _sts = (info or {}).get("unplaced_status") or {}
    _why = (f" Solver verdicts: {_sts}." if _sts else "")
    return (f"PLACEMENT DEGRADED - CP-SAT could not place {n_inf} of {n_wk} "
            f"week(s) ({pct:.0f}% of the horizon). Those weeks were laid out by "
            f"the tank-pick FALLBACK, which enforces no per-tank density cap - "
            f"their per-tank densities and per-system loads are NOT "
            f"solver-verified and may exceed the cap. This run is NOT an "
            f"optimal placement." + _why)


def _solve_cpsat_q(result, facility, system_limits, control, time_limit,
                   workers: int = 8, det_time: float = 30.0, initial_tb=None):
    """Run the CP-SAT placement on L1's standing and return
    ({week: {(batch, tank): kg}}, info) for the grow-out layout.

    NOT full-horizon and NOT 0-swap (both claims corrected 2026-08-14): the
    callee is `solve_cpsat_perweek`, which solves ONE model per week seeded by
    last week's occupancy, and in it same-week swaps are a soft objective term
    (`+ 3 * sum(tr_swap)`) — the cheapest term in the objective — not a
    constraint. The hard 0-swap formulations in that module (full-horizon /
    rolling-window) are NOT what this path calls.

    `info` carries the solver's own self-report — crucially `n_infeasible`, the
    number of weeks CP-SAT could NOT place. Those weeks come back EMPTY
    (global_placement_milp_poc: `q_by_w[w] = {}`) and silently fall through to
    the tank pick's uncapped fallback, so the caller MUST surface them: a run
    with infeasible weeks is not an optimal placement and must never be graded
    as one (see run_global)."""
    from collections import defaultdict
    from forecast.global_placement_milp_poc import solve_cpsat_perweek
    from forecast.sixn import SIXN_MAIN_TANKS
    # Grow-out placement pool = the 11 nursery+grow-out systems PLUS OG6N's 3 MAIN
    # tanks (61/63/65); the per-week solver adds those mains only in production-mode
    # weeks (33->36) and leaves them out in purge weeks. OG6N sisters (67/69/71)
    # are never production, so they are excluded entirely here.
    def _in_pool(t):
        return t.type == "OG" and (t.system_id != "OG6N"
                                   or t.tank_id in SIXN_MAIN_TANKS)
    og = {t.tank_id: t.system_id for t in facility.tanks if _in_pool(t)}
    tvol = {t.tank_id: t.max_density_kg_m3 * t.volume_m3
            for t in facility.tanks if _in_pool(t)}
    vol = {t.tank_id: t.volume_m3 for t in facility.tanks if _in_pool(t)}
    by_week, wl_of = defaultdict(dict), {}
    for r in result.final_l1.batch_standing:
        if getattr(r, "in_purge", False) or r.biomass_kg <= 1e-9:
            continue
        # SUM a batch's multiple non-purge rows for one week (on-feed population +
        # off-feed in-place hold on grading weeks) instead of overwriting — else
        # CP-SAT sizes the batch on ONE row's biomass while the specific-tank pick
        # applies the SUM, cramming the full biomass into too few tanks (the
        # single-tank / 3x-per-system-cap collapse). Blend avg_wt count-consistent.
        cur = by_week[r.week].get(r.batch_id)
        if cur is None:
            by_week[r.week][r.batch_id] = (r.biomass_kg, r.feed_kg_day, r.avg_wt_g)
        else:
            b0, f0, a0 = cur
            c0 = (b0 * 1000.0 / a0) if a0 > 1e-9 else 0.0
            c1 = (r.biomass_kg * 1000.0 / r.avg_wt_g) if r.avg_wt_g > 1e-9 else 0.0
            b = b0 + r.biomass_kg
            c = c0 + c1
            by_week[r.week][r.batch_id] = (
                b, f0 + r.feed_kg_day, (b * 1000.0 / c) if c > 1e-9 else a0)
        wl_of[r.week] = r.week_label
    # Deterministic budget (det_time) is the binding stop criterion; the passed
    # wall-clock time_limit is only a per-week safety cap (was hardcoded to 10.0,
    # which ignored the caller's cpsat_time AND left the ~2.6% gap open).
    # Seed the solver with where the fish ACTUALLY are at the handoff, so R6
    # (">= 1 kg fish MAY remain in entry-tier tanks") can apply from week 0.
    _tanks = getattr(initial_tb, "tanks_by_id", None) or {}
    _init_tb = {int(t): st.batch_id for t, st in _tanks.items()
                if getattr(st, "batch_id", None)}
    q, info = solve_cpsat_perweek(by_week, og, tvol, vol, wl_of, system_limits,
                                  control, det_time=float(det_time),
                                  time_limit=float(time_limit),
                                  workers=int(workers), verbose=True,
                                  initial_tb=_init_tb)
    print(f"  [CP-SAT per-week placement] worst_gap={info['worst_gap']*100:.2f}% "
          f"infeasible={info['n_infeasible']} slack={info.get('slack_kg'):,.0f} kg "
          f"solve={info.get('solve_s'):.0f}s")
    return q, info


def _emit_workbook(gft, result, batches, tables, control, facility,
                   facility_limits, cons, out_path: Path,
                   initial_state=None, manual_week_states=None,
                   system_limits=None, manual_warnings=None) -> None:
    from openpyxl import Workbook
    from forecast.excel_io import (
        write_advisory, write_batch_locations, write_batch_plan,
        write_facility_map, write_feed_forecast_monthly,
        write_feed_forecast_weekly, write_harvest_plan_output,
        write_harvest_plan_report, write_harvest_report,
        write_input_conservation_audit, write_monthly_report,
        write_system_limits_audit, write_tank_continuity_audit,
        write_transfer_plan_output, write_validation_log,
        write_weekly_report,
    )

    batch_by_id = {b.batch_id: b for b in batches}
    fs_date = gft.forecast_start
    hog = control.default_hog_yield
    bl = gft.batch_locations
    hv = gft.harvest_events
    fw_states_by_batch: dict[str, list] = {}
    for s in gft.fw_states:
        fw_states_by_batch.setdefault(s.batch_id, []).append(s)

    # ---- TankContinuityAudit: build the sheet, then scan it for drift flags ----
    # so the RunConfig stamp can report the drift counts (must be 0/0). The audit
    # reconciles the REAL BatchLocations + Transfers + Harvest events; the
    # specific-tank pick supplies `realized_biology` (per-tank net growth/mort)
    # and `mort_states` (per-(batch, week) weekly mortality %) so both the count
    # and biomass balances close exactly. No PR initial tank state is passed
    # (the global pick stocks every batch from empty), so first-week opens are 0.
    _audit_wb = Workbook()
    write_tank_continuity_audit(
        _audit_wb, bl, gft.mort_states, hv, gft.transfer_events,
        grade_events=[], tranog_events=gft.tranog_events,
        initial_state=initial_state,
        realized_biology=gft.realized_biology)
    n_tank_drift, n_bio_drift, fac_count_signed, fac_count_abs = \
        _scan_audit_drift(_audit_wb["TankContinuityAudit"])
    _audit_wb.close()
    # NOTE: the authoritative conservation proof for the GLOBAL (LP) method is the
    # ReconciliationReport (batch-level seeded == harvested + standing + mort + cull;
    # see conservation_summary). This per-tank TankContinuityAudit reconciles the
    # CONTROLLER's tank-to-tank Transfer/Harvest EVENT stream, which the LP placement
    # does not fully emit (it re-solves the whole-facility layout each week rather
    # than moving fish tank-by-tank) — so it can OVER-REPORT drift on LP output even
    # when every fish is accounted for. It is a cross-check, not the source of truth.
    _recon_ok = abs(cons.get("residual_pct", 0.0)) < 0.01
    # The per-tank audit is a REAL proof: with the specific-tank pick's complete
    # event stream (stocking, transfers, harvests, per-tank mortality) it must
    # reconcile open+events==close for every (tank, week). 0 TANK_DRIFT means
    # every fish is accounted per tank; nonzero is a genuine gap to INVESTIGATE,
    # not an artifact. (BIO_DRIFT is the kg cross-check; batch-level conservation
    # is separately proven by the ReconciliationReport.)
    _tank_clean = (n_tank_drift == 0 and abs(fac_count_signed) < 1.0)
    _verdict = ("PASS — every fish accounted per tank, per week"
                if _tank_clean else
                "INVESTIGATE — per-tank count flow does not reconcile")
    tank_drift_note = (
        f"TankContinuityAudit (per-tank cross-check): {_verdict}. "
        f"TANK_DRIFT={n_tank_drift}, BIO_DRIFT={n_bio_drift}, facility count "
        f"signed/abs {fac_count_signed:.0f}/{fac_count_abs:.0f}. Batch-level "
        f"ReconciliationReport residual {cons.get('residual_pct', 0.0):+.4f}% "
        f"({'CONSERVES' if _recon_ok else 'CHECK RECON'}).")
    print("  " + tank_drift_note)

    wb = Workbook()

    # ---- RunConfig: the METHOD STAMP + model flags + conservation. ----
    ws = wb.active
    ws.title = "RunConfig"
    # A1 is the sheet-KIND discriminator (forecast.config_snapshot.run_config_kind):
    # this is a method STAMP — a record of what ran — not the controller's
    # re-importable YAML snapshot that shares the sheet name. Keep the prefix in
    # sync with config_snapshot.KIND_STAMP_MARK (a test pins the pair).
    ws["A1"] = _cs.KIND_STAMP_MARK
    rows = [
        ("(sheet kind)", "method stamp — a record of what ran; NOT an "
                         "importable config snapshot"),
        ("planning_method", gf.METHOD_STAMP),
        ("model_purge_hold", "True (2-week off-feed 6N purge depuration flow)"),
        ("model_full_facility", "True (FW + OG + 6N purge counted vs the cap)"),
        ("fw_inflight", "hydrated from PR (FW-phase not under-counted)"),
        ("forecast_start", str(fs_date)),
        ("horizon_weeks", control.horizon_weeks),
        ("biomass_cap_kg", control.max_biomass_kg),
        ("feed_cap_kg_day", control.max_feed_per_day_kg),
        ("loop_converged", result.converged),
        ("loop_iterations", result.n_iterations),
        ("peak_facility_biomass_kg", round(result.peak_biomass_kg, 0)),
        ("peak_pct_of_cap", f"{result.pct_of_cap_peak:.1f}%"),
        ("total_HOG_kg", round(result.total_hog_kg, 0)),
        ("CONSERVATION seeded", round(cons["seeded"], 0)),
        ("  = harvested", round(cons["harvested"], 0)),
        ("  + standing (incl FW + 6N hold)", round(cons["standing"], 0)),
        ("  + mortality", round(cons["mortality"], 0)),
        ("  + cull", round(cons["cull"], 0)),
        ("  residual_pct", f"{cons['residual_pct']:.4f}%"),
        ("SPECIFIC-TANK PICK", gf.TANKPICK_STAMP),
        ("tank_to_tank_transfers", gft.n_transfers),
        ("tank_week_rows", len(gft.batch_locations)),
        ("over_subscribed_weeks",
         f"{len(gft.oversub_weeks)} (double-stacked {gft.n_oversub_rows} rows): "
         + (", ".join(gft.oversub_weeks) if gft.oversub_weeks else "none")),
        ("tank_continuity", tank_drift_note),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 56

    # ---- FULL sheets (the L1/L3 layers support these). ----
    write_harvest_plan_output(wb, hv, default_hog_yield=hog,
                              facility_limits_hog={})
    write_harvest_plan_report(wb, hv, scenario_name=control.scenario_name,
                              default_hog_yield=hog, facility_limits_hog={},
                              forecast_start=fs_date)
    write_harvest_report(wb, hv, default_hog_yield=hog,
                         facility_limits_hog={}, forecast_start=fs_date)
    write_batch_plan(wb, bl, hv, default_hog_yield=hog)
    write_feed_forecast_weekly(wb, bl, fw_states_by_batch, fs_date, tables,
                               batch_by_id)
    write_feed_forecast_monthly(wb, bl, fw_states_by_batch, fs_date, tables,
                                batch_by_id)
    write_advisory(wb, bl, hv, facility_limits, control,
                   batches=batch_by_id, tables=tables)
    # Manual-window synth states FIRST so a genuine fw_state on a colliding
    # (batch, week) overwrites them (last-write-wins in the ledger's bio_state /
    # mortpct); base run has manual_week_states=[] -> byte-identical to fw_states.
    _bws = list(manual_week_states or []) + list(gft.fw_states)
    write_weekly_report(wb, bl, hv, _bws,
                        batches=batch_by_id, tables=tables,
                        scenario_name=control.scenario_name, hog_yield=hog,
                        tranog_events=gft.tranog_events,
                        og_mort_states=gft.mort_states)
    write_monthly_report(wb, bl, hv, _bws,
                         batches=batch_by_id, tables=tables,
                         scenario_name=control.scenario_name, hog_yield=hog,
                         forecast_start=fs_date,
                         tranog_events=gft.tranog_events,
                         og_mort_states=gft.mort_states)

    # ---- L1-native StandingTrace + ReconciliationReport (the conservation). ----
    _write_standing_trace(wb, gft, control)
    _write_reconciliation(wb, gft, cons)

    # ---- REAL specific-tank sheets (step #2): physical tanks + transfers. ----
    write_facility_map(wb, bl, facility, batches=batch_by_id, tables=tables)
    write_batch_locations(wb, bl)
    write_transfer_plan_output(wb, gft.transfer_events,
                               tranog_events=gft.tranog_events, grade_events=[])
    # Per-batch input conservation (every stocked batch has a realized fate) +
    # per-(week, system) realized biomass/feed vs the system caps. These two
    # audits are written for the CONTROLLER too; emitting them here gives the
    # GLOBAL workbook the SAME sheet set, so the cross-method RunComparison scorer
    # (forecast.optimize.metrics_from_workbook) reads real per-system + transfer
    # figures instead of silently scoring zeros for the missing sheets.
    write_input_conservation_audit(wb, batches, bl, hv, control,
                                   tranog_events=gft.tranog_events)
    if system_limits is not None:
        write_system_limits_audit(wb, bl, batch_by_id, tables,
                                  system_limits, control)
    # The REAL continuity audit over the emitted BatchLocations + Transfers +
    # TranOG + Harvest events (proves 0 TANK_DRIFT / 0 BIO_DRIFT).
    write_tank_continuity_audit(
        wb, bl, gft.mort_states, hv, gft.transfer_events,
        grade_events=[], tranog_events=gft.tranog_events,
        initial_state=initial_state,
        realized_biology=gft.realized_biology)
    # Manual-window narration (MANUAL EVENT OK / MANUAL WINDOW lints): the
    # workbook must SELF-DESCRIBE its operator-scripted window weeks — the
    # cross-method scorer (forecast.window_weeks.manual_window_weeks) reads
    # them back to exclude window weeks from harvest-compliance metrics, same
    # as the controller's workbook. Only written when a window ran, so
    # no-window global outputs keep their exact previous sheet set.
    if manual_warnings:
        write_validation_log(wb, invariant_warnings=list(manual_warnings))

    # Order: RunConfig first.
    wb.move_sheet("RunConfig", -(wb.sheetnames.index("RunConfig")))
    wb.save(out_path)
    wb.close()


def _scan_audit_drift(ws):
    """Scan a written TankContinuityAudit sheet for TANK_DRIFT / BIO_DRIFT flags
    + the facility count signed/abs totals. Columns (1-based): Flag=15,
    Bio_Flag=28; the facility summary row 'Count (fish)' carries signed/abs.

    The facility totals start as NaN, not 0.0: a summary row that is missing
    or unparseable must FAIL the caller's `abs(fac_signed) < 1.0` cleanliness
    test (NaN comparisons are False), never default to the passing value — a
    verdict read off a row that was never read is the forbidden class."""
    n_tank = n_bio = 0
    fac_signed = fac_abs = float("nan")
    fac_row_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if len(row) >= 28:
            if row[14] == "TANK_DRIFT":
                n_tank += 1
            if row[27] == "BIO_DRIFT":
                n_bio += 1
        if row and row[0] == "Count (fish)" and len(row) >= 3:
            fac_row_seen = True
            try:
                fac_signed = float(row[1] or 0.0)
                fac_abs = float(row[2] or 0.0)
            except (TypeError, ValueError):
                print("  WARN: TankContinuityAudit facility 'Count (fish)' row "
                      "is unparseable — verdict reads INVESTIGATE, not PASS")
    if not fac_row_seen:
        print("  WARN: TankContinuityAudit facility 'Count (fish)' summary row "
              "not found — verdict reads INVESTIGATE, not PASS")
    return n_tank, n_bio, fac_signed, fac_abs


def _write_standing_trace(wb, gft, control) -> None:
    ws = wb.create_sheet("StandingTrace")
    ws.append(["FACILITY STANDING TRACE (L1; TRUE total = FW + OG + 6N purge)"])
    ws.append([gf.METHOD_STAMP])
    ws.append([])
    ws.append(["Week", "Week_Label", "Standing_TOTAL (kg)", "Biomass_Cap (kg)",
               "Feed (kg/day)", "Feed_Cap (kg/day)", "FW (kg)", "OG (kg)",
               "Purge_6N (kg)", "Harvested (kg)", "Legal", "Binding"])
    for r in gft.trace:
        ws.append([r.week, r.week_label, round(r.standing_biomass_kg, 0),
                   round(r.biomass_cap, 0), round(r.feed_kg_day, 0),
                   round(r.feed_cap, 0), round(r.fw_biomass_kg, 0),
                   round(r.og_biomass_kg, 0), round(r.purge_biomass_kg, 0),
                   round(r.harvested_kg, 0), r.legal, r.binding])
    for c, w in {1: 6, 2: 11, 3: 19, 4: 16, 5: 14, 6: 17, 7: 12, 8: 12,
                 9: 14, 10: 14, 11: 7, 12: 10}.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_reconciliation(wb, gft, cons) -> None:
    ws = wb.create_sheet("ReconciliationReport")
    ws.append(["RECONCILIATION REPORT (L1 conservation; FW counted in the total)"])
    ws.append([gf.METHOD_STAMP])
    ws.append(["Per batch: seeded == harvested + standing + mortality + cull. "
               "Standing@horizon folds any fish still in the 6N purge hold."])
    ws.append([])
    ws.append(["Batch", "Seeded", "Harvested", "Standing", "Mortality", "Cull",
               "Accounted", "Residual", "Residual_%"])
    for bid in sorted(gft.conservation):
        c = gft.conservation[bid]
        ws.append([bid, round(c["seeded_count"], 0),
                   round(c["harvested_count"], 0), round(c["standing_count"], 0),
                   round(c["mortality_count"], 0), round(c["cull_count"], 0),
                   round(c["accounted_count"], 0), round(c["residual_count"], 3),
                   round(c["residual_pct"], 4)])
    ws.append([])
    ws.append(["FACILITY", round(cons["seeded"], 0), round(cons["harvested"], 0),
               round(cons["standing"], 0), round(cons["mortality"], 0),
               round(cons["cull"], 0), round(cons["accounted"], 0),
               round(cons["residual"], 3), round(cons["residual_pct"], 4)])
    from openpyxl.utils import get_column_letter
    for c in range(1, 10):
        ws.column_dimensions[get_column_letter(c)].width = 12


if __name__ == "__main__":
    raise SystemExit(main())
