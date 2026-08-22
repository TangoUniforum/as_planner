"""Safety net for changes to the Global engine: prove no fish were destroyed.

    python tools/check_global_invariants.py out.xlsx
    python tools/check_global_invariants.py out.xlsx --save-baseline
    python tools/check_global_invariants.py out.xlsx --baseline tools/global_baseline.json

WHY THIS EXISTS
---------------
The Global tank picker has a KNOWN way to lose fish, and it has already
happened once. From global_tank_pick_poc.py's own notes: enforcing the
6N-only harvest rule at the draw, without first reconciling L1's
`harvest_by_bw` (harvest demand, keyed by week LABEL) against L1's separate
`purge_rows` (6N parking, keyed by INTEGER week), took TANK_DRIFT from 0 to 3
and made the facility lose 39,077 fish -- exactly the fish that used to
overflow into production tanks. They did not stay in their tanks; they
VANISHED, because `new_state` was built assuming L1's harvest executes, so a
fish the draw declines to take has no allocated home.

A plan that silently destroys 39,000 fish still LOOKS fine: the harvest
series is smooth, the density chart is calmer, and the 6N-rule probe reads a
reassuring 100%. That is precisely why this check is mechanical rather than a
matter of reading the output and forming an impression.

WHAT IS A HARD FAILURE (exit 1)
-------------------------------
  * any COUNT drift in TankContinuityAudit -- fish appearing or disappearing
    inside a tank-week
  * any Fish_At_Risk in InputConservationAudit
  * total harvested fish falling more than --harvest-tol below the baseline
    (a fix that "improves" the plan by quietly harvesting fewer fish is the
    exact failure this guards)

WHAT IS REPORTED BUT NOT FAILED
-------------------------------
  * BIO_DRIFT rows (a biomass expectation-formula residual; the count side is
    what proves conservation)
  * empty harvest weeks and the 6N-rule overflow -- these are the things a fix
    is TRYING to improve, so they are shown as deltas, never as pass/fail.

Run it before a change to record the baseline, and after to compare.
"""
from __future__ import annotations

import argparse
import collections
import json
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl                                                   # noqa: E402

DEFAULT_BASELINE = _ROOT / "tools" / "global_baseline.json"
# Count drift is judged in whole fish: the audit balances exactly, so anything
# above rounding is a real loss, not a tolerance question.
COUNT_EPS = 1.0


def _rows(ws, header_first_cell):
    """Rows of a sheet as dicts, keyed off the row whose first cell matches."""
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if r and r[0] == header_first_cell:
            hdr = list(r)
            continue
        if hdr and r and r[0] is not None:
            yield dict(zip(hdr, r))


def measure(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = {"workbook": path.name}

        count_drift_rows = 0
        count_drift_fish = 0.0
        bio_drift_rows = 0
        tank_rows = 0
        if "TankContinuityAudit" in wb.sheetnames:
            for d in _rows(wb["TankContinuityAudit"], "Week"):
                tank_rows += 1
                delta = d.get("Delta")
                if isinstance(delta, (int, float)) and abs(delta) > COUNT_EPS:
                    count_drift_rows += 1
                    count_drift_fish += abs(float(delta))
                if d.get("Bio_Flag"):
                    bio_drift_rows += 1
        out.update(tank_week_rows=tank_rows,
                   count_drift_rows=count_drift_rows,
                   count_drift_fish=round(count_drift_fish, 2),
                   bio_drift_rows=bio_drift_rows)

        at_risk = 0.0
        statuses = collections.Counter()
        if "InputConservationAudit" in wb.sheetnames:
            for d in _rows(wb["InputConservationAudit"], "Batch"):
                statuses[str(d.get("Status"))] += 1
                v = d.get("Fish_At_Risk (fish)")
                if isinstance(v, (int, float)):
                    at_risk += float(v)
        out.update(fish_at_risk=round(at_risk, 2), batch_status=dict(statuses))

        weeks = set()
        harvest = collections.defaultdict(float)
        if "BatchLocations" in wb.sheetnames:
            for i, r in enumerate(
                    wb["BatchLocations"].iter_rows(max_col=2, values_only=True), 1):
                if i > 4 and r and r[0]:
                    weeks.add(str(r[0]))
        if "HarvestPlan" in wb.sheetnames:
            for i, r in enumerate(
                    wb["HarvestPlan"].iter_rows(max_col=4, values_only=True), 1):
                if i > 4 and r and r[0] and isinstance(r[3], (int, float)) and r[3] > 0:
                    harvest[str(r[0])] += float(r[3])
                    weeks.add(str(r[0]))
        empty = sorted(w for w in weeks if harvest.get(w, 0.0) <= 0.0)
        out.update(horizon_weeks=len(weeks),
                   harvest_fish=round(sum(harvest.values()), 2),
                   empty_weeks=len(empty),
                   empty_week_labels=empty[:12])
        return out
    finally:
        wb.close()


def compare(now: dict, base: dict, harvest_tol: float) -> list[str]:
    """Hard failures only. Improvements and neutral drifts are not failures."""
    fails = []
    if now["count_drift_rows"] > 0:
        fails.append(
            f"COUNT DRIFT: {now['count_drift_rows']} tank-week(s) do not "
            f"balance, {now['count_drift_fish']:,.0f} fish. Fish are being "
            f"created or destroyed inside a tank-week.")
    if now["fish_at_risk"] > 0:
        fails.append(f"FISH AT RISK: {now['fish_at_risk']:,.0f} in "
                     f"InputConservationAudit (baseline "
                     f"{base.get('fish_at_risk', 0):,.0f}).")
    if base:
        b = float(base.get("harvest_fish") or 0.0)
        n = float(now.get("harvest_fish") or 0.0)
        if b > 0 and n < b * (1.0 - harvest_tol):
            fails.append(
                f"HARVEST FELL: {n:,.0f} vs baseline {b:,.0f} "
                f"({100.0 * (n - b) / b:+.2f}%, tolerance -{harvest_tol * 100:.2f}%). "
                f"A change that harvests fewer fish is how the known "
                f"39,077-fish loss presented.")
    return fails


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--baseline", default=str(DEFAULT_BASELINE))
    ap.add_argument("--save-baseline", action="store_true")
    ap.add_argument("--harvest-tol", type=float, default=0.005,
                    help="fractional harvest drop tolerated (default 0.5%%)")
    args = ap.parse_args()

    wb_path = Path(args.workbook)
    if not wb_path.is_file():
        print(f"workbook not found: {wb_path}")
        return 2
    now = measure(wb_path)

    bpath = Path(args.baseline)
    if args.save_baseline:
        bpath.parent.mkdir(parents=True, exist_ok=True)
        bpath.write_text(json.dumps(now, indent=2), encoding="utf-8")
        print(f"baseline saved -> {bpath}")

    base = {}
    if bpath.is_file() and not args.save_baseline:
        try:
            base = json.loads(bpath.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            base = {}

    print(f"\n{'metric':>22}{'now':>16}{'baseline':>16}")
    print("-" * 54)
    for k in ("tank_week_rows", "count_drift_rows", "count_drift_fish",
              "bio_drift_rows", "fish_at_risk", "horizon_weeks",
              "harvest_fish", "empty_weeks"):
        b = base.get(k, "-")
        bs = f"{b:,.0f}" if isinstance(b, (int, float)) else str(b)
        print(f"{k:>22}{now[k]:>16,.0f}{bs:>16}")
    if now["empty_week_labels"]:
        print(f"\n  empty harvest weeks: {', '.join(now['empty_week_labels'])}")

    fails = compare(now, base, args.harvest_tol)
    if fails:
        print("\nHARD INVARIANT FAILURES:")
        for f in fails:
            print(f"  * {f}")
        return 1
    print("\nOK — no fish created or destroyed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
