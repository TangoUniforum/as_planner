"""Convert a legacy VBA forecast workbook's INPUT sheets into the Python
pipeline's config/scenario YAML, so the Python pipeline can be run on the
exact same scenario the VBA workbook describes and the two outputs compared.

What it reads from the workbook (openpyxl, data_only=True):
  - Control        -> ControlParams (the 9 VBA-owned knobs + a few mappings)
  - FacilityConfig -> FacilityConfig (tanks)
  - BatchRegistry  -> list[BatchInput]
  - Tables         -> BiologyTables (SGR / FCR / mortality / feed / culling)
  - SystemLimits   -> SystemLimits  (per-week per-system caps; usually empty)
  - FacilityLimits -> FacilityLimits (per-week facility overrides; usually empty)

What it writes (reusing the pipeline's own dump functions so the YAML format
is guaranteed correct — never hand-written):
  <out>/config/control.yaml  + biology.yaml + facility.yaml   (dump_config)
  <out>/scenario/batches.yaml + limits.yaml                   (dump_scenario)

Design notes / decisions baked in (see module docstring of the run script and
the task brief):
  * forecast_start is DERIVED from the ProductionReport at run time, so the
    Control "Forecast Start Date" is only a harmless seed.
  * "Target Biomass %" (e.g. 0.99) -> facility_biomass_deviation_pct = 1 - x.
  * "6N growth" == 'yes'/'no'       -> sixn_growth bool.
  * Every Python-only knob the VBA has no concept of (tran_og_default_tanks,
    density_target_pct, rebalance_*, harvest_*, placement_method,
    sixn_production_start, ...) is taken from the LIVE config/control.yaml so
    the run uses the same planner tuning as the live app; only the VBA-owned
    fields are overridden from the workbook.
  * FacilityConfig SystemID for OG tanks is normalized to the "OG<X>" form the
    Python expects. In this workbook it is already "OG1N".."OG6N", but a bare
    "<X>" (e.g. "1N") is also handled by prepending "OG".
  * tank_id is forced unique across the facility: the VBA TankID is used when
    it is unique, otherwise a sequential id is synthesized; location_id keeps
    the human label.
  * Text dates in BatchRegistry are parsed MONTH-first ("%m/%d/%Y"), matching
    the only date-text parser in the codebase (production_report.py) and the
    data's own chronology (B37 "2/6/2024" = 6 Feb 2024, which must precede its
    4 May 2024 TranSF). A day-first parse is tried only as a fallback when
    month-first is impossible (month > 12).

Idempotent + re-runnable: writing to a caller-supplied dir overwrites cleanly;
with no dir it mkdtemp's a fresh one. It never touches the live config/ or
scenario/ and never writes to the source workbook.
"""
from __future__ import annotations

import argparse
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

# Import the pipeline's own dataclasses + dump functions so the YAML is
# produced by the same serializers the live app uses.
import sys
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.caps import FacilityLimits, SystemLimits  # noqa: E402
from forecast.config_io import dump_config, load_control  # noqa: E402
from forecast.models import (  # noqa: E402
    BatchInput, BiologyTables, ControlParams, FacilityConfig, TankConfig,
)
from forecast.scenario_io import dump_scenario  # noqa: E402


# ---------------------------------------------------------------------------
# cell / value helpers
# ---------------------------------------------------------------------------

def _num(v) -> Optional[float]:
    """Coerce a cell to float, or None if blank/non-numeric."""
    if v is None:
        return None
    if isinstance(v, bool):  # guard: bools are ints in python
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(v) -> Optional[datetime]:
    """Parse a cell into a datetime (date at midnight), or None if blank.

    datetime / date objects pass through. Text is parsed MONTH-first
    ("%m/%d/%Y"), falling back to DAY-first only when month-first is
    impossible (so an unambiguous "13/2/2024" still works).
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    if s == "":
        return None
    # ISO first (handles '2024-02-06' text).
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            d = datetime.strptime(s, fmt)
            return datetime(d.year, d.month, d.day)
        except ValueError:
            pass
    # Slash/dash d-m-y ambiguity: month-first to match the codebase
    # (production_report.py uses %m/%d/%Y), day-first as fallback.
    seps = ("/", "-")
    for sep in seps:
        if sep in s:
            for fmt in (f"%m{sep}%d{sep}%Y", f"%d{sep}%m{sep}%Y",
                        f"%m{sep}%d{sep}%y", f"%d{sep}%m{sep}%y"):
                try:
                    d = datetime.strptime(s, fmt)
                    return datetime(d.year, d.month, d.day)
                except ValueError:
                    continue
    raise ValueError(f"Unparseable date cell: {v!r}")


def _norm(s) -> str:
    return "" if s is None else str(s).strip()


def _find_header_row(ws, required: list[str], search_rows: int = 12) -> int:
    """Return the 1-based row index whose cells contain all `required`
    header labels (case/space-insensitive). Raises if not found.
    """
    req = {r.lower() for r in required}
    for r in range(1, min(ws.max_row, search_rows) + 1):
        present = {
            _norm(ws.cell(r, c).value).lower()
            for c in range(1, ws.max_column + 1)
        }
        if req.issubset(present):
            return r
    raise ValueError(
        f"Sheet {ws.title!r}: header row with {required} not found in first "
        f"{search_rows} rows")


def _header_map(ws, hdr_row: int) -> dict[str, int]:
    """Map normalized header label -> 1-based column index for `hdr_row`."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        label = _norm(ws.cell(hdr_row, c).value)
        if label and label not in out:
            out[label] = c
    return out


def _col(hmap: dict[str, int], *names: str) -> Optional[int]:
    """First matching column index for any of `names` (case-insensitive)."""
    low = {k.lower(): v for k, v in hmap.items()}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    return None


# ---------------------------------------------------------------------------
# Control
# ---------------------------------------------------------------------------

def read_control(ws, *, live_config_dir: Path) -> ControlParams:
    """Build ControlParams: start from the LIVE control.yaml (so every
    Python-only planner knob matches the live app), then override only the
    fields the VBA Control sheet owns.
    """
    base = load_control(live_config_dir)  # ControlParams with all live knobs

    # key/value scan of cols A (key) / B (value).
    kv: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        key = _norm(ws.cell(r, 1).value)
        if not key:
            continue
        kv.setdefault(key.lower().rstrip(" :"), ws.cell(r, 2).value)

    def g(*keys):
        for k in keys:
            kk = k.lower().rstrip(" :")
            if kk in kv:
                return kv[kk]
        return None

    # VBA-owned scalar caps / settings.
    base.horizon_weeks = int(_num(g("forecast horizon (weeks)", "horizon")) or base.horizon_weeks)
    sc = g("scenario name")
    if sc:
        base.scenario_name = str(sc).strip()
    base.max_feed_per_day_kg = _num(g("max feed/day (kg)")) or base.max_feed_per_day_kg
    base.max_biomass_kg = _num(g("max biomass (kg)")) or base.max_biomass_kg
    base.max_harvest_per_week = _num(g("max harvest/week")) or base.max_harvest_per_week
    base.min_harvest_weight_g = _num(g("min harvest weight (g)")) or base.min_harvest_weight_g
    base.min_harvest_per_week = _num(g("min harvest/week")) or base.min_harvest_per_week
    base.min_tank_control = _num(g("min tank control")) or base.min_tank_control
    base.default_hog_yield = _num(g("default hog yield")) or base.default_hog_yield
    hm = _num(g("handling mortality"))
    if hm is not None:
        base.handling_mortality_pct = hm

    # Mapping: Target Biomass % (e.g. 0.99) -> deviation (1 - x).
    tbp = _num(g("target biomass %"))
    if tbp is not None:
        base.facility_biomass_deviation_pct = round(1.0 - tbp, 10)

    # Mapping: 6N growth yes/no -> bool.
    sixn = g("6n growth")
    if sixn is not None:
        s = str(sixn).strip().lower()
        base.sixn_growth = s in ("yes", "true", "y", "1")

    # forecast_start: leave as the live seed; run.py derives it from the PR.
    # (We intentionally do NOT copy Control "Forecast Start Date".)
    return base


# ---------------------------------------------------------------------------
# Facility
# ---------------------------------------------------------------------------

def read_facility(ws) -> FacilityConfig:
    hdr = _find_header_row(ws, ["SystemID", "TankID", "Volume_m3", "Type"])
    h = _header_map(ws, hdr)
    c_loc = _col(h, "LocationID")
    c_sys = _col(h, "SystemID")
    c_tank = _col(h, "TankID")
    c_vol = _col(h, "Volume_m3", "Volume_m³", "Volume")
    c_den = _col(h, "MaxDensity_kg/m3", "MaxDensity_kg/m³", "MaxDensity_kg", "MaxDensity")
    c_feed = _col(h, "MaxFeed_kg/day", "MaxFeed_kg", "MaxFeed")
    c_type = _col(h, "Type")

    tanks: list[TankConfig] = []
    used_ids: set[int] = set()
    next_seq = 1
    for r in range(hdr + 1, ws.max_row + 1):
        sysid = _norm(ws.cell(r, c_sys).value)
        typ = _norm(ws.cell(r, c_type).value).upper()
        vol = _num(ws.cell(r, c_vol).value)
        if not sysid or vol is None:
            # Skip blank rows / the Summary side-block (no SystemID/Volume).
            continue
        if typ not in ("FW", "OG"):
            # A row with a real SystemID + Volume is a tank even if its Type
            # cell is blank. The May workbooks leave the 6N depuration sisters
            # 67/69/71 Type-blank (template artifact) though they are real OG
            # purge tanks the VBA tool uses — dropping them halves the 6N
            # purge throughput and (in purge mode, 6N-only harvest) forces
            # incoming TranOG batches to be dropped. Infer Type from the
            # SystemID: an OG* system is an OG tank, otherwise FW.
            typ = "OG" if sysid.upper().startswith("OG") else "FW"
        # OG system-id normalization: ensure an "OG" prefix on OG systems.
        if typ == "OG" and not sysid.upper().startswith("OG"):
            sysid = "OG" + sysid
        # tank_id: prefer the VBA TankID when unique+integer, else sequential.
        raw_tank = _num(ws.cell(r, c_tank).value) if c_tank else None
        tank_id: Optional[int] = int(raw_tank) if raw_tank is not None else None
        if tank_id is None or tank_id in used_ids:
            while next_seq in used_ids:
                next_seq += 1
            tank_id = next_seq
        used_ids.add(tank_id)
        next_seq = max(next_seq, tank_id + 1)

        loc = _norm(ws.cell(r, c_loc).value) if c_loc else ""
        if not loc:
            loc = f"{sysid}-{tank_id}"
        tanks.append(TankConfig(
            location_id=loc,
            system_id=sysid,
            tank_id=tank_id,
            volume_m3=float(vol),
            max_density_kg_m3=float(_num(ws.cell(r, c_den).value) or 0.0) if c_den else 0.0,
            max_feed_kg_day=float(_num(ws.cell(r, c_feed).value) or 0.0) if c_feed else 0.0,
            type=typ,
        ))
    if not tanks:
        raise ValueError("FacilityConfig: no tank rows parsed")
    return FacilityConfig(tanks=tanks)


# ---------------------------------------------------------------------------
# Batches
# ---------------------------------------------------------------------------

def read_batches(ws) -> list[BatchInput]:
    hdr = _find_header_row(ws, ["Batch_ID", "Input_Date", "Input_Count"])
    h = _header_map(ws, hdr)
    c_id = _col(h, "Batch_ID")
    c_in = _col(h, "Input_Date")
    c_cnt = _col(h, "Input_Count")
    c_sf = _col(h, "TranSF_date", "TranSF_Date")
    c_og = _col(h, "TranOG_Date")
    c_ogc = _col(h, "TranOG_Count")
    c_cv = _col(h, "TranOG_CV")
    c_wt = _col(h, "TranOG_AvgWt")
    c_fcr = _col(h, "FCR_Model")
    c_sgr = _col(h, "SGR_Correction")
    c_fw = _col(h, "FW_Correction")
    c_notes = _col(h, "Notes")

    out: list[BatchInput] = []
    for r in range(hdr + 1, ws.max_row + 1):
        bid = _norm(ws.cell(r, c_id).value)
        if not bid:
            continue
        in_count = _num(ws.cell(r, c_cnt).value)
        og_count = _num(ws.cell(r, c_ogc).value) if c_ogc else None
        cv = _num(ws.cell(r, c_cv).value) if c_cv else None
        wt = _num(ws.cell(r, c_wt).value) if c_wt else None
        sgr = _num(ws.cell(r, c_sgr).value) if c_sgr else None
        fw = _num(ws.cell(r, c_fw).value) if c_fw else None
        fcr = _norm(ws.cell(r, c_fcr).value) if c_fcr else ""
        notes = _norm(ws.cell(r, c_notes).value) if c_notes else ""
        out.append(BatchInput(
            batch_id=bid,
            input_date=_parse_date(ws.cell(r, c_in).value),
            input_count=int(in_count or 0),
            tran_sf_date=_parse_date(ws.cell(r, c_sf).value) if c_sf else None,
            tran_og_date=_parse_date(ws.cell(r, c_og).value) if c_og else None,
            tran_og_count=int(og_count) if og_count is not None else None,
            tran_og_avg_wt_g=float(wt) if wt is not None else None,
            tran_og_cv=float(cv) if cv is not None else 16.0,
            fcr_model=fcr or "FCR_116_Quick",
            fw_correction=float(fw) if fw is not None else 1.0,
            sgr_correction=float(sgr) if sgr is not None else 1.0,
            notes=notes,
        ))
    if not out:
        raise ValueError("BatchRegistry: no batch rows parsed")
    return out


# ---------------------------------------------------------------------------
# Biology (Tables)
# ---------------------------------------------------------------------------

def _read_col_until_blank(ws, hdr_row: int, col: int) -> list:
    """Read a single column downward from hdr_row+1 until the first blank."""
    vals = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            break
        vals.append(v)
    return vals


def read_biology(ws) -> BiologyTables:
    hdr = _find_header_row(ws, ["Size", "SGR_FW", "SGR_SW"])
    h = _header_map(ws, hdr)
    c_size = _col(h, "Size")
    c_fw = _col(h, "SGR_FW")
    c_sw = _col(h, "SGR_SW")
    c_f121 = _col(h, "FCR_1.21")
    c_f118 = _col(h, "FCR_1.18")
    c_f116 = _col(h, "FCR_1.16")
    c_wfi = _col(h, "Week_From_Input")
    c_mort = _col(h, "Mortality_%", "Mortality_%", "Mortality")
    c_ft = _col(h, "Feed Type")
    c_ms = _col(h, "Max Size", "Max Size ")
    c_dsi = _col(h, "Days Since Input")
    c_cull = _col(h, "Culling %", "Culling%")

    size = [float(x) for x in _read_col_until_blank(ws, hdr, c_size)]
    n = len(size)

    def _aligned(col):
        """Read a per-size column; pad with None up to len(size)."""
        raw = []
        for r in range(hdr + 1, hdr + 1 + n):
            raw.append(_num(ws.cell(r, col).value))
        return raw

    sgr_fw = _aligned(c_fw)
    sgr_sw = _aligned(c_sw)
    f121 = [float(x) for x in _aligned(c_f121)]
    f118 = [float(x) for x in _aligned(c_f118)]
    f116 = [float(x) for x in _aligned(c_f116)]

    wfi = [int(x) for x in _read_col_until_blank(ws, hdr, c_wfi)]
    mort = [float(_num(x)) for x in _read_col_until_blank(ws, hdr, c_mort)]
    # truncate to common length (defensive)
    m = min(len(wfi), len(mort))
    wfi, mort = wfi[:m], mort[:m]

    feed_names = _read_col_until_blank(ws, hdr, c_ft)
    feed_max = _read_col_until_blank(ws, hdr, c_ms)
    fm = min(len(feed_names), len(feed_max))
    feed_types = [(float(_num(feed_max[i])), str(feed_names[i]).strip())
                  for i in range(fm)]

    dsi = _read_col_until_blank(ws, hdr, c_dsi)
    cull = _read_col_until_blank(ws, hdr, c_cull)
    cm = min(len(dsi), len(cull))
    culling = [(int(_num(dsi[i])), float(_num(cull[i]))) for i in range(cm)]

    return BiologyTables(
        sgr_size_g=size,
        sgr_fw_pct_day=sgr_fw,
        sgr_sw_pct_day=sgr_sw,
        fcr_size_g=list(size),
        fcr_by_model={"1.21": f121, "1.18": f118, "1.16": f116},
        mortality_week_from_input=wfi,
        mortality_pct_weekly=mort,
        feed_types=feed_types,
        culling=culling,
    )


# ---------------------------------------------------------------------------
# Limits (SystemLimits + FacilityLimits) — usually empty in this scenario
# ---------------------------------------------------------------------------

# VBA metric label -> Python metric token.
_METRIC_MAP = {
    "feed/day": "feed_per_day",
    "feed/day (kg/day)": "feed_per_day",
    "biomass": "biomass",
    "biomass (kg)": "biomass",
    "max harvest/week": "max_harvest_per_week",
    "min harvest/week": "min_harvest_per_week",
    "hog yield": "hog_yield",
}


def _week_label_for_col(ws, hdr_row: int, col: int) -> Optional[str]:
    """Resolve the ISO week label for a per-week data column.

    Prefers an explicit 'YYYY-Www' label sitting one row above the
    Metric header; otherwise derives it from a date in the header row.
    """
    from forecast.time_grid import label_for_date
    above = ws.cell(hdr_row - 1, col).value if hdr_row >= 2 else None
    if isinstance(above, str) and above.strip().count("-W") == 1:
        return above.strip()
    cell = ws.cell(hdr_row, col).value
    d = None
    try:
        d = _parse_date(cell)
    except ValueError:
        d = None
    if d is not None:
        return label_for_date(d.date())
    return None


def read_system_limits(ws) -> SystemLimits:
    try:
        hdr = _find_header_row(ws, ["System", "Metric"])
    except ValueError:
        return SystemLimits(caps={})
    h = _header_map(ws, hdr)
    c_sys = _col(h, "System")
    c_metric = _col(h, "Metric")
    first_data_col = max(c_sys, c_metric) + 1
    caps: dict[tuple[str, str, str], float] = {}
    for r in range(hdr + 1, ws.max_row + 1):
        sysid = _norm(ws.cell(r, c_sys).value)
        metric_raw = _norm(ws.cell(r, c_metric).value).lower()
        if not sysid or not metric_raw:
            continue
        metric = _METRIC_MAP.get(metric_raw)
        if metric is None:
            continue
        if not sysid.upper().startswith("OG") and sysid[:1].isdigit():
            sysid = "OG" + sysid
        for c in range(first_data_col, ws.max_column + 1):
            val = _num(ws.cell(r, c).value)
            if val is None or val <= 0:
                continue
            wk = _week_label_for_col(ws, hdr, c)
            if wk:
                caps[(wk, sysid, metric)] = float(val)
    return SystemLimits(caps=caps)


def read_facility_limits(ws) -> FacilityLimits:
    try:
        hdr = _find_header_row(ws, ["Facility", "Metric"])
    except ValueError:
        return FacilityLimits(overrides={})
    h = _header_map(ws, hdr)
    c_metric = _col(h, "Metric")
    c_fac = _col(h, "Facility")
    first_data_col = max(c_metric, c_fac or 1) + 1
    overrides: dict[tuple[str, str], float] = {}
    for r in range(hdr + 1, ws.max_row + 1):
        metric_raw = _norm(ws.cell(r, c_metric).value).lower()
        metric = _METRIC_MAP.get(metric_raw)
        if metric is None:
            continue
        for c in range(first_data_col, ws.max_column + 1):
            val = _num(ws.cell(r, c).value)
            if val is None or val <= 0:
                continue
            wk = _week_label_for_col(ws, hdr, c)
            if wk:
                overrides[(wk, metric)] = float(val)
    return FacilityLimits(overrides=overrides)


# ---------------------------------------------------------------------------
# Top-level convert
# ---------------------------------------------------------------------------

def convert(
    workbook_path: str | Path,
    out_dir: str | Path | None = None,
    *,
    live_config_dir: str | Path | None = None,
) -> dict:
    """Read the VBA workbook's input sheets and write config/scenario YAML.

    Returns a dict with the produced paths + the parsed dataclasses (for
    verification/printing).
    """
    wb_path = Path(workbook_path)
    root = Path(__file__).resolve().parent.parent
    live_cfg = Path(live_config_dir) if live_config_dir else root / "config"

    out = Path(out_dir) if out_dir else Path(tempfile.mkdtemp(prefix="vba_to_config_"))
    config_dir = out / "config"
    scenario_dir = out / "scenario"
    config_dir.mkdir(parents=True, exist_ok=True)
    scenario_dir.mkdir(parents=True, exist_ok=True)

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(wb_path, data_only=True)

    control = read_control(wb["Control"], live_config_dir=live_cfg)
    facility = read_facility(wb["FacilityConfig"])
    batches = read_batches(wb["BatchRegistry"])
    tables = read_biology(wb["Tables"])
    system_limits = (read_system_limits(wb["SystemLimits"])
                     if "SystemLimits" in wb.sheetnames else SystemLimits(caps={}))
    facility_limits = (read_facility_limits(wb["FacilityLimits"])
                       if "FacilityLimits" in wb.sheetnames else FacilityLimits(overrides={}))
    wb.close()

    dump_config(config_dir, control=control, tables=tables, facility=facility)
    dump_scenario(scenario_dir, batches=batches,
                  facility_limits=facility_limits, system_limits=system_limits)

    return {
        "out_dir": out,
        "config_dir": config_dir,
        "scenario_dir": scenario_dir,
        "control": control,
        "facility": facility,
        "batches": batches,
        "tables": tables,
        "system_limits": system_limits,
        "facility_limits": facility_limits,
    }


def _cli() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("workbook", help="Path to the VBA .xlsm workbook (input).")
    p.add_argument("--out-dir", default=None,
                   help="Directory to write config/ + scenario/ into. "
                        "Default: a fresh temp dir.")
    p.add_argument("--live-config-dir", default=None,
                   help="Live config dir to seed Python-only knobs from "
                        "(default: repo config/).")
    a = p.parse_args()
    res = convert(a.workbook, a.out_dir, live_config_dir=a.live_config_dir)
    print(f"config_dir   = {res['config_dir']}")
    print(f"scenario_dir = {res['scenario_dir']}")
    print(f"tanks={len(res['facility'].tanks)} batches={len(res['batches'])} "
          f"sgr_rows={len(res['tables'].sgr_size_g)} "
          f"system_caps={len(res['system_limits'].caps)} "
          f"facility_overrides={len(res['facility_limits'].overrides)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_cli())
