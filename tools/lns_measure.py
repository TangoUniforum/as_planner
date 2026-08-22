"""Measure LNS-placement vs greedy on the live config + Forecast.xlsm.

Runs the FULL pipeline under each `placement_method`, reads the resulting
workbook, and prints the conservation + hot-spot comparison. This is the
measure-or-revert harness for the LNS build: the engine must keep 0 drift /
0 dropped and never raise `system_peak` above greedy.

    python -m tools.lns_measure                 # greedy vs lns
    python -m tools.lns_measure --methods greedy lns --det

`--det` re-runs `lns` under PYTHONHASHSEED 0 and 1 (separate processes) and
asserts the BatchLocations density signature is identical (determinism gate).
"""
from __future__ import annotations

import argparse
import contextlib
import io
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "Forecast.xlsm"
CFG = ROOT / "config"
SCN = ROOT / "scenario"


def _run_method(method: str):
    """Run the pipeline with placement_method=method; return (rc, metrics, dropped, overprod, out_path)."""
    import forecast.run as run_mod
    from forecast import optimize

    cdir = optimize.config_dir_with_overrides(str(CFG), {"placement_method": method})
    tmp_in = Path(tempfile.mkdtemp(prefix="lns_in_")) / "Forecast.xlsm"
    shutil.copy(WB, tmp_in)
    out = Path(tempfile.mkdtemp(prefix="lns_out_")) / "out.xlsm"
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        rc = run_mod.main(str(tmp_in), str(out), config_dir=cdir, scenario_dir=str(SCN))
    cap = optimize._harvest_cap(str(CFG), {})
    # welfare_density: omitting it took the 80.0 module default while the
    # operator's density_welfare_threshold_kg_m3 is what every other caller
    # (tournament, compare, robustness) resolves and passes.
    m, dropped, overprod = optimize.metrics_from_workbook(
        str(out), cap, welfare_density=optimize._welfare_density(str(CFG), {}))
    return rc, m, dropped, overprod, out


def _drift_counts(out_path):
    """(count_drift_rows, bio_drift_rows) from TankContinuityAudit — the audit the regression locks."""
    import openpyxl
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb["TankContinuityAudit"]
    cd = bd = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row:
            continue
        if len(row) > 14 and row[14] == "TANK_DRIFT":
            cd += 1
        if len(row) > 27 and row[27] == "BIO_DRIFT":
            bd += 1
    return cd, bd


_DET_CODE = """
import shutil, tempfile, os, io, contextlib, openpyxl
import forecast.run as r
from forecast import optimize
cdir = optimize.config_dir_with_overrides(os.environ["CFG"], {"placement_method": os.environ["M"]})
t = os.path.join(tempfile.gettempdir(), "lnsdet%d.xlsm" % os.getpid())
o = os.path.join(tempfile.gettempdir(), "lnsdeto%d.xlsm" % os.getpid())
shutil.copy(os.environ["WB"], t)
with contextlib.redirect_stdout(io.StringIO()):
    r.main(t, o, config_dir=cdir, scenario_dir=os.environ["SCN"])
wb = openpyxl.load_workbook(o, data_only=True); ws = wb["BatchLocations"]
v = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i < 5 or not row: continue
    d = row[8]
    if isinstance(d, (int, float)) and d > 95: v.append(round(d, 2))
print("%d|%.2f|%.2f" % (len(v), max(v, default=0.0), round(sum(v), 2)))
"""


def _det_signature(method, seed):
    env = dict(os.environ, WB=str(WB), CFG=str(CFG), SCN=str(SCN), M=method,
               PYTHONHASHSEED=str(seed))
    out = subprocess.run([sys.executable, "-c", _DET_CODE], cwd=str(ROOT),
                         capture_output=True, text=True, env=env)
    assert out.returncode == 0, f"{method} seed {seed} failed:\n{out.stderr[-800:]}"
    return out.stdout.strip().splitlines()[-1]


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--methods", nargs="+", default=["greedy", "lns"])
    ap.add_argument("--det", action="store_true",
                    help="also run a PYTHONHASHSEED 0 vs 1 determinism check on lns")
    args = ap.parse_args(argv)

    if not WB.exists():
        sys.exit("Forecast.xlsm not found at repo root")

    base_peak = None
    print(f"{'method':<8} {'rc':>3} {'sys_peak':>9} {'wk>55k':>7} "
          f"{'dropped':>8} {'overprod':>9} {'cnt_drift':>10} {'bio_drift':>10}  verdict")
    for method in args.methods:
        rc, m, dropped, overprod, out = _run_method(method)
        cd, bd = _drift_counts(out)
        if base_peak is None:
            base_peak = m.system_peak
        verdict = []
        if cd or bd or dropped or overprod:
            verdict.append("CONSERVATION-FAIL")
        if m.system_peak > base_peak + 1e-9:
            verdict.append("WORSE-THAN-GREEDY")
        v = " ".join(verdict) or "ok"
        print(f"{method:<8} {rc:>3} {m.system_peak:>9.3f} "
              f"{m.weeks_over_harvest_cap:>7} {dropped:>8} {overprod:>9} "
              f"{cd:>10} {bd:>10}  {v}")

    if args.det:
        s0 = _det_signature("lns", 0)
        s1 = _det_signature("lns", 1)
        ok = "OK" if s0 == s1 else "FAIL"
        print(f"\ndeterminism(lns): seed0={s0} seed1={s1}  {ok}")


if __name__ == "__main__":
    main()
