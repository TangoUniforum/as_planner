"""Runner for the tankless GLOBAL system assigner (Layer 2) proof-of-concept.

METHOD: GLOBAL L2 (system assignment)

Loads the repo config/ + scenario/, hydrates in-flight OG state from the
ProductionReport in Forecast.xlsm, runs L1 (forecast.global_planner_poc.plan
with record_standing=True) to get the per-(batch, week) standing population,
then runs L2 (forecast.global_planner_l2_poc.assign) to assign that population
to SYSTEMS honouring the conveyor + per-system biomass/feed caps. Writes a small
.xlsx (CSV fallback) with:

  1. SystemLoadTrace — per (system, week): biomass + feed/day vs caps + OVER flag.
  2. Overflow        — which weeks/tiers overflowed and by how much.
  3. Assignment      — per (batch, week): system + biomass there.
  4. Conservation    — per week: Sum(systems) biomass == L1 facility standing.

Usage:
    python -m tools.run_l2_poc
    python -m tools.run_l2_poc --workbook Forecast.xlsm --out l2_out.xlsx
    python -m tools.run_l2_poc --no-pr      # skip PR; incoming-only

Creates ONLY this runner + forecast/global_planner_l2_poc.py (plus an additive
record_standing flag on L1). Touches no production file; not imported by the
pipeline.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.config_io import load_config
from forecast.scenario_io import load_batches, load_limits
from forecast import global_planner_poc as gpp
from forecast import global_planner_l2_poc as l2
from tools.run_global_poc import _hydrate_inflight_og


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config-dir", default=str(_ROOT / "config"))
    ap.add_argument("--scenario-dir", default=str(_ROOT / "scenario"))
    ap.add_argument("--workbook", default=str(_ROOT / "Forecast.xlsm"))
    ap.add_argument("--out", default=str(_ROOT / "global_l2_poc_out.xlsx"))
    ap.add_argument("--harvest-tank-density-pct", type=float, default=1.25)
    ap.add_argument("--no-pr", action="store_true",
                    help="skip ProductionReport hydration (incoming batches only)")
    args = ap.parse_args()

    print("=" * 72)
    print("  METHOD: GLOBAL L2 (system assignment)")
    print("=" * 72)

    control, tables, facility = load_config(args.config_dir)
    batches = load_batches(args.scenario_dir)
    _facility_limits, system_limits = load_limits(args.scenario_dir, control)
    print(f"  Config:   {args.config_dir}")
    print(f"  Scenario: {len(batches)} batches from {args.scenario_dir}")

    inflight = {}
    if not args.no_pr:
        inflight, derived_start = _hydrate_inflight_og(Path(args.workbook), batches)
        if derived_start is not None:
            control.forecast_start = derived_start
            print(f"  ForecastStart: derived {derived_start.date()} from PR closing +1d")
        print(f"  In-flight OG batches from PR: {len(inflight)}")

    fs = control.forecast_start
    fs_date = fs.date() if hasattr(fs, "date") else fs
    print(f"  forecast_start={fs_date}, horizon={control.horizon_weeks}w")

    # ---- L1: envelope + standing (record_standing=True exposes per-batch-week)
    # OG-ONLY instant-removal diagnostic (L2 greedy water-filler does not route
    # the 6N in_purge pool). Pin both whole-facility flags False so the new
    # defaults do not change this diagnostic.
    l1 = gpp.plan(batches, tables, control, facility,
                  inflight_og=inflight,
                  harvest_tank_density_pct=args.harvest_tank_density_pct,
                  record_standing=True,
                  model_purge_hold=False, model_full_facility=False)
    peak_bio = max((r.standing_biomass_kg for r in l1.trace), default=0.0)
    print(f"  L1 peak standing biomass: {peak_bio:,.0f} kg "
          f"({100*peak_bio/control.max_biomass_kg:.1f}% of "
          f"{control.max_biomass_kg:,.0f} cap); "
          f"L1 feasible={l1.feasible}")
    print(f"  L1 per-(batch,week) standing rows: {len(l1.batch_standing)}")

    # ---- L2: assign standing to systems
    res = l2.assign(l1, control, facility, system_limits)
    print(f"  L2 systems: {', '.join(res.systems)}")

    # ---- Per-system feasibility verdict ----
    print("\n  PER-SYSTEM FEASIBILITY VERDICT")
    n_sys_weeks = len(res.loads)
    over_b = [r for r in res.loads if r.over_biomass]
    over_f = [r for r in res.loads if r.over_feed]
    print(f"    system-weeks evaluated: {n_sys_weeks}")
    print(f"    over BIOMASS cap: {len(over_b)} system-weeks")
    print(f"    over FEED cap:    {len(over_f)} system-weeks")
    print(f"    weeks with >=1 system over a cap: {res.over_system_weeks}")
    print(f"    worst biomass fill ratio: {res.worst_biomass_ratio:.3f} "
          f"({'OK' if res.worst_biomass_ratio <= 1.0 else 'OVER'})")
    print(f"    worst feed fill ratio:    {res.worst_feed_ratio:.3f} "
          f"({'OK' if res.worst_feed_ratio <= 1.0 else 'OVER'})")
    if over_b or over_f:
        from collections import Counter, defaultdict
        sysc = Counter(r.system_id for r in (over_b + over_f))
        print("    systems over a cap (system: over-weeks):")
        for sid, n in sorted(sysc.items(), key=lambda kv: -kv[1]):
            print(f"      {sid}: {n}")
        # Worst SINGLE-week total over-cap (a snapshot, not the horizon sum).
        wk_b = defaultdict(float)
        wk_f = defaultdict(float)
        for r in res.loads:
            if r.over_biomass and r.biomass_cap:
                wk_b[r.week_label] += r.biomass_kg - r.biomass_cap
            if r.over_feed and r.feed_cap:
                wk_f[r.week_label] += r.feed_kg_day - r.feed_cap
        if wk_b:
            wlab, wval = max(wk_b.items(), key=lambda kv: kv[1])
            print(f"    worst single-week biomass over-cap: {wval:,.0f} kg ({wlab})")
        if wk_f:
            wlab, wval = max(wk_f.items(), key=lambda kv: kv[1])
            print(f"    worst single-week feed over-cap: {wval:,.0f} kg/day ({wlab})")

    # ---- Overflow report ----
    print("\n  OVERFLOW REPORT")
    if not res.overflows:
        print("    no overflows: every tier had system headroom every week.")
    else:
        from collections import Counter
        ow_weeks = sorted({o.week_label for o in res.overflows})
        ow_kg = sum(o.biomass_kg for o in res.overflows)
        print(f"    {len(res.overflows)} overflow events across "
              f"{len(ow_weeks)} weeks")
        print(f"    (cumulative batch-weeks placed over-cap: {ow_kg:,.0f} kg "
              f"summed over the horizon — a churn measure, not a snapshot)")
        reasons = Counter(o.reason for o in res.overflows)
        for reason, n in reasons.most_common():
            print(f"      {n:>4}x  {reason}")
        print(f"    overflow weeks: {', '.join(ow_weeks[:12])}"
              + (" ..." if len(ow_weeks) > 12 else ""))

    # ---- Conservation ----
    print("\n  CONSERVATION (per week: Sum(systems) biomass == L1 standing)")
    worst = max((abs(c["diff_kg"]) for c in res.conservation), default=0.0)
    worst_pct = max((abs(c["diff_kg"]) / c["l1_standing_kg"] * 100.0
                     for c in res.conservation if c["l1_standing_kg"] > 0),
                    default=0.0)
    print(f"    worst |Sum(systems) - L1 standing|: {worst:.6f} kg "
          f"({worst_pct:.6f}%) — {'OK conserves' if worst < 1e-3 else 'CHECK'}")

    # ---- How much of L1's envelope survives ----
    print("\n  ENVELOPE REALISABILITY")
    placed = sum(a.biomass_kg for a in res.assignments if not a.overflowed)
    spilled = sum(a.biomass_kg for a in res.assignments if a.overflowed)
    total = placed + spilled
    if total > 0:
        print(f"    standing placed within tier+caps: {placed:,.0f} kg "
              f"({100*placed/total:.2f}%)")
        print(f"    standing requiring overflow/over-cap: {spilled:,.0f} kg "
              f"({100*spilled/total:.2f}%)")

    _write_outputs(res, control, Path(args.out))
    return 0


def _write_outputs(res: l2.L2Result, control, out_path: Path):
    try:
        from openpyxl import Workbook
    except Exception as e:  # noqa: BLE001
        print(f"\n  (openpyxl unavailable: {e}); writing CSVs instead")
        _write_csvs(res, out_path.with_suffix(""))
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "METHOD"
    ws["A1"] = "METHOD: GLOBAL L2 (system assignment)"
    ws["A2"] = f"forecast_start={control.forecast_start}"
    ws["A3"] = f"horizon_weeks={control.horizon_weeks}"
    ws["A4"] = f"systems={','.join(res.systems)}"
    ws["A5"] = f"feasible_all_systems_under_caps={res.feasible}"
    ws["A6"] = f"over_system_weeks={res.over_system_weeks}"
    ws["A7"] = f"worst_biomass_ratio={res.worst_biomass_ratio:.4f}"
    ws["A8"] = f"worst_feed_ratio={res.worst_feed_ratio:.4f}"

    wl = wb.create_sheet("SystemLoadTrace")
    wl.append(["week", "week_label", "system_id", "tier", "biomass_kg",
               "biomass_cap", "feed_kg_day", "feed_cap", "n_batches",
               "over_biomass", "over_feed", "biomass_ratio", "feed_ratio"])
    for r in res.loads:
        wl.append([r.week, r.week_label, r.system_id, r.tier,
                   round(r.biomass_kg, 1), r.biomass_cap,
                   round(r.feed_kg_day, 1), r.feed_cap, r.n_batches,
                   r.over_biomass, r.over_feed,
                   round(r.biomass_ratio, 4), round(r.feed_ratio, 4)])

    wo = wb.create_sheet("Overflow")
    wo.append(["week", "week_label", "batch_id", "from_tier", "biomass_kg",
               "spilled_to_system", "reason"])
    for o in res.overflows:
        wo.append([o.week, o.week_label, o.batch_id, o.from_tier,
                   round(o.biomass_kg, 1), o.spilled_to_system, o.reason])

    wa = wb.create_sheet("Assignment")
    wa.append(["week", "week_label", "batch_id", "tier", "system_id",
               "biomass_kg", "feed_kg_day", "avg_wt_g", "overflowed"])
    for a in res.assignments:
        wa.append([a.week, a.week_label, a.batch_id, a.tier, a.system_id,
                   round(a.biomass_kg, 1), round(a.feed_kg_day, 1),
                   round(a.avg_wt_g, 1), a.overflowed])

    wc = wb.create_sheet("Conservation")
    wc.append(["week", "week_label", "system_total_kg", "l1_standing_kg",
               "diff_kg"])
    for c in res.conservation:
        wc.append([c["week"], c["week_label"], round(c["system_total_kg"], 3),
                   round(c["l1_standing_kg"], 3), round(c["diff_kg"], 6)])

    try:
        wb.save(out_path)
        print(f"\n  Wrote {out_path}")
    except Exception as e:  # noqa: BLE001
        print(f"\n  (could not save xlsx: {e}); writing CSVs instead")
        _write_csvs(res, out_path.with_suffix(""))


def _write_csvs(res: l2.L2Result, stem: Path):
    import csv
    stem.parent.mkdir(parents=True, exist_ok=True)
    with open(f"{stem}_systemload.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["week", "week_label", "system_id", "tier", "biomass_kg",
                    "biomass_cap", "feed_kg_day", "feed_cap", "n_batches",
                    "over_biomass", "over_feed", "biomass_ratio", "feed_ratio"])
        for r in res.loads:
            w.writerow([r.week, r.week_label, r.system_id, r.tier, r.biomass_kg,
                        r.biomass_cap, r.feed_kg_day, r.feed_cap, r.n_batches,
                        r.over_biomass, r.over_feed, r.biomass_ratio, r.feed_ratio])
    with open(f"{stem}_overflow.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["week", "week_label", "batch_id", "from_tier", "biomass_kg",
                    "spilled_to_system", "reason"])
        for o in res.overflows:
            w.writerow([o.week, o.week_label, o.batch_id, o.from_tier,
                        o.biomass_kg, o.spilled_to_system, o.reason])
    print(f"  Wrote {stem}_systemload.csv and {stem}_overflow.csv")


if __name__ == "__main__":
    raise SystemExit(main())
