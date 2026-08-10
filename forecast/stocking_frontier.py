"""Stocking-for-quality frontier (pure Python, UI-free).

On a capacity-bound facility the density KNOBS can't lower density (no free
tanks). The real lever for product quality is stocking FEWER fish. This sweeps a
stocking-reduction fraction over the FUTURE batch schedule and, for each, runs
the full forecast and measures the quality-vs-volume trade: fewer fish are reared
gentler (lower experienced density) but yield less harvest tonnage.

Only batches whose INPUT (stocking) date is AFTER forecast_start are scaled —
you can only choose to stock less in the FUTURE; fish already in the facility
(from the PR) are fixed, and that includes FW-in-flight batches that merely
haven't reached seawater yet (their TranOG date is future but the fish are
already swimming). The caller's config / scenario / PR are never touched: every
point runs in a throwaway temp copy.

Decoupled from any UI (mirrors forecast.copilot / forecast.optimize) — the app is
a thin shell over `stocking_frontier()`.
"""
from __future__ import annotations

import contextlib
import io
import os
import shutil
import tempfile
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path


@dataclass
class FrontierPoint:
    reduction: float               # stocking cut fraction (0.10 = 10% fewer future fish)
    harvest_t: float               # total gross (live) tonnes harvested — the VOLUME
    harvest_fish: float            # total fish harvested
    mean_rearing_density: float    # biomass-weighted kg/m3 — QUALITY (lower = gentler)
    crowded_biomass_fraction: float
    worst_density: float
    conserves: bool
    scaled_batches: int            # how many future batches were reduced
    error: str | None = None


def _forecast_start(input_path):
    from openpyxl import load_workbook
    from forecast.production_report import read_production_report
    wb = load_workbook(str(input_path), data_only=True)
    pc, _og, _fw = read_production_report(wb)
    wb.close()
    if pc is None:                      # tolerant PR parse — fail with WHY, not .year
        raise ValueError(
            "ProductionReport has no readable 'Closing Month' date — the frontier "
            "cannot derive forecast_start (fix the PR's closing-date cell)")
    return date(pc.year, pc.month, pc.day) + timedelta(days=1)


def scale_future_batches(batches, fs, keep: float) -> int:
    """Scale the stocking of every batch NOT YET STOCKED at forecast start `fs`
    by `keep` (in place); return how many batches were scaled.

    Keys on the INPUT (stocking) date, not the TranOG (seawater-entry) date: a
    batch whose input_date <= fs is already swimming (possibly still in FW,
    awaiting seawater entry) — 'fish in the facility are fixed', so scaling or
    culling it is not a stocking decision. Only input_date > fs is a future
    stocking the operator can still reduce."""
    scaled = 0
    for b in batches:
        ind = b.input_date
        indd = ind.date() if hasattr(ind, "date") else ind
        if indd and indd > fs:                  # future stocking only
            b.input_count = int(round((b.input_count or 0) * keep))
            if b.tran_og_count:
                b.tran_og_count = int(round(b.tran_og_count * keep))
            scaled += 1
    return scaled


def _harvest_and_worst(wb):
    """(total gross tonnes, total fish, worst non-6N per-tank density) from output."""
    ht = hf = 0.0
    if "HarvestPlan" in wb.sheetnames:
        for r in wb["HarvestPlan"].iter_rows(values_only=True):
            if (r and isinstance(r[0], str) and "-W" in r[0] and len(r) > 5
                    and isinstance(r[3], (int, float)) and isinstance(r[5], (int, float))):
                hf += r[3]
                ht += r[5]
    worst = 0.0
    if "BatchLocations" in wb.sheetnames:
        for i, r in enumerate(wb["BatchLocations"].iter_rows(values_only=True), 1):
            if i < 5 or not r or r[0] is None:
                continue
            if len(r) > 4 and r[4] == "OG6N":
                continue
            d = r[8] if len(r) > 8 else None
            if isinstance(d, (int, float)):
                worst = max(worst, d)
    return ht / 1000.0, hf, worst


def _run_scaled(input_path, config_dir, scenario_dir, reduction, fs, welfare):
    """One frontier point: temp-copy config+scenario, scale FUTURE batch counts
    by (1-reduction), run the forecast, and measure quality + volume."""
    import yaml
    from openpyxl import load_workbook
    from forecast import optimize as _opt
    from forecast.run import main as run_pipeline
    from forecast.scenario_io import load_batches, batches_to_list, BATCHES_FILE
    from forecast.tuning import _conservation

    work = tempfile.mkdtemp(prefix="as_stock_")
    try:
        cdir = os.path.join(work, "config")
        sdir = os.path.join(work, "scenario")
        shutil.copytree(config_dir, cdir)
        shutil.copytree(scenario_dir, sdir)

        scaled = 0
        if reduction > 0:
            batches = load_batches(sdir)
            scaled = scale_future_batches(batches, fs, 1.0 - reduction)
            (Path(sdir) / BATCHES_FILE).write_text(
                yaml.safe_dump({"batches": batches_to_list(batches)},
                               sort_keys=False, allow_unicode=True),
                encoding="utf-8")

        inp = os.path.join(work, os.path.basename(str(input_path)))
        out = os.path.join(work, "out.xlsm")
        shutil.copy(str(input_path), inp)
        with contextlib.redirect_stdout(io.StringIO()):
            run_pipeline(inp, out, config_dir=cdir, scenario_dir=sdir)

        wb = load_workbook(out, data_only=True)
        mean_d, _fw, frac = _opt._density_quality(wb, welfare)
        ht, hf, worst = _harvest_and_worst(wb)
        wb.close()
        dropped, overprod = _conservation(out)
        return FrontierPoint(
            reduction=reduction, harvest_t=ht, harvest_fish=hf,
            mean_rearing_density=mean_d, crowded_biomass_fraction=frac,
            worst_density=worst, conserves=(dropped == 0 and overprod == 0),
            scaled_batches=scaled)
    except Exception as e:  # noqa: BLE001 — one bad point must not kill the sweep
        return FrontierPoint(reduction=reduction, harvest_t=0.0, harvest_fish=0.0,
                             mean_rearing_density=0.0, crowded_biomass_fraction=0.0,
                             worst_density=0.0, conserves=False, scaled_batches=0,
                             error=f"{type(e).__name__}: {e}")
    finally:
        shutil.rmtree(work, ignore_errors=True)


def stocking_frontier(input_path, config_dir, scenario_dir, *,
                      reductions=(0.0, 0.05, 0.10, 0.15),
                      welfare_density=80.0) -> list[FrontierPoint]:
    """Run one forecast per stocking-reduction fraction and return the quality-vs-
    volume frontier (see module docstring). Each point runs the full pipeline
    (~20s), so keep `reductions` short for interactive use."""
    fs = _forecast_start(input_path)
    return [_run_scaled(input_path, config_dir, scenario_dir, f, fs, welfare_density)
            for f in reductions]
