"""Populatable config template workbook.

A standalone Excel file for BUILDING the forecast config — models, facility,
control, forward batches, and limits — offline and separately from the
ProductionReport. Export the current config as a template (or a blank one),
edit it in Excel, then import it back into config/ + scenario/ YAML.

Layout: one sheet per config piece, each a simple header-row table that maps
1:1 to the YAML. Uses openpyxl only (no pandas) so it works in the engine.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from .config_io import (
    load_control, load_biology_tables, load_facility_config,
    control_to_dict, control_from_dict, biology_to_dict, biology_from_dict,
    facility_to_dict, facility_from_dict, dump_config,
)
from .scenario_io import (
    load_batches, load_limits, batches_to_list, batches_from_list,
    facility_limits_to_list, system_limits_to_list, facility_limits_from_list,
    system_limits_from_list, dump_scenario,
)

S_CONTROL = "Control"
S_GROWTH = "Biology_Growth"
S_MORT = "Biology_Mortality"
S_FEED = "Biology_Feed"
S_CULL = "Biology_Cull"
S_FAC = "Facility"
S_BATCH = "Batches"
S_FLIM = "FacilityLimits"
S_SLIM = "SystemLimits"
# The standing per-system capacities (defaults + mode defaults). SystemLimits
# above stays the per-week EXCEPTION grid; this sheet is where a capacity
# normally lives, so a template round-trip carries it in one legible row per
# system instead of 130 identical columns.
S_SCAP = "SystemCapacities"
_SCAP_COLS = ["system", "mode", "metric", "value"]

# Headers for the empty-template / fixed-schema sheets.
_FAC_COLS = ["location_id", "system_id", "tank_id",
             "volume_m3", "max_density_kg_m3", "max_feed_kg_day", "type"]
_BATCH_COLS = ["batch_id", "input_date", "input_count", "tran_sf_date",
               "tran_og_date", "tran_og_count", "tran_og_avg_wt_g", "tran_og_cv",
               "fcr_model", "fw_correction", "sgr_correction", "notes"]
_GROWTH_COLS = ["size_g", "SGR_FW", "SGR_SW", "FCR_1.21", "FCR_1.18",
                "FCR_1.16", "FCR_1.15"]
_MORT_COLS = ["week_from_input", "mortality_pct"]
_FEED_COLS = ["max_size_g", "feed_name"]
_CULL_COLS = ["days_since_input", "cull_pct"]
_FLIM_COLS = ["week", "metric", "value"]
_SLIM_COLS = ["week", "system", "metric", "value"]


def _write_table(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(list(r))
    for i, _ in enumerate(header, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16


_DEFAULT_OG_SYSTEMS = ["OG1N", "OG1S", "OG2N", "OG2S", "OG3N", "OG3S",
                       "OG4N", "OG4S", "OG5N", "OG5S", "OG6N", "OG6S"]


def _og_systems(config_dir):
    """OG system ids from facility config, or the standard 12 as a default."""
    if config_dir and (Path(config_dir) / "facility.yaml").exists():
        fac = load_facility_config(config_dir)
        systems = sorted({t.system_id for t in fac.tanks
                          if t.type == "OG" and t.system_id})
        if systems:
            return systems
    return list(_DEFAULT_OG_SYSTEMS)


def write_config_template(out_path, config_dir=None, scenario_dir=None,
                          horizon_weeks=None, forecast_start=None) -> Path:
    """Write a config-template workbook to out_path.

    Pre-filled with the current config where available (edit-from-current),
    else blank headers. When `horizon_weeks` + `forecast_start` are given, the
    FacilityLimits/SystemLimits sheets are generated as a FULL grid — a row for
    every (week, metric) and (week, system, metric) over the horizon, value
    pre-filled where the current config has one and blank otherwise — so every
    cap slot the engine can use is laid out ready to populate.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    info = wb.create_sheet("README")
    info.append(["CONFIG TEMPLATE — populate these sheets and import in the app "
                 "(Configure > Import) to build the forecast config."])
    info.append(["One sheet per config piece. Keep the header row. "
                 "Dates as YYYY-MM-DD. Blank cells = default/none."])
    info.append(["Sheets: Control, Biology_Growth/Mortality/Feed/Cull, "
                 "Facility, Batches, FacilityLimits, SystemCapacities, "
                 "SystemLimits."])
    info.append(["SystemCapacities holds each system's standing cap (blank "
                 "mode = every mode) — edit a capacity THERE. SystemLimits is "
                 "for one-off weeks only; blank means 'use the capacity'."])
    info.column_dimensions["A"].width = 100

    # ---- Control ----
    if config_dir and (Path(config_dir) / "control.yaml").exists():
        d = control_to_dict(load_control(config_dir))
        ctrl_rows = [[k, "" if v is None else v] for k, v in d.items()]
    else:
        ctrl_rows = []
    _write_table(wb.create_sheet(S_CONTROL), ["field", "value"], ctrl_rows)

    # ---- Biology ----
    if config_dir and (Path(config_dir) / "biology.yaml").exists():
        bd = biology_to_dict(load_biology_tables(config_dir))
        models = sorted(bd["fcr_by_model"].keys())
        gcols = ["size_g", "SGR_FW", "SGR_SW"] + [f"FCR_{m}" for m in models]
        n = len(bd["sgr_size_g"])
        grows = []
        for i in range(n):
            row = [bd["sgr_size_g"][i], bd["sgr_fw_pct_day"][i], bd["sgr_sw_pct_day"][i]]
            for m in models:
                col = bd["fcr_by_model"].get(m, [])
                row.append(col[i] if i < len(col) else None)
            grows.append(row)
        mrows = list(zip(bd["mortality_week_from_input"], bd["mortality_pct_weekly"]))
        frows = [list(x) for x in bd["feed_types"]]
        crows = [list(x) for x in bd["culling"]]
    else:
        gcols, grows, mrows, frows, crows = _GROWTH_COLS, [], [], [], []
    _write_table(wb.create_sheet(S_GROWTH), gcols, grows)
    _write_table(wb.create_sheet(S_MORT), _MORT_COLS, mrows)
    _write_table(wb.create_sheet(S_FEED), _FEED_COLS, frows)
    _write_table(wb.create_sheet(S_CULL), _CULL_COLS, crows)

    # ---- Facility ----
    if config_dir and (Path(config_dir) / "facility.yaml").exists():
        tanks = facility_to_dict(load_facility_config(config_dir))["tanks"]
        frows = [[t.get(c) for c in _FAC_COLS] for t in tanks]
    else:
        frows = []
    _write_table(wb.create_sheet(S_FAC), _FAC_COLS, frows)

    # ---- Batches ----
    if scenario_dir and (Path(scenario_dir) / "batches.yaml").exists():
        brows = [[b.get(c) for c in _BATCH_COLS]
                 for b in batches_to_list(load_batches(scenario_dir))]
    else:
        brows = []
    _write_table(wb.create_sheet(S_BATCH), _BATCH_COLS, brows)

    # ---- Limits (full grid when a horizon is given) ----
    from .caps import (METRIC_BIOMASS, METRIC_FEED_DAY, METRIC_MAX_HARVEST,
                       METRIC_MIN_HARVEST, METRIC_HOG_YIELD)
    from .time_grid import forecast_week_labels
    fl_metrics = [METRIC_BIOMASS, METRIC_FEED_DAY, METRIC_MAX_HARVEST,
                  METRIC_MIN_HARVEST, METRIC_HOG_YIELD]
    sl_metrics = [METRIC_BIOMASS, METRIC_FEED_DAY]
    fl_cur, sl_cur = {}, {}
    sl_defaults, sl_mode_defaults = {}, {}
    if scenario_dir and (Path(scenario_dir) / "limits.yaml").exists():
        fl, sl = load_limits(scenario_dir)
        fl_cur = {(r["week"], r["metric"]): r["value"]
                  for r in facility_limits_to_list(fl)}
        sl_cur = {(r["week"], r["system"], r["metric"]): r["value"]
                  for r in system_limits_to_list(sl)}
        sl_defaults, sl_mode_defaults = dict(sl.defaults), dict(sl.mode_defaults)
    if horizon_weeks and forecast_start is not None:
        weeks = forecast_week_labels(forecast_start, int(horizon_weeks))
    else:
        weeks = sorted({k[0] for k in fl_cur} | {k[0] for k in sl_cur})
    systems = _og_systems(config_dir)
    # WIDE layout: weeks as columns, one row per parameter (matches the VBA
    # FacilityLimits/SystemLimits sheets). Easy to fill across weeks in Excel.
    _write_table(
        wb.create_sheet(S_FLIM), ["metric"] + weeks,
        [[m] + [fl_cur.get((wk, m)) for wk in weeks] for m in fl_metrics],
    )
    _write_table(
        wb.create_sheet(S_SLIM), ["system", "metric"] + weeks,
        [[s, m] + [sl_cur.get((wk, s, m)) for wk in weeks]
         for s in systems for m in sl_metrics],
    )
    # Standing capacities: one row per (system, mode, metric). `mode` blank =
    # applies in every mode. Written AFTER SystemLimits so the exception grid
    # keeps its familiar position.
    scap_rows = [[s, "", m, sl_defaults.get((s, m))]
                 for s in systems for m in sl_metrics]
    scap_rows += [[s, mode, m, v]
                  for (s, mode, m), v in sorted(sl_mode_defaults.items())]
    _write_table(wb.create_sheet(S_SCAP), _SCAP_COLS, scap_rows)

    out = Path(out_path)
    wb.save(out)
    wb.close()
    return out


def _read_table(wb, name) -> tuple[list[str], list[dict]]:
    if name not in wb.sheetnames:
        return [], []
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({header[i]: (r[i] if i < len(r) else None)
                    for i in range(len(header))})
    return header, out


def _facility_limits_records(wb) -> list[dict]:
    """Read FacilityLimits (WIDE: metric rows × week columns), or old LONG
    (week, metric, value). Returns long {week, metric, value} records,
    blank values skipped."""
    header, rows = _read_table(wb, S_FLIM)
    if not rows:
        return []
    lower = [h.lower() for h in header]
    if "value" in lower:  # legacy long format
        return [{"week": r["week"], "metric": r["metric"], "value": r["value"]}
                for r in rows
                if r.get("week") and r.get("value") not in (None, "")]
    weekcols = [h for h in header if h.lower() != "metric"]
    out = []
    for r in rows:
        m = r.get("metric")
        if not m:
            continue
        for wk in weekcols:
            v = r.get(wk)
            if v not in (None, ""):
                out.append({"week": wk, "metric": m, "value": v})
    return out


def _system_limits_records(wb) -> list[dict]:
    """Read SystemLimits (WIDE: system+metric rows × week columns), or old
    LONG. Returns long {week, system, metric, value} records."""
    header, rows = _read_table(wb, S_SLIM)
    if not rows:
        return []
    lower = [h.lower() for h in header]
    if "value" in lower:  # legacy long format
        return [{"week": r["week"], "system": r["system"], "metric": r["metric"],
                 "value": r["value"]} for r in rows
                if r.get("week") and r.get("value") not in (None, "")]
    weekcols = [h for h in header if h.lower() not in ("system", "metric")]
    out = []
    for r in rows:
        s, m = r.get("system"), r.get("metric")
        if not s or not m:
            continue
        for wk in weekcols:
            v = r.get(wk)
            if v not in (None, ""):
                out.append({"week": wk, "system": s, "metric": m, "value": v})
    return out


def _system_capacity_records(wb) -> list[dict]:
    """Read SystemCapacities → {system, mode, metric, value} records.

    Blank `mode` means "in every mode". Blank value = no default for that
    (system, metric), which is a real state (no cap) and NOT the same as an
    absent sheet — see import_config_template.
    """
    _header, rows = _read_table(wb, S_SCAP)
    out = []
    for r in rows or []:
        s, m = r.get("system"), r.get("metric")
        if not s or not m:
            continue
        if r.get("value") in (None, ""):
            continue
        out.append({"system": str(s).strip(),
                    "mode": str(r.get("mode") or "").strip(),
                    "metric": str(m).strip(),
                    "value": float(r["value"])})
    return out


def import_config_template(wb, config_dir, scenario_dir) -> list[str]:
    """Read a config-template workbook and write config/ + scenario/ YAML.

    Returns the list of pieces written.
    """
    from .models import BiologyTables

    written: list[str] = []

    # ---- Control ----
    _, ctrl = _read_table(wb, S_CONTROL)
    control = None
    if ctrl:
        d = {row["field"]: (None if row.get("value") in (None, "") else row["value"])
             for row in ctrl if row.get("field")}
        control = control_from_dict(d)

    # ---- Biology ----
    gh, growth = _read_table(wb, S_GROWTH)
    _, mort = _read_table(wb, S_MORT)
    _, feed = _read_table(wb, S_FEED)
    _, cull = _read_table(wb, S_CULL)
    tables = None
    if growth:
        models = [c[4:] for c in gh if c.startswith("FCR_")]

        def _f(v):
            return None if v in (None, "") else float(v)

        sizes = [float(r["size_g"]) for r in growth if r.get("size_g") not in (None, "")]
        tables = BiologyTables(
            sgr_size_g=sizes,
            sgr_fw_pct_day=[_f(r.get("SGR_FW")) for r in growth if r.get("size_g") not in (None, "")],
            sgr_sw_pct_day=[_f(r.get("SGR_SW")) for r in growth if r.get("size_g") not in (None, "")],
            fcr_size_g=list(sizes),
            fcr_by_model={
                m: [float("nan") if r.get(f"FCR_{m}") in (None, "") else float(r[f"FCR_{m}"])
                    for r in growth if r.get("size_g") not in (None, "")]
                for m in models
            },
            mortality_week_from_input=[int(r["week_from_input"]) for r in mort
                                       if r.get("week_from_input") not in (None, "")],
            mortality_pct_weekly=[float(r["mortality_pct"]) for r in mort
                                  if r.get("week_from_input") not in (None, "")],
            feed_types=[(float(r["max_size_g"]), str(r["feed_name"])) for r in feed
                        if r.get("max_size_g") not in (None, "")],
            culling=[(int(r["days_since_input"]), float(r["cull_pct"])) for r in cull
                     if r.get("days_since_input") not in (None, "")],
        )

    # ---- Facility ----
    _, fac = _read_table(wb, S_FAC)
    facility = facility_from_dict({"tanks": fac}) if fac else None

    # Config needs all three together to dump; fall back to current files for
    # any piece the template left empty.
    if control or tables or facility:
        cur_control = control or (load_control(config_dir)
                                  if (Path(config_dir) / "control.yaml").exists() else None)
        cur_tables = tables or (load_biology_tables(config_dir)
                                if (Path(config_dir) / "biology.yaml").exists() else None)
        cur_fac = facility or (load_facility_config(config_dir)
                               if (Path(config_dir) / "facility.yaml").exists() else None)
        if cur_control and cur_tables and cur_fac:
            dump_config(config_dir, control=cur_control, tables=cur_tables, facility=cur_fac)
            written += ["config/control.yaml", "config/biology.yaml", "config/facility.yaml"]

    # ---- Scenario: batches + limits ----
    _, batch = _read_table(wb, S_BATCH)
    flim = _facility_limits_records(wb)   # long records, blanks already skipped
    slim = _system_limits_records(wb)
    scap = _system_capacity_records(wb)
    if batch or flim or slim or scap:
        batches = (batches_from_list(batch) if batch
                   else (load_batches(scenario_dir)
                         if (Path(scenario_dir) / "batches.yaml").exists() else []))
        if flim or slim or scap:
            fl = facility_limits_from_list(flim)
            sl = system_limits_from_list(slim)
        else:
            fl, sl = load_limits(scenario_dir)
        if scap:
            for r in scap:
                if r["mode"]:
                    sl.mode_defaults[(r["system"], r["mode"], r["metric"])] = r["value"]
                else:
                    sl.defaults[(r["system"], r["metric"])] = r["value"]
        elif (flim or slim) and (Path(scenario_dir) / "limits.yaml").exists():
            # An OLDER template (or one built before SystemCapacities existed)
            # carries per-week rows but no capacities sheet. Rebuilding from it
            # alone would silently delete every standing capacity and leave the
            # facility uncapped — so keep what is on disk. A template that DOES
            # carry the sheet is authoritative and replaces them.
            _fl0, sl0 = load_limits(scenario_dir)
            sl.defaults = dict(sl0.defaults)
            sl.mode_defaults = dict(sl0.mode_defaults)
        dump_scenario(scenario_dir, batches=batches, facility_limits=fl, system_limits=sl)
        written += ["scenario/batches.yaml", "scenario/limits.yaml"]

    return written


def is_config_template(wb) -> bool:
    """True if the workbook looks like a config template (has the marker sheets)."""
    names = set(wb.sheetnames)
    return S_CONTROL in names and (S_GROWTH in names or S_BATCH in names or S_FLIM in names)
