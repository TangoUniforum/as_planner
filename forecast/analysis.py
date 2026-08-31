"""Analysis layer — the composition that turns many runs into ONE decision.

The app's modes (Compare & Choose, Optimize, Tune) each answer a PIECE of the
operator's real question: "which engine, with which knobs, gives the best plan
that passes the hard rules?" This module holds the pieces the composition
needs that don't exist anywhere else:

  * hard/soft GATES as a registry (conservation, never-an-empty-week, caps,
    harvest targets) — the checklist that makes "did I miss something?"
    structurally impossible to answer wrong;
  * harvest TARGETS (monthly/yearly kg) — config-owned, penalized not
    hard-gated (operator decision 2026-08-05);
  * ECONOMICS (price per fish size) — turns harvest kg into revenue;
  * the PROMOTED DEFAULT — the operator-blessed best candidate, stored in
    config/ (versioned, exported with config snapshots, never in an output
    workbook) so it survives sessions and cannot be lost by a run.

Extensibility rule: the analysis flow iterates registries — it does not know
what the gates/objectives are. A future lever (facility expansion, stocking
frequency/size) or objective (growth, revenue emphasis) is a register() call
plus its implementation, not a rewrite of the flow. Same pattern as
forecast.methods.REGISTRY.

Everything here is pure logic + file IO; Streamlit stays in app.py.
"""
from __future__ import annotations

import datetime as _dt
import re as _re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import yaml

from .yaml_atomic import read_text_resilient, write_text_atomic

TARGETS_FILE = "targets.yaml"
ECONOMICS_FILE = "economics.yaml"
DEFAULTS_FILE = "analysis_defaults.yaml"


# --------------------------------------------------------------------------- #
# Config files: targets, economics, promoted default
# --------------------------------------------------------------------------- #
def _load_yaml_or_none(path: Path) -> Optional[dict]:
    if not path.exists():
        return None
    data = yaml.safe_load(read_text_resilient(path))
    return data if isinstance(data, dict) else None


def _dump_yaml(path: Path, data: dict) -> None:
    write_text_atomic(path, yaml.safe_dump(data, sort_keys=False,
                                           allow_unicode=True))


def load_targets(config_dir) -> Optional[dict]:
    """{basis: 'hog'|'gross', tolerance_pct: float,
        monthly: {'YYYY-MM': kg}, yearly: {'YYYY': kg}} or None if unset.

    Missing/empty file -> None (the targets gate reports N/A, never FAIL:
    absent targets must not block analysis)."""
    d = _load_yaml_or_none(Path(config_dir) / TARGETS_FILE)
    if not d:
        return None
    out = {
        "basis": str(d.get("basis", "hog")).lower(),
        "tolerance_pct": float(d.get("tolerance_pct", 5.0)),
        "monthly": {str(k): float(v) for k, v in (d.get("monthly") or {}).items()
                    if v is not None},
        "yearly": {str(k): float(v) for k, v in (d.get("yearly") or {}).items()
                   if v is not None},
    }
    return out if (out["monthly"] or out["yearly"]) else None


def save_targets(config_dir, targets: dict) -> None:
    _dump_yaml(Path(config_dir) / TARGETS_FILE, targets)


def load_economics(config_dir) -> Optional[dict]:
    """{currency, basis: 'hog'|'gross', model_cv_pct,
        price_bands: [{min_kg, max_kg, price_per_kg, monthly: {YYYY-MM: p}}]}
    or None.

    model_cv_pct — the SALES-side harvest weight-distribution CV (%), the
    operator's Model_CV: each harvest event's kg is spread across the bands
    with a size-biased lognormal around the event's average weight (the
    operator's own Excel method). Re-tuned against historical harvest results,
    hence a config variable, not a constant. Distinct from the biological
    grading CV (tran_og_cv).

    price_per_kg is the default price; `monthly` optionally overrides it per
    forecast month. kg falling outside every band is reported unpriced — a
    loud gap beats silently inventing a price."""
    d = _load_yaml_or_none(Path(config_dir) / ECONOMICS_FILE)
    if not d:
        return None
    bands = []
    for b in d.get("price_bands") or []:
        try:
            monthly = {str(k): float(v)
                       for k, v in (b.get("monthly") or {}).items()
                       if v is not None}
            bands.append({"min_kg": float(b["min_kg"]),
                          "max_kg": float(b["max_kg"]),
                          "price_per_kg": float(b["price_per_kg"]),
                          "monthly": monthly})
        except (KeyError, TypeError, ValueError):
            print(f"WARN: economics.yaml: malformed price band skipped: {b!r}")
            continue
    if not bands:
        return None
    return {"currency": str(d.get("currency", "USD")),
            "basis": str(d.get("basis", "hog")).lower(),
            "model_cv_pct": float(d.get("model_cv_pct", 18.0) or 18.0),
            "price_bands": sorted(bands, key=lambda b: b["min_kg"])}


def save_economics(config_dir, economics: dict) -> None:
    _dump_yaml(Path(config_dir) / ECONOMICS_FILE, economics)


def load_promoted_default(config_dir) -> Optional[dict]:
    """The operator-blessed analysis candidate:
    {method, overrides, promoted_ts, note, evidence} or None."""
    d = _load_yaml_or_none(Path(config_dir) / DEFAULTS_FILE)
    return d if d and d.get("method") else None


def save_promoted_default(config_dir, method: str, overrides: dict,
                          promoted_ts: str, note: str = "",
                          evidence: Optional[dict] = None) -> None:
    _dump_yaml(Path(config_dir) / DEFAULTS_FILE, {
        "method": method,
        "overrides": overrides or {},
        "promoted_ts": promoted_ts,
        "note": note,
        "evidence": evidence or {},
    })


# --------------------------------------------------------------------------- #
# Harvest readers — HarvestPlan rows -> period totals + revenue
# --------------------------------------------------------------------------- #
def week_to_month(week_label: str) -> Optional[str]:
    """'2026-W31' -> '2026-07' (month of the ISO week's Monday — the SAME
    convention as the app's monthly harvest table, so targets and the Harvest
    tab can never disagree about which month a week belongs to)."""
    try:
        y, w = int(str(week_label)[:4]), int(str(week_label)[6:8])
        return _dt.date.fromisocalendar(y, w, 1).strftime("%Y-%m")
    except (ValueError, TypeError):
        return None


def harvest_rows(out_path) -> list[dict]:
    """Per-event harvest rows from the output workbook's HarvestPlan sheet:
    [{week, count, gross_avg_kg, gross_kg, hog_kg, hog_avg_kg}, ...]."""
    import openpyxl
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    try:
        if "HarvestPlan" not in wb.sheetnames:
            return []
        ws = wb["HarvestPlan"]
        header = None
        rows: list[dict] = []
        for r in ws.iter_rows(values_only=True):
            if header is None:
                if r and str(r[0]).strip() == "Week" and any(
                        str(c).strip() == "Batch" for c in r if c):
                    header = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not r or not str(r[0]).startswith("20"):
                continue

            def _num(prefix, _r=r, _h=header):
                k = next((c for c in _h if c.startswith(prefix)), None)
                v = _r[_h[k]] if k is not None and _h[k] < len(_r) else None
                return float(v) if isinstance(v, (int, float)) else 0.0

            count = _num("Count")
            hog_kg = _num("HOG_Biomass")
            rows.append({
                "week": str(r[0]).strip(),
                "count": count,
                "gross_avg_kg": _num("Gross_AvgWt"),
                "gross_kg": _num("Gross_Biomass"),
                "hog_kg": hog_kg,
                "hog_avg_kg": (hog_kg / count) if count > 0 else 0.0,
            })
        return rows
    finally:
        wb.close()


def harvest_by_period(rows: list[dict], basis: str = "hog"
                      ) -> tuple[dict, dict]:
    """({'YYYY-MM': kg}, {'YYYY': kg}) on the given basis ('hog'|'gross')."""
    key = "hog_kg" if basis == "hog" else "gross_kg"
    monthly: dict[str, float] = {}
    yearly: dict[str, float] = {}
    for r in rows:
        m = week_to_month(r["week"])
        if m is None:
            continue
        monthly[m] = monthly.get(m, 0.0) + r[key]
        y = m[:4]
        yearly[y] = yearly.get(y, 0.0) + r[key]
    return monthly, yearly


def review_targets(monthly: dict, yearly: dict, targets: dict) -> dict:
    """Score actuals against targets — PENALIZED, not hard-gated (operator
    decision): a miss beyond tolerance is flagged, never disqualifying.

    Only periods the plan's horizon actually reaches are judged: a target for
    a month with zero recorded harvest AND no neighboring in-horizon month is
    still judged (0 vs target) IF any harvest month >= it exists — i.e. we
    judge every target period up to the last month with any harvest, so a
    blackout month inside the horizon shows as MISSED, while targets beyond
    the horizon end are N/A rather than false misses.

    Returns {rows: [{period, target_kg, actual_kg, pct, status}],
             judged, met, close, missed, worst_pct, total_shortfall_kg}."""
    tol = float(targets.get("tolerance_pct", 5.0))
    horizon_end_m = max(monthly) if monthly else ""
    horizon_end_y = max(yearly) if yearly else ""
    rows = []

    def _judge(period, target_kg, actual_kg, in_horizon):
        if not in_horizon:
            return {"period": period, "target_kg": target_kg,
                    "actual_kg": actual_kg, "pct": None, "status": "N/A"}
        pct = (actual_kg / target_kg * 100.0) if target_kg > 0 else 100.0
        status = ("MET" if pct >= 100.0 - 1e-9
                  else "CLOSE" if pct >= 100.0 - tol else "MISSED")
        return {"period": period, "target_kg": target_kg,
                "actual_kg": actual_kg, "pct": pct, "status": status}

    for period in sorted(targets.get("monthly") or {}):
        rows.append(_judge(period, targets["monthly"][period],
                           monthly.get(period, 0.0),
                           bool(horizon_end_m) and period <= horizon_end_m))
    for period in sorted(targets.get("yearly") or {}):
        rows.append(_judge(period, targets["yearly"][period],
                           yearly.get(period, 0.0),
                           bool(horizon_end_y) and period <= horizon_end_y))

    judged = [r for r in rows if r["status"] != "N/A"]
    shortfall = sum(max(0.0, r["target_kg"] - r["actual_kg"]) for r in judged)
    worst = min((r["pct"] for r in judged), default=None)
    return {
        "rows": rows,
        "judged": len(judged),
        "met": sum(1 for r in judged if r["status"] == "MET"),
        "close": sum(1 for r in judged if r["status"] == "CLOSE"),
        "missed": sum(1 for r in judged if r["status"] == "MISSED"),
        "worst_pct": worst,
        "total_shortfall_kg": shortfall,
    }


def biomass_band_fraction(mean_kg: float, cv: float,
                          lo_kg: float, hi_kg: float) -> float:
    """Fraction of a harvest event's BIOMASS falling in [lo, hi) — the
    operator's Excel method, verbatim: fish weights ~ lognormal around the
    event mean with the sales-model CV; the biomass (not count) share uses
    the size-biased distribution, i.e. LogN(mu + s^2, s):

        s  = sqrt(ln(1 + cv^2))
        mu = ln(mean) - s^2/2
        share = F(hi; mu+s^2, s) - F(lo; mu+s^2, s)

    cv is a FRACTION here (0.18, not 18). Scale-invariant: mean and band
    edges just need the same unit."""
    import math
    from statistics import NormalDist
    if mean_kg <= 0 or cv < 0:
        return 0.0
    if cv == 0:
        return 1.0 if lo_kg <= mean_kg < hi_kg else 0.0
    s = math.sqrt(math.log(1.0 + cv * cv))
    mu_b = math.log(mean_kg) - 0.5 * s * s + s * s   # mu + s^2 (size-biased)
    nd = NormalDist()

    def _F(x):
        return nd.cdf((math.log(x) - mu_b) / s) if x > 0 else 0.0

    return max(0.0, _F(hi_kg) - _F(lo_kg))


def sixn_outbound_transfers(out_path, production_start_iso: str = ""
                            ) -> Optional[int]:
    """R7 lens: TransferPlan rows moving fish OUT of a 6N tank
    (61/63/65/67/69/71) during the DEPURATION era (weeks before the 6N
    production start). In production mode the mains are ordinary grow-out,
    so later outbound moves are legal rebalancing, not violations.

    Returns the violation count, or None when the sheet is absent (the gate
    reports N/A, never a false verdict)."""
    import openpyxl
    # The 6N tank ids have ONE definition (sixn.SIXN_PAIRS -> SIXN_ALL_TANKS).
    # A literal set here silently disagrees the day a pair changes.
    from .sixn import SIXN_ALL_TANKS
    sixn = {str(t) for t in SIXN_ALL_TANKS}
    cutoff = ""
    if production_start_iso:
        try:
            d = _dt.date.fromisoformat(str(production_start_iso)[:10])
            iso = d.isocalendar()
            cutoff = f"{iso[0]}-W{iso[1]:02d}"
        except (ValueError, TypeError):
            cutoff = ""
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    try:
        if "TransferPlan" not in wb.sheetnames:
            return None
        ws = wb["TransferPlan"]
        header = None
        n = 0
        for r in ws.iter_rows(values_only=True):
            if header is None:
                if r and str(r[0]).strip() == "Week" and any(
                        str(c).strip() == "From_Tank" for c in r if c):
                    header = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not r or not str(r[0]).startswith("20"):
                continue
            week = str(r[0]).strip()
            if cutoff and week >= cutoff:
                continue                      # production era — legal moves
            typ = str(r[header.get("Type", -1)] or "") if "Type" in header else ""
            if typ != "Transfer":
                continue
            ft = r[header["From_Tank"]] if "From_Tank" in header else None
            try:
                ft_s = str(int(float(ft)))
            except (TypeError, ValueError):
                continue
            if ft_s in sixn:
                n += 1
        return n
    finally:
        wb.close()


def density_review(out_path) -> Optional[dict]:
    """Per-batch peak-density distribution for one plan — the diagnostic that
    was Tune mode's reason to exist, now a per-candidate lens. Reuses Tune's
    exact math (TransferTemplate Section B peaks; severe = >=1.3x cap).

    KEY READING RULE (hard-won): severe batches clustering in time and peaking
    mid-grow-out = a STOCKING/CAPACITY problem — no knob fixes it; the stocking
    lever does. Scattered mild overshoot near 1.0-1.1x = normal near-cap
    operation. Returns None when the sheet is absent (e.g. Global outputs
    without Section B) — the gate reports N/A, never a false verdict."""
    from . import tuning
    try:
        peaks, detail = tuning._peaks_and_detail(out_path)
    except Exception:  # noqa: BLE001 — sheet absent/foreign shape -> no lens
        return None
    if not peaks:
        return None
    d = tuning.analyze(peaks)
    return {"n": d.n, "over": d.over, "severe": d.severe,
            "worst": float(d.worst), "median": float(d.median),
            "buckets": dict(d.buckets), "severe_rows": list(detail)}


def week_was_forced(excess_kg: float, harvest_wt_kg: float,
                   feed_capacity: list, feed_mature: list) -> bool:
    """Could the weeks that FED this one have erased its excess?

    `excess_kg`     how far over the cap the week ran.
    `harvest_wt_kg` realised harvest weight, to turn kg into fish.
    `feed_capacity` per feeding week, plant capacity still unused that week.
    `feed_mature`   per feeding week, mature fish the pick could actually SEE
                    (the PRIOR week's standing stock outside 6N).

    Each feeding week can contribute only the smaller of the two: capacity with
    no fish is useless, and fish with no capacity cannot be processed. If their
    sum cannot cover the fish the excess represents, no plan could have erased
    it -> FORCED. Pure arithmetic on workbook numbers, so it reads identically
    for every engine.
    """
    if harvest_wt_kg <= 0:
        return True                       # cannot convert -> never charge
    need = max(0.0, excess_kg) / harvest_wt_kg
    possible = sum(min(c, m) for c, m in zip(feed_capacity, feed_mature))
    return possible < need


# Staging leads harvest: fish move into 6N, sit the purge hold, then drain. A red
# week is judged against the mature inventory available in the weeks that could
# still have fed it — the hold (~2 weeks) plus one week of rotation slack.
_CONVERGE_FEED_LEAD = 3


def convergence_review(out_path) -> Optional[dict]:
    """Does the plan WORK ITS WAY into bounds, and then HOLD there?

    The starting state is an operator INPUT, not a planner choice. A PR handed
    over at 101% of the biomass cap makes week 0 red whichever engine runs, so
    judging candidates on PEAK biomass scores the inherited state rather than
    the plan: every candidate fails the same way and the ranking says nothing.
    The question that actually separates plans is the operator's own —

        start red  ->  how fast does it reach green  ->  can it STAY green?

    A plan that inherits 101%, is under the cap by week 22 and never returns is
    doing its job. A plan that reaches green and relapses has not converged; it
    is oscillating, and the relapse is a planning failure in a way the red
    START never was.

    Judged against the per-week RESOLVED cap (Biomass_Limit moves — 3.80M at
    2026-W35, 3.65M by W37), so "green" means inside THAT week's limit, never a
    constant. Uses the Advisory sheet's own Biomass_Excess, the same column the
    operator reads, so the gate and the workbook can never disagree.

    FORCED vs AVOIDABLE red weeks (2026-08-27)
    ------------------------------------------
    The same logic that excuses an inherited red START has to be applied one
    level deeper, or the gate stops discriminating. Fish reach the plant only
    through the 6N purge, and only fish at/above `min_harvest_weight_g` may be
    staged. When a cohort has passed through and the next has not yet grown into
    the window, there is NO mature inventory: harvest cannot rise, biomass
    climbs, and NO plan — from any engine, under any knobs — can do otherwise.
    Measured on the 8.23.26 PR: 2026-W44..W49 hold ZERO mature fish outside 6N
    while 244k-440k sit just under the line, and the facility runs to 107% of
    cap. That excursion is a property of the smolt/growth calendar, which the
    operator does not control either (the site grows its own smolts).

    So a red week is FORCED when the weeks that could still have FED it (staging
    leads harvest by `_CONVERGE_FEED_LEAD`) could not between them have supplied
    the fish needed to erase the excess — each contributing the smaller of its
    remaining plant capacity and the mature fish the pick could actually see.
    Otherwise it is AVOIDABLE and counts against the plan. The gate judges
    AVOIDABLE weeks; forced ones are reported, never charged, and
    operator-scripted window weeks are excluded from both.

    The yardstick is deliberately the EXCESS, not the engine's own target. An
    intent-based test ("did it get what it asked for?") would judge the
    controller on a number it records and Global on plant capacity, which is the
    same asymmetry that let Global's OG-only biomass flatter it for months. This
    form reads only workbook data and applies identically to both.

    BASIS, stated plainly: maturity is measured on the TANK AVERAGE weight, so
    the graded tail of a sub-threshold batch (the engine can pull the top of a
    3,083 g batch across a 3,500 g line) is NOT counted as available. FORCED
    therefore means "no batch stood at harvest weight", which is the physical
    condition; where the engine ALSO logged its own shortfall, that is recorded
    in `engine_shortfall_weeks` as corroboration.

    Returns None when Advisory is absent -> the gate reports N/A, never a false
    verdict. `forced_judged` is False when BatchLocations/RunConfig are missing,
    in which case every red week is treated as AVOIDABLE (never silently
    excused).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — unreadable/foreign file -> no lens
        return None
    try:
        if "Advisory" not in wb.sheetnames:
            return None
        header, rows, adv_harvest = None, [], []
        for r in wb["Advisory"].iter_rows(values_only=True):
            if header is None:
                if (r and str(r[0]).strip() == "Week" and len(r) > 2
                        and "Total_Biomass" in str(r[2] or "")):
                    header = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not r or not str(r[0] or "").startswith("20"):
                continue
            def _num(key):
                i = next((j for k, j in header.items() if k.startswith(key)), None)
                v = r[i] if i is not None and i < len(r) else None
                return float(v) if isinstance(v, (int, float)) else None
            bio, cap = _num("Total_Biomass"), _num("Biomass_Limit")
            if bio is None or not cap:
                continue
            rows.append((str(r[0]).strip(), bio, cap))
            adv_harvest.append((str(r[0]).strip(), bio, cap,
                                _num("Harvest_Count"), _num("Harvest_Biomass")))

        # Operator inputs, taken from the workbook's own RunConfig stamp so the
        # gate judges against the numbers THIS run used, never today's config.
        min_wt_g = min_hv = max_hv = None
        inputs_from = "workbook"
        if "RunConfig" in wb.sheetnames:
            # Two stamp DIALECTS: the controller writes a YAML snapshot
            # ("key: value" in one cell), Global writes a two-column method
            # stamp (key in A, value in B). Read both — a parser that knows only
            # one silently leaves the other engine unjudged, which is how a
            # comparison stops being apples-to-apples.
            for r in wb["RunConfig"].iter_rows(values_only=True):
                cells = list(r or ())
                for key, setter in (("min_harvest_weight_g", "wt"),
                                    ("min_harvest_per_week", "hv"),
                                    ("max_harvest_per_week", "mx")):
                    v = None
                    for i, c in enumerate(cells):
                        t = str(c or "").strip()
                        if t.startswith(key + ":"):
                            try:
                                v = float(t.split(":", 1)[1].strip())
                            except (ValueError, IndexError):
                                v = None
                        elif t == key and i + 1 < len(cells):
                            try:
                                v = float(cells[i + 1])
                            except (TypeError, ValueError):
                                v = None
                        if v is not None:
                            break
                    if v is None:
                        continue
                    if setter == "wt":
                        min_wt_g = v
                    elif setter == "hv":
                        min_hv = v
                    else:
                        max_hv = v

        # Not every engine stamps its inputs (Global's RunConfig omits the
        # harvest block). Without them nothing can be excused and that engine
        # would be charged for every red week while another is excused — the
        # apples-to-apples failure this whole lens exists to prevent. Fall back
        # to the SHIPPED control defaults and say so in `inputs_from`, so a
        # reader can see the numbers did not come from the run itself.
        if min_wt_g is None or min_hv is None or max_hv is None:
            try:
                _cfg = yaml.safe_load(
                    (Path(__file__).resolve().parent.parent / "config"
                     / "control.yaml").read_text(encoding="utf-8")) or {}
                min_wt_g = min_wt_g or float(_cfg.get("min_harvest_weight_g") or 0) or None
                min_hv = min_hv or float(_cfg.get("min_harvest_per_week") or 0) or None
                max_hv = max_hv or float(_cfg.get("max_harvest_per_week") or 0) or None
                inputs_from = "shipped config (workbook did not stamp them)"
            except Exception:  # noqa: BLE001 — no config -> stays unjudged
                pass

        # Mature inventory OUTSIDE 6N: the only fish that can still be staged.
        # Fish already in 6N are committed — counting them is the double-count
        # that made an early read of this data show 355k "waiting" fish that
        # were in fact the 6N contents themselves.
        mature = {}
        if min_wt_g and "BatchLocations" in wb.sheetnames:
            from .sixn import SIXN_ALL_TANKS
            sixn = {str(t) for t in SIXN_ALL_TANKS}
            bh = None
            for r in wb["BatchLocations"].iter_rows(values_only=True):
                if bh is None:
                    if (r and str(r[0]).strip() == "Week"
                            and any(str(c).strip() == "Tank" for c in r if c)):
                        bh = {str(c).strip(): i for i, c in enumerate(r) if c}
                    continue
                if not r or not str(r[0] or "").startswith("20"):
                    continue
                try:
                    tank = str(int(float(r[bh["Tank"]])))
                    awt = float(r[bh["AvgWt (kg)"]] or 0)
                    cnt = float(r[bh["Count (fish)"]] or 0)
                except (TypeError, ValueError, KeyError):
                    continue
                if tank in sixn or awt * 1000.0 < min_wt_g:
                    continue
                wk = str(r[0]).strip()
                mature[wk] = mature.get(wk, 0.0) + cnt

        # What was actually STAGED into 6N each week. A week where the plan
        # already filled to the processing limit had no legal move left, however
        # many mature fish stood in the tanks — without this the lens calls
        # 2026-W52 "avoidable" while the plan is pinned at 55,000.
        staged = {}
        if "TransferPlan" in wb.sheetnames:
            from .sixn import SIXN_ALL_TANKS as _S
            _sx = {str(t) for t in _S}
            th = None
            for r in wb["TransferPlan"].iter_rows(values_only=True):
                if th is None:
                    if (r and str(r[0]).strip() == "Week"
                            and any(str(c).strip() == "From_Tank" for c in r if c)):
                        th = {str(c).strip(): i for i, c in enumerate(r) if c}
                    continue
                if not r or not str(r[0] or "").startswith("20"):
                    continue
                if str(r[th.get("Type", 0)] or "").strip() != "Transfer":
                    continue
                try:
                    ft = str(int(float(r[th["From_Tank"]])))
                    tt = str(int(float(r[th["To_Tank"]])))
                    cnt = float(r[th["Count (fish)"]] or 0)
                except (TypeError, ValueError, KeyError):
                    continue
                if tt in _sx and ft not in _sx:
                    wk = str(r[0]).strip()
                    staged[wk] = staged.get(wk, 0.0) + cnt

        # Realised harvest weight per week, to convert an excess in KG into the
        # number of FISH that would have had to be harvested to erase it.
        # Advisory carries both columns, so this needs no second sheet.
        hv_wt = {}
        for w, _b, _c, hc, hkg in adv_harvest:
            if hc and hkg and hc > 0:
                hv_wt[w] = hkg / hc

        # Operator-scripted MANUAL WINDOW weeks execute only the operator's own
        # events — no engine may add or trim a harvest there — so they are not
        # the planner's to answer for. Every harvest gate already excludes them
        # (forecast.window_weeks); this one must too, or a deliberately scripted
        # week is charged as a planning failure. Measured: the July'26 PR
        # scripts 2026-W31..W33 and W33 was landing in the avoidable list.
        try:
            from . import window_weeks as _ww
            scripted = set(_ww.manual_window_weeks(wb) or ())
        except Exception:  # noqa: BLE001 — no window info -> exclude nothing
            scripted = set()

        # The engine's OWN admission, where it made one — corroboration for the
        # inventory measure, never a substitute (it only logs on weeks where a
        # fill was actually attempted).
        shortfall_weeks = set()
        if "ValidationLog" in wb.sheetnames:
            for r in wb["ValidationLog"].iter_rows(values_only=True):
                txt = " ".join(str(c) for c in (r or ()) if c is not None)
                if "mature inventory" in txt or "short of min_hv" in txt:
                    m = _re.search(r"\b(\d{4}-W\d{2})\b", txt)
                    if m:
                        shortfall_weeks.add(m.group(1))
    finally:
        wb.close()
    if not rows:
        return None

    pct = [(w, b / c * 100.0) for w, b, c in rows]
    red = [p > 100.0 for _w, p in pct]
    n = len(pct)
    # TOUCHING green and SETTLING into it are different events, and conflating
    # them is how "green by week 5" hides a December bulge to 107%. `first_green`
    # is the first week under the cap; `settled` is the week after the LAST red
    # one -- the point from which the plan actually MAINTAINS green, which is the
    # operator's question. None when the horizon ends red: it never settles.
    first_green = next((i for i, r in enumerate(red) if not r), None)
    last_red = max((i for i, r in enumerate(red) if r), default=None)
    settled = None if last_red is None else (
        last_red + 1 if last_red + 1 < n else None)
    if last_red is None:
        settled = 0
    # A relapse is a GREEN -> RED transition after the first green week: the plan
    # had it and gave it back. Counting red weeks instead would punish the
    # inherited start, which is exactly the confound this lens exists to remove.
    relapses = sum(1 for i in range(1, n)
                   if first_green is not None and i > first_green
                   and red[i] and not red[i - 1])
    # FORCED: no plan could have harvested more into this week, because the
    # weeks that could still have FED it held no batch at harvest weight.
    # Without the inputs to judge that (foreign/old workbook), nothing is
    # excused — `forced_judged` False and every red week stays AVOIDABLE.
    forced_judged = bool(mature) and bool(min_hv)
    scripted_red = [red[i] and pct[i][0] in scripted for i in range(n)]
    forced = [False] * n
    if forced_judged:
        _wts = [v for v in hv_wt.values() if v > 0]
        _mean_wt = (sum(_wts) / len(_wts)) if _wts else 0.0
        for i in range(n):
            if not red[i] or scripted_red[i]:
                continue
            lo = max(0, i - _CONVERGE_FEED_LEAD)
            # AVOIDABILITY, not "was there slack". The earlier form asked
            # whether staging sat below the PLANT limit while mature fish
            # existed — which charges a feedback controller for correctly
            # staging the floor three weeks before biomass ever crossed the cap.
            # Measured: it called 19 weeks avoidable on the July'26 PR where a
            # direct sweep found 3 real staging defects.
            #
            # Ask the operator's question instead: could ENOUGH MORE FISH have
            # been staged, across the weeks that could still have fed this one,
            # to erase the excess? Each feeding week contributes the smaller of
            # its remaining plant capacity and the mature fish the pick could
            # actually see. If their sum cannot cover the excess, no plan could
            # have erased it and the week is FORCED. Uses only workbook data, so
            # it applies IDENTICALLY to every engine — an intent-based yardstick
            # would judge the controller on what it asked for while Global, which
            # records no such number, kept being judged on capacity.
            #
            # MATURITY IS READ ONE WEEK BACK. BatchLocations is a week-END
            # snapshot, so week j's row counts fish that crossed the harvest
            # weight AFTER week j's pick had already run — charging the pick for
            # fish it could not yet see. Measured on the 8.23.26 PR: 2026-W51
            # staged 27,445 and its own week shows 231,737 mature (a damning
            # -looking miss), while the snapshot the pick actually saw held
            # exactly 27,445 — it took every fish available. Reading w-1 is also
            # conservative: mid-week growth can only ADD, so this never invents
            # an opportunity that did not exist.
            _caps, _mats = [], []
            for j in range(max(1, lo), i):
                _caps.append(max(0.0, (max_hv or 0.0) - staged.get(pct[j][0], 0.0)))
                _mats.append(mature.get(pct[j - 1][0], 0.0))
            forced[i] = week_was_forced(
                max(0.0, rows[i][1] - rows[i][2]),
                hv_wt.get(pct[i][0]) or _mean_wt, _caps, _mats)
    avoidable = [red[i] and not forced[i] and not scripted_red[i]
                 for i in range(n)]

    relapse_pcts = [p for i, (_w, p) in enumerate(pct)
                    if first_green is not None and i > first_green and red[i]]
    steady = pct[settled:] if settled is not None else []
    steady_worst = max(steady, key=lambda x: x[1]) if steady else None
    return {
        "n_weeks": n,
        "start_pct": pct[0][1],
        "red_start": red[0],
        "weeks_red": sum(red),
        "converged": first_green is not None,
        "first_green_i": first_green,
        "first_green_week": pct[first_green][0] if first_green is not None else None,
        "settled_i": settled,
        "settled_week": pct[settled][0] if settled is not None else None,
        "relapses": relapses,
        # Relapses that begin on an AVOIDABLE week — the ones the plan owns.
        "relapses_avoidable": sum(
            1 for i in range(1, n)
            if first_green is not None and i > first_green
            and avoidable[i] and not red[i - 1]),
        "forced_judged": forced_judged,
        "inputs_from": inputs_from,
        "weeks_red_forced": sum(forced),
        "weeks_red_avoidable": sum(avoidable),
        "forced_weeks": [pct[i][0] for i in range(n) if forced[i]],
        "avoidable_weeks": [pct[i][0] for i in range(n) if avoidable[i]],
        "engine_shortfall_weeks": sorted(shortfall_weeks),
        "weeks_red_scripted": sum(scripted_red),
        "scripted_weeks": [pct[i][0] for i in range(n) if scripted_red[i]],
        "weeks_red_after_green": (sum(red[first_green:])
                                  if first_green is not None else sum(red)),
        "worst_pct": max(p for _w, p in pct),
        "worst_relapse_pct": max(relapse_pcts) if relapse_pcts else None,
        "steady_worst_pct": steady_worst[1] if steady_worst else None,
        "steady_worst_week": steady_worst[0] if steady_worst else None,
    }


def system_feed_review(out_path) -> Optional[dict]:
    """Per-SYSTEM feed load against each system's own cap.

    The facility feed cap has a gate; the per-system caps had none, so a plan
    could sit over a system's feeder capacity indefinitely and nothing would
    say so. Measured on the 8.23.26 PR: 67 system-weeks over, worst 1.318x,
    entirely unreported -- and the objective has no term for it either, which
    means a knob search is actively DRAWN to configurations that breach it,
    because an unmeasured violation is free score. (The relief pass that would
    shave these is `for _ in range(rebalance_balance_budget)`, so setting that
    budget to 0 -- which scores well -- switches the relief off completely.)

    Feed is a per-DAY rate against a per-day cap, so a breach means that system
    physically cannot deliver the day's ration, not that a total was exceeded.

    Reads SystemLimitsAudit, which both engine families write. Returns None when
    the sheet is absent -> the gate reports N/A, never a false verdict.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — unreadable/foreign file -> no lens
        return None
    try:
        if "SystemLimitsAudit" not in wb.sheetnames:
            return None
        header, over, worst, n = None, {}, {}, 0
        for r in wb["SystemLimitsAudit"].iter_rows(values_only=True):
            if header is None:
                if (r and str(r[0]).strip() == "Week"
                        and any(str(c).strip() == "System" for c in r if c)):
                    header = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not r or not str(r[0] or "").startswith("20"):
                continue
            def _n(key):
                i = header.get(key)
                v = r[i] if i is not None and i < len(r) else None
                return float(v) if isinstance(v, (int, float)) else None
            f, cap = _n("Feed_kg_day"), _n("Feed_cap")
            if f is None or not cap:
                continue
            sysid = str(r[header["System"]]).strip()
            n += 1
            if f > cap:
                over[sysid] = over.get(sysid, 0) + 1
                worst[sysid] = max(worst.get(sysid, 0.0), f / cap)
    finally:
        wb.close()
    if not n:
        return None
    tot = sum(over.values())
    return {"system_weeks": n, "over": tot,
            "over_pct": (tot / n * 100.0) if n else 0.0,
            "worst": (max(worst.values()) if worst else 1.0),
            "worst_system": (max(worst, key=worst.get) if worst else None),
            "systems_breaching": len(over),
            "by_system": dict(sorted(over.items(), key=lambda kv: -kv[1]))}


def revenue_for(rows: list[dict], economics: dict) -> dict:
    """Revenue for a plan: each harvest event's kg is SPREAD across the price
    bands with the size-biased lognormal (model_cv_pct), then priced per band
    — with the band's monthly price override when one exists for the event's
    month, else its default price. kg in no band (distribution tails outside
    the ladder) is unpriced, reported loudly.

    Returns {total, priced_kg, unpriced_kg, currency, by_band}."""
    basis = economics.get("basis", "hog")
    wt_key = "hog_avg_kg" if basis == "hog" else "gross_avg_kg"
    kg_key = "hog_kg" if basis == "hog" else "gross_kg"
    cv = float(economics.get("model_cv_pct", 18.0)) / 100.0
    bands = economics["price_bands"]
    by_band = [{"band": f"{b['min_kg']:g}–{b['max_kg']:g} kg",
                "price_per_kg": b["price_per_kg"], "kg": 0.0, "revenue": 0.0}
               for b in bands]
    total = priced = unpriced = 0.0
    for r in rows:
        w, kg = r[wt_key], r[kg_key]
        if kg <= 0 or w <= 0:
            continue
        month = week_to_month(r["week"])
        event_priced = 0.0
        for i, b in enumerate(bands):
            frac = biomass_band_fraction(w, cv, b["min_kg"], b["max_kg"])
            if frac <= 0:
                continue
            price = (b.get("monthly") or {}).get(month, b["price_per_kg"])
            part = kg * frac
            by_band[i]["kg"] += part
            by_band[i]["revenue"] += part * price
            total += part * price
            event_priced += part
        priced += event_priced
        unpriced += max(0.0, kg - event_priced)
    return {"total": total, "priced_kg": priced, "unpriced_kg": unpriced,
            "currency": economics.get("currency", "USD"), "by_band": by_band}


# --------------------------------------------------------------------------- #
# Result cache — finished runs survive reloads, frozen tabs, and new sessions
# --------------------------------------------------------------------------- #
# Streamlit session_state dies with the browser session; a finished CP-SAT leg
# is 30 minutes of compute. Entries are pickled OUTSIDE OneDrive (multi-MB
# binaries would churn sync) and keyed by name; staleness is the CALLER's
# problem — every stored entry carries its input signature and is checked at
# use, so an old entry is simply re-run, never wrongly trusted.
def _default_cache_dir() -> Path:
    import os
    import tempfile
    base = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    return Path(base) / "as_planner" / "result_cache"


def cache_save(name: str, obj, cache_dir=None, keep: int = 40) -> None:
    """Pickle `obj` under `name` atomically; evict oldest beyond `keep`."""
    import os
    import pickle
    d = Path(cache_dir) if cache_dir else _default_cache_dir()
    d.mkdir(parents=True, exist_ok=True)
    tmp = d / f"{name}.pkl.tmp-{os.getpid()}"
    with tmp.open("wb") as fh:
        pickle.dump(obj, fh, protocol=pickle.HIGHEST_PROTOCOL)
    os.replace(tmp, d / f"{name}.pkl")
    files = sorted(d.glob("*.pkl"), key=lambda p: p.stat().st_mtime,
                   reverse=True)
    for p in files[keep:]:
        try:
            p.unlink()
        except OSError:
            pass


def cache_load_all(cache_dir=None, prefix: str = "") -> dict:
    """{name: obj} for every readable cached entry (optionally filtered by
    name prefix). A corrupt/unreadable file is SKIPPED, never fatal — the
    worst outcome of a bad cache must be a re-run, not a crash."""
    import pickle
    d = Path(cache_dir) if cache_dir else _default_cache_dir()
    out: dict = {}
    if not d.exists():
        return out
    for p in sorted(d.glob(f"{prefix}*.pkl")):
        try:
            with p.open("rb") as fh:
                out[p.stem] = pickle.load(fh)
        except Exception as e:  # noqa: BLE001 — see docstring
            # Skipping is the right degrade (worst case = a re-run), but a
            # silent skip is how a "finished" leg quietly vanishes — say so.
            print(f"WARN: result cache: skipping unreadable {p.name} "
                  f"({type(e).__name__}) — that entry will re-run")
            continue
    return out


def dirs_fingerprint(dirs, exclude=frozenset()) -> str:
    """CONTENT hash of every file under `dirs` (recursive), minus names listed
    in `exclude`. Content, not mtimes: the 2026-08-10 stale-board incident was
    a scenario edit (a W33 manual harvest) that the previous name+mtime scan
    did not register, so disk-cached engine legs keyed on the mtime proxy
    replayed the pre-edit scenario as if current. Bytes cannot lie, and the
    files here are small YAML — hashing them outright costs nothing.
    tools/run_tuned_tournament.py already keys its caches on content; this
    aligns the app. An unreadable file degrades to its stat identity rather
    than failing the whole fingerprint."""
    import hashlib
    h = hashlib.md5()
    for d in dirs:
        d = Path(d)
        if not d.exists():
            continue
        for p in sorted(d.rglob("*"), key=lambda x: str(x)):
            if p.is_file() and p.name not in exclude:
                h.update(str(p.relative_to(d)).encode())
                try:
                    h.update(p.read_bytes())
                except OSError:
                    h.update(str(p.stat().st_mtime_ns).encode())
    return h.hexdigest()


def code_fingerprint(dirs) -> str:
    """CONTENT hash of the ENGINE SOURCE (*.py) under `dirs`.

    A cached result is only reusable when the CODE that produced it is
    unchanged too — inputs alone are not its identity. The 2026-08-12 incident:
    the Global engines were rebuilt over two days while config/ and scenario/
    stayed byte-identical, so the input fingerprint never moved and FOUR OF FIVE
    board legs replayed pre-repair results. The board presented a week-old
    engine as current, showing 230 kg/m3 and 2 unplaced batches for a method
    that had since been fixed.

    __pycache__ is skipped: .pyc files carry build metadata and would make the
    hash unstable run to run. An unreadable file degrades to a marker rather
    than failing the whole fingerprint."""
    import hashlib
    h = hashlib.md5()
    for d in dirs:
        d = Path(d)
        if not d.exists():
            continue
        for p in sorted(d.rglob("*.py"), key=lambda x: str(x)):
            if "__pycache__" in p.parts:
                continue
            h.update(str(p.relative_to(d)).replace("\\", "/").encode())
            try:
                h.update(p.read_bytes())
            except OSError:
                h.update(b"<unreadable>")
    return h.hexdigest()


def board_leg_current(entry, expected_sig: str) -> bool:
    """Whether a cached board engine leg may be REPLAYED for the inputs on
    screen. A leg is usable only when it carries the exact signature of the
    current inputs — anything else (old-format entry with no stored sig, a
    sig mismatch after a PR/config/scenario change, a malformed entry) is
    treated as ABSENT so the caller re-runs the engine instead of replaying
    a plan built for different inputs. Never raises on junk."""
    try:
        return bool(entry) and entry.get("sig") == expected_sig \
            and isinstance(entry.get("res"), dict)
    except AttributeError:  # entry isn't a dict — corrupt/foreign cache object
        return False


def drop_stale_grades(res, schema: str) -> bool:
    """Strip grading artifacts computed under an older METRICS_SCHEMA from a
    result dict, leaving the ENGINE output (workbook bytes/path, run stats)
    untouched. A schema bump changes how a run is JUDGED, not what the engine
    produced — so the remedy is a re-grade from the cached workbook, never an
    engine re-run. Grades from before schema-stamping existed (no "schema"
    key) are stale by definition. Returns True when anything was dropped."""
    if not isinstance(res, dict):
        return False
    stale = False
    sc = res.get("_score")
    if sc is not None and sc.get("schema") != schema:
        res.pop("_score", None)
        res.pop("_score_err", None)
        stale = True
    for k in ("_ana_rows", "_ana_density"):
        c = res.get(k)
        if c is not None and c.get("schema") != schema:
            res.pop(k, None)
            stale = True
    return stale


# --------------------------------------------------------------------------- #
# Gate registry — the checklist
# --------------------------------------------------------------------------- #
@dataclass
class Gate:
    key: str
    label: str
    hard: bool                     # hard = a FAIL disqualifies the candidate
    fn: Callable[[dict], tuple]    # ctx -> (status, detail); status in
    #                                PASS / WARN / FAIL / N/A


GATES: list[Gate] = []


def register_gate(key: str, label: str, hard: bool,
                  fn: Callable[[dict], tuple]) -> None:
    """Add a gate to the checklist. The analysis flow evaluates EVERY
    registered gate on every candidate — new rules become part of the
    checklist by registering, not by editing the flow."""
    GATES.append(Gate(key=key, label=label, hard=hard, fn=fn))


def evaluate_gates(ctx: dict) -> list[dict]:
    """Run every registered gate against one candidate's context dict.
    ctx keys used by the built-ins: dropped, overprod, zero_weeks,
    weeks_over_cap, weeks_over_harvest_cap, peak_pct_of_cap, worst_density,
    targets_review (from review_targets, or None)."""
    out = []
    for g in GATES:
        try:
            status, detail = g.fn(ctx)
        except Exception as e:  # noqa: BLE001 — a broken gate must be VISIBLE
            status, detail = "FAIL", f"gate error: {type(e).__name__}: {e}"
        out.append({"key": g.key, "label": g.label, "hard": g.hard,
                    "status": status, "detail": detail})
    return out


def _gate_conservation(ctx):
    d, o = int(ctx.get("dropped") or 0), int(ctx.get("overprod") or 0)
    if d == 0 and o == 0:
        return "PASS", "0 dropped / 0 over-produced"
    return "FAIL", f"{d} dropped / {o} over-produced fish"


def _gate_no_empty_week(ctx):
    """Judges the PLANNER on the weeks the planner controls: operator-scripted
    manual-window weeks are excluded upstream (the counts in ctx already leave
    them out — see forecast.window_weeks) and reported here so the verdict
    says what it judged. The MANUAL WINDOW ValidationLog lints police the
    scripted weeks themselves."""
    z = ctx.get("zero_weeks")
    if z is None:
        return "N/A", "zero-week count unavailable"
    ex = int(ctx.get("zero_weeks_excluded") or 0)
    scope = (f" ({ex} operator-scripted window week(s) excluded — "
             f"see the ValidationLog MANUAL WINDOW lints)" if ex else "")
    if int(z) == 0:
        return "PASS", (f"harvests something every planner week{scope}"
                        if ex else "harvests something every week")
    return "FAIL", f"{int(z)} totally empty harvest week(s){scope}"


def _gate_harvest_floor(ctx):
    """The steady-harvest CONTRACT floor (`min_harvest_per_week`).

    `no_empty_week` above only catches the DEGENERATE case — a week that
    harvests literally nothing. The contract the operator actually signed is
    a weekly FLOOR, and until 2026-08-12 nothing in the tool judged it: the
    count was computed (run_compare._harvest_extras), written to the
    RunComparison sheet, and then read by no gate and no score component.
    Measured consequence on the 7.29 PR: the tuned tournament promoted a
    controller knob set that cut the worst harvest week 20,526 -> 16,185
    fish, because the emphasis objective is statistically blind to the floor
    (corr(worst week, score) = -0.03 over a 40-variant search).

    SOFT by design: on a capacity-bound facility every plan has some
    sub-floor weeks, so a FAIL here would fail everything and teach the
    operator to ignore it. It reports, and `tournament.pick_winner` is what
    refuses to PROMOTE a regression.

    ctx keys: `weeks_below_floor`, `min_week`, `min_harvest` (the floor);
    `zero_weeks_excluded` = scripted window weeks left out of the counts."""
    n = ctx.get("weeks_below_floor")
    floor = ctx.get("min_harvest")
    if n is None or not floor:
        return "N/A", ("no harvest floor configured (min_harvest_per_week)"
                       if not floor else "weekly harvest series unavailable")
    n = int(n)
    ex = int(ctx.get("zero_weeks_excluded") or 0)
    scope = (f" ({ex} operator-scripted window week(s) excluded)" if ex else "")
    worst = ctx.get("min_week")
    worst_txt = (f", worst week {float(worst):,.0f}" if worst is not None else "")
    if n == 0:
        return "PASS", (f"every planner week meets the {float(floor):,.0f}-fish "
                        f"contract floor{worst_txt}{scope}")
    return "WARN", (f"{n} planner week(s) below the {float(floor):,.0f}-fish "
                    f"contract floor{worst_txt} — the steady-harvest contract "
                    f"is the hardest business rule; compare candidates on this "
                    f"number, not only on 'never an empty week'{scope}")


def _gate_biomass_cap(ctx):
    p = ctx.get("peak_pct_of_cap")
    if p is None:
        return "N/A", "peak biomass unavailable"
    p = float(p)
    if p <= 100.0:
        return "PASS", f"peak {p:.1f}% of cap"
    return ("WARN" if p <= 110.0 else "FAIL"), f"peak {p:.1f}% of cap"


def _gate_convergence(ctx):
    """Red -> green -> STAY green, charging only what the plan could control.

    `biomass_cap` judges the PEAK, which on an inherited over-cap starting state
    is a property of the INPUT: every engine peaks in week 0 and the gate cannot
    tell them apart. This gate judges the TRAJECTORY, and charges a red week
    only when the plan had a legal move it did not take — mature fish standing
    outside 6N in a week that was not already staging at the processing limit.
    Weeks with neither are FORCED (a maturity trough, or the plant already
    flat out) and are reported, never charged: no engine under any knobs could
    have done better, so charging them makes every plan WARN for the same
    reason and the gate stops discriminating.

    Touching green is not holding it, so the gate reports the week the plan
    SETTLES (the week after its last red one).

    PASS — never red; or every red week forced; or settles with real headroom.
    WARN — avoidable relapses, thin headroom, or ends red with no legal move.
    FAIL — the horizon ENDS red with avoidable weeks: it never settles and the
           plan owns it.
    """
    cr = ctx.get("convergence")
    if not cr:
        return "N/A", "weekly biomass-vs-cap series unavailable for this plan"
    n, worst = cr["n_weeks"], cr["worst_pct"]
    if cr["weeks_red"] == 0:
        return "PASS", (f"in bounds from week 0 and stays there — worst "
                        f"{worst:.1f}% of cap across {n} weeks")
    judged = cr.get("forced_judged")
    av = cr.get("weeks_red_avoidable", cr["weeks_red"]) if judged else cr["weeks_red"]
    fc = cr.get("weeks_red_forced", 0) if judged else 0
    fw = cr.get("forced_weeks") or []
    sc_n = cr.get("weeks_red_scripted", 0)
    sc_note = (f"; {sc_n} red week(s) were operator-SCRIPTED "
               f"({', '.join(cr.get('scripted_weeks') or [])}) and are excluded — "
               f"the engine may not add or trim a harvest there"
               if sc_n else "")
    forced_note = ""
    if fc:
        span = (f"{fw[0]}..{fw[-1]}" if len(fw) > 1 else fw[0])
        forced_note = (f"; {fc} further red week(s) ({span}) were FORCED — no "
                       f"mature fish outside 6N, or already staging at the "
                       f"processing limit — and are not charged")
    start = (f"starts at {cr['start_pct']:.1f}% (inherited starting state)"
             if cr["red_start"] else f"starts green at {cr['start_pct']:.1f}%")
    if cr["settled_i"] is None:
        if av == 0 and judged:
            return "WARN", (f"{start} and is still over the cap in the final "
                            f"week — but every one of its {cr['weeks_red']} red "
                            f"week(s) was forced: no plan could have harvested "
                            f"more. Worst {worst:.1f}%. This is a capacity/"
                            f"maturity problem, not a planning one")
        return "FAIL", (f"{start} and is still over the cap in the final week "
                        f"({av} avoidable red week(s), worst {worst:.1f}%): "
                        f"the plan never settles into green{forced_note}{sc_note}")
    sw, si = cr["settled_week"], cr["settled_i"]
    stw = cr["steady_worst_pct"]
    head = (f"; once settled it holds {100.0 - stw:.1f}% clear of the cap "
            f"(worst {stw:.1f}% at {cr['steady_worst_week']})"
            if stw is not None else "")
    if av == 0 and judged:
        return "PASS", (f"{start}; every one of its {cr['weeks_red']} red week(s) "
                        f"was FORCED — no mature fish outside 6N, or already "
                        f"staging at the processing limit — so no plan could "
                        f"have done better. Settles at {sw} (week {si} of {n})"
                        f"{head}")
    rel = cr.get("relapses_avoidable", cr["relapses"]) if judged else cr["relapses"]
    if rel:
        wr = cr["worst_relapse_pct"]
        wr_s = f", worst relapse {wr:.1f}%" if wr is not None else ""
        return "WARN", (f"{start}, first touches green at {cr['first_green_week']} "
                        f"but gives it back {rel}x{wr_s}, settling only at {sw} "
                        f"(week {si} of {n}). {av} red week(s) were AVOIDABLE — "
                        f"mature fish stood outside 6N in a week that was not "
                        f"already staging at the limit{forced_note}{sc_note}{head}")
    if stw is not None and stw > 99.5:
        return "WARN", (f"{start}, settles by {sw} (week {si} of {n}) with no "
                        f"avoidable relapse — but only {100.0 - stw:.1f}% "
                        f"headroom at {cr['steady_worst_week']}; it is riding "
                        f"the cap, so any surprise puts it back over"
                        f"{forced_note}")
    return "PASS", (f"{start}, works down to green by {sw} (week {si} of {n}) "
                    f"with {av} avoidable red week(s) and no avoidable relapse"
                    f"{forced_note}{head}")


def _gate_harvest_cap(ctx):
    """Weekly processing limit + pressure relief (operator semantics
    2026-08-09). `max_harvest_per_week` is THE processing limit; the relief
    band above it (up to limit * (1 + harvest_relief_pct)) is for
    EXCEPTIONAL weeks only — "allowed more, but not every time".

    PASS — no week over the limit.
    WARN — 1..3 relief weeks (over the limit, within the derived ceiling):
           acceptable if exceptional.
    FAIL — any week above the derived relief ceiling, OR more than 3 relief
           weeks: the plan leans on relief as capacity — harvests must ramp
           up earlier instead.
    Like the empty-week gate, this judges the PLANNER's weeks: operator-
    scripted manual-window weeks are excluded upstream (forecast.window_weeks)
    and the verdict says so, so a deliberately big scripted harvest cannot
    fail every engine at once.

    ctx keys: `weeks_over_harvest_cap` = weeks over the LIMIT (relief usage);
    `weeks_over_relief_ceiling` = weeks over the derived ceiling (0/absent
    when the workbook never breaches it); `zero_weeks_excluded` = how many
    scripted window weeks were left out of both counts."""
    wc = ctx.get("weeks_over_harvest_cap")
    if wc is None:
        return "N/A", "weekly harvest series unavailable"
    wc = int(wc)
    wr = int(ctx.get("weeks_over_relief_ceiling") or 0)
    ex = int(ctx.get("zero_weeks_excluded") or 0)
    scope = (f" ({ex} operator-scripted window week(s) excluded — judged by "
             f"the ValidationLog MANUAL WINDOW lints instead)" if ex else "")
    if wr > 0:
        return "FAIL", (f"{wr} planner week(s) above the relief ceiling "
                        f"(limit + harvest_relief_pct) — the plant cannot "
                        f"take it; the plan must ramp harvests up earlier"
                        f"{scope}")
    if wc == 0:
        return "PASS", (f"no planner week over the weekly processing limit"
                        f"{scope}" if ex
                        else "no week over the weekly processing limit")
    if wc <= 3:
        return "WARN", (f"pressure relief used {wc}x (planner weeks over the "
                        f"processing limit, within the relief ceiling) — "
                        f"acceptable if exceptional{scope}")
    return "FAIL", (f"{wc} relief week(s) — the plan leans on the relief "
                    f"band as everyday capacity; relief must stay "
                    f"exceptional: ramp harvests up earlier instead{scope}")


def _gate_system_feed(ctx):
    """Per-system feed capacity — the plan must be FEEDABLE system by system.

    Soft, like the facility biomass cap: it ranks a plan down rather than
    disqualifying it, because a system slightly over on a few days can be
    absorbed by the feeding schedule. What it must not do is stay SILENT, which
    is what it did before this gate existed -- and silence is what lets a knob
    search choose a breach for free.

    PASS — no system-week over its own feed cap.
    WARN — breaches, but the worst is within 10% and they are not systemic.
    FAIL — worst above 1.10x, or breaches on more than a quarter of
           system-weeks: that is a plan the feed system cannot deliver.
    """
    sf = ctx.get("system_feed")
    if not sf:
        return "N/A", "per-system feed series unavailable for this plan"
    if sf["over"] == 0:
        return "PASS", (f"every system within its feed cap across "
                        f"{sf['system_weeks']:,} system-week(s)")
    where = ", ".join(f"{k} x{v}" for k, v in list(sf["by_system"].items())[:3])
    tail = (f"{sf['over']} of {sf['system_weeks']:,} system-week(s) over "
            f"({sf['over_pct']:.0f}%), worst {sf['worst']:.3f}x on "
            f"{sf['worst_system']}; {sf['systems_breaching']} system(s) affected "
            f"— {where}")
    if sf["worst"] > 1.10 or sf["over_pct"] > 25.0:
        return "FAIL", (tail + ". The feed system cannot deliver this: check "
                        "the rebalancer budget, whose relief pass shaves "
                        "over-cap systems and does nothing when set to 0")
    return "WARN", tail


def _gate_targets(ctx):
    tr = ctx.get("targets_review")
    if not tr or not tr.get("judged"):
        return "N/A", "no harvest targets configured (Configure → Targets)"
    if tr["missed"] == 0 and tr["close"] == 0:
        return "PASS", f"all {tr['judged']} target period(s) met"
    if tr["missed"] == 0:
        return "WARN", (f"{tr['close']} period(s) within tolerance, "
                        f"worst {tr['worst_pct']:.0f}% of target")
    return "WARN", (f"{tr['missed']} period(s) missed "
                    f"(worst {tr['worst_pct']:.0f}% of target, "
                    f"short {tr['total_shortfall_kg'] / 1000.0:,.0f} t) — "
                    f"penalized, not disqualifying")


register_gate("conservation", "Conservation (no fish created or lost)",
              hard=True, fn=_gate_conservation)
register_gate("no_empty_week", "Never an empty harvest week", hard=True,
              fn=_gate_no_empty_week)
register_gate("harvest_floor", "Weekly contract floor (min harvest/week)",
              hard=False, fn=_gate_harvest_floor)
register_gate("biomass_cap", "Facility biomass cap", hard=False,
              fn=_gate_biomass_cap)
register_gate("convergence", "Converges: red -> green -> stays green",
              hard=False, fn=_gate_convergence)
register_gate("system_feed", "Per-system feed capacity", hard=False,
              fn=_gate_system_feed)
register_gate("harvest_cap", "Weekly processing limit + relief",
              hard=False, fn=_gate_harvest_cap)
register_gate("targets", "Harvest targets (monthly/yearly)", hard=False,
              fn=_gate_targets)


def _gate_density_quality(ctx):
    dr = ctx.get("density_review")
    if not dr:
        return "N/A", "per-batch peak-density data unavailable for this plan"
    if dr["severe"] == 0:
        return "PASS", (f"no batch over 1.3× cap (worst {dr['worst']:.2f}×, "
                        f"{dr['over']}/{dr['n']} touch the cap — normal near "
                        f"full utilisation)")
    return "WARN", (f"{dr['severe']} batch(es) peak ≥1.3× cap (worst "
                    f"{dr['worst']:.2f}×) — if these cluster in time and peak "
                    f"mid-grow-out it's a STOCKING/capacity problem, not a "
                    f"knob: see the stocking lever, don't re-tune")


register_gate("density_quality", "Per-batch density quality", hard=False,
              fn=_gate_density_quality)


def _gate_sixn_one_way(ctx):
    """R7 — 6N one-way commitment: fish moved into 6N depuration may never
    transfer out (only harvest). Counts depuration-era outbound 6N moves."""
    n = ctx.get("sixn_outbound_purge")
    if n is None:
        return "N/A", "TransferPlan unavailable for this plan"
    n = int(n)
    if n == 0:
        return "PASS", "no fish left a 6N depuration tank except by harvest"
    return "FAIL", (f"{n} transfer(s) moved fish OUT of a 6N depuration tank "
                    f"— the one-way commitment (R7) forbids this; those fish "
                    f"may only be harvested")


# HARD (operator ruling 2026-08-23: "every method needs to follow the same
# rules"). R7 is a RULE, not a preference: fish in depuration may leave only by
# harvest, because a transfer out means they are harvested WITHOUT completing
# the purge hold -- food safety, not bookkeeping.
#
# As a SOFT gate it merely ranked a method down, so an engine that broke R7
# still competed and could be promoted. That is how the Global arms placed in
# the 2026-08-23 tournament while moving 44,838 fish out of 6N, and their
# apparent advantage (555 t more harvest, zero density breaches) was measured
# against a controller that obeys the rule. Comparing a rule-follower with a
# rule-breaker is not a comparison.
#
# The controller family passes this gate at ZERO outbound moves, so making it
# binding costs nothing that obeys the rule -- it disqualifies exactly the
# methods that do not.
register_gate("sixn_one_way", "6N one-way commitment (R7)", hard=True,
              fn=_gate_sixn_one_way)


def _gate_handling_budget(ctx):
    """Operator rule 4 — the weekly handling budget (`max_transfers_per_week`).
    FAIL: any week over the budget; WARN: any week over ~80% of it;
    PASS: every week within.

    The budget number comes from ctx (`move_cap`) — the same value the counts
    were measured against — never from a literal here: a gate that hardcodes
    "15" starts lying the moment the knob is retuned.

    A Controller engine defers its quality passes to hold the budget, so an
    overrun there means ESSENTIAL moves alone (arrival make-room, rotation
    fills, the plan-diff) exceeded it that week. The Global engines never read
    the budget at all, so an overrun there is simply unbudgeted planning."""
    over = ctx.get("weeks_moves_over_cap")
    warn = ctx.get("weeks_moves_warn")
    if over is None and warn is None:
        return "N/A", "weekly transfer counts unavailable"
    over, warn = int(over or 0), int(warn or 0)
    cap = int(ctx.get("move_cap") or 0)
    cap_s = f"{cap}-move " if cap else ""
    warn_s = f" (>{int(0.8 * cap)} moves)" if cap else ""
    mx = ctx.get("moves_week_max")
    mx_s = f" (worst week {int(mx)} moves)" if mx else ""
    if over > 0:
        return "FAIL", (f"{over} week(s) over the {cap_s}handling budget"
                        f"{mx_s} — on a Controller plan that means essential "
                        f"moves alone exceeded it; the Global engines do not "
                        f"read the budget at all")
    if warn > 0:
        return "WARN", (f"{warn} week(s) above ~80% of the {cap_s}handling "
                        f"budget{warn_s}{mx_s}")
    return "PASS", f"every week within the {cap_s}handling budget{mx_s}"


register_gate("handling_budget", "Weekly handling budget", hard=False,
              fn=_gate_handling_budget)


# --------------------------------------------------------------------------- #
# Ranking — the operator-approved pick order
# --------------------------------------------------------------------------- #
def rank_key(candidate: dict) -> tuple:
    """Sort key for candidates (ascending; first = best). Encodes the
    operator-approved ordering (2026-08-05): hard gates absolutely first,
    then soft-gate failures, then target shortfall, then the emphasis score.
    Revenue enters through the card display (and later as a scored component
    once real price data is in economics.yaml), not the pick order."""
    gates = candidate.get("gates") or []
    hard_fails = sum(1 for g in gates if g["hard"] and g["status"] == "FAIL")
    soft_fails = sum(1 for g in gates if not g["hard"] and g["status"] == "FAIL")
    warns = sum(1 for g in gates if g["status"] == "WARN")
    tr = candidate.get("targets_review") or {}
    shortfall_kg = float(tr.get("total_shortfall_kg") or 0.0)
    score = float(candidate.get("score") if candidate.get("score") is not None
                  else 1e9)
    return (hard_fails, soft_fails, warns, round(shortfall_kg), score)


# --------------------------------------------------------------------------- #
# Adoption eligibility — the LAST door a plan passes on its way to config
# --------------------------------------------------------------------------- #
# `rank_key` above RANKS on gate failures; it does not FILTER on them. That is
# correct for a table (every candidate must be visible and comparable) but it
# is not a guard: the top row of a ranking is still the top row when every
# candidate breaks a rule, and `harvest_cap` — the weekly processing limit and
# its relief ceiling — is registered SOFT, so a ceiling breach lowers a plan's
# rank without ever disqualifying it.
#
# Analyze's ✅ Adopt writes the winning knobs into control.yaml and ⭐ Promote
# writes method + knobs into analysis_defaults.yaml (what ⚡ Quick run replays).
# Both are the same destination `tournament.pick_winner` and `optimize.recommend`
# already refuse to send an ineligible plan to. This section closes that third
# door with the SAME predicates — imported, never re-implemented.
#
# The difference from the other two doors is deliberate: those pick a winner
# automatically, so they exclude. This one is the operator's own decision
# surface — they can see the plan and its checklist and may have a reason — so
# it does not exclude, it REFUSES TO BE SILENT. See `adoption_blocked`.

# Hard gates whose failure the imported predicates already report in their own
# words (`optimize.ineligibility_reasons`). Any OTHER hard gate — including one
# registered later — is picked up from the candidate's own checklist, so the
# registry stays the single place a hard rule is declared.
_PREDICATE_GATE_KEYS = frozenset({"conservation", "no_empty_week"})


def adoption_variant(candidate: dict):
    """Adapt a GRADED analysis candidate to the `OptVariant` shape the
    winner-eligibility predicates read.

    Analyze grades candidates into gate dicts + a `Metrics` instance; the
    tournament grades `OptVariant`s. Rather than write a second copy of the
    rules against the gate shape, this builds the shape the existing predicates
    already understand — an adapter, not a rulebook.

    `candidate` keys used: `metrics` (a `Metrics`, or None), and
    `res["_score"]["verdict"]` for the conservation counts (`dropped`,
    `overprod`) — the same two numbers the `conservation` gate judges.

    An UNGRADED candidate (no metrics at all) yields a variant whose three
    guarded measurements are explicitly None — UNKNOWN, which the predicates
    read as "never a pass". A zeroed sentinel would read as a clean sweep."""
    import dataclasses

    from . import optimize
    sc = (candidate.get("res") or {}).get("_score") or {}
    verdict = sc.get("verdict") or {}
    m = candidate.get("metrics")
    if m is None:
        m = sc.get("metrics")
    if m is None:
        m = dataclasses.replace(optimize._infeasible_metrics(),
                                harvest_zero_weeks=None,
                                weeks_over_relief_ceiling=None,
                                harvest_min_week=None)
    return optimize.OptVariant(
        label=str(candidate.get("label") or candidate.get("key") or "candidate"),
        overrides=dict(candidate.get("overrides") or {}),
        metrics=m,
        dropped=int(verdict.get("dropped") or 0),
        overprod=int(verdict.get("overprod") or 0))


def stock_reference_min_week(candidate: dict, candidates) -> Optional[float]:
    """The contract-floor guard's baseline for one adoption candidate: the
    worst planner harvest week of the SAME METHOD run at STOCK config (no knob
    overrides) on the same board.

    `floor_eligible` asks "does tuning make the leanest week worse than not
    tuning?", so the reference must be the candidate's own un-tuned run. A
    candidate with no overrides IS its own stock run (nothing to regress
    against), and a tuned candidate whose stock leg isn't on the board has no
    reference at all — both return None, which leaves the guard OFF exactly as
    `optimize.recommend` does for a seeded search. A guard never invents a
    baseline; an unmeasured worst week is not one either."""
    if not candidate.get("overrides"):
        return None
    key = candidate.get("key")
    for c in candidates or ():
        if c is candidate or c.get("key") != key or c.get("overrides"):
            continue
        mw = getattr(c.get("metrics"), "harvest_min_week", None)
        if mw is not None:
            try:
                return float(mw)
            except (TypeError, ValueError):
                return None
    return None


def adoption_breaches(candidate: dict, baseline_min_week=None) -> list:
    """Every reason this graded candidate may not be adopted or promoted
    SILENTLY. Empty = nothing to confirm (the common case — a guard that cries
    wolf on clean plans teaches the operator to click through it).

    Two sources, each already the single copy of its own rule:

      * `optimize.ineligibility_reasons` — the imported winner-eligibility
        predicates (`tournament.variant_hard_ok`, `ceiling_eligible`,
        `floor_eligible`): conservation, never-an-empty-week, the relief
        ceiling, and the contract-floor no-regression when a baseline exists.
        The ceiling is the one this door adds that the checklist cannot: its
        gate is SOFT, so `rank_key` ranks a breach down and then adopts it.
      * the candidate's OWN checklist for any hard gate the predicates don't
        cover (`_PREDICATE_GATE_KEYS`). A hard rule registered tomorrow guards
        this door the day it is registered — registry, not rewrite."""
    from . import optimize
    out = list(optimize.ineligibility_reasons(adoption_variant(candidate),
                                              baseline_min_week))
    for g in candidate.get("gates") or []:
        if (g.get("hard") and g.get("status") == "FAIL"
                and g.get("key") not in _PREDICATE_GATE_KEYS):
            out.append(f"{g.get('label')} — {g.get('detail')}")
    return out


def adoption_blocked(breaches, acknowledged: bool) -> bool:
    """Whether an adopt/promote must NOT write.

    The operator is never LOCKED OUT of a plan they can see and understand —
    this is their decision surface, not an automatic winner-pick, and a guard
    they cannot override on their own judgement would be wrong here. But a
    guard they can trip WITHOUT NOTICING is exactly what this closes: a plan
    carrying a breach writes only after an explicit acknowledgement that names
    the breach, and the breach is recorded with whatever is saved."""
    return bool(breaches) and not bool(acknowledged)


DEFAULT_ADOPTION_LOG = "adoption_history.jsonl"


def adoption_record(candidate: dict, *, ts: str, action: str, method: str,
                    overrides: Optional[dict] = None, breaches=None,
                    source: str = "") -> dict:
    """One durable record of WHAT was adopted/promoted — including any breach
    that was knowingly accepted. Anything the tool decides silently is a defect
    in this project; anything the OPERATOR decides silently is one too, so the
    accepted breach travels with the decision instead of living only in the
    session that made it."""
    br = [str(b) for b in (breaches or [])]
    return {
        "ts": ts,
        "action": action,
        "candidate": str(candidate.get("label") or candidate.get("key") or ""),
        "method": method,
        "overrides": dict(overrides or {}),
        "gates": {g["key"]: g["status"] for g in (candidate.get("gates") or [])},
        "breaches": br,
        "accepted_with_breach": bool(br),
        "source": source,
    }


def append_adoption_log(record: dict,
                        log_path: str = DEFAULT_ADOPTION_LOG) -> None:
    """Append one adoption record as a JSON line. Reuses the optimizer's
    best-effort JSONL writer: a logging failure must never break an adoption,
    and it must never be the reason a plan cannot be saved."""
    from . import optimize
    optimize.append_run_log(record, log_path)


def read_adoption_log(log_path: str = DEFAULT_ADOPTION_LOG, n: int = 20) -> list:
    """The last `n` adoption records (oldest→newest), or [] if none yet."""
    from . import optimize
    return optimize.read_run_log(log_path, n)


# --------------------------------------------------------------------------- #
# Realized-plan audit — judge the PLAN, not an intermediate pass
# --------------------------------------------------------------------------- #
def realized_plan_audit(
    harvest_events,
    transfer_events,
    facility_limits,
    control,
    window_weeks=frozenset(),
) -> list[str]:
    """Warnings measured on the FINAL plan, for the ValidationLog.

    Every other harvest-floor warning in the tool is raised mid-plan, by the
    pass that happened to notice a shortfall — the scheduler warns from its
    own demand pass, before make-room, level-load and the 6N fallback ladder
    have run. Those later passes both FIX weeks it flagged and BREAK weeks it
    did not, so the log ends up describing a plan that was never produced.
    Measured on the 8.13 PR: 29 weeks of the realized plan were under the
    floor, the log named 3 of them, and one of those 3 (2026-W47) was fine in
    the plan that actually shipped. An operator reading that log to answer
    "which weeks are short?" got the wrong answer in both directions.

    This runs last, over the events that were actually emitted, and uses the
    PER-WEEK resolved floor — `scenario/limits.yaml` overrides the Control
    default week by week, and a check against the flat default silently passes
    every week the operator raised.

    Operator-scripted manual-window weeks are EXCLUDED, with a note saying so.
    Those weeks execute only the script, so the planner neither chose nor can
    fix what they harvest, and their harvests are stitched in separately (they
    are not in `harvest_events` at all, so counting them here would report a
    phantom zero). The MANUAL WINDOW lints already police the script itself.

    Returns prefixed strings; `write_validation_log` maps the prefixes to
    categories. Pure measurement — reads the plan, changes nothing.
    """
    from collections import defaultdict
    from .caps import (METRIC_MAX_HARVEST, METRIC_MIN_HARVEST,
                       resolve_facility_cap)
    from .time_grid import iso_week_label

    out: list[str] = []
    window = set(window_weeks or ())

    def _wk(ev):
        d = getattr(ev, "event_date", None)
        return iso_week_label(d) if d is not None else None

    # ---- harvest floor / ceiling, on the realized events -------------------
    hv: dict[str, float] = defaultdict(float)
    for ev in harvest_events or ():
        w = _wk(ev)
        if w:
            hv[w] += float(getattr(ev, "count", 0.0) or 0.0)
    skipped = sorted(w for w in hv if w in window)
    for w in sorted(hv):
        if w in window:
            continue
        got = hv[w]
        floor = resolve_facility_cap(METRIC_MIN_HARVEST, w, facility_limits, control)
        ceil_ = resolve_facility_cap(METRIC_MAX_HARVEST, w, facility_limits, control)
        if floor and got < floor - 1.0:
            short = floor - got
            # Separate a real shortfall from the known ~0.2% mortality-pad
            # artifact: a fill sized to exactly the floor drains a few dozen
            # fish short two weeks later. Both are reported — suppressing one
            # would be the same sin as the old detector — but an operator
            # scanning the category must be able to see which is which, or
            # eight "72 fish" lines teach them to ignore all of them.
            tag = ("  [rounding-scale: under 0.5% of the floor]"
                   if short < max(1.0, 0.005 * floor) else "")
            out.append(
                f"HARVEST FLOOR - {w}: realized harvest {got:,.0f} fish is "
                f"{short:,.0f} under the {floor:,.0f} floor in force this "
                f"week{tag}")
        if ceil_ and got > ceil_ + 1.0:
            out.append(
                f"HARVEST CEILING - {w}: realized harvest {got:,.0f} fish is "
                f"{got - ceil_:,.0f} over the {ceil_:,.0f} weekly processing "
                f"limit")

    if window:
        out.append(
            f"HARVEST FLOOR - scope: {len(window)} operator-scripted manual-"
            f"window week(s) excluded from this audit "
            f"({', '.join(sorted(window)[:6])}"
            f"{', ...' if len(window) > 6 else ''}) — those weeks run only your "
            f"scripted events, so the planner did not choose their harvest; the "
            f"MANUAL WINDOW entries police the script itself"
            + (f". {len(skipped)} of them emitted planner harvest events."
               if skipped else "."))

    # ---- weekly handling budget, on the realized moves ---------------------
    cap = float(getattr(control, "max_transfers_per_week", 0.0) or 0.0)
    if cap > 0:
        from .events import Transfer as _Transfer
        # Same UNIT the planner clamps to (placement `_moves_left`): distinct
        # applied (source, dest) tank pairs — one physical pumping event each.
        # A multi-destination Transfer is as many moves as it has destinations;
        # two legs of the same pair are one move; TranOG/Grade rows are not
        # moves. Counting events instead would report a different number than
        # the budget was enforced against, which is how a "breach" becomes
        # unarguable-looking and wrong.
        mv: dict[str, set] = defaultdict(set)
        for ev in transfer_events or ():
            if not isinstance(ev, _Transfer):
                continue
            if float(getattr(ev, "count_transferred", 0.0) or 0.0) <= 0:
                continue          # refused/no-op events are not handling
            w = _wk(ev)
            if not w:
                continue
            for a in getattr(ev, "destinations", ()) or ():
                if float(getattr(a, "count", 0.0) or 0.0) >= 0.5:
                    mv[w].add((ev.source_tank_id, a.tank_id))
        for w in sorted(mv):
            n = len(mv[w])
            if n > cap:
                out.append(
                    f"HANDLING BUDGET - {w}: {n} moves planned against the "
                    f"{cap:,.0f}-move weekly budget (max_transfers_per_week) — "
                    f"the deferrable quality passes clamp to this budget, so an "
                    f"overrun means ESSENTIAL moves alone (arrival make-room, "
                    f"rotation fills, the plan diff) exceeded it")
    return out
