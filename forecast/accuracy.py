"""FORECAST vs ACTUALS — grade a past forecast against the PR that followed it.

Why this exists
---------------
The suite proves BOOKKEEPING (no fish created or lost, rules respected). It has
never proved the BIOLOGY — whether the growth model matches the facility. The
operator re-anchors every month, so model error never ACCUMULATES; but it was
also never MEASURED, because each month's prediction was replaced instead of
graded.

The measurement already exists and was being discarded:
  * each month's ProductionReport IS the actuals (real per-tank counts and
    weights at a real closing date);
  * the PREVIOUS run's output workbook holds the PREDICTION for that same date
    (the ``BatchLocations`` sheet).
So this module is a COMPARISON, not new modelling, and it needs no new habit
from the operator beyond keeping the previous output workbook.

What is measured, and what is NOT
---------------------------------
BATCH level is the primary, honest view of the BIOLOGY. Fish are compared per
batch, summed over whatever tanks they ended up in, so the operator having
placed fish differently from the plan does not register as model error.

TANK level measures PLAN ADHERENCE, not model error. If the operator moved a
batch to a different tank than the plan proposed, that is a decision, not a bad
prediction. The two are reported separately and labelled, because conflating
them makes the report worse than useless.

Within the batch view the metrics are not equally clean:
  * MEAN WEIGHT is the closest thing to a pure growth-model score. It is the
    headline.
  * COUNT and BIOMASS mix model error with EXECUTION — a batch harvested,
    culled, split or transferred differently than planned moves those numbers
    without the growth model being wrong at all.

Hard limits, restated in ``AccuracyReport.limits`` so they travel with the
numbers:
  * HARVEST EXECUTION IS NOT MEASURABLE HERE. A PR is a snapshot of what is in
    the water; fish already sold are simply absent. A batch harvested early
    looks like a catastrophic count miss and is not one.
  * SEAWATER ONLY. ``BatchLocations`` snapshots the OG tank state; freshwater
    is projected separately and never appears there, while the PR's FW units
    ("PostS.01") have no counterpart. FW model error is measured elsewhere, by
    the auto-calibration history in this same module.
  * OVERLAP ONLY. Anything outside the weeks the old workbook covers, and any
    batch present in only one of the two sources, is reported as coverage —
    never silently folded into an average.

Alignment
---------
A ``BatchLocations`` row is a live tank-state snapshot taken at the END of its
week but LABELLED with that week's start date (``forecast/placement.py``, the
"Snapshot BatchLocations for this week" block). A PR closing date is likewise
an end-of-period state, so closing is compared to closing.

Weekly snapshots rarely land on the PR's closing date, and on this facility the
typical weight error moves about 0.8 percentage points per day of gap —
measured, not assumed: grading one forecast against one PR at three consecutive
weeks returns 0.95%, 6.36%, 13.38%. Grading against a snapshot up to 3.5 days
away would therefore charge the calendar gap to the growth model and swamp the
signal. So the batch (biology) view reads the prediction at the EXACT closing
date, by linear interpolation between the two bracketing week-end snapshots —
a value the forecast already implies between two points it already produced,
not a growth assumption. ``AccuracyReport.basis`` records which weeks were used
and ``sensitivity`` shows what the neighbouring weeks would have said.

The tank (adherence) view is NOT interpolated — occupancy is discrete — so it
uses the single nearest snapshot and discloses ``alignment_offset_days``.

Nothing here mutates anything: both workbooks are opened read-only.
"""
from __future__ import annotations

import json
import statistics
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Optional

# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #


def _as_date(d):
    """date | datetime | None -> date | None."""
    if d is None:
        return None
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    return None


def _pct(err, base):
    """err as a % of base, or None when base is 0 (never a divide-by-zero)."""
    if base in (None, 0) or base == 0.0:
        return None
    return 100.0 * err / base


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


# --------------------------------------------------------------------------- #
# Reading the PREDICTION out of a previous output workbook
# --------------------------------------------------------------------------- #

def _locate_columns(header_row) -> dict:
    """Map the BatchLocations header row to column indices, BY NAME.

    Positional fallbacks match the writer's current order, so this keeps
    working if a column is inserted and still reads the right metric if one is
    renamed. "Week" is pinned by exact match because it is also a prefix of
    "Week_Start" — a prefix search would silently read dates as labels.
    """
    named = {str(c).strip(): i for i, c in enumerate(header_row) if c is not None}

    def pick(*prefixes, default=None):
        for name, i in named.items():
            low = name.lower()
            if any(low.startswith(p.lower()) for p in prefixes):
                return i
        return default

    return {
        "label": named.get("Week", 0),
        "start": pick("Week_Start", default=1),
        "batch": pick("Batch", default=2),
        "tank": pick("Tank", default=3),
        "system": pick("System", default=4),
        "count": pick("Count", default=5),
        "biomass": pick("Biomass", default=7),
        "stage": pick("Stage", default=9),
    }


def read_forecast_locations(src) -> list[dict]:
    """Per-(week, batch, tank) rows from an output workbook's BatchLocations.

    `src` is a path or a file-like object. Columns are located BY NAME with a
    positional fallback (mirrors ``optimize._batchloc_rows``) so a column
    reorder in the writer cannot silently feed this the wrong metric.

    Weight is normalised to GRAMS here: the sheet writes AvgWt in kg while the
    PR reports grams, and that mismatch is the single easiest way to produce a
    confident 1000x wrong answer.
    """
    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    try:
        if "BatchLocations" not in wb.sheetnames:
            raise ValueError(
                "That workbook has no BatchLocations sheet — it does not look "
                "like a forecast output workbook.")
        ws = wb["BatchLocations"]
        rows: list[dict] = []
        cols: Optional[dict] = None
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if cols is None:
                if str(row[0]).strip() == "Week" and any(
                        str(c).strip().startswith("Biomass") for c in row if c):
                    cols = _locate_columns(row)
                continue
            if row[0] is None or not str(row[0]).strip():
                continue

            def _at(i, _row=row):
                return _row[i] if i is not None and len(_row) > i else None

            cnt = _num(_at(cols["count"]))
            bio = _num(_at(cols["biomass"]))
            if cnt <= 0:
                continue
            rows.append({
                "week": str(_at(cols["label"]) or ""),
                "week_start": _as_date(_at(cols["start"])),
                "batch": str(_at(cols["batch"]) or ""),
                "tank": _at(cols["tank"]),
                "system": str(_at(cols["system"]) or ""),
                "count": cnt,
                "biomass_kg": bio,
                "avg_wt_g": (bio / cnt * 1000.0) if cnt > 0 else 0.0,
                "stage": str(_at(cols["stage"]) or ""),
            })
        if not rows:
            raise ValueError(
                "BatchLocations is present but empty — nothing to grade.")
        return rows
    finally:
        wb.close()


def forecast_weeks(rows: list[dict]) -> list[dict]:
    """Distinct forecast weeks, chronological, each with its start AND end date.

    The end date is what a PR closing date is compared against. Week i ends
    where week i+1 starts (the grid allows a short first week when
    forecast_start is not a Monday); the final week is assumed a full 7 days.
    """
    seen: dict[str, date] = {}
    for r in rows:
        ws = r.get("week_start")
        if r["week"] and ws is not None and r["week"] not in seen:
            seen[r["week"]] = ws
    ordered = sorted(seen.items(), key=lambda kv: kv[1])
    out = []
    for i, (label, start) in enumerate(ordered):
        end = ordered[i + 1][1] if i + 1 < len(ordered) else start + timedelta(days=7)
        out.append({"week": label, "start": start, "end": end})
    return out


def align_week(rows: list[dict], closing: date) -> Optional[dict]:
    """The forecast week whose END is nearest `closing` (closing vs closing).

    Returns the week dict plus ``offset_days`` = week end - PR closing (signed;
    positive means the forecast snapshot sits AFTER the actuals date, so it has
    a little extra growth in it). None when there are no weeks at all.

    This names the week the tank-level (plan-adherence) view is taken from, and
    provides provenance for the batch view. The batch view itself does NOT stop
    here — see ``predicted_at_date``.
    """
    weeks = forecast_weeks(rows)
    if not weeks or closing is None:
        return None
    best = min(weeks, key=lambda w: abs((w["end"] - closing).days))
    return dict(best, offset_days=(best["end"] - closing).days)


def bracket_weeks(weeks: list[dict], closing: date) -> tuple:
    """The two forecast weeks whose ENDS bracket `closing`: (before, after).

    Either side is None when `closing` falls outside the horizon.
    """
    before = [w for w in weeks if w["end"] <= closing]
    after = [w for w in weeks if w["end"] >= closing]
    return (max(before, key=lambda w: w["end"]) if before else None,
            min(after, key=lambda w: w["end"]) if after else None)


def _rollup(rows: list[dict], week_label: str) -> dict:
    """{batch: [count, biomass_kg]} for one forecast week."""
    agg: dict[str, list] = {}
    for r in rows:
        if r["week"] != week_label:
            continue
        e = agg.setdefault(r["batch"], [0.0, 0.0])
        e[0] += r["count"]
        e[1] += r["biomass_kg"]
    return agg


def predicted_at_date(fc_rows: list[dict], weeks: list[dict],
                      closing: date) -> tuple[dict, dict]:
    """What the forecast predicted ON the PR's closing date, per batch.

    Weekly snapshots land on week ends, which rarely coincide with a PR closing
    date. On this facility the typical weight error moves ~0.8 percentage
    points per day of gap, so grading against a snapshot up to 3.5 days away
    would swamp the very signal being measured — the alignment gap would be
    charged to the growth model. That is the measurement-bug pattern this
    project has been bitten by repeatedly, so it is closed here: the prediction
    is read off the model's OWN curve at the exact date, by linear
    interpolation between the two bracketing week-end snapshots.

    This is not a growth assumption: it reads a value the forecast already
    implies between two points it already produced. When the date falls outside
    the horizon (or on a week end exactly) it degrades to that single snapshot,
    and ``method`` says which happened.

    Returns ({batch: {"count", "biomass_kg", "wt_g"}}, meta).
    """
    lo, hi = bracket_weeks(weeks, closing)
    if lo is None and hi is None:
        return {}, {"method": "none"}
    if lo is None or hi is None or lo["week"] == hi["week"]:
        wk = lo or hi
        agg = _rollup(fc_rows, wk["week"])
        out = {b: {"count": c, "biomass_kg": m,
                   "wt_g": (m / c * 1000.0) if c > 0 else 0.0}
               for b, (c, m) in agg.items()}
        return out, {
            "method": "snapshot",
            "weeks": [wk["week"]],
            "offset_days": (wk["end"] - closing).days,
        }

    span = (hi["end"] - lo["end"]).days
    t = ((closing - lo["end"]).days / span) if span else 0.0
    a, b = _rollup(fc_rows, lo["week"]), _rollup(fc_rows, hi["week"])
    out = {}
    for bid in set(a) | set(b):
        # A batch present in only one side of the bracket has entered or left
        # the facility mid-span; interpolating across that would invent fish,
        # so its single real snapshot is used unchanged.
        if bid in a and bid in b:
            ca, ma = a[bid]
            cb, mb = b[bid]
            wa = (ma / ca * 1000.0) if ca > 0 else 0.0
            wb_ = (mb / cb * 1000.0) if cb > 0 else 0.0
            out[bid] = {
                "count": ca + (cb - ca) * t,
                "biomass_kg": ma + (mb - ma) * t,
                # Weight is interpolated DIRECTLY, not derived from the
                # interpolated count/biomass: counts step at harvests, and the
                # biology score must not inherit that step.
                "wt_g": wa + (wb_ - wa) * t,
            }
        else:
            c, m = (a.get(bid) or b.get(bid))
            out[bid] = {"count": c, "biomass_kg": m,
                        "wt_g": (m / c * 1000.0) if c > 0 else 0.0}
    return out, {
        "method": "interpolated",
        "weeks": [lo["week"], hi["week"]],
        "week_ends": [lo["end"], hi["end"]],
        "fraction": round(t, 3),
        "offset_days": 0,
    }


# --------------------------------------------------------------------------- #
# Reading the ACTUALS out of a ProductionReport
# --------------------------------------------------------------------------- #

def read_actual_locations(src) -> tuple[Optional[date], list[dict]]:
    """(closing_date, per-(batch, tank) actual rows) from a ProductionReport.

    Reuses the production PR parser so this can never drift from what the
    engine itself hydrates from. Seawater (OG) rows only — see the module
    docstring; the FW records the parser also returns have no counterpart in
    BatchLocations and are deliberately not compared.
    """
    import openpyxl
    from .production_report import read_production_report
    wb = openpyxl.load_workbook(src, read_only=False, data_only=True)
    try:
        from .production_report import find_pr_sheet
        if find_pr_sheet(wb) is None:
            raise ValueError(
                "That workbook has no ProductionReport sheet — it does not "
                "look like a PR.")
        closing, og, _fw = read_production_report(wb)
        rows = [{
            "batch": r.batch_id,
            "tank": r.tank_id,
            "count": float(r.closing_count),
            "biomass_kg": float(r.closing_biomass_kg),
            "avg_wt_g": float(r.closing_avg_wt_g),
        } for r in og if r.closing_count > 0]
        return _as_date(closing), rows
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# The comparison
# --------------------------------------------------------------------------- #

@dataclass
class BatchAccuracy:
    """One batch, predicted vs actual, summed over whatever tanks it is in.

    ``wt_err_pct`` is the biology score. ``count_err_pct`` and
    ``biomass_err_pct`` also carry execution (harvest/cull/transfer) and must
    not be read as pure model error.
    """
    batch_id: str
    present: str = "both"           # "both" | "forecast-only" | "actual-only"
    # True when the FORECAST removed a material share of this batch between the
    # anchor week and the graded week — i.e. the model harvested (or culled) it
    # over the interval. Such a batch is EXECUTION-CONFOUNDED and its weight
    # error is not a clean biology score: a partial harvest takes the BIGGEST
    # fish, so the survivors' mean weight drops for reasons that have nothing
    # to do with the growth model, and the actual facility almost never
    # harvested the same slice on the same day. See summarize_bias, which
    # reports the clean subset alongside the all-batch figure.
    exec_confounded: bool = False
    pred_count: float = 0.0
    act_count: float = 0.0
    pred_biomass_kg: float = 0.0
    act_biomass_kg: float = 0.0
    pred_wt_g: float = 0.0
    act_wt_g: float = 0.0

    @property
    def count_err(self) -> float:
        return self.pred_count - self.act_count

    @property
    def count_err_pct(self):
        return _pct(self.count_err, self.act_count)

    @property
    def biomass_err(self) -> float:
        return self.pred_biomass_kg - self.act_biomass_kg

    @property
    def biomass_err_pct(self):
        return _pct(self.biomass_err, self.act_biomass_kg)

    @property
    def wt_err_g(self) -> float:
        return self.pred_wt_g - self.act_wt_g

    @property
    def wt_err_pct(self):
        return _pct(self.wt_err_g, self.act_wt_g)


@dataclass
class TankAccuracy:
    """One (batch, tank) cell — PLAN ADHERENCE, not model error.

    A mismatch here means the fish are somewhere other than the plan put them.
    That is an operator decision (or a plan the operator improved on), and it
    is reported separately from the biology for exactly that reason.
    """
    batch_id: str
    tank_id: object
    present: str = "both"           # "both" | "forecast-only" | "actual-only"
    pred_count: float = 0.0
    act_count: float = 0.0
    pred_biomass_kg: float = 0.0
    act_biomass_kg: float = 0.0

    @property
    def count_err(self) -> float:
        return self.pred_count - self.act_count


@dataclass
class AccuracyReport:
    """Everything the comparison found, plus what it could not see."""
    actual_closing: Optional[date] = None
    forecast_anchor: Optional[date] = None      # first forecast week start
    aligned_week: str = ""
    aligned_week_end: Optional[date] = None
    alignment_offset_days: int = 0
    elapsed_days: int = 0
    batches: list = field(default_factory=list)
    tanks: list = field(default_factory=list)
    facility: dict = field(default_factory=dict)
    basis: dict = field(default_factory=dict)   # how the prediction was read
    bias: dict = field(default_factory=dict)
    sensitivity: dict = field(default_factory=dict)
    coverage: dict = field(default_factory=dict)
    notes: list = field(default_factory=list)
    limits: list = field(default_factory=list)

    @property
    def elapsed_weeks(self) -> float:
        return round(self.elapsed_days / 7.0, 1)

    @property
    def graded(self) -> list:
        """Batches present in BOTH sources — the only ones that can be scored."""
        return [b for b in self.batches if b.present == "both"]


# Stated in one place so the app, the tests and the report all quote the same
# limits instead of three prose copies drifting apart.
LIMITS = [
    "Harvest execution is NOT graded. A PR shows what is in the water; fish "
    "already sold are absent, so a batch harvested earlier or later than "
    "planned shows up as a count miss that is not a model error.",
    "Seawater only. BatchLocations snapshots OG tanks; freshwater is projected "
    "separately and is in neither side of this comparison. FW model error is "
    "tracked instead by the auto-calibration history.",
    "Overlap only. Batches present in just one source are listed under "
    "coverage and excluded from every average, never quietly averaged in.",
    "Mean weight is the biology score. Count and biomass also move with "
    "harvest, culling, grading and transfers, so they mix model error with "
    "execution.",
]


# A batch whose forecast count falls by more than this over the graded interval
# was HARVESTED or culled by the model, not merely thinned by mortality.
# Weekly mortality runs well under 1%, so a few percent over a multi-week
# interval is comfortably above the noise while still catching a real draw.
_EXEC_DRAW_PCT = 0.05


def compare(forecast_src, actual_src) -> AccuracyReport:
    """Grade the forecast in `forecast_src` against the actuals in `actual_src`.

    `forecast_src` is a previous run's OUTPUT workbook (needs BatchLocations);
    `actual_src` is the ProductionReport whose closing date falls inside that
    forecast's horizon. Both are read-only.
    """
    fc_rows = read_forecast_locations(forecast_src)
    closing, act_rows = read_actual_locations(actual_src)
    if closing is None:
        raise ValueError(
            "That ProductionReport has no parseable 'Closing Month' date, so "
            "there is no date to grade the forecast at.")

    weeks = forecast_weeks(fc_rows)
    # Per-batch forecast counts at the FIRST week and at the graded week. Their
    # difference is how much the MODEL took out over the interval, which is the
    # execution-confound detector (see BatchAccuracy.exec_confounded). Read off
    # the forecast alone, so it needs no second ProductionReport and works for
    # any caller of compare(), not just the backtest driver.
    _fc_anchor_count: dict[str, float] = {}
    _fc_aligned_count: dict[str, float] = {}
    _anchor_wk = weeks[0]["week"] if weeks else None
    anchor = weeks[0]["start"] if weeks else None
    horizon_end = weeks[-1]["end"] if weeks else None
    aligned = align_week(fc_rows, closing)
    if aligned is None:
        raise ValueError("That forecast workbook has no dated weeks to grade.")
    for _r in fc_rows:
        _w, _b, _c = _r.get("week"), _r.get("batch"), _r.get("count") or 0.0
        if _w == _anchor_wk:
            _fc_anchor_count[_b] = _fc_anchor_count.get(_b, 0.0) + _c
        if _w == aligned["week"]:
            _fc_aligned_count[_b] = _fc_aligned_count.get(_b, 0.0) + _c

    rep = AccuracyReport(
        actual_closing=closing,
        forecast_anchor=anchor,
        aligned_week=aligned["week"],
        aligned_week_end=aligned["end"],
        alignment_offset_days=aligned["offset_days"],
        elapsed_days=((closing - anchor).days if anchor else 0),
        limits=list(LIMITS),
    )

    # ---- guards: is this pair even comparable? -----------------------------
    if anchor is not None and closing < anchor:
        rep.notes.append(
            f"The PR closes {anchor - closing} day(s) BEFORE this forecast "
            f"even starts ({anchor:%Y-%m-%d}) — these two files are the wrong "
            f"way round, or from different periods.")
    elif horizon_end is not None and closing > horizon_end:
        rep.notes.append(
            f"The PR closing date is past the end of this forecast's horizon "
            f"({horizon_end:%Y-%m-%d}); the last forecast week was graded "
            f"instead.")
    if rep.elapsed_days <= 0:
        rep.notes.append(
            "Zero elapsed time between the forecast anchor and this PR — this "
            "grades the hand-off, not the growth model.")

    # ---- batch level: the biology view -------------------------------------
    # Read at the EXACT PR closing date, not at the nearest weekly snapshot.
    p_batch, rep.basis = predicted_at_date(fc_rows, weeks, closing)
    a_batch: dict[str, list] = {}
    for r in act_rows:
        e = a_batch.setdefault(r["batch"], [0.0, 0.0])
        e[0] += r["count"]
        e[1] += r["biomass_kg"]

    for bid in sorted(set(p_batch) | set(a_batch)):
        p = p_batch.get(bid)
        ac, ab = a_batch.get(bid, [0.0, 0.0])
        present = ("both" if p is not None and bid in a_batch
                   else "forecast-only" if p is not None else "actual-only")
        _a0 = _fc_anchor_count.get(bid, 0.0)
        _a1 = _fc_aligned_count.get(bid, 0.0)
        _drew = (_a0 > 0.0) and ((_a0 - _a1) / _a0) > _EXEC_DRAW_PCT
        rep.batches.append(BatchAccuracy(
            batch_id=bid, present=present, exec_confounded=_drew,
            pred_count=(p or {}).get("count", 0.0),
            act_count=ac,
            pred_biomass_kg=(p or {}).get("biomass_kg", 0.0),
            act_biomass_kg=ab,
            pred_wt_g=(p or {}).get("wt_g", 0.0),
            act_wt_g=(ab / ac * 1000.0) if ac > 0 else 0.0,
        ))

    # ---- tank level: the plan-adherence view --------------------------------
    # Tanks are NOT interpolated: occupancy is discrete, and a fish is either
    # in tank 44 or it is not. This view is taken from the single nearest
    # weekly snapshot (`aligned_week`), so read it with `alignment_offset_days`.
    if rep.basis.get("method") == "interpolated":
        rep.notes.append(
            f"Batch (biology) figures are read at the exact closing date by "
            f"interpolating between weeks {rep.basis['weeks'][0]} and "
            f"{rep.basis['weeks'][1]}. The tank (adherence) table below is a "
            f"single snapshot from {rep.aligned_week}, "
            f"{rep.alignment_offset_days:+d} days off.")
    elif abs(rep.alignment_offset_days) > 3:
        rep.notes.append(
            f"The PR date sits outside the forecast's weekly grid, so the "
            f"nearest snapshot ({rep.aligned_week}, ending "
            f"{rep.alignment_offset_days:+d} days away) was graded directly — "
            f"that gap is growth this comparison charges to the model.")

    fc_week = [r for r in fc_rows if r["week"] == aligned["week"]]
    p_tank: dict[tuple, list] = {}
    for r in fc_week:
        e = p_tank.setdefault((r["batch"], r["tank"]), [0.0, 0.0])
        e[0] += r["count"]
        e[1] += r["biomass_kg"]
    a_tank: dict[tuple, list] = {}
    for r in act_rows:
        e = a_tank.setdefault((r["batch"], r["tank"]), [0.0, 0.0])
        e[0] += r["count"]
        e[1] += r["biomass_kg"]
    for key in sorted(set(p_tank) | set(a_tank), key=lambda k: (str(k[0]), str(k[1]))):
        pc, pb = p_tank.get(key, [0.0, 0.0])
        ac, ab = a_tank.get(key, [0.0, 0.0])
        rep.tanks.append(TankAccuracy(
            batch_id=key[0], tank_id=key[1],
            present=("both" if key in p_tank and key in a_tank
                     else "forecast-only" if key in p_tank else "actual-only"),
            pred_count=pc, act_count=ac,
            pred_biomass_kg=pb, act_biomass_kg=ab,
        ))

    graded = rep.graded
    matched = sum(1 for t in rep.tanks if t.present == "both")
    rep.coverage = {
        "batches_graded": len(graded),
        "batches_forecast_only": [b.batch_id for b in rep.batches
                                  if b.present == "forecast-only"],
        "batches_actual_only": [b.batch_id for b in rep.batches
                                if b.present == "actual-only"],
        "tank_cells_total": len(rep.tanks),
        "tank_cells_matched": matched,
        "tank_adherence_pct": (100.0 * matched / len(rep.tanks)
                               if rep.tanks else None),
    }

    # ---- facility totals ----------------------------------------------------
    # Totals are over GRADED batches only: folding in a batch that exists on
    # one side alone would book its entire mass as model error.
    pc_t = sum(b.pred_count for b in graded)
    ac_t = sum(b.act_count for b in graded)
    pb_t = sum(b.pred_biomass_kg for b in graded)
    ab_t = sum(b.act_biomass_kg for b in graded)
    rep.facility = {
        "pred_count": pc_t, "act_count": ac_t,
        "count_err": pc_t - ac_t, "count_err_pct": _pct(pc_t - ac_t, ac_t),
        "pred_biomass_kg": pb_t, "act_biomass_kg": ab_t,
        "biomass_err_kg": pb_t - ab_t, "biomass_err_pct": _pct(pb_t - ab_t, ab_t),
        "pred_wt_g": (pb_t / pc_t * 1000.0) if pc_t > 0 else 0.0,
        "act_wt_g": (ab_t / ac_t * 1000.0) if ac_t > 0 else 0.0,
    }
    rep.facility["wt_err_pct"] = _pct(
        rep.facility["pred_wt_g"] - rep.facility["act_wt_g"],
        rep.facility["act_wt_g"])

    rep.bias = summarize_bias(graded)
    rep.sensitivity = _alignment_sensitivity(fc_rows, weeks, aligned, a_batch)
    return rep


def _alignment_sensitivity(fc_rows, weeks, aligned, a_batch) -> dict:
    """How much of the error is just the alignment gap?

    The graded week rarely ends exactly on the PR closing date, and at ~1%/day
    growth a 3-day gap is worth ~3% of weight — enough to be mistaken for model
    error. Rather than assume it away (or silently "correct" it with a growth
    assumption, which would be modelling inside a measurement), this re-scores
    the SAME actuals against the neighbouring forecast weeks and reports the
    spread. The operator can then see directly how much the date gap is worth.
    """
    labels = [w["week"] for w in weeks]
    try:
        i = labels.index(aligned["week"])
    except ValueError:
        return {}

    def typical_for(idx):
        if idx < 0 or idx >= len(weeks):
            return None
        lbl = labels[idx]
        agg: dict[str, list] = {}
        for r in fc_rows:
            if r["week"] != lbl:
                continue
            e = agg.setdefault(r["batch"], [0.0, 0.0])
            e[0] += r["count"]
            e[1] += r["biomass_kg"]
        errs = []
        for bid, (pc, pb) in agg.items():
            if bid not in a_batch or pc <= 0:
                continue
            ac, ab = a_batch[bid]
            if ac <= 0 or ab <= 0:
                continue
            pw, aw = pb / pc * 1000.0, ab / ac * 1000.0
            errs.append(abs(pw - aw) / aw * 100.0)
        if not errs:
            return None
        return {"week": lbl, "week_end": weeks[idx]["end"],
                "typical_wt_err_pct": statistics.median(errs)}

    return {"graded": typical_for(i), "previous": typical_for(i - 1),
            "next": typical_for(i + 1)}


def summarize_bias(graded: list) -> dict:
    """SIGNED bias over graded batches — is the model consistently hot or cold?

    A one-signed error is the finding that matters most: a systematic bias can
    be corrected at source, whereas symmetric scatter is just noise the monthly
    re-anchor already absorbs. So this reports the SIGNED median alongside the
    typical magnitude, and how lopsided the signs are.
    """
    wt_all = [b.wt_err_pct for b in graded if b.wt_err_pct is not None]
    cn = [b.count_err_pct for b in graded if b.count_err_pct is not None]
    if not wt_all:
        return {"n": 0, "verdict": "Not enough overlap to judge bias."}

    # EXECUTION CONFOUND. A batch the model harvested over the interval is not
    # a clean biology score: a partial harvest takes the BIGGEST fish, so the
    # survivors' mean weight falls for reasons unrelated to growth, and the
    # real facility almost never took the same slice on the same day. Judge the
    # bias on the batches the model did NOT draw from, and report both so the
    # gap between them is visible rather than hidden.
    wt_clean = [b.wt_err_pct for b in graded
                if b.wt_err_pct is not None
                and not getattr(b, "exec_confounded", False)]
    _n_conf = len(wt_all) - len(wt_clean)
    # Fall back to the full set only when the clean subset is too small to say
    # anything — and label it, so a confounded number is never read as clean.
    _using_clean = len(wt_clean) >= 3
    wt = wt_clean if _using_clean else wt_all

    med = statistics.median(wt)
    typical = statistics.median([abs(x) for x in wt])
    worst = max(wt, key=abs)
    over = sum(1 for x in wt if x > 0)
    under = sum(1 for x in wt if x < 0)
    share = 100.0 * max(over, under) / len(wt)

    # "Systematic" = the signs agree AND the median is bigger than a rounding
    # wobble. Both conditions matter: 8 of 8 batches at +0.1% is not a finding.
    if share >= 75.0 and abs(med) >= 1.0:
        direction = "HOT (over-predicts growth)" if med > 0 else "COLD (under-predicts growth)"
        verdict = (f"Systematic bias: the model runs {direction}. "
                   f"{max(over, under)} of {len(wt)} batches miss the same way, "
                   f"median {med:+.1f}%.")
    elif abs(med) < 1.0 and share < 75.0:
        verdict = (f"No systematic bias — errors scatter both ways "
                   f"(median {med:+.1f}%, {over} high / {under} low).")
    else:
        verdict = (f"Mixed: median {med:+.1f}% with {over} high / {under} low — "
                   f"leaning but not one-signed.")

    return {
        "n": len(wt),
        "wt_median_signed_pct": med,
        "wt_typical_abs_pct": typical,
        "wt_worst_signed_pct": worst,
        "wt_max_abs_pct": max(abs(x) for x in wt),
        "over_predicted": over,
        "under_predicted": under,
        "one_sided_share_pct": share,
        "count_median_signed_pct": statistics.median(cn) if cn else None,
        "count_typical_abs_pct": (statistics.median([abs(x) for x in cn])
                                  if cn else None),
        # Provenance of the number above, so a confounded reading cannot be
        # mistaken for a clean one.
        "n_all": len(wt_all),
        "n_exec_confounded": _n_conf,
        "bias_basis": ("clean (batches the model did not harvest)"
                       if _using_clean else
                       "ALL batches — too few unharvested to judge separately; "
                       "this figure carries harvest-execution error"),
        "wt_median_signed_pct_all": statistics.median(wt_all),
        "verdict": verdict,
    }


def headline(rep: AccuracyReport) -> dict:
    """The one answer to "how much should I trust this forecast?".

    Typical and worst batch-level weight error, with the elapsed time they were
    measured over — an error only means something next to the horizon it
    accumulated across.
    """
    b = rep.bias
    return {
        "elapsed_days": rep.elapsed_days,
        "elapsed_weeks": rep.elapsed_weeks,
        "batches_graded": b.get("n", 0),
        "typical_wt_err_pct": b.get("wt_typical_abs_pct"),
        "worst_wt_err_pct": b.get("wt_max_abs_pct"),
        "signed_median_pct": b.get("wt_median_signed_pct"),
        # Which batches the headline is computed over. A bias quoted without
        # this is not safe to act on: batches the model harvested carry the
        # grading decision, not the growth model.
        "bias_basis": b.get("bias_basis"),
        "n_exec_confounded": b.get("n_exec_confounded"),
        "verdict": b.get("verdict", ""),
    }


# --------------------------------------------------------------------------- #
# Calibration history — the FW model error the tool already measures every run
# --------------------------------------------------------------------------- #
#
# Every run with auto_calibrate_fw on back-solves each FW batch's fw_correction
# and REWRITES it ("B37: fw_correction 1.000 -> 0.774" = the model grew that
# batch 29% faster than reality). Those numbers scrolled past in stdout and the
# ValidationLog and were never retained, so a correction the model has needed
# every month for six months looked exactly like a one-off. Persisted here, it
# becomes a standing model error to fix at SOURCE rather than re-correct
# monthly.

DEFAULT_CALIB_LOG = "fw_calibration_history.jsonl"


def calibration_record(batch_id: str, *, ts: str, configured, applied, solved,
                       target_wt_g, clamped: bool, converged: bool,
                       lo=None, hi=None, pr_closing=None,
                       source: str = "") -> dict:
    """One durable record of a single fw_correction rewrite.

    `configured` is what the scenario asked for, `applied` what the run used;
    the gap between them, repeated month after month, IS the standing error.
    """
    return {
        "ts": ts,
        "batch": str(batch_id),
        "configured": (round(float(configured), 4)
                       if isinstance(configured, (int, float)) else None),
        "applied": (round(float(applied), 4)
                    if isinstance(applied, (int, float)) else None),
        "solved": (round(float(solved), 4)
                   if isinstance(solved, (int, float)) else None),
        "target_wt_g": (round(float(target_wt_g), 1)
                        if isinstance(target_wt_g, (int, float)) else None),
        "clamped": bool(clamped),
        "converged": bool(converged),
        "clamp_lo": lo, "clamp_hi": hi,
        "pr_closing": (pr_closing.isoformat()
                       if hasattr(pr_closing, "isoformat") else pr_closing),
        "source": source,
    }


def append_calibration_log(records, log_path: str = DEFAULT_CALIB_LOG) -> None:
    """Append calibration records as JSON lines.

    Best-effort by design, matching ``optimize.append_run_log``: a logging
    failure must never break a forecast run. History is a diagnostic, and a
    diagnostic that can take the pipeline down is worse than no diagnostic.
    """
    recs = list(records or [])
    if not recs:
        return
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            for r in recs:
                f.write(json.dumps(r) + "\n")
    except Exception:  # noqa: BLE001 — see docstring
        pass


def read_calibration_log(log_path: str = DEFAULT_CALIB_LOG, n: int = 2000) -> list:
    """The last `n` calibration records (oldest -> newest), or [] if no log."""
    try:
        with open(log_path, encoding="utf-8") as f:
            lines = [ln for ln in f if ln.strip()]
        out = []
        for ln in lines[-n:]:
            try:
                out.append(json.loads(ln))
            except ValueError:
                continue        # one corrupt line must not lose the history
        return out
    except Exception:  # noqa: BLE001
        return []


def calibration_drift(records: list, min_runs: int = 3,
                      gap_threshold: float = 0.05) -> list:
    """Per-batch drift view over the calibration history.

    ``persistent`` marks a batch whose applied correction has sat away from its
    configured value across at least `min_runs` runs — i.e. the model has
    needed the same correction repeatedly. That is a standing model error and
    belongs in the biology config, not in a monthly re-correction.
    """
    by_batch: dict[str, list] = {}
    for r in records or []:
        if r.get("batch"):
            by_batch.setdefault(str(r["batch"]), []).append(r)

    out = []
    for bid, recs in sorted(by_batch.items()):
        applied = [r["applied"] for r in recs
                   if isinstance(r.get("applied"), (int, float))]
        conf = [r["configured"] for r in recs
                if isinstance(r.get("configured"), (int, float))]
        if not applied:
            continue
        med_app = statistics.median(applied)
        med_conf = statistics.median(conf) if conf else None
        gap = (med_app - med_conf) if med_conf is not None else None
        out.append({
            "batch": bid,
            "runs": len(recs),
            "first_seen": recs[0].get("ts"),
            "last_seen": recs[-1].get("ts"),
            "median_applied": round(med_app, 4),
            "latest_applied": applied[-1],
            "median_configured": (round(med_conf, 4)
                                  if med_conf is not None else None),
            "gap": (round(gap, 4) if gap is not None else None),
            "spread": round(max(applied) - min(applied), 4),
            "clamped_runs": sum(1 for r in recs if r.get("clamped")),
            "not_converged_runs": sum(1 for r in recs
                                      if not r.get("converged", True)),
            "persistent": bool(
                len(recs) >= min_runs and gap is not None
                and abs(gap) >= gap_threshold),
        })
    return out
