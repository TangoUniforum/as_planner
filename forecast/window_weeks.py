"""Which output weeks were OPERATOR-SCRIPTED (the manual override window)?

The workbook self-describes its window: since d1b26ea/33dd203 every scripted
manual event writes a loud ValidationLog line ("MANUAL EVENT OK — 2026-W31:
..." / "MANUAL EVENT REFUSED — ..."), and a window week that schedules no
harvest gets a "MANUAL WINDOW — 2026-W33 schedules NO harvest ..." warning.
So the set of window weeks can be recovered from ANY produced workbook with no
caller plumbing — old workbooks (no such rows, or no ValidationLog at all)
simply yield the empty set and nothing changes.

Why callers want this: window weeks execute ONLY the operator's scripted
events — no planning engine may add or trim a harvest there. Compliance gates
judge the PLANNER on the weeks the planner controls, so harvest-compliance
metrics (zero weeks, min week, over-cap/over-ceiling counts) exclude the
window weeks; the separate MANUAL WINDOW lints already police the script
itself. Conservation stays whole-horizon — fish are fish everywhere.

Robustness note: the log lines carry an em-dash that survives some encodings
as '�', so matching keys on the stable parts only — the 'MANUAL EVENT' /
'MANUAL WINDOW' prefix and the ISO week label pattern — never on the dash.
"""
from __future__ import annotations

import re

_WEEK_LABEL_RE = re.compile(r"\b(\d{4}-W\d{2})\b")


def manual_window_weeks(wb_or_path) -> set:
    """Set of week labels (e.g. {"2026-W31", "2026-W32"}) the workbook's
    ValidationLog marks as manual-override-window weeks.

    Accepts an already-open openpyxl workbook (so callers that have one open
    don't pay a second load) or a path. Missing sheet / no manual rows ->
    empty set (old workbooks keep their exact previous behavior)."""
    wb = wb_or_path
    close_after = False
    if not hasattr(wb, "sheetnames"):
        import openpyxl
        wb = openpyxl.load_workbook(wb_or_path, read_only=True, data_only=True)
        close_after = True
    try:
        if "ValidationLog" not in wb.sheetnames:
            return set()
        weeks: set = set()
        for row in wb["ValidationLog"].iter_rows(values_only=True):
            if not row:
                continue
            # Rows are (#, Category, Detail); scan every text cell so a
            # category/detail column shuffle can't blind the reader.
            joined = " ".join(str(c) for c in row if c is not None)
            up = joined.upper()
            if "MANUAL EVENT" in up or "MANUAL WINDOW" in up:
                m = _WEEK_LABEL_RE.search(joined)
                if m:
                    weeks.add(m.group(1))
        return weeks
    finally:
        if close_after:
            wb.close()
