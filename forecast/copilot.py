"""Co-pilot engine — the human-in-the-loop transfer/harvest recommender.

Given the operator's manual override window (scenario/manual_events.yaml), this
runs the planners forward and surfaces the NEXT week's recommended moves:

  * harvest + OG->6N staging  ← the validated CONTROLLER (models the 3-pair 6N
    fallow rotation + dual-limit harvest setpoint), tagged priority 1/2.
  * ranked OG<->OG transfers   ← the GLOBAL optimizer (genuinely optimized layout;
    v1a surfaces the single global-LP plan, v1b adds ranked alternatives).

Respect mode: the operator's manual transfers are FIXED. Both engines already
honor the manual window (they advance through it, then plan forward), so the
recommendation is simply the first week AFTER the window — and the optimizer's
own transfer-minimisation keeps it continuous with the last manual week.

This module is deliberately UI-free (no Streamlit): it is the durable engine a
thin shell renders. `propose_next_week()` returns a plain `Proposal`.
"""
from __future__ import annotations

import tempfile
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path


@dataclass
class Move:
    """One recommended operation for the handoff week."""
    kind: str            # "harvest" | "to_6n" | "og_transfer"
    engine: str          # "controller" | "global-lp"
    priority: int        # 1 harvest/contract · 2 6N staging/cap · 3 transfer/balance
    from_tank: int
    to_tank: int | None
    from_loc: str
    to_loc: str | None
    batch: str
    count: float
    avg_wt_kg: float
    note: str            # human-readable "why"


@dataclass
class TransferOption:
    """One ranked OG<->OG relocation alternative from the global optimizer."""
    label: str
    why: str
    moves: list          # list[Move]


@dataclass
class Proposal:
    week_label: str          # ISO label of the handoff week (e.g. "2026-W39")
    window_week: int         # forecast-relative week# to script approved moves as
    harvest_recs: list       # list[Move] — controller harvests (priority 1)
    sixn_recs: list          # list[Move] — controller OG->6N staging (priority 2)
    transfer_options: list   # list[TransferOption] — global OG<->OG (priority 3)
    warnings: list = field(default_factory=list)

    def is_empty(self) -> bool:
        return not (self.harvest_recs or self.sixn_recs
                    or any(o.moves for o in self.transfer_options))


def _handoff(input_path, config_dir, scenario_dir):
    """(handoff ISO week label, forecast-relative week#, handoff week-start date,
    tank->system, tank->loc).

    The manual window runs weeks 1..N from the PR-derived start; the next week to
    recommend is N+1, whose ISO label filters the planners' output sheets. The
    returned date opens week N+1, so callers can step it +7 days to label the
    look-ahead weeks N+2, N+3, ... off the SAME planner run."""
    from openpyxl import load_workbook
    from forecast.config_io import load_config
    from forecast.manual_events import load_manual_events
    from forecast.production_report import read_production_report
    from forecast.time_grid import iso_week_label
    _c, _t, facility = load_config(str(config_dir))
    wb = load_workbook(str(input_path), data_only=True)
    pc, _og, _fw = read_production_report(wb)
    wb.close()
    fs0 = date(pc.year, pc.month, pc.day) + timedelta(days=1)
    events = load_manual_events(str(scenario_dir),
                                pr_closing=date(pc.year, pc.month, pc.day))
    n = max((e.week or 1) for e in events) if events else 0
    handoff_date = fs0 + timedelta(days=7 * n)
    handoff = iso_week_label(handoff_date)
    tank_sys = {t.tank_id: t.system_id for t in facility.tanks}
    tank_loc = {t.tank_id: t.location_id for t in facility.tanks}
    return handoff, n + 1, handoff_date, tank_sys, tank_loc


def _extract_harvests(wb, wk_label, tank_loc):
    """HarvestPlan rows for `wk_label` -> harvest Moves.
    Columns: Week|Batch|Tank|Count|Gross_AvgWt|Gross_Biomass|HOG_Yield|HOG_AvgWt|HOG_Biomass."""
    if "HarvestPlan" not in wb.sheetnames:
        return []
    out = []
    for r in wb["HarvestPlan"].iter_rows(values_only=True):
        if not r or not isinstance(r[0], str) or "-W" not in r[0] or str(r[0]) != wk_label:
            continue
        if len(r) < 5 or not isinstance(r[3], (int, float)) or not isinstance(r[2], (int, float)):
            continue
        tank = int(r[2])
        avg = r[4] if isinstance(r[4], (int, float)) else 0.0
        out.append(Move(kind="harvest", engine="controller", priority=1,
                        from_tank=tank, to_tank=None,
                        from_loc=tank_loc.get(tank, f"#{tank}"), to_loc=None,
                        batch=str(r[1]), count=float(r[3]), avg_wt_kg=float(avg),
                        note="meets the weekly harvest contract / holds facility caps"))
    return out


def _extract_transfers(wb, wk_label, tank_sys, tank_loc, is6n, *, only_to_6n, engine):
    """TransferPlan rows for `wk_label` -> Moves. `only_to_6n` keeps OG->6N staging;
    else keeps OG<->OG relocations (both ends real non-6N OG tanks).
    Columns: Week|Batch|Type|From_Tank|To_Tank|Count|Avg_Weight|Grade|CV."""
    if "TransferPlan" not in wb.sheetnames:
        return []
    out = []
    for r in wb["TransferPlan"].iter_rows(values_only=True):
        if not r or not isinstance(r[0], str) or "-W" not in r[0] or str(r[0]) != wk_label:
            continue
        if len(r) < 6:
            continue
        try:
            ft, tt = int(r[3]), int(r[4])
        except (TypeError, ValueError):
            continue
        if not isinstance(r[5], (int, float)):
            continue
        avg = r[6] if len(r) > 6 and isinstance(r[6], (int, float)) else 0.0
        from_is6n, to_is6n = ft in is6n, tt in is6n
        if only_to_6n:
            if not (ft in tank_sys and not from_is6n and to_is6n):
                continue
            kind, pri, note = "to_6n", 2, "stage into 6N for harvest (purge rotation)"
        else:
            if not (ft in tank_sys and tt in tank_sys and not from_is6n and not to_is6n):
                continue
            fs, ts = tank_sys.get(ft), tank_sys.get(tt)
            kind, pri = "og_transfer", 3
            note = (f"rebalance within {fs}" if fs == ts
                    else f"move {fs} → {ts} to relieve load")
        out.append(Move(kind=kind, engine=engine, priority=pri,
                        from_tank=ft, to_tank=tt,
                        from_loc=tank_loc.get(ft, f"#{ft}"),
                        to_loc=tank_loc.get(tt, f"#{tt}"),
                        batch=str(r[1]), count=float(r[5]), avg_wt_kg=float(avg),
                        note=note))
    return out


def _short_horizon_config(config_dir, window_n, n_weeks, buffer):
    """(config_dir_to_use, temp_dir_or_None). The co-pilot only surfaces the next
    n_weeks, so running the planners over the FULL config horizon — especially the
    global optimiser's full-horizon placement solve — is minutes of wasted work.

    If (window_n + n_weeks + buffer) is STRICTLY shorter than the config's
    horizon_weeks, copy config_dir to a temp dir with that shortened horizon and
    return it (caller deletes the temp dir); else return config_dir unchanged. The
    forward planner subtracts the manual window (window_n) itself, so the effective
    look-ahead is n_weeks + buffer. The buffer preserves near-term fidelity — the
    6N purge-rotation lead (~2-3 wks), harvest anticipation (~1 wk) and any TranOG
    arrivals just past the window — so the handoff-week recommendations don't change
    (verified in tests/test_copilot_short_horizon)."""
    import os
    import re
    import shutil
    import tempfile

    import yaml
    ctrl_src = os.path.join(str(config_dir), "control.yaml")
    try:
        with open(ctrl_src, encoding="utf-8") as f:
            full = int((yaml.safe_load(f) or {}).get("horizon_weeks") or 0)
    except (OSError, ValueError, TypeError):
        return str(config_dir), None
    want = int(window_n) + int(n_weeks) + int(buffer)
    if full <= 0 or want >= full:
        return str(config_dir), None          # already short enough — leave it alone
    tmp = tempfile.mkdtemp(prefix="as_copilot_cfg_")
    cdir = os.path.join(tmp, "config")
    shutil.copytree(str(config_dir), cdir)
    ctrl_dst = os.path.join(cdir, "control.yaml")
    with open(ctrl_dst, encoding="utf-8") as f:
        text = f.read()
    new_text, n_sub = re.subn(r"(?m)^horizon_weeks:.*$",
                              f"horizon_weeks: {want}", text)
    if n_sub == 0:                            # unexpected layout — fall back to full
        shutil.rmtree(tmp, ignore_errors=True)
        return str(config_dir), None
    with open(ctrl_dst, "w", encoding="utf-8") as f:
        f.write(new_text)
    return cdir, tmp


def propose_upcoming(input_path, config_dir, scenario_dir, *,
                     n_weeks=6, include_global=True, horizon_buffer=20) -> list:
    """Run the planners forward ONCE from the current manual window and return a
    `Proposal` for each of the next `n_weeks` weeks (handoff = index 0, then the
    look-ahead weeks N+2, N+3, ...). Reads the CURRENT scenario/manual_events.yaml
    — the caller must save the operator's edits first. Each planner runs to a
    single throwaway temp workbook that every week is extracted from; no production
    file is touched.

    Because only the next `n_weeks` are surfaced, the planners run over a SHORT
    horizon (manual window + n_weeks + `horizon_buffer`) instead of the full config
    horizon — the global optimiser's full-horizon solve is otherwise minutes of
    wasted work. With the default buffer (20) the HANDOFF (the only approvable week)
    is byte-identical to the full-horizon run and the look-ahead previews stay within
    ~20 fish with no structural change, at ~5x the speed (measured on the live config:
    full ~90s -> ~18s). Smaller buffers are faster but drift the look-ahead more (a
    10-week buffer split one week's 6N staging differently). See _short_horizon_config.

    Only the handoff (index 0) is approvable in Sequential mode — the look-ahead
    weeks are projections from the current plan and will refresh once the nearer
    weeks are scripted (see the app shell)."""
    import shutil

    from openpyxl import load_workbook
    from forecast.run import main as run_pipeline
    from forecast.sixn import SIXN_ALL_TANKS
    from forecast.time_grid import iso_week_label

    handoff, window_week, handoff_date, tank_sys, tank_loc = _handoff(
        input_path, config_dir, scenario_dir)
    is6n = set(SIXN_ALL_TANKS)
    n_weeks = max(1, int(n_weeks))
    weeks = [(iso_week_label(handoff_date + timedelta(days=7 * j)), window_week + j)
             for j in range(n_weeks)]
    warnings: list[str] = []

    # Run the planners over just (manual window + look-ahead + buffer) weeks, not the
    # full config horizon — the co-pilot never reads past week n_weeks.
    plan_cfg, _cfg_tmp = _short_horizon_config(
        config_dir, window_week - 1, n_weeks, horizon_buffer)
    try:
        # 1) CONTROLLER — the validated harvest + 6N-staging recommendation, extracted
        #    for every upcoming week from a single forward run.
        ctrl: dict = {}
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "copilot_controller.xlsm"
            run_pipeline(input_path=str(input_path), output_path=str(out),
                         config_dir=str(plan_cfg), scenario_dir=str(scenario_dir))
            wb = load_workbook(str(out), read_only=True, data_only=True)
            for wk, _ww in weeks:
                ctrl[wk] = (
                    _extract_harvests(wb, wk, tank_loc),
                    _extract_transfers(wb, wk, tank_sys, tank_loc, is6n,
                                       only_to_6n=True, engine="controller"),
                )
            wb.close()

        # 2) GLOBAL-LP — the optimized OG<->OG relocation plan, same single run.
        glob: dict = {}
        if include_global:
            try:
                from tools.run_global_forecast import run_global
                with tempfile.TemporaryDirectory() as td:
                    out = Path(td) / "copilot_global.xlsm"
                    run_global(str(input_path), str(out), str(plan_cfg),
                               str(scenario_dir), optimal=False)
                    wb = load_workbook(str(out), read_only=True, data_only=True)
                    for wk, _ww in weeks:
                        glob[wk] = _extract_transfers(wb, wk, tank_sys, tank_loc, is6n,
                                                      only_to_6n=False, engine="global-lp")
                    wb.close()
            except Exception as e:  # noqa: BLE001 — optimizer is optional; degrade gracefully
                warnings.append(f"global optimizer unavailable "
                                f"({type(e).__name__}: {e}) — harvest/6N only")
    finally:
        if _cfg_tmp:
            shutil.rmtree(_cfg_tmp, ignore_errors=True)

    proposals: list[Proposal] = []
    for wk, ww in weeks:
        harvest_recs, sixn_recs = ctrl.get(wk, ([], []))
        transfer_options: list[TransferOption] = []
        og = glob.get(wk, [])
        if og:
            transfer_options.append(TransferOption(
                label="Global LP",
                why="minimises cap breaches first, then number of transfers",
                moves=og))
        proposals.append(Proposal(
            week_label=wk, window_week=ww,
            harvest_recs=harvest_recs, sixn_recs=sixn_recs,
            transfer_options=transfer_options, warnings=list(warnings)))
    return proposals


def propose_next_week(input_path, config_dir, scenario_dir, *,
                      include_global=True) -> Proposal:
    """Handoff week only — thin wrapper over `propose_upcoming` (see its docstring).
    Kept for callers that want just the next approvable week."""
    return propose_upcoming(input_path, config_dir, scenario_dir,
                            n_weeks=1, include_global=include_global)[0]


def to_manual_events(moves, window_week):
    """Convert approved `Move`s into ManualEvents for `window_week` so they append
    straight into the operator's scripted window (extending it by one week).

    The move's fish count MUST ride on the destination (`ManualDest.count`) for
    og_transfer / og_to_6n: their appliers (_apply_og_transfer / _apply_og_to_6n)
    read only the per-destination count and treat a single count=None dest as
    "all remaining", so an approved PARTIAL move would silently drain the WHOLE
    source tank. `harvest` is the exception — _apply_harvest reads ManualEvent.count
    (None = whole tank), so the count belongs on the event there."""
    from forecast.manual_events import ManualEvent, ManualDest
    evs = []
    for m in moves:
        if m.kind == "harvest":
            evs.append(ManualEvent(type="harvest", week=window_week,
                                   from_tank=m.from_tank, count=m.count))
        elif m.kind == "to_6n":
            evs.append(ManualEvent(type="og_to_6n", week=window_week,
                                   from_tank=m.from_tank, count=m.count,
                                   destinations=[ManualDest(tank=m.to_tank,
                                                            count=m.count)]))
        elif m.kind == "og_transfer":
            evs.append(ManualEvent(type="og_transfer", week=window_week,
                                   from_tank=m.from_tank, count=m.count,
                                   destinations=[ManualDest(tank=m.to_tank,
                                                            count=m.count)]))
    return evs
