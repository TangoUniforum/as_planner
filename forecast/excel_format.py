"""Presentation pass over the finished export workbook.

This runs AFTER every writer has appended its rows, and only ever touches
*presentation* (fonts, fills, number formats, freeze panes, filters, tab
colours). It never reads, writes, moves or rounds a value — so it cannot
change a forecast number, and a formatting bug can only ever look wrong,
never BE wrong.

Why a post-pass instead of styling inside each writer: there are ~30 writer
functions and they would drift apart. One pass that infers each sheet's shape
gives the whole workbook a single visual grammar, and there is exactly one
place to change when the operator wants a different look.

The colour vocabulary is inherited from `write_run_comparison` so the two
don't read as different products:
    dark blue   = title banner
    pale blue   = column header
    red         = a breach (over a cap, a drift, a lost fish)
    amber       = worth a look (non-zero residual, a soft miss)
    green       = clean / OK

Conditional formatting is used in preference to per-cell fills wherever a
rule can express the test. That keeps the file small, keeps the highlight
LIVE if the operator edits a cell, and — the reason that matters here — it
never bakes a threshold into a static colour that a later run would leave
stale.
"""
from __future__ import annotations

from datetime import date, datetime

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- palette (shared with write_run_comparison) ----------
C_TITLE = PatternFill("solid", fgColor="1F4E78")
C_HDR = PatternFill("solid", fgColor="D9E1F2")
C_BAD = PatternFill("solid", fgColor="FFC7CE")
C_WARN = PatternFill("solid", fgColor="FFEB9C")
C_GOOD = PatternFill("solid", fgColor="C6EFCE")

F_TITLE = Font(bold=True, color="FFFFFF", size=13)
F_NOTE = Font(italic=True, color="444444", size=9)
F_HDR = Font(bold=True, color="1F4E78", size=10)
F_BAD = Font(bold=True, color="9C0006")
F_WARN = Font(bold=True, color="9C6500")
F_GOOD = Font(bold=True, color="006100")
F_MONO = Font(name="Consolas", size=9)

_thin = Side(style="thin", color="BFBFBF")
_med = Side(style="medium", color="1F4E78")
B_HDR = Border(bottom=_med)
A_HDR = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

# Sheets whose layout is not a simple table, or which are operator INPUT and
# should keep whatever formatting the operator gave them.
SKIP = frozenset({
    "ProductionReport",   # input: merged, banded, multi-row header
    "TransferTemplate",   # prose reference card, hand-laid-out
    "Control",            # input cells the VBA and the app both write
})

# Tab colours by role, so a 24-sheet workbook is navigable at a glance.
TAB_PLAN = "1F4E78"    # the plan itself
TAB_REPORT = "2E75B6"  # derived reporting
TAB_AUDIT = "C55A11"   # gates and audits
TAB_INPUT = "808080"   # inputs / provenance

TAB_COLOURS = {
    "HarvestPlan": TAB_PLAN, "TransferPlan": TAB_PLAN, "Batch Plan": TAB_PLAN,
    "BatchLocations": TAB_PLAN, "FacilityMap": TAB_PLAN,
    "Daily Harvest Schedule": TAB_PLAN,
    "HarvestReport": TAB_REPORT, "HarvestPlan Report": TAB_REPORT,
    "WeeklyReport": TAB_REPORT, "MonthlyReport": TAB_REPORT,
    "YearlySummary": TAB_REPORT, "BiologyProjection": TAB_REPORT,
    "FeedForecastWeekly": TAB_REPORT, "FeedForecastMonthly": TAB_REPORT,
    "Advisory": TAB_AUDIT, "ValidationLog": TAB_AUDIT,
    "ReconciliationReport": TAB_AUDIT, "InputConservationAudit": TAB_AUDIT,
    "TankContinuityAudit": TAB_AUDIT, "SystemLimitsAudit": TAB_AUDIT,
    "Diagnostics": TAB_AUDIT,
    "RunConfig": TAB_INPUT, "ProductionReport": TAB_INPUT,
    "TransferTemplate": TAB_INPUT,
}

# Text tokens that mean "this row is a breach" / "this row is fine". Matched
# case-sensitively against whole cell values in flag-ish columns.
BAD_TOKENS = (
    "BIOMASS_OVER", "FEED_OVER", "BIO_DRIFT", "COUNT_DRIFT", "DROPPED",
    "REDUCE BIOMASS", "REDUCE FEED", "FAIL", "OVER", "UNMET", "BREACH",
)
WARN_TOKENS = ("FW UNDER plan", "PARTIAL", "NOTE", "WARNING", "pre-start", "SHORT")
GOOD_TOKENS = ("OK", "PASS", "PLACED", "CLEAN")


# ---------- number-format inference ----------

def _number_format(header: str) -> str | None:
    """Pick a display format from the column's own header text.

    Ordered most-specific first: 'Year' and 'Tank' must be tested before the
    generic integer rule or they render as 2,026 and 1,2 — which reads as a
    quantity rather than a label.
    """
    h = (header or "").strip().lower()
    if not h:
        return None

    # Labels that happen to be numeric — never separator-grouped.
    if h in ("year", "tank", "unit", "week", "month", "#", "no", "row"):
        return "0"
    if "date" in h or h.endswith("_start") or h == "weekstart":
        return "yyyy-mm-dd"

    # Rates and ratios, before the generic rules claim them.
    if "fcr" in h or "ratio" in h or "yield" in h:
        return "0.000"
    if "sgr" in h or "sfr" in h or "%/day" in h:
        return "0.0000"
    if "(%)" in h or "_pct" in h or "pct_" in h or "percent" in h or h.endswith("%"):
        return "0.00"
    if "density" in h or "kg/m" in h:
        return "0.0"

    # Average weights: grams are whole numbers, kilos need 3 places to show a
    # 5 g move — the same quantity, two very different useful precisions.
    if "avgwt" in h or "avg_wt" in h or "avg wt" in h or "avg weight" in h or "avg live" in h:
        return "#,##0" if "(g" in h or h.endswith("_g") else "0.000"
    if h.endswith("_g") or "(g)" in h:
        return "#,##0"

    # Everything else that is a quantity.
    if any(k in h for k in (
        "count", "fish", "kg", "bio", "feed", "growth", "harvest", "delta",
        "cap", "limit", "excess", "mort", "cull", "(t)", "residual",
        "transfer", "xfer", "input", "open", "close", "expected", "actual",
    )):
        return "#,##0"
    return None


def _is_flagish(header: str) -> bool:
    h = (header or "").strip().lower()
    return h in ("flag", "bio_flag", "feed_flag", "fw_flag", "status",
                 "advisory", "verdict", "gate", "result", "category", "severity")


def _is_deltaish(header: str) -> bool:
    """Columns where any non-zero value is worth the eye, and a positive one
    is worse (an excess over a cap) than a merely non-zero one (a residual)."""
    h = (header or "").strip().lower()
    return ("delta" in h or "excess" in h or "residual" in h
            or "_check" in h or "at_risk" in h or "over_cap" in h)


# ---------- sheet shape detection ----------

def _find_header_row(ws, max_scan: int = 10):
    """Return the 1-based row index of the column-header row, or None.

    Three shapes exist in this workbook: header on row 1; a title/notes
    preamble then a header around row 4; or a pivot-style sheet whose header
    is a row of PERIODS (FeedForecastMonthly, HarvestPlan Report). All three
    are found by the same test — the widest row of labels with data under it,
    where a label is text or a date. Admitting dates is what catches the
    pivot sheets; a genuine data row is still excluded because it mixes in
    plain numbers.
    """
    best, best_score = None, 0
    limit = min(max_scan, ws.max_row - 1)
    for r in range(1, limit + 1):
        vals = [c.value for c in ws[r]]
        filled = [v for v in vals if v is not None and str(v).strip() != ""]
        if len(filled) < 2:
            continue
        if not all(isinstance(v, (str, datetime, date)) for v in filled):
            continue                      # a data row, not a header
        nxt = [c.value for c in ws[r + 1]]
        if not any(v is not None and str(v).strip() != "" for v in nxt):
            continue                      # nothing underneath it
        if len(filled) > best_score:
            best, best_score = r, len(filled)
    return best


def _header_block_top(ws, header_row: int) -> int:
    """First row of the header BLOCK ending at `header_row`.

    FacilityMap stacks week labels over week-start dates; HarvestPlan Report
    stacks a title over its months. Walking up while the row is still all
    labels AND still roughly as wide keeps a genuine second header row inside
    the block, while leaving a one-cell title or section caption outside it.
    """
    width = sum(1 for c in ws[header_row]
                if c.value is not None and str(c.value).strip() != "")
    top = header_row
    r = header_row - 1
    while r >= 1:
        vals = [c.value for c in ws[r]]
        filled = [v for v in vals if v is not None and str(v).strip() != ""]
        if not filled or not all(isinstance(v, (str, datetime, date)) for v in filled):
            break
        if len(filled) < width * 0.5:
            break                          # a title or caption, not a header
        top = r
        r -= 1
    return top


def _last_data_col(ws, header_row: int) -> int:
    last = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value not in (None, ""):
            last = c
    return last or ws.max_column


# ---------- the pass ----------

def _format_table(ws, header_row: int) -> None:
    ncol = _last_data_col(ws, header_row)
    nrow = ws.max_row
    raw = [ws.cell(header_row, c).value for c in range(1, ncol + 1)]
    headers = [str(v or "") for v in raw]
    # Columns headed by a period (the pivot-style sheets) always hold a
    # quantity, so they get the count format that name-inference can't reach.
    period = {i + 1 for i, v in enumerate(raw) if isinstance(v, (datetime, date))}

    top = _header_block_top(ws, header_row)

    # Title banner + note lines above the header block.
    if top > 1:
        for c in range(1, ncol + 1):
            ws.cell(1, c).fill = C_TITLE
        ws.cell(1, 1).font = F_TITLE
        for r in range(2, top):
            for c in range(1, ncol + 1):
                if ws.cell(r, c).value is not None:
                    ws.cell(r, c).font = F_NOTE

    # Header block. A period header is shown as 'Aug 2026' / '2026-08-14'
    # rather than a raw midnight timestamp.
    for r in range(top, header_row + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.fill = C_HDR
            cell.font = F_HDR
            cell.alignment = A_HDR
            if r == header_row:
                cell.border = B_HDR
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = ("mmm yyyy" if getattr(cell.value, "day", 0) == 1
                                      else "yyyy-mm-dd")
    ws.row_dimensions[header_row].height = 30

    ws.freeze_panes = ws.cell(header_row + 1, 1)
    if nrow > header_row:
        ws.auto_filter.ref = (f"A{header_row}:"
                              f"{get_column_letter(ncol)}{nrow}")

    if nrow <= header_row:
        return

    # Number formats, applied only to cells that hold a number so a text note
    # parked in a numeric column is left legible.
    first, last = header_row + 1, nrow
    for c, head in enumerate(headers, start=1):
        fmt = "#,##0" if c in period else _number_format(head)
        if not fmt:
            continue
        for r in range(first, last + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = fmt

    # Conditional rules. Every one of these tests a value the ENGINE computed
    # (a flag it raised, a delta it measured) — none invents a threshold of
    # its own. Per-tank density caps vary 30..95 kg/m3 by tier, so density is
    # shaded relatively rather than cut at a line that would be wrong for the
    # smolt tanks.
    for c, head in enumerate(headers, start=1):
        col = get_column_letter(c)
        rng = f"{col}{first}:{col}{last}"

        if _is_flagish(head):
            for tok in BAD_TOKENS:
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'ISNUMBER(SEARCH("{tok}",{col}{first}))'],
                    fill=C_BAD, font=F_BAD, stopIfTrue=False))
            for tok in WARN_TOKENS:
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'ISNUMBER(SEARCH("{tok}",{col}{first}))'],
                    fill=C_WARN, font=F_WARN, stopIfTrue=False))
            for tok in GOOD_TOKENS:
                ws.conditional_formatting.add(rng, FormulaRule(
                    formula=[f'EXACT({col}{first},"{tok}")'],
                    fill=C_GOOD, font=F_GOOD, stopIfTrue=False))

        elif _is_deltaish(head):
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=["0"], fill=C_BAD, font=F_BAD))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="lessThan", formula=["0"], fill=C_WARN, font=F_WARN))

        elif "density" in head.lower():
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=70, mid_color="FFEB9C",
                end_type="max", end_color="FFC7CE"))

        elif head.strip().lower() in ("count (fish)", "harvest_count",
                                      "harv_count (fish)"):
            ws.conditional_formatting.add(rng, DataBarRule(
                start_type="num", start_value=0, end_type="max",
                color="8EA9DB", showValue=True))

    # Widths: respect what the writers already chose, only rescue the ones
    # that would clip their own header.
    for c, head in enumerate(headers, start=1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions[letter]
        # A period header stringifies to a 19-char timestamp but displays as
        # 'Aug 2026', so size it to what is actually shown.
        want = 11 if c in period else min(max(len(head) + 3, 9), 24)
        if not dim.width or dim.width < want:
            dim.width = want

    # Print setup — these get printed for the production meeting.
    ws.print_title_rows = f"{top}:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _format_runconfig(ws) -> None:
    """RunConfig is a YAML dump in one column: monospace it and colour the
    comment lines so the seed-vs-derived caveats are readable."""
    ws.column_dimensions["A"].width = 110
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1)
        val = cell.value
        if not isinstance(val, str):
            continue
        cell.font = F_NOTE if val.lstrip().startswith("#") else F_MONO
    ws.freeze_panes = "A2"


def apply_workbook_formatting(wb, skip=SKIP) -> int:
    """Style every table-shaped sheet in `wb`. Returns the number formatted.

    Presentation only — no cell VALUE is read for a decision that changes it,
    and none is written.
    """
    done = 0
    for ws in wb.worksheets:
        colour = TAB_COLOURS.get(ws.title)
        if colour:
            ws.sheet_properties.tabColor = colour

        if ws.title in skip:
            continue
        if ws.title == "RunConfig":
            _format_runconfig(ws)
            done += 1
            continue
        if ws.max_row < 2 or ws.max_column < 2:
            continue

        header_row = _find_header_row(ws)
        if header_row is None:
            continue
        _format_table(ws, header_row)
        done += 1
    return done
