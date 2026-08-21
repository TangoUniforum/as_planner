"""Selectable multi-objective production optimizer — a tool beside the tuner.

Sweeps Control knobs and ranks variants on a SELECTABLE, weighted objective, so
the operator can choose what "best" means for a scenario. The objective is to
WALK THE LINE: hold the production targets (biomass, harvest throughput) close to
their limits AND flat — high, stable utilization with no lumps and no breaches —
not to minimize biomass. Components (all "less is better"):

  biomass_overshoot  peak/weeks of facility biomass OVER its cap          -> no breach
  biomass_var        mean per-system CV + facility weekly swing           -> flat
  biomass_util_gap   (cap - mean biomass)/cap                             -> close to the limit
  harvest_var        weekly-harvest fish CV                               -> flat harvest
  harvest_overshoot  fraction of weeks over max_harvest_per_week (55k)    -> no processing breach
  feed_load          mean daily feed (the one cost target, pushed DOWN)   -> minimize
  feed_var           feed CV + swing                                       -> flat feed
  transfers_per_fish avg tank-to-tank moves a fish experiences            -> minimize handling

HARD GATE (never traded): tank continuity + zero unaccounted count/biomass drift
+ all fish accounted (reused from forecast.tuning._conservation). Conservation-
failing variants are excluded from ranking, normalization, and recommendation.

Reuses forecast.tuning's run harness (`_run_in_tempdir`) and conservation gate —
one run harness, two thin metric layers. The transfer/density trade is REAL (the
rebalancer cuts biomass variability by ADDING transfers), so there is no single
optimum — the operator picks the emphasis; nothing is auto-decided.
"""
from __future__ import annotations

import os
import statistics
from dataclasses import dataclass, field, fields as _dc_fields

import openpyxl
import yaml

from . import tuning
from . import window_weeks

# Bumped whenever the MEANING of a measured metric changes, so caches keyed by
# input content alone also fold this in and never reuse a measurement graded
# under old rules (the 2026-08 lesson: stale cached metrics defeat a metrics
# fix silently). v2 = harvest-compliance counts exclude operator-scripted
# manual-window weeks (forecast.window_weeks).
METRICS_SCHEMA = "metrics-v3-harvest-floor"

# Knob grid: (label, {control-knob: value}); baseline first. Every variant
# inherits the caller's config (both leveling defaults — rebalance_level +
# harvest_level_load — ON) and changes only the listed knobs. Built to span the
# real FEED<->HARVEST trade so the emphasis has spread to choose between:
#   - tran_og_default_tanks is the strongest single lever (3 tanks/arrival spreads
#     feed thinner -> fewer feed breaches, but tightens the facility -> bigger
#     make-room harvest dumps; 2 is the reverse). The sweep MUST carry it.
#   - density_target_pct (packing) and the harvest setpoint/K knobs tune each side.
#   - the two CONTROLS — `density-only` (rebalance_level off) and `reactive-harvest`
#     (harvest_level_load off) — let the sweep VERIFY each leveling default earns
#     its keep rather than assuming it.
# NOTE: variants set knobs EXPLICITLY (both endpoints of a lever), never relying
# on the baseline's current value — else if config already sits at one end the
# trade is invisible (e.g. a config at tran_og=3 makes a "tran_og=3" variant a
# no-op duplicate of baseline). baseline = "where your config is now".
OPT_QUICK_GRID = [
    ("baseline", {}),
    ("tran_og=2", {"tran_og_default_tanks": 2}),   # harvest-favoring end
    ("tran_og=3", {"tran_og_default_tanks": 3}),   # feed-favoring end
    ("density-only", {"rebalance_level": False}),
]
OPT_FULL_GRID = [
    ("baseline", {}),
    # tran_og — the strongest feed<->harvest lever; test BOTH ends explicitly,
    # plus 3 paired with a tighter cap band to try to claw harvest back.
    ("tran_og=2", {"tran_og_default_tanks": 2}),
    ("tran_og=3", {"tran_og_default_tanks": 3}),
    ("tran_og=3,dev=0.005", {"tran_og_default_tanks": 3,
                             "facility_biomass_deviation_pct": 0.005}),
    # packing density
    ("density=0.85", {"density_target_pct": 0.85}),
    ("density=0.95", {"density_target_pct": 0.95}),
    # harvest controller — facility_biomass_deviation_pct is now THE "how close to the
    # cap" knob (the dual-limit setpoint runs one band below the binding cap): a TIGHT
    # band runs closer to the cap (more breach risk), a LOOSE one keeps more headroom.
    # harvest_setpoint_lookahead_weeks is vestigial post-redesign, so it is not swept.
    ("dev=0.005 (tight)", {"facility_biomass_deviation_pct": 0.005}),
    ("dev=0.02 (loose)", {"facility_biomass_deviation_pct": 0.02}),
    ("smooth:K12", {"harvest_smooth_lookahead_weeks": 12}),
    # ---- density relief / consolidation policy -------------------------------
    # Added 2026-08-21. These were hand-tuned on ONE workbook and the reference
    # fixture disagreed with the real workbook about the best values, so they
    # were promoted from module constants to knobs precisely so the search — not
    # a developer's judgement — settles them. Both endpoints are set explicitly,
    # per the convention above, so a lever is not invisible when the live config
    # already sits at one end.
    #
    # chronic_pressure_frac must stay CLEAR of density_relief_pct: when the
    # trigger sat exactly on the level relief drives tanks to, relieved tanks
    # piled up on it and flipped on a 0.01% perturbation. The 0.88 row probes
    # below the relief target on purpose — if it wins, that interaction is
    # worth a second look rather than a straight adoption.
    ("chronic=0.88", {"chronic_pressure_frac": 0.88}),
    ("chronic=0.95", {"chronic_pressure_frac": 0.95}),
    ("chronic_wk=3", {"chronic_pressure_weeks": 3}),
    ("chronic_wk=6", {"chronic_pressure_weeks": 6}),
    # 0 disables anticipatory freeing entirely (urgent over-cap tanks still act)
    # — the honest "is the chronic trigger worth its handling?" control.
    ("chronic_frees=0", {"chronic_max_frees_per_week": 0}),
    ("chronic_frees=2", {"chronic_max_frees_per_week": 2}),
    ("relief=0.85", {"density_relief_pct": 0.85}),
    ("relief=0.95", {"density_relief_pct": 0.95}),
    ("consol=0.75", {"consolidation_fill_pct": 0.75}),
    ("consol=0.85", {"consolidation_fill_pct": 0.85}),
    # rebalancer effort. Both of these are DEFERRABLE-pass budgets, clamped to
    # min(knob, quality budget <= max_transfers_per_week) — so every value at or
    # above 15 produces the identical plan. "balance=60" was therefore a no-op
    # duplicate of baseline (the live config sits at 30), and "varqty=20" was
    # one too (live config 20). Both now sweep values the search can actually
    # tell apart, and varqty gets BOTH endpoints for the same reason
    # rebalance_level does — so the trade shows up wherever the config sits.
    ("balance=8", {"rebalance_balance_budget": 8}),
    ("varqty=off", {"rebalance_varqty_budget": 0}),
    ("varqty=max", {"rebalance_varqty_budget": 15}),
    # leveling controls — test BOTH endpoints explicitly so the trade shows up no
    # matter where the baseline config sits (else, if the config already has
    # rebalance_level off, "density-only" is a no-op duplicate of baseline and the
    # genuinely-different ON case is never tried — same lesson as tran_og's two ends).
    ("density-only", {"rebalance_level": False}),     # OFF endpoint
    ("level-rebalance", {"rebalance_level": True}),    # ON endpoint (feed-leveling)
    ("reactive-harvest", {"harvest_level_load": False}),
    ("level-harvest", {"harvest_level_load": True}),
    ("handling:balance=0", {"rebalance_balance_budget": 0}),
    # LNS placement refinement (opt-in) — lets the optimizer decide, per PR, whether
    # the realized relocate/swap pass beats greedy on the chosen emphasis. No-ops
    # (== greedy) when the facility is capacity-bound; helps when there's tank room.
    ("lns-placement", {"placement_method": "lns"}),
]


def opt_grid_for(quick: bool):
    return OPT_QUICK_GRID if quick else OPT_FULL_GRID


# Objective components in display order. All "less is better".
COMPONENTS = [
    "biomass_overshoot", "biomass_var", "biomass_util_gap",
    "harvest_var", "harvest_overshoot", "feed_load", "feed_var",
    "transfers_per_fish",
    "system_overshoot",    # per-system feed+biomass over-cap (compliance)
    "density_overshoot",   # per-tank density over-cap (compliance)
    "system_peak",         # hottest single (system, week) load — the HOT SPOT
    "crowded_biomass_fraction",  # product QUALITY: grow-out biomass reared over the welfare line
]

# Selectable emphasis presets (component -> weight; 0 drops out).
EMPHASIS_PRESETS = {
    # Lumpiness/fluctuation is the headline complaint, so flatness (var) and
    # not-breaching (overshoot) dominate; closeness-to-the-limit (util_gap) is
    # secondary (don't waste capacity, but never at the cost of breaching).
    "Walk the line": {"biomass_var": 3, "harvest_var": 3,
                      "biomass_overshoot": 2, "harvest_overshoot": 2,
                      "system_overshoot": 2, "density_overshoot": 2,
                      "biomass_util_gap": 1,
                      "feed_load": 0.5, "feed_var": 0.5, "transfers_per_fish": 0.5},
    "Flatten biomass": {"biomass_var": 3, "biomass_overshoot": 2, "harvest_var": 2,
                        "biomass_util_gap": 1, "harvest_overshoot": 1,
                        "system_overshoot": 1, "density_overshoot": 1,
                        "feed_var": 0.5, "feed_load": 0.5, "transfers_per_fish": 0.25},
    "Minimize feed": {"feed_load": 3, "feed_var": 2, "system_overshoot": 2,
                      "biomass_var": 1, "biomass_util_gap": 0.5, "biomass_overshoot": 0.5,
                      "density_overshoot": 1,
                      "harvest_var": 0.5, "harvest_overshoot": 0.5,
                      "transfers_per_fish": 0.5},
    "Minimize handling": {"transfers_per_fish": 3, "biomass_var": 1, "harvest_var": 1,
                          "biomass_util_gap": 1, "biomass_overshoot": 1,
                          "harvest_overshoot": 1, "system_overshoot": 1,
                          "density_overshoot": 1, "feed_load": 0.5, "feed_var": 0.5},
    # Respect caps: minimize ALL over-cap excursions (per-system feed/biomass,
    # per-tank density, facility biomass + harvest) above everything else — the
    # emphasis for judging the leveling knob's compliance trade.
    "Respect caps": {"system_overshoot": 3, "density_overshoot": 3,
                     "biomass_overshoot": 2, "harvest_overshoot": 2,
                     "biomass_var": 1, "harvest_var": 1,
                     "feed_load": 0.5, "feed_var": 0.5,
                     "biomass_util_gap": 0.5, "transfers_per_fish": 0.5},
    # Minimize loads / no hot spots: keep every system's biomass+feed as LOW and
    # EVEN as possible (minimize the peak per-system load + all CVs), minimize feed
    # and handling — and explicitly DROP biomass_util_gap, the "press to the cap"
    # reward, because the goal here is the opposite (run cool, lots of headroom).
    "Minimize loads": {"system_peak": 3, "system_overshoot": 2,
                       "feed_load": 2, "feed_var": 2,
                       "biomass_var": 2, "transfers_per_fish": 2,
                       "harvest_var": 1, "harvest_overshoot": 1,
                       "biomass_overshoot": 1, "density_overshoot": 1,
                       "biomass_util_gap": 0},
    # Product quality / welfare: minimize the biomass reared above the welfare
    # density line above all else, then compliance + flatness — and DROP the
    # "press to the cap" reward (util_gap), because gentler rearing is the goal,
    # not packing. The deliberate counterweight to throughput/footprint.
    "Product quality": {"crowded_biomass_fraction": 3, "density_overshoot": 2,
                        "system_overshoot": 1, "biomass_overshoot": 1,
                        "biomass_var": 1, "harvest_var": 1,
                        "transfers_per_fish": 0.5, "feed_load": 0.5,
                        "feed_var": 0.5, "biomass_util_gap": 0},
    "Balanced": {c: 1 for c in COMPONENTS},
}
DEFAULT_EMPHASIS = "Walk the line"


@dataclass
class Metrics:
    # objective components (less is better)
    biomass_overshoot: float
    biomass_var: float
    biomass_util_gap: float
    harvest_var: float
    harvest_overshoot: float
    feed_load: float
    feed_var: float
    transfers_per_fish: float
    system_overshoot: float
    density_overshoot: float
    system_peak: float
    # display-only context
    overall_peak_biomass: float
    overall_mean_biomass: float
    biomass_cap: float
    system_load: float
    feed_peak: float
    density_peak: float
    # Weeks over the weekly processing LIMIT (max_harvest_per_week) — i.e.
    # weeks that USED the pressure-relief band. Relief is exceptional: the
    # harvest gate WARNs at 1-3 and FAILs beyond.
    weeks_over_harvest_cap: int
    # Weeks above the DERIVED relief ceiling (limit * (1 + harvest_relief_pct))
    # — never legal; any such week FAILs the harvest gate. Subset of
    # weeks_over_harvest_cap.
    weeks_over_relief_ceiling: int = 0
    # HANDLING BUDGET (rule 4): weeks whose TransferPlan 'Transfer' row count
    # exceeds the weekly move cap / the ~80% warn line, + the worst week.
    weeks_moves_over_cap: int = 0
    weeks_moves_warn: int = 0
    moves_week_max: int = 0
    transfers_by_type: dict = field(default_factory=dict)
    per_system: dict = field(default_factory=dict)
    # --- additive comparison metrics: tank usage + inter/intra-system balance ---
    # Grow-out tank FOOTPRINT (OG6N depuration excluded): how many tanks the plan
    # actually occupies FW->OG. Peak = the busiest week, mean = average over weeks.
    tank_footprint_peak: float = 0.0
    tank_footprint_mean: float = 0.0
    # Per-batch tank PATH: distinct grow-out tanks a batch passes through over its
    # life (relocation footprint, complements transfers_per_fish). Mean + worst.
    batch_tank_path_mean: float = 0.0
    batch_tank_path_max: float = 0.0
    # BETWEEN-system balance: per week, spread of biomass/feed ACROSS OG systems
    # (CV + max-min range), reported mean-over-weeks and peak-week. High = one
    # system carrying far more than another (a placement-balance failure).
    between_system: dict = field(default_factory=dict)
    # WITHIN-system balance: per (system, week), spread of biomass/feed ACROSS the
    # tanks of a system (CV + range), aggregated mean + peak. Per-tank feed is the
    # system's reported feed apportioned by biomass x a size-declining rate shape.
    within_system: dict = field(default_factory=dict)
    # --- product-QUALITY (welfare) view of density; OG6N depuration excluded ---
    # crowded_biomass_fraction (objective) = fraction of grow-out biomass-weeks
    # reared ABOVE the welfare density line (~80 kg/m3, below the 95 hard cap).
    # mean_rearing_density = biomass-weighted mean density the product experienced.
    # crowded_fish_weeks = Σ fish x weeks spent over the line (operator-legible).
    crowded_biomass_fraction: float = 0.0
    mean_rearing_density: float = 0.0
    crowded_fish_weeks: float = 0.0
    # --- the "never an empty harvest week" HARD rule, per variant ---
    # Zero-harvest weeks + the emptiest week over the FULL horizon (the
    # 34ecbaf series, zero weeks included). Display/gate-only, never scored.
    # None = measured before these fields existed (an old cache entry): the
    # tournament's probe treats None as UNKNOWN, never as a pass.
    harvest_zero_weeks: "int | None" = None
    harvest_min_week: "float | None" = None
    # --- the CONTRACT FLOOR (min_harvest_per_week), measured but NOT scored ---
    # The steady-harvest contract is a weekly FLOOR, and "not literally zero"
    # (harvest_zero_weeks) is only its degenerate case. Measured 2026-08-12 on
    # the 7.29 PR: across a 40-variant controller search the worst harvest week
    # ranged 7,855..27,462 fish while the objective's only harvest term
    # (harvest_var, a CV) moved 0.233..0.403 with corr(min_week, harvest_var)
    # = +0.04 and corr(min_week, score) = -0.03 — i.e. the emphasis score is
    # statistically BLIND to the floor, and the pool's best-floor plan ranked
    # 36th of 40. These two fields make the floor VISIBLE (to the gate
    # checklist and to tournament.pick_winner's no-regression guard) without
    # silently re-weighting anyone's objective. None = not measured (no floor
    # supplied, or an old cache entry) — never read as "the floor was met".
    #   harvest_weeks_below_floor  planner weeks under min_harvest_per_week
    #   harvest_floor_gap          mean shortfall below the floor as a fraction
    #                              of it (0 = never below; scale-free)
    harvest_weeks_below_floor: "int | None" = None
    harvest_floor_gap: "float | None" = None

    def component(self, name):
        return getattr(self, name)


@dataclass
class OptVariant:
    label: str
    overrides: dict
    metrics: "Metrics"
    dropped: int
    overprod: int
    score: float = 0.0
    norm: dict = field(default_factory=dict)
    # First line of the error when this variant could not be planned (e.g. an
    # infeasible TranOG arrival — the engine's "refuse to drop fish" guard). A
    # failed variant is EXCLUDED from selection (never chosen) but does NOT abort
    # the sweep, so the search still returns the best FEASIBLE variant.
    failed: str | None = None

    @property
    def conservation_ok(self) -> bool:
        return self.failed is None and self.dropped == 0 and self.overprod == 0


@dataclass
class OptRecommendation:
    best_label: str
    emphasis: str
    score: float
    is_capacity_bound: bool
    text: str
    # The winner's actual knobs. Labels are NOT unique — coordinate_descent
    # names each candidate for the one knob it changed that step, so the same
    # label recurs across rounds carrying different accumulated overrides.
    # Callers must apply/save THIS, never a by-label re-lookup, or a round-1
    # partial set can be persisted in place of the winning combination.
    overrides: dict = field(default_factory=dict)
    # Winner-ELIGIBILITY audit (see `eligible_pool`). `guards` maps each guard
    # to 'applied' | 'stood-down' | 'off'; `guard_notes` are the operator-facing
    # sentences already folded into `text` (kept separately so a caller can
    # style them). Empty = every guard passed with nothing to report.
    guards: dict = field(default_factory=dict)
    guard_notes: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Workbook parsing
# --------------------------------------------------------------------------- #
def _table(ws, is_header, is_data):
    """Yield dicts for the data rows under the first row matching `is_header`."""
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            if is_header(row):
                hdr = [str(c) if c is not None else "" for c in row]
            continue
        if is_data(row):
            yield {hdr[i]: row[i] for i in range(min(len(hdr), len(row)))}


def _col_map(ws, is_header):
    """{header_name: col_index} from the first row matching is_header, else {}."""
    for row in ws.iter_rows(values_only=True):
        if is_header(row):
            return {str(c).strip(): i for i, c in enumerate(row) if c is not None}
    return {}


def _find_col(colmap, *prefixes, default=None):
    """Index of the first column whose header starts with any prefix (case-insens).

    Lets the workbook parsers locate columns BY NAME instead of a hard-coded
    position, so a report-writer column reorder doesn't silently feed the
    optimizer the wrong metric. Falls back to `default` if not found."""
    for name, idx in colmap.items():
        low = name.lower()
        if any(low.startswith(p.lower()) for p in prefixes):
            return idx
    return default


def _cv(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    m = statistics.mean(xs) if xs else 0.0
    return (statistics.pstdev(xs) / m) if m else 0.0


def _swing(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    if len(xs) < 2:
        return 0.0
    m = statistics.mean(xs)
    return (statistics.mean(abs(b - a) for a, b in zip(xs, xs[1:])) / m) if m else 0.0


def _is_week(v):
    return isinstance(v, str) and "-W" in v


def _biomass_and_feed(wb):
    """Facility biomass + feed series + caps from Advisory; per-system biomass
    series + caps from SystemLimitsAudit (for the smoothness term)."""
    adv = list(_table(
        wb["Advisory"],
        lambda r: r and r[0] == "Week" and len(r) > 2 and "Total_Biomass" in str(r[2]),
        lambda r: _is_week(r[0])))
    bcol = next((k for k in (adv[0] if adv else {}) if k.startswith("Total_Biomass")), None)
    blim = next((k for k in (adv[0] if adv else {}) if k.startswith("Biomass_Limit")), None)
    fcol = next((k for k in (adv[0] if adv else {}) if k.startswith("Total_Feed")), None)
    flim = next((k for k in (adv[0] if adv else {}) if k.startswith("Feed_Limit")), None)
    bio = [d[bcol] for d in adv if isinstance(d.get(bcol), (int, float))]
    feed = [d[fcol] for d in adv if isinstance(d.get(fcol), (int, float))]
    bcap = next((d[blim] for d in adv if isinstance(d.get(blim), (int, float)) and d[blim] > 0), 0.0)
    fcap = next((d[flim] for d in adv if isinstance(d.get(flim), (int, float)) and d[flim] > 0), 0.0)

    # per-system biomass series (exclude OG6N depuration from the smoothness term)
    per = {}
    if "SystemLimitsAudit" in wb.sheetnames:
        for d in _table(
                wb["SystemLimitsAudit"],
                lambda r: r and r[0] == "Week" and len(r) > 1 and r[1] == "System",
                lambda r: _is_week(r[0])):
            sysid = d.get("System")
            b = d.get("Biomass_kg")
            cap = d.get("Biomass_cap")
            if not isinstance(b, (int, float)):
                continue
            e = per.setdefault(sysid, {"series": [], "cap": cap})
            e["series"].append(b)
    return bio, feed, bcap, fcap, per


def _harvest_weekly_fish(wb):
    """Fish harvested per week, over the FULL horizon — zero weeks included.

    A week in which nothing is harvested writes NO rows to HarvestPlan, so
    keying the series off that sheet alone silently DROPPED blackout weeks
    instead of recording them as 0. Every harvest metric inherited the blind
    spot: `zero_weeks` could never exceed 0, `min_week` could never be 0,
    `weeks_below_min` undercounted, and the crater regression test plus the
    Compare board's "No empty week" gate — the guards for the operator's
    hardest rule, never an empty week — could not see the emptiest possible
    week. Measured on a real PR: two consecutive blackout weeks (2026-W47/W48,
    biomass climbing through the cap meanwhile) reported as zero_weeks=0,
    weeks_below_min 21 instead of 23, min_week 19,070 instead of 0.

    The horizon comes from Advisory (one row per week, written every run). If
    that sheet is missing we fall back to the harvested weeks alone — the old,
    optimistic behaviour — rather than inventing a horizon.
    """
    return _harvest_weekly_series(wb)[1]


def _harvest_weekly_series(wb):
    """(week_labels, fish) — the full-horizon weekly harvest series of
    `_harvest_weekly_fish` WITH its aligned week labels, so callers can tell
    operator-scripted manual-window weeks apart from planner weeks."""
    ws = wb["HarvestPlan"]
    cols = _col_map(ws, lambda r: r and str(r[0]).strip() == "Week"
                    and any(str(c).strip() == "Batch" for c in r if c))
    ci = _find_col(cols, "Count", default=3)   # "Count (fish)"
    weekly = {}
    for row in ws.iter_rows(values_only=True):
        if _is_week(row[0]) and len(row) > ci and isinstance(row[ci], (int, float)):
            weekly[row[0]] = weekly.get(row[0], 0.0) + row[ci]

    horizon = set()
    if "Advisory" in wb.sheetnames:
        for row in wb["Advisory"].iter_rows(values_only=True):
            if row and _is_week(row[0]):
                horizon.add(row[0])
    if not horizon and weekly:
        # Without the Advisory horizon the series only spans HARVESTED weeks,
        # so zero-harvest weeks become invisible to the graders (the pre-fix
        # "optimistic" behaviour). Both engines write Advisory; if it's ever
        # absent the operator must know the empty-week gate is blind.
        print("WARN: Advisory sheet missing — harvest series spans harvested "
              "weeks only; zero-harvest weeks CANNOT be detected for this run")
    weeks = sorted(horizon | set(weekly)) if horizon else sorted(weekly)
    return weeks, [weekly.get(w, 0.0) for w in weeks]


def _transfers_per_fish(wb):
    by_type = {"TranOG": 0.0, "Transfer": 0.0, "Grade": 0.0}
    for d in _table(
            wb["TransferPlan"],
            lambda r: r and r[0] == "Week" and len(r) > 2 and r[2] == "Type",
            lambda r: _is_week(r[0])):
        t = d.get("Type")
        c = d.get("Count (fish)")
        if t in by_type and isinstance(c, (int, float)):
            by_type[t] += c
    total_in = 0.0
    if "InputConservationAudit" in wb.sheetnames:
        for d in _table(
                wb["InputConservationAudit"],
                lambda r: r and r[0] == "Batch" and len(r) > 1 and "Input_Count" in str(r[1]),
                lambda r: isinstance(r[0], str) and len(r[0]) > 1
                          and r[0][0] == "B" and r[0][1:].isdigit()):
            ic = d.get("Input_Count (fish)")
            if isinstance(ic, (int, float)):
                total_in += ic
    tpf = (sum(by_type.values()) / total_in) if total_in else 0.0
    return tpf, by_type


def _weekly_move_counts(wb):
    """Per-week count of TransferPlan 'Transfer' rows — the operator's
    handling-budget unit (TranOG/Grade rows are not moves). A 0-fish row
    (float-residue leg on a workbook written before the writer merged/
    dropped them) is not a move and is skipped."""
    counts: dict[str, int] = {}
    for d in _table(
            wb["TransferPlan"],
            lambda r: r and r[0] == "Week" and len(r) > 2 and r[2] == "Type",
            lambda r: _is_week(r[0])):
        if d.get("Type") != "Transfer":
            continue
        n = d.get("Count (fish)")
        if isinstance(n, (int, float)) and n < 0.5:
            continue
        wk = str(d.get("Week"))
        counts[wk] = counts.get(wk, 0) + 1
    return list(counts.values())


def _system_overshoot(wb):
    """Fraction of (system, week) cells over their per-system biomass OR feed cap.
    The per-system compliance dimension the leveling knob targets — the optimizer
    needs this to 'see' whether leveling actually helps."""
    if "SystemLimitsAudit" not in wb.sheetnames:
        return 0.0
    over = tot = 0
    for d in _table(
            wb["SystemLimitsAudit"],
            lambda r: r and r[0] == "Week" and len(r) > 1 and r[1] == "System",
            lambda r: _is_week(r[0])):
        b = d.get("Biomass_kg"); bc = d.get("Biomass_cap")
        f = d.get("Feed_kg_day"); fc = d.get("Feed_cap")
        has_bc = isinstance(bc, (int, float)) and bc > 0
        has_fc = isinstance(fc, (int, float)) and fc > 0
        if not (has_bc or has_fc):
            continue
        tot += 1
        bover = has_bc and isinstance(b, (int, float)) and b > bc
        fover = has_fc and isinstance(f, (int, float)) and f > fc
        if bover or fover:
            over += 1
    return (over / tot) if tot else 0.0


def _is_purge_row(row, si, sti):
    """True for a BatchLocations row that is off-feed purge/depuration — excluded
    from the density metrics (harvest-size fish held at high density just before
    shipping is expected, not a compliance/welfare breach).

    Discriminates by Stage == STARVE, NOT by System == OG6N: from 2028 the 6N
    MAINS are grow-out production tanks whose fish are reared and fed — excluding
    the whole system would blind the metrics to real crowding there (and an OG
    tank starving in place IS purge, wherever it sits). Workbooks without a Stage
    value fall back to the legacy System == OG6N rule."""
    stg = row[sti] if sti is not None and len(row) > sti else None
    if isinstance(stg, str) and stg.strip():
        return stg.strip().upper() == "STARVE"
    return len(row) > si and row[si] == "OG6N"


def _density_overshoot(wb):
    """Fraction of (tank, week) cells over the ~95 kg/m3 density cap (per-tank
    density compliance — the trade leveling can create)."""
    if "BatchLocations" not in wb.sheetnames:
        return 0.0
    ws = wb["BatchLocations"]
    cols = _col_map(ws, lambda r: r and str(r[0]).strip() == "Week"
                    and any(str(c).strip().startswith("Density") for c in r if c))
    di = _find_col(cols, "Density", default=8)   # "Density (kg/m3)"
    si = _find_col(cols, "System", default=4)
    sti = _find_col(cols, "Stage", default=9)
    over = tot = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row or row[0] is None:
            continue
        # EXCLUDE off-feed purge/depuration rows (see _is_purge_row — consistent
        # with the engine, the app alert, and the smoothness term above).
        if _is_purge_row(row, si, sti):
            continue
        dens = row[di] if len(row) > di else None
        if isinstance(dens, (int, float)):
            tot += 1
            if dens > 95:
                over += 1
    return (over / tot) if tot else 0.0


def _density_peak(wb):
    """The single HIGHEST per-tank density (kg/m3) across BatchLocations (off-feed
    purge rows excluded, see _is_purge_row). A COMPLEMENT to _density_overshoot (a
    fraction): a PEAK surfaces one physically-impossible tank — e.g. a placement
    collapse cramming a whole batch into one tank at 600+ kg/m3 — that a fraction
    dilutes to ~0."""
    if "BatchLocations" not in wb.sheetnames:
        return 0.0
    ws = wb["BatchLocations"]
    cols = _col_map(ws, lambda r: r and str(r[0]).strip() == "Week"
                    and any(str(c).strip().startswith("Density") for c in r if c))
    di = _find_col(cols, "Density", default=8)
    si = _find_col(cols, "System", default=4)
    sti = _find_col(cols, "Stage", default=9)
    peak = 0.0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row or row[0] is None:
            continue
        if _is_purge_row(row, si, sti):
            continue
        dens = row[di] if len(row) > di else None
        if isinstance(dens, (int, float)):
            peak = max(peak, dens)
    return peak


WELFARE_DENSITY_KG_M3 = 80.0   # soft welfare/quality line, below the ~95 hard cap


def _density_quality(wb, welfare=WELFARE_DENSITY_KG_M3):
    """Product-QUALITY view of per-tank density (off-feed purge rows excluded,
    see _is_purge_row — from 2028 the 6N mains rear production fish and DO count).
    Reads the realized density on every BatchLocations (batch, week, tank) row and, vs
    the `welfare` threshold (kg/m3 — a SOFT line below the ~95 hard cap), returns:

      mean_density  — biomass-weighted mean rearing density: the density the
                      PRODUCT actually experienced (lower = gentler rearing =
                      better welfare / flesh quality);
      crowded_fw    — crowded FISH-WEEKS: Σ fish in tank-weeks over the line
                      ("how many fish spent how long crowded");
      crowded_frac  — fraction of grow-out BIOMASS-weeks spent over the line
                      (scale-free — the objective term).

    All three drop when a plan keeps fish lower/longer, which costs throughput
    (fewer fish / more tanks): that trade is exactly the point."""
    if "BatchLocations" not in wb.sheetnames:
        return (0.0, 0.0, 0.0)
    ws = wb["BatchLocations"]
    cols = _col_map(ws, lambda r: r and str(r[0]).strip() == "Week"
                    and any(str(c).strip().startswith("Density") for c in r if c))
    di = _find_col(cols, "Density", default=8)
    si = _find_col(cols, "System", default=4)
    sti = _find_col(cols, "Stage", default=9)
    mi = _find_col(cols, "Biomass", default=7)
    ci = _find_col(cols, "Count", default=5)
    sum_bd = sum_b = crowded_b = crowded_fw = 0.0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row or row[0] is None:
            continue
        if _is_purge_row(row, si, sti):
            continue
        dens = row[di] if len(row) > di else None
        bio = row[mi] if len(row) > mi else None
        if not isinstance(dens, (int, float)) or not isinstance(bio, (int, float)) or bio <= 0:
            continue
        sum_bd += dens * bio
        sum_b += bio
        if dens > welfare:
            crowded_b += bio
            cnt = row[ci] if len(row) > ci and isinstance(row[ci], (int, float)) else 0.0
            crowded_fw += cnt          # each row ≈ one week of occupancy -> fish-weeks
    mean_d = (sum_bd / sum_b) if sum_b else 0.0
    frac = (crowded_b / sum_b) if sum_b else 0.0
    return (mean_d, crowded_fw, frac)


def _system_peak(wb):
    """The single HOTTEST (system, week) load across biomass AND feed, as a
    fraction of cap (OG6N depuration excluded). Minimizing this = no hot spots —
    keep every system's load as low as possible, evenly. Unlike system_overshoot
    (a COUNT of breaches), this is the worst peak, so the optimizer can drive the
    tallest spike down even while it stays under cap."""
    if "SystemLimitsAudit" not in wb.sheetnames:
        return 0.0
    peak = 0.0
    for d in _table(
            wb["SystemLimitsAudit"],
            lambda r: r and r[0] == "Week" and len(r) > 1 and r[1] == "System",
            lambda r: _is_week(r[0])):
        if d.get("System") == "OG6N":
            continue
        b = d.get("Biomass_kg"); bc = d.get("Biomass_cap")
        f = d.get("Feed_kg_day"); fc = d.get("Feed_cap")
        if isinstance(bc, (int, float)) and bc > 0 and isinstance(b, (int, float)):
            peak = max(peak, b / bc)
        if isinstance(fc, (int, float)) and fc > 0 and isinstance(f, (int, float)):
            peak = max(peak, f / fc)
    return peak


def _batchloc_rows(wb):
    """Yield per-tank {week, batch, tank, system, avgwt, biomass} from
    BatchLocations, OG6N depuration EXCLUDED (consistent with the density/system
    metrics) and empty rows skipped. Columns located by name (reorder-safe)."""
    if "BatchLocations" not in wb.sheetnames:
        return
    ws = wb["BatchLocations"]
    cols = _col_map(ws, lambda r: r and str(r[0]).strip() == "Week"
                    and any(str(c).strip().startswith("Biomass") for c in r if c))
    wi = _find_col(cols, "Week", default=0)
    bi = _find_col(cols, "Batch", default=2)
    ti = _find_col(cols, "Tank", default=3)
    si = _find_col(cols, "System", default=4)
    ai = _find_col(cols, "AvgWt", "Avg", default=6)
    mi = _find_col(cols, "Biomass", default=7)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row or row[wi] is None:
            continue
        if len(row) > si and row[si] == "OG6N":
            continue
        bio = row[mi] if len(row) > mi else None
        if not isinstance(bio, (int, float)) or bio <= 0:
            continue
        w = row[ai] if len(row) > ai and isinstance(row[ai], (int, float)) else 0.0
        yield {"week": row[wi], "batch": row[bi] if len(row) > bi else None,
               "tank": row[ti] if len(row) > ti else None,
               "system": row[si] if len(row) > si else None,
               "avgwt": w, "biomass": bio}


def _tank_footprint(wb):
    """Grow-out tank FOOTPRINT (OG6N excluded): (peak, mean) count of distinct
    occupied tanks per week — how many tanks the plan actually uses FW->OG."""
    by_week = {}
    for r in _batchloc_rows(wb):
        by_week.setdefault(r["week"], set()).add(r["tank"])
    counts = [len(s) for s in by_week.values()]
    return (float(max(counts)), statistics.mean(counts)) if counts else (0.0, 0.0)


def _batch_tank_path(wb):
    """Per-batch tank PATH (OG6N excluded): (mean, max) distinct grow-out tanks a
    batch passes through over its life — the relocation footprint per batch."""
    by_batch = {}
    for r in _batchloc_rows(wb):
        by_batch.setdefault(r["batch"], set()).add(r["tank"])
    paths = [len(s) for s in by_batch.values() if s]
    return (statistics.mean(paths), float(max(paths))) if paths else (0.0, 0.0)


def _system_by_week(wb):
    """{week: {system: (biomass_kg, feed_kg_day)}} from SystemLimitsAudit, OG6N
    depuration excluded (a purge hold, not a grow-out load)."""
    out = {}
    if "SystemLimitsAudit" not in wb.sheetnames:
        return out
    for d in _table(
            wb["SystemLimitsAudit"],
            lambda r: r and r[0] == "Week" and len(r) > 1 and r[1] == "System",
            lambda r: _is_week(r[0])):
        sysid = d.get("System")
        if sysid == "OG6N":
            continue
        b = d.get("Biomass_kg"); f = d.get("Feed_kg_day")
        out.setdefault(d.get("Week"), {})[sysid] = (
            b if isinstance(b, (int, float)) else 0.0,
            f if isinstance(f, (int, float)) else 0.0)
    return out


def _mean_peak(xs):
    return (statistics.mean(xs), max(xs)) if xs else (0.0, 0.0)


def _between_system_spread(wb):
    """How EVEN load is ACROSS systems: per week, CV and range (max-min) across the
    OG systems' biomass and feed; mean-over-weeks + peak-week for each."""
    bcv, brng, fcv, frng = [], [], [], []
    for sysmap in _system_by_week(wb).values():
        bios = [v[0] for v in sysmap.values()]
        feeds = [v[1] for v in sysmap.values()]
        if len(bios) >= 2:
            bcv.append(_cv(bios)); brng.append(max(bios) - min(bios))
            fcv.append(_cv(feeds)); frng.append(max(feeds) - min(feeds))
    bcm, bcp = _mean_peak(bcv); brm, brp = _mean_peak(brng)
    fcm, fcp = _mean_peak(fcv); frm, frp = _mean_peak(frng)
    return {"bio_cv_mean": bcm, "bio_cv_peak": bcp,
            "bio_range_mean": brm, "bio_range_peak": brp,
            "feed_cv_mean": fcm, "feed_cv_peak": fcp,
            "feed_range_mean": frm, "feed_range_peak": frp}


def _within_system_variation(wb):
    """How EVEN load is WITHIN a system, across its tanks: for each (system, week),
    CV and range of per-tank biomass and per-tank feed, aggregated mean + peak.
    Per-tank feed = the system's REPORTED feed apportioned across its tanks by
    biomass x a size-declining rate shape (w**-1/3), so it stays calibrated to the
    real system total while reflecting that big-fish tanks eat less per kg."""
    sbw = _system_by_week(wb)                 # {week: {sys: (bio, feed)}}
    cells = {}                                # (system, week) -> [(biomass, avgwt)]
    for r in _batchloc_rows(wb):
        cells.setdefault((r["system"], r["week"]), []).append((r["biomass"], r["avgwt"]))
    bcv, brng, fcv, frng = [], [], [], []
    for (sysid, wk), tanks in cells.items():
        if len(tanks) < 2:
            continue
        bios = [t[0] for t in tanks]
        bcv.append(_cv(bios)); brng.append(max(bios) - min(bios))
        sys_feed = sbw.get(wk, {}).get(sysid, (0.0, 0.0))[1]
        shape = [b * (w ** (-1.0 / 3.0) if w and w > 0 else 0.0) for b, w in tanks]
        tot = sum(shape)
        if sys_feed > 0 and tot > 0:
            feeds = [sys_feed * s / tot for s in shape]
            fcv.append(_cv(feeds)); frng.append(max(feeds) - min(feeds))
    bcm, bcp = _mean_peak(bcv); brm, brp = _mean_peak(brng)
    fcm, fcp = _mean_peak(fcv); frm, frp = _mean_peak(frng)
    return {"bio_cv_mean": bcm, "bio_cv_peak": bcp,
            "bio_range_mean": brm, "bio_range_peak": brp,
            "feed_cv_mean": fcm, "feed_cv_peak": fcp,
            "feed_range_mean": frm, "feed_range_peak": frp}


def metrics_from_workbook(out_path, harvest_cap,
                          welfare_density=WELFARE_DENSITY_KG_M3,
                          relief_ceiling=None,
                          move_cap=None, min_harvest=None) -> tuple["Metrics", int, int]:
    """`harvest_cap` = the weekly processing LIMIT (max_harvest_per_week);
    weeks over it are relief-band weeks. `relief_ceiling` = the derived
    absolute ceiling (limit * (1 + harvest_relief_pct)) for the never-legal
    count; None (legacy callers) counts ceiling breaches against the limit.
    `move_cap` = the weekly handling budget (15) for the move-count fields;
    None leaves them 0 (legacy callers). `min_harvest` = the contract FLOOR
    (min_harvest_per_week) for the floor fields; None (legacy callers) leaves
    them None, which every reader must treat as UNKNOWN, never as a pass."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    bio, feed, bcap, fcap, per = _biomass_and_feed(wb)
    hv_weeks, fish = _harvest_weekly_series(wb)
    # Operator-scripted manual-window weeks (self-described by the workbook's
    # ValidationLog; empty set on old workbooks). Window weeks execute ONLY
    # the operator's scripted events — no engine may add or trim a harvest
    # there — so the harvest-COMPLIANCE counts below (zero weeks, min week,
    # over-limit / over-ceiling weeks) judge the planner on planner weeks
    # only. Everything else (totals, variability, conservation) stays
    # whole-horizon. Reuses the already-open workbook — no second load.
    _window = window_weeks.manual_window_weeks(wb)
    planner_fish = ([x for wk, x in zip(hv_weeks, fish) if wk not in _window]
                    if _window else fish)
    tpf, by_type = _transfers_per_fish(wb)
    _mv_counts = _weekly_move_counts(wb)

    peak_b = max(bio) if bio else 0.0
    mean_b = statistics.mean(bio) if bio else 0.0
    overshoot = 0.0
    util_gap = 0.0
    system_load = 0.0
    if bcap > 0:
        overshoot = max(0.0, peak_b - bcap) / bcap \
            + (sum(1 for x in bio if x > bcap) / len(bio) if bio else 0.0)
        util_gap = max(0.0, (bcap - mean_b) / bcap)
    sys_cvs = []
    per_out = {}
    for sysid, e in per.items():
        s = e["series"]
        cap = e["cap"] if isinstance(e["cap"], (int, float)) else 0.0
        mean_s = statistics.mean(s) if s else 0.0
        peak_s = max(s) if s else 0.0
        cv_s = _cv(s)
        per_out[sysid] = {"mean": mean_s, "peak": peak_s, "cv": cv_s, "cap": cap}
        if sysid != "OG6N":
            sys_cvs.append(cv_s)
            if cap > 0:
                system_load = max(system_load, peak_s / cap)
    biomass_var = (statistics.mean(sys_cvs) if sys_cvs else 0.0) + _swing(bio)

    fp_peak, fp_mean = _tank_footprint(wb)
    path_mean, path_max = _batch_tank_path(wb)
    between = _between_system_spread(wb)
    within = _within_system_variation(wb)
    mean_rear_d, crowded_fw, crowded_frac = _density_quality(wb, welfare_density)

    metrics = Metrics(
        biomass_overshoot=overshoot,
        biomass_var=biomass_var,
        biomass_util_gap=util_gap,
        harvest_var=_cv(fish),
        harvest_overshoot=(sum(1 for x in fish if x > harvest_cap) / len(fish)) if fish else 0.0,
        feed_load=statistics.mean(feed) if feed else 0.0,
        feed_var=_cv(feed) + _swing(feed),
        transfers_per_fish=tpf,
        system_overshoot=_system_overshoot(wb),
        density_overshoot=_density_overshoot(wb),
        system_peak=_system_peak(wb),
        overall_peak_biomass=peak_b,
        overall_mean_biomass=mean_b,
        biomass_cap=bcap,
        system_load=system_load,
        feed_peak=max(feed) if feed else 0.0,
        density_peak=_density_peak(wb),
        weeks_over_harvest_cap=sum(1 for x in planner_fish if x > harvest_cap),
        weeks_over_relief_ceiling=sum(
            1 for x in planner_fish if x > (relief_ceiling or harvest_cap)),
        weeks_moves_over_cap=(
            sum(1 for n in _mv_counts if n > move_cap) if move_cap else 0),
        weeks_moves_warn=(
            sum(1 for n in _mv_counts if n > int(0.8 * move_cap))
            if move_cap else 0),
        moves_week_max=(max(_mv_counts) if _mv_counts else 0),
        transfers_by_type=by_type,
        per_system=per_out,
        tank_footprint_peak=fp_peak,
        tank_footprint_mean=fp_mean,
        batch_tank_path_mean=path_mean,
        batch_tank_path_max=path_max,
        between_system=between,
        within_system=within,
        crowded_biomass_fraction=crowded_frac,
        mean_rearing_density=mean_rear_d,
        crowded_fish_weeks=crowded_fw,
        harvest_zero_weeks=sum(1 for x in planner_fish if x < 1.0),
        harvest_min_week=(min(planner_fish) if planner_fish else 0.0),
        harvest_weeks_below_floor=(
            sum(1 for x in planner_fish if x < float(min_harvest))
            if (min_harvest and planner_fish) else None),
        harvest_floor_gap=(
            sum(max(0.0, float(min_harvest) - x) for x in planner_fish)
            / (float(min_harvest) * len(planner_fish))
            if (min_harvest and planner_fish) else None),
    )
    dropped, overprod = tuning._conservation(out_path)
    return metrics, dropped, overprod


# --------------------------------------------------------------------------- #
# Scoring
# --------------------------------------------------------------------------- #
def weights_for(emphasis, custom=None) -> dict:
    if custom:
        return {c: float(custom.get(c, 0.0)) for c in COMPONENTS}
    base = EMPHASIS_PRESETS.get(emphasis, EMPHASIS_PRESETS[DEFAULT_EMPHASIS])
    return {c: float(base.get(c, 0.0)) for c in COMPONENTS}


def score_variants(variants, weights) -> None:
    """Fill each variant's .norm and .score in place. Only conservation-OK
    variants participate in normalization (a rejected variant never skews
    another's score). Normalization is per-component max over OK variants."""
    ok = [v for v in variants if v.conservation_ok]
    maxima = {c: max((v.metrics.component(c) for v in ok), default=0.0) for c in COMPONENTS}
    for v in variants:
        v.norm = {c: (v.metrics.component(c) / maxima[c]) if maxima[c] > 0 else 0.0
                  for c in COMPONENTS}
        v.score = sum(weights.get(c, 0.0) * v.norm[c] for c in COMPONENTS)


def _as_float(x):
    """`x` as a float, or None when it isn't one (an unusable baseline)."""
    if x is None:
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def eligible_pool(variants, baseline_min_week=None):
    """The variants ELIGIBLE to win, and an audit of which guards were in force.

    This is Optimize's half of the winner-eligibility rules the tuned
    tournament already applies in `tournament.pick_winner`. The predicates
    themselves are IMPORTED from tournament, never re-implemented — two copies
    of a business rule drift, and a drifted copy is how an exemption survives.

    Applied in the operator's priority order, each STANDING DOWN rather than
    emptying the pool (a search must still return its best, and the caller
    reports which stood down):

      1. hard gates — `variant_hard_ok`: conserves fish AND never plans an
         empty harvest week.
      2. `ceiling_eligible` — no week above the derived relief ceiling. The
         objective protects the processing limit with a single term whose
         weight is 0 in the shipped "Product quality" preset, so a weight
         cannot be the enforcement.
      3. `floor_eligible` — no worse leanest harvest week than
         `baseline_min_week`. The emphasis score is statistically blind to the
         floor (corr = -0.03 over a 40-variant pool), so it must be a rank.

    An UNMEASURED metric is UNKNOWN, never a pass — that rule lives inside the
    imported predicates.

    Returns `(pool, guards)` where guards maps 'hard'/'ceiling'/'floor' to
    'applied' (the guard chose the pool), 'stood-down' (nothing cleared it, so
    it was dropped) or 'off' (not armed — only the floor, which needs a
    baseline)."""
    from . import tournament
    guards = {"hard": "off", "ceiling": "off", "floor": "off"}
    ok = [v for v in variants if v.conservation_ok]
    if not ok:
        return [], guards

    full = [v for v in ok if tournament.variant_hard_ok(v) is True]
    pool = full or ok
    guards["hard"] = "applied" if full else "stood-down"

    ceil = tournament.ceiling_eligible(pool)
    pool = ceil or pool
    guards["ceiling"] = "applied" if ceil else "stood-down"

    base = _as_float(baseline_min_week)
    if base is not None:
        flr = tournament.floor_eligible(pool, base)
        pool = flr or pool
        guards["floor"] = "applied" if flr else "stood-down"
    return pool, guards


def ineligibility_reasons(v, baseline_min_week=None) -> list:
    """Plain-language reasons `v` may not be crowned — the sentences the
    recommendation quotes when a guard excludes the best-scoring candidate.
    Silence about a rejected candidate is the failure mode this project keeps
    closing, so the reason travels with the decision. Empty = eligible."""
    from . import tournament
    out = []
    hard = tournament.variant_hard_ok(v)
    if hard is False:
        if not v.conservation_ok:
            out.append("it does not conserve fish (dropped/over-produced, or the "
                       "engine refused to plan it)")
        else:
            _z = getattr(v.metrics, "harvest_zero_weeks", None)
            out.append(f"it plans {int(_z)} empty harvest week(s) — the "
                       f"steady-harvest contract is a hard rule")
    elif hard is None:
        out.append("its empty-week count was never measured (unknown is never "
                   "read as a pass)")

    n = getattr(v.metrics, "weeks_over_relief_ceiling", None)
    if n is None:
        out.append("its relief-ceiling breaches were never measured (unknown is "
                   "never read as a pass)")
    elif int(n) > 0:
        out.append(f"it breaches the relief ceiling in {int(n)} week(s) — over "
                   f"the processing limit's relief band, never legal")

    base = _as_float(baseline_min_week)
    if base is not None:
        mw = getattr(v.metrics, "harvest_min_week", None)
        if mw is None:
            out.append("its worst harvest week was never measured (unknown is "
                       "never read as a pass)")
        elif float(mw) < base:
            out.append(f"it lowers the worst harvest week to {float(mw):,.0f} "
                       f"fish, below baseline's {base:,.0f}")
    return out


def recommend(variants, emphasis=DEFAULT_EMPHASIS, weights=None) -> OptRecommendation:
    """The Optimize winner: emphasis-best among the ELIGIBLE variants.

    Selection is NOT score alone. `eligible_pool` applies the same
    winner-eligibility guards the tuned tournament uses (conservation +
    never-an-empty-week, then the relief ceiling, then the contract floor vs
    the `baseline` variant), because this winner is APPLIED — the app and
    tools/auto_optimize.py can write it straight into control.yaml, where it
    becomes the standing config for every later run. A plan that breaks a hard
    rule must never get there.

    When a guard excludes the best-scoring candidate, or when a guard stands
    down because nothing cleared it, `text` says so by name."""
    w = weights or weights_for(emphasis)
    score_variants(variants, w)
    ok = [v for v in variants if v.conservation_ok]
    if not ok:
        return OptRecommendation("(none)", emphasis, float("inf"), False,
                                 "No variant held conservation — investigate before optimizing.",
                                 {}, {"hard": "off", "ceiling": "off", "floor": "off"}, [])

    def _rank(v):
        return (v.score, 0 if v.label == "baseline" else 1, v.label)

    baseline = next((v for v in variants if v.label == "baseline"), None)
    # The floor guard's reference is the method's own un-tuned run. Here that
    # is the `baseline` variant — the one `recommend` already looks up. A
    # seeded coordinate descent labels its seed "seed" and so has none: the
    # guard then stays OFF rather than inventing a baseline. An infeasible
    # baseline carries the _infeasible_metrics sentinel, which is not a
    # measurement either.
    floor_base = (getattr(baseline.metrics, "harvest_min_week", None)
                  if (baseline is not None and baseline.conservation_ok) else None)

    pool, guards = eligible_pool(ok, floor_base)
    top = min(ok, key=_rank)          # what score ALONE would have crowned
    best = min(pool, key=_rank)

    notes = []
    if best is not top:
        why = ineligibility_reasons(top, floor_base) or [
            "it failed a winner-eligibility guard"]
        notes.append(
            f"⚠ The best-scoring variant `{top.label}` (score {top.score:.3f}) "
            f"was EXCLUDED: {'; '.join(why)}.")
    for key, msg in (
            ("hard", "no variant avoided an empty harvest week"),
            ("ceiling", "no variant stayed inside the relief ceiling"),
            ("floor", "no variant held the baseline's worst harvest week")):
        if guards.get(key) == "stood-down":
            notes.append(f"⚠ The {key} guard STOOD DOWN — {msg}, so the winner "
                         f"does not clear it. Read the gate checklist before "
                         f"adopting or saving these knobs.")
    # Is the baseline itself a legal option? If a guard threw it out, saying
    # "baseline stands" would recommend an ineligible plan.
    baseline_eligible = (baseline is not None and baseline.conservation_ok
                         and any(v is baseline for v in pool))
    if (baseline is not None and baseline.conservation_ok
            and not baseline_eligible and baseline is not top):
        notes.append("⚠ The baseline config itself is ineligible: "
                     + "; ".join(ineligibility_reasons(baseline, floor_base))
                     + ".")

    capacity_bound = (baseline_eligible and best.score >= baseline.score - 1e-9)
    if capacity_bound:
        text = (f"No variant beats baseline on the '{emphasis}' objective — the "
                "remaining lumpiness/over-cap is a stocking/capacity limit, not a "
                "knob (see USER_GUIDE). Baseline stands.")
    else:
        # `baseline` is absent when the caller seeded the search (coordinate_descent
        # labels a non-empty seed "seed"), so quote it only when it exists.
        _vs = f" vs baseline {baseline.score:.3f}" if baseline is not None else ""
        text = (f"Best for '{emphasis}': {best.label} (score {best.score:.3f}{_vs}). "
                f"Use the Apply & verify panel below "
                f"to run it now (or paste the knobs into Configure → Control to keep them).")
    if notes:
        text = text + " " + " ".join(notes)
    return OptRecommendation(best.label, emphasis, best.score, capacity_bound, text,
                             dict(best.overrides), guards, notes)


# --------------------------------------------------------------------------- #
# The sweep
# --------------------------------------------------------------------------- #
def _harvest_cap(config_dir, overrides):
    """Effective max_harvest_per_week (fish/week) for a variant."""
    cap = 55000.0
    try:
        with open(os.path.join(config_dir, "control.yaml")) as f:
            cap = float(yaml.safe_load(f).get("max_harvest_per_week", cap) or cap)
    except (OSError, ValueError, TypeError) as e:
        print(f"WARN: control.yaml unreadable for grading "
              f"({type(e).__name__}) — judging vs default harvest cap {cap:g}")
    if "max_harvest_per_week" in overrides:
        cap = float(overrides["max_harvest_per_week"])
    return cap


def _move_cap(config_dir, overrides):
    """Effective max_transfers_per_week (moves/week) for a variant — the
    handling budget. None when 0/unset (off)."""
    cap = 15.0
    try:
        with open(os.path.join(config_dir, "control.yaml")) as f:
            cap = float(yaml.safe_load(f).get("max_transfers_per_week", cap) or 0)
    except (OSError, ValueError, TypeError) as e:
        print(f"WARN: control.yaml unreadable for grading "
              f"({type(e).__name__}) — judging vs default move cap {cap:g}")
    if "max_transfers_per_week" in overrides:
        try:
            cap = float(overrides["max_transfers_per_week"])
        except (ValueError, TypeError):
            pass
    return int(cap) if cap > 0 else None


def _min_harvest(config_dir, overrides):
    """Effective min_harvest_per_week (fish/week) for a variant — the steady-
    harvest CONTRACT floor. None when 0/unset (no floor configured), in which
    case the floor metrics stay None (unknown) rather than reporting a pass."""
    mn = 30000.0
    try:
        with open(os.path.join(config_dir, "control.yaml")) as f:
            mn = float(yaml.safe_load(f).get("min_harvest_per_week", mn) or 0)
    except (OSError, ValueError, TypeError) as e:
        print(f"WARN: control.yaml unreadable for grading "
              f"({type(e).__name__}) — judging vs default harvest floor {mn:g}")
    if "min_harvest_per_week" in overrides:
        try:
            mn = float(overrides["min_harvest_per_week"] or 0)
        except (ValueError, TypeError):
            pass
    return mn if mn > 0 else None


def _relief_ceiling(config_dir, overrides):
    """Derived absolute relief ceiling (fish/week) for a variant:
    max_harvest_per_week * (1 + harvest_relief_pct). None when the relief
    band is off (pct 0/unset) — ceiling breaches then count vs the limit."""
    pct = 0.10                              # models.ControlParams default
    try:
        with open(os.path.join(config_dir, "control.yaml")) as f:
            v = yaml.safe_load(f).get("harvest_relief_pct", pct)
            pct = float(v) if v is not None else 0.0
    except (OSError, ValueError, TypeError) as e:
        print(f"WARN: control.yaml unreadable for grading "
              f"({type(e).__name__}) — judging vs default relief pct {pct:g}")
    if "harvest_relief_pct" in overrides:
        try:
            pct = float(overrides["harvest_relief_pct"])
        except (ValueError, TypeError):
            pass
    if pct <= 0:
        return None
    return _harvest_cap(config_dir, overrides) * (1.0 + pct)


def _welfare_density(config_dir, overrides):
    """Effective welfare density line (kg/m3) for a variant — the quality metric's
    soft threshold, from control.yaml (or a sweep override), default 80."""
    wl = WELFARE_DENSITY_KG_M3
    try:
        with open(os.path.join(config_dir, "control.yaml")) as f:
            wl = float(yaml.safe_load(f).get("density_welfare_threshold_kg_m3", wl) or wl)
    except (OSError, ValueError, TypeError) as e:
        print(f"WARN: control.yaml unreadable for grading "
              f"({type(e).__name__}) — judging vs default welfare line {wl:g}")
    if "density_welfare_threshold_kg_m3" in overrides:
        # `or default`: a 0/None override resolves to the default 80, matching
        # the config path above and every app surface — previously an explicit
        # 0 here (and only here) meant "everything is crowded".
        wl = float(overrides["density_welfare_threshold_kg_m3"]
                   or WELFARE_DENSITY_KG_M3)
    return wl


def _infeasible_metrics() -> "Metrics":
    """A sentinel Metrics for a variant that FAILED (couldn't be planned): every
    objective component is a huge finite value so it can never win, and the
    display fields are zeroed. The variant is also flagged `failed` and excluded
    from selection via conservation_ok — this object only keeps scoring/display
    from crashing on a variant that never produced a workbook. Built by
    introspecting the dataclass so it survives new Metrics fields."""
    big = 1e18
    kw = {}
    for f in _dc_fields(Metrics):
        if f.name in COMPONENTS:
            kw[f.name] = big
        elif f.name in ("transfers_by_type", "per_system",
                        "between_system", "within_system"):
            continue                     # use the default_factory (empty dict)
        elif f.name in ("weeks_over_harvest_cap", "weeks_over_relief_ceiling",
                        "weeks_moves_over_cap", "weeks_moves_warn",
                        "moves_week_max"):
            kw[f.name] = 0
        elif f.name in ("harvest_zero_weeks", "harvest_min_week",
                        "harvest_weeks_below_floor", "harvest_floor_gap"):
            kw[f.name] = None      # unknown, NOT "passes the empty-week/floor rule"
        else:
            kw[f.name] = 0.0
    return Metrics(**kw)


def run_variant(label, overrides, config_dir, scenario_dir, input_path) -> OptVariant:
    # A single infeasible/errored variant must NOT abort the whole sweep — the
    # optimizer's job is to find the best FEASIBLE variant, so a variant that the
    # engine refuses to plan (e.g. a TranOG arrival with no free tanks — the
    # "refuse to drop fish silently" guard) is caught, recorded as `failed`, and
    # excluded from selection. The search continues over the remaining variants.
    try:
        out = tuning._run_in_tempdir(label, overrides, config_dir, scenario_dir, input_path)
        metrics, dropped, overprod = metrics_from_workbook(
            out, _harvest_cap(config_dir, overrides),
            welfare_density=_welfare_density(config_dir, overrides),
            relief_ceiling=_relief_ceiling(config_dir, overrides),
            move_cap=_move_cap(config_dir, overrides),
            min_harvest=_min_harvest(config_dir, overrides))
        return OptVariant(label=label, overrides=dict(overrides),
                          metrics=metrics, dropped=dropped, overprod=overprod)
    except Exception as e:  # noqa: BLE001 — reject-and-continue, don't crash the sweep
        msg = str(e).strip()
        reason = (msg.splitlines()[0] if msg else type(e).__name__)[:400]
        return OptVariant(label=label, overrides=dict(overrides),
                          metrics=_infeasible_metrics(), dropped=0, overprod=0,
                          failed=reason)


def overrides_yaml(overrides) -> str:
    """Render a recommendation's knob overrides as a control.yaml snippet."""
    if not overrides:
        return "# (baseline — no knob changes)"
    return "\n".join(f"{k}: {('null' if v is None else str(v).lower() if isinstance(v, bool) else v)}"
                     for k, v in overrides.items())


def run_full_forecast(input_path, config_dir, scenario_dir, overrides) -> str:
    """Apply `overrides` onto control.yaml and run the FULL pipeline, returning
    the output workbook path — i.e. feed a recommendation straight back into the
    forecast. Reuses the shared run harness; nothing mutates the caller's config."""
    return tuning._run_in_tempdir("optimized", overrides or {},
                                  config_dir, scenario_dir, input_path)


def save_overrides_to_config(config_dir, overrides) -> None:
    """Persist `overrides` by merging them into the REAL config_dir/control.yaml,
    so every later normal run (and the Configure editor) uses them. This is how a
    recommendation becomes the standing config — without it, the knobs live only
    in the optimizer's temp run and a normal 'Run forecast' reverts to baseline."""
    cy = os.path.join(config_dir, "control.yaml")
    with open(cy) as f:
        cfg = yaml.safe_load(f)
    cfg.update(overrides or {})
    with open(cy, "w") as f:
        yaml.safe_dump(cfg, f, sort_keys=False)


# --------------------------------------------------------------------------- #
# Run log — a durable record of WHAT was run and WHAT it produced
# --------------------------------------------------------------------------- #
DEFAULT_RUN_LOG = "optimize_history.jsonl"


def make_run_record(best, method, emphasis, *, ts, saved, source,
                    dropped=None, overprod=None) -> dict:
    """Build one optimize/auto-optimize log record: the SETTINGS used (method,
    emphasis, the winning knobs) + the RESULTS (key metrics + conservation). `ts` is
    an ISO timestamp the caller supplies (so this stays import-light)."""
    m = best.metrics if best else None
    return {
        "ts": ts,
        "source": source,
        "method": method,
        "emphasis": emphasis,
        "winning_knobs": dict(best.overrides) if best and best.overrides else {},
        "saved_to_config": bool(saved),
        "dropped": dropped if dropped is not None else (best.dropped if best else None),
        "overprod": overprod if overprod is not None else (best.overprod if best else None),
        "metrics": {
            "system_peak": round(m.system_peak, 3),
            "feed_load": round(m.feed_load),
            "weeks_over_harvest_cap": m.weeks_over_harvest_cap,
            "harvest_cv": round(m.harvest_var, 3),
            "system_overshoot": round(m.system_overshoot, 3),
        } if m else {},
    }


def append_run_log(record: dict, log_path: str = DEFAULT_RUN_LOG) -> None:
    """Append one run record as a JSON line — a durable, reviewable history of what
    the optimizer was asked to do and what it produced. Best-effort: a logging
    failure never breaks the run."""
    import json
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
    except Exception:  # noqa: BLE001
        pass


def read_run_log(log_path: str = DEFAULT_RUN_LOG, n: int = 20) -> list:
    """Return the last `n` run records (oldest→newest), or [] if no log yet."""
    import json
    try:
        with open(log_path, encoding="utf-8") as f:
            lines = [ln for ln in f if ln.strip()]
        return [json.loads(ln) for ln in lines[-n:]]
    except Exception:  # noqa: BLE001
        return []


def config_dir_with_overrides(config_dir, overrides) -> str:
    """Return a TEMP copy of `config_dir` with `overrides` merged into
    control.yaml — so a caller (e.g. the app) can run the full pipeline against
    it and parse the result for visualization. Nothing mutates `config_dir`."""
    import shutil
    import tempfile
    tmp = tempfile.mkdtemp(prefix="as_optcfg_")
    dst = os.path.join(tmp, "config")
    shutil.copytree(config_dir, dst)
    cy = os.path.join(dst, "control.yaml")
    with open(cy) as f:
        cfg = yaml.safe_load(f)
    cfg.update(overrides or {})
    with open(cy, "w") as f:
        yaml.safe_dump(cfg, f)
    return dst


def _default_workers(n_tasks: int) -> int:
    """Process-pool size for the sweep: leave a core for the UI, never more workers
    than tasks, capped at 8 (each worker runs a full pipeline — memory + diminishing
    returns past that)."""
    try:
        cpu = os.cpu_count() or 2
    except Exception:  # noqa: BLE001
        cpu = 2
    return max(1, min(n_tasks, cpu - 1, 8))


def _vc_get(vc, overrides, label):
    """Fetch a finished measurement from the cross-run variant cache (keyed by
    the deterministic overrides key) and relabel it for THIS requester — the
    same knob set can be reached under different labels by grid vs descent.

    Entries are PLAIN DICTS (see _vc_put); the OptVariant/Metrics are rebuilt
    from the CURRENT classes here, so a cache written before a source
    hot-reload stays loadable. Legacy object entries are still honored."""
    import dataclasses
    if vc is None:
        return None
    v = vc.get(_overrides_key(overrides))
    if v is None:
        return None
    if isinstance(v, OptVariant):        # legacy pre-plain-data entry
        return OptVariant(label=label, overrides=dict(v.overrides),
                          metrics=v.metrics, dropped=v.dropped,
                          overprod=v.overprod, failed=v.failed)
    try:
        m = None
        if v.get("metrics") is not None:
            fields = {f.name for f in dataclasses.fields(Metrics)}
            m = Metrics(**{k: x for k, x in v["metrics"].items()
                           if k in fields})
        return OptVariant(label=label, overrides=dict(v["overrides"]),
                          metrics=m, dropped=v["dropped"],
                          overprod=v["overprod"], failed=v.get("failed"))
    except Exception:  # noqa: BLE001 — schema drift -> cache miss, re-run
        return None


def _vc_put(vc, variant) -> None:
    """Record as PLAIN DATA — class instances tied to a module generation
    become unpicklable after a hot-reload (the 2026-08-07 disk-cache
    failures); dicts never do."""
    import dataclasses
    if vc is None:
        return
    vc[_overrides_key(variant.overrides)] = {
        "overrides": dict(variant.overrides),
        "metrics": (dataclasses.asdict(variant.metrics)
                    if dataclasses.is_dataclass(variant.metrics) else None),
        "dropped": variant.dropped,
        "overprod": variant.overprod,
        "failed": variant.failed,
    }


def sweep(input_path, config_dir, scenario_dir, grid=None, progress=None,
          parallel=True, max_workers=None, variant_cache=None) -> list[OptVariant]:
    """Run every grid row and return per-variant results (unscored — call
    recommend()/score_variants() with an emphasis). Nothing mutates the caller's
    config; each variant runs in its own temp copy.

    Each variant is an INDEPENDENT full forecast, so by default they run across a
    PROCESS POOL — the pipeline is CPU-bound + deterministic, so parallel execution
    yields IDENTICAL results (sorted back to grid order), just N× faster. Falls back
    to sequential automatically if a pool can't start (restricted env). Pass
    parallel=False to force the old one-at-a-time path.

    `variant_cache` (optional MutableMapping) persists finished measurements
    across searches and crashes: a variant whose overrides-key is already in
    the cache is reused, every fresh run is recorded. The CALLER owns validity
    (key the cache store by the input signature) and durability (pass a
    write-through mapping to survive a mid-search crash)."""
    grid = grid or OPT_FULL_GRID
    n = len(grid)
    if max_workers is None:
        max_workers = _default_workers(n)

    if not parallel or max_workers <= 1 or n <= 1:
        results = []
        for i, (label, overrides) in enumerate(grid):
            if progress is not None:
                progress(i, n, label)
            v = _vc_get(variant_cache, overrides, label)
            if v is None:
                v = run_variant(label, overrides, config_dir, scenario_dir,
                                input_path)
                _vc_put(variant_cache, v)
            results.append(v)
        return results

    from concurrent.futures import ProcessPoolExecutor, as_completed
    order = {label: i for i, (label, _) in enumerate(grid)}
    results: list = []
    todo = []
    done = 0
    for label, ov in grid:
        v = _vc_get(variant_cache, ov, label)
        if v is not None:
            done += 1
            if progress is not None:
                progress(done, n, f"{label} (cached)")
            results.append(v)
        else:
            todo.append((label, ov))
    if not todo:
        results.sort(key=lambda v: order.get(v.label, 1 << 30))
        return results
    from pickle import PicklingError
    try:
        with ProcessPoolExecutor(max_workers=max_workers) as ex:
            futs = {ex.submit(run_variant, label, ov, config_dir, scenario_dir,
                              input_path): label for label, ov in todo}
            for fut in as_completed(futs):
                done += 1
                if progress is not None:
                    progress(done, n, futs[fut])
                v = fut.result()
                _vc_put(variant_cache, v)
                results.append(v)
    except PicklingError:
        # Source hot-reload under a live run broke pool serialization (module
        # identity changed). Direct calls still work — finish sequentially;
        # the variant cache skips everything that already completed. SAY SO:
        # a principled fallback the operator can't see is still a silent one.
        if progress is not None:
            progress(done, n, "process pool broke (source hot-reload) — "
                              "finishing the sweep sequentially")
        print("WARN: sweep process pool broke (PicklingError, likely a source "
              "hot-reload) — finishing sequentially")
        return sweep(input_path, config_dir, scenario_dir, grid=grid,
                     progress=progress, parallel=False,
                     variant_cache=variant_cache)
    except Exception as e:  # noqa: BLE001
        # A pool that can't even start (sandboxed env) -> sequential. But if some
        # variants already ran, a failure is a real variant error -> surface it.
        if len(results) > n - len(todo):
            raise
        if progress is not None:
            progress(done, n, "process pool unavailable — running the sweep "
                              "sequentially")
        print(f"WARN: sweep process pool unavailable "
              f"({type(e).__name__}: {e}) — running sequentially")
        return sweep(input_path, config_dir, scenario_dir, grid=grid,
                     progress=progress, parallel=False,
                     variant_cache=variant_cache)
    results.sort(key=lambda v: order.get(v.label, 1 << 30))   # deterministic order
    return results


# Knob search space for the deep (coordinate-descent) search: each knob + the
# candidate values to try. Spans the levers that move the Minimize-loads objective
# — placement (tran_og), the harvest controller (setpoint/K), packing density, and
# the rebalancer budgets. Edit here to widen/narrow the search.
# Revised 2026-08-12 from the 106-run control-parameter study. Two changes, both
# measured rather than guessed:
#   * The rebalance budgets were exploring only their NON-BINDING side. 60 was
#     bit-identical to 30 and 12 bit-identical to 8 on every metric — five
#     duplicate score rows in one controller search, ~40% of the descent's budget
#     spent proving that a slack constraint is slack. The binding side is BELOW
#     the current value (0 changes the plan materially), so the axes now bracket
#     it instead.
#   * `global_buffer_pct` and `min_tank_control` were in NO search space despite
#     measurably moving the operator's hardest metric: global_buffer 0.05 -> 0.0
#     bought +3,700 fish on the worst harvest week, and min_tank_control 7000 ->
#     12000 bought +2,900 AND a better density peak (106 -> 101.6). Both are
#     planner preferences, not business rules, so they are legitimately tunable.
# `min_transfer_count` stays OUT: 7000 measured as a genuine optimum with sharp
# loss either side (3500 -> 10,450 worst week), so a descent axis would only
# rediscover it at the cost of runs.
CD_KNOB_SPACE = [
    ("tran_og_default_tanks", [2, 3]),
    ("facility_biomass_deviation_pct", [0.005, 0.01, 0.02]),
    ("harvest_smooth_lookahead_weeks", [6, 12]),
    ("density_target_pct", [0.85, 0.90, 0.95]),
    ("global_buffer_pct", [0.0, 0.05, 0.10]),
    ("min_tank_control", [7000, 12000]),
    # 2026-08-13: the de-duplication above was left half-done. A deferrable
    # pass gets min(knob, quality budget) and the quality budget can never
    # exceed max_transfers_per_week (15), so 15 and 30 are the SAME plan —
    # measured bit-identical on all 7 real starting states, which is how a
    # single-PR search came to report "30 -> 15" as a tuned win. 15 is the
    # highest value the search can distinguish; 8 brackets the binding side.
    ("rebalance_balance_budget", [0, 8, 15]),
    ("rebalance_split_budget", [0, 4, 8]),
]


def _overrides_key(ov):
    """Deterministic, order-independent sort key for an overrides dict — so a tie on
    score breaks the SAME way regardless of evaluation order (needed once variants
    are evaluated in parallel)."""
    return tuple(sorted((str(k), str(v)) for k, v in ov.items()))


def coordinate_descent(input_path, config_dir, scenario_dir, emphasis=DEFAULT_EMPHASIS,
                       weights=None, knob_space=None, max_rounds=3, seed=None,
                       progress=None, parallel=True,
                       max_workers=None, variant_cache=None) -> list[OptVariant]:
    """Greedy local search that finds COMBINATIONS the grid can't.

    From the `seed` config (default = current config / baseline), improve ONE knob
    at a time: try each candidate value for a knob (holding the others at the current
    best), keep the value that scores best, move to the next knob. Loop over all knobs
    each round; stop when a full round makes no improvement (local optimum) or
    max_rounds.

    The candidate VALUES of one knob are independent, so they're evaluated across a
    PROCESS POOL (one pool reused for the whole descent). The cyclic knob order +
    accept-best-per-knob trajectory is UNCHANGED, and ties break by a deterministic
    override key, so the result is identical to (and as reproducible as) the
    sequential version — just faster. Conservation-OK variants only ever win. Returns
    ALL evaluated variants (same shape as sweep()). Nothing mutates the caller's config."""
    knob_space = knob_space or CD_KNOB_SPACE
    w = weights or weights_for(emphasis)
    cache: dict = {}
    evaluated: list = []

    def _key(ov):
        return tuple(sorted(ov.items()))

    workers = max_workers if max_workers is not None else _default_workers(8)
    pool = None
    if parallel and workers > 1:
        from concurrent.futures import ProcessPoolExecutor
        try:
            pool = ProcessPoolExecutor(max_workers=workers)
        except Exception as e:  # noqa: BLE001 — restricted env -> sequential
            print(f"WARN: descent process pool unavailable "
                  f"({type(e).__name__}: {e}) — running sequentially")
            pool = None

    def _record(k, v):
        cache[k] = v
        evaluated.append(v)
        if progress is not None:
            progress(len(evaluated), None, v.label)

    def _eval(ov, label):
        k = _key(ov)
        if k in cache:
            return cache[k]
        cv = _vc_get(variant_cache, ov, label)
        if cv is not None:
            _record(k, cv)
            return cache[k]
        v = run_variant(label, dict(ov), config_dir, scenario_dir, input_path)
        _vc_put(variant_cache, v)
        _record(k, v)
        return cache[k]

    def _eval_many(items):
        """Evaluate a list of (key, overrides, label) — all uncached in THIS
        descent — reusing the cross-run variant cache, the rest in parallel."""
        nonlocal pool
        if not items:
            return
        run_items = []
        for k, ov, label in items:
            cv = _vc_get(variant_cache, ov, label)
            if cv is not None:
                _record(k, cv)
            else:
                run_items.append((k, ov, label))
        if not run_items:
            return
        if pool is None or len(run_items) == 1:
            for k, ov, label in run_items:
                _eval(ov, label)
            return
        from concurrent.futures import as_completed
        from pickle import PicklingError
        try:
            futs = {pool.submit(run_variant, label, dict(ov), config_dir,
                                scenario_dir, input_path): k
                    for k, ov, label in run_items}
            for fut in as_completed(futs):
                v = fut.result()
                _vc_put(variant_cache, v)
                _record(futs[fut], v)
        except PicklingError:
            # A source hot-reload under a LIVE run (Streamlit re-imports local
            # modules) changes run_variant's identity and the pool can no
            # longer serialize it. Direct calls still work — finish this batch
            # sequentially and stay sequential for the rest of the descent
            # instead of dying (crashed once for real: 2026-08-07). SAY SO.
            print("WARN: descent process pool broke (PicklingError, likely a "
                  "source hot-reload) — continuing sequentially")
            if progress is not None:
                progress(len(evaluated), None,
                         "pool broke — descent continues sequentially")
            pool = None
            for k, ov, label in run_items:
                if k not in cache:
                    _eval(ov, label)

    def _best_overrides():
        # Re-score the whole evaluated set (consistent normalization over all OK
        # variants) and return the lowest-scoring conservation-OK config; ties break
        # deterministically so parallel == sequential.
        score_variants(evaluated, w)
        ok = [v for v in evaluated if v.conservation_ok]
        best = (min(ok, key=lambda v: (v.score, _overrides_key(v.overrides)))
                if ok else evaluated[0])
        return dict(best.overrides)

    try:
        seed = dict(seed or {})
        _eval(seed, "seed" if seed else "baseline")   # warm-start (seq: it's evaluated[0])
        current = _best_overrides()
        for _ in range(max_rounds):
            improved = False
            for knob, values in knob_space:
                todo = []
                for val in values:
                    ov = dict(current); ov[knob] = val
                    k = _key(ov)
                    if k != _key(current) and k not in cache:
                        todo.append((k, ov, f"{knob}={val}"))
                _eval_many(todo)
                nb = _best_overrides()
                if _key(nb) != _key(current):
                    current = nb
                    improved = True
            if not improved:
                break
    finally:
        if pool is not None:
            pool.shutdown()
    return evaluated


def deep_search_combined(input_path, config_dir, scenario_dir, emphasis=DEFAULT_EMPHASIS,
                         weights=None, grid=None, max_rounds=3,
                         progress=None, max_workers=None,
                         variant_cache=None, knob_space=None) -> list[OptVariant]:
    """Best-of-both for ANY emphasis: run the full GRID (broad, diverse coverage —
    incl. the off-by-default controls), then coordinate descent SEEDED FROM the
    grid's best (local refinement that finds combinations around the broadly-best
    point). Pool every evaluated variant, dedup by overrides, and return the lot —
    recommend() then picks the GLOBAL best across both methods. Grid explores,
    descent exploits; the winner is whichever wins. Same OptVariant list shape, so
    the app/score-table/Pareto/apply-verify consume it unchanged.

    `knob_space` restricts BOTH halves to a caller-chosen search space (the
    tuned tournament passes a Method's own axes); None keeps the defaults
    (CD_KNOB_SPACE for the descent, OPT_FULL_GRID for the grid)."""
    w = weights or weights_for(emphasis)
    gvars = sweep(input_path, config_dir, scenario_dir,
                  grid=grid or OPT_FULL_GRID, progress=progress,
                  max_workers=max_workers, variant_cache=variant_cache)
    score_variants(gvars, w)
    ok = [v for v in gvars if v.conservation_ok]
    seed = dict((min(ok, key=lambda v: v.score) if ok else gvars[0]).overrides)
    dvars = coordinate_descent(input_path, config_dir, scenario_dir, emphasis=emphasis,
                               weights=w, seed=seed, max_rounds=max_rounds, progress=progress,
                               max_workers=max_workers, variant_cache=variant_cache,
                               knob_space=knob_space)
    pool = {}
    for v in gvars + dvars:        # dedup by overrides; identical configs collapse
        pool[tuple(sorted(v.overrides.items()))] = v
    return list(pool.values())
