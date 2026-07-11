"""Read inputs and write outputs against Forecast.xlsm."""
from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .models import CalibrationResidual, BatchWeekState


# ---------- Writer: BiologyProjection sheet ----------

def write_biology_projection(wb, states: Iterable[BatchWeekState], sheet_name: str = "BiologyProjection") -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Batch", "Week", "WeekStart", "DaysSinceInput", "WeekFromInput",
        "Stage", "Count", "AvgWt_g", "Biomass_kg", "SGR_pct_day", "FCR",
        "FeedType", "Mortality_pct_wk", "Cull_pct", "Cull_Count",
        "Cull_Biomass_kg", "Feed_kg_day", "Feed_kg_week",
    ]
    ws.append(headers)
    # Sort deterministically (batch, then chronological week) so the sheet is
    # byte-reproducible run-to-run. The incoming `states` order depends on dict
    # iteration, which PYTHONHASHSEED randomizes across processes — leaving the
    # rows shuffled (identical data, different order), which breaks output diffs.
    for s in sorted(states, key=lambda s: (s.batch_id, s.week_label)):
        ws.append([
            s.batch_id, s.week_label, s.week_start, s.days_since_input,
            s.week_from_input, s.stage, round(s.count, 1), round(s.avg_weight_g, 3),
            round(s.biomass_kg, 1), round(s.sgr_pct_day, 4), round(s.fcr, 4),
            s.feed_type, round(s.mortality_pct_weekly, 4), round(s.cull_event_pct, 4),
            round(s.cull_count_week, 0) if s.cull_count_week > 0 else None,
            round(s.cull_biomass_kg_week, 1) if s.cull_biomass_kg_week > 0 else None,
            round(s.feed_kg_day, 2), round(s.feed_kg_week, 1),
        ])

    widths = {1: 8, 2: 12, 3: 14, 4: 14, 5: 14, 6: 6, 7: 12, 8: 10,
              9: 13, 10: 12, 11: 8, 12: 18, 13: 16, 14: 9,
              15: 12, 16: 16, 17: 12, 18: 13}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_batch_locations(wb, batch_locations, sheet_name: str = "BatchLocations") -> None:
    """Per-(week, batch, tank) occupancy from the placement plan."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["BATCH LOCATIONS"])
    ws.append(["Per-tank batch occupancy from the forecast plan. Auto-generated."])
    ws.append([])
    ws.append([
        "Week", "Week_Start", "Batch", "Tank", "System",
        "Count (fish)", "AvgWt (kg)", "Biomass (kg)", "Density (kg/m3)", "Stage",
    ])
    for r in batch_locations:
        ws.append([
            r.week_label, r.week_start, r.batch_id, r.tank_id, r.system_id,
            round(r.count, 0),
            round(r.avg_wt_g / 1000.0, 3),
            round(r.biomass_kg, 0),
            round(r.density_kg_m3, 1),
            getattr(r, "stage", ""),
        ])
    widths = {1: 11, 2: 12, 3: 8, 4: 6, 5: 9, 6: 12, 7: 11, 8: 13, 9: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


_BP_TIER = {
    "OG1N": "Nursery (OG1/2)", "OG1S": "Nursery (OG1/2)",
    "OG2N": "Nursery (OG1/2)", "OG2S": "Nursery (OG1/2)",
    "OG3N": "Grow-out OG3", "OG3S": "Grow-out OG3",
    "OG4N": "Grow-out OG4", "OG4S": "Grow-out OG4",
    "OG5N": "Grow-out OG5", "OG5S": "Grow-out OG5",
    "OG6S": "Finishing OG6", "OG6N": "Finishing/depuration OG6N",
}
_BP_ORDER = ["Nursery (OG1/2)", "Grow-out OG3", "Grow-out OG4", "Grow-out OG5",
             "Finishing OG6", "Finishing/depuration OG6N"]


def write_batch_plan(wb, batch_locations, harvest_events, default_hog_yield: float = 0.81,
                     sheet_name: str = "Batch Plan") -> None:
    """Per-batch journey: a summary header + the milestone timeline (each conveyor
    tier the batch enters — SW entry → grow-out → finishing → harvest — with week,
    systems, weight, tank count), derived from BatchLocations + harvest events. The
    'where each batch is + how it got there' as a shareable Excel sheet; mirrors the
    app's Plan-tab per-batch plan."""
    import collections
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["BATCH PLAN"])
    ws.append(["Per-batch journey: summary header + milestone timeline (each conveyor "
               "tier entered, at week/weight, through to harvest). Auto-generated."])
    ws.append([])

    by_batch = collections.defaultdict(list)
    for r in batch_locations:
        by_batch[r.batch_id].append(r)
    hv = collections.defaultdict(lambda: {"weeks": [], "hog": 0.0, "wt": []})
    for ev in harvest_events:
        h = hv[ev.batch_id]
        h["weeks"].append(iso_week_label(ev.event_date))
        h["hog"] += ev.count * (ev.avg_wt_g / 1000.0) * default_hog_yield
        h["wt"].append(ev.avg_wt_g / 1000.0)

    plans = []
    for bid, recs in by_batch.items():
        recs = sorted(recs, key=lambda r: r.week_label)
        weeks = list(dict.fromkeys(r.week_label for r in recs))
        wk_tanks = collections.defaultdict(set)
        by_week = collections.defaultdict(list)
        for r in recs:
            wk_tanks[r.week_label].add(r.tank_id)
            by_week[r.week_label].append(r)
        peak = max((len(s) for s in wk_tanks.values()), default=0)
        milestones, seen = [], set()
        for wk in weeks:
            tiers_here = collections.defaultdict(list)
            for r in by_week[wk]:
                tiers_here[_BP_TIER.get(r.system_id, r.system_id)].append(r)
            for tier in _BP_ORDER:
                if tier in tiers_here and tier not in seen:
                    seen.add(tier)
                    sub = tiers_here[tier]
                    avgwt = sum(r.avg_wt_g for r in sub) / len(sub) / 1000.0
                    if not milestones:   # first appearance: real entry vs in-flight
                        label = ("Seawater entry (TranOG)" if avgwt < 0.6
                                 else "In-flight at forecast start")
                    else:
                        label = f"-> {tier}"
                    # Tank count = the batch's TOTAL concurrent footprint that
                    # week (not just tanks in the newly-entered tier) — so the
                    # column grows with biomass as expected, instead of reading
                    # "1" at every tier entry. Biomass + peak density (over all
                    # the batch's tanks that week) let the operator sanity-check
                    # the density at each milestone.
                    wk_recs = by_week[wk]
                    tot_bio = sum(r.biomass_kg for r in wk_recs)
                    peak_dens = max((r.density_kg_m3 for r in wk_recs), default=0.0)
                    milestones.append((wk, label,
                                       ", ".join(sorted({r.system_id for r in sub})),
                                       round(avgwt, 2), len(wk_tanks[wk]),
                                       round(tot_bio, 0), round(peak_dens, 1)))
        h = hv.get(bid)
        hw = (f"{min(h['weeks'])}-{max(h['weeks'])}" if h and h["weeks"] else "-")
        hog_t = (h["hog"] / 1000.0) if h else 0.0
        if h and h["weeks"]:
            milestones.append((hw, "Harvest", "-> harvest",
                               round(sum(h["wt"]) / len(h["wt"]), 2), "", "", ""))
        plans.append({"batch": bid, "sw": weeks[0] if weeks else "-",
                      "peak": peak, "hw": hw, "hog_t": hog_t, "ms": milestones})
    plans.sort(key=lambda p: p["sw"])

    ws.append(["SUMMARY - one row per batch"])
    ws.append(["Batch", "SW_entry", "Peak_tanks", "Harvest_window", "HOG_tonnes"])
    for p in plans:
        ws.append([p["batch"], p["sw"], p["peak"], p["hw"], round(p["hog_t"], 0)])
    ws.append([])
    ws.append(["MILESTONES - the journey per batch"])
    ws.append(["Batch", "Week", "Event", "Systems", "AvgWt (kg)",
               "Total_Tanks (that week)", "Total_Biomass (kg)", "Peak_Density (kg/m3)"])
    for p in plans:
        for (wk, label, systems, avgwt, tanks, bio_kg, dens) in p["ms"]:
            ws.append([p["batch"], wk, label, systems, avgwt, tanks, bio_kg, dens])
        ws.append([])
    widths = {1: 8, 2: 13, 3: 26, 4: 18, 5: 11, 6: 22, 7: 18, 8: 18}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_harvest_plan_output(
    wb,
    harvest_events,
    default_hog_yield: float,
    facility_limits_hog: dict,
    sheet_name: str = "HarvestPlan",
) -> None:
    """Per-event harvest plan as a single table (matches reference format).

    Columns: Week, Batch, Tank, Count (fish), Gross_AvgWt (kg),
    Gross_Biomass (kg), HOG_Yield (ratio), HOG_AvgWt (kg), HOG_Biomass (kg).

    `facility_limits_hog` is a dict `{week_label: hog_yield}` for per-week HOG
    yield overrides; default falls back to `default_hog_yield`. `harvest_events`
    are the realized closed-loop harvests.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["HARVEST PLAN"])
    ws.append(["For each harvest: specify batch, date or week#, tank, counts and biomass."])
    ws.append([])
    ws.append([
        "Week", "Batch", "Tank", "Count (fish)",
        "Gross_AvgWt (kg)", "Gross_Biomass (kg)",
        "HOG_Yield (ratio)", "HOG_AvgWt (kg)", "HOG_Biomass (kg)",
    ])

    events_sorted = sorted(harvest_events, key=lambda e: (e.event_date, e.source_tank_id))
    for ev in events_sorted:
        wk = iso_week_label(ev.event_date)
        gross_avg_kg = ev.avg_wt_g / 1000.0
        gross_biomass = ev.count * gross_avg_kg
        hog_yield = facility_limits_hog.get(wk, default_hog_yield)
        ws.append([
            wk, ev.batch_id, ev.source_tank_id,
            round(ev.count, 0),
            round(gross_avg_kg, 2),
            round(gross_biomass, 0),
            round(hog_yield, 2),
            round(gross_avg_kg * hog_yield, 2),
            round(gross_biomass * hog_yield, 0),
        ])
    widths = {1: 11, 2: 8, 3: 6, 4: 13, 5: 16, 6: 17, 7: 17, 8: 14, 9: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_transfer_plan_output(
    wb,
    transfer_events,
    tranog_events,
    grade_events=None,
    sheet_name: str = "TransferPlan",
) -> None:
    """Per-event transfer + TranOG + Grade plan as a single table (matches
    reference format).

    Columns: Week, Batch, Type, From_Tank, To_Tank, Count (fish), Avg_Weight (kg),
    Grade, CV (%). Type is TranOG (FW->seawater) / Transfer / Grade — filter on it
    to find the FW->OG arrivals. From_Tank is 'FW' for TranOG entries, and those
    arrivals are size-class split, so the Grade column shows A (big class) / B
    (small class); a regular transfer is blank; a Grade event's multi-tank source
    is comma-separated. Rejected transfer attempts (count_transferred == 0) are
    omitted — this is the actionable plan, not the attempt log (the audit sheets
    carry rejected attempts).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["TRANSFER PLAN"])
    ws.append(["For each transfer: specify batch, date or week#, from/to tanks, count, avg weight. "
               "Grade A = big size class, B = small (size-class split)."])
    ws.append([])
    ws.append([
        "Week", "Batch", "Type", "From_Tank", "To_Tank",
        "Count (fish)", "Avg_Weight (kg)", "Grade", "CV (%)",
    ])

    def _grade(dest):
        # TankAllocation.size_class is "big" / "small" / "mixed" / "". Map to the
        # A/B grade convention; mixed/blank -> no grade.
        return {"big": "A", "small": "B"}.get(getattr(dest, "size_class", "") or "", "")

    rows: list[tuple] = []
    for ev in tranog_events:
        wk = iso_week_label(ev.event_date)
        for dest in ev.destinations:
            rows.append((
                ev.event_date, wk, ev.batch_id, "FW", dest.tank_id,
                dest.count, dest.avg_wt_g / 1000.0, _grade(dest), dest.cv_pct,
                "TranOG",
            ))
    for ev in transfer_events:
        # GradedHarvest (Event 5) rides in transfer_events with a different shape
        # (no .destinations): 1 source -> pickup + retention. Emit its two legs
        # instead of crashing on ev.destinations.
        if hasattr(ev, "pickup_tank_id"):
            wk = iso_week_label(ev.event_date)
            rows.append((ev.event_date, wk, ev.batch_id, str(ev.source_tank_id),
                         ev.pickup_tank_id, ev.pickup_count,
                         ev.pickup_avg_wt_g / 1000.0, "pickup", ev.cv_pct, "Grade"))
            rows.append((ev.event_date, wk, ev.batch_id, str(ev.source_tank_id),
                         ev.retention_tank_id, ev.retention_count,
                         ev.retention_avg_wt_g / 1000.0, "retention", ev.cv_pct, "Grade"))
            continue
        ct = getattr(ev, "count_transferred", None)
        if ct is not None and ct <= 0:
            continue  # rejected attempt — not part of the actionable plan
        wk = iso_week_label(ev.event_date)
        for dest in ev.destinations:
            rows.append((
                ev.event_date, wk, ev.batch_id, str(ev.source_tank_id), dest.tank_id,
                dest.count, dest.avg_wt_g / 1000.0, _grade(dest), dest.cv_pct,
                "Transfer",
            ))
    for ev in (grade_events or []):
        wk = iso_week_label(ev.event_date)
        src_str = ",".join(str(t) for t in ev.source_tank_ids)
        for dest in ev.destinations:
            rows.append((
                ev.event_date, wk, ev.batch_id, src_str, dest.tank_id,
                dest.count, dest.avg_wt_g / 1000.0,
                _grade(dest), dest.cv_pct, "Grade",
            ))
    rows.sort(key=lambda r: (r[0], r[2]))

    for r in rows:
        ws.append([
            r[1], r[2], r[9],   # Week, Batch, Type
            r[3], r[4],         # From_Tank, To_Tank
            round(r[5], 0),
            round(r[6], 3),
            r[7],
            round(r[8], 1) if r[8] else None,
        ])
    widths = {1: 11, 2: 8, 3: 10, 4: 10, 5: 8, 6: 13, 7: 14, 8: 9, 9: 8}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_transfer_template(
    wb,
    batch_locations,
    harvest_events,
    tranog_events,
    control,
    facility,
    sheet_name: str = "TransferTemplate",
) -> None:
    """Generalized batch-flow template + a per-batch plan summary.

    Two sections:
      A) OVERALL TEMPLATE — the canonical seawater journey every batch follows
         (entry → nursery → fan-out grow-out → finishing/depuration → harvest),
         with relative week offsets, weights, systems, and tank counts.
      B) PER-BATCH SUMMARY — one row per batch: when it enters seawater (week +
         weeks from forecast start), its entry weight/count/density, its PEAK
         tank footprint + PEAK density (the density-risk indicator) and when, its
         harvest window (weeks from entry) and weight, and a Density_Status flag.
         Keeps the facility aware of each batch's footprint, density risk, and
         harvest timing at a glance.
    """
    from collections import defaultdict
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ---- Section A: overall template ----
    ws.append(["TRANSFER PLAN TEMPLATE"])
    ws.append([])
    ws.append(["A. OVERALL TEMPLATE — the canonical seawater journey every batch "
               "follows (timing varies by stocking date; the shape does not)."])
    ws.append(["Stage", "Week from SW entry", "Avg weight (kg)", "From", "To",
               "Grade", "Tanks", "Purpose"])
    for row in [
        ("1. Seawater entry (TranOG)", "0", "~0.37", "FW", "OG1 / OG2",
         "A big + B small", "2", "size-class split into the nursery"),
        ("2. Nursery hold", "0–11", "0.37 → 1.0", "OG1/2", "OG1/2",
         "—", "1–2", "grow to 1 kg (the <1 kg nursery lock)"),
        ("3. Cross the 1 kg lock", "~11", "~1.0", "OG1/2", "OG3",
         "—", "+1", "first grow-out move"),
        ("4. Grow-out fan-out", "~11–31", "1.0 → 2.9", "OG3", "OG4 → OG5",
         "—", "6–7", "spread across systems to hold density"),
        ("5. Finishing / depuration", "~31–35", "2.9 → 3.8", "OG5", "OG6 / 6N / 6S",
         "—", "~8 (peak)", "enter top systems; 6N = in-place depuration purge"),
        ("6. Harvest drain", "~35–43", "3.8 → 4.4", "OG6 / 6N", "harvest",
         "—", "drains down", "~4.2 kg over ~6 weeks; ~43 wk SW phase total"),
    ]:
        ws.append(list(row))
    ws.append([])
    ws.append([])

    # ---- Section B: per-batch summary ----
    ws.append(["B. PER-BATCH PLAN SUMMARY — entry timing, footprint, density risk, "
               "and harvest window per batch."])
    ws.append([
        "Batch", "SW_Entry_Week", "Wks_from_Start", "Entry_AvgWt (kg)",
        "Entry_Count (fish)", "Entry_Density (×cap)", "Peak_Tanks",
        "Peak_Density (×cap)", "Peak_Wk (from entry)", "Harvest_Start",
        "Wks_Entry→Harvest", "Harvest_AvgWt (kg)", "Density_Status",
    ])

    cap = {t.tank_id: t.max_density_kg_m3 for t in facility.tanks}
    tsys = {t.tank_id: t.system_id for t in facility.tanks}
    weeks = sorted({r.week_label for r in batch_locations})
    widx = {w: i for i, w in enumerate(weeks)}

    bw = defaultdict(lambda: {"cnt": 0.0, "wsum": 0.0, "tanks": set(), "maxratio": 0.0})
    for r in batch_locations:
        e = bw[(r.batch_id, r.week_label)]
        e["cnt"] += r.count
        e["wsum"] += r.avg_wt_g * r.count
        e["tanks"].add(r.tank_id)
        c = cap.get(r.tank_id)
        # Peak density EXCLUDES the OG6N depuration/purge pool — harvest-size fish
        # held off-feed at high density before shipping is expected, not a stocking
        # problem, so it must not dominate a batch's peak (consistent with the
        # engine + the app/optimizer density alerts).
        if c and r.density_kg_m3 and tsys.get(r.tank_id) != "OG6N":
            e["maxratio"] = max(e["maxratio"], r.density_kg_m3 / c)

    tog_week = {}
    for ev in (tranog_events or []):
        tog_week[ev.batch_id] = iso_week_label(ev.event_date)

    hv = defaultdict(lambda: {"first": None, "cnt": 0.0, "wsum": 0.0})
    for ev in harvest_events:
        wk = iso_week_label(ev.event_date)
        h = hv[ev.batch_id]
        if h["first"] is None or wk < h["first"]:
            h["first"] = wk
        h["cnt"] += ev.count
        h["wsum"] += ev.avg_wt_g * ev.count

    weeks_by_batch = defaultdict(list)
    for (b, w) in bw:
        weeks_by_batch[b].append(w)

    for b in sorted(weeks_by_batch):
        bws = sorted(weeks_by_batch[b], key=lambda w: widx[w])
        # SW entry = the forecast TranOG week if any, else first-seen (in-flight).
        entry_wk = tog_week.get(b, bws[0])
        if entry_wk not in widx:
            entry_wk = bws[0]
        in_flight = b not in tog_week
        e0 = bw[(b, entry_wk)]
        entry_wt = (e0["wsum"] / e0["cnt"] / 1000.0) if e0["cnt"] else 0.0
        peak_tanks = max(len(bw[(b, w)]["tanks"]) for w in bws)
        peak_wk = max(bws, key=lambda w: bw[(b, w)]["maxratio"])
        peak_ratio = bw[(b, peak_wk)]["maxratio"]
        h = hv.get(b)
        hstart = h["first"] if h and h["first"] else ""
        h_off = (widx[hstart] - widx[entry_wk]) if (hstart and hstart in widx) else ""
        h_wt = (h["wsum"] / h["cnt"] / 1000.0) if (h and h["cnt"]) else ""
        ws.append([
            b, entry_wk + (" (in-flight)" if in_flight else ""),
            widx[entry_wk], round(entry_wt, 2), round(e0["cnt"], 0),
            round(e0["maxratio"], 2) if e0["maxratio"] else "",
            peak_tanks, round(peak_ratio, 2) if peak_ratio else "",
            widx[peak_wk] - widx[entry_wk],
            hstart, h_off, round(h_wt, 2) if h_wt != "" else "",
            "OVER CAP" if peak_ratio > 1.0 else "OK",
        ])

    ws.column_dimensions["A"].width = 18
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 16


def write_harvest_plan_report(
    wb,
    harvest_events,
    scenario_name: str,
    default_hog_yield: float,
    facility_limits_hog: dict,
    forecast_start=None,
    sheet_name: str = "HarvestPlan Report",
) -> None:
    """Annual per-batch harvest summary (matches reference format).

    One block per year. Each block: a "<scenario> <year>" title row, a month-
    header row (12 month-start columns + "TOTAL <year>"), then three rows per
    batch — Units, Av Weight - Kg HOG, Biomass - Tons HOG — with monthly values
    and a year total. Blank cells for months with no harvest.

    Boundary weeks (a week straddling a month boundary) are split between the
    two months by WORKING-DAY fraction (see time_grid.working_day_month_split),
    so a smooth weekly harvest maps to its true ~21-22 working-day share per
    month instead of dumping the whole week into the week-start's month.
    """
    from collections import defaultdict
    from datetime import date as _date
    from .time_grid import working_day_month_split

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Aggregate per (year, batch, month) -> count + HOG biomass (kg).
    agg: dict[tuple, dict] = defaultdict(lambda: {"count": 0.0, "hog_kg": 0.0})
    years: set[int] = set()
    batches_by_year: dict[int, set] = defaultdict(set)
    for ev in harvest_events:
        hog_yield = facility_limits_hog.get(iso_week_label(ev.event_date), default_hog_yield)
        hog_kg = ev.count * ev.avg_wt_g / 1000.0 * hog_yield
        # Split the week's harvest across the months its Mon-Fri working days
        # fall into (boundary weeks split by working-day fraction). No
        # forecast_start clip: manual override-window harvests are dated BEFORE
        # the shifted forecast_start, and clipping would dump a whole boundary
        # week into one month/year (same class as the Daily Harvest Schedule bug).
        for (yr, mo), frac in working_day_month_split(ev.event_date).items():
            e = agg[(yr, ev.batch_id, mo)]
            e["count"] += ev.count * frac
            e["hog_kg"] += hog_kg * frac
            years.add(yr)
            batches_by_year[yr].add(ev.batch_id)

    for year in sorted(years):
        ws.append([f"{scenario_name} {year}"])
        ws.append(["", ""] + [_date(year, m, 1) for m in range(1, 13)] + [f"TOTAL {year}"])
        # Per-month totals across all batches (the bottom TOTAL block) — what
        # sales planning reads: how much HOG lands each month.
        mo_count = [0.0] * 12
        mo_hog = [0.0] * 12
        for b in sorted(batches_by_year[year]):
            months = [agg.get((year, b, m)) for m in range(1, 13)]
            tot_count = sum(m["count"] for m in months if m)
            tot_hog = sum(m["hog_kg"] for m in months if m)
            units = [round(m["count"], 0) if m else "" for m in months]
            avgwt = [round(m["hog_kg"] / m["count"], 2) if m and m["count"] else "" for m in months]
            tons = [round(m["hog_kg"] / 1000.0, 0) if m else "" for m in months]
            ws.append([b, "Units"] + units + [round(tot_count, 0)])
            ws.append(["", "Av Weight - Kg HOG"] + avgwt
                      + [round(tot_hog / tot_count, 2) if tot_count else ""])
            ws.append(["", "Biomass - Tons HOG"] + tons + [round(tot_hog / 1000.0, 0)])
            for i, m in enumerate(months):
                if m:
                    mo_count[i] += m["count"]
                    mo_hog[i] += m["hog_kg"]
        # Bottom TOTAL block: monthly sums across batches (Units / weighted Av
        # Weight / Biomass tons), plus the year total in the last column.
        yr_count = sum(mo_count)
        yr_hog = sum(mo_hog)
        ws.append(["TOTAL", "Units"]
                  + [round(c, 0) if c else "" for c in mo_count]
                  + [round(yr_count, 0)])
        ws.append(["", "Av Weight - Kg HOG"]
                  + [round(mo_hog[i] / mo_count[i], 2) if mo_count[i] else "" for i in range(12)]
                  + [round(yr_hog / yr_count, 2) if yr_count else ""])
        ws.append(["", "Biomass - Tons HOG"]
                  + [round(h / 1000.0, 0) if h else "" for h in mo_hog]
                  + [round(yr_hog / 1000.0, 0)])
        ws.append([])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    for c in range(3, 16):
        ws.column_dimensions[get_column_letter(c)].width = 11


def _row_feed_kg_day(r, batches, tables):
    """Realized feed/day (kg) for one BatchLocation row — 0 for STARVE (6N
    depuration), empty, or when tables is absent.

    Single source for the per-row feed every feed accumulator sums; it was
    hand-inlined with subtly different guards in ~6 writers (the SystemLimitsAudit
    copy even omitted the `tables is not None` guard). Callers add the FW/EGG
    projected feed and the 6N move-in add-back themselves (those aren't per-row).
    """
    from .biology import realized_feed_kg_day
    if tables is None or getattr(r, "stage", "") == "STARVE":
        return 0.0
    return realized_feed_kg_day(
        r.avg_wt_g, r.biomass_kg, (batches or {}).get(r.batch_id), tables)


def write_yearly_summary(
    wb,
    batch_locations,
    harvest_events,
    facility_limits,
    control,
    batches=None,
    tables=None,
    default_hog_yield: float = 0.0,
    hog_overrides=None,
    sixn_move_in_feed=None,
    biology_states_by_batch=None,
    sheet_name: str = "YearlySummary",
) -> None:
    """Facility-wide per-year rollup for at-a-glance yearly trends.

    One row per calendar year: harvest (count, HOG tonnes, gross tonnes, avg HOG
    weight), feed (realized tonnes), and facility biomass (peak / mean tonnes,
    mean utilisation vs the per-week cap). Aggregates the same realized data the
    other sheets use (HarvestReport harvest, realized_feed_kg_day feed,
    BatchLocations biomass) so it reconciles with them.

    Feed (realized tonnes) = OG/SW realized feed + the 6N purge move-in 4-day
    feed + FW/EGG projected feed -- the SAME three sources FeedForecast and the
    Weekly/Monthly ledger sum, so the annual feed total ties out across all of
    them (FW feed is hatchery feed, part of total facility feed ordered).
    """
    from collections import defaultdict
    from .caps import resolve_facility_cap, METRIC_BIOMASS
    from .biology import realized_feed_kg_day

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Per-week facility biomass + realized feed (kg/day), with each week's year.
    wk_bio: dict[str, float] = defaultdict(float)
    wk_feed: dict[str, float] = defaultdict(float)
    wk_year: dict[str, int] = {}
    for r in batch_locations:
        wk_bio[r.week_label] += r.biomass_kg
        if r.week_label not in wk_year and hasattr(r.week_start, "year"):
            wk_year[r.week_label] = r.week_start.year
        # STARVE tank-weeks (6N depuration) eat nothing — handled by the helper
        # (biomass still counts above; feed does not).
        wk_feed[r.week_label] += _row_feed_kg_day(r, batches, tables)

    # 6N purge move-in fish ate 4 pre-transfer days in their source tank (shown
    # in 6N = STARVE above) — add that real feed (already weekly kg) per week so
    # the yearly feed tonnes match the FeedForecast / ledger totals.
    wk_movein: dict[str, float] = defaultdict(float)
    for (_bid, wk_, _ftype), kg in (sixn_move_in_feed or {}).items():
        if kg:
            wk_movein[wk_] += kg

    # FW/EGG projected feed (hatchery) AND biomass per week. FW fish live in FW
    # tanks, absent from batch_locations, so both are sourced here from the biology
    # projection. The biomass is folded into wk_bio because FW/EGG biomass is real
    # facility biomass counted against the 3.8M cap (audit H2) — so Peak/Mean/
    # Utilisation are FW-inclusive, mirroring the long-standing FW-feed correction.
    wk_fwfeed: dict[str, float] = defaultdict(float)
    for states in (biology_states_by_batch or {}).values():
        for s in states:
            if s.stage not in ("FW", "EGG"):
                continue
            if s.feed_kg_week:
                wk_fwfeed[s.week_label] += s.feed_kg_week
            wk_bio[s.week_label] += s.biomass_kg
            wk_year.setdefault(s.week_label,
                               s.week_start.year if hasattr(s.week_start, "year") else None)

    # Harvest per year (HOG via per-week override or default).
    hog_overrides = hog_overrides or {}
    yr_hc: dict[int, float] = defaultdict(float)
    yr_gross: dict[int, float] = defaultdict(float)
    yr_hog: dict[int, float] = defaultdict(float)
    for ev in harvest_events:
        d = ev.event_date.date() if hasattr(ev.event_date, "date") else ev.event_date
        y = d.year
        gross = ev.count * ev.avg_wt_g / 1000.0
        hy = hog_overrides.get(iso_week_label(ev.event_date), default_hog_yield)
        yr_hc[y] += ev.count
        yr_gross[y] += gross
        yr_hog[y] += gross * hy

    # Feed + biomass + utilisation per year.
    yr_feed: dict[int, float] = defaultdict(float)
    yr_bio: dict[int, list] = defaultdict(list)
    yr_util: dict[int, list] = defaultdict(list)
    for wk, bio in wk_bio.items():
        y = wk_year.get(wk)
        if y is None:
            continue
        yr_bio[y].append(bio)
        yr_feed[y] += wk_feed.get(wk, 0.0) * 7.0 + wk_movein.get(wk, 0.0)
        cap = resolve_facility_cap(METRIC_BIOMASS, wk, facility_limits, control)
        if cap:
            yr_util[y].append(100.0 * bio / cap)
    # FW/EGG feed in a separate pass so FW-only weeks (no OG biomass, absent from
    # wk_bio) still contribute their hatchery feed to the annual total.
    for wk, fw in wk_fwfeed.items():
        y = wk_year.get(wk)
        if y is not None:
            yr_feed[y] += fw

    years = sorted(set(yr_hc) | set(yr_bio))
    ws.append(["YEARLY SUMMARY (facility-wide)"])
    ws.append([])
    ws.append([
        "Year", "Harvest_Count (fish)", "Harvest_HOG (t)", "Harvest_Gross (t)",
        "Avg_HOG_Wt (kg)", "Feed (t)", "Peak_Biomass (t)", "Mean_Biomass (t)",
        "Mean_Utilisation (%)",
    ])
    for y in years:
        hc = yr_hc.get(y, 0.0)
        hog_t = yr_hog.get(y, 0.0) / 1000.0
        gross_t = yr_gross.get(y, 0.0) / 1000.0
        avg_hog = (yr_hog.get(y, 0.0) / hc) if hc > 0 else 0.0
        bios = yr_bio.get(y, [])
        utils = yr_util.get(y, [])
        ws.append([
            y, round(hc, 0), round(hog_t, 0), round(gross_t, 0),
            round(avg_hog, 2), round(yr_feed.get(y, 0.0) / 1000.0, 0),
            round(max(bios) / 1000.0, 0) if bios else 0,
            round((sum(bios) / len(bios)) / 1000.0, 0) if bios else 0,
            round(sum(utils) / len(utils), 1) if utils else 0,
        ])
    widths = {1: 8, 2: 20, 3: 16, 4: 17, 5: 16, 6: 12, 7: 17, 8: 17, 9: 20}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_daily_harvest_schedule(
    wb,
    harvest_events,
    forecast_start,
    default_hog_yield: float,
    facility_limits_hog: dict,
    sheet_name: str = "Daily Harvest Schedule",
) -> None:
    """Mon-Fri split of each week's COMBINED harvest with HOG conversions.

    All Harvest events in the same ISO week are combined: their count + live
    biomass are summed and distributed evenly across that week's five Mon-Fri
    operating days, with blended average weights (total biomass / total fish),
    a per-week Total row, and a blank separator. The Tank/Batch columns list
    every tank/batch that contributed. No forecast_start clip (manual override
    weeks are dated before the shifted start and must show their full 5 days).
    """
    from collections import defaultdict
    from datetime import timedelta
    from openpyxl.styles import Font
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["DAILY HARVEST SCHEDULE"])
    ws.append([f"Each week's harvest (all tanks combined) split Mon-Fri. "
               f"Forecast start {forecast_start}"])
    ws.append([])
    ws.append([
        "Year", "Week", "Date", "Tank", "Batch", "Count (fish)",
        "Weight (kg HOG)", "Avg Weight (kg HOG)",
        "Live Weight (kg)", "Avg Weight (kg live)",
    ])

    # Combine every harvest event by ISO week.
    by_week: dict = defaultdict(
        lambda: {"count": 0.0, "live_kg": 0.0, "tanks": set(), "batches": set(),
                 "ev_date": None})
    for ev in harvest_events:
        ev_date = ev.event_date.date() if hasattr(ev.event_date, "date") else ev.event_date
        rec = by_week[iso_week_label(ev_date)]
        rec["count"] += ev.count
        rec["live_kg"] += ev.count * ev.avg_wt_g / 1000.0
        rec["tanks"].add(ev.source_tank_id)
        rec["batches"].add(ev.batch_id)
        if rec["ev_date"] is None or ev_date < rec["ev_date"]:
            rec["ev_date"] = ev_date

    for wk_label in sorted(by_week):
        rec = by_week[wk_label]
        ev_date = rec["ev_date"]
        monday = ev_date - timedelta(days=ev_date.weekday())
        mon_fri = [monday + timedelta(days=i) for i in range(5)]
        n_days = len(mon_fri)
        cnt, live_kg = rec["count"], rec["live_kg"]
        hog_yield = facility_limits_hog.get(wk_label, default_hog_yield)
        hog_kg = live_kg * hog_yield
        live_avg_kg = (live_kg / cnt) if cnt else 0.0            # blended live kg/fish
        hog_avg_kg = live_avg_kg * hog_yield                     # blended HOG kg/fish
        tanks = ", ".join(str(t) for t in sorted(rec["tanks"]))
        batches = ", ".join(sorted(rec["batches"]))
        iso_y, iso_w, _ = ev_date.isocalendar()
        for d in mon_fri:
            ws.append([
                iso_y, iso_w, d, tanks, batches,
                round(cnt / n_days, 0),
                round(hog_kg / n_days, 0),
                round(hog_avg_kg, 3),
                round(live_kg / n_days, 0),
                round(live_avg_kg, 3),
            ])
        ws.append([
            iso_y, iso_w, "Total", tanks, batches,
            round(cnt, 0), round(hog_kg, 0), round(hog_avg_kg, 3),
            round(live_kg, 0), round(live_avg_kg, 3),
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        ws.append([])
    widths = {1: 6, 2: 6, 3: 12, 4: 14, 5: 14, 6: 12, 7: 15, 8: 17, 9: 14, 10: 17}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_harvest_report(
    wb,
    harvest_events,
    default_hog_yield: float,
    facility_limits_hog: dict,
    forecast_start=None,
    sheet_name: str = "HarvestReport",
) -> None:
    """Per-event harvest forecast (one row per tank harvest), matches reference.

    Columns: Year, Month, Week, Date, Tank, Batch, Count (fish),
    Gross Biomass (kg), HOG Biomass (kg), Avg Live Wt (kg).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["HARVEST FORECAST"])
    fs = forecast_start.date() if hasattr(forecast_start, "date") else forecast_start
    ws.append([f"Generated from forecast starting {fs}" if fs else "Generated from forecast"])
    ws.append([])
    ws.append([
        "Year", "Month", "Week", "Date", "Tank", "Batch",
        "Count (fish)", "Gross Biomass (kg)", "HOG Biomass (kg)", "Avg Live Wt (kg)",
    ])

    def _d(ev_date):
        return ev_date.date() if hasattr(ev_date, "date") else ev_date

    evs = sorted(harvest_events, key=lambda e: (_d(e.event_date), e.source_tank_id))
    for ev in evs:
        d = _d(ev.event_date)
        wk = iso_week_label(ev.event_date)
        hog_yield = facility_limits_hog.get(wk, default_hog_yield)
        gross = ev.count * ev.avg_wt_g / 1000.0
        ws.append([
            d.year,
            d.replace(day=1),
            wk,
            d,
            ev.source_tank_id,
            ev.batch_id,
            round(ev.count, 0),
            round(gross, 0),
            round(gross * hog_yield, 0),
            round(ev.avg_wt_g / 1000.0, 2),
        ])
    widths = {1: 7, 2: 12, 3: 10, 4: 12, 5: 7, 6: 8, 7: 13, 8: 18, 9: 17, 10: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _feed_by_type_week(batch_locations, biology_states_by_batch, tables,
                       batches=None, sixn_move_in_feed=None):
    """(feed_name, week_label) -> kg/week, plus a week_label -> week_start map.

    OG/SW feed comes from REALIZED batch_locations via realized_feed_kg_day, so
    totals match the WeeklyReport Feed column and the Advisory feed/day. FW/EGG
    feed comes from the projection (FW fish live in FW tanks, which are absent
    from batch_locations). Phantom unharvested SW projection fish are excluded —
    they never appear in batch_locations — so later-year feed no longer balloons.

    STARVE tank-weeks (6N purge-mode depuration) eat nothing, so they are
    excluded from realized feed; the 4 pre-transfer feed-days each 6N move-in
    cohort ate in its source tank are added back from `sixn_move_in_feed`
    ((batch, week, feed_type) -> kg).
    """
    from collections import defaultdict
    from .biology import realized_feed_kg_day, _feed_type_for_size
    ftw: dict[tuple[str, str], float] = defaultdict(float)
    wk_start: dict[str, object] = {}
    for r in batch_locations:
        wk_start.setdefault(r.week_label, r.week_start)
        if getattr(r, "stage", "") == "STARVE":
            continue  # off-feed depuration tank-week
        fkg = _row_feed_kg_day(r, batches, tables) * 7.0
        if fkg:
            ftw[(_feed_type_for_size(tables, r.avg_wt_g), r.week_label)] += fkg
    for (bid, wk, ftype), kg in (sixn_move_in_feed or {}).items():
        if kg:
            ftw[(ftype, wk)] += kg
    for states in (biology_states_by_batch or {}).values():
        for s in states:
            if s.stage in ("FW", "EGG") and s.feed_kg_week:
                wk_start.setdefault(s.week_label, s.week_start)
                ftw[(s.feed_type, s.week_label)] += s.feed_kg_week
    return ftw, wk_start


def _feed_by_batch_type_week(batch_locations, biology_states_by_batch, tables,
                             batches=None, sixn_move_in_feed=None):
    """(batch_id, feed_name, week_label) -> kg/week, plus week_label -> week_start.

    Same realized-OG/SW + projected-FW feed sourcing as `_feed_by_type_week`
    (so per-batch totals reconcile to the by-type block), but keyed by batch
    so the operator can see feed per batch (the legacy "FEED BY BATCH & TYPE"
    block). FW/EGG projection rows are keyed by their owning batch. STARVE 6N
    depuration tank-weeks are excluded; the 4 pre-transfer move-in feed-days are
    added back from `sixn_move_in_feed`.
    """
    from collections import defaultdict
    from .biology import realized_feed_kg_day, _feed_type_for_size
    fbtw: dict[tuple[str, str, str], float] = defaultdict(float)
    wk_start: dict[str, object] = {}
    for r in batch_locations:
        wk_start.setdefault(r.week_label, r.week_start)
        if getattr(r, "stage", "") == "STARVE":
            continue  # off-feed depuration tank-week
        fkg = _row_feed_kg_day(r, batches, tables) * 7.0
        if fkg:
            fbtw[(r.batch_id, _feed_type_for_size(tables, r.avg_wt_g),
                  r.week_label)] += fkg
    for (bid, wk, ftype), kg in (sixn_move_in_feed or {}).items():
        if kg:
            fbtw[(bid, ftype, wk)] += kg
    for batch_id, states in (biology_states_by_batch or {}).items():
        for s in states:
            if s.stage in ("FW", "EGG") and s.feed_kg_week:
                wk_start.setdefault(s.week_label, s.week_start)
                fbtw[(batch_id, s.feed_type, s.week_label)] += s.feed_kg_week
    return fbtw, wk_start


def write_feed_forecast_weekly(
    wb,
    batch_locations,
    biology_states_by_batch,
    forecast_start,
    tables=None,
    batches=None,
    sheet_name: str = "FeedForecastWeekly",
    sixn_move_in_feed=None,
) -> None:
    """Per-week feed forecast as a Feed Type x Week matrix (kg/week).

    Matches the reference report: rows are feed types (ordered by Max Size),
    columns are ISO weeks with a week-start date sub-row, cells are the weekly
    feed (kg) consumed by fish in that feed-type size band, plus a Total row.
    Feed source = realized OG/SW + projected FW (see _feed_by_type_week).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ftw, wk_start = _feed_by_type_week(
        batch_locations, biology_states_by_batch, tables, batches,
        sixn_move_in_feed=sixn_move_in_feed)
    weeks = sorted(wk_start.keys())
    # Feed-type row order (by Max Size) from biology tables; fall back to the
    # feed_type values present if tables weren't passed.
    if tables is not None and getattr(tables, "feed_types", None):
        ftypes = sorted(tables.feed_types, key=lambda x: x[0])  # (max_size_g, name)
    else:
        ftypes = [(None, n) for n in sorted({ft for (ft, _) in ftw})]

    ws.append(["WEEKLY FEED FORECAST BY TYPE (kg)"])
    ws.append([])
    ws.append(["Feed Type", "Max Size (g)"] + weeks)
    ws.append(["", ""] + [wk_start[w] for w in weeks])
    totals = [0.0] * len(weeks)
    for max_size, name in ftypes:
        row = [name, max_size]
        for i, w in enumerate(weeks):
            v = ftw.get((name, w), 0.0)
            row.append(round(v, 0))
            totals[i] += v
        ws.append(row)
    ws.append(["Total (kg)", ""] + [round(t, 0) for t in totals])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12


def write_feed_forecast_monthly(
    wb,
    batch_locations,
    biology_states_by_batch,
    forecast_start,
    tables=None,
    batches=None,
    sheet_name: str = "FeedForecastMonthly",
    sixn_move_in_feed=None,
) -> None:
    """Per-month feed forecast as a Feed Type x Month matrix (kg/month).

    Matches the reference: rows are feed types (by Max Size), columns are
    month-start dates, cells are monthly feed (kg) by feed-type band, plus a
    Grand Total row. Feed source = realized OG/SW + projected FW (rolled up
    from the same per-week aggregation as the weekly sheet).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    from collections import defaultdict
    from datetime import date as _date
    from .time_grid import calendar_day_month_split
    ftw, wk_start = _feed_by_type_week(
        batch_locations, biology_states_by_batch, tables, batches,
        sixn_move_in_feed=sixn_move_in_feed)
    # Feed is a DAILY flow, so a week straddling a month boundary is split
    # between the two months by calendar-day fraction (not dumped into the
    # week-start's month). Pure calendar attribution — totals are unchanged.
    ftm: dict[tuple[str, str], float] = defaultdict(float)  # (feed_type, month) -> kg
    months: set[str] = set()
    for (name, wk), v in ftw.items():
        ws_ = wk_start.get(wk)
        for (yr, mon), frac in calendar_day_month_split(ws_).items():
            mo = f"{yr}-{mon:02d}"
            months.add(mo)
            ftm[(name, mo)] += v * frac
    months_sorted = sorted(months)
    mo_dates = [_date(int(m[:4]), int(m[5:7]), 1) for m in months_sorted]
    if tables is not None and getattr(tables, "feed_types", None):
        ftypes = sorted(tables.feed_types, key=lambda x: x[0])
    else:
        ftypes = [(None, n) for n in sorted({ft for (ft, _) in ftm})]

    ws.append(["MONTHLY FEED FORECAST BY TYPE (kg)"])
    ws.append([])
    ws.append(["Feed Type", "Max Size (g)"] + mo_dates)
    totals = [0.0] * len(months_sorted)
    for max_size, name in ftypes:
        row = [name, max_size]
        for i, m in enumerate(months_sorted):
            v = ftm.get((name, m), 0.0)
            row.append(round(v, 0))
            totals[i] += v
        ws.append(row)
    ws.append(["Grand Total", ""] + [round(t, 0) for t in totals])

    # ----- FEED BY BATCH & TYPE block (matches the legacy layout) -----
    # One row per (batch, feed-type) the batch consumes, grouped by batch
    # (batch label on the first row of each group), feed types in size order,
    # month columns identical to the by-type block above. Same feed source,
    # so per-batch totals reconcile to the by-type grand total.
    fbtw, _ = _feed_by_batch_type_week(
        batch_locations, biology_states_by_batch, tables, batches,
        sixn_move_in_feed=sixn_move_in_feed)
    fbtm: dict[tuple[str, str, str], float] = defaultdict(float)  # (batch, type, month)
    for (bid, name, wk), v in fbtw.items():
        ws_ = wk_start.get(wk)
        for (yr, mon), frac in calendar_day_month_split(ws_).items():
            fbtm[(bid, name, f"{yr}-{mon:02d}")] += v * frac
    # Max-size order for feed types (same ordering basis as the by-type block).
    size_of = {name: (ms if ms is not None else 0.0) for ms, name in ftypes}

    def _batch_sort_key(b):
        # Order B41, B42, ... numerically; non-standard ids sort after, by name.
        s = str(b)
        if s[:1] == "B" and s[1:].isdigit():
            return (0, int(s[1:]), "")
        return (1, 0, s)

    batch_ids = sorted({bid for (bid, _n, _m) in fbtm}, key=_batch_sort_key)
    ws.append([])
    ws.append(["FEED BY BATCH & TYPE (kg)"])
    ws.append(["Batch", "Feed Type"] + mo_dates)
    for bid in batch_ids:
        names = sorted({n for (b, n, _m) in fbtm if b == bid},
                       key=lambda n: (size_of.get(n, 0.0), n))
        first = True
        for name in names:
            row = [bid if first else None, name]
            for m in months_sorted:
                row.append(round(fbtm.get((bid, name, m), 0.0), 0))
            ws.append(row)
            first = False

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16


# Open/Close ledger column headers (shared by Weekly + Monthly reports).
_LEDGER_COLS = [
    "Open_Count (fish)", "Open_AvgWt (g)", "Open_Bio (kg)",
    "Close_Count (fish)", "Close_AvgWt (g)", "Close_Bio (kg)",
    "Density (kg/m³)", "SGR (%/day)", "Gross_Growth (kg)",
    "Net_Production (kg)", "Feed (kg)", "SFR (%/day)",
    "Bio_FCR (ratio)", "Econ_FCR (ratio)",
    "Mort_Count (fish)", "Mort_Bio (kg)",
    "Harv_Count (fish)", "Harv_Gross (kg)", "Harv_HOG (kg)", "Harv_AvgWt_HOG (g)",
    "Cull_Count (fish)", "Cull_Bio (kg)",
    "Input_Count (fish)", "Xfer_In (fish)", "Xfer_Out (fish)",
    "Count_Check (fish)", "Bio_Check (kg)",
]


def _build_batch_week_ledger(
    batch_locations, harvest_events, batch_week_states,
    transfer_events=None, batches=None, tables=None, hog_yield=0.0,
    hog_overrides=None, sixn_move_in_feed=None,
    tranog_events=None, og_mort_states=None,
):
    """Assemble a per-(batch, week) open/close production ledger.

    Open = prior week's realized close (chained); Close = this week's realized
    state (BatchLocations aggregate, falling back to the biology projection for
    pre-TranOG FW weeks with no tank rows). Growth/feed/FCR/mortality columns
    are derived to mirror the reference report definitions:
      Gross_Growth = (close_bio - open_bio) + mort_bio + harv_gross + cull_bio - input_bio
      Net_Production = Gross_Growth - mort_bio
      SGR = ln(close_wt/open_wt)/7*100 ; SFR = feed / avg_bio / 7 * 100
      Bio_FCR = feed / gross_growth ; Econ_FCR = feed / net_production
    Bio_Check is 0 by construction; Count_Check carries the small count residual.
    Returns a list of row dicts ordered by (batch, week).
    """
    from collections import defaultdict
    from math import log
    from .biology import realized_feed_kg_day

    # Realized close per (batch, week) from BatchLocations.
    rl: dict[tuple, dict] = defaultdict(
        lambda: {"count": 0.0, "bio": 0.0, "wt_sum": 0.0, "week_start": None})
    feed: dict[tuple, float] = defaultdict(float)
    for r in batch_locations:
        key = (r.batch_id, r.week_label)
        e = rl[key]
        e["count"] += r.count
        e["bio"] += r.biomass_kg
        e["wt_sum"] += r.avg_wt_g * r.count
        e["week_start"] = r.week_start
        # STARVE tank-weeks (6N depuration) eat nothing (helper returns 0).
        feed[key] += _row_feed_kg_day(r, batches, tables) * 7.0
    # 6N purge move-in fish ate 4 pre-transfer days in their source tank (now
    # shown in 6N = STARVE, excluded above) — add that real feed back so the
    # ledger Feed column matches the FeedForecast / YearlySummary totals.
    for (bid, _wk, _ftype), kg in (sixn_move_in_feed or {}).items():
        if kg:
            feed[(bid, _wk)] += kg

    # Harvest per (batch, week).
    harv: dict[tuple, dict] = defaultdict(lambda: {"count": 0.0, "gross": 0.0, "wt_sum": 0.0})
    for ev in harvest_events:
        key = (ev.batch_id, iso_week_label(ev.event_date))
        e = harv[key]
        e["count"] += ev.count
        e["gross"] += ev.count * ev.avg_wt_g / 1000.0
        e["wt_sum"] += ev.avg_wt_g * ev.count

    # Transfers per (batch, week): intra-batch moves, so In == Out (net 0).
    xfer: dict[tuple, float] = defaultdict(float)
    for ev in (transfer_events or ()):
        moved = sum(getattr(d, "count", 0.0) for d in getattr(ev, "destinations", []))
        if not moved:
            moved = getattr(ev, "count_transferred", 0.0)
        xfer[(ev.batch_id, iso_week_label(ev.event_date))] += moved

    # TranOG fresh-stocking inflow per (batch, week): fish ENTERING OG from FW /
    # appearing in-flight, with NO chained OG predecessor (opening OG balance 0).
    # Credited as input at a genuine FW->OG boundary week only (open reset to 0
    # there) — a two-engine handoff where the FW projection's count does not flow
    # by count into the pick's realized OG entry. (Global reports only; the
    # controller passes no tranog_events -> this map stays empty.)
    tranog_in: dict[tuple, float] = defaultdict(float)
    for ev in (tranog_events or ()):
        wk = iso_week_label(ev.event_date)
        for d in getattr(ev, "destinations", []):
            tranog_in[(ev.batch_id, wk)] += getattr(d, "count", 0.0)

    # Cull / mortality% / input / biology fallback, keyed by (batch, week).
    cull: dict[tuple, dict] = defaultdict(lambda: {"count": 0.0, "bio": 0.0})
    mortpct: dict[tuple, float] = {}
    mortc: dict[tuple, float] = {}   # realized mortality count (fish) per week
    inputc: dict[tuple, float] = defaultdict(float)
    bio_state: dict[tuple, object] = {}
    for s in batch_week_states or ():
        key = (s.batch_id, s.week_label)
        bio_state[key] = s
        mortpct[key] = s.mortality_pct_weekly
        mortc[key] = getattr(s, "mort_count_week", 0.0)
        # Credit an FW-projection cull ONLY on a week NOT realized in OG
        # (BatchLocations). fw_states are FW/EGG-only and rl is OG-only, so a
        # genuine FW->OG reconciliation cull is consumed on the FW-projection
        # fallback close at a non-OG week; if `key` IS in rl the close already
        # comes from realized OG (no cull), so an fw_states cull there is spurious
        # double-tracking (a batch advanced into OG yet still FW-projected — e.g.
        # a manual-window in-flight batch left in fw_inflight). Mirrors the same
        # `key not in rl` guard already used for FW feed below.
        if s.cull_count_week > 0 and key not in rl:
            cull[key]["count"] += s.cull_count_week
            cull[key]["bio"] += s.cull_biomass_kg_week
        if s.week_from_input == 0:
            # STOCKING week: the input flow is the count that was STOCKED =
            # the closing balance plus this week's mortality + culls (open is 0).
            # Using the weekly-mean count instead left a residual in Count_Check.
            inputc[key] += (getattr(s, "close_count", 0.0) or s.count) \
                + getattr(s, "mort_count_week", 0.0) + s.cull_count_week
        # FW (pre-TranOG) feed: FW fish live in FW tanks, which are NOT in
        # batch_locations (OG-only). Pull their feed from the projection so the
        # ledger + FeedForecast cover small-fish feed. SW/OG feed always comes
        # from realized batch_locations above — never the projection, which
        # would re-introduce phantom unharvested fish (see close_vals gate).
        if s.stage in ("FW", "EGG") and s.feed_kg_week and key not in rl:
            feed[key] += s.feed_kg_week

    # OG-phase weekly mortality % (the pick's per-OG-batch-week rate). fw_states
    # cover only FW weeks, so OG weeks would otherwise read mort=0 and leak the
    # ~decline into Count_Check. Fill OG weeks only; setdefault never overrides an
    # FW-state rate (FW weeks and OG weeks don't collide anyway). Controller passes
    # no og_mort_states -> no-op.
    for ms in (og_mort_states or ()):
        mortpct.setdefault((ms.batch_id, ms.week_label), ms.mortality_pct_weekly)

    def close_vals(key):
        e = rl.get(key)
        if e and e["count"] > 0:
            return e["count"], e["wt_sum"] / e["count"], e["bio"], e["week_start"]
        # Biology fallback ONLY for genuine pre-TranOG FW/EGG weeks. A projected
        # SW/STARVE week with no realized placement is a phantom (batch never
        # placed, or already harvested out) and must not contribute biomass —
        # else facility totals balloon far past the real ~4M kg.
        s = bio_state.get(key)
        if s and s.stage in ("FW", "EGG"):
            # Close = END-of-week balance (post mid-week cull), not the weekly
            # mean, so the FW ledger chains open->close consistently across the
            # FW->OG TranOG reconciliation-cull week (mean under-counts the drop
            # and leaks a ~cull-sized Count_Check residual into the next week).
            cc = s.close_count if s.close_count > 0 else s.count
            cw = s.close_avg_weight_g if s.close_avg_weight_g > 0 else s.avg_weight_g
            cb = s.close_biomass_kg if s.close_biomass_kg > 0 else s.biomass_kg
            return cc, cw, cb, s.week_start
        return 0.0, 0.0, 0.0, None

    # All (batch, week) cells: union of realized + projected weeks.
    by_batch: dict[str, set] = defaultdict(set)
    for (b, wk) in set(rl) | set(bio_state) | set(harv) | set(cull):
        by_batch[b].add(wk)

    rows = []
    for b in sorted(by_batch):
        weeks = sorted(by_batch[b])
        for i, wk in enumerate(weeks):
            cc, cwt, cbio, ws_date = close_vals((b, wk))
            if i == 0:
                s0 = bio_state.get((b, wk))
                if s0 and getattr(s0, "week_from_input", -1) == 0:
                    # STOCKING week: the batch had no opening balance — the
                    # input_count flow CREATES the fish (open 0 -> close). Using
                    # the post-input biology count as the open would double-count
                    # the input in Count_Check (open already includes the input
                    # AND input_count adds it again -> a spurious ~input residual
                    # on every batch's first week). Open is 0 before stocking.
                    oc, owt, obio = 0.0, 0.0, 0.0
                elif s0:
                    # Start-of-week balance (before the week's losses), not the
                    # weekly mean — the mean mis-states the open on week 0 (no
                    # prior-week close to chain from), leaving a ~half-week
                    # mortality residual in Count_Check. Fall back to mean if the
                    # open fields aren't populated.
                    oc = s0.open_count if getattr(s0, "open_count", 0.0) > 0 else s0.count
                    owt = (s0.open_avg_weight_g if getattr(s0, "open_avg_weight_g", 0.0) > 0
                           else s0.avg_weight_g)
                    obio = (s0.open_biomass_kg if getattr(s0, "open_biomass_kg", 0.0) > 0
                            else s0.biomass_kg)
                elif (b, wk) in tranog_in:
                    # In-flight OG batch's first ledger week entered via TranOG with
                    # no opening balance -> reset open to 0 (inflow credited below).
                    oc, owt, obio = 0.0, 0.0, 0.0
                else:
                    oc, owt, obio = cc, cwt, cbio
            else:
                prev_wk = weeks[i - 1]
                prev_s = bio_state.get((b, prev_wk))
                prev_is_fw_proj = (prev_s is not None
                                   and prev_s.stage in ("FW", "EGG")
                                   and (b, prev_wk) not in rl)
                if (b, wk) in tranog_in and (b, wk) in rl and prev_is_fw_proj:
                    # FW->OG boundary: the pick FRESH-STOCKS the whole OG entry via
                    # TranOG; the OG opening balance is 0 (the FW projection is a
                    # separate track whose close does not flow by COUNT into OG).
                    # Reset open + credit the inflow (below); chaining the FW close
                    # here would leave the two-engine handoff gap as a residual.
                    oc, owt, obio = 0.0, 0.0, 0.0
                else:
                    oc, owt, obio, _ = close_vals((b, prev_wk))
            h = harv.get((b, wk), {"count": 0.0, "gross": 0.0, "wt_sum": 0.0})
            cu = cull.get((b, wk), {"count": 0.0, "bio": 0.0})
            # Mortality count: for a FW/EGG PROJECTION week (close comes from the
            # biology, not realized BatchLocations) use the REALIZED mortality the
            # daily sim actually applied — open*weekly_rate% mis-counts when the
            # mortality table steps mid-week (early FW). Realized OG weeks keep the
            # rate-based estimate (their close is the realized placement).
            _s_st = bio_state.get((b, wk))
            _rlz = rl.get((b, wk))
            if (_s_st is not None and _s_st.stage in ("FW", "EGG")
                    and not (_rlz and _rlz.get("count", 0) > 0)):
                mort_count = mortc.get((b, wk), oc * mortpct.get((b, wk), 0.0) / 100.0)
            else:
                mort_count = oc * mortpct.get((b, wk), 0.0) / 100.0
            mort_bio = mort_count * owt / 1000.0
            input_count = inputc.get((b, wk), 0.0) + tranog_in.get((b, wk), 0.0)
            input_bio = input_count * owt / 1000.0
            xf = xfer.get((b, wk), 0.0)
            harv_gross = h["gross"]
            # Per-week HOG yield override (matches HarvestReport/HarvestPlan);
            # falls back to the scalar default when no override for the week.
            _hy = hog_overrides.get(wk, hog_yield) if hog_overrides else hog_yield
            harv_hog = harv_gross * _hy
            harv_avg_hog = ((h["wt_sum"] / h["count"]) * _hy) if h["count"] > 0 else 0.0
            gross_growth = (cbio - obio) + mort_bio + harv_gross + cu["bio"] - input_bio
            net_prod = gross_growth - mort_bio
            f = feed.get((b, wk), 0.0)
            avg_bio = (obio + cbio) / 2.0
            sgr = (log(cwt / owt) / 7.0 * 100.0) if owt > 0 and cwt > 0 else 0.0
            sfr = (f / avg_bio / 7.0 * 100.0) if avg_bio > 0 else 0.0
            bio_fcr = (f / gross_growth) if gross_growth > 0 else 0.0
            econ_fcr = (f / net_prod) if net_prod > 0 else 0.0
            count_check = (oc - mort_count - h["count"] - cu["count"]
                           + input_count + xf - xf - cc)
            # Bio_Check is 0 by construction (gross_growth balances the ledger).
            bio_check = (obio + gross_growth - mort_bio - harv_gross - cu["bio"]
                         + input_bio - cbio)
            if oc <= 0 and cc <= 0 and h["count"] <= 0 and cu["count"] <= 0:
                continue
            rows.append({
                "batch": b, "week": wk, "week_start": ws_date,
                "open_count": oc, "open_wt": owt, "open_bio": obio,
                "close_count": cc, "close_wt": cwt, "close_bio": cbio,
                "sgr": sgr, "gross_growth": gross_growth, "net_prod": net_prod,
                "feed": f, "sfr": sfr, "bio_fcr": bio_fcr, "econ_fcr": econ_fcr,
                "mort_count": mort_count, "mort_bio": mort_bio,
                "harv_count": h["count"], "harv_gross": harv_gross,
                "harv_hog": harv_hog, "harv_avg_hog": harv_avg_hog,
                "cull_count": cu["count"], "cull_bio": cu["bio"],
                "input_count": input_count, "xfer_in": xf, "xfer_out": xf,
                "count_check": count_check, "bio_check": bio_check,
            })
    return rows


def _ledger_value_cells(d: dict) -> list:
    """Format the 27 shared open/close ledger value columns from a row dict."""
    return [
        round(d["open_count"], 0), round(d["open_wt"], 1), round(d["open_bio"], 0),
        round(d["close_count"], 0), round(d["close_wt"], 1), round(d["close_bio"], 0),
        0, round(d["sgr"], 4), round(d["gross_growth"], 0),
        round(d["net_prod"], 0), round(d["feed"], 0), round(d["sfr"], 4),
        round(d["bio_fcr"], 2), round(d["econ_fcr"], 2),
        round(d["mort_count"], 0), round(d["mort_bio"], 1),
        round(d["harv_count"], 0), round(d["harv_gross"], 1),
        round(d["harv_hog"], 1), round(d["harv_avg_hog"], 1),
        round(d["cull_count"], 0), round(d["cull_bio"], 1),
        round(d["input_count"], 0), round(d["xfer_in"], 0), round(d["xfer_out"], 0),
        round(d["count_check"], 0), round(d["bio_check"], 0),
    ]


def write_weekly_report(
    wb,
    batch_locations,
    harvest_events,
    batch_week_states=None,
    transfer_events=None,
    batches=None,
    tables=None,
    scenario_name: str = "",
    hog_yield: float = 0.0,
    hog_overrides=None,
    sixn_move_in_feed=None,
    tranog_events=None,
    og_mort_states=None,
    sheet_name: str = "WeeklyReport",
) -> None:
    """Per-(week, batch) open/close production ledger (matches reference format).

    Columns: Scenario, Week, Week_Start, Batch, then the 27 shared open/close
    ledger columns (open/close count-avgwt-bio, density, SGR, growth, feed, FCR,
    mortality, harvest, cull, transfers, and reconciliation checks).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append([f"{sheet_name} - populated by RunForecast"])
    ws.append([])
    ws.append([])
    ws.append(["Scenario", "Week", "Week_Start", "Batch"] + _LEDGER_COLS)

    rows = _build_batch_week_ledger(
        batch_locations, harvest_events, batch_week_states,
        transfer_events, batches, tables, hog_yield, hog_overrides,
        sixn_move_in_feed=sixn_move_in_feed,
        tranog_events=tranog_events, og_mort_states=og_mort_states)
    for d in rows:
        ws.append([scenario_name, d["week"], d["week_start"], d["batch"]]
                  + _ledger_value_cells(d))
    for c in range(1, 5 + len(_LEDGER_COLS)):
        ws.column_dimensions[get_column_letter(c)].width = 14


def write_monthly_report(
    wb,
    batch_locations,
    harvest_events,
    batch_week_states=None,
    transfer_events=None,
    batches=None,
    tables=None,
    scenario_name: str = "",
    hog_yield: float = 0.0,
    hog_overrides=None,
    forecast_start=None,
    sixn_move_in_feed=None,
    tranog_events=None,
    og_mort_states=None,
    sheet_name: str = "MonthlyReport",
) -> None:
    """Per-(month, batch) open/close production ledger (matches reference format).

    Rolls the weekly ledger up to months: Open = first week's open in the month,
    Close = last week's close; flows (growth, feed, mortality, harvest, cull,
    transfers, input) are summed; SGR/SFR/FCR are recomputed from the monthly
    aggregates. Columns mirror the weekly report minus Week_Start.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append([f"{sheet_name} - populated by RunForecast"])
    ws.append([])
    ws.append([])
    ws.append(["Scenario", "Month", "Batch"] + _LEDGER_COLS)

    from collections import defaultdict
    from math import log

    weekly = _build_batch_week_ledger(
        batch_locations, harvest_events, batch_week_states,
        transfer_events, batches, tables, hog_yield, hog_overrides,
        sixn_move_in_feed=sixn_move_in_feed,
        tranog_events=tranog_events, og_mort_states=og_mort_states)

    # Roll the weekly ledger up to calendar months, splitting any week that
    # straddles a month boundary into its true month. CONTINUOUS flows (growth,
    # feed, mortality, cull, input, transfers) happen every calendar day, so they
    # split by CALENDAR-DAY fraction. HARVEST happens only Mon-Fri, so it splits
    # by WORKING-DAY fraction -- identical to the HarvestPlan Report, so the two
    # sheets' monthly harvest now tie out exactly. The two fractions differ on
    # boundary weeks, so the month-boundary Open/Close is advanced by the harvest
    # part and the non-harvest part SEPARATELY, each at its own cumulative
    # fraction. Close is therefore exactly open + (the month's actual flows), so
    # the monthly Count_Check / Bio_Check stay ~0 (the ledger stays internally
    # consistent) even though harvest and the daily flows attribute differently.
    from .time_grid import calendar_day_month_split, working_day_month_split
    from datetime import date as _date
    FLOW_KEYS = ("gross_growth", "net_prod", "feed", "mort_count", "mort_bio",
                 "harv_count", "harv_gross", "harv_hog", "cull_count", "cull_bio",
                 "input_count", "xfer_in", "xfer_out", "count_check", "bio_check")
    HARVEST_KEYS = ("harv_count", "harv_gross", "harv_hog")

    def _wk_date(w):
        # Week-start date for proration. Some ledger rows (e.g. harvest-only)
        # carry no week_start; fall back to the ISO week label's Monday (matches
        # the prior _month_key fallback — d["week"][6:8] is the week NUMBER).
        d = w["week_start"]
        if d is not None and not isinstance(d, str):
            return d.date() if hasattr(d, "date") else d
        try:
            return _date.fromisocalendar(int(w["week"][:4]), int(w["week"][6:8]), 1)
        except Exception:
            return None

    by_batch: dict[str, list] = defaultdict(list)
    for d in weekly:
        by_batch[d["batch"]].append(d)

    rows_out: list[tuple] = []  # (month, batch, agg-accumulator)
    for b, wks in by_batch.items():
        acc: dict[str, dict] = {}
        for w in sorted(wks, key=lambda x: x["week"]):
            wkd = _wk_date(w)
            if wkd is None:
                continue  # unparseable week — cannot attribute, skip (rare)
            split_c = calendar_day_month_split(wkd)   # daily flows
            split_w = working_day_month_split(wkd)     # harvest (no fs clip —
            #        pre-start manual weeks must split by working day like the rest)
            oc, ob = w["open_count"], w["open_bio"]
            hc, hg = w["harv_count"], w["harv_gross"]
            dc, db = w["close_count"] - oc, w["close_bio"] - ob
            dc_h, db_h = -hc, -hg              # harvest part of the net delta
            dc_n, db_n = dc - dc_h, db - db_h  # everything-else part of the delta
            cum_c = cum_w = 0.0
            for (yr, mon) in sorted(set(split_c) | set(split_w)):
                fc = split_c.get((yr, mon), 0.0)   # calendar-day fraction
                fw = split_w.get((yr, mon), 0.0)   # working-day fraction
                mo = f"{yr}-{mon:02d}"
                a = acc.get(mo)
                if a is None:
                    a = {k: 0.0 for k in FLOW_KEYS}
                    a["days"] = 0.0
                    # state at month START: advance harvest + non-harvest deltas
                    # each by its own cumulative fraction
                    a["open_count"] = oc + cum_c * dc_n + cum_w * dc_h
                    a["open_bio"] = ob + cum_c * db_n + cum_w * db_h
                    acc[mo] = a
                a["close_count"] = oc + (cum_c + fc) * dc_n + (cum_w + fw) * dc_h
                a["close_bio"] = ob + (cum_c + fc) * db_n + (cum_w + fw) * db_h
                a["days"] += fc * 7.0
                for k in FLOW_KEYS:
                    a[k] += w[k] * (fw if k in HARVEST_KEYS else fc)
                cum_c += fc
                cum_w += fw
        for mo, a in acc.items():
            rows_out.append((mo, b, a))

    for mo, b, a in sorted(rows_out, key=lambda x: (x[0], x[1])):
        open_count, open_bio = a["open_count"], a["open_bio"]
        close_count, close_bio = a["close_count"], a["close_bio"]
        open_wt = (open_bio / open_count * 1000.0) if open_count > 0 else 0.0
        close_wt = (close_bio / close_count * 1000.0) if close_count > 0 else 0.0
        gross_growth, net_prod, f = a["gross_growth"], a["net_prod"], a["feed"]
        avg_bio = (open_bio + close_bio) / 2.0
        days = a["days"] or 7.0
        sgr = (log(close_wt / open_wt) / days * 100.0) if open_wt > 0 and close_wt > 0 else 0.0
        sfr = (f / avg_bio / days * 100.0) if avg_bio > 0 else 0.0
        harv_count = a["harv_count"]
        agg = {
            "batch": b, "week": mo, "week_start": None,
            "open_count": open_count, "open_wt": open_wt, "open_bio": open_bio,
            "close_count": close_count, "close_wt": close_wt, "close_bio": close_bio,
            "sgr": sgr, "gross_growth": gross_growth, "net_prod": net_prod,
            "feed": f, "sfr": sfr,
            "bio_fcr": (f / gross_growth) if gross_growth > 0 else 0.0,
            "econ_fcr": (f / net_prod) if net_prod > 0 else 0.0,
            "mort_count": a["mort_count"], "mort_bio": a["mort_bio"],
            "harv_count": harv_count, "harv_gross": a["harv_gross"],
            "harv_hog": a["harv_hog"],
            "harv_avg_hog": (a["harv_hog"] * 1000.0 / harv_count) if harv_count > 0 else 0.0,
            "cull_count": a["cull_count"], "cull_bio": a["cull_bio"],
            "input_count": a["input_count"], "xfer_in": a["xfer_in"], "xfer_out": a["xfer_out"],
            "count_check": a["count_check"], "bio_check": a["bio_check"],
        }
        ws.append([scenario_name, mo, b] + _ledger_value_cells(agg))
    for c in range(1, 4 + len(_LEDGER_COLS)):
        ws.column_dimensions[get_column_letter(c)].width = 14


def write_input_conservation_audit(
    wb,
    batches,
    batch_locations,
    harvest_events,
    control,
    tranog_events=None,
    biology_states_by_batch=None,
    manual_fw_balance=None,
    sheet_name: str = "InputConservationAudit",
) -> None:
    """Input-fish conservation: every stocked batch must have a realized fate.

    The TankContinuityAudit proves 0 drift for fish that ARE in tanks, but it is
    BLIND to a batch that is never placed — a dropped TranOG arrival creates no
    tank-week row, so it never unbalances per-tank continuity. This audit closes
    that gap: every batch whose TranOG falls within the forecast horizon MUST
    appear in the realized placement (BatchLocations). Any in-horizon batch with
    no placement is DROPPED — its stocked fish vanished from the plan.

    It ALSO reconciles the otherwise-unaudited FRESHWATER phase: the realized
    count entering seawater (TranOG) vs the operator's planned tran_og_count. A
    material shortfall = the FW survival model delivered fewer smolts than planned
    (a calibration gap) — not lost fish (the realized count is conserved
    downstream), but a real production divergence that was previously only buried
    in a FW-Calibration warning. The FW_Flag surfaces it as a clear row.
    """
    from datetime import timedelta
    _FW_DIVERGENCE_THRESH = 0.05   # flag |realized - planned| / planned beyond this
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    fs = control.forecast_start
    fs = fs.date() if hasattr(fs, "date") else fs
    horizon_end = fs + timedelta(weeks=control.horizon_weeks)
    # A batch only enters OG at the first forecast-week boundary ON/AFTER its
    # TranOG_Date (the OG-entry week — see time_grid.og_entry_week_start). The
    # placement engine can only place a cohort in a week it actually generates,
    # i.e. weeks [0, horizon_weeks). A TranOG_Date that is < horizon_end but
    # whose OG-ENTRY week is week `horizon_weeks` (the boundary one past the last
    # forecast week) lands beyond the plannable window — the engine never sees
    # it, so it is "future", NOT dropped. Use the last plannable week start as
    # the real cutoff so this horizon-edge off-by-one isn't mis-reported as a
    # silent fish loss. (horizon_weeks-1 is the last week index; its start is the
    # latest week a cohort can enter and still be placed.)
    from .time_grid import og_entry_week_start, week_start as _wk_start
    last_week_start = _wk_start(control.horizon_weeks - 1, fs)

    placed_first = {}
    placed_last = {}
    standing = {}
    last_wk = max((r.week_label for r in batch_locations), default=None)
    for r in batch_locations:
        b = r.batch_id
        if b not in placed_first or r.week_label < placed_first[b]:
            placed_first[b] = r.week_label
        if b not in placed_last or r.week_label > placed_last[b]:
            placed_last[b] = r.week_label
        if r.week_label == last_wk:
            standing[b] = standing.get(b, 0.0) + r.count
    harv = {}
    for ev in harvest_events:
        harv[ev.batch_id] = harv.get(ev.batch_id, 0.0) + ev.count
    # Realized count placed into seawater at TranOG per batch (sum of destination
    # counts) — the FW-phase reconciliation reference.
    tranog_placed = {}
    for ev in (tranog_events or []):
        tranog_placed[ev.batch_id] = tranog_placed.get(ev.batch_id, 0.0) + sum(
            getattr(d, "count", 0.0) for d in getattr(ev, "destinations", []))

    # CLOSED FW MASS-BALANCE (audit I2): the projected FRESHWATER phase must
    # conserve fish — the count entering FW reduces to the realized seawater entry
    # ONLY through the modeled losses:
    #   first_FW_count == realized_TranOG + FW_mortality + FW_culls.
    # We reconcile from each batch's FIRST projected FW count, NOT the egg seed:
    # for in-flight batches the egg->startfeed phase is pre-horizon (absent from the
    # states), so input_count (eggs) cannot be reconciled here. But the projected FW
    # phase can, and a fish leak or a mortality/cull-accounting error inside it would
    # otherwise pass every gate — TankContinuity only starts at OG, and the drop/
    # over guards never balanced FW losses. Pre-TranOG there is no harvest, so the
    # balance is clean (no M4 pre/post-harvest mortality ambiguity); the few-percent
    # tolerance absorbs the FW->SW transition-week boundary.
    _FW_BALANCE_THRESH = 0.02   # flag |residual| / first_FW_count beyond 2%
    fw_loss = {}   # batch_id -> (first_fw_count, mortality, cull) over FW/EGG weeks
    for b_id, sl in (biology_states_by_batch or {}).items():
        fws = sorted((s for s in sl if s.stage in ("FW", "EGG")),
                     key=lambda s: s.week_label)
        if not fws:
            continue
        m = sum(getattr(s, "mort_count_week", 0.0) for s in fws)
        c = sum(getattr(s, "cull_count_week", 0.0) for s in fws)
        fw_loss[b_id] = (fws[0].count, m, c)
    fw_bal_base = 0.0       # summed first-FW count of crossed-in-horizon batches
    fw_bal_residual = 0.0   # summed signed residual (first - tranog - mort - cull)
    fw_bal_abs = 0.0        # summed |residual|
    fw_unbalanced = []      # (batch_id, residual, residual_pct)

    dropped_fish = 0.0
    dropped_batches = 0
    in_horizon_input = 0.0
    over_produced = []   # harvested + standing > input (impossible: fish created)
    fw_divergent = []    # realized seawater entry materially off the planned tran_og_count
    rowbuf = []
    for bt in sorted(batches, key=lambda x: x.batch_id):
        bid = bt.batch_id
        tog = bt.tran_og_date
        togd = (tog.date() if hasattr(tog, "date") else tog) if tog else None
        is_placed = bid in placed_first
        hv = harv.get(bid, 0.0)
        _st = standing.get(bid, 0.0)
        # Over-production guard: a batch cannot harvest + still-hold MORE fish
        # than were stocked (mortality + culls only remove). Catches count
        # CREATION — the opposite of a drop, a different conservation breach.
        if (bt.input_count or 0) > 0 and hv + _st > (bt.input_count or 0) * 1.001:
            over_produced.append(bid)
        # OG-entry week start for this batch: the week it can first be placed.
        og_entry = og_entry_week_start(togd, fs) if togd is not None else None
        beyond_plannable = og_entry is not None and og_entry > last_week_start
        in_h = (togd is not None and fs <= togd <= horizon_end
                and not beyond_plannable)
        if is_placed or hv > 0:
            status = "PLACED"
        elif togd is None:
            status = "FW-only (no TranOG)"
        elif togd > horizon_end or beyond_plannable:
            # TranOG_Date past horizon, OR OG-entry week one past the last
            # forecast week (horizon-edge): the engine never gets a week to
            # place it — outside the plannable window, not a dropped fish.
            status = "future (TranOG beyond horizon)"
        elif togd < fs:
            status = "pre-start"
        else:
            status = "*** DROPPED ***"
        at_risk = 0.0
        if in_h:
            in_horizon_input += bt.input_count or 0
        if status == "*** DROPPED ***":
            dropped_batches += 1
            at_risk = bt.input_count or 0.0
            dropped_fish += at_risk
        # FW/TranOG reconciliation: realized seawater entry vs planned tran_og_count.
        planned_tog = bt.tran_og_count or 0
        realized_tog = tranog_placed.get(bid, 0.0)
        fw_surv = (100.0 * realized_tog / bt.input_count) if (realized_tog and bt.input_count) else None
        # A manual fw_to_og batch was transferred to OG in the override window at
        # an OPERATOR-CHOSEN target count — its realized-vs-planned gap is
        # intentional, NOT an FW-survival calibration miss, so it must not be
        # mislabeled "FW UNDER/OVER plan" or counted in fw_divergent.
        _man_fw = (manual_fw_balance or {}).get(bid)
        fw_flag = ""
        if _man_fw is not None:
            fw_flag = "manual fw_to_og"
        elif realized_tog > 0 and planned_tog > 0:
            _div = (realized_tog - planned_tog) / planned_tog
            if _div < -_FW_DIVERGENCE_THRESH:
                fw_flag = "FW UNDER plan"
                fw_divergent.append((bid, _div))
            elif _div > _FW_DIVERGENCE_THRESH:
                fw_flag = "FW OVER plan"
                fw_divergent.append((bid, _div))
        # Closed FW mass-balance for this batch (only meaningful once it has
        # crossed TranOG in-horizon, so there is a realized seawater entry to
        # reconcile the seed against through the modeled FW losses).
        _fwbase, fwm, fwc = fw_loss.get(bid, (0.0, 0.0, 0.0))
        if _man_fw is not None and _fwbase <= 0:
            # Manually transferred FW->OG: the FW phase + cull ran inside the
            # override window (handling mortality + reconcile-to-target cull),
            # so this batch has NO FW biology states for the gate above. The
            # window captured (fw_count_at_transfer, culled); reconcile from
            # those: fw_count == realized_TranOG (placed) + culled. Reported in
            # the FW_Cull column (the cull includes handling mortality).
            _fwbase, fwm, fwc = _man_fw[0], 0.0, _man_fw[1]
        fw_resid = None
        if realized_tog > 0 and _fwbase > 0:
            fw_resid = _fwbase - realized_tog - fwm - fwc
            fw_bal_base += _fwbase
            fw_bal_residual += fw_resid
            fw_bal_abs += abs(fw_resid)
            if abs(fw_resid) > _fwbase * _FW_BALANCE_THRESH:
                fw_unbalanced.append((bid, fw_resid, 100.0 * fw_resid / _fwbase))
        rowbuf.append([
            bid, round(bt.input_count or 0, 0),
            togd, "Y" if in_h else "N",
            "Y" if is_placed else "N",
            round(hv, 0) if hv else 0,
            round(standing.get(bid, 0.0), 0),
            status, round(at_risk, 0) if at_risk else "",
            round(planned_tog, 0) if planned_tog else "",
            round(realized_tog, 0) if realized_tog else "",
            round(fw_surv, 1) if fw_surv is not None else "",
            fw_flag,
            round(fwm, 0) if realized_tog > 0 else "",
            round(fwc, 0) if realized_tog > 0 else "",
            round(fw_resid, 0) if fw_resid is not None else "",
        ])

    pct = (100.0 * dropped_fish / in_horizon_input) if in_horizon_input > 0 else 0.0
    ws.append(["INPUT-FISH CONSERVATION AUDIT"])
    ws.append([f"Generated: {datetime.now().isoformat(timespec='seconds')}"])
    if dropped_fish > 0:
        ws.append([f"*** {dropped_batches} batch(es) DROPPED — {dropped_fish:,.0f} stocked "
                   f"fish ({pct:.1f}% of in-horizon input) never placed. NOT caught by "
                   f"TankContinuityAudit (a never-placed batch has no tank rows). ***"])
    else:
        ws.append(["OK — every in-horizon batch reached the realized facility (0 dropped fish)."])
    if over_produced:
        ws.append([f"*** {len(over_produced)} batch(es) OVER-PRODUCED (harvested + standing > "
                   f"stocked input — fish created): {', '.join(over_produced)} ***"])
    if fw_divergent:
        _d = ", ".join(f"{b} ({100 * v:+.0f}%)" for b, v in sorted(fw_divergent, key=lambda x: x[1]))
        ws.append([f"NOTE: {len(fw_divergent)} batch(es) reached seawater >"
                   f"{_FW_DIVERGENCE_THRESH * 100:.0f}% off the planned tran_og_count — FW "
                   f"survival calibration gap (NOT lost fish; realized count is conserved): {_d}"])
    # Closed FW mass-balance gate (audit I2): first_FW_count == realized_TranOG +
    # FW_mort + FW_cull (reconciles the projected FW phase; see note above).
    fw_bal_pct = (100.0 * fw_bal_residual / fw_bal_base) if fw_bal_base > 0 else 0.0
    fw_bal_abs_pct = (100.0 * fw_bal_abs / fw_bal_base) if fw_bal_base > 0 else 0.0
    if fw_unbalanced:
        _u = ", ".join(f"{b} ({p:+.1f}%)"
                       for b, _r, p in sorted(fw_unbalanced, key=lambda x: -abs(x[1])))
        ws.append([f"*** FW MASS-BALANCE BREACH: {len(fw_unbalanced)} batch(es) where the FW "
                   f"phase does not conserve (first_FW_count != realized_TranOG + FW_mort + "
                   f"FW_cull) beyond {_FW_BALANCE_THRESH * 100:.0f}%: {_u}. A fish leak or a FW "
                   f"mortality/cull-accounting error shifts smolts -> harvest tonnage with no "
                   f"other gate catching it. ***"])
    elif fw_bal_base > 0:
        ws.append([f"FW mass-balance OK — projected FW phase conserves (first_FW_count == "
                   f"realized_TranOG + FW_mort + FW_cull; facility net {fw_bal_residual:+,.0f} "
                   f"fish, {fw_bal_pct:+.2f}%; abs {fw_bal_abs_pct:.2f}%)."])
    ws.append([
        "Batch", "Input_Count (fish)", "TranOG_Date", "In_Horizon",
        "Placed", "Harvested (fish)", "Standing@Horizon (fish)",
        "Status", "Fish_At_Risk (fish)",
        "Planned_TranOG (fish)", "Realized_TranOG (fish)", "FW_Survival (%)", "FW_Flag",
        "FW_Mort (fish)", "FW_Cull (fish)", "FW_Bal_Residual (fish)",
    ])
    for r in rowbuf:
        ws.append(r)
    widths = {1: 8, 2: 17, 3: 13, 4: 11, 5: 8, 6: 16, 7: 20, 8: 28, 9: 18,
              10: 18, 11: 18, 12: 14, 13: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_reconciliation_report(
    wb,
    batch_locations,
    batch_week_states,
    harvest_events,
    tranog_events,
    initial_state,
    realized_biology=None,
    transfer_events=None,
    grade_events=None,
    sheet_name: str = "ReconciliationReport",
) -> None:
    """Per-(batch, week) count + biomass balance check (OG side).

    Formula: open - mortality - harvest + input = expected_close.
    (Cull is FW-side: applied before fish reach OG; the `input` count
    is already POST-cull, so cull doesn't enter this OG-side balance.
    Shown in the output for transparency only.)

    Growth + mortality use the RECORDED realized biology (the net
    growth-minus-mortality biomass and mortality count the daily walker
    actually applied, summed over the batch's tanks) -- the SAME ground
    truth as TankContinuityAudit -- so the COUNT close reconciles EXACTLY
    instead of drifting on a coarse weekly-SGR re-estimate. A coarse SGR
    estimate is used only as a fallback when no realized biology is given.

    The balance includes the NET transfer + grade biomass for the batch (intra-
    batch moves net to ~0, EXCEPT a 6N purge move-in, which debits the source at
    the week-open weight but credits 6N at the +4-day grown weight — a real
    biomass injection that rides on the transfer). Without that term the injection
    showed up as a spurious ~0.5% Biomass_Delta on grow-out batches feeding 6N;
    with it, BOTH the count and biomass balances reconcile to 0 (the per-batch
    realized-biology + transfer terms are exact). TankContinuityAudit remains the
    authoritative per-tank check.

    Open for first week = PR-hydrated initial count (in-flight batches)
    or 0 (incoming batches arrive via TranOG events).
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["RECONCILIATION REPORT"])
    ws.append([
        "Per-(batch, week) count + biomass balance. open - mortality - cull - "
        "harvest + input = expected_close. Mismatches above tolerance are flagged."
    ])
    ws.append([])
    ws.append([
        "Week", "Batch",
        "Open_Count", "Mortality_Count", "Cull_Count", "Harvest_Count",
        "Input_Count", "Expected_Close", "Actual_Close",
        "Count_Delta",
        "Open_Bio_kg", "Growth_kg", "Mort_kg", "Cull_kg",
        "Harvest_kg", "Input_kg",
        "Expected_Bio_kg", "Actual_Bio_kg", "Biomass_Delta_kg",
        "Flag",
    ])

    from collections import defaultdict

    # Initial counts/biomass from PR-hydrated state.
    pr_count: dict[str, float] = defaultdict(float)
    pr_biomass: dict[str, float] = defaultdict(float)
    if initial_state is not None:
        for tank in initial_state.tanks_by_id.values():
            if tank.batch_id:
                pr_count[tank.batch_id] += tank.count
                pr_biomass[tank.batch_id] += tank.biomass_kg

    # Per-(batch, week) aggregates from BatchLocations.
    loc_count: dict[tuple[str, str], float] = defaultdict(float)
    loc_biomass: dict[tuple[str, str], float] = defaultdict(float)
    # STARVE (6N production in-place purge) tanks neither grow nor take mortality;
    # track their count/biomass per (batch, week) so the reconciliation excludes
    # them from the growth + mortality expectation (else they read as drift).
    starve_count: dict[tuple[str, str], float] = defaultdict(float)
    starve_biomass: dict[tuple[str, str], float] = defaultdict(float)
    weeks_seen: set[str] = set()
    week_start_by_label: dict[str, object] = {}
    for r in batch_locations:
        loc_count[(r.batch_id, r.week_label)] += r.count
        loc_biomass[(r.batch_id, r.week_label)] += r.biomass_kg
        if getattr(r, "stage", "") == "STARVE":
            starve_count[(r.batch_id, r.week_label)] += r.count
            starve_biomass[(r.batch_id, r.week_label)] += r.biomass_kg
        weeks_seen.add(r.week_label)
        week_start_by_label[r.week_label] = r.week_start
    weeks = sorted(weeks_seen)

    # Per-(batch, week) mortality % + cull count + cull biomass + SGR.
    mort_pct: dict[tuple[str, str], float] = {}
    sgr_pct_day: dict[tuple[str, str], float] = {}
    cull_count: dict[tuple[str, str], float] = {}
    cull_biomass: dict[tuple[str, str], float] = {}
    for s in (batch_week_states or []):
        mort_pct[(s.batch_id, s.week_label)] = s.mortality_pct_weekly
        sgr_pct_day[(s.batch_id, s.week_label)] = s.sgr_pct_day
        cull_count[(s.batch_id, s.week_label)] = s.cull_count_week or 0
        cull_biomass[(s.batch_id, s.week_label)] = s.cull_biomass_kg_week or 0

    # Recorded realized biology, aggregated over the batch's tanks:
    # (batch, week) -> net growth-minus-mortality biomass (kg) + mortality count.
    # Same source TankContinuityAudit uses, so both reconcile to the same truth.
    rbio_bw: dict[tuple[str, str], float] = defaultdict(float)
    rmort_bw: dict[tuple[str, str], float] = defaultdict(float)
    _have_rbio: set[tuple[str, str]] = set()
    for (tid_, wk_, b_), v in (realized_biology or {}).items():
        rbio_bw[(b_, wk_)] += v[0]
        rmort_bw[(b_, wk_)] += v[1]
        _have_rbio.add((b_, wk_))

    # Per-(batch, week) harvest count + biomass from events.
    from .time_grid import iso_week_label
    harv_count: dict[tuple[str, str], float] = defaultdict(float)
    harv_biomass: dict[tuple[str, str], float] = defaultdict(float)
    for ev in (harvest_events or []):
        wk = iso_week_label(ev.event_date)
        harv_count[(ev.batch_id, wk)] += ev.count
        harv_biomass[(ev.batch_id, wk)] += ev.count * ev.avg_wt_g / 1000.0

    # Per-(batch, week) input count + biomass from TranOG events.
    tin_count: dict[tuple[str, str], float] = defaultdict(float)
    tin_biomass: dict[tuple[str, str], float] = defaultdict(float)
    for ev in (tranog_events or []):
        wk = iso_week_label(ev.event_date)
        total_c = sum(d.count for d in ev.destinations)
        total_b = sum(d.count * d.avg_wt_g / 1000.0 for d in ev.destinations)
        tin_count[(ev.batch_id, wk)] += total_c
        tin_biomass[(ev.batch_id, wk)] += total_b

    # Per-(batch, week) NET transfer + grade biomass (kg). Transfers/grades are
    # intra-batch, so in == out (net 0) for an ordinary move; the exception is a
    # 6N purge move-in, which debits the source at the week-open weight but
    # credits 6N at the +4-day grown weight, leaving a real biomass injection.
    # Mirrors TankContinuityAudit's per-tank transfer/grade kg (source_avg_wt_g /
    # pickup_source_avg_wt_g), summed to the batch. (TranOG is NOT included here —
    # it's already the `input` term above.)
    xfer_net_kg: dict[tuple[str, str], float] = defaultdict(float)
    for ev in (transfer_events or []):
        wk = iso_week_label(ev.event_date)
        if hasattr(ev, "pickup_tank_id"):                 # GradedHarvest
            pk_src = getattr(ev, "pickup_source_avg_wt_g", None) or ev.pickup_avg_wt_g
            out_kg = (ev.pickup_count * pk_src + ev.retention_count * ev.retention_avg_wt_g) / 1000.0
            in_kg = (ev.pickup_count * ev.pickup_avg_wt_g
                     + ev.retention_count * ev.retention_avg_wt_g) / 1000.0
            xfer_net_kg[(ev.batch_id, wk)] += in_kg - out_kg
            continue
        ct = getattr(ev, "count_transferred", None)
        if ct is None or ct <= 0:
            continue
        total_planned = sum(d.count for d in ev.destinations) or 1.0
        avg_wt_g = sum(d.count * d.avg_wt_g for d in ev.destinations) / total_planned
        src_wt = getattr(ev, "source_avg_wt_g", None)
        out_wt = src_wt if src_wt is not None else avg_wt_g
        xfer_net_kg[(ev.batch_id, wk)] += ct * (avg_wt_g - out_wt) / 1000.0
    for ev in (grade_events or []):
        wk = iso_week_label(ev.event_date)
        total_dest = sum(d.count for d in ev.destinations)
        avg_wt_g = (sum(d.count * d.avg_wt_g for d in ev.destinations) / total_dest
                    if total_dest > 0 else 0.0)
        in_kg = sum(d.count * d.avg_wt_g for d in ev.destinations) / 1000.0
        xfer_net_kg[(ev.batch_id, wk)] += in_kg - total_dest * avg_wt_g / 1000.0

    # Walk batches × weeks, write rows + flag mismatches.
    TOLERANCE = 100.0   # fish; below this absolute, treat as numerical noise
    all_batches = sorted({b for (b, _) in loc_count} | set(pr_count.keys()))
    for batch in all_batches:
        prev_count = pr_count.get(batch, 0.0)
        prev_biomass = pr_biomass.get(batch, 0.0)
        for wk in weeks:
            actual_c = loc_count.get((batch, wk), 0.0)
            actual_b = loc_biomass.get((batch, wk), 0.0)
            if prev_count == 0 and actual_c == 0:
                continue
            m_pct = mort_pct.get((batch, wk), 0.0) or 0.0
            # STARVE fish this week neither grow nor take mortality.
            st_b = starve_biomass.get((batch, wk), 0.0)
            st_c = starve_count.get((batch, wk), 0.0)
            cull = cull_count.get((batch, wk), 0.0) or 0.0
            cull_b = cull_biomass.get((batch, wk), 0.0) or 0.0
            hv_c = harv_count.get((batch, wk), 0.0)
            hv_b = harv_biomass.get((batch, wk), 0.0)
            in_c = tin_count.get((batch, wk), 0.0)
            in_b = tin_biomass.get((batch, wk), 0.0)
            # Biomass after the pre-biology events (harvest out, TranOG in).
            # TranOG entries land on the OG-entry WEEK START, so the entered fish
            # are present the full week and take full-week growth + mortality.
            bio_full_growth = prev_biomass - hv_b + in_b
            if (batch, wk) in _have_rbio:
                # GROUND TRUTH: realized growth-minus-mortality biomass + mortality
                # count the daily walker actually applied (summed over the batch's
                # tanks) -- so the close reconciles exactly. Growth/Mort display
                # columns are split via the recorded mort count at the open weight
                # (growth = net + mort), matching TankContinuityAudit.
                mort = rmort_bw[(batch, wk)]
                rb = rbio_bw[(batch, wk)]
                open_wt_g = (prev_biomass / prev_count * 1000.0) if prev_count > 0 else 0.0
                mort_kg = mort * open_wt_g / 1000.0
                growth_kg = rb + mort_kg
                expected_b = bio_full_growth + rb
            else:
                # Fallback (no recorded biology): coarse weekly-SGR estimate on the
                # at-week-open biomass; only the NON-starve biomass grows / dies.
                mort = max(0.0, prev_count + in_c - st_c) * (m_pct / 100.0)
                sgr = sgr_pct_day.get((batch, wk), 0.0)
                growth_factor = (1.0 + sgr / 100.0) ** 7
                grow_bio = max(0.0, bio_full_growth - st_b)
                growth_kg = grow_bio * (growth_factor - 1.0)
                mort_kg = grow_bio * (m_pct / 100.0)
                expected_b = bio_full_growth + growth_kg - mort_kg
            # Net transfer + grade biomass (intra-batch ~0 except the 6N move-in
            # grown-weight injection); without this the injection reads as drift.
            expected_b += xfer_net_kg.get((batch, wk), 0.0)
            # OG-side balance: cull is FW-side (input is already post-cull),
            # so cull is shown as informational but not subtracted.
            expected_c = prev_count - mort - hv_c + in_c
            delta_c = actual_c - expected_c
            delta_b = actual_b - expected_b
            flag = ""
            if abs(delta_c) > TOLERANCE and abs(delta_c) > 0.005 * max(actual_c, prev_count, 1):
                flag = "COUNT_DRIFT"
            elif abs(delta_b) > 1000 and abs(delta_b) > 0.02 * max(actual_b, prev_biomass, 1):
                flag = "BIO_DRIFT"
            ws.append([
                wk, batch,
                round(prev_count, 0),
                round(mort, 0) if mort > 0 else None,
                round(cull, 0) if cull > 0 else None,
                round(hv_c, 0) if hv_c > 0 else None,
                round(in_c, 0) if in_c > 0 else None,
                round(expected_c, 0),
                round(actual_c, 0),
                round(delta_c, 0),
                round(prev_biomass, 0),
                round(growth_kg, 0) if growth_kg > 0 else None,
                round(mort_kg, 0) if mort_kg > 0 else None,
                round(cull_b, 0) if cull_b > 0 else None,
                round(hv_b, 0) if hv_b > 0 else None,
                round(in_b, 0) if in_b > 0 else None,
                round(expected_b, 0),
                round(actual_b, 0),
                round(delta_b, 0),
                flag,
            ])
            prev_count = actual_c
            prev_biomass = actual_b

    widths = {1: 11, 2: 7, 3: 11, 4: 13, 5: 11, 6: 13, 7: 11, 8: 14,
              9: 13, 10: 12,
              11: 12, 12: 11, 13: 10, 14: 9, 15: 12, 16: 11,
              17: 16, 18: 14, 19: 16, 20: 11}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_tank_continuity_audit(
    wb,
    batch_locations,
    batch_week_states,
    harvest_events,
    transfer_events,
    grade_events,
    tranog_events,
    initial_state,
    realized_biology=None,
    sheet_name: str = "TankContinuityAudit",
) -> None:
    """Per-(tank, week, batch) reconciliation.

    Formula: open - mortality - harvest_out - transfer_out + transfer_in
             - grade_out + grade_in + tranog_in = expected_close
    Compares to BatchLocations actual close. Flags any drift > tolerance.

    Open for first week = PR-hydrated initial tank count (in-flight).
    Batch transitions in a tank: when tank batch changes week-over-week,
    each batch gets its own row showing its arrival/departure path.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["TANK CONTINUITY AUDIT"])
    ws.append([
        "Per-(tank, week) count balance: every count change is accounted for "
        "by an event (mortality, harvest, transfer, grade, TranOG)."
    ])
    ws.append([])
    ws.append([
        # Count balance
        "Week", "Tank", "Batch", "Open_Count", "Mortality",
        "Harvest_Out", "Transfer_Out", "Transfer_In",
        "Grade_Out", "Grade_In", "TranOG_In",
        "Expected_Close", "Actual_Close", "Delta", "Flag",
        # Biomass balance
        "Open_Bio_kg", "Growth_kg", "Mort_kg",
        "Harvest_Out_kg", "Transfer_Out_kg", "Transfer_In_kg",
        "Grade_Out_kg", "Grade_In_kg", "TranOG_In_kg",
        "Expected_Close_kg", "Actual_Close_kg", "Delta_kg", "Bio_Flag",
    ])

    from collections import defaultdict
    from .time_grid import iso_week_label

    # Per-(tank, week) state from BatchLocations.
    tank_wk_state: dict[tuple[int, str], tuple[str, float]] = {}
    weeks_seen: set[str] = set()
    for r in batch_locations:
        tank_wk_state[(r.tank_id, r.week_label)] = (r.batch_id, r.count)
        weeks_seen.add(r.week_label)
    weeks = sorted(weeks_seen)
    all_tanks = sorted({t for (t, _) in tank_wk_state})

    # PR-hydrated initial tank state.
    pr_tank: dict[int, tuple[str | None, float]] = {}
    if initial_state is not None:
        for tid, tank in initial_state.tanks_by_id.items():
            if not tank.is_empty:
                pr_tank[tid] = (tank.batch_id, tank.count)

    # Per-batch per-week mortality % + SGR (from BatchWeekState). Used only as a
    # FALLBACK growth estimate when realized biology isn't available.
    mort_pct: dict[tuple[str, str], float] = {}
    sgr_pct_day: dict[tuple[str, str], float] = {}
    for s in (batch_week_states or []):
        mort_pct[(s.batch_id, s.week_label)] = s.mortality_pct_weekly
        sgr_pct_day[(s.batch_id, s.week_label)] = s.sgr_pct_day

    # Realized biology biomass delta + mortality count per (tank, week) — the
    # GROUND TRUTH growth-minus-mortality the daily walker actually applied
    # (summed across any batches that shared the tank that week). The audit
    # reconciles against this instead of re-estimating growth from the coarse
    # weekly SGR, so split-off sub-populations don't false-positive a BIO_DRIFT.
    rbio_tw: dict[tuple[int, str], float] = defaultdict(float)
    rmort_tw: dict[tuple[int, str], float] = defaultdict(float)
    for (tid_, wk_, _b), v in (realized_biology or {}).items():
        rbio_tw[(tid_, wk_)] += v[0]
        rmort_tw[(tid_, wk_)] += v[1]
    # When the daily walker's realized biology is available (the main run), the
    # count balance reconciles mortality against its RECORDED value (ground truth)
    # rather than a re-modelled estimate. Absent it (e.g. the LNS drift_count gate,
    # which passes realized_biology=None) the original modelled path runs unchanged.
    _have_realized = realized_biology is not None

    # Per-(tank, week) event aggregates — counts AND biomass (kg).
    harvest_out: dict[tuple[int, str], float] = defaultdict(float)
    harvest_out_kg: dict[tuple[int, str], float] = defaultdict(float)
    for ev in (harvest_events or []):
        wk = iso_week_label(ev.event_date)
        harvest_out[(ev.source_tank_id, wk)] += ev.count
        harvest_out_kg[(ev.source_tank_id, wk)] += ev.count * ev.avg_wt_g / 1000.0

    transfer_out: dict[tuple[int, str], float] = defaultdict(float)
    transfer_out_kg: dict[tuple[int, str], float] = defaultdict(float)
    transfer_in: dict[tuple[int, str], float] = defaultdict(float)
    transfer_in_kg: dict[tuple[int, str], float] = defaultdict(float)
    for ev in (transfer_events or []):
        # GradedHarvest (Event 5) rides in transfer_events but has a different
        # shape (1 source -> pickup + retention, no .destinations). Account its
        # moves explicitly so the audit conserves if it ever fires — otherwise it
        # is invisible (no count_transferred) and the source/pickup/retention
        # tank-weeks silently mis-reconcile.
        if hasattr(ev, "pickup_tank_id"):
            wk = iso_week_label(ev.event_date)
            # Debit the SOURCE at the pickup's PRE-GROWTH weight (when the pickup
            # was grown a few SW days for the 6N transfer, pickup_avg_wt_g is the
            # grown weight but the source held it lighter). The grown weight is
            # credited to the (frozen) pickup tank, so the few-day growth shows
            # as injected biomass there, not an over-debit on the source — same
            # accounting as a Transfer's source_avg_wt_g.
            pk_src_wt = getattr(ev, "pickup_source_avg_wt_g", None)
            if pk_src_wt is None:
                pk_src_wt = ev.pickup_avg_wt_g
            transfer_out[(ev.source_tank_id, wk)] += ev.pickup_count + ev.retention_count
            transfer_out_kg[(ev.source_tank_id, wk)] += (
                ev.pickup_count * pk_src_wt
                + ev.retention_count * ev.retention_avg_wt_g) / 1000.0
            transfer_in[(ev.pickup_tank_id, wk)] += ev.pickup_count
            transfer_in_kg[(ev.pickup_tank_id, wk)] += ev.pickup_count * ev.pickup_avg_wt_g / 1000.0
            transfer_in[(ev.retention_tank_id, wk)] += ev.retention_count
            transfer_in_kg[(ev.retention_tank_id, wk)] += ev.retention_count * ev.retention_avg_wt_g / 1000.0
            continue
        ct = getattr(ev, "count_transferred", None)
        if ct is None or ct <= 0:
            continue   # rejected transfers don't count
        wk = iso_week_label(ev.event_date)
        # Split count_transferred proportionally across destinations.
        total_planned = sum(d.count for d in ev.destinations) or 1.0
        # Weighted-average DEST avg_wt (~source avg_wt for a normal transfer).
        avg_wt_g = sum(d.count * d.avg_wt_g for d in ev.destinations) / total_planned
        # 6N purge move-ins place the GROWN (mid-week) weight on the destination
        # but drain the source by count at its WEEK-OPEN weight; debit the source
        # at that source weight so it balances (the 4-day growth then shows as
        # real injected biomass on the frozen 6N tank, not a source over-debit).
        src_wt_g = getattr(ev, "source_avg_wt_g", None)
        out_wt_g = src_wt_g if src_wt_g is not None else avg_wt_g
        transfer_out[(ev.source_tank_id, wk)] += ct
        transfer_out_kg[(ev.source_tank_id, wk)] += ct * out_wt_g / 1000.0
        for d in ev.destinations:
            share = d.count / total_planned
            transfer_in[(d.tank_id, wk)] += ct * share
            transfer_in_kg[(d.tank_id, wk)] += ct * share * d.avg_wt_g / 1000.0

    grade_out: dict[tuple[int, str], float] = defaultdict(float)
    grade_out_kg: dict[tuple[int, str], float] = defaultdict(float)
    grade_in: dict[tuple[int, str], float] = defaultdict(float)
    grade_in_kg: dict[tuple[int, str], float] = defaultdict(float)
    for ev in (grade_events or []):
        wk = iso_week_label(ev.event_date)
        total_dest = sum(d.count for d in ev.destinations)
        avg_wt_g = (sum(d.count * d.avg_wt_g for d in ev.destinations) / total_dest
                    if total_dest > 0 else 0.0)
        for src_tid in ev.source_tank_ids:
            grade_out[(src_tid, wk)] += total_dest / len(ev.source_tank_ids)
            grade_out_kg[(src_tid, wk)] += (
                total_dest / len(ev.source_tank_ids) * avg_wt_g / 1000.0
            )
        for d in ev.destinations:
            grade_in[(d.tank_id, wk)] += d.count
            grade_in_kg[(d.tank_id, wk)] += d.count * d.avg_wt_g / 1000.0

    tranog_in: dict[tuple[int, str], float] = defaultdict(float)
    tranog_in_kg: dict[tuple[int, str], float] = defaultdict(float)
    for ev in (tranog_events or []):
        wk = iso_week_label(ev.event_date)
        for d in ev.destinations:
            tranog_in[(d.tank_id, wk)] += d.count
            tranog_in_kg[(d.tank_id, wk)] += d.count * d.avg_wt_g / 1000.0

    # Per-(tank, week) biomass from BatchLocations.
    tank_wk_bio: dict[tuple[int, str], float] = {}
    # STARVE (in-place purge) tank-weeks: no growth, no mortality.
    tank_wk_starve: dict[tuple[int, str], bool] = {}
    for r in batch_locations:
        tank_wk_bio[(r.tank_id, r.week_label)] = r.biomass_kg
        if getattr(r, "stage", "") == "STARVE":
            tank_wk_starve[(r.tank_id, r.week_label)] = True

    pr_tank_bio: dict[int, float] = {}
    if initial_state is not None:
        for tid, tank in initial_state.tanks_by_id.items():
            if not tank.is_empty:
                pr_tank_bio[tid] = tank.biomass_kg

    TOLERANCE = 50.0      # fish
    BIO_TOLERANCE = 500.0  # kg

    # Facility-level conservation accumulators. Per-row tolerances pass small
    # same-sign drifts; summing every tank-week delta catches a DISTRIBUTED leak
    # (many tanks each under tolerance) that the per-row flags miss — the audit-
    # blind-spot class that hid the dropped batches. Count must cancel to ~0
    # (fish conserved); biomass carries a known systematic + bias from the
    # weekly-vs-daily growth approximation (reported, not asserted).
    _fac_dc_signed = _fac_dc_abs = 0.0
    _fac_db_signed = _fac_db_abs = 0.0
    for tid in all_tanks:
        prev_batch, prev_count = pr_tank.get(tid, (None, 0.0))
        prev_biomass = pr_tank_bio.get(tid, 0.0)
        for wk in weeks:
            cur = tank_wk_state.get((tid, wk))
            cur_batch, cur_count = cur if cur else (None, 0.0)
            cur_biomass = tank_wk_bio.get((tid, wk), 0.0)
            if prev_count == 0 and cur_count == 0:
                continue
            # STARVE (in-place purge) this week: no growth, no mortality.
            is_starve = tank_wk_starve.get((tid, wk), False)

            # ---- Count balance ----
            # The batch driving this week's biology: the current occupant,
            # or the prior one if the tank emptied. TranOG entries land on
            # the OG-entry WEEK START, so entered fish are present the full
            # week and take full-week mortality + growth — modelled exactly
            # like fish present at week-open (and like pre-biology transfers).
            bio_batch = cur_batch or prev_batch
            m_pct = mort_pct.get((bio_batch, wk), 0.0) if bio_batch else 0.0
            tn_in = tranog_in.get((tid, wk), 0.0)
            # Mortality. With realized biology available, use the RECORDED mortality
            # the daily walker actually applied — ground truth on the true (post
            # pre-biology-transfer) population. This fixes two opposite-signed audit
            # drifts the old model produced: (+) full-week mortality over-charged on
            # fish that departed mid-week pre-biology, and (-) STARVE (6N depuration)
            # mortality zeroed while the biology really killed those fish. Both
            # netted only by coincidence, so any 6N move broke the cancellation.
            # Without realized biology (LNS drift_count gate) keep the old modelled
            # estimate verbatim so that path stays byte-identical.
            if _have_realized:
                mort = rmort_tw.get((tid, wk), 0.0)
            else:
                mort = 0.0 if is_starve else (prev_count + tn_in) * (m_pct / 100.0)
            h_out = harvest_out.get((tid, wk), 0.0)
            t_out = transfer_out.get((tid, wk), 0.0)
            t_in = transfer_in.get((tid, wk), 0.0)
            g_out = grade_out.get((tid, wk), 0.0)
            g_in = grade_in.get((tid, wk), 0.0)
            expected_count = prev_count - mort - h_out - t_out + t_in - g_out + g_in + tn_in
            delta_count = cur_count - expected_count
            flag = ""
            if abs(delta_count) > TOLERANCE and abs(delta_count) > 0.005 * max(cur_count, prev_count, 1):
                flag = "TANK_DRIFT"

            # ---- Biomass balance ----
            # Phase D event order per week:
            #   1) 6N purge harvests + Layer 2 harvest demands (pre-biology)
            #   2) Migration transfers (pre-biology)
            #   3) Day-by-day biology: mortality + growth + TranOG entries
            #   4) Density-trigger Grade events (post-biology)
            h_out_kg = harvest_out_kg.get((tid, wk), 0.0)
            t_out_kg = transfer_out_kg.get((tid, wk), 0.0)
            t_in_kg = transfer_in_kg.get((tid, wk), 0.0)
            g_out_kg = grade_out_kg.get((tid, wk), 0.0)
            g_in_kg = grade_in_kg.get((tid, wk), 0.0)
            tn_in_kg = tranog_in_kg.get((tid, wk), 0.0)
            # Biomass after the pre-biology events (harvest/transfer/TranOG).
            bio_full_growth = prev_biomass - h_out_kg - t_out_kg + t_in_kg + tn_in_kg
            rb = rbio_tw.get((tid, wk))
            if rb is not None and not is_starve:
                # GROUND TRUTH: the net growth-minus-mortality biomass the daily
                # walker actually applied to this tank-week. Split into Growth/Mort
                # display columns via the recorded mortality count at the open
                # weight (growth = net + mort, so growth - mort == the recorded
                # net and the close reconciles exactly).
                open_wt_g = (prev_biomass / prev_count * 1000.0) if prev_count > 0 else 0.0
                mort_kg = rmort_tw.get((tid, wk), 0.0) * open_wt_g / 1000.0
                growth_kg = rb + mort_kg
                bio_after_biology = bio_full_growth + rb
            else:
                # Fallback (STARVE = no biology, or no recorded biology): coarse
                # weekly-SGR estimate on the at-week-open biomass.
                sgr = sgr_pct_day.get((bio_batch, wk), 0.0) if bio_batch else 0.0
                growth_factor = (1.0 + sgr / 100.0) ** 7
                growth_kg = 0.0 if is_starve else bio_full_growth * (growth_factor - 1.0)
                mort_kg = 0.0 if is_starve else bio_full_growth * (m_pct / 100.0)
                bio_after_biology = bio_full_growth + growth_kg - mort_kg
            expected_bio = bio_after_biology - g_out_kg + g_in_kg
            delta_bio = cur_biomass - expected_bio
            bio_flag = ""
            if abs(delta_bio) > BIO_TOLERANCE and abs(delta_bio) > 0.01 * max(cur_biomass, prev_biomass, 1):
                bio_flag = "BIO_DRIFT"
            _fac_dc_signed += delta_count
            _fac_dc_abs += abs(delta_count)
            _fac_db_signed += delta_bio
            _fac_db_abs += abs(delta_bio)

            display_batch = cur_batch or (prev_batch if prev_count > 0 else "")
            ws.append([
                wk, tid, display_batch,
                round(prev_count, 0),
                round(mort, 0) if mort > 0 else None,
                round(h_out, 0) if h_out > 0 else None,
                round(t_out, 0) if t_out > 0 else None,
                round(t_in, 0) if t_in > 0 else None,
                round(g_out, 0) if g_out > 0 else None,
                round(g_in, 0) if g_in > 0 else None,
                round(tn_in, 0) if tn_in > 0 else None,
                round(expected_count, 0),
                round(cur_count, 0),
                round(delta_count, 0),
                flag,
                round(prev_biomass, 0),
                round(growth_kg, 0) if growth_kg > 0 else None,
                round(mort_kg, 0) if mort_kg > 0 else None,
                round(h_out_kg, 0) if h_out_kg > 0 else None,
                round(t_out_kg, 0) if t_out_kg > 0 else None,
                round(t_in_kg, 0) if t_in_kg > 0 else None,
                round(g_out_kg, 0) if g_out_kg > 0 else None,
                round(g_in_kg, 0) if g_in_kg > 0 else None,
                round(tn_in_kg, 0) if tn_in_kg > 0 else None,
                round(expected_bio, 0),
                round(cur_biomass, 0),
                round(delta_bio, 0),
                bio_flag,
            ])
            prev_batch = cur_batch
            prev_count = cur_count
            prev_biomass = cur_biomass

    # Facility-level conservation summary (catches DISTRIBUTED drift the per-row
    # tolerances miss). Count must cancel to ~0; the |signed|/|abs| ratio is the
    # scale-free leak gauge — near 0 = random/cancelling (conserved), near 1 =
    # systematic one-way loss. Biomass carries a known + bias (weekly-vs-daily
    # growth approximation), surfaced here as a caveat, not a defect.
    _dc_ratio = (_fac_dc_signed / _fac_dc_abs) if _fac_dc_abs else 0.0
    _db_ratio = (_fac_db_signed / _fac_db_abs) if _fac_db_abs else 0.0
    ws.append([])
    ws.append(["FACILITY CONSERVATION SUMMARY (sum over all tank-weeks)"])
    ws.append(["Metric", "Signed_Sum", "Abs_Sum", "Signed/Abs_ratio", "Note"])
    ws.append(["Count (fish)", round(_fac_dc_signed, 0), round(_fac_dc_abs, 0),
               round(_dc_ratio, 4),
               "must cancel to ~0 (|ratio|<0.3); near 1 = distributed fish loss"])
    ws.append(["Biomass (kg)", round(_fac_db_signed, 0), round(_fac_db_abs, 0),
               round(_db_ratio, 4),
               "known + bias from weekly-vs-daily growth approximation (not a leak)"])

    widths = {1: 11, 2: 6, 3: 7, 4: 11, 5: 10, 6: 11, 7: 11, 8: 11,
              9: 9, 10: 9, 11: 10, 12: 14, 13: 12, 14: 9, 15: 11,
              16: 12, 17: 11, 18: 11, 19: 13, 20: 14, 21: 13,
              22: 11, 23: 11, 24: 11, 25: 16, 26: 14, 27: 11, 28: 9}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def write_facility_map(
    wb,
    batch_locations,
    facility,
    batches=None,
    tables=None,
    biology_states_by_batch=None,
    sheet_name: str = "FacilityMap",
) -> None:
    """Tank × Week matrix showing which batch occupies each tank each week.

    Cell value is the batch_id (or blank if empty). Rows are tanks
    ordered by system then tank_id; columns are forecast weeks in
    chronological order.

    Below the tank grid: two per-SYSTEM × week summaries — total planned feed
    (kg/day) and total biomass (kg) per system — each with a FACILITY total row
    so the operator can read system loads and check them against the caps.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["FACILITY MAP"])
    ws.append(["Each cell: Batch# AvgWt(kg) / Density(kg/m³). Color = batch."])

    # Order tanks by system + tank_id (OG tanks only for compactness).
    og_tanks = sorted(
        [t for t in facility.tanks if t.type == "OG"],
        key=lambda t: (t.system_id, t.tank_id),
    )
    # Weeks in chronological order + their start dates.
    weeks = sorted({r.week_label for r in batch_locations})
    wk_start: dict[str, object] = {}
    for r in batch_locations:
        wk_start.setdefault(r.week_label, r.week_start)

    # Build occupancy map (tank, week) → (batch_id, avg_wt_g, density).
    occ: dict[tuple[int, str], tuple] = {}
    for r in batch_locations:
        occ[(r.tank_id, r.week_label)] = (r.batch_id, r.avg_wt_g, r.density_kg_m3)

    # Two-row header: week labels then week-start dates.
    ws.append(["Week", ""] + weeks)
    ws.append(["Tank", "Sys"] + [wk_start.get(w) for w in weeks])

    for t in og_tanks:
        sys = t.system_id[2:] if t.system_id.startswith("OG") else t.system_id
        row = [t.tank_id, sys]
        for wk in weeks:
            cell = occ.get((t.tank_id, wk))
            if cell:
                bid, wt_g, dens = cell
                bnum = bid[1:] if bid and bid[:1] == "B" else bid
                row.append(f"{bnum} {wt_g / 1000.0:.1f}/{dens:.0f}")
            else:
                row.append("")
        ws.append(row)

    # ---- Per-system summaries below the tank grid ----
    from collections import defaultdict
    systems = sorted({t.system_id for t in og_tanks})
    sys_feed: dict[tuple[str, str], float] = defaultdict(float)
    sys_bio: dict[tuple[str, str], float] = defaultdict(float)
    for r in batch_locations:
        if r.system_id in systems:
            sys_feed[(r.system_id, r.week_label)] += _row_feed_kg_day(r, batches, tables)
            sys_bio[(r.system_id, r.week_label)] += r.biomass_kg

    # FW/EGG (hatchery) biomass per week — real facility biomass against the cap
    # (audit H2) but not in any OG system (FW fish live in FW tanks). Shown as its
    # own row and folded into the BIOMASS block's FACILITY total so that total is
    # FW-inclusive, matching the engine's cap basis and the Advisory.
    fw_bio: dict[str, float] = defaultdict(float)
    for states in (biology_states_by_batch or {}).values():
        for s in states:
            if s.stage in ("FW", "EGG"):
                fw_bio[s.week_label] += s.biomass_kg

    def _sys_label(sysid):
        return sysid[2:] if sysid.startswith("OG") else sysid

    def _block(title, data, fmt, extra=None):
        ws.append([])
        ws.append([title])
        ws.append(["System", ""] + weeks)
        for sysid in systems:
            ws.append([_sys_label(sysid), ""]
                      + [fmt(data.get((sysid, w), 0.0)) for w in weeks])
        if extra is not None:
            elabel, edata = extra
            ws.append([elabel, ""] + [fmt(edata.get(w, 0.0)) for w in weeks])
        ws.append(["FACILITY", ""]
                  + [fmt(sum(data.get((s, w), 0.0) for s in systems)
                         + (extra[1].get(w, 0.0) if extra else 0.0)) for w in weeks])

    _block("TOTAL PLANNED FEED PER DAY (kg/day) — per system (STARVE/6N-purge = 0)",
           sys_feed, lambda v: round(v, 0))
    _block("BIOMASS (kg) — per system per week; FW = hatchery, FACILITY = total vs cap",
           sys_bio, lambda v: round(v, 0), extra=("FW (hatchery)", fw_bio))

    ws.column_dimensions[get_column_letter(1)].width = 9
    ws.column_dimensions[get_column_letter(2)].width = 6
    for c in range(3, 3 + len(weeks)):
        ws.column_dimensions[get_column_letter(c)].width = 12


def write_advisory(
    wb,
    batch_locations,
    harvest_events,
    facility_limits,
    control,
    batches=None,
    tables=None,
    biology_states_by_batch=None,
    sheet_name: str = "Advisory",
) -> None:
    """Per-week capacity advisory + harvest recommendations (matches reference).

    Caps summary header, then one row per week: facility biomass + feed vs their
    (per-week resolved) limits with excess, that week's harvest, and an advisory
    flag (OK / REDUCE ...). Realized feed (kg/day) via realized_feed_kg_day so
    the totals match what fish in the tanks actually eat.
    """
    from collections import defaultdict
    from .caps import resolve_facility_cap, METRIC_BIOMASS, METRIC_FEED_DAY
    from .biology import realized_feed_kg_day

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["CAPACITY ADVISORY - HARVEST RECOMMENDATIONS"])
    ws.append(["Max Feed/Day:", f"{control.max_feed_per_day_kg:,.0f} kg/day"])
    ws.append(["Max Facility Biomass:", f"{control.max_biomass_kg:,.0f} kg"])
    ws.append([])
    ws.append([
        "Week", "Week_Start", "Total_Biomass (kg)", "Biomass_Limit (kg)",
        "Biomass_Excess (kg)", "Total_Feed (kg/day)", "Feed_Limit (kg/day)",
        "Feed_Excess (kg/day)", "Harvest_Count", "Harvest_Biomass (kg)",
        "Advisory", "Harvest_Batch", "Harvest_Recommended (kg)",
    ])

    bio: dict[str, float] = defaultdict(float)
    feed: dict[str, float] = defaultdict(float)
    wk_start: dict[str, object] = {}
    for r in batch_locations:
        bio[r.week_label] += r.biomass_kg
        wk_start.setdefault(r.week_label, r.week_start)
        # STARVE eats nothing (helper returns 0). NOTE: this is a per-DAY feed-rate
        # cap check, so it uses the steady realized rate only — the 6N move-in's
        # 4-day pre-transfer feed is a TOTAL-feed accounting item (in the
        # FeedForecast / ledger / YearlySummary totals), not a per-day rate.
        feed[r.week_label] += _row_feed_kg_day(r, batches, tables)
    # FW-INCLUSIVE biomass + feed (audit H2/M3): FW/EGG fish live in FW tanks
    # (absent from batch_locations) but are real facility biomass against the 3.8M
    # cap AND eat real hatchery feed against the feed/day cap. Add both so
    # Total_Biomass / Total_Feed / the excess + OK-vs-REDUCE flag report the TOTAL
    # facility figures — the same dual-limit basis the harvest engine now enforces —
    # instead of OG-only (which can show false headroom on either limit).
    for states in (biology_states_by_batch or {}).values():
        for s in states:
            if s.stage in ("FW", "EGG"):
                bio[s.week_label] += s.biomass_kg
                feed[s.week_label] += getattr(s, "feed_kg_day", 0.0)
                wk_start.setdefault(s.week_label, s.week_start)

    harv_c: dict[str, float] = defaultdict(float)
    harv_b: dict[str, float] = defaultdict(float)
    for ev in harvest_events:
        wk = iso_week_label(ev.event_date)
        harv_c[wk] += ev.count
        harv_b[wk] += ev.count * ev.avg_wt_g / 1000.0

    for wk in sorted(set(bio) | set(feed) | set(harv_c)):
        tb = bio.get(wk, 0.0)
        tf = feed.get(wk, 0.0)
        bcap = resolve_facility_cap(METRIC_BIOMASS, wk, facility_limits, control) or 0.0
        fcap = resolve_facility_cap(METRIC_FEED_DAY, wk, facility_limits, control) or 0.0
        bex = max(0.0, tb - bcap)
        fex = max(0.0, tf - fcap)
        if bex > 0 and fex > 0:
            adv = "REDUCE BIOMASS + FEED"
        elif bex > 0:
            adv = "REDUCE BIOMASS"
        elif fex > 0:
            adv = "REDUCE FEED"
        else:
            adv = "OK"
        ws.append([
            wk, wk_start.get(wk),
            round(tb, 0), round(bcap, 0), round(bex, 0),
            round(tf, 0), round(fcap, 0), round(fex, 0),
            round(harv_c.get(wk, 0.0), 0), round(harv_b.get(wk, 0.0), 0),
            adv, "", round(bex, 0) if bex > 0 else "",
        ])
    widths = {1: 10, 2: 12, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18,
              9: 14, 10: 18, 11: 22, 12: 14, 13: 22}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# Regex helpers used by both write_advisory categorization and ValidationLog
# message parsing.
_RE_WEEK = re.compile(r"\b(\d{4}-W\d{2})\b")
_RE_BATCH = re.compile(r"\bB\d{2,3}\b")
_RE_TANK = re.compile(r"\b(OG[1-6][NS])-(\d+)\b")


def _vlog_parse(msg: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Best-effort extraction of (week_label, batch_id, location_id) from
    a free-form warning string."""
    week = (_RE_WEEK.search(msg).group(1) if _RE_WEEK.search(msg) else None)
    batch = (_RE_BATCH.search(msg).group(0) if _RE_BATCH.search(msg) else None)
    tm = _RE_TANK.search(msg)
    loc = tm.group(0) if tm else None
    return week, batch, loc


def write_validation_log(
    wb,
    residuals=None,
    placement_warnings=None,
    scheduler_warnings=None,
    bottlenecks=None,
    density_violations=None,
    invariant_warnings=None,
    placed_batches=None,
    sheet_name: str = "ValidationLog",
) -> None:
    """Numbered validation issue stream (matches reference format).

    Title rows + a "# | Category | Detail" table, one row per
    warning/diagnostic from every layer.
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    entries: list[tuple[str, str]] = []
    for r in residuals or ():
        if abs(r.residual_pct) >= 0.5:
            sug = (f"; suggested FW_Correction={r.suggested_fw_correction:.3f}"
                   if r.suggested_fw_correction is not None else "")
            entries.append((
                "WARNING - FW Calibration",
                f"Batch {r.batch_id} TranOG {r.tran_og_date.date()}: "
                f"residual {r.residual_pct:+.2f}%{sug}"))
    # Bottlenecks are PRECALC static-plan predictions (supply vs demand on the
    # canvas, before placement runs). A TranOG bottleneck the precalc coordinator
    # could not seat may still be resolved downstream by the Phase-D make-room
    # harvest. Annotate each with its ACTUAL outcome (placed vs dropped) so the
    # log doesn't cry wolf about a problem the executor already handled.
    _re_tranog_b = re.compile(r"TranOG\s+(B\d+)")
    for b in bottlenecks or ():
        detail = f"{b.week_label}: {b.detail}"
        if placed_batches is not None and "tranog" in b.kind:
            m = _re_tranog_b.search(b.detail or "")
            if m:
                detail += ("  [RESOLVED at placement — make-room harvest freed a "
                           "tank; batch placed, not dropped]"
                           if m.group(1) in placed_batches
                           else "  [NOT resolved — batch DROPPED; facility saturated]")
        entries.append((f"WARNING - Bottleneck/{b.kind}", detail))
    for w in scheduler_warnings or ():
        entries.append(("WARNING - Harvest Scheduler", w))
    for w in invariant_warnings or ():
        if "INV-1" in w:
            cat = "WARNING - INV-1 (one-batch-per-tank)"
        elif "INV-5" in w:
            cat = "WARNING - INV-5 (min_tank_control)"
        elif "Density" in w:
            cat = "WARNING - Hydration Density"
        else:
            cat = "WARNING - Hydration"
        entries.append((cat, w))
    for v in density_violations or ():
        wk, loc, bid, d, cap = v
        entries.append((
            "WARNING - Density",
            f"{wk}: {loc} (batch {bid}) at {d:.1f} kg/m³ > cap {cap:.1f}"))
    for w in placement_warnings or ():
        if "INV-4" in w:
            cat = "WARNING - INV-4 (1 kg rule)"
        elif "INV-5" in w:
            cat = "WARNING - INV-5 (min_tank_control)"
        elif "INV-1" in w:
            cat = "WARNING - INV-1 (one-batch-per-tank)"
        elif "TranOG" in w:
            cat = "WARNING - TranOG Entry"
        elif "6N" in w or "purge" in w.lower():
            cat = "WARNING - 6N Pipeline"
        elif w.startswith("[B]"):
            cat = "WARNING - Placement/Phase B"
        elif w.startswith("[C]"):
            cat = "WARNING - Placement/Phase C"
        elif w.startswith("[D]"):
            cat = "WARNING - Placement/Phase D"
        else:
            cat = "WARNING - Placement"
        entries.append((cat, w))

    ws.append(["VALIDATION LOG"])
    ws.append([f"Generated: {datetime.now().isoformat(timespec='seconds')}"])
    ws.append([f"{len(entries)} issue(s) found - review below" if entries
               else "No issues found"])
    ws.append(["#", "Category", "Detail"])
    for i, (cat, detail) in enumerate(entries, 1):
        ws.append([i, cat, detail])

    widths = {1: 6, 2: 34, 3: 120}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"


def write_forecast_start(wb, forecast_start, sheet_name: str = "Control") -> None:
    """Write the derived forecast start back into the Control INPUT cell.

    Mirrors the VBA DetectForecastStart(), which writes the derived start
    (= ProductionReport closing + 1 day) into Control B3 on every run so
    the input cell never drifts from the ProductionReport. This closes the
    gap between a stale, un-reconciled input sheet and a post-run sheet:
    after a run, `B3 == PR closing + 1` is an invariant that means
    "reconciled". The status-echo row (R11, written by
    write_control_status) is informational only — this writer keeps the
    actual INPUT cell honest.

    Finds the 'Forecast Start Date' label in column A (the same label
    read_control matches), falling back to the VBA's B3 convention.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    target_row = None
    for r in range(1, 30):
        a = ws.cell(row=r, column=1).value
        if (isinstance(a, str)
                and re.sub(r"\s+", " ", a.strip().lower()).rstrip(":")
                == "forecast start date"):
            target_row = r
            break
    if target_row is None:
        target_row = 3  # VBA convention: Control!B3
    ws.cell(row=target_row, column=2).value = forecast_start


def write_control_status(
    wb,
    *,
    status: str,
    scenario: str,
    forecast_start,
    horizon_weeks: int,
    batches: int,
    og_tanks: int,
    elapsed_s: float,
    warnings: int,
    sheet_name: str = "Control",
) -> None:
    """Overwrite Control R8-R16 column B with the run summary (DESIGN §1).

    Row layout (label in A, value in B) is fixed by the workbook
    template: R8 Last Run, R9 Timestamp, R10 Scenario, R11 Forecast
    Start, R12 Horizon, R13 Batches, R14 OG Tanks, R15 Elapsed,
    R16 Warnings.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    ws.cell(row=8, column=2).value = status
    ws.cell(row=9, column=2).value = datetime.now()
    ws.cell(row=10, column=2).value = scenario
    ws.cell(row=11, column=2).value = forecast_start
    ws.cell(row=12, column=2).value = f"{horizon_weeks} weeks"
    ws.cell(row=13, column=2).value = batches
    ws.cell(row=14, column=2).value = og_tanks
    ws.cell(row=15, column=2).value = f"{elapsed_s:.1f}s"
    ws.cell(row=16, column=2).value = warnings


def write_calibration_diagnostics(
    wb,
    residuals: Iterable[CalibrationResidual],
    sheet_name: str = "Diagnostics",
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["FW Calibration: projected pre-cull avg wt vs target at TranOG_Date, plus suggested FW_Correction"])
    ws.append([
        "Batch", "TranOG_Date", "Target_AvgWt_g",
        "Current_FW_Correction", "Projected_PreCull_AvgWt_g", "Residual_pct",
        "Suggested_FW_Correction",
    ])
    for r in residuals:
        ws.append([
            r.batch_id, r.tran_og_date, round(r.target_avg_wt_g, 2),
            round(r.current_fw_correction, 4),
            round(r.projected_pre_cull_avg_wt_g, 2), round(r.residual_pct, 2),
            None if r.suggested_fw_correction is None else round(r.suggested_fw_correction, 4),
        ])
    widths = {1: 8, 2: 12, 3: 16, 4: 22, 5: 26, 6: 12, 7: 24}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------- Workbook helpers ----------

def load_workbook(path: Path):
    return openpyxl.load_workbook(path, keep_vba=True, data_only=True)


def is_macro_enabled_workbook(src) -> bool:
    """True if the workbook (bytes or path) is macro-enabled — its content type is
    macroEnabled or it carries a vbaProject. The controller loads with keep_vba=True,
    so a macro-enabled INPUT yields a macro-enabled OUTPUT that MUST be saved as
    `.xlsm`: Excel refuses a macro-enabled workbook that wears a `.xlsx` extension
    (content-type / extension mismatch). Use this to pick the output extension."""
    import zipfile, io as _io
    try:
        srcobj = _io.BytesIO(src) if isinstance(src, (bytes, bytearray)) else str(src)
        with zipfile.ZipFile(srcobj) as z:
            if any("vbaProject" in n for n in z.namelist()):
                return True
            return "macroEnabled" in z.read("[Content_Types].xml").decode("utf-8", "replace")
    except Exception:
        return False


def iso_week_label(d: datetime) -> str:
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def write_system_limits_audit(
    wb,
    batch_locations,
    batch_by_id,
    tables,
    system_limits,
    control,
    sheet_name: str = "SystemLimitsAudit",
):
    """Per-(week, system) REALIZED biomass + feed vs the system caps.

    The engine checks tank density but never checked per-system biomass/feed
    against the SystemLimits caps — so over-cap systems went unsurfaced. This
    audit closes that gap, mirroring TankContinuityAudit. Feed is the realized
    tank feed (biomass × SGR × FCR at the tank's actual weight), NOT the raw
    projection (which over-counts un-harvested fish). Caps are carried forward
    past the data's last week and buffered by Control R29. OG6N (depuration)
    has no biomass/feed caps in purge mode, so it's reported but unflagged.

    Returns (n_biomass_over, n_feed_over, worst_biomass_ratio, worst_feed_ratio).
    """
    from collections import defaultdict
    from .biology import realized_feed_kg_day

    buf = 1.0 + (getattr(control, "global_buffer_pct", 0.0) or 0.0)

    # Carry-forward cap lookup per (system, metric).
    smw: dict = defaultdict(list)
    for (wk, sysid, metric), val in system_limits.caps.items():
        smw[(sysid, metric)].append((wk, val))
    for k in smw:
        smw[k].sort()

    def _cap(wk, sysid, metric):
        lst = smw.get((sysid, metric))
        if not lst:
            return None
        best = lst[0][1]
        for w, v in lst:
            if w <= wk:
                best = v
            else:
                break
        return best

    sb: dict = defaultdict(float)
    sf: dict = defaultdict(float)
    for r in batch_locations:
        if r.count <= 0:
            continue
        sb[(r.week_label, r.system_id)] += r.biomass_kg
        # STARVE = in-place purge: biomass counts to the system, but no feed
        # (helper returns 0). Per-DAY feed-rate cap check -> steady realized rate
        # only; the 6N move-in 4-day pre-transfer feed is a total-feed accounting
        # item (FeedForecast / ledger / YearlySummary), not a per-day rate.
        sf[(r.week_label, r.system_id)] += _row_feed_kg_day(r, batch_by_id, tables)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["SYSTEM LIMITS AUDIT"])
    ws.append(["Realized per-system biomass + feed vs caps "
               "(carry-forward, +R29 buffer). OG6N = depuration (no caps)."])
    ws.append([])
    ws.append(["Week", "System", "Biomass_kg", "Biomass_cap", "Bio_flag",
               "Feed_kg_day", "Feed_cap", "Feed_flag"])

    nb = nf = 0
    worst_b = worst_f = 0.0
    for (wk, sysid) in sorted(sb.keys()):
        bio = sb[(wk, sysid)]
        feed = sf[(wk, sysid)]
        bcap = _cap(wk, sysid, "biomass")
        fcap = _cap(wk, sysid, "feed_per_day")
        bflag = fflag = ""
        # OG6N is the depuration pool (purge mode): fish starve (no feed) and
        # the system is intentionally uncapped. Report its rows but never flag.
        flaggable = sysid != "OG6N"
        if bcap and flaggable:
            worst_b = max(worst_b, bio / bcap)
            if bio > bcap * buf:
                bflag = "BIOMASS_OVER"
                nb += 1
        if fcap and flaggable:
            worst_f = max(worst_f, feed / fcap)
            if feed > fcap * buf:
                fflag = "FEED_OVER"
                nf += 1
        ws.append([wk, sysid, round(bio, 0),
                   round(bcap, 0) if bcap else None, bflag,
                   round(feed, 1), round(fcap, 0) if fcap else None, fflag])

    widths = {1: 10, 2: 8, 3: 12, 4: 12, 5: 13, 6: 12, 7: 10, 8: 11}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    return nb, nf, worst_b, worst_f


def write_run_comparison(wb, records, *, pr_name="", generated=None,
                         sheet_name="RunComparison"):
    """Legible side-by-side comparison of several planning methods run on the
    SAME inputs (same manual override window + control rules; only the method
    differs). One column per method, one labeled row per metric.

    `records` is a list of per-method dicts (see tools/run_compare.py), each:
        key, label, family, blurb : method identity
        failed        : None, or an error string (run/scoring failed)
        elapsed       : wall seconds
        workbook      : filename of this method's full workbook (for drill-in)
        dropped, overprod : conservation gate (both 0 == PASS)
        metrics       : forecast.optimize.Metrics (or None if failed)
        harvest       : dict with n_weeks / min_week / max_week /
                        weeks_below_min / zero_weeks / min_harvest

    Winner per metric is highlighted GREEN, but ONLY among methods that PASS
    conservation (a method that loses fish is never called 'best'). There is no
    single overall winner — which method is best depends on the operator's
    objective (hold the cap vs flatten vs minimize feed/handling), so the sheet
    shows every dimension rather than collapsing to one score.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ---- palette ----
    C_TITLE = PatternFill("solid", fgColor="1F4E78")
    C_FAMILY = PatternFill("solid", fgColor="DDEBF7")
    C_SECT = PatternFill("solid", fgColor="D9E1F2")
    C_WIN = PatternFill("solid", fgColor="C6EFCE")
    C_PASS = PatternFill("solid", fgColor="C6EFCE")
    C_PART = PatternFill("solid", fgColor="FFEB9C")
    C_FAIL = PatternFill("solid", fgColor="FFC7CE")
    F_TITLE = Font(bold=True, color="FFFFFF", size=13)
    F_SUB = Font(italic=True, color="444444", size=9)
    F_HDR = Font(bold=True)
    F_SECT = Font(bold=True, color="1F4E78")
    F_PASS = Font(bold=True, color="006100")
    F_PART = Font(bold=True, color="9C6500")
    F_FAIL = Font(bold=True, color="9C0006")
    F_GREY = Font(color="B0B0B0")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    RIGHT = Alignment(horizontal="right")
    LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CENTER = Alignment(horizontal="center")

    n = len(records)
    first_m = 3                      # first method column (A=label, B=unit)
    last_m = first_m + n - 1
    last_col = get_column_letter(last_m)

    def _gate(rec):
        """PASS / PARTIAL / FAIL, using the driver's per-method verdict when
        present (it reads each method's OWN authoritative conservation proof —
        the batch-level ReconciliationReport for Global, the tank-by-tank audit
        for the Controller). Falls back to dropped/overprod for synthetic rows.
        PARTIAL = mass conserved but some stocked fish left unplaced in tanks."""
        g = rec.get("gate")
        if g:
            return g
        if rec.get("failed") is not None or rec.get("metrics") is None:
            return "FAIL"
        return ("PASS" if rec.get("dropped", 0) == 0
                and rec.get("overprod", 0) == 0 else "FAIL")

    def _ok(rec):
        # Winner-eligible = a COMPLETE, mass-conserving plan. A PARTIAL plan
        # (fish unplaced) or a FAIL is never crowned "best" on any metric.
        return rec.get("metrics") is not None and _gate(rec) == "PASS"

    # ---- title + subtitle ----
    gen = generated or datetime.now().isoformat(timespec="seconds")
    ws["A1"] = (f"RUN COMPARISON  —  {n} method(s), identical inputs "
                f"(same manual override window + control rules)")
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].fill = C_TITLE
    ws["A1"].font = F_TITLE
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws["A2"] = (f"Generated {gen}"
                + (f"  ·  PR: {pr_name}" if pr_name else "")
                + "   ·   Winner per metric highlighted GREEN among methods that "
                  "PASS conservation.  'Best overall' depends on your objective — "
                  "compare the dimensions below.")
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].font = F_SUB
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 28

    # ---- family header (row 4) + method header (row 5) ----
    r_fam, r_hdr = 4, 5
    ws.cell(r_fam, 1, "").fill = C_FAMILY
    # merge contiguous same-family runs
    i = 0
    while i < n:
        fam = records[i].get("family", "")
        j = i
        while j + 1 < n and records[j + 1].get("family", "") == fam:
            j += 1
        c0, c1 = first_m + i, first_m + j
        cell = ws.cell(r_fam, c0, fam)
        if c1 > c0:
            ws.merge_cells(start_row=r_fam, start_column=c0,
                           end_row=r_fam, end_column=c1)
        cell.fill = C_FAMILY
        cell.font = F_HDR
        cell.alignment = CENTER
        i = j + 1

    ws.cell(r_hdr, 1, "Metric").font = F_HDR
    ws.cell(r_hdr, 2, "Unit / better").font = F_HDR
    ws.cell(r_hdr, 2).alignment = LEFT
    for k, rec in enumerate(records):
        cell = ws.cell(r_hdr, first_m + k, rec.get("label", rec["key"]))
        cell.font = F_HDR
        cell.alignment = Alignment(wrap_text=True, horizontal="center",
                                   vertical="top")
        cell.border = BORDER

    # ---- conservation gate (row 6) — the hard gate, most prominent ----
    r = 6
    ws.cell(r, 1, "CONSERVATION GATE").font = Font(bold=True, size=11)
    ws.cell(r, 2, "PASS = conserves + all placed").font = F_SUB
    ws.cell(r, 2).alignment = LEFT
    for k, rec in enumerate(records):
        cell = ws.cell(r, first_m + k)
        cell.border = BORDER
        cell.alignment = CENTER
        if rec.get("failed") is not None:
            cell.value = "RUN FAILED"
            cell.fill = C_FAIL
            cell.font = F_FAIL
        else:
            g = _gate(rec)
            if g == "PASS":
                cell.value = "PASS"
                cell.fill = C_PASS
                cell.font = F_PASS
            elif g == "PARTIAL":
                ub = (rec.get("placement") or {}).get("unplaced_batches", 0)
                cell.value = f"PARTIAL — {ub} batch(es) unplaced"
                cell.fill = C_PART
                cell.font = F_PART
            else:
                d, o = rec.get("dropped", 0), rec.get("overprod", 0)
                cell.value = f"FAIL (drop {d}, over {o})"
                cell.fill = C_FAIL
                cell.font = F_FAIL

    # ---- metric rows ----
    def g_metric(attr):
        return lambda rec: (getattr(rec["metrics"], attr)
                            if rec.get("metrics") is not None else None)

    def g_pct_cap_peak(rec):
        m = rec.get("metrics")
        if m is None or not m.biomass_cap:
            return None
        return 100.0 * m.overall_peak_biomass / m.biomass_cap

    def g_util(rec):
        m = rec.get("metrics")
        if m is None or not m.biomass_cap:
            return None
        return 100.0 * m.overall_mean_biomass / m.biomass_cap

    def g_harv(field):
        return lambda rec: (rec.get("harvest") or {}).get(field)

    def g_place(field):
        return lambda rec: (rec.get("placement") or {}).get(field)

    def g_between(field):
        return lambda rec: (rec["metrics"].between_system.get(field)
                            if rec.get("metrics") is not None else None)

    def g_within(field):
        return lambda rec: (rec["metrics"].within_system.get(field)
                            if rec.get("metrics") is not None else None)

    f_kg = lambda v: f"{v:,.0f}"
    f_int = lambda v: f"{v:,.0f}"
    f_pct = lambda v: f"{v:.1f}%"
    f_cv = lambda v: f"{v:.3f}"
    f_ratio = lambda v: f"{v:.3f}"
    f_1 = lambda v: f"{v:,.1f}"
    f_sec = lambda v: f"{v:,.0f}s"

    # (kind, ...) — "sect": section header; "row": metric row
    # metric row = (label, unit_dir, direction, getter, fmt)
    #   direction: "low" lower-is-better, "high" higher-is-better, None neutral
    SPEC = [
        ("sect", "REALIZED PLACEMENT (are all stocked fish in tanks?)"),
        ("row", "Batches not placed in tanks", "count · ↓ · 0 = complete", "low", g_place("unplaced_batches"), f_int),
        ("row", "Fish not placed in tanks", "stocked fish · ↓", "low", g_place("unplaced_fish"), f_int),
        ("sect", "BIOMASS vs the facility cap"),
        ("row", "Peak facility biomass", "kg", None, g_metric("overall_peak_biomass"), f_kg),
        ("row", "Peak biomass vs cap", "% of cap · ≤100 · ↓", "low", g_pct_cap_peak, f_pct),
        ("row", "Mean biomass utilization", "% of cap · ↑ (use it)", "high", g_util, f_pct),
        ("sect", "HARVEST — every week, min→max (contract)"),
        ("row", "Harvest weeks", "weeks", None, g_harv("n_weeks"), f_int),
        ("row", "Min weekly harvest", "fish · ↑ (never starve)", "high", g_harv("min_week"), f_int),
        ("row", "Max weekly harvest", "fish", None, g_harv("max_week"), f_int),
        ("row", "Weeks below min-harvest", "weeks · ↓", "low", g_harv("weeks_below_min"), f_int),
        ("row", "Zero-harvest weeks", "weeks · must be 0", "low", g_harv("zero_weeks"), f_int),
        ("row", "Weeks over harvest cap", "weeks · ↓", "low", g_metric("weeks_over_harvest_cap"), f_int),
        ("sect", "CAP COMPLIANCE (per-system + per-tank)"),
        ("row", "System over-cap cells", "% of cells · ↓", "low",
         lambda r: (None if r.get("metrics") is None else 100.0 * r["metrics"].system_overshoot), f_pct),
        ("row", "Density over-cap cells", "% of cells · ↓", "low",
         lambda r: (None if r.get("metrics") is None else 100.0 * r["metrics"].density_overshoot), f_pct),
        ("row", "Hottest system load", "% of cap · ↓", "low",
         lambda r: (None if r.get("metrics") is None else 100.0 * r["metrics"].system_peak), f_pct),
        ("row", "Max per-tank density", "kg/m³ · ↓ · ≤95", "low", g_metric("density_peak"), f_kg),
        ("sect", "TANK USAGE (FW→OG footprint + moves; OG6N excluded)"),
        ("row", "Grow-out tanks used (mean)", "tanks · context", None, g_metric("tank_footprint_mean"), f_1),
        ("row", "Grow-out tanks used (peak)", "tanks · context", None, g_metric("tank_footprint_peak"), f_int),
        ("row", "Tanks per batch (mean)", "tanks/batch · ↓ (fewer moves)", "low", g_metric("batch_tank_path_mean"), f_1),
        ("row", "Tanks per batch (worst)", "tanks/batch · ↓", "low", g_metric("batch_tank_path_max"), f_int),
        ("sect", "STEADINESS (flatness — lower = smoother)"),
        ("row", "Biomass variability", "CV+swing · ↓", "low", g_metric("biomass_var"), f_cv),
        ("row", "Harvest variability", "CV · ↓", "low", g_metric("harvest_var"), f_cv),
        ("row", "Feed variability", "CV+swing · ↓", "low", g_metric("feed_var"), f_cv),
        ("sect", "BALANCE — between systems (even load across systems)"),
        ("row", "Between-system biomass CV", "CV · ↓ (even)", "low", g_between("bio_cv_mean"), f_cv),
        ("row", "Between-system feed CV", "CV · ↓ (even)", "low", g_between("feed_cv_mean"), f_cv),
        ("row", "Between-system biomass gap (peak)", "kg · ↓", "low", g_between("bio_range_peak"), f_kg),
        ("sect", "BALANCE — within systems (even load across tanks)"),
        ("row", "Within-system biomass CV", "CV · ↓ (even)", "low", g_within("bio_cv_mean"), f_cv),
        ("row", "Within-system feed CV", "CV · ↓ (even)", "low", g_within("feed_cv_mean"), f_cv),
        ("sect", "COST / HANDLING"),
        ("row", "Mean daily feed", "kg/day · ↓", "low", g_metric("feed_load"), f_kg),
        ("row", "Transfers per fish", "moves/fish · ↓", "low", g_metric("transfers_per_fish"), f_ratio),
        ("sect", "RUN"),
        ("row", "Wall time", "seconds", None, lambda r: r.get("elapsed"), f_sec),
    ]

    r = 8
    for item in SPEC:
        if item[0] == "sect":
            ws.cell(r, 1, item[1]).font = F_SECT
            ws.cell(r, 1).fill = C_SECT
            for c in range(2, last_m + 1):
                ws.cell(r, c).fill = C_SECT
            r += 1
            continue
        _, label, unit, direction, getter, fmt = item
        ws.cell(r, 1, label).alignment = LEFT
        ws.cell(r, 2, unit).font = F_SUB
        ws.cell(r, 2).alignment = LEFT
        # winners among conservation-OK methods
        vals = {}
        for k, rec in enumerate(records):
            if not _ok(rec):
                continue
            v = getter(rec)
            if isinstance(v, (int, float)):
                vals[k] = v
        winners = set()
        if direction and vals:
            best = min(vals.values()) if direction == "low" else max(vals.values())
            winners = {k for k, v in vals.items() if abs(v - best) < 1e-9}
        for k, rec in enumerate(records):
            cell = ws.cell(r, first_m + k)
            cell.border = BORDER
            cell.alignment = RIGHT
            v = getter(rec)
            if not isinstance(v, (int, float)):
                cell.value = "—"
                cell.font = F_GREY
            elif k in winners:
                cell.value = fmt(v)
                cell.fill = C_WIN
                cell.font = Font(bold=True, color="006100")
            elif not _ok(rec):
                # PARTIAL / FAIL plan: show the number but grey it — computed on
                # an incomplete or non-conserving plan, so not directly comparable.
                cell.value = fmt(v)
                cell.font = F_GREY
            else:
                cell.value = fmt(v)
        r += 1

    # ---- method legend (plain-language) ----
    r += 1
    ws.cell(r, 1, "METHODS").font = Font(bold=True, size=11)
    r += 1
    ws.cell(r, 1, "Key").font = F_HDR
    ws.cell(r, 2, "Method").font = F_HDR
    ws.cell(r, 3, "How it plans  ·  full workbook  ·  runtime").font = F_HDR
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=last_m)
    r += 1
    for rec in records:
        ws.cell(r, 1, rec["key"]).alignment = LEFT
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2, rec.get("label", "")).alignment = LEFT
        detail = rec.get("blurb", "")
        if rec.get("failed") is not None:
            detail += f"   [RUN FAILED: {rec['failed']}]"
        else:
            wb_name = rec.get("workbook", "")
            el = rec.get("elapsed")
            detail += (f"   ·   workbook: {wb_name}"
                       + (f"   ·   {el:,.0f}s" if isinstance(el, (int, float)) else ""))
        cell = ws.cell(r, 3, detail)
        cell.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=last_m)
        ws.row_dimensions[r].height = 30
        r += 1

    # ---- notes ----
    r += 1
    for note in (
        "Notes:",
        "• All methods run from the SAME manual override window (scenario/manual_events.yaml) "
        "and the SAME control rules — only the planning method differs, so columns are comparable.",
        "• CONSERVATION GATE — PASS = every stocked fish is both conserved (mass) AND placed in a "
        "tank. PARTIAL = mass conserved but some fish left unplaced in the realized layout (see the "
        "'not placed in tanks' rows). FAIL = fish lost/created, or the run errored. Each method is "
        "judged on its OWN authoritative proof (Global: batch-level ReconciliationReport; "
        "Controller: tank-by-tank audit).",
        "• GREEN = best on that row among PASS methods only. PARTIAL/FAIL numbers are greyed — "
        "they are computed on an incomplete or non-conserving plan, so they are not a valid 'best'.",
        "• There is no single 'best' column: hold-the-cap (peak vs cap, utilization), flatness "
        "(variability rows) and cost (feed, transfers) trade against each other. Pick the method "
        "that wins the dimensions you care about, among PASS methods.",
        "• Open a method's own workbook (see METHODS above) for its full sheets "
        "(Advisory, HarvestPlan, per-tank continuity, ReconciliationReport, etc.).",
    ):
        ws.cell(r, 1, note).font = F_SUB
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_m)
        ws.cell(r, 1).alignment = LEFT
        ws.row_dimensions[r].height = 26
        r += 1

    # ---- widths + freeze ----
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    for k in range(n):
        ws.column_dimensions[get_column_letter(first_m + k)].width = 18
    ws.freeze_panes = ws.cell(r_hdr + 1, first_m)
    return ws
