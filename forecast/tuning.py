"""Density-tuning sweep — the one clean method shared by the CLI and the app.

Per-batch density over-cap (the Plan tab / TransferTemplate Section B) is, on a
tank-constrained config, a stocking/capacity problem rather than a controller-
tuning one. This module runs the forecast across a grid of Control-knob values
and reports, for each variant, the PER-BATCH PEAK-DENSITY DISTRIBUTION plus the
conservation gates — the empirical recipe for tuning a new scenario.

Read the *distribution*, not the raw "OVER CAP" count: running near capacity means
many batches peak right AT the cap (the structural between-weekly-check touch,
~1.0-1.1, normal). Only the SEVERE rows (>1.3x) matter, and when no knob reduces
them they are a capacity collision (stagger entries / reduce counts / add tanks),
not slack. See docs/USER_GUIDE.md Section 7.1.

Public API:
    sweep(input_path, config_dir, scenario_dir, grid=DEFAULT_GRID, progress=None)
        -> list[VariantResult]
    analyze(peaks) -> Distribution
    recommend(results) -> Recommendation

The forecast is re-run once per grid row (~90s each); nothing here mutates the
caller's config — each variant runs in its own temp copy.
"""
from __future__ import annotations

import contextlib
import io
import os
import shutil
import statistics
import tempfile
from dataclasses import dataclass, field

import openpyxl
import yaml

from . import run as _run

# Each grid row is (label, {control-knob: value, ...}); the first is the baseline
# (no overrides). These are the knobs that plausibly move per-batch density; on a
# tank-constrained config the baseline usually wins (see module docstring).
#
# FULL_GRID sweeps both directions of every relevant knob. QUICK_GRID is the
# cheapest informative subset — baseline plus the one dominant lever on each axis
# (sizing, finishing/harvest) — for a fast read before committing to the full run.
FULL_GRID: list[tuple[str, dict]] = [
    ("baseline", {}),
    ("density=0.90", {"density_target_pct": 0.90}),
    ("density=0.85", {"density_target_pct": 0.85}),
    ("varqty=20", {"rebalance_varqty_budget": 20}),
    ("balance=60", {"rebalance_balance_budget": 60}),
    # facility_biomass_deviation_pct is the live harvest-tightness knob (it
    # superseded harvest_setpoint_lookahead_weeks, which is now vestigial).
    ("deviation=0.005", {"facility_biomass_deviation_pct": 0.005}),
    ("deviation=0.02", {"facility_biomass_deviation_pct": 0.02}),
]
QUICK_GRID: list[tuple[str, dict]] = [
    ("baseline", {}),
    ("density=0.85", {"density_target_pct": 0.85}),
    ("deviation=0.005", {"facility_biomass_deviation_pct": 0.005}),
]
# Back-compat default.
DEFAULT_GRID = FULL_GRID


def grid_for(quick: bool) -> list:
    """The grid for a 'quick' (cheap subset) or full sweep."""
    return QUICK_GRID if quick else FULL_GRID

# A peak density at/above this fraction of cap is "severe" — the only band worth
# acting on. Below it is the normal between-check overshoot of running near cap.
SEVERE_RATIO = 1.3
# Batches at/above this ratio are listed in the per-variant detail.
DETAIL_RATIO = 1.2


@dataclass
class Distribution:
    """Peak-density distribution for one variant."""
    n: int
    over: int          # peak > 1.0 (touches the cap at all)
    severe: int        # peak >= SEVERE_RATIO
    worst: float
    median: float
    buckets: dict      # {"<=1.0", "1.0-1.1", "1.1-1.3", ">1.3"} -> count


@dataclass
class VariantResult:
    label: str
    overrides: dict
    dist: Distribution
    dropped: int
    overprod: int
    severe_rows: list[dict] = field(default_factory=list)  # Batch/peak/peak_wk/entry_wk

    @property
    def conservation_ok(self) -> bool:
        return self.dropped == 0 and self.overprod == 0


@dataclass
class Recommendation:
    best_label: str
    is_capacity_bound: bool   # True => no knob beats baseline on severe count
    text: str


# --------------------------------------------------------------------------- #
# Extraction from an output workbook
# --------------------------------------------------------------------------- #
def _section_b_rows(out_path):
    """Yield each batch's plan-summary dict from TransferTemplate Section B."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    hdr = None
    for row in wb["TransferTemplate"].iter_rows(values_only=True):
        if row and row[0] == "Batch":
            hdr = [str(c) for c in row if c is not None]
            continue
        if (hdr and row and isinstance(row[0], str) and row[0].startswith("B")
                and len(row[0]) > 1 and row[0][1].isdigit()):
            yield {hdr[i]: row[i] for i in range(min(len(hdr), len(row)))}, hdr


def _col(hdr, prefix):
    return next((c for c in hdr if c.startswith(prefix)), None)


def _peaks_and_detail(out_path):
    peaks = []
    detail = []
    for d, hdr in _section_b_rows(out_path):
        pc = _col(hdr, "Peak_Density")
        wc = _col(hdr, "Peak_Wk")
        ec = _col(hdr, "Wks_from_Start")
        if pc is None:
            # No Peak_Density column: without this guard every batch parsed as
            # peak 0.0 and the distribution reported "all clean" — a missing
            # MEASUREMENT must read as "no data" (gate N/A), never as a pass.
            print("WARN: TransferTemplate Section B has no Peak_Density column "
                  "— per-batch density review unavailable for this workbook")
            return [], []
        try:
            pk = float(d.get(pc) or 0)
        except (TypeError, ValueError):
            continue
        peaks.append(pk)
        if pk >= DETAIL_RATIO:
            detail.append({
                "Batch": d.get("Batch"),
                "Peak_density": round(pk, 2),
                "Peak_wk_from_entry": d.get(wc),
                "Entry_wk_from_start": d.get(ec),
            })
    detail.sort(key=lambda r: r["Peak_density"], reverse=True)
    return peaks, detail


def _conservation(out_path):
    """Scan the audit sheets for dropped / over-produced fish (both must be 0)."""
    wb = openpyxl.load_workbook(out_path, data_only=True)
    dropped = overprod = 0
    if not any(sh in wb.sheetnames
               for sh in ("TankContinuityAudit", "InputConservationAudit")):
        # Absence of evidence must not read as evidence of conservation: both
        # engines write both sheets, so a workbook with neither is foreign —
        # the (0, 0) returned below is "unverified", not "verified clean".
        print(f"WARN: {out_path}: no conservation audit sheet found — "
              f"dropped/over-produced counts are UNVERIFIED (reported as 0)")
    for sh in ("TankContinuityAudit", "InputConservationAudit"):
        if sh not in wb.sheetnames:
            continue
        for row in wb[sh].iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            line = " ".join(cells).upper()
            for c in cells:
                try:
                    v = int(float(c))
                except (TypeError, ValueError):
                    continue
                if v > 0 and "DROP" in line:
                    dropped = max(dropped, v)
                if v > 0 and "OVER-PRODUCED" in line:
                    overprod = max(overprod, v)
    return dropped, overprod


# --------------------------------------------------------------------------- #
# Analysis
# --------------------------------------------------------------------------- #
def analyze(peaks) -> Distribution:
    """Bucket a list of per-batch peak-density ratios."""
    v = list(peaks)
    if not v:
        return Distribution(0, 0, 0, 0.0, 0.0,
                            {"<=1.0": 0, "1.0-1.1": 0, "1.1-1.3": 0, ">1.3": 0})
    b = lambda lo, hi: sum(1 for x in v if lo <= x < hi)
    return Distribution(
        n=len(v),
        over=sum(1 for x in v if x > 1.0001),
        severe=sum(1 for x in v if x >= SEVERE_RATIO),
        worst=max(v),
        median=statistics.median(v),
        buckets={
            "<=1.0": b(0, 1.0001),
            "1.0-1.1": b(1.0001, 1.1),
            "1.1-1.3": b(1.1, SEVERE_RATIO),
            ">1.3": sum(1 for x in v if x >= SEVERE_RATIO),
        },
    )


def recommend(results) -> Recommendation:
    """Pick the variant minimising severe over-cap while conservation holds.

    If the baseline is the (joint) winner, the residual severe peaks are a
    capacity collision, not a tuning problem.
    """
    ok = [r for r in results if r.conservation_ok]
    if not ok:
        return Recommendation("(none)", False,
                              "No variant held conservation — investigate before tuning.")
    # Lower severe is better; tie-break on worst, then on the baseline.
    best = min(ok, key=lambda r: (r.dist.severe, r.dist.worst,
                                  0 if r.label == "baseline" else 1))
    baseline = next((r for r in results if r.label == "baseline"), None)
    capacity_bound = (
        baseline is not None and baseline.conservation_ok
        and best.dist.severe >= baseline.dist.severe
    )
    if capacity_bound:
        text = (
            f"No knob beats baseline ({baseline.dist.severe} severe, "
            f"worst {baseline.dist.worst:.2f}x). The severe peaks are a CAPACITY "
            "collision, not a tuning problem — stagger batch entries, reduce input "
            "counts, or add grow-out tanks (USER_GUIDE Section 7.1)."
        )
    else:
        text = (
            f"Best: {best.label} -- {best.dist.severe} severe (worst "
            f"{best.dist.worst:.2f}x) vs baseline "
            f"{baseline.dist.severe if baseline else '?'}. Apply these knobs in "
            "Configure, then re-run the forecast."
        )
    return Recommendation(best.label, capacity_bound, text)


# --------------------------------------------------------------------------- #
# The sweep
# --------------------------------------------------------------------------- #
def _run_in_tempdir(label, overrides, base_config_dir, base_scenario_dir,
                    input_path) -> str:
    """Run the pipeline once with `overrides` applied onto control.yaml in an
    isolated temp copy of the config/scenario, and return the output workbook
    path. Shared run harness — used by both this module's tuning sweep and
    forecast.optimize's multi-objective sweep, so the temp-copy + override + run
    plumbing lives in exactly one place. Nothing mutates the caller's dirs."""
    work = tempfile.mkdtemp(prefix="as_run_")
    cdir = os.path.join(work, "config")
    sdir = os.path.join(work, "scenario")
    shutil.copytree(base_config_dir, cdir)
    shutil.copytree(base_scenario_dir, sdir)
    cy = os.path.join(cdir, "control.yaml")
    with open(cy) as f:
        cfg = yaml.safe_load(f)
    cfg.update(overrides)
    with open(cy, "w") as f:
        yaml.safe_dump(cfg, f)

    inp = os.path.join(work, os.path.basename(input_path))
    out = os.path.join(work, "out.xlsm")
    shutil.copy(input_path, inp)
    with contextlib.redirect_stdout(io.StringIO()):
        _run.main(inp, out, config_dir=cdir, scenario_dir=sdir)
    return out


def run_variant(label, overrides, base_config_dir, base_scenario_dir,
                input_path) -> VariantResult:
    """Run the pipeline once with `overrides` applied to control.yaml."""
    out = _run_in_tempdir(label, overrides, base_config_dir, base_scenario_dir,
                          input_path)
    peaks, detail = _peaks_and_detail(out)
    dropped, overprod = _conservation(out)
    return VariantResult(
        label=label,
        overrides=dict(overrides),
        dist=analyze(peaks),
        dropped=dropped,
        overprod=overprod,
        severe_rows=detail,
    )


def sweep(input_path, config_dir, scenario_dir, grid=None,
          progress=None) -> list[VariantResult]:
    """Run every grid row and return the per-variant results.

    `progress(i, n, label)` is called before each run, if given (for a UI bar).
    Nothing mutates `config_dir`/`scenario_dir`; each variant runs in a temp copy.
    """
    grid = grid or DEFAULT_GRID
    results = []
    n = len(grid)
    for i, (label, overrides) in enumerate(grid):
        if progress is not None:
            progress(i, n, label)
        results.append(run_variant(label, overrides, config_dir,
                                    scenario_dir, input_path))
    return results
