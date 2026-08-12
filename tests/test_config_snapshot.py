"""The `RunConfig` sheet: what travels in a workbook, and which sheet it is.

Two DIFFERENT sheets are both called `RunConfig`:
  * the controller family writes a re-importable YAML snapshot
    (forecast.config_snapshot), and
  * tools/run_global_forecast.py writes a Global METHOD STAMP — a key/value
    record of what ran, with nothing to restore.

Importing a Global workbook therefore restored nothing, and said so as though
the file had no RunConfig at all. These pin the disambiguation, the honest
message, and the DELIBERATE exclusion of the analysis overlays.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from forecast import config_snapshot as cs


def _seed(tmp_path) -> tuple[Path, Path]:
    cfg = tmp_path / "config"
    scn = tmp_path / "scenario"
    cfg.mkdir()
    scn.mkdir()
    (cfg / "control.yaml").write_text("scenario_name: pinned\n", encoding="utf-8")
    (cfg / "biology.yaml").write_text("sgr: [1, 2]\n", encoding="utf-8")
    (cfg / "facility.yaml").write_text("tanks: []\n", encoding="utf-8")
    (scn / "batches.yaml").write_text("batches: []\n", encoding="utf-8")
    (scn / "limits.yaml").write_text("limits: {}\n", encoding="utf-8")
    # The analysis overlays exist on disk and must NOT be picked up.
    (cfg / "analysis_defaults.yaml").write_text(
        "method: global-lp\noverrides: {hybrid_follow: off}\n", encoding="utf-8")
    (cfg / "targets.yaml").write_text("monthly: {'2026-08': 1}\n", encoding="utf-8")
    (cfg / "economics.yaml").write_text("currency: USD\n", encoding="utf-8")
    return cfg, scn


def _stamp_wb():
    """A Global method-stamp workbook, exactly as run_global_forecast writes it."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cs.SNAPSHOT_SHEET
    ws["A1"] = cs.KIND_STAMP_MARK
    ws["A2"], ws["B2"] = "planning_method", "global-lp"
    ws["A3"], ws["B3"] = "biomass_cap_kg", 3_900_000
    return wb


class TestRoundTrip:
    def test_snapshot_round_trips_the_engine_inputs(self, tmp_path):
        cfg, scn = _seed(tmp_path)
        wb = openpyxl.Workbook()
        cs.write_config_snapshot(wb, config_dir=cfg, scenario_dir=scn)
        out_c, out_s = tmp_path / "c2", tmp_path / "s2"
        restored = cs.import_config_snapshot(wb, out_c, out_s)
        assert set(restored) == {
            "config/control.yaml", "config/biology.yaml", "config/facility.yaml",
            "scenario/batches.yaml", "scenario/limits.yaml"}
        assert (out_c / "control.yaml").read_text(
            encoding="utf-8").strip() == "scenario_name: pinned"


class TestAnalysisOverlaysAreExcludedOnPurpose:
    """DECISION: the analysis overlays stay OUT of the snapshot — they steer how
    a run is judged, not what it computes, and restoring analysis_defaults.yaml
    from a workbook would silently re-point today's Quick-run card at a plan
    that won a tournament on a different PR. The exclusion is only defensible if
    it is VISIBLE, so the sheet has to say it."""

    def test_the_overlays_are_not_restored(self, tmp_path):
        cfg, scn = _seed(tmp_path)
        wb = openpyxl.Workbook()
        cs.write_config_snapshot(wb, config_dir=cfg, scenario_dir=scn)
        out_c, out_s = tmp_path / "c2", tmp_path / "s2"
        restored = cs.import_config_snapshot(wb, out_c, out_s)
        for f in ("analysis_defaults.yaml", "targets.yaml", "economics.yaml"):
            assert f"config/{f}" not in restored
            assert not (out_c / f).exists(), \
                f"{f} was restored from a workbook — it must never be"

    def test_an_import_cannot_overwrite_a_promoted_quick_run_default(self, tmp_path):
        """The concrete harm: the operator promotes a default, then imports an
        older workbook's config. Their promotion must survive."""
        cfg, scn = _seed(tmp_path)
        wb = openpyxl.Workbook()
        cs.write_config_snapshot(wb, config_dir=cfg, scenario_dir=scn)
        (cfg / "analysis_defaults.yaml").write_text(
            "method: controller\n", encoding="utf-8")   # today's promotion
        cs.import_config_snapshot(wb, cfg, scn)
        assert (cfg / "analysis_defaults.yaml").read_text(
            encoding="utf-8").strip() == "method: controller"

    def test_the_sheet_names_what_it_left_out(self, tmp_path):
        """A silent omission is indistinguishable from a bug. The workbook must
        tell its reader which files did NOT travel with it."""
        cfg, scn = _seed(tmp_path)
        wb = openpyxl.Workbook()
        cs.write_config_snapshot(wb, config_dir=cfg, scenario_dir=scn)
        text = "\n".join(
            str(r[0]) for r in wb[cs.SNAPSHOT_SHEET].iter_rows(values_only=True)
            if r and r[0] is not None)
        assert "NOT included" in text
        for f in ("analysis_defaults.yaml", "targets.yaml", "economics.yaml"):
            assert f in text, f"the sheet never mentions omitting {f}"


class TestTheTwoRunConfigSheetsAreToldApart:
    def test_kinds(self, tmp_path):
        cfg, scn = _seed(tmp_path)
        snap = openpyxl.Workbook()
        cs.write_config_snapshot(snap, config_dir=cfg, scenario_dir=scn)
        assert cs.run_config_kind(snap) == "snapshot"
        assert cs.run_config_kind(_stamp_wb()) == "stamp"
        assert cs.run_config_kind(openpyxl.Workbook()) is None

    def test_importing_a_global_stamp_says_which_sheet_it_found(self, tmp_path):
        """The silent failure: import_config_snapshot returned [] and the app
        reported "no RunConfig snapshot found" for a workbook that plainly has
        a RunConfig sheet. It must name the sheet type instead."""
        with pytest.raises(ValueError) as e:
            cs.import_config_snapshot(_stamp_wb(), tmp_path / "c", tmp_path / "s")
        msg = str(e.value)
        assert "METHOD STAMP" in msg
        assert "RunConfig" in msg
        assert "controller" in msg, "the message must say where to get a real one"

    def test_describe_covers_every_case(self, tmp_path):
        cfg, scn = _seed(tmp_path)
        snap = openpyxl.Workbook()
        cs.write_config_snapshot(snap, config_dir=cfg, scenario_dir=scn)
        assert "re-importable" in cs.describe_run_config_sheet(snap)
        assert "METHOD STAMP" in cs.describe_run_config_sheet(_stamp_wb())
        assert "no 'RunConfig' sheet" in cs.describe_run_config_sheet(
            openpyxl.Workbook())

    def test_the_global_writer_still_stamps_the_marker_it_is_matched_on(self):
        """run_config_kind() reads cell A1. If the Global writer's A1 text
        drifts from KIND_STAMP_MARK, every Global workbook silently becomes
        'unknown' again — so the two live in one assertion."""
        import inspect
        from tools import run_global_forecast as rgf
        src = inspect.getsource(rgf)
        assert '_cs.KIND_STAMP_MARK' in src
        assert cs.KIND_STAMP_MARK == "RUN CONFIG — GLOBAL METHOD EXPORT"


def test_an_old_workbook_without_the_a1_marker_still_imports(tmp_path):
    """Back-compat: snapshots written before the marker existed have no kind in
    A1 and must still be recognised by content."""
    cfg, scn = _seed(tmp_path)
    wb = openpyxl.Workbook()
    cs.write_config_snapshot(wb, config_dir=cfg, scenario_dir=scn)
    wb[cs.SNAPSHOT_SHEET]["A1"] = "some older header line"
    assert cs.run_config_kind(wb) == "snapshot"
    assert cs.import_config_snapshot(wb, tmp_path / "c3", tmp_path / "s3")
