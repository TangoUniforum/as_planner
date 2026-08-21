"""NEGATIVE-CONTROL SUITE — proof that every alarm can actually ring.

Policy (operator-endorsed): checks exist to DETECT defects, never to coerce
results — and a check that cannot fire is itself a defect. Twice this project
a gate could not physically report the failure it existed to catch (the
zero-week counter dropped empty weeks by construction; board gates trusted
stale grades). This suite is the systematic fix: for EVERY detection surface
there is a minimal synthetic input containing exactly the defect that check
exists to catch, plus a matching positive control proving the check stays
quiet on clean input. A registry meta-guard fails CI when a new gate is
registered without an alarm test.

Surfaces covered (see the per-section headers):
  1. the analysis gate registry (forecast/analysis.py GATES) — meta-guarded;
  2. the workbook audits (forecast/excel_io.py): InputConservationAudit,
     TankContinuityAudit (incl. the distributed-loss ratio + GradedHarvest
     accounting), SystemLimitsAudit;
  3. the compare harness (tools/run_compare.py): _conservation_verdict and
     _harvest_extras, plus the RunComparison sheet's gate rendering and
     winner-eligibility;
  4. the manual-window lints (zero-harvest week, over-ceiling, MANUAL EVENT
     REFUSED) end-to-end into the ValidationLog and back out through
     forecast.window_weeks;
  5. the tournament hard-gate predicates (variant_hard_ok, probe_outcome,
     hard_gate_fails) incl. None-is-not-a-pass;
  6. the board cache staleness predicates (board_leg_current,
     drop_stale_grades);
  7. the workbook sensors the gates read (sixn_outbound_transfers).

HONESTY RULE baked in: no check was weakened to make a control pass. One
control exposed a real can't-fire bug (tuning._conservation was structurally
blind to the OVER-PRODUCED headline); the CHECK was fixed, and the control
here proves the alarm now rings — see
test_overproduction_alarm_rings_from_the_writers_own_sheet.
"""
from __future__ import annotations

from datetime import date, timedelta
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from forecast import analysis, tournament, tuning
from forecast import excel_io
from forecast.caps import SystemLimits
from forecast.time_grid import forecast_week_labels, iso_week_label


# --------------------------------------------------------------------------- #
# Shared minimal machinery
# --------------------------------------------------------------------------- #
_MON = date(2026, 7, 6)                    # a Monday — clean ISO week starts
_W1 = iso_week_label(_MON)
_W2 = iso_week_label(_MON + timedelta(days=7))


def _sheet_lines(wb, name):
    """Every row of a sheet joined to one searchable text line."""
    return [" ".join(str(c) for c in row if c is not None)
            for row in wb[name].iter_rows(values_only=True)]


def _sheet_text(wb, name):
    return "\n".join(_sheet_lines(wb, name))


def _loc(week_label, tank_id, batch_id, count, biomass_kg=None,
         system_id="OG3", stage="SW", avg_wt_g=1000.0):
    """Minimal BatchLocationRow stand-in (only the attrs the audits read)."""
    return SimpleNamespace(
        week_label=week_label, tank_id=tank_id, batch_id=batch_id,
        count=float(count),
        biomass_kg=float(biomass_kg if biomass_kg is not None else count),
        system_id=system_id, stage=stage, avg_wt_g=avg_wt_g)


def _batch(batch_id="B1", input_count=1000.0, tran_og_date=_MON + timedelta(days=7),
           tran_og_count=None):
    return SimpleNamespace(batch_id=batch_id, input_count=input_count,
                           tran_og_date=tran_og_date,
                           tran_og_count=(input_count if tran_og_count is None
                                          else tran_og_count))


_CONTROL = SimpleNamespace(forecast_start=_MON, horizon_weeks=10)


def _initial_state(tank_id=40, batch_id="B1", count=10000.0, biomass_kg=None):
    tank = SimpleNamespace(is_empty=False, batch_id=batch_id, count=float(count),
                           biomass_kg=float(biomass_kg if biomass_kg is not None
                                            else count))
    return SimpleNamespace(tanks_by_id={tank_id: tank})


def _harvest_workbook(path, weeks, harvest, manual_weeks=()):
    """Smallest workbook the harvest metrics readers accept: Advisory
    (horizon), HarvestPlan (weekly fish), optional ValidationLog narration."""
    wb = Workbook()
    adv = wb.active
    adv.title = "Advisory"
    adv.append(["Week", "Phase"])
    for w in weeks:
        adv.append([w, ""])
    hp = wb.create_sheet("HarvestPlan")
    hp.append(["Week", "Batch", "Tank", "Count (fish)"])
    for w, n in harvest.items():
        hp.append([w, "B1", "T1", n])
    if manual_weeks:
        vl = wb.create_sheet("ValidationLog")
        vl.append(["#", "Category", "Detail"])
        for i, w in enumerate(manual_weeks, 1):
            vl.append([i, "INFO - Manual window (executed)",
                       f"MANUAL EVENT OK — {w}: harvest #1 (scripted)"])
    wb.save(path)
    wb.close()
    return str(path)


# =========================================================================== #
# 1. GATE REGISTRY (forecast/analysis.py) — controls + meta-guard
# =========================================================================== #
# For every registered gate: a ctx containing exactly the defect the gate
# exists to catch ("neg", must land in `fires`), and a clean ctx ("pos",
# must be PASS). A new gate registered without an entry here turns CI red
# (test_every_registered_gate_has_a_negative_control).
GATE_CONTROLS = {
    "conservation": dict(
        neg={"dropped": 30000, "overprod": 0}, fires={"FAIL"},
        pos={"dropped": 0, "overprod": 0}),
    "no_empty_week": dict(
        neg={"zero_weeks": 2}, fires={"FAIL"},
        pos={"zero_weeks": 0}),
    # The steady-harvest CONTRACT floor. Soft by design: near full
    # utilisation every real plan misses it sometimes, and a gate that always
    # FAILs is a gate the operator learns to ignore — so WARN is "fires".
    # (The alarm that DISQUALIFIES is tournament.floor_eligible, which forbids
    # promoting a tuned winner that lowers the worst week.)
    "harvest_floor": dict(
        neg={"weeks_below_floor": 9, "min_week": 16185.0,
             "min_harvest": 30000.0}, fires={"WARN"},
        pos={"weeks_below_floor": 0, "min_week": 31000.0,
             "min_harvest": 30000.0}),
    "biomass_cap": dict(
        neg={"peak_pct_of_cap": 115.0}, fires={"FAIL"},
        pos={"peak_pct_of_cap": 97.0}),
    "harvest_cap": dict(
        neg={"weeks_over_harvest_cap": 5}, fires={"FAIL"},
        pos={"weeks_over_harvest_cap": 0}),
    # Operator decision 2026-08-05: targets are penalized, never disqualifying
    # — the alarm for a missed target is WARN by design, so WARN is "fires".
    "targets": dict(
        neg={"targets_review": {"judged": 2, "met": 1, "close": 0, "missed": 1,
                                "worst_pct": 60.0,
                                "total_shortfall_kg": 40000.0}},
        fires={"WARN"},
        pos={"targets_review": {"judged": 2, "met": 2, "close": 0, "missed": 0,
                                "worst_pct": 100.0,
                                "total_shortfall_kg": 0.0}}),
    # Density quality is a diagnostic lens: severe clustering = stocking
    # problem, flagged WARN by design (no knob fixes it, so no FAIL).
    "density_quality": dict(
        neg={"density_review": {"n": 10, "over": 3, "severe": 2, "worst": 1.52,
                                "median": 1.0, "buckets": {},
                                "severe_rows": []}},
        fires={"WARN"},
        pos={"density_review": {"n": 10, "over": 1, "severe": 0, "worst": 1.05,
                                "median": 0.9, "buckets": {},
                                "severe_rows": []}}),
    "sixn_one_way": dict(
        neg={"sixn_outbound_purge": 3}, fires={"FAIL"},
        pos={"sixn_outbound_purge": 0}),
    "handling_budget": dict(
        neg={"weeks_moves_over_cap": 2, "weeks_moves_warn": 2,
             "moves_week_max": 18}, fires={"FAIL"},
        pos={"weeks_moves_over_cap": 0, "weeks_moves_warn": 0,
             "moves_week_max": 10}),
}


def _gate_row(ctx, key):
    rows = analysis.evaluate_gates(ctx)
    return next(r for r in rows if r["key"] == key)


def test_every_registered_gate_has_a_negative_control():
    """META-GUARD: a gate registered without an alarm-proof is red CI."""
    registered = [g.key for g in analysis.GATES]
    missing = [k for k in registered if k not in GATE_CONTROLS]
    assert not missing, (
        f"gate(s) {missing} registered in forecast.analysis.GATES without a "
        f"negative control — every detection surface ships with a proof it "
        f"can fire. Add entries to GATE_CONTROLS in {__name__}.")
    stale = [k for k in GATE_CONTROLS if k not in registered]
    assert not stale, f"controls {stale} no longer match a registered gate"


@pytest.mark.parametrize("key", sorted(GATE_CONTROLS))
def test_gate_negative_control_fires(key):
    ctrl = GATE_CONTROLS[key]
    row = _gate_row(ctrl["neg"], key)
    assert row["status"] in ctrl["fires"], (
        f"gate {key!r} did NOT fire on its own defect "
        f"(got {row['status']}: {row['detail']}) — a can't-fire bug")


@pytest.mark.parametrize("key", sorted(GATE_CONTROLS))
def test_gate_positive_control_stays_quiet(key):
    ctrl = GATE_CONTROLS[key]
    row = _gate_row(ctrl["pos"], key)
    assert row["status"] == "PASS", (
        f"gate {key!r} cried wolf on clean input "
        f"(got {row['status']}: {row['detail']})")


@pytest.mark.parametrize("key,na_ctx", [
    ("no_empty_week", {}),                     # zero-week count unavailable
    ("biomass_cap", {}),
    ("harvest_cap", {}),
    ("targets", {}),
    ("density_quality", {}),
    ("sixn_one_way", {}),
    ("handling_budget", {}),
])
def test_gate_missing_data_is_na_never_a_verdict(key, na_ctx):
    """Absent evidence must be N/A — neither a false PASS nor a false FAIL."""
    assert _gate_row(na_ctx, key)["status"] == "N/A"


def test_gate_warn_bands_distinguish_relief_from_abuse():
    # 1-3 relief weeks = WARN (exceptional use); ceiling breach = always FAIL.
    assert _gate_row({"weeks_over_harvest_cap": 2}, "harvest_cap")["status"] == "WARN"
    r = _gate_row({"weeks_over_harvest_cap": 1, "weeks_over_relief_ceiling": 1},
                  "harvest_cap")
    assert r["status"] == "FAIL"
    assert _gate_row({"peak_pct_of_cap": 105.0}, "biomass_cap")["status"] == "WARN"


def test_broken_gate_is_reported_as_fail_not_swallowed():
    """A gate whose own code raises must surface as a loud FAIL row."""
    def _boom(ctx):
        raise RuntimeError("sensor exploded")
    analysis.register_gate("nc_tmp_boom", "temp", hard=False, fn=_boom)
    try:
        row = _gate_row({}, "nc_tmp_boom")
        assert row["status"] == "FAIL" and "gate error" in row["detail"]
    finally:
        analysis.GATES[:] = [g for g in analysis.GATES
                             if g.key != "nc_tmp_boom"]


def test_tournament_hard_gate_keys_match_the_registry():
    """tournament.HARD_GATE_KEYS documents itself as 'analysis.GATES keys with
    hard=True' — a hard gate added to the registry without teaching the
    tournament (or vice versa) silently unsyncs the probe logic."""
    assert set(tournament.HARD_GATE_KEYS) == {
        g.key for g in analysis.GATES if g.hard}


# =========================================================================== #
# 2. WORKBOOK AUDITS (forecast/excel_io.py)
# =========================================================================== #
class TestInputConservationAudit:
    def _write(self, batches, locs, harv, **kw):
        wb = Workbook()
        wb.remove(wb.active)
        excel_io.write_input_conservation_audit(wb, batches, locs, harv,
                                                _CONTROL, **kw)
        return wb

    def test_dropped_batch_fires(self):
        # In-horizon TranOG, never placed, never harvested -> DROPPED.
        wb = self._write([_batch(input_count=30000.0)], [], [])
        text = _sheet_text(wb, "InputConservationAudit")
        assert "*** DROPPED ***" in text
        assert "DROPPED" in text and "30,000" in text     # loud headline

    def test_clean_batch_stays_quiet(self):
        wb = self._write([_batch()], [_loc(_W1, 40, "B1", 1000)], [])
        text = _sheet_text(wb, "InputConservationAudit")
        assert "DROPPED" not in text
        assert "OK — every in-horizon batch" in text

    def test_over_production_fires(self):
        # harvested + standing > stocked input: fish were CREATED.
        wb = self._write([_batch(input_count=1000.0)],
                         [_loc(_W1, 40, "B1", 200)],
                         [SimpleNamespace(batch_id="B1", count=1500.0)])
        assert "OVER-PRODUCED" in _sheet_text(wb, "InputConservationAudit")

    def test_fw_divergence_fires(self):
        # Realized seawater entry 20% under the planned tran_og_count.
        tog = [SimpleNamespace(batch_id="B1",
                               destinations=[SimpleNamespace(count=80000.0)])]
        wb = self._write([_batch(input_count=100000.0)],
                         [_loc(_W1, 40, "B1", 80000)], [],
                         tranog_events=tog)
        assert "FW UNDER plan" in _sheet_text(wb, "InputConservationAudit")

    def test_fw_mass_balance_breach_fires(self):
        # FW phase leaks fish: first_FW_count != TranOG + mort + cull (20% gap;
        # planned == realized so the divergence flag can't mask this one).
        tog = [SimpleNamespace(batch_id="B1",
                               destinations=[SimpleNamespace(count=80000.0)])]
        states = {"B1": [SimpleNamespace(stage="FW", week_label="2026-W20",
                                         count=100000.0, mort_count_week=0.0,
                                         cull_count_week=0.0)]}
        wb = self._write([_batch(input_count=100000.0, tran_og_count=80000.0)],
                         [_loc(_W1, 40, "B1", 80000)], [],
                         tranog_events=tog, biology_states_by_batch=states)
        assert "FW MASS-BALANCE BREACH" in _sheet_text(wb, "InputConservationAudit")

    def test_fw_mass_balance_conserving_stays_quiet(self):
        tog = [SimpleNamespace(batch_id="B1",
                               destinations=[SimpleNamespace(count=80000.0)])]
        states = {"B1": [SimpleNamespace(stage="FW", week_label="2026-W20",
                                         count=100000.0,
                                         mort_count_week=15000.0,
                                         cull_count_week=5000.0)]}
        wb = self._write([_batch(input_count=100000.0, tran_og_count=80000.0)],
                         [_loc(_W1, 40, "B1", 80000)], [],
                         tranog_events=tog, biology_states_by_batch=states)
        text = _sheet_text(wb, "InputConservationAudit")
        assert "FW MASS-BALANCE BREACH" not in text
        assert "FW mass-balance OK" in text


class TestFwBalanceCrossingWeek:
    """The FW->SW crossing week must not decide the verdict.

    A week's `stage` is its CLOSING stage and the FW->SW flip always lands on
    the first day of a week, so a TranOG_Date that IS a week start puts the
    reconciliation cull into a week labelled SW. Summing only FW/EGG weeks then
    loses that cull and the gate fires on a conserving batch — the balance
    swung on which weekday the operator picked. (Live case: B49 read +14.4%
    purely from moving tran_og_date 2026-08-27 -> 2026-09-14.)
    """

    def _write(self, batches, locs, harv, **kw):
        wb = Workbook()
        wb.remove(wb.active)
        excel_io.write_input_conservation_audit(wb, batches, locs, harv,
                                                _CONTROL, **kw)
        return wb

    @staticmethod
    def _wk(label, stage, count=0.0, mort=0.0, cull=0.0):
        return SimpleNamespace(stage=stage, week_label=label, count=count,
                               mort_count_week=mort, cull_count_week=cull)

    def _audit(self, states):
        tog = [SimpleNamespace(batch_id="B1",
                               destinations=[SimpleNamespace(count=80000.0)])]
        wb = self._write([_batch(input_count=100000.0, tran_og_count=80000.0)],
                         [_loc(_W1, 40, "B1", 80000)], [],
                         tranog_events=tog, biology_states_by_batch=states)
        return _sheet_text(wb, "InputConservationAudit")

    def test_cull_in_the_sw_crossing_week_is_counted(self):
        """TranOG_Date on a week start: cull + flip land the same day, so the
        cull is recorded in an SW-labelled week. It still closes the balance."""
        text = self._audit({"B1": [
            self._wk("2026-W20", "FW", count=100000.0, mort=15000.0),
            # Crossing week: reads SW, carries the reconciliation cull.
            self._wk("2026-W21", "SW", count=80000.0, cull=5000.0),
        ]})
        assert "FW MASS-BALANCE BREACH" not in text
        assert "FW mass-balance OK" in text

    def test_crossing_week_mortality_is_NOT_credited(self):
        """Only the crossing week's CULL carries over. Its mortality is applied
        after OG entry (TankContinuity's to account for) and `realized_tog` is
        measured before it — crediting it here would under-report a real leak."""
        text = self._audit({"B1": [
            self._wk("2026-W20", "FW", count=100000.0),
            # 20k of mortality in the crossing week must not fill the 20k hole.
            self._wk("2026-W21", "SW", count=80000.0, mort=20000.0),
        ]})
        assert "FW MASS-BALANCE BREACH" in text

    def test_verdict_does_not_depend_on_the_tranog_weekday(self):
        """The same conserving batch, cull booked either side of the boundary,
        must get the same verdict. This is the invariant B49 broke."""
        mid_week = {"B1": [                       # TranOG_Date mid-week
            self._wk("2026-W20", "FW", count=100000.0, mort=15000.0, cull=5000.0),
            self._wk("2026-W21", "SW", count=80000.0),
        ]}
        week_start = {"B1": [                     # TranOG_Date on the boundary
            self._wk("2026-W20", "FW", count=100000.0, mort=15000.0),
            self._wk("2026-W21", "SW", count=80000.0, cull=5000.0),
        ]}
        mid_text, start_text = self._audit(mid_week), self._audit(week_start)
        breach = "FW MASS-BALANCE BREACH"
        assert (breach in mid_text) == (breach in start_text)
        assert breach not in start_text
        assert "FW mass-balance OK" in start_text

    def test_a_genuine_fw_leak_still_fires(self):
        """The carry-over must not blunt the gate: no cull anywhere, 20k gone."""
        text = self._audit({"B1": [
            self._wk("2026-W20", "FW", count=100000.0),
            self._wk("2026-W21", "SW", count=80000.0),
        ]})
        assert "FW MASS-BALANCE BREACH" in text


class TestTankContinuityAudit:
    def _write(self, locs, initial, realized, transfer_events=None,
               harvest_events=None):
        wb = Workbook()
        wb.remove(wb.active)
        excel_io.write_tank_continuity_audit(
            wb, locs, [], harvest_events or [], transfer_events or [], [], [],
            initial, realized_biology=realized)
        return wb

    def test_unexplained_count_loss_fires_tank_drift(self):
        # 1,060 fish vanish week-over-week with no event and no mortality.
        locs = [_loc(_W1, 40, "B1", 10000), _loc(_W2, 40, "B1", 8940)]
        wb = self._write(locs, _initial_state(count=10000), realized={})
        assert "TANK_DRIFT" in _sheet_text(wb, "TankContinuityAudit")

    def test_unexplained_biomass_jump_fires_bio_drift(self):
        # +2,000 kg appear with zero recorded growth.
        locs = [_loc(_W1, 40, "B1", 10000, 10000),
                _loc(_W2, 40, "B1", 10000, 12000)]
        realized = {(40, _W1, "B1"): [0.0, 0.0], (40, _W2, "B1"): [0.0, 0.0]}
        wb = self._write(locs, _initial_state(count=10000), realized)
        assert "BIO_DRIFT" in _sheet_text(wb, "TankContinuityAudit")

    def test_clean_tank_history_stays_quiet(self):
        locs = [_loc(_W1, 40, "B1", 10000, 10000),
                _loc(_W2, 40, "B1", 10000, 10000)]
        realized = {(40, _W1, "B1"): [0.0, 0.0], (40, _W2, "B1"): [0.0, 0.0]}
        wb = self._write(locs, _initial_state(count=10000), realized)
        text = _sheet_text(wb, "TankContinuityAudit")
        assert "TANK_DRIFT" not in text and "BIO_DRIFT" not in text

    @staticmethod
    def _facility_count_ratio(wb):
        for row in wb["TankContinuityAudit"].iter_rows(values_only=True):
            if row and str(row[0]) == "Count (fish)":
                return float(row[3])
        raise AssertionError("facility conservation summary row missing")

    def test_distributed_leak_fires_the_ratio_where_rows_are_blind(self):
        # 30 tanks each lose 40 fish/week — UNDER the 50-fish per-row
        # tolerance, so no row flags — but the same-signed facility sum
        # must peg the signed/abs leak gauge at ~1 (systematic loss).
        locs, init_tanks = [], {}
        for tid in range(101, 131):
            init_tanks[tid] = SimpleNamespace(is_empty=False, batch_id="B1",
                                              count=10040.0, biomass_kg=10040.0)
            locs.append(_loc(_W1, tid, "B1", 10000, 10000))
            locs.append(_loc(_W2, tid, "B1", 9960, 9960))
        wb = self._write(locs, SimpleNamespace(tanks_by_id=init_tanks),
                         realized={(t, w, "B1"): [0.0, 0.0]
                                   for t in range(101, 131)
                                   for w in (_W1, _W2)})
        text = _sheet_text(wb, "TankContinuityAudit")
        assert "TANK_DRIFT" not in text          # per-row alarms blind, by design
        ratio = self._facility_count_ratio(wb)
        assert abs(ratio) > 0.3, (
            f"distributed 2,400-fish leak left the facility ratio at {ratio} "
            f"— the distributed-loss gauge cannot fire")

    def test_clean_facility_ratio_stays_near_zero(self):
        locs = [_loc(_W1, 40, "B1", 10000, 10000),
                _loc(_W2, 40, "B1", 10000, 10000)]
        realized = {(40, _W1, "B1"): [0.0, 0.0], (40, _W2, "B1"): [0.0, 0.0]}
        wb = self._write(locs, _initial_state(count=10000), realized)
        assert abs(self._facility_count_ratio(wb)) < 0.3

    @staticmethod
    def _graded(pickup_count=6000.0, retention_count=4000.0):
        """GradedHarvest-shaped event (rides in transfer_events)."""
        return SimpleNamespace(
            pickup_tank_id=61, retention_tank_id=40, source_tank_id=40,
            event_date=_MON, pickup_count=pickup_count,
            pickup_avg_wt_g=1000.0, pickup_source_avg_wt_g=1000.0,
            retention_count=retention_count, retention_avg_wt_g=1000.0)

    def test_graded_harvest_split_is_accounted_no_phantom_drift(self):
        # POSITIVE: a correct 6,000/4,000 graded split reconciles exactly. If
        # GradedHarvest accounting ever went missing, THIS control trips
        # (source shows 10,000 fish leaving with no recorded event).
        locs = [_loc(_W1, 40, "B1", 4000, 4000),
                _loc(_W1, 61, "B1", 6000, 6000, system_id="OG6N",
                     stage="STARVE")]
        wb = self._write(locs, _initial_state(count=10000),
                         realized={}, transfer_events=[self._graded()])
        assert "TANK_DRIFT" not in _sheet_text(wb, "TankContinuityAudit")

    def test_graded_harvest_leg_that_loses_fish_fires(self):
        # NEGATIVE: the pickup tank lands 3,000 fish short of the credited leg.
        locs = [_loc(_W1, 40, "B1", 4000, 4000),
                _loc(_W1, 61, "B1", 3000, 3000, system_id="OG6N",
                     stage="STARVE")]
        wb = self._write(locs, _initial_state(count=10000),
                         realized={}, transfer_events=[self._graded()])
        assert "TANK_DRIFT" in _sheet_text(wb, "TankContinuityAudit")


class TestSystemLimitsAudit:
    def _write(self, locs, caps, tables=None, control=None, **sl_kwargs):
        # A real SystemLimits, not a duck-typed stand-in: the audit resolves
        # caps through caps.carry_forward_cap_lookup (per-week exception >
        # system+mode default > system default), so a stub carrying only
        # `.caps` would test a lookup the product does not use.
        wb = Workbook()
        wb.remove(wb.active)
        res = excel_io.write_system_limits_audit(
            wb, locs, {}, tables, SystemLimits(caps=caps, **sl_kwargs),
            control or SimpleNamespace(global_buffer_pct=0.0))
        return wb, res

    def test_system_biomass_over_cap_fires(self):
        wb, (nb, nf, worst_b, _) = self._write(
            [_loc(_W1, 40, "B1", 1500, 1500)],
            {(_W1, "OG3", "biomass"): 1000.0})
        assert nb == 1 and worst_b > 1.0
        assert "BIOMASS_OVER" in _sheet_text(wb, "SystemLimitsAudit")

    def test_system_within_cap_stays_quiet(self):
        wb, (nb, nf, _, _) = self._write(
            [_loc(_W1, 40, "B1", 900, 900)],
            {(_W1, "OG3", "biomass"): 1000.0})
        assert (nb, nf) == (0, 0)
        assert "BIOMASS_OVER" not in _sheet_text(wb, "SystemLimitsAudit")

    def test_feed_over_cap_fires(self, monkeypatch):
        # The flag logic is under test, not the feed math — pin the per-row
        # realized feed above the cap.
        monkeypatch.setattr(excel_io, "_row_feed_kg_day",
                            lambda r, b, t: 50.0)
        wb, (nb, nf, _, worst_f) = self._write(
            [_loc(_W1, 40, "B1", 900, 900)],
            {(_W1, "OG3", "feed_per_day"): 30.0})
        assert nf == 1 and worst_f > 1.0
        assert "FEED_OVER" in _sheet_text(wb, "SystemLimitsAudit")

    def test_og6n_biomass_cap_exempt_in_purge_but_fires_in_production(self):
        # SPEC CHANGED 2026-08-20 (operator), superseding the 2026-08-14 ruling
        # this test used to encode. 6N in PURGE mode has NO biomass cap: what
        # bounds a depuration tank is the HARVEST SCHEDULE, not a kg figure.
        # (The 600,000 kg that had been configured was a placeholder the
        # operator picked "just to add a number", never a real limit.) In
        # PRODUCTION mode 6N is an ordinary system and every cap applies.
        #
        # BOTH halves are pinned on purpose. An exemption that is really a
        # check-that-can-never-fire is the exact defect class this suite
        # exists to catch, so the production half proves the check still works
        # and the purge half proves the carve-out is scoped to purge alone.
        over = [_loc(_W1, 61, "B1", 5000, 5000, system_id="OG6N",
                     stage="STARVE")]
        caps = {(_W1, "OG6N", "biomass"): 1000.0}

        # PURGE: sixn_production_start None -> purge_mode_on True. Exempt.
        purge_ctl = SimpleNamespace(global_buffer_pct=0.0,
                                    sixn_growth=False,
                                    sixn_production_start=None)
        wb, (nb, _nf, worst_b, _) = self._write(over, caps, control=purge_ctl)
        assert nb == 0 and worst_b == 0.0
        text = _sheet_text(wb, "SystemLimitsAudit")
        assert "BIOMASS_OVER" not in text
        # The tonnage and its cap are still WRITTEN — exempt from FLAGGING is
        # not the same as hidden, and the operator must still see the number.
        assert "OG6N" in text

        # PRODUCTION: sixn_growth True -> purge_mode_on False. Cap binds.
        prod_ctl = SimpleNamespace(global_buffer_pct=0.0,
                                   sixn_growth=True,
                                   sixn_production_start=None)
        wb2, (nb2, _nf2, worst_b2, _) = self._write(over, caps,
                                                    control=prod_ctl)
        assert nb2 == 1 and worst_b2 > 1.0
        assert "BIOMASS_OVER" in _sheet_text(wb2, "SystemLimitsAudit")

    def test_og6n_feed_stays_exempt_for_a_physical_reason(self):
        # The ONE exemption that survives, and it is physics not policy: purge
        # fish are STARVE, `_row_feed_kg_day` returns 0 for them, so a feed-RATE
        # cap on a system that by construction eats nothing could only ever
        # report 0. Pinned so nobody "fixes" it into a check that cannot fire —
        # the exact defect class this suite exists to catch.
        wb, (nb, nf, _, _) = self._write(
            [_loc(_W1, 61, "B1", 5000, 5000, system_id="OG6N",
                  stage="STARVE")],
            {(_W1, "OG6N", "feed_per_day"): 0.001})
        assert nf == 0


# =========================================================================== #
# 3. COMPARE HARNESS (tools/run_compare.py + RunComparison sheet)
# =========================================================================== #
class TestConservationVerdict:
    @staticmethod
    def _save(wb, path):
        wb.save(path)
        wb.close()
        return str(path)

    def _controller_wb(self, path, status_row=None, headline=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "InputConservationAudit"
        ws.append(["INPUT-FISH CONSERVATION AUDIT"])
        if headline:
            ws.append([headline])
        ws.append(["Batch", "Input_Count (fish)", "TranOG_Date", "In_Horizon",
                   "Placed", "Harvested (fish)", "Standing@Horizon (fish)",
                   "Status", "Fish_At_Risk (fish)"])
        if status_row:
            ws.append(status_row)
        return self._save(wb, path)

    def _global_wb(self, path, residual, headline=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "ReconciliationReport"
        ws.append(["Batch", "Seeded", "Residual_pct"])
        ws.append(["FACILITY", 1000000, residual])
        if headline:
            ica = wb.create_sheet("InputConservationAudit")
            ica.append([headline])
        return self._save(wb, path)

    def test_controller_dropped_fish_fail(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        p = self._controller_wb(
            tmp_path / "c_neg.xlsx",
            status_row=["B1", 30000, "2026-07-13", "Y", "N", 0, 0,
                        "*** DROPPED ***", 30000],
            headline="*** 1 batch(es) DROPPED — 30,000 stocked fish "
                     "(17.4% of in-horizon input) never placed. ***")
        v = _conservation_verdict(p)
        assert v["gate"] == "FAIL" and v["dropped"] > 0

    def test_controller_clean_pass(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        p = self._controller_wb(
            tmp_path / "c_pos.xlsx",
            status_row=["B1", 30000, "2026-07-13", "Y", "Y", 30000, 0,
                        "PLACED", ""],
            headline="OK — every in-horizon batch reached the realized "
                     "facility (0 dropped fish).")
        v = _conservation_verdict(p)
        assert v["gate"] == "PASS" and v["dropped"] == 0

    def test_overproduction_alarm_rings_from_the_writers_own_sheet(self, tmp_path):
        """THE FIXED CAN'T-FIRE BUG (2026-08-10): the audit writer states
        over-production in a single text headline cell; tuning._conservation
        only read numeric CELLS, so overprod was structurally always 0 and
        fish CREATION passed the conservation gate everywhere. This control
        runs the REAL writer, then asserts the alarm actually rings."""
        wb = Workbook()
        wb.remove(wb.active)
        excel_io.write_input_conservation_audit(
            wb, [_batch(input_count=1000.0)], [_loc(_W1, 40, "B1", 200)],
            [SimpleNamespace(batch_id="B1", count=1500.0)], _CONTROL)
        p = self._save(wb, tmp_path / "overprod.xlsx")
        dropped, overprod = tuning._conservation(p)
        assert overprod > 0, "over-production written by the audit but invisible " \
                             "to the conservation check — can't-fire bug is back"
        from tools.run_compare import _conservation_verdict
        assert _conservation_verdict(p)["overprod"] > 0

    def test_global_mass_residual_fail(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        v = _conservation_verdict(self._global_wb(tmp_path / "g_neg.xlsx", 0.5))
        assert v["gate"] == "FAIL" and v["residual_pct"] == 0.5

    def test_global_unparseable_residual_is_fail_not_pass(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        v = _conservation_verdict(self._global_wb(tmp_path / "g_bad.xlsx", "n/a"))
        assert v["gate"] == "FAIL"

    def test_global_unplaced_fish_partial(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        p = self._global_wb(
            tmp_path / "g_part.xlsx", 0.0,
            headline="*** 2 batch(es) DROPPED — 60,000 stocked fish "
                     "(10.0% of in-horizon input) never placed. ***")
        v = _conservation_verdict(p)
        assert v["gate"] == "PARTIAL" and v["unplaced_batches"] == 2

    def test_global_clean_pass(self, tmp_path):
        from tools.run_compare import _conservation_verdict
        v = _conservation_verdict(self._global_wb(tmp_path / "g_pos.xlsx", 0.0))
        assert v["gate"] == "PASS"


class TestHarvestExtras:
    _WEEKS = ["2026-W28", "2026-W29", "2026-W30"]

    def test_zero_week_fires(self, tmp_path):
        from tools.run_compare import _harvest_extras
        p = _harvest_workbook(tmp_path / "z.xlsx", self._WEEKS,
                              {"2026-W28": 10000.0, "2026-W30": 10000.0})
        h = _harvest_extras(p, min_harvest=5000.0)
        assert h["zero_weeks"] == 1 and h["min_week"] == 0.0
        assert h["weeks_below_min"] == 1

    def test_below_min_floor_fires(self, tmp_path):
        from tools.run_compare import _harvest_extras
        p = _harvest_workbook(tmp_path / "bm.xlsx", self._WEEKS,
                              {w: 4000.0 for w in self._WEEKS})
        assert _harvest_extras(p, min_harvest=5000.0)["weeks_below_min"] == 3

    def test_steady_weeks_stay_quiet(self, tmp_path):
        from tools.run_compare import _harvest_extras
        p = _harvest_workbook(tmp_path / "ok.xlsx", self._WEEKS,
                              {w: 10000.0 for w in self._WEEKS})
        h = _harvest_extras(p, min_harvest=5000.0)
        assert h["zero_weeks"] == 0 and h["weeks_below_min"] == 0

    def test_window_never_blinds_planner_weeks(self, tmp_path):
        # An engine-planned empty week OUTSIDE the operator window still
        # counts — the exclusion must not become a blindfold.
        from tools.run_compare import _harvest_extras
        p = _harvest_workbook(tmp_path / "wb.xlsx", self._WEEKS,
                              {"2026-W28": 10000.0, "2026-W29": 10000.0},
                              manual_weeks=("2026-W28",))
        h = _harvest_extras(p, min_harvest=5000.0)
        assert h["zero_weeks"] == 1                # W30, the planner's own
        assert h["window_weeks_excluded"] == 1


class TestRunComparisonRendering:
    @staticmethod
    def _metrics():
        return SimpleNamespace(
            overall_peak_biomass=100000.0, overall_mean_biomass=90000.0,
            biomass_cap=120000.0, weeks_over_harvest_cap=0,
            system_overshoot=0.0, density_overshoot=0.0, system_peak=0.9,
            density_peak=80.0, tank_footprint_mean=10.0,
            tank_footprint_peak=12, batch_tank_path_mean=2.0,
            batch_tank_path_max=3, biomass_var=0.1, harvest_var=0.1,
            feed_var=0.1, feed_load=1000.0, transfers_per_fish=0.5,
            between_system={}, within_system={})

    @classmethod
    def _rec(cls, key, gate, **over):
        rec = {"key": key, "label": key.upper(), "family": "F", "blurb": "",
               "workbook": f"{key}.xlsx", "failed": None, "elapsed": 10.0,
               "dropped": 0, "overprod": 0, "gate": gate,
               "metrics": cls._metrics(),
               "harvest": {"n_weeks": 10, "min_week": 5000.0,
                           "max_week": 20000.0, "weeks_below_min": 0,
                           "zero_weeks": 0, "min_harvest": 1000.0},
               "placement": {"unplaced_batches": 0, "unplaced_fish": 0}}
        rec.update(over)
        return rec

    def _sheet(self, records):
        wb = Workbook()
        wb.remove(wb.active)
        return excel_io.write_run_comparison(wb, records)

    def test_gate_row_shows_fail_partial_and_run_failed(self):
        ws = self._sheet([
            self._rec("a", "PASS"),
            self._rec("b", "FAIL", dropped=1000),
            self._rec("c", "PARTIAL",
                      placement={"unplaced_batches": 2, "unplaced_fish": 60000}),
            self._rec("d", None, failed="boom", metrics=None, harvest=None,
                      placement=None),
        ])
        cells = [ws.cell(6, 3 + k).value for k in range(4)]
        assert cells[0] == "PASS"
        assert str(cells[1]).startswith("FAIL")
        assert "PARTIAL" in str(cells[2]) and "2" in str(cells[2])
        assert cells[3] == "RUN FAILED"

    def test_failing_method_is_never_crowned_winner(self):
        # B posts the 'better' zero-week number but FAILS conservation —
        # the green crown must go to the only PASS method.
        rec_a = self._rec("a", "PASS")
        rec_a["harvest"]["zero_weeks"] = 1
        rec_b = self._rec("b", "FAIL", dropped=1000)
        rec_b["harvest"]["zero_weeks"] = 0
        ws = self._sheet([rec_a, rec_b])
        row = next(r for r in range(1, ws.max_row + 1)
                   if ws.cell(r, 1).value == "Zero-harvest weeks")
        green = lambda c: str(c.fill.start_color.rgb).endswith("C6EFCE")
        assert green(ws.cell(row, 3)), "PASS method should be the winner"
        assert not green(ws.cell(row, 4)), \
            "a FAILING method was crowned best — the gate does not bind"


# =========================================================================== #
# 4. MANUAL-WINDOW LINTS (forecast/manual_window.py -> ValidationLog -> reader)
# =========================================================================== #
def _window_state():
    from forecast.state import FacilityState, TankState
    t = TankState(location_id="OG3-40", tank_id=40, system_id="OG3",
                  volume_m3=500.0, max_density_kg_m3=95.0,
                  max_feed_kg_day_cap=1000.0, type="OG")
    t.assign("B1", 5000.0, 1000.0, 10.0, "SW")
    return FacilityState(today=_MON, tanks=[t])


_WIN_CTRL = SimpleNamespace(handling_mortality_pct=0.0, sixn_growth=True,
                            sixn_production_start=None,
                            max_harvest_per_week=1000.0,
                            harvest_relief_pct=0.10)


def _run_window(events, n_weeks=1):
    from forecast.manual_window import advance_facility_window
    return advance_facility_window(_window_state(), {}, None, _MON, n_weeks,
                                   events=events, control=_WIN_CTRL)


class TestManualWindowLints:
    def test_window_week_without_harvest_fires(self):
        from forecast.manual_events import ManualEvent
        res = _run_window([ManualEvent(type="harvest", week=2, from_tank=40)],
                          n_weeks=1)
        assert any("schedules NO harvest" in w for w in res["warnings"]), \
            "a scripted zero-harvest week produced no lint"

    def test_a_window_with_NO_events_at_all_still_fires_every_week(self):
        """The emptiest window of all — `--advance-weeks N` with nothing
        scripted (run.py: window_n = max(advance_weeks, last event week), so a
        window can exist with an EMPTY event list). Those N weeks harvest
        nothing, which is precisely what this lint exists to catch.

        NEGATIVE CONTROL: the lint used to sit inside `if events:`, so the one
        case it could never fire on was the one where EVERY week is empty. On
        the parent commit this returns zero warnings."""
        res = _run_window([], n_weeks=3)
        hits = [w for w in res["warnings"] if "schedules NO harvest" in w]
        assert len(hits) == 3, \
            f"an all-empty 3-week window produced {len(hits)} lints, not 3"

    def test_impossible_event_is_refused_loudly(self):
        from forecast.manual_events import ManualEvent
        res = _run_window([ManualEvent(type="harvest", week=1, from_tank=99)])
        assert any(w.startswith("MANUAL EVENT REFUSED") for w in res["warnings"])

    def test_over_ceiling_scripted_harvest_fires(self):
        from forecast.manual_events import ManualEvent
        res = _run_window([ManualEvent(type="harvest", week=1, from_tank=40,
                                       count=5000.0)])
        assert any("above the plant ceiling" in w for w in res["warnings"])

    def test_legal_scripted_harvest_stays_quiet(self):
        from forecast.manual_events import ManualEvent
        res = _run_window([ManualEvent(type="harvest", week=1, from_tank=40,
                                       count=500.0)])
        assert any(w.startswith("MANUAL EVENT OK") for w in res["warnings"])
        assert not any("schedules NO harvest" in w for w in res["warnings"])
        assert not any("REFUSED" in w for w in res["warnings"])
        assert not any("above the plant ceiling" in w for w in res["warnings"])

    def test_lints_land_in_validation_log_and_mark_the_window(self):
        # End-to-end: window warnings -> ValidationLog categories -> the
        # window-week reader every compliance metric depends on.
        from forecast.manual_events import ManualEvent
        from forecast.window_weeks import manual_window_weeks
        wk1 = forecast_week_labels(_MON, 1)[0]
        res = _run_window([
            ManualEvent(type="harvest", week=1, from_tank=40, count=500.0),
            ManualEvent(type="harvest", week=1, from_tank=99),
        ])
        wb = Workbook()
        wb.remove(wb.active)
        excel_io.write_validation_log(wb, invariant_warnings=res["warnings"])
        cats = {str(row[1]) for row in
                wb["ValidationLog"].iter_rows(values_only=True)
                if row and row[1]}
        assert "ERROR - Manual window (REFUSED)" in cats, \
            "a refused scripted event is not escalated to ERROR"
        assert "INFO - Manual window (executed)" in cats
        assert manual_window_weeks(wb) == {wk1}


class TestDarkHandoff:
    def test_window_that_drains_6n_fires_dark_weeks(self):
        from forecast.manual_events import ManualEvent
        from forecast.manual_window import dark_handoff_weeks
        dark = dark_handoff_weeks(
            {61: 20000.0},
            [ManualEvent(type="harvest", week=1, from_tank=61)],
            window_weeks=2, hold_weeks=2)
        assert dark == [3, 4], "a drained 6N pipeline must go dark at handoff"

    def test_untouched_pr_fish_cover_the_handoff(self):
        from forecast.manual_window import dark_handoff_weeks
        assert dark_handoff_weeks({61: 20000.0}, [], window_weeks=2,
                                  hold_weeks=2) == []


# =========================================================================== #
# 5. TOURNAMENT HARD-GATE PREDICATES (forecast/tournament.py)
# =========================================================================== #
class TestTournamentPredicates:
    def test_non_conserving_variant_fails(self):
        v = SimpleNamespace(conservation_ok=False, metrics=None)
        assert tournament.variant_hard_ok(v) is False

    def test_empty_week_variant_fails(self):
        v = SimpleNamespace(conservation_ok=True,
                            metrics=SimpleNamespace(harvest_zero_weeks=2))
        assert tournament.variant_hard_ok(v) is False

    def test_clean_variant_passes(self):
        v = SimpleNamespace(conservation_ok=True,
                            metrics=SimpleNamespace(harvest_zero_weeks=0))
        assert tournament.variant_hard_ok(v) is True

    def test_pre_schema_metrics_are_unknowable_not_a_pass(self):
        # Old cache entries predate harvest_zero_weeks: verdict None, and a
        # probe must NEVER treat None as a fix — a gate is only cleared by
        # a measurement.
        v = SimpleNamespace(conservation_ok=True, metrics=SimpleNamespace())
        assert tournament.variant_hard_ok(v) is None
        assert tournament.probe_outcome([None, None]) == "gate-bound"
        assert tournament.probe_outcome([None, False]) == "gate-bound"
        assert tournament.probe_outcome([None, True]) == "fixable"

    def test_hard_gate_fails_extracts_only_hard_failures(self):
        gates = analysis.evaluate_gates(
            {"dropped": 5, "zero_weeks": 1, "peak_pct_of_cap": 130.0})
        fails = tournament.hard_gate_fails(gates)
        assert set(fails) == {"conservation", "no_empty_week"}, \
            "hard-gate extraction must fire on hard FAILs and only those"
        clean = analysis.evaluate_gates({"dropped": 0, "zero_weeks": 0})
        assert tournament.hard_gate_fails(clean) == []


# =========================================================================== #
# 6. BOARD CACHE STALENESS (forecast/analysis.py predicates)
# =========================================================================== #
class TestBoardStaleness:
    def test_stale_or_foreign_leg_is_never_replayed(self):
        good = {"sig": "A", "res": {"ok": 1}}
        assert analysis.board_leg_current(good, "A") is True     # positive
        assert analysis.board_leg_current(good, "B") is False    # sig mismatch
        assert analysis.board_leg_current({"res": {}}, "A") is False   # no sig
        assert analysis.board_leg_current({"sig": "A", "res": "x"}, "A") is False
        assert analysis.board_leg_current(None, "A") is False
        assert analysis.board_leg_current(["junk"], "A") is False

    def test_old_schema_grades_are_dropped_engine_output_kept(self):
        res = {"_score": {"schema": "v1", "total": 5},
               "_ana_rows": {"schema": "v1"},
               "workbook_bytes": b"engine-output"}
        assert analysis.drop_stale_grades(res, "v2") is True
        assert "_score" not in res and "_ana_rows" not in res
        assert res["workbook_bytes"] == b"engine-output"

    def test_current_schema_grades_survive(self):
        res = {"_score": {"schema": "v2", "total": 5}}
        assert analysis.drop_stale_grades(res, "v2") is False
        assert res["_score"]["total"] == 5

    def test_unstamped_grades_are_stale_by_definition(self):
        res = {"_score": {"total": 5}}
        assert analysis.drop_stale_grades(res, "v2") is True
        assert "_score" not in res


# =========================================================================== #
# 7. WORKBOOK SENSORS the gates read
# =========================================================================== #
class TestSixnOutboundSensor:
    @staticmethod
    def _wb(path, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "TransferPlan"
        ws.append(["Week", "Type", "From_Tank", "Count (fish)"])
        for r in rows:
            ws.append(r)
        wb.save(path)
        wb.close()
        return str(path)

    def test_depuration_era_outbound_6n_move_is_counted(self, tmp_path):
        p = self._wb(tmp_path / "r7.xlsx",
                     [["2026-W29", "Transfer", 61, 5000]])
        assert analysis.sixn_outbound_transfers(p) == 1

    def test_production_era_moves_are_legal(self, tmp_path):
        p = self._wb(tmp_path / "r7ok.xlsx",
                     [["2026-W29", "Transfer", 61, 5000]])
        assert analysis.sixn_outbound_transfers(
            p, production_start_iso="2026-01-01") == 0

    def test_non_6n_and_non_transfer_rows_stay_quiet(self, tmp_path):
        p = self._wb(tmp_path / "r7q.xlsx",
                     [["2026-W29", "Transfer", 40, 5000],
                      ["2026-W29", "TranOG", 61, 5000]])
        assert analysis.sixn_outbound_transfers(p) == 0

    def test_missing_sheet_is_none_never_a_verdict(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Other"
        p = tmp_path / "nosheet.xlsx"
        wb.save(p)
        wb.close()
        assert analysis.sixn_outbound_transfers(str(p)) is None


# =========================================================================== #
# 8. AUDIT SURFACE META-GUARD
# =========================================================================== #
# The workbook audits have no runtime registry (they are writer functions), so
# the coverage map is maintained here BY NAME: each audit writer this suite
# controls, with the test class that proves it can fire. A rename/removal of a
# writer, or a mapping entry pointing at a vanished test class, turns CI red —
# the closest enumerable equivalent of the GATES meta-guard.
AUDIT_CONTROLS = {
    "write_input_conservation_audit": "TestInputConservationAudit",
    "write_tank_continuity_audit": "TestTankContinuityAudit",
    "write_system_limits_audit": "TestSystemLimitsAudit",
    "write_validation_log": "TestManualWindowLints",
    "write_run_comparison": "TestRunComparisonRendering",
}


def test_every_mapped_audit_writer_and_control_exists():
    for fn, test_cls in AUDIT_CONTROLS.items():
        assert callable(getattr(excel_io, fn, None)), (
            f"audit writer excel_io.{fn} vanished/renamed — update "
            f"AUDIT_CONTROLS and its negative controls together")
        assert test_cls in globals(), (
            f"control class {test_cls} for excel_io.{fn} is missing")


# --------------------------------------------------------------------------- #
# 8. The REALIZED-PLAN audit (forecast/analysis.realized_plan_audit)
#
# These three alarms exist BECAUSE the pre-existing ones could not answer the
# question they appeared to answer. The harvest-floor warning was raised by the
# scheduler's demand pass, which runs before make-room, level-loading and the
# 6N fallback ladder; those later passes fix weeks it flagged and break weeks it
# did not. On the 8.13 PR the realized plan was under the floor in 29 weeks, the
# log named 3, and one of those 3 was comfortably fine in the plan that shipped.
# A detector that is wrong in BOTH directions is the exact failure this suite
# exists to prevent, so these replacements ship with their own controls.
# --------------------------------------------------------------------------- #
class TestRealizedPlanAudit:
    """Each alarm fires on a plan containing exactly its defect, and stays
    silent on a clean plan. No alarm may depend on a mid-plan pass."""

    class _Ev:
        """Minimal harvest event: the audit reads only date + count."""
        def __init__(self, d, count):
            self.event_date, self.count = d, count

    class _Dest:
        def __init__(self, tank_id, count):
            self.tank_id, self.count = tank_id, count

    def _control(self, floor=30_000.0, ceiling=60_000.0, moves=15):
        return SimpleNamespace(
            min_harvest_per_week=floor, max_harvest_per_week=ceiling,
            max_transfers_per_week=moves,
            facility_biomass_deviation_pct=0.0, max_biomass_kg=1e9,
            max_feed_per_day_kg=1e9, default_hog_yield=0.81,
            forecast_start=_MON, horizon_weeks=4)

    def _run(self, harvests=(), transfers=(), control=None, limits=None,
             window=frozenset()):
        from forecast.analysis import realized_plan_audit
        from forecast.caps import FacilityLimits
        return realized_plan_audit(
            list(harvests), list(transfers), limits or FacilityLimits(),
            control or self._control(), window_weeks=window)

    # ---- harvest floor ----------------------------------------------------
    def test_floor_alarm_rings_on_a_short_week(self):
        out = self._run(harvests=[self._Ev(_MON, 10_000.0)])
        assert any("HARVEST FLOOR" in m for m in out), out
        assert any("20,000 under" in m for m in out), out

    def test_floor_alarm_stays_quiet_when_the_floor_is_met(self):
        out = self._run(harvests=[self._Ev(_MON, 30_000.0)])
        assert not [m for m in out if "HARVEST FLOOR" in m], out

    def test_floor_uses_the_PER_WEEK_resolved_cap_not_the_control_default(self):
        """A check against the flat default silently passes every week the
        operator RAISED in scenario/limits.yaml — the defect that let 26 short
        weeks go unreported."""
        from forecast.caps import FacilityLimits, METRIC_MIN_HARVEST
        from forecast.time_grid import iso_week_label
        wk = iso_week_label(_MON)
        lim = FacilityLimits(overrides={(wk, METRIC_MIN_HARVEST): 48_000.0})
        # 40,000 clears the 30,000 default but not this week's raised floor.
        assert not [m for m in self._run(harvests=[self._Ev(_MON, 40_000.0)])
                    if "HARVEST FLOOR" in m]
        out = self._run(harvests=[self._Ev(_MON, 40_000.0)], limits=lim)
        assert any("HARVEST FLOOR" in m and "48,000" in m for m in out), out

    def test_a_rounding_scale_miss_is_reported_but_marked(self):
        """Suppressing it would repeat the old sin; leaving eight identical
        '72 fish' lines unmarked trains the reader to ignore the category."""
        out = [m for m in self._run(harvests=[self._Ev(_MON, 29_950.0)])
               if "HARVEST FLOOR" in m]
        assert out and "rounding-scale" in out[0], out

    def test_a_material_miss_is_NOT_marked_rounding_scale(self):
        out = [m for m in self._run(harvests=[self._Ev(_MON, 10_000.0)])
               if "HARVEST FLOOR" in m]
        assert out and "rounding-scale" not in out[0], out

    def test_operator_scripted_window_weeks_are_excluded_and_say_so(self):
        """The planner neither chose nor can fix those weeks, and their
        harvests are stitched in separately — counting them would report a
        phantom zero."""
        from forecast.time_grid import iso_week_label
        wk = iso_week_label(_MON)
        out = self._run(harvests=[self._Ev(_MON, 1.0)], window={wk})
        assert not [m for m in out if m.startswith(f"HARVEST FLOOR - {wk}:")], out
        assert any("excluded from this audit" in m for m in out), out

    # ---- harvest ceiling --------------------------------------------------
    def test_ceiling_alarm_rings_when_the_week_exceeds_the_processing_limit(self):
        out = self._run(harvests=[self._Ev(_MON, 70_000.0)])
        assert any("HARVEST CEILING" in m for m in out), out

    def test_ceiling_alarm_stays_quiet_at_the_limit(self):
        out = self._run(harvests=[self._Ev(_MON, 60_000.0)])
        assert not [m for m in out if "HARVEST CEILING" in m], out

    # ---- handling budget --------------------------------------------------
    def _move(self, d, pairs, applied=True):
        from forecast.events import Transfer
        ev = Transfer(batch_id="B1", event_date=d, source_tank_id=pairs[0][0],
                      destinations=[self._Dest(t, 100.0) for _, t in pairs])
        ev.count_transferred = 100.0 if applied else 0.0
        return ev

    def test_handling_alarm_rings_when_the_week_exceeds_the_budget(self):
        moves = [self._move(_MON, [(1, 100 + i)]) for i in range(16)]
        out = self._run(harvests=[self._Ev(_MON, 30_000.0)], transfers=moves,
                        control=self._control(moves=15))
        assert any("HANDLING BUDGET" in m and "16 moves" in m for m in out), out

    def test_handling_alarm_stays_quiet_at_the_budget(self):
        moves = [self._move(_MON, [(1, 100 + i)]) for i in range(15)]
        out = self._run(harvests=[self._Ev(_MON, 30_000.0)], transfers=moves,
                        control=self._control(moves=15))
        assert not [m for m in out if "HANDLING BUDGET" in m], out

    def test_refused_transfers_are_not_handling(self):
        """count_transferred == 0 means the move never happened; counting it
        would invent a breach out of a refusal."""
        moves = [self._move(_MON, [(1, 100 + i)], applied=(i < 10))
                 for i in range(20)]
        out = self._run(harvests=[self._Ev(_MON, 30_000.0)], transfers=moves,
                        control=self._control(moves=15))
        assert not [m for m in out if "HANDLING BUDGET" in m], out

    def test_a_multi_destination_move_counts_once_per_destination(self):
        """The unit must match what placement clamps to — distinct applied
        (source, dest) pairs. Counting EVENTS would report a different number
        than the budget was enforced against."""
        moves = [self._move(_MON, [(1, 200 + i) for i in range(16)])]
        out = self._run(harvests=[self._Ev(_MON, 30_000.0)], transfers=moves,
                        control=self._control(moves=15))
        assert any("HANDLING BUDGET" in m and "16 moves" in m for m in out), out
