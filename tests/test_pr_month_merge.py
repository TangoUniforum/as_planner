"""A mid-month ProductionReport must complete its own month.

The PR's closing date is the day BEFORE forecast_start. When it falls mid-month
the month is split across two sources — the days the PR already reported, and
the forecast that picks up the next day — and reporting only the forecast half
understates the month badly. Measured on the 8.13 PR: August showed 70,444 of
its 134,289 harvested fish, i.e. 48% of the real tonnage, on the two sheets
sales planning actually reads.

Operator rule (2026-08-18): merge ONLY when the PR closes mid-month. A PR
closing on a month's LAST day needs nothing — the forecast then starts on the
1st and already covers the whole month.

Scope is the reporting layer. The audits deliberately never see this: they
prove the FORECAST conserves, and feeding actuals into them would break their
identities and mask real defects.
"""
from __future__ import annotations

from datetime import date

import openpyxl

from forecast import excel_io
from forecast.placement import BatchLocationRow
from forecast.production_report import PRBatchPeriod, PRPeriod

_W1 = date(2026, 8, 10)          # 2026-W33  } both inside August, so the
_W2 = date(2026, 8, 17)          # 2026-W34  } month has an open and a close


def _loc(week, wkstart, batch, tank, count, wt_g, bio):
    return BatchLocationRow(
        week_label=week, week_start=wkstart, batch_id=batch, tank_id=tank,
        location_id=f"OG3-{tank}", system_id="OG3", count=count,
        avg_wt_g=wt_g, biomass_kg=bio, density_kg_m3=50.0)


_LOCS = [
    _loc("2026-W33", _W1, "B1", 40, 100_000, 4000.0, 400_000.0),
    _loc("2026-W34", _W2, "B1", 40,  99_000, 4100.0, 405_900.0),
]


def _period(closing, harv=20_000.0, harv_kg=80_000.0, feed=50_000.0):
    return PRPeriod(
        closing_date=closing,
        batches={"B1": PRBatchPeriod(
            batch_id="B1", open_count=130_000.0, open_bio_kg=470_000.0,
            open_avg_wt_g=3615.0, growth_kg=30_000.0, feed_kg=feed,
            harv_count=harv, harv_gross_kg=harv_kg,
            mort_count=500.0, mort_bio_kg=1_800.0,
            cull_count=0.0, cull_bio_kg=0.0)})


def _monthly(pr_period):
    wb = openpyxl.Workbook()
    excel_io.write_monthly_report(wb, _LOCS, [], [], pr_period=pr_period)
    ws = wb["MonthlyReport"]
    rows = [r for r in ws.iter_rows(values_only=True)]
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "Scenario")
    hdr = [str(c) if c else "" for c in rows[hi]]
    # NB: filter on Batch, not on the row's first cell — scenario_name is ""
    # here, so a truthiness test on column A silently drops every data row.
    return [d for d in (dict(zip(hdr, r)) for r in rows[hi + 1:])
            if d.get("Batch")]


def _num(row, key):
    try:
        return float(row.get(key) or 0.0)
    except (TypeError, ValueError):
        return 0.0


class TestTheMonthEndRule:
    def test_mid_month_close_needs_the_merge(self):
        assert _period(date(2026, 8, 13)).is_mid_month is True
        assert _period(date(2026, 2, 27)).is_mid_month is True

    def test_last_day_of_month_needs_nothing(self):
        """The forecast starts on the 1st, so the month is already whole."""
        for d in (date(2026, 8, 31), date(2026, 2, 28), date(2026, 12, 31),
                  date(2026, 4, 30)):
            assert _period(d).is_mid_month is False, d

    def test_leap_year_february(self):
        assert _period(date(2028, 2, 29)).is_mid_month is False
        assert _period(date(2028, 2, 28)).is_mid_month is True


class TestTheMergeItself:
    def test_the_elapsed_harvest_and_feed_land_in_the_month(self):
        base = _monthly(None)
        merged = _monthly(_period(date(2026, 8, 13)))
        b = next(r for r in base if r["Batch"] == "B1")
        m = next(r for r in merged if r["Batch"] == "B1")
        assert _num(m, "Harv_Count (fish)") - _num(b, "Harv_Count (fish)") == 20_000
        assert _num(m, "Harv_Gross (kg)") - _num(b, "Harv_Gross (kg)") == 80_000
        assert _num(m, "Feed (kg)") - _num(b, "Feed (kg)") == 50_000

    def test_the_month_opens_where_the_PR_opened(self):
        """Not where the forecast picked up mid-month — otherwise the row shows
        a full month of flows against a half month's opening."""
        m = next(r for r in _monthly(_period(date(2026, 8, 13)))
                 if r["Batch"] == "B1")
        assert _num(m, "Open_Count (fish)") == 130_000

    def test_a_month_end_PR_changes_nothing(self):
        base = _monthly(None)
        same = _monthly(_period(date(2026, 8, 31)))
        assert base == same

    def test_count_check_stays_true_to_the_columns_shown(self):
        """The PR carries its own 'Deviation count in period', which is not a
        fish flow and has no column here. Count_Check must surface it rather
        than keep reporting the pre-merge residual — a column whose job is to
        expose discrepancies must not hide one."""
        m = next(r for r in _monthly(_period(date(2026, 8, 13)))
                 if r["Batch"] == "B1")
        residual = (_num(m, "Open_Count (fish)") - _num(m, "Mort_Count (fish)")
                    - _num(m, "Harv_Count (fish)") - _num(m, "Cull_Count (fish)")
                    + _num(m, "Input_Count (fish)") + _num(m, "Xfer_In (fish)")
                    - _num(m, "Xfer_Out (fish)") - _num(m, "Close_Count (fish)"))
        assert abs(residual - _num(m, "Count_Check (fish)")) <= 1.0

    def test_a_batch_the_forecast_never_carries_still_appears(self):
        """Harvested out before the forecast starts — its landings are real and
        must not vanish from the month."""
        p = _period(date(2026, 8, 13))
        p.batches["B_GONE"] = PRBatchPeriod(
            batch_id="B_GONE", open_count=9_000.0, open_bio_kg=36_000.0,
            open_avg_wt_g=4000.0, growth_kg=0.0, feed_kg=1_000.0,
            harv_count=9_000.0, harv_gross_kg=36_000.0,
            mort_count=0.0, mort_bio_kg=0.0, cull_count=0.0, cull_bio_kg=0.0)
        rows = _monthly(p)
        gone = [r for r in rows if r["Batch"] == "B_GONE"]
        assert len(gone) == 1
        assert _num(gone[0], "Harv_Count (fish)") == 9_000


class TestHarvestPlanReport:
    def _hp(self, pr_period):
        wb = openpyxl.Workbook()
        excel_io.write_harvest_plan_report(
            wb, [], scenario_name="Forecast", default_hog_yield=0.81,
            facility_limits_hog={}, pr_period=pr_period)
        return [r for r in wb["HarvestPlan Report"].iter_rows(values_only=True)]

    def test_elapsed_landings_reach_the_sales_sheet(self):
        rows = self._hp(_period(date(2026, 8, 13)))
        flat = [c for r in rows for c in r if c is not None]
        assert any(isinstance(c, (int, float)) and abs(c - 20_000) < 1
                   for c in flat), "the PR's 20,000 landed fish are missing"

    def test_month_end_PR_adds_nothing(self):
        assert self._hp(_period(date(2026, 8, 31))) == self._hp(None)
