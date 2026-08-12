"""Weekly handling budget (operator rule 4): max 15 transfer moves/week.

Deferrable quality passes stop at the cap (their work waits for a calmer
week); essential moves are never blocked — an essential-only overrun is
reported by the handling gate, not silently truncated.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from forecast.analysis import _gate_handling_budget
from forecast.events import TankAllocation, Transfer
from forecast.excel_io import write_transfer_plan_output
from forecast.models import ControlParams
from forecast.optimize import _weekly_move_counts
from forecast import placement as _pl
from forecast.placement import (
    _consolidate_remnants,
    _emit_transfers_for_batch_diff,
    _entry_makeroom_move_cost,
    _even_out_density,
    _pacing_may_defer,
    _quality_moves_left,
)
from forecast.state import FacilityState, TankState

TODAY = date(2026, 8, 3)


def _mk_state():
    return FacilityState(TODAY, [
        TankState("OG3N-31", 31, "OG3N", 1000.0, 95.0, 1000.0, "OG"),
        TankState("OG3N-32", 32, "OG3N", 1000.0, 95.0, 1000.0, "OG"),
        TankState("OG3N-33", 33, "OG3N", 1000.0, 95.0, 1000.0, "OG"),
        TankState("OG4N-41", 41, "OG4N", 1000.0, 95.0, 1000.0, "OG"),
        TankState("OG4N-42", 42, "OG4N", 1000.0, 95.0, 1000.0, "OG"),
    ])


class TestGate:
    def test_truth_table(self):
        assert _gate_handling_budget({"weeks_moves_over_cap": 0,
                                      "weeks_moves_warn": 0})[0] == "PASS"
        s, d = _gate_handling_budget({"weeks_moves_over_cap": 0,
                                      "weeks_moves_warn": 3,
                                      "moves_week_max": 14})
        assert s == "WARN" and "14" in d
        s, d = _gate_handling_budget({"weeks_moves_over_cap": 2,
                                      "weeks_moves_warn": 5,
                                      "moves_week_max": 19})
        assert s == "FAIL" and "19" in d
        assert _gate_handling_budget({})[0] == "N/A"

    def test_knob_default(self):
        assert ControlParams.__dataclass_fields__[
            "max_transfers_per_week"].default == 15


class TestEvenOutBudget:
    def _crowded(self):
        s = _mk_state()
        # over-cap source (25,001 fish @ 3.8kg in 1000 m3 = 95.004) + two
        # near-empty same-batch destinations -> would emit 2 equalize moves
        s.tanks_by_id[31].assign("B50", 25100, 3800.0, 16.0, "SW")
        s.tanks_by_id[32].assign("B50", 8000, 3800.0, 16.0, "SW")
        s.tanks_by_id[33].assign("B50", 8000, 3800.0, 16.0, "SW")
        return s

    def test_unlimited_emits_all(self):
        s = self._crowded()
        evs, warns = [], []
        _even_out_density(s, "B50", TODAY, evs, warns)
        assert len([e for e in evs if e.count_transferred > 0]) == 2

    def test_budget_one_stops_after_one(self):
        s = self._crowded()
        evs, warns = [], []
        _even_out_density(s, "B50", TODAY, evs, warns, max_moves=1)
        assert len([e for e in evs if e.count_transferred > 0]) == 1

    def test_budget_zero_emits_none(self):
        s = self._crowded()
        evs, warns = [], []
        _even_out_density(s, "B50", TODAY, evs, warns, max_moves=0)
        assert evs == []


class TestRemnantSweepBudget:
    def _with_remnants(self):
        s = _mk_state()
        # two independent remnants, each with an absorbing same-batch tank
        s.tanks_by_id[31].assign("B50", 500, 3800.0, 16.0, "SW")
        s.tanks_by_id[32].assign("B50", 20000, 3800.0, 16.0, "SW")
        s.tanks_by_id[41].assign("B51", 600, 3800.0, 16.0, "SW")
        s.tanks_by_id[42].assign("B51", 20000, 3800.0, 16.0, "SW")
        return s

    def test_unlimited_folds_both(self):
        s = self._with_remnants()
        evs, warns = [], []
        folds = _consolidate_remnants(s, TODAY, "2026-W40", evs, warns, 7000.0)
        assert folds == 2

    def test_budget_one_folds_one_and_defers(self):
        s = self._with_remnants()
        evs, warns = [], []
        folds = _consolidate_remnants(s, TODAY, "2026-W40", evs, warns, 7000.0,
                                      max_moves=1)
        assert folds == 1
        # the deferred remnant is intact, not half-moved
        left = [t for t in (s.tanks_by_id[31], s.tanks_by_id[41])
                if not t.is_empty]
        assert len(left) == 1 and left[0].count in (500, 600)


class TestDiffBudget:
    """Plan-diff: SOURCE drains are essential (never blocked); kept-tank
    EVENING top-ups are quality moves that yield to the handling budget."""

    def test_source_drain_never_blocked(self):
        s = _mk_state()
        s.tanks_by_id[31].assign("B50", 20000, 3800.0, 16.0, "SW")
        s.tanks_by_id[32].assign("B50", 20000, 3800.0, 16.0, "SW")
        evs, warns = [], []
        # 31 leaves the plan (source), 41 joins (new dest). Budget exhausted.
        _emit_transfers_for_batch_diff(
            s, "B50", {31, 32}, {32, 41}, TODAY, evs, warns,
            moves_left=lambda: 0)
        applied = [e for e in evs if e.count_transferred > 0]
        assert applied, "essential source drain was blocked by the budget"
        assert all(e.source_tank_id == 31 for e in applied)
        assert s.tanks_by_id[31].is_empty

    def test_kept_evening_yields_to_budget(self):
        s = _mk_state()
        for tid in (31, 32, 33):
            s.tanks_by_id[tid].assign("B50", 21000, 3800.0, 16.0, "SW")
        evs, warns = [], []
        # No sources; batch gains tank 41 -> the 3 kept tanks would each top
        # it up (quality evening). Budget spent -> nothing moves.
        _emit_transfers_for_batch_diff(
            s, "B50", {31, 32, 33}, {31, 32, 33, 41}, TODAY, evs, warns,
            moves_left=lambda: 0)
        assert [e for e in evs if e.count_transferred > 0] == []
        assert s.tanks_by_id[41].is_empty

    def test_kept_evening_partial_budget(self):
        s = _mk_state()
        for tid in (31, 32, 33):
            s.tanks_by_id[tid].assign("B50", 21000, 3800.0, 16.0, "SW")
        evs, warns = [], []
        _emit_transfers_for_batch_diff(
            s, "B50", {31, 32, 33}, {31, 32, 33, 41}, TODAY, evs, warns,
            moves_left=lambda: max(0, 1 - len(evs)))
        assert len([e for e in evs if e.count_transferred > 0]) == 1

    def test_no_budget_arg_is_unlimited(self):
        s = _mk_state()
        for tid in (31, 32, 33):
            s.tanks_by_id[tid].assign("B50", 21000, 3800.0, 16.0, "SW")
        evs, warns = [], []
        _emit_transfers_for_batch_diff(
            s, "B50", {31, 32, 33}, {31, 32, 33, 41}, TODAY, evs, warns)
        assert len([e for e in evs if e.count_transferred > 0]) == 3
        assert not s.tanks_by_id[41].is_empty


class TestEntryMakeroomReserve:
    """ANTICIPATORY RESERVE, layer A. The arrival-week entry-tier make-room
    runs AFTER the deferrable quality passes, so its cost is priced up front
    and held back from them. Pricing must be exact: over-reserving starves the
    leveling for nothing, under-reserving lets the week go over cap again."""

    def test_no_congestion_reserves_nothing(self):
        # NEGATIVE CONTROL: the entry tier already has room for the cohort —
        # nothing to vacate, so quality keeps its full budget.
        assert _entry_makeroom_move_cost(
            need_tanks=2, empty_entry=2, free_growout=0, vacatable_entry=6) == 0
        assert _entry_makeroom_move_cost(
            need_tanks=2, empty_entry=5, free_growout=0, vacatable_entry=6) == 0

    def test_no_arrival_at_all_reserves_nothing(self):
        assert _entry_makeroom_move_cost(0, 0, 0, 6) == 0

    def test_one_vacate_per_short_tank_when_growout_is_free(self):
        # 2 tanks short, 2 grow-out slots standing free -> 2 forward vacates.
        assert _entry_makeroom_move_cost(
            need_tanks=2, empty_entry=0, free_growout=2, vacatable_entry=6) == 2

    def test_no_free_growout_doubles_the_cost(self):
        # Each vacate must first free a grow-out slot (a 6N purge move-in),
        # so a short tank costs TWO moves, not one.
        assert _entry_makeroom_move_cost(
            need_tanks=2, empty_entry=0, free_growout=0, vacatable_entry=6) == 4

    def test_partial_growout_headroom_is_consumed_in_order(self):
        # 3 short, only 1 grow-out slot free: 1 cheap vacate + 2 expensive.
        assert _entry_makeroom_move_cost(
            need_tanks=3, empty_entry=0, free_growout=1, vacatable_entry=6) == 5

    def test_never_reserves_for_work_it_cannot_do(self):
        # 4 tanks short but only ONE entry occupant is movable -> price 1
        # vacate, not 4. (A depurating entry tank is not vacatable.)
        assert _entry_makeroom_move_cost(
            need_tanks=4, empty_entry=0, free_growout=9, vacatable_entry=1) == 1
        assert _entry_makeroom_move_cost(
            need_tanks=4, empty_entry=0, free_growout=9, vacatable_entry=0) == 0

    def test_reserve_is_never_negative(self):
        assert _entry_makeroom_move_cost(1, 9, 0, 9) == 0

    def test_matches_the_measured_2026_w43_shape(self):
        # The week that motivated this: 7.29.26 PR, 2026-W43 emitted 17 moves
        # = 7 essential + 10 quality, and 2 of those essential moves were the
        # entry-tier vacates that ran after quality had spent the budget.
        # With 2 entry tanks short and grow-out slots free, the reserve is
        # exactly those 2 moves — quality is capped at 8 and the week lands
        # on 15.
        assert _entry_makeroom_move_cost(
            need_tanks=2, empty_entry=0, free_growout=4, vacatable_entry=8) == 2


class TestPacingDeferral:
    """ANTICIPATORY RESERVE, layer B, WITH THE SWITCH ON. The purge-pacing pass
    walks a multi-week lookahead, so in a week that is already at budget it
    stands down and a calmer week inside the window pre-frees the tank instead.
    Work moves EARLIER/LATER — it is never refused.

    The switch ships OFF (see TestAnticipatoryLayerSwitches for why); these
    tests pin the policy so re-enabling it is a one-line, covered change."""

    @pytest.fixture(autouse=True)
    def _layer_b_on(self, monkeypatch):
        monkeypatch.setattr(_pl, "_ANTICIPATE_PACING_DEFER", True)

    def test_defers_when_at_budget_and_the_window_has_slack(self):
        assert _pacing_may_defer(weeks_out=4, moves_left=0) is True
        assert _pacing_may_defer(weeks_out=2, moves_left=0) is True

    def test_does_not_defer_with_budget_to_spare(self):
        # NEGATIVE CONTROL: no congestion -> the pass behaves exactly as before.
        assert _pacing_may_defer(weeks_out=4, moves_left=1) is False
        assert _pacing_may_defer(weeks_out=4, moves_left=15) is False

    def test_never_defers_the_last_chance(self):
        # An arrival landing NEXT week pre-stages regardless of the budget:
        # essential work is never blocked, only re-timed.
        assert _pacing_may_defer(weeks_out=1, moves_left=0) is False
        assert _pacing_may_defer(weeks_out=0, moves_left=0) is False

    def test_over_budget_week_still_defers(self):
        # _moves_left() floors at 0, but be robust if a caller passes negative.
        assert _pacing_may_defer(weeks_out=3, moves_left=-2) is True


class TestQualityReserveWithSwitchOn:
    """ANTICIPATORY RESERVE, layer A, WITH THE SWITCH ON: the deferrable
    quality passes get the remaining budget MINUS the priced make-room."""

    @pytest.fixture(autouse=True)
    def _layer_a_on(self, monkeypatch):
        monkeypatch.setattr(_pl, "_ANTICIPATE_ARRIVAL_RESERVE", True)

    def test_reserve_is_subtracted(self):
        assert _quality_moves_left(moves_left=10, reserve=2, move_cap=15) == 8

    def test_zero_reserve_leaves_the_budget_whole(self):
        assert _quality_moves_left(moves_left=10, reserve=0, move_cap=15) == 10

    def test_never_goes_negative(self):
        assert _quality_moves_left(moves_left=1, reserve=9, move_cap=15) == 0

    def test_reserve_is_clamped_to_the_whole_budget(self):
        # A week whose essential work alone blows the cap is a capacity signal
        # the handling gate must report — not a reason to over-subtract.
        assert _quality_moves_left(moves_left=15, reserve=99, move_cap=15) == 0

    def test_budget_off_is_untouched(self):
        # move_cap <= 0 means the handling budget is disabled entirely.
        assert _quality_moves_left(moves_left=10 ** 9, reserve=4,
                                   move_cap=0) == 10 ** 9


class TestAnticipatoryLayerSwitches:
    """THE SHIPPED DECISION: both anticipatory layers are OFF.

    Measured by a 4-arm (both off / A only / B only / both on) x 3-PR x
    2-knob-set ablation, with the both-off arm verified against a physically
    pre-layer placement.py. Layer A buys the 15-move budget compliance and pays
    for it out of the 30,000/week harvest floor — on the operator's own PR it
    moves weeks-under-floor 3 -> 5 and more than doubles the shortfall, and on
    7.9 with tuned knobs it produces a 69,677-fish week, past the 60,500 relief
    ceiling. Layer B is plan-identical to both-off on 2 of the 3 PRs. The
    operator's rule order (steady harvest HARD, handling flexible) makes that
    trade a loss, so both ship off — and these tests hold that line, because a
    silent flip back would cost the contract floor."""

    def test_both_layers_ship_off(self):
        assert _pl._ANTICIPATE_ARRIVAL_RESERVE is False
        assert _pl._ANTICIPATE_PACING_DEFER is False

    def test_layer_a_off_is_the_identity_on_the_budget(self):
        # Quality spends the whole remaining budget — exactly the pre-layer
        # behaviour, whatever the make-room would have cost.
        for reserve in (0, 1, 5, 99):
            assert _quality_moves_left(
                moves_left=10, reserve=reserve, move_cap=15) == 10

    def test_layer_b_off_never_defers(self):
        # Constantly False, including the cases the ON policy would defer.
        assert _pacing_may_defer(weeks_out=4, moves_left=0) is False
        assert _pacing_may_defer(weeks_out=2, moves_left=0) is False
        assert _pacing_may_defer(weeks_out=3, moves_left=-2) is False

    def test_switches_are_not_operator_config(self):
        # They are an engineering result, not a knob: no control.yaml key and
        # no ControlParams field may shadow them (an operator-flippable layer
        # would put the harvest floor back at risk from the app).
        cfg_text = (Path(__file__).resolve().parents[1]
                    / "config" / "control.yaml").read_text()
        for name in ("anticipate_arrival_reserve", "anticipate_pacing_defer",
                     "handling_reserve", "pacing_defer"):
            assert name not in cfg_text
            assert not hasattr(ControlParams, name)


class TestReserveOnlySubtracts:
    """The whole mechanism can only REDUCE what a deferrable pass may emit —
    it never authorises a move. That is why it cannot introduce a topology
    violation, a remnant, or a second batch in a tank: it emits nothing."""

    def test_cost_is_monotone_in_the_deficit(self):
        prev = -1
        for need in range(0, 8):
            c = _entry_makeroom_move_cost(need, 0, 2, 8)
            assert c >= prev, "reserve must not shrink as the deficit grows"
            prev = c

    def test_cost_is_non_negative_over_a_grid(self):
        for need in range(0, 6):
            for empty in range(0, 6):
                for free in range(0, 6):
                    for vac in range(0, 6):
                        c = _entry_makeroom_move_cost(need, empty, free, vac)
                        assert c >= 0
                        # never prices more than 2 moves per vacatable tank
                        assert c <= 2 * vac


class TestTransferPlanRowHonesty:
    """A TransferPlan 'Transfer' row = one real src->dst tank move: 0-fish
    float-residue legs are dropped and same-week duplicate (batch, src, dst)
    legs merge into one row — the handling gate counts these rows."""

    def _events(self):
        e1 = Transfer(
            batch_id="B50", event_date=TODAY, source_tank_id=31,
            destinations=[
                TankAllocation(tank_id=41, count=0.3, avg_wt_g=3800.0,
                               cv_pct=16.0),
                TankAllocation(tank_id=42, count=5000, avg_wt_g=3800.0,
                               cv_pct=16.0),
            ])
        e1.count_transferred = 5000.3
        e2 = Transfer(
            batch_id="B50", event_date=TODAY, source_tank_id=32,
            destinations=[TankAllocation(tank_id=41, count=1000,
                                         avg_wt_g=3600.0, cv_pct=16.0)])
        e2.count_transferred = 1000
        e3 = Transfer(
            batch_id="B50", event_date=TODAY, source_tank_id=32,
            destinations=[TankAllocation(tank_id=41, count=2000,
                                         avg_wt_g=3900.0, cv_pct=16.0)])
        e3.count_transferred = 2000
        return [e1, e2, e3]

    def _rows(self, wb):
        ws = wb["TransferPlan"]
        rows = list(ws.iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows) if r and r[0] == "Week")
        return [dict(zip(rows[hi], r)) for r in rows[hi + 1:]
                if r and r[0]]

    def test_zero_leg_dropped_and_duplicates_merged(self):
        wb = openpyxl.Workbook()
        write_transfer_plan_output(wb, self._events(), [], [])
        rows = self._rows(wb)
        assert len(rows) == 2
        moves = {(r["From_Tank"], r["To_Tank"]): r for r in rows}
        assert ("31", 42) in moves            # the real leg survives
        assert ("31", 41) not in moves        # 0.3-fish leg is not a move
        merged = moves[("32", 41)]
        assert merged["Count (fish)"] == 3000
        # fish-weighted avg weight: (1000*3.6 + 2000*3.9) / 3000 = 3.8
        assert abs(merged["Avg_Weight (kg)"] - 3.8) < 1e-6

    def test_gate_counter_matches_rows(self):
        wb = openpyxl.Workbook()
        write_transfer_plan_output(wb, self._events(), [], [])
        assert _weekly_move_counts(wb) == [2]

    def test_gate_counter_skips_zero_rows_in_legacy_workbooks(self):
        # A workbook written before the writer fix may carry 0-fish rows.
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("TransferPlan")
        ws.append(["Week", "Batch", "Type", "From_Tank", "To_Tank",
                   "Count (fish)", "Avg_Weight (kg)", "Grade", "CV (%)"])
        ws.append(["2026-W40", "B50", "Transfer", "31", 41, 0, 3.8, None, 16])
        ws.append(["2026-W40", "B50", "Transfer", "31", 42, 5000, 3.8, None, 16])
        assert _weekly_move_counts(wb) == [1]
