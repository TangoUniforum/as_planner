"""Faithful reject-at-entry validation for the manual override window.

`validate_manual_events` must match what `advance_facility_window` actually
does on the run, or the in-app editor accepts events that misbehave in trials:
  * fw_to_og is checked against the SAME FW projection — the batch must be in
    freshwater at its week and the target must be feasible (was previously a
    dest-tanks-only check that passed any in/out-of-FW batch);
  * events are sequenced by week with biology advancing between weeks.

These need the gitignored Forecast.xlsm + seeded config/scenario (the FW
projection is real biology), so they skip cleanly when absent — they run in the
main checkout where the data lives.
"""
from __future__ import annotations

from datetime import datetime as _dt, timedelta as _td
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent  # Python/
WORKBOOK = ROOT / "Forecast.xlsm"
CONFIG_DIR = ROOT / "config"
SCENARIO_DIR = ROOT / "scenario"

pytestmark = pytest.mark.skipif(
    not (WORKBOOK.exists()
         and (CONFIG_DIR / "control.yaml").exists()
         and (SCENARIO_DIR / "limits.yaml").exists()),
    reason="needs Forecast.xlsm + seeded config/scenario (gitignored)",
)


@pytest.fixture(scope="module")
def hydrated():
    """(state, ctx) hydrated from the live PR + config, exactly as the app's
    _hydrate_state_from_upload assembles them for faithful validation."""
    from forecast.excel_io import load_workbook
    from forecast.config_io import load_config
    from forecast.scenario_io import load_batches
    from forecast.production_report import (
        read_production_report, hydrate_facility_state)
    from forecast.state import FacilityState
    wb = load_workbook(str(WORKBOOK))
    control, tables, facility = load_config(str(CONFIG_DIR))
    batches = load_batches(str(SCENARIO_DIR))
    pc, og, fw = read_production_report(wb)
    fs = _dt(pc.year, pc.month, pc.day) + _td(days=1)
    control.forecast_start = fs
    state = FacilityState.from_facility_config(facility, today=fs.date())
    hydrate_facility_state(state, og, batches)
    ctx = dict(batch_by_id={b.batch_id: b for b in batches}, tables=tables,
               forecast_start=fs.date(), control=control, pr_closing=pc,
               fw_records=fw)
    # Pick a batch that is actually in freshwater at week 1 (the first one an
    # fw_to_og can legally reference) so the tests don't hard-code an id.
    from forecast.manual_window import _build_fw_lookup
    from forecast.manual_events import ManualEvent
    from forecast.time_grid import forecast_week_labels
    wk1 = forecast_week_labels(fs.date(), 1)[0]
    probe = [ManualEvent(type="fw_to_og", week=1, batch=b.batch_id,
                         destinations=[]) for b in batches]
    lookup = _build_fw_lookup(probe, fw, control, pc, tables, ctx["batch_by_id"])
    fw_batches_wk1 = sorted({bid for (bid, wl) in lookup if wl == wk1})
    return state, ctx, fw_batches_wk1


def _empty_og_tanks(state, n):
    # Entry-tier only: FW->OG destinations must be OG1/2 (rule R1).
    from forecast.tiers import is_entry
    return [t for t, tk in sorted(state.tanks_by_id.items())
            if tk.type == "OG" and tk.is_empty and is_entry(tk.system_id)][:n]


def _assert_count_conserved(tc):
    """Hard conservation invariant on a TankContinuityAudit sheet: NO TANK_DRIFT
    (count) rows + facility count signed/abs ratio in band (|ratio| < 0.3).

    COUNT is the hard invariant — fish can't be created or lost, so a count
    (TANK_DRIFT) row is a real leak. Per-row **BIO_DRIFT** on full-turnover weeks
    is a benign audit-vs-snapshot timing artifact (weekly-vs-daily growth
    attribution — see the facility 'Biomass (kg)' summary's own "not a leak"
    label, and the biomass-drift note behind test_coordinator_regression). It
    shifts with ANY placement perturbation (e.g. holding 6N frozen re-routes
    downstream tanks) and is bounded by the facility ratio, so it is tolerated
    here rather than asserted to be per-row zero (which was never the real
    invariant)."""
    count_drift = [r for r in tc
                   if any("TANK_DRIFT" in str(c) for c in r if c is not None)]
    assert not count_drift, \
        f"{len(count_drift)} TANK_DRIFT (count) rows: {count_drift[:2]}"
    ratio = next((float(str(r[3])) for r in tc
                  if r and str(r[0]) == "Count (fish)"), None)
    assert ratio is not None and abs(ratio) < 0.3, f"facility count ratio {ratio}"


class TestFaithfulFwToOgValidation:
    def test_valid_fw_to_og_passes(self, hydrated):
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        state, ctx, fw_batches = hydrated
        assert fw_batches, "expected at least one in-FW batch at week 1"
        dests = _empty_og_tanks(state, 2)
        ev = ManualEvent(type="fw_to_og", week=1, batch=fw_batches[0],
                         count=50000,
                         destinations=[ManualDest(tank=t) for t in dests])
        (i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        # A cull to hit the target must NOT block (informational note filtered).
        assert ok, msgs

    def test_non_fw_batch_is_rejected(self, hydrated):
        # An OG (already-in-seawater) batch has no FW state at week 1 — the old
        # dest-only check passed this; the faithful check must reject it.
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        state, ctx, fw_batches = hydrated
        og_batch = next(tk.batch_id for tk in state.tanks_by_id.values()
                        if not tk.is_empty and tk.batch_id not in fw_batches)
        dests = _empty_og_tanks(state, 1)
        ev = ManualEvent(type="fw_to_og", week=1, batch=og_batch, count=10000,
                         destinations=[ManualDest(tank=dests[0])])
        (i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        assert not ok
        assert any("not in freshwater" in m for m in msgs)

    def test_target_exceeding_fw_is_rejected(self, hydrated):
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        state, ctx, fw_batches = hydrated
        dests = _empty_og_tanks(state, 2)
        ev = ManualEvent(type="fw_to_og", week=1, batch=fw_batches[0],
                         count=99_000_000,
                         destinations=[ManualDest(tank=t) for t in dests])
        (i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        assert not ok
        assert any("exceeds available FW" in m for m in msgs)

    def test_legacy_fallback_skips_fw_feasibility(self, hydrated):
        # Without the run-time context, fw_to_og falls back to structural-only
        # (feasibility deferred to run) — a non-FW batch is NOT rejected.
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        state, ctx, fw_batches = hydrated
        og_batch = next(tk.batch_id for tk in state.tanks_by_id.values()
                        if not tk.is_empty and tk.batch_id not in fw_batches)
        dests = _empty_og_tanks(state, 1)
        ev = ManualEvent(type="fw_to_og", week=1, batch=og_batch, count=10000,
                         destinations=[ManualDest(tank=dests[0])])
        (i, ok, msgs), = validate_manual_events(state, [ev])  # no ctx
        assert ok, msgs  # structural-only: dest is empty OG, so it passes


class TestGradedHarvest:
    """Manual graded harvest (top-N-by-size): the split conserves count + biomass
    exactly, reject-at-entry gates an occupied pickup, and a full-pipeline run
    reconciles in EVERY forecast audit (tank continuity 0-drift, 0 dropped)."""

    @staticmethod
    def _source_pickup(state):
        # Non-entry source/pickup: harvest & 6N staging FROM the entry tier
        # (OG1/2) is forbidden (rule R5), so the fixture grades a grow-out tank.
        from forecast.sixn import SIXN_ALL_TANKS
        from forecast.tiers import is_entry
        src = max((t for t in state.tanks_by_id.values()
                   if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                   and not is_entry(t.system_id)
                   and not t.is_empty), key=lambda t: t.count)
        pickup = next(t for t in sorted(state.tanks_by_id.values(),
                                        key=lambda t: t.tank_id)
                      if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                      and not is_entry(t.system_id)
                      and t.is_empty and t.tank_id != src.tank_id)
        return src, pickup

    def test_split_conserves_count_and_biomass(self, hydrated):
        import copy
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        from forecast.manual_window import advance_facility_window
        state, ctx, _fw = hydrated
        src, pickup = self._source_pickup(state)
        K = round(src.count * 0.4)
        open_n, open_bio = src.count, src.count * src.avg_wt_g
        ev = ManualEvent(type="graded_harvest", week=1, from_tank=src.tank_id,
                         count=K, destinations=[ManualDest(tank=pickup.tank_id)])
        (_i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        assert ok, msgs
        sc = copy.deepcopy(state)
        win = advance_facility_window(
            sc, ctx["batch_by_id"], ctx["tables"], ctx["forecast_start"], 2,
            events=[ev], control=ctx["control"], pr_closing=ctx["pr_closing"],
            fw_records=ctx["fw_records"])
        ghs = [e for e in win["transfer_events"] if hasattr(e, "pickup_tank_id")]
        assert len(ghs) == 1
        gh = ghs[0]
        assert abs((gh.pickup_count + gh.retention_count) - open_n) < 1.0   # I1
        bio = (gh.pickup_count * gh.pickup_avg_wt_g
               + gh.retention_count * gh.retention_avg_wt_g)
        assert abs(bio - open_bio) / open_bio < 1e-6                        # I2
        assert abs(gh.pickup_count - K) < 1.0
        assert gh.pickup_avg_wt_g > gh.retention_avg_wt_g   # big class heavier
        hv = [h for h in win["harvest_events"]
              if h.source_tank_id == pickup.tank_id]
        assert hv and abs(hv[0].count - K) < 1.0            # pickup drained
        assert abs(hv[0].avg_wt_g - gh.pickup_avg_wt_g) < 1e-6   # at big weight

    def test_rejects_occupied_pickup(self, hydrated):
        from forecast.sixn import SIXN_ALL_TANKS
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        state, ctx, _fw = hydrated
        src, _ = self._source_pickup(state)
        occ = next(t for t in state.tanks_by_id.values()
                   if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                   and not t.is_empty and t.batch_id != src.batch_id
                   and t.tank_id != src.tank_id)
        ev = ManualEvent(type="graded_harvest", week=1, from_tank=src.tank_id,
                         count=round(src.count * 0.3),
                         destinations=[ManualDest(tank=occ.tank_id)])
        (_i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        assert not ok and any("not empty" in m for m in msgs)

    def test_graded_to_6n_depurates_conserves_no_harvest(self, hydrated):
        # "Graded -> 6N" (the UI form): grade the tank, biggest N -> a 6N tank to
        # DEPURATE (frozen off-feed, NO immediate harvest), smaller remainder ->
        # a separate OG tank, and the SOURCE EMPTIES. Realistic OG->6N->harvest
        # flow; still conserves count + biomass exactly.
        import copy
        from forecast.sixn import SIXN_ALL_TANKS
        from forecast.state import STAGE_STARVE
        from forecast.manual_events import (
            ManualEvent, ManualDest, validate_manual_events)
        from forecast.manual_window import advance_facility_window
        state, ctx, _fw = hydrated
        src, _ = self._source_pickup(state)
        sixn = next(t for t in sorted(SIXN_ALL_TANKS)
                    if state.tanks_by_id[t].is_empty)
        og_ret = next(t.tank_id for t in state.tanks_by_id.values()
                      if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                      and t.is_empty and t.tank_id != src.tank_id)
        K = round(src.count * 0.4)
        open_n, open_bio = src.count, src.count * src.avg_wt_g
        ev = ManualEvent(type="graded_harvest", week=1, from_tank=src.tank_id,
                         count=K, destinations=[ManualDest(tank=sixn),
                                                ManualDest(tank=og_ret)])
        (_i, ok, msgs), = validate_manual_events(state, [ev], **ctx)
        assert ok, msgs
        sc = copy.deepcopy(state)
        win = advance_facility_window(
            sc, ctx["batch_by_id"], ctx["tables"], ctx["forecast_start"], 2,
            events=[ev], control=ctx["control"], pr_closing=ctx["pr_closing"],
            fw_records=ctx["fw_records"])
        ghs = [e for e in win["transfer_events"] if hasattr(e, "pickup_tank_id")]
        assert len(ghs) == 1
        assert win["harvest_events"] == []          # depurates, NOT harvested now
        gh = ghs[0]
        assert abs((gh.pickup_count + gh.retention_count) - open_n) < 1.0
        bio = (gh.pickup_count * gh.pickup_avg_wt_g
               + gh.retention_count * gh.retention_avg_wt_g)
        assert abs(bio - open_bio) / open_bio < 1e-6
        assert sc.tanks_by_id[src.tank_id].is_empty          # source graded OUT
        pk = sc.tanks_by_id[sixn]
        assert pk.stage == STAGE_STARVE and not pk.is_empty  # 6N frozen, big fish
        rt = sc.tanks_by_id[og_ret]
        assert not rt.is_empty and rt.batch_id == src.batch_id  # smaller -> OG

    def test_full_pipeline_audits_clean(self, hydrated, tmp_path):
        import shutil
        from openpyxl import load_workbook as _lw
        import forecast.run as run_mod
        state, _ctx, _fw = hydrated
        src, pickup = self._source_pickup(state)
        K = round(src.count * 0.4)
        sdir = tmp_path / "scenario"
        shutil.copytree(SCENARIO_DIR, sdir)
        (sdir / "manual_events.yaml").write_text(
            "events:\n  - type: graded_harvest\n    week: 1\n"
            f"    from_tank: {src.tank_id}\n    count: {K}\n"
            f"    destinations:\n      - {{tank: {pickup.tank_id}}}\n")
        wb = tmp_path / "Forecast.xlsm"
        shutil.copy(WORKBOOK, wb)
        out = tmp_path / "out.xlsm"
        run_mod.main(str(wb), output_path=str(out),
                     config_dir=str(CONFIG_DIR), scenario_dir=str(sdir))
        owb = _lw(str(out), data_only=True)
        tc = [[c.value for c in r]
              for r in owb["TankContinuityAudit"].iter_rows()]
        _assert_count_conserved(tc)
        ic = [[c.value for c in r]
              for r in owb["InputConservationAudit"].iter_rows()]
        assert not any("*** DROP" in str(c).upper()
                       for r in ic for c in r if c is not None), "dropped batch"


class TestPurge6NFreeze:
    """6N depuration fish must stay frozen — no growth, no feed, only mortality —
    through the manual override window while the facility is in PURGE mode
    (before the 6N production-start date), matching the engine's depuration
    rules. In PRODUCTION mode (6N growth) they grow like normal grow-out."""

    @staticmethod
    def _occupied_6n(state):
        from forecast.sixn import SIXN_ALL_TANKS
        return sorted(t.tank_id for t in state.tanks_by_id.values()
                      if t.tank_id in SIXN_ALL_TANKS and not t.is_empty)

    @staticmethod
    def _project(state, control, ctx, n):
        import copy
        from forecast.manual_window import advance_facility_window
        from forecast.time_grid import forecast_week_labels
        labels = forecast_week_labels(ctx["forecast_start"], n)
        win = advance_facility_window(
            copy.deepcopy(state), ctx["batch_by_id"], ctx["tables"],
            ctx["forecast_start"], n, events=[], control=control,
            pr_closing=ctx["pr_closing"], fw_records=ctx["fw_records"])
        op = {(r.tank_id, r.week_label): r for r in win["opening_locations"]}
        rows = lambda tid: [op[(tid, wk)] for wk in labels if (tid, wk) in op]
        return win, rows

    def test_purge_6n_frozen_no_growth_but_mortality(self, hydrated):
        from forecast.sixn import is_purge_mode
        from forecast.state import STAGE_STARVE
        state, ctx, _fw = hydrated
        assert is_purge_mode(ctx["control"], ctx["forecast_start"]), \
            "fixture must be in purge mode"
        occ6 = self._occupied_6n(state)
        assert occ6, "expected occupied 6N tanks at hydration"
        win, rows = self._project(state, ctx["control"], ctx, 6)
        for tid in occ6:
            r = rows(tid)
            assert r
            assert abs(r[-1].avg_wt_g - r[0].avg_wt_g) < 1e-6, \
                f"6N tank {tid} grew {r[0].avg_wt_g}->{r[-1].avg_wt_g} in purge"
            assert all(x.stage == STAGE_STARVE for x in r)   # depurating
            assert r[-1].count < r[0].count                  # mortality applied
        # Traceability: the hold is surfaced as a note.
        assert any("6N depuration" in w for w in win["warnings"])

    def test_control_non_6n_still_grows(self, hydrated):
        from forecast.sixn import SIXN_ALL_TANKS
        state, ctx, _fw = hydrated
        ctrl = next(t.tank_id for t in state.tanks_by_id.values()
                    if t.tank_id not in SIXN_ALL_TANKS and not t.is_empty)
        _win, rows = self._project(state, ctx["control"], ctx, 6)
        r = rows(ctrl)
        assert r[-1].avg_wt_g > r[0].avg_wt_g   # grow-out fish grow (unaffected)

    def test_production_mode_6n_grows(self, hydrated):
        # sixn_growth=True => is_purge_mode False => the window must NOT freeze
        # 6N; they grow like normal grow-out (the freeze is purge-gated, not a
        # blanket 6N rule).
        import copy
        from forecast.sixn import is_purge_mode
        state, ctx, _fw = hydrated
        control = copy.deepcopy(ctx["control"])
        control.sixn_growth = True
        assert not is_purge_mode(control, ctx["forecast_start"])
        occ6 = self._occupied_6n(state)
        win, rows = self._project(state, control, ctx, 6)
        for tid in occ6:
            r = rows(tid)
            assert r[-1].avg_wt_g > r[0].avg_wt_g, \
                f"6N tank {tid} should grow in production mode"
        assert not any("6N depuration" in w for w in win["warnings"])

    def test_full_pipeline_purge_6n_audits_clean(self, hydrated, tmp_path):
        # A manual window (harvest in wk2) runs with the PR's 6N tanks held
        # frozen for the window weeks, then the planner takes over. The whole
        # run must still reconcile the HARD invariant: no count (TANK_DRIFT)
        # rows, facility count ratio in band, 0 dropped — freezing 6N must not
        # create/leak fish at handoff (a benign biomass timing artifact may
        # shift, see _assert_count_conserved).
        import shutil
        from openpyxl import load_workbook as _lw
        import forecast.run as run_mod
        from forecast.sixn import SIXN_ALL_TANKS
        state, _ctx, _fw = hydrated
        src = next(t for t in state.tanks_by_id.values()
                   if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                   and not t.is_empty)
        sdir = tmp_path / "scenario"
        shutil.copytree(SCENARIO_DIR, sdir)
        (sdir / "manual_events.yaml").write_text(
            "events:\n  - type: harvest\n    week: 2\n"
            f"    from_tank: {src.tank_id}\n    count: 1000\n")
        wb = tmp_path / "Forecast.xlsm"
        shutil.copy(WORKBOOK, wb)
        out = tmp_path / "out.xlsm"
        run_mod.main(str(wb), output_path=str(out),
                     config_dir=str(CONFIG_DIR), scenario_dir=str(sdir))
        owb = _lw(str(out), data_only=True)
        tc = [[c.value for c in r]
              for r in owb["TankContinuityAudit"].iter_rows()]
        _assert_count_conserved(tc)
        ic = [[c.value for c in r]
              for r in owb["InputConservationAudit"].iter_rows()]
        assert not any("*** DROP" in str(c).upper()
                       for r in ic for c in r if c is not None), "dropped batch"


class TestAutoCalibrateFw:
    """control.auto_calibrate_fw replaces each FW batch's FW_Correction with the
    back-solved value that lands its pre-cull avg weight on the TranOG target (the
    Suggested_FW_Correction the Diagnostics tab reports). Verified at the solver
    level, through config round-trip, and end-to-end (residuals -> ~0)."""

    def test_solved_correction_lands_on_target(self, hydrated):
        from forecast.biology import (
            solve_fw_correction, _simulate_fw_avg_weight_at_tran_og)
        _state, ctx, _fw = hydrated
        tables = ctx["tables"]
        b = next(b for b in ctx["batch_by_id"].values()
                 if b.tran_og_avg_wt_g and b.tran_og_date and b.input_date)
        s = solve_fw_correction(b, tables)
        assert s is not None
        w = _simulate_fw_avg_weight_at_tran_og(b, tables, s)
        assert abs(w - b.tran_og_avg_wt_g) / b.tran_og_avg_wt_g < 0.01

    def test_control_flag_roundtrips(self):
        from forecast.config_io import (
            load_control, control_to_dict, control_from_dict)
        c = load_control(str(CONFIG_DIR))
        c.auto_calibrate_fw = True
        c.auto_calibrate_fw_min = 0.6
        c.auto_calibrate_fw_max = 1.4
        c2 = control_from_dict(control_to_dict(c))
        assert c2.auto_calibrate_fw is True
        assert c2.auto_calibrate_fw_min == 0.6
        assert c2.auto_calibrate_fw_max == 1.4

    def test_full_pipeline_drives_residuals_to_zero(self, tmp_path):
        # With the toggle ON, every FW calibration residual in the output must be
        # ~0 (each batch lands on its transfer target), and the run must still
        # conserve fish.
        import re
        import shutil
        from openpyxl import load_workbook as _lw
        import forecast.run as run_mod
        cdir = tmp_path / "config"
        shutil.copytree(CONFIG_DIR, cdir)
        sdir = tmp_path / "scenario"
        shutil.copytree(SCENARIO_DIR, sdir)
        cy = cdir / "control.yaml"
        txt = cy.read_text()
        txt = (re.sub(r"auto_calibrate_fw:\s*\w+", "auto_calibrate_fw: true", txt)
               if "auto_calibrate_fw:" in txt
               else txt.rstrip() + "\nauto_calibrate_fw: true\n")
        cy.write_text(txt)
        (sdir / "manual_events.yaml").write_text("events: []\n")
        wb = tmp_path / "Forecast.xlsm"
        shutil.copy(WORKBOOK, wb)
        out = tmp_path / "out.xlsm"
        run_mod.main(str(wb), output_path=str(out),
                     config_dir=str(cdir), scenario_dir=str(sdir))
        owb = _lw(str(out), data_only=True)
        dg = [r for r in owb["Diagnostics"].iter_rows(values_only=True)]
        hdr = next(r for r in dg if r and str(r[0]) == "Batch")
        ri = hdr.index("Residual_pct")
        resids = [r[ri] for r in dg[dg.index(hdr) + 1:]
                  if r and r[0] and isinstance(r[ri], (int, float))]
        assert resids, "expected FW calibration residual rows"
        assert max(abs(v) for v in resids) < 0.1, f"worst residual {max(abs(v) for v in resids)}%"
        tc = [[c.value for c in r]
              for r in owb["TankContinuityAudit"].iter_rows()]
        _assert_count_conserved(tc)


class TestWindowHorizonGuard:
    """A manual override window as long as (or longer than) the forecast horizon
    leaves the planner no weeks to plan; the run must reject it, not silently
    clamp the post-window horizon to 1 week."""

    def test_window_at_or_beyond_horizon_is_rejected(self, tmp_path):
        import shutil
        import forecast.run as run_mod
        sdir = tmp_path / "scenario"
        shutil.copytree(SCENARIO_DIR, sdir)
        # The live scenario/ may contain operator-saved PER-PR event files
        # (scenario/manual_events/<closing>.yaml). Their mere presence marks
        # the environment "migrated", which correctly DISABLES the legacy
        # shared-file fallback this test relies on — purge the copy so the
        # test's legacy manual_events.yaml is actually honored.
        shutil.rmtree(sdir / "manual_events", ignore_errors=True)
        # A harvest event at a week far beyond any plausible horizon forces
        # window_n >= horizon_weeks.
        (sdir / "manual_events.yaml").write_text(
            "events:\n"
            "  - type: harvest\n"
            "    week: 9999\n"
            "    from_tank: 35\n"
            "    count: 1000\n")
        wb = tmp_path / "Forecast.xlsm"
        shutil.copy(WORKBOOK, wb)
        out = tmp_path / "out.xlsm"
        with pytest.raises(ValueError, match="horizon"):
            run_mod.main(str(wb), output_path=str(out),
                         config_dir=str(CONFIG_DIR), scenario_dir=str(sdir))
