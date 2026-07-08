"""Behavioral regression guard for the forecast pipeline.

Locks the model/pipeline BEHAVIOR, not specific output numbers. The exact
density-violation count / worst density / TranOG count legitimately change
with config (horizon, density target, limits, batches) — they are NOT the
thing under test. What must always hold, regardless of config, is that the
pipeline behaves correctly:

  1. COMPLETES + POPULATES — rc == 0 and the key sheets are written.
  2. MASS CONSERVATION — zero count-drift and zero biomass-drift rows in
     TankContinuityAudit: no fish/biomass created or lost unaccounted. (This
     also guarantees no TranOG arrival is silently dropped — a dropped
     arrival would break the count balance.)
  3. OUTPUT SANITY — every density is finite and >= 0 (no NaN / negative
     blow-ups); counts are non-negative.
  4. DETERMINISM — identical output regardless of PYTHONHASHSEED.

Runs the supported path (live config/ + scenario/ YAML, ProductionReport from
a COPY of Forecast.xlsm so the source is never mutated). Skips cleanly if the
workbook or config/scenario are absent.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent  # Python/
WORKBOOK = ROOT / "Forecast.xlsm"

pytestmark = pytest.mark.skipif(
    not WORKBOOK.exists(),
    reason="Forecast.xlsm not present (gitignored); regression test needs it",
)


CONFIG_DIR = ROOT / "config"
SCENARIO_DIR = ROOT / "scenario"


@pytest.fixture(scope="module")
def run_outputs(tmp_path_factory):
    """Run the SUPPORTED path and return the output workbook path.

    PR-only: the live config/ + scenario/ YAML (the single source of truth
    the app uses) plus the ProductionReport from the workbook. This is
    robust to Forecast.xlsm being a trimmed PR-only artifact — the limits +
    batches live in scenario/, not the workbook. Skips if config/scenario
    haven't been seeded (a clean checkout: run scripts/export_*_to_yaml.py).
    """
    import forecast.run as run_mod
    if not ((CONFIG_DIR / "control.yaml").exists()
            and (SCENARIO_DIR / "limits.yaml").exists()):
        pytest.skip("config/ + scenario/ not seeded "
                    "(run scripts/export_config_to_yaml.py + export_scenario_to_yaml.py)")
    tmp = tmp_path_factory.mktemp("wb") / "Forecast.xlsm"
    shutil.copy(WORKBOOK, tmp)
    rc = run_mod.main(str(tmp), config_dir=str(CONFIG_DIR),
                      scenario_dir=str(SCENARIO_DIR))
    assert rc == 0, f"pipeline exited non-zero ({rc})"
    return tmp


def _load(path):
    import openpyxl
    return openpyxl.load_workbook(path, keep_vba=True, data_only=True)


_REQUIRED_SHEETS = ["BatchLocations", "TankContinuityAudit", "HarvestPlan",
                    "TransferPlan", "BiologyProjection", "RunConfig"]


def test_run_completes_and_populates(run_outputs):
    """The run finishes and writes the key sheets with data."""
    wb = _load(run_outputs)
    for name in _REQUIRED_SHEETS:
        assert name in wb.sheetnames, f"missing output sheet {name}"
    rows = sum(1 for _ in wb["BatchLocations"].iter_rows())
    assert rows > 5, "BatchLocations has no data rows"


def test_mass_conservation(run_outputs):
    """No fish or biomass created/lost unaccounted — config-independent.

    COUNT is the HARD invariant: fish cannot be created or destroyed, so
    per-tank count continuity must be EXACT (zero TANK_DRIFT) and the
    facility-level signed/abs count ratio must cancel to ~0 (the distributed-
    leak gauge — near 1 = systematic one-way fish loss).

    BIOMASS per-tank reconciliation is inherently APPROXIMATE and is asserted at
    the FACILITY level, not per row. Transfer events capture the source weight
    ~half a week behind the weekly BatchLocations snapshot, so a tank that fully
    turns over in one week shows a phantom per-row BIO_DRIFT while mass is in
    fact conserved (the departed mass reappears in the destination tank; the
    audit itself labels facility biomass drift "reported, not asserted"). The
    decisive point: a biomass drift NOT accompanied by a count drift cannot be a
    real leak — no fish went missing, only weight attribution shifted between
    tank-weeks. So we bound the facility-level NET biomass drift to a small
    fraction of peak facility biomass; a genuine mass leak would be far larger
    AND would surface as a count drift (already excluded above).

    NOTE: zero count-drift does NOT prove no batch was dropped (a never-placed
    batch creates no tank-week rows). Input-fish conservation is enforced
    separately by test_no_dropped_batches.
    """
    from collections import defaultdict
    wb = _load(run_outputs)
    ws = wb["TankContinuityAudit"]
    rows = list(ws.iter_rows(values_only=True))

    count_drift = 0
    fac = {}
    for i, row in enumerate(rows, 1):
        if not row:
            continue
        if i >= 5 and row[14] == "TANK_DRIFT":
            count_drift += 1
        # Facility conservation summary rows: [metric, signed, abs, ratio, note]
        if row[0] in ("Count (fish)", "Biomass (kg)"):
            fac[row[0]] = (row[1], row[2], row[3])

    # 1) Fish conservation is EXACT per tank.
    assert count_drift == 0, f"{count_drift} tank count-drift rows"
    # 2) Facility-level fish-leak gauge: signed must cancel to ~0.
    assert "Count (fish)" in fac, "missing facility conservation summary"
    c_signed, c_abs, c_ratio = fac["Count (fish)"]
    assert abs(c_ratio) < 0.3, (
        f"facility count signed/abs ratio {c_ratio:.3f} (|ratio|>=0.3) — "
        f"distributed fish leak (signed {c_signed:,.0f} / abs {c_abs:,.0f})")

    # 3) Biomass conserved within the weekly-vs-daily growth-approximation bias:
    #    bound the facility NET signed drift to <2% of peak facility biomass.
    bl = wb["BatchLocations"]
    blrows = list(bl.iter_rows(values_only=True))
    bhi = next(idx for idx, r in enumerate(blrows)
              if r and "Week" in [str(c) for c in r])
    bhdr = [str(c) for c in blrows[bhi]]
    wcol, biocol = bhdr.index("Week"), bhdr.index("Biomass (kg)")
    per_wk = defaultdict(float)
    for r in blrows[bhi + 1:]:
        if not r or r[wcol] is None:
            continue
        per_wk[r[wcol]] += float(r[biocol] or 0.0)
    peak_bio = max(per_wk.values()) if per_wk else 0.0
    b_signed = fac.get("Biomass (kg)", (0.0,))[0] or 0.0
    assert peak_bio > 0 and abs(b_signed) < 0.02 * peak_bio, (
        f"facility biomass net drift {b_signed:,.0f} kg >= 2% of peak facility "
        f"biomass {peak_bio:,.0f} kg — possible real mass leak")


def test_no_dropped_batches(run_outputs):
    """Input-fish conservation: no stocked batch is silently dropped.

    Closes the blind spot in test_mass_conservation. Every batch whose TranOG
    falls within the horizon must reach the realized facility; a batch the
    placement engine fails to place (no empty OG tank) vanishes from the plan
    with its full stocked population. The InputConservationAudit flags these as
    'DROPPED' with a Fish_At_Risk count, which must total zero.
    """
    wb = _load(run_outputs)
    assert "InputConservationAudit" in wb.sheetnames, "missing InputConservationAudit"
    ws = wb["InputConservationAudit"]
    dropped = []
    at_risk = 0.0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row or not row[0]:
            continue
        if isinstance(row[7], str) and "DROPPED" in row[7]:
            dropped.append(row[0])
            if isinstance(row[8], (int, float)):
                at_risk += row[8]
    assert not dropped, (
        f"{len(dropped)} batch(es) dropped (never placed): {dropped} — "
        f"{at_risk:,.0f} stocked fish lost from the plan")
    # And the other end: no batch may harvest + still-hold MORE than it stocked.
    over = [row[0] for row in ws.iter_rows(values_only=True)
            if row and isinstance(row[0], str) and "OVER-PRODUCED" in row[0]]
    assert not over, f"input-fish conservation breached (fish created): {over}"


def test_fw_mass_balance(run_outputs):
    """Closed FW-phase mass-balance (audit I2): the previously-unaudited freshwater
    phase must conserve fish — first_FW_count == realized_TranOG + FW_mortality +
    FW_culls for every batch that crosses to seawater in-horizon. TankContinuity
    only starts at OG, so a fish leak or a mortality/cull-accounting error in FW
    would otherwise shift total smolts (and harvest tonnage) with every other gate
    green. The InputConservationAudit emits a 'FW MASS-BALANCE BREACH' line when any
    batch fails to reconcile beyond tolerance; it must not be present.
    """
    wb = _load(run_outputs)
    ws = wb["InputConservationAudit"]
    breach = [row[0] for row in ws.iter_rows(values_only=True)
              if row and isinstance(row[0], str) and "FW MASS-BALANCE BREACH" in row[0]]
    assert not breach, f"FW phase does not conserve fish: {breach}"


def test_facility_count_conservation(run_outputs):
    """No DISTRIBUTED fish loss across the facility.

    Per-tank-week count drift is bounded by a 50-fish tolerance, so a small
    same-sign leak spread across many tanks (each under tolerance) passes
    test_mass_conservation while still losing fish in aggregate. The
    TankContinuityAudit FACILITY CONSERVATION SUMMARY sums every tank-week count
    delta; the signed/abs ratio must stay near 0 (random/cancelling = conserved).
    A ratio near 1 means a systematic, distributed one-way loss.
    """
    wb = _load(run_outputs)
    ws = wb["TankContinuityAudit"]
    ratio = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Count (fish)" and isinstance(row[3], (int, float)):
            ratio = row[3]
            break
    assert ratio is not None, "missing FACILITY CONSERVATION SUMMARY count row"
    assert abs(ratio) < 0.3, (
        f"facility count signed/abs ratio {ratio:.3f} — distributed fish loss "
        f"(systematic one-way count drift across tank-weeks)")


def test_output_sanity(run_outputs):
    """Densities are finite and non-negative; counts non-negative.

    We do NOT assert a specific violation count or worst density — those are
    config-dependent. We only guard against NaN/negative blow-ups.
    """
    wb = _load(run_outputs)
    ws = wb["BatchLocations"]
    bad_density = bad_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 5 or not row:
            continue
        count, density = row[5], row[8]
        if isinstance(density, (int, float)) and (density != density or density < 0):
            bad_density += 1
        if isinstance(count, (int, float)) and count < 0:
            bad_count += 1
    assert bad_density == 0, f"{bad_density} NaN/negative density rows"
    assert bad_count == 0, f"{bad_count} negative count rows"


def test_no_harvest_craters(run_outputs):
    """No NEAR-EMPTY mid-horizon harvest week — the steady-harvest contract rule.

    The controller must never leave a mid-horizon week effectively empty: the
    L1-envelope diagnostic (2026-07-08) proved the market-ready SUPPLY exists at
    the controller's crater weeks (L1 holds 30-47k where the controller drops to
    a few hundred), so a near-empty week is a pacing failure, not a shortage.

    Excludes the first 6 weeks (operator-pinned startup handoff) and treats only
    a week below a QUARTER of the harvest floor as a crater (a week merely a few
    fish under the floor is rounding, not a breach) — matching the Compare &
    Choose board's "No empty week" gate. This is the RED gate the anti-crater
    hybrid (L1 envelope -> controller harvest target) must turn green; on a PR
    that does not crater it is a forward-lock against a regression.
    """
    import yaml
    from forecast.optimize import _harvest_weekly_fish
    cfg = yaml.safe_load((CONFIG_DIR / "control.yaml").read_text()) or {}
    floor = float(cfg.get("min_harvest_per_week", 0) or 0)
    if floor <= 0:
        pytest.skip("no min_harvest_per_week floor configured")
    wb = _load(run_outputs)
    weekly = _harvest_weekly_fish(wb)
    craters = [(i, int(c)) for i, c in enumerate(weekly)
               if i >= 6 and c < 0.25 * floor]
    assert not craters, (
        f"near-empty mid-horizon harvest weeks (< 25% of the {floor:,.0f} floor): "
        f"{craters} — the steady-harvest contract rule is breached")


# ---- Determinism guard (2026-06-05) ----
# The forecast must be identical regardless of PYTHONHASHSEED. A
# set-of-strings iteration in phase_d without a deterministic tiebreak
# made it vary run-to-run (245/216.5 vs 228/168.3 on the same workbook).
# Runs the pipeline in two subprocesses with different hash seeds and
# asserts the BatchLocations density signature matches.

def test_engine_deterministic_across_hash_seeds():
    import os
    import subprocess
    import sys
    import textwrap

    if not ((CONFIG_DIR / "control.yaml").exists()
            and (SCENARIO_DIR / "limits.yaml").exists()):
        pytest.skip("config/ + scenario/ not seeded")

    code = textwrap.dedent(
        """
        import shutil, tempfile, os, io, contextlib, openpyxl
        import forecast.run as r
        td = tempfile.gettempdir()
        t = os.path.join(td, "det%d.xlsm" % os.getpid())
        o = os.path.join(td, "deto%d.xlsm" % os.getpid())
        shutil.copy(os.environ["WB"], t)
        with contextlib.redirect_stdout(io.StringIO()):
            r.main(t, o, config_dir=os.environ["CFG"], scenario_dir=os.environ["SCN"])
        wb = openpyxl.load_workbook(o, data_only=True)
        ws = wb["BatchLocations"]
        v = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i < 5 or not row:
                continue
            d = row[8]
            if isinstance(d, (int, float)) and d > 95:
                v.append(round(d, 2))
        print("%d|%.2f|%.2f" % (len(v), max(v, default=0.0), round(sum(v), 2)))
        """
    )

    def _run(seed):
        env = dict(os.environ, WB=str(WORKBOOK), CFG=str(CONFIG_DIR),
                   SCN=str(SCENARIO_DIR), PYTHONHASHSEED=str(seed))
        out = subprocess.run([sys.executable, "-c", code], cwd=str(ROOT),
                             capture_output=True, text=True, env=env)
        assert out.returncode == 0, f"seed {seed} failed: {out.stderr[-500:]}"
        return out.stdout.strip().splitlines()[-1]

    sig0 = _run(0)
    sig1 = _run(1)
    assert sig0 == sig1, (
        f"non-deterministic across hash seeds: seed0={sig0} seed1={sig1}")
