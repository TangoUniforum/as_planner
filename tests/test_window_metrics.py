"""Harvest-COMPLIANCE metrics judge the PLANNER on planner weeks only.

Operator-hit 2026-08: a 3-week manual window whose week 3 deliberately
scheduled no harvest (scripted staging only) made EVERY method fail the hard
`no_empty_week` gate — the zero-week counters counted the operator's own
scripted week against all engines, none of which may touch a window week. The
board went degenerate: no winner, 14 useless tuning probes per method.

The fix under test: the workbook SELF-DESCRIBES its window (the MANUAL EVENT
OK / MANUAL WINDOW ValidationLog lines shipped in 33dd203), and the harvest-
compliance counters (zero weeks, min week, over-limit / over-ceiling weeks)
exclude those weeks. Old workbooks without the log rows behave exactly as
before, and a genuine out-of-window empty week still fails — the gate must
never go blind.
"""
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from forecast.optimize import metrics_from_workbook
from forecast.window_weeks import manual_window_weeks


def _make_workbook(path: Path, weeks, harvest, manual_weeks=(),
                   dash="—") -> str:
    """Minimal workbook metrics_from_workbook can read: Advisory (horizon +
    caps), HarvestPlan (weekly fish), an empty TransferPlan, and — when
    `manual_weeks` given — a ValidationLog with the manual-window narration
    rows (dash swappable to simulate the '—' -> '�' encoding variance)."""
    wb = Workbook()
    adv = wb.active
    adv.title = "Advisory"
    adv.append(["Week", "Phase", "Total_Biomass (kg)", "Biomass_Limit (kg)",
                "Total_Feed (kg/day)", "Feed_Limit (kg/day)"])
    for w in weeks:
        adv.append([w, "", 1_000.0, 2_000.0, 10.0, 20.0])
    hp = wb.create_sheet("HarvestPlan")
    hp.append(["Week", "Batch", "Tank", "Count (fish)"])
    for w, n in harvest.items():
        hp.append([w, "B1", "T1", n])
    tp = wb.create_sheet("TransferPlan")
    tp.append(["Week", "From", "Type", "Count (fish)"])
    if manual_weeks:
        vl = wb.create_sheet("ValidationLog")
        vl.append(["VALIDATION LOG"])
        vl.append(["#", "Category", "Detail"])
        for i, w in enumerate(manual_weeks, 1):
            vl.append([i, "INFO - Manual window (executed)",
                       f"MANUAL EVENT OK {dash} {w}: harvest #1 from tank "
                       f"#45 (scripted)"])
    wb.save(path)
    wb.close()
    return str(path)


_WEEKS = ["2026-W31", "2026-W32", "2026-W33", "2026-W34", "2026-W35"]
_WINDOW = ("2026-W31", "2026-W32", "2026-W33")


class TestManualWindowWeeksReader:
    def test_reads_week_labels_from_validation_log(self, tmp_path):
        p = _make_workbook(tmp_path / "a.xlsx", _WEEKS,
                           {w: 10_000.0 for w in _WEEKS},
                           manual_weeks=_WINDOW)
        assert manual_window_weeks(p) == set(_WINDOW)
        # Also accepts an already-open workbook (the metrics reader's path).
        wb = load_workbook(p, data_only=True)
        try:
            assert manual_window_weeks(wb) == set(_WINDOW)
        finally:
            wb.close()

    def test_old_workbook_without_log_rows_is_empty_set(self, tmp_path):
        p = _make_workbook(tmp_path / "b.xlsx", _WEEKS,
                           {w: 10_000.0 for w in _WEEKS})
        assert manual_window_weeks(p) == set()

    def test_survives_dash_encoding_variance(self, tmp_path):
        # The em-dash sometimes round-trips as '�'; the reader keys on the
        # MANUAL prefix + the ISO week pattern, never on the dash.
        p = _make_workbook(tmp_path / "c.xlsx", _WEEKS,
                           {w: 10_000.0 for w in _WEEKS},
                           manual_weeks=_WINDOW, dash="�")
        assert manual_window_weeks(p) == set(_WINDOW)


class TestZeroWeekExclusion:
    def test_scripted_zero_harvest_week_not_counted(self, tmp_path):
        # W33 is the operator's deliberately harvest-free window week.
        harvest = {w: 10_000.0 for w in _WEEKS if w != "2026-W33"}
        p = _make_workbook(tmp_path / "win.xlsx", _WEEKS, harvest,
                           manual_weeks=_WINDOW)
        m, dropped, overprod = metrics_from_workbook(p, 55_000.0)
        assert m.harvest_zero_weeks == 0
        assert m.harvest_min_week == 10_000.0
        assert (dropped, overprod) == (0, 0)

    def test_same_workbook_without_log_rows_counts_the_week(self, tmp_path):
        # Old workbook (no narration): exact previous behavior — W33 counts.
        harvest = {w: 10_000.0 for w in _WEEKS if w != "2026-W33"}
        p = _make_workbook(tmp_path / "old.xlsx", _WEEKS, harvest)
        m, _, _ = metrics_from_workbook(p, 55_000.0)
        assert m.harvest_zero_weeks == 1
        assert m.harvest_min_week == 0.0

    def test_genuine_out_of_window_empty_week_still_fails(self, tmp_path):
        # The gate must not go blind: an engine-planned empty week OUTSIDE
        # the window still counts even when a window is present.
        weeks = _WEEKS + ["2026-W40"]          # W40 in horizon, no harvest
        harvest = {w: 10_000.0 for w in _WEEKS if w != "2026-W33"}
        p = _make_workbook(tmp_path / "blind.xlsx", weeks, harvest,
                           manual_weeks=_WINDOW)
        m, _, _ = metrics_from_workbook(p, 55_000.0)
        assert m.harvest_zero_weeks == 1       # W40 only; W33 excluded


class TestContractFloorMetrics:
    """The steady-harvest CONTRACT floor (min_harvest_per_week), measured.

    `harvest_zero_weeks` only ever caught the DEGENERATE case (a week that
    harvests literally nothing). The floor itself reached Metrics on
    2026-08-12, after a tuned tournament on the operator's 7.29 PR promoted
    knobs that cut the worst harvest week 20,526 -> 16,185 fish — a change
    the objective could not see because it scores no floor term.
    """

    def test_floor_fields_are_none_without_a_floor(self, tmp_path):
        p = _make_workbook(tmp_path / "nofloor.xlsx", _WEEKS,
                           {w: 10_000.0 for w in _WEEKS})
        m, _, _ = metrics_from_workbook(p, 55_000.0)
        # UNKNOWN, never "the floor was met" — the same rule as zero_weeks.
        assert m.harvest_weeks_below_floor is None
        assert m.harvest_floor_gap is None

    def test_counts_sub_floor_weeks_and_the_mean_shortfall(self, tmp_path):
        harvest = {w: 40_000.0 for w in _WEEKS}
        harvest["2026-W34"] = 15_000.0        # 15,000 short of a 30,000 floor
        harvest["2026-W35"] = 27_000.0        # 3,000 short
        p = _make_workbook(tmp_path / "floor.xlsx", _WEEKS, harvest)
        m, _, _ = metrics_from_workbook(p, 55_000.0, min_harvest=30_000.0)
        assert m.harvest_weeks_below_floor == 2
        assert m.harvest_min_week == 15_000.0
        # mean shortfall as a fraction of the floor: 18,000 / (30,000 x 5)
        assert m.harvest_floor_gap == pytest.approx(0.12)

    def test_a_plan_that_always_clears_the_floor_scores_zero_gap(self, tmp_path):
        p = _make_workbook(tmp_path / "clear.xlsx", _WEEKS,
                           {w: 30_000.0 for w in _WEEKS})
        m, _, _ = metrics_from_workbook(p, 55_000.0, min_harvest=30_000.0)
        assert m.harvest_weeks_below_floor == 0
        assert m.harvest_floor_gap == 0.0

    def test_scripted_window_weeks_are_not_blamed_on_the_planner(self, tmp_path):
        # Same exclusion rule as every other compliance count: a lean week the
        # OPERATOR scripted is not a planner floor miss.
        harvest = {w: 40_000.0 for w in _WEEKS}
        harvest["2026-W31"] = 1_000.0          # inside the manual window
        p = _make_workbook(tmp_path / "wfloor.xlsx", _WEEKS, harvest,
                           manual_weeks=_WINDOW)
        m, _, _ = metrics_from_workbook(p, 55_000.0, min_harvest=30_000.0)
        assert m.harvest_weeks_below_floor == 0 and m.harvest_floor_gap == 0.0


class TestOverCeilingExclusion:
    def test_scripted_over_ceiling_week_excluded(self, tmp_path):
        # The operator scripted a 100k blowout week (their choice, both
        # directions); planner weeks stay legal.
        harvest = {w: 10_000.0 for w in _WEEKS}
        harvest["2026-W31"] = 100_000.0
        p = _make_workbook(tmp_path / "ceil.xlsx", _WEEKS, harvest,
                           manual_weeks=_WINDOW)
        m, _, _ = metrics_from_workbook(p, 55_000.0, relief_ceiling=60_500.0)
        assert m.weeks_over_harvest_cap == 0
        assert m.weeks_over_relief_ceiling == 0
        # Without the narration the same week is blamed on the planner.
        p2 = _make_workbook(tmp_path / "ceil_old.xlsx", _WEEKS, harvest)
        m2, _, _ = metrics_from_workbook(p2, 55_000.0, relief_ceiling=60_500.0)
        assert m2.weeks_over_harvest_cap == 1
        assert m2.weeks_over_relief_ceiling == 1

    def test_planner_over_ceiling_week_still_counts(self, tmp_path):
        harvest = {w: 10_000.0 for w in _WEEKS}
        harvest["2026-W35"] = 100_000.0        # planner week, over ceiling
        p = _make_workbook(tmp_path / "pceil.xlsx", _WEEKS, harvest,
                           manual_weeks=_WINDOW)
        m, _, _ = metrics_from_workbook(p, 55_000.0, relief_ceiling=60_500.0)
        assert m.weeks_over_harvest_cap == 1
        assert m.weeks_over_relief_ceiling == 1


class TestCompareHarness:
    def test_harvest_extras_exclude_window(self, tmp_path):
        from tools.run_compare import _harvest_extras
        harvest = {w: 10_000.0 for w in _WEEKS if w != "2026-W33"}
        p = _make_workbook(tmp_path / "cmp.xlsx", _WEEKS, harvest,
                           manual_weeks=_WINDOW)
        h = _harvest_extras(p, min_harvest=5_000.0)
        assert h["zero_weeks"] == 0
        assert h["weeks_below_min"] == 0
        assert h["min_week"] == 10_000.0
        assert h["window_weeks_excluded"] == 3
        assert h["n_weeks"] == len(_WEEKS)     # display stays whole-horizon
        # Old workbook: previous behavior byte-for-byte.
        p2 = _make_workbook(tmp_path / "cmp_old.xlsx", _WEEKS, harvest)
        h2 = _harvest_extras(p2, min_harvest=5_000.0)
        assert h2["zero_weeks"] == 1
        assert h2["weeks_below_min"] == 1
        assert h2["min_week"] == 0.0
        assert h2["window_weeks_excluded"] == 0


class TestGateWording:
    def test_gate_says_what_it_judged(self):
        from forecast.analysis import _gate_no_empty_week
        s, d = _gate_no_empty_week({"zero_weeks": 0, "zero_weeks_excluded": 3})
        assert s == "PASS"
        assert "3" in d and "excluded" in d
        s, d = _gate_no_empty_week({"zero_weeks": 0})
        assert (s, d) == ("PASS", "harvests something every week")
        s, d = _gate_no_empty_week({"zero_weeks": 2, "zero_weeks_excluded": 3})
        assert s == "FAIL" and "2" in d and "excluded" in d
