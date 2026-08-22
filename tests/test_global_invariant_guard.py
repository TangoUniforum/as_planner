"""Prove the Global safety net actually goes red.

tools/check_global_invariants.py exists to make one specific historical
failure impossible to ship again: enforcing the 6N-only harvest rule in the
tank picker without first reconciling `harvest_by_bw` against `purge_rows`
took TANK_DRIFT 0 -> 3 and destroyed 39,077 fish, while making every surface
an operator looks at read BETTER (smoother harvest, calmer density, a
6N-rule probe of 100%).

A net that never goes red is worse than no net, because it is trusted. So
these tests do not check that a good run passes -- they check that each way
of losing fish is CAUGHT, and that a genuine improvement is NOT flagged.
"""
from __future__ import annotations

import openpyxl
import pytest

from tools.check_global_invariants import COUNT_EPS, compare, measure

# The real run this guard was calibrated on.
BASE = {
    "tank_week_rows": 2777, "count_drift_rows": 0, "count_drift_fish": 0.0,
    "bio_drift_rows": 31, "fish_at_risk": 0.0, "horizon_weeks": 85,
    "harvest_fish": 3548489.0, "empty_weeks": 1,
}
TOL = 0.005


def _now(**over):
    d = dict(BASE)
    d.update(over)
    return d


class TestCatchesFishLoss:
    """Each of these is a way the known defect actually presented."""

    def test_count_drift_is_a_hard_failure(self):
        fails = compare(_now(count_drift_rows=3, count_drift_fish=39077.0),
                        BASE, TOL)
        assert fails, "3 drifting tank-weeks must fail"
        assert "COUNT DRIFT" in fails[0]
        assert "39,077" in fails[0]

    def test_fish_at_risk_is_a_hard_failure(self):
        fails = compare(_now(fish_at_risk=39077.0), BASE, TOL)
        assert any("FISH AT RISK" in f for f in fails)

    def test_silent_harvest_drop_is_a_hard_failure(self):
        """The dangerous shape: nothing drifts, the plan just harvests fewer
        fish. Every other signal looks fine or better."""
        lost = BASE["harvest_fish"] - 39077.0
        fails = compare(_now(harvest_fish=lost, empty_weeks=0), BASE, TOL)
        assert any("HARVEST FELL" in f for f in fails), (
            "losing 39,077 fish while REMOVING the empty week must still fail")

    def test_all_three_failures_are_reported_together(self):
        fails = compare(_now(count_drift_rows=3, count_drift_fish=39077.0,
                             fish_at_risk=39077.0,
                             harvest_fish=BASE["harvest_fish"] - 39077.0),
                        BASE, TOL)
        assert len(fails) == 3, "each independent signal must be reported"


class TestDoesNotBlockRealFixes:
    """The net must not punish the improvement it is meant to enable."""

    def test_removing_the_empty_week_alone_passes(self):
        assert compare(_now(empty_weeks=0, empty_week_labels=[]), BASE, TOL) == []

    def test_harvesting_more_fish_passes(self):
        assert compare(_now(harvest_fish=BASE["harvest_fish"] * 1.02),
                       BASE, TOL) == []

    def test_bio_drift_change_is_not_a_failure(self):
        """BIO_DRIFT is a biomass expectation residual; the COUNT side is what
        proves conservation. Failing on it would block unrelated work."""
        assert compare(_now(bio_drift_rows=99), BASE, TOL) == []

    def test_harvest_dip_inside_tolerance_passes(self):
        inside = BASE["harvest_fish"] * (1.0 - TOL / 2.0)
        assert compare(_now(harvest_fish=inside), BASE, TOL) == []

    def test_no_baseline_still_catches_absolute_failures(self):
        """A first run has nothing to compare against, but drift is drift."""
        fails = compare(_now(count_drift_rows=1, count_drift_fish=5.0), {}, TOL)
        assert any("COUNT DRIFT" in f for f in fails)


class TestMeasuresARealWorkbook:
    """measure() must read the sheets the way the audits actually write them."""

    def _wb(self, tmp_path, delta):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TankContinuityAudit"
        ws.append(["TANK CONTINUITY AUDIT"])
        ws.append([])
        ws.append([])
        ws.append(["Week", "Tank", "Batch", "Open_Count", "Expected_Close",
                   "Actual_Close", "Delta", "Flag", "Bio_Flag"])
        ws.append(["2026-W33", 61, "B41", 100.0, 100.0, 100.0, 0.0, None, None])
        ws.append(["2026-W34", 61, "B41", 100.0, 100.0, 100.0 - delta, delta,
                   "TANK_DRIFT" if delta else None, None])
        ic = wb.create_sheet("InputConservationAudit")
        ic.append(["INPUT-FISH CONSERVATION AUDIT"])
        ic.append([])
        ic.append([])
        ic.append(["Batch", "Input_Count (fish)", "Status", "Fish_At_Risk (fish)"])
        ic.append(["B41", 1000, "PLACED", 0])
        bl = wb.create_sheet("BatchLocations")
        for _ in range(4):
            bl.append(["Week", "x"])
        bl.append(["2026-W33", "x"])
        bl.append(["2026-W34", "x"])
        hp = wb.create_sheet("HarvestPlan")
        for _ in range(4):
            hp.append(["Week", "Batch", "Tank", "Count (fish)"])
        hp.append(["2026-W34", "B41", 61, 500.0])
        p = tmp_path / f"g{int(delta)}.xlsx"
        wb.save(p)
        return p

    def test_clean_workbook_measures_zero_drift(self, tmp_path):
        m = measure(self._wb(tmp_path, 0.0))
        assert m["count_drift_rows"] == 0
        assert m["harvest_fish"] == 500.0
        # W33 appears in BatchLocations but has no harvest -> one empty week.
        assert m["empty_weeks"] == 1
        assert m["empty_week_labels"] == ["2026-W33"]

    def test_drifting_workbook_is_measured_and_fails(self, tmp_path):
        m = measure(self._wb(tmp_path, 39077.0))
        assert m["count_drift_rows"] == 1
        assert m["count_drift_fish"] == pytest.approx(39077.0)
        assert compare(m, BASE, TOL), "a drifting workbook must fail"

    def test_rounding_noise_is_not_drift(self, tmp_path):
        m = measure(self._wb(tmp_path, COUNT_EPS / 2.0))
        assert m["count_drift_rows"] == 0
