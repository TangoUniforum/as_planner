"""Extract the numbers the regression baseline pins, from a run's output workbook.

Shared by the freezer (`freeze_golden.py`) and the test
(`tests/test_reference_baseline.py`) so the two can never drift.

WHAT IS PINNED, AND WHY THESE
  totals        the headline the business reads: fish, gross kg, HOG kg.
  monthly_hog   the number finance signs off. Pinned per month because the
                horizon total can stay put while the monthly profile moves —
                measured on the VBA/Python comparison, annual tonnage agreed
                to ~2% while individual months moved up to 40%.
  weekly_fish   the shape. Catches a scheduler policy change (a FIFO->LIFO
                inversion moved this and nothing else caught it).
  compliance    the rule counts — weeks over the processing cap, weeks under
                the sales floor, tank-weeks over density. These are the ones
                that would embarrass you if they moved silently, so they are
                compared EXACTLY.
  batches       per-batch harvest, so a change that merely reshuffles which
                cohort was taken is still visible.

Everything is read from the SHIPPED workbook, not from engine internals — the
baseline must break when the artifact the operator receives changes, which is
not the same thing as when a data structure changes.
"""
from __future__ import annotations

import collections
import datetime as dt
from pathlib import Path

import openpyxl


# A tank at 95.004 kg/m3 against a 95 cap is a rounding artefact, not a
# welfare breach. 0.05 kg/m3 is ~0.05% of the cap and well below anything an
# operator would act on.
_DENSITY_EPS = 0.05


def _monday(d):
    if isinstance(d, dt.datetime):
        d = d.date()
    if not isinstance(d, dt.date):
        return None
    return d - dt.timedelta(days=d.weekday())


def extract(path) -> dict:
    """Read one output workbook into the pinned-metric dict."""
    wb = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)

    # ---- HarvestReport: totals, monthly, weekly, per batch ----------------
    fish = gross = hog = 0.0
    events = 0
    monthly: dict[str, float] = collections.defaultdict(float)
    weekly: dict[str, float] = collections.defaultdict(float)
    by_batch: dict[str, float] = collections.defaultdict(float)
    for i, row in enumerate(
            wb["HarvestReport"].iter_rows(min_row=1, max_col=10, values_only=True), 1):
        if i <= 4:
            continue
        d, batch, cnt, gr, hg = row[3], row[5], row[6], row[7], row[8]
        if not isinstance(cnt, (int, float)) or not isinstance(d, dt.datetime):
            continue
        events += 1
        fish += cnt
        gross += gr or 0.0
        hog += hg or 0.0
        monthly[d.strftime("%Y-%m")] += hg or 0.0
        weekly[str(_monday(d))] += cnt
        by_batch[str(batch)] += cnt

    # ---- BatchLocations: density + per-system biomass ---------------------
    tank_weeks_over_density = 0
    peak_density = 0.0
    sys_week_bio: dict[tuple, float] = collections.defaultdict(float)
    caps_by_tank = _tank_density_caps()
    for i, row in enumerate(
            wb["BatchLocations"].iter_rows(min_row=1, max_col=9, values_only=True), 1):
        if i <= 4:
            continue
        ws_, tank, sysid, bio, dens = row[1], row[3], row[4], row[7], row[8]
        if not isinstance(bio, (int, float)) or not isinstance(ws_, dt.datetime):
            continue
        if isinstance(dens, (int, float)):
            # PURGE-MODE 6N IS EXEMPT (operator, 2026-08-20) — the harvest
            # schedule bounds a depuration tank, not kg/m3. run.py's own
            # density audit already skips it; this baseline must judge by the
            # same rule or it measures a constraint the engine does not have.
            # The fixture horizon is entirely pre-production, so OG6N here is
            # always purge.
            if str(sysid) != "OG6N":
                peak_density = max(peak_density, dens)
                # Per-TANK cap, from the fixture facility config. This used to
                # take one number scraped out of the RunConfig text, which
                # matched a FRESHWATER tank at 30 kg/m3 and then judged all 39
                # grow-out tanks (cap 95) against it — reporting 691 breaches
                # where there was 1. A cap read from the wrong tank is not a
                # stricter test, it is a meaningless one.
                try:
                    _cap = caps_by_tank.get(int(tank), 0.0)
                except (TypeError, ValueError):
                    _cap = 0.0
                # Epsilon, not a bare `>`. The planner fills deliberately TO
                # the cap, so without it this counts float noise.
                if _cap > 0 and dens > _cap + _DENSITY_EPS:
                    tank_weeks_over_density += 1
        sys_week_bio[(str(_monday(ws_)), str(sysid))] += bio

    wb.close()
    return {
        "totals": {
            "harvest_events": events,
            "fish": round(fish, 3),
            "gross_kg": round(gross, 3),
            "hog_kg": round(hog, 3),
        },
        "monthly_hog_kg": {k: round(v, 3) for k, v in sorted(monthly.items())},
        "weekly_fish": {k: round(v, 3) for k, v in sorted(weekly.items())},
        "batch_fish": {k: round(v, 3) for k, v in sorted(by_batch.items())},
        "compliance": {
            "tank_weeks_over_density": tank_weeks_over_density,
            "peak_density_kg_m3": round(peak_density, 2),
            "occupied_system_weeks": len(sys_week_bio),
        },
    }


def _tank_density_caps() -> dict:
    """Per-tank `max_density_kg_m3`, from the reference fixture's own config.

    Read from config rather than scraped out of the RunConfig sheet text: the
    scrape took the FIRST number it found after "max_density_kg_m3", which is
    a freshwater tank's 30, and then judged every grow-out tank against it.
    Tank ids are unique across the fixture facility, so one flat map is enough.
    """
    import sys as _sys
    _root = Path(__file__).resolve().parents[2]
    if str(_root) not in _sys.path:
        _sys.path.insert(0, str(_root))
    from forecast.config_io import load_facility_config
    fac = load_facility_config(Path(__file__).resolve().parent
                               / "reference" / "config")
    return {t.tank_id: t.max_density_kg_m3 for t in fac.tanks}


def compare(golden: dict, actual: dict) -> list[str]:
    """Return a list of human-readable differences. Empty list == identical.

    Tolerances (agreed with the operator 2026-08-20):
      counts and compliance   EXACT
      tonnage                 0.1%
    """
    diffs: list[str] = []

    def _exact(path, g, a):
        if g != a:
            diffs.append(f"{path}: golden {g!r} != actual {a!r}")

    def _tol(path, g, a, pct=0.001):
        if g == 0:
            if abs(a) > 1e-6:
                diffs.append(f"{path}: golden 0 != actual {a:,.3f}")
            return
        rel = abs(a - g) / abs(g)
        if rel > pct:
            diffs.append(
                f"{path}: golden {g:,.3f} != actual {a:,.3f} "
                f"({(a - g) / g * 100:+.3f}%, tolerance {pct * 100:g}%)")

    _exact("totals.harvest_events",
           golden["totals"]["harvest_events"], actual["totals"]["harvest_events"])
    _exact("totals.fish", golden["totals"]["fish"], actual["totals"]["fish"])
    _tol("totals.gross_kg", golden["totals"]["gross_kg"], actual["totals"]["gross_kg"])
    _tol("totals.hog_kg", golden["totals"]["hog_kg"], actual["totals"]["hog_kg"])

    for section, cmp_fn in (("monthly_hog_kg", _tol), ("weekly_fish", _exact),
                            ("batch_fish", _exact)):
        g, a = golden[section], actual[section]
        for key in sorted(set(g) | set(a)):
            if key not in g:
                diffs.append(f"{section}[{key}]: not in golden, actual {a[key]:,.3f}")
            elif key not in a:
                diffs.append(f"{section}[{key}]: golden {g[key]:,.3f}, absent from actual")
            else:
                cmp_fn(f"{section}[{key}]", g[key], a[key])

    for key in sorted(set(golden["compliance"]) | set(actual["compliance"])):
        _exact(f"compliance.{key}",
               golden["compliance"].get(key), actual["compliance"].get(key))
    return diffs
