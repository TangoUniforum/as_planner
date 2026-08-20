"""Build the REFERENCE FIXTURE: a synthetic facility snapshot the regression
baseline is measured against.

Run it:  python tests/fixtures/build_reference.py

WHY THIS EXISTS
---------------
The suite proves the machinery works; it does not prove the ANSWERS are right.
Demonstrated 2026-08-20: inverting the harvest scheduler from FIFO to LIFO — a
reversal of which batch gets harvested first — passed all 762 tests while the
forecast moved 0.92% on fish and worst tank density went 100 -> 134 kg/m3.
A numeric baseline is what turns "a number changed" into a test failure.

It is deliberately NOT a frozen copy of a production workbook:
  * production workbooks are ~3.2 MB .xlsm and `*.xlsm` is gitignored, so the
    32 tests that need one silently vanish on a clean clone (measured coverage
    77.6% -> 48.4%, run.py 88% -> 3.1%);
  * nobody understands every number in a real workbook, so nobody can say
    whether a diff is a regression or an improvement.

Here every input is chosen on purpose and the choice is written down.

WHAT IS REAL AND WHAT IS SYNTHETIC
----------------------------------
REAL      the facility geometry and the biology tables are copied from the
          live config/ — the operator asked for "real shape", and real
          geometry means the true per-system cap interactions are exercised.
SYNTHETIC the stock, the batches and the limits are authored here in round
          numbers, so the plan can be reasoned about by hand.

The starting state is supplied exactly the way production supplies it — a
ProductionReport sheet with one row per (batch, tank). That is deliberate:
per-tank rows are what carry TANK CONTINUITY, and using the same input path
as the real biomass-software export means this fixture tests the real
mechanism rather than a test-only shortcut.

INPUT CONTRACT NOTES (learned the hard way; see the docstrings in
forecast/production_report.py)
  * "Closing Month: m/d/yyyy" must be a STRING in column A. A real Excel date
    cell is ignored, and only %m/%d/%Y parses — 5/31/26, 31/5/2026 and
    2026-05-31 all silently yield no closing date.
  * The batch row ("Fish group name: Bnn", column C) must appear ABOVE its
    unit rows, and the id must match B\\d+.
  * "Unit: <id>" is column D. The literal replaced is "Unit:" with NO space.
  * Closing Count (col G) must be a NUMBER — a numeric string drops the row.
  * For OG rows the AVG WEIGHT (col K) is authoritative and biomass is
    decorative; for FW rows it is the other way round. We keep them
    consistent anyway.
  * Formulas are read with data_only=True, so only literal values survive.
  * The 6N tank ids are hardcoded in forecast/sixn.py as pairs
    (61,67) (63,69) (65,71) — they are the one part of the layout that is
    not free.
"""
from __future__ import annotations

import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[2]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from forecast.caps import FacilityLimits, SystemLimits          # noqa: E402
from forecast.config_io import (dump_config, load_biology_tables,    # noqa: E402
                                load_control, load_facility_config)
from forecast.models import BatchInput                          # noqa: E402
from forecast.scenario_io import dump_scenario                  # noqa: E402

HERE = Path(__file__).resolve().parent
OUT = HERE / "reference"

# --------------------------------------------------------------------------
# The calendar. PR closes on a month end and the forecast starts the next day,
# which is the production convention (DESIGN §1) and also skips the partial
# first-month merge, keeping the baseline's monthly numbers whole months.
# 2026-06-01 is a Monday, so week 0 is a full 7-day week.
#
# 26 weeks, not 12. The 6N rotation is three pairs with a 14-day depuration
# hold, so a 12-week horizon fits barely three drains per pair and the
# baseline was dominated by fish still sitting in the hold at the end -- 2
# harvest events, most pinned metrics zero or constant, i.e. a weak detector.
# 26 weeks gives the rotation room to reach steady state while still running
# in well under a second.
# --------------------------------------------------------------------------
PR_CLOSING = date(2026, 5, 31)
FORECAST_START = PR_CLOSING + timedelta(days=1)      # 2026-06-01, a Monday
HORIZON_WEEKS = 26

# --------------------------------------------------------------------------
# The stock. One entry per PR row: (batch, tank, fish, avg weight g).
#
# Sized so the 12-week horizon is comfortably FEASIBLE against the sales floor.
# An earlier draft used 390,000 fish and produced 8 harvest events with 5 empty
# weeks — a baseline dominated by shortfall handling rather than by normal
# operation, which is not what a reference run should pin.
#
# Round numbers throughout: 40,000 / 50,000 / 60,000 fish per tank, weights to
# the hundred grams, so every figure below can be checked by hand.
# --------------------------------------------------------------------------
STOCK: list[tuple[str, int, int, float]] = [
    # B01 — mature, harvest-eligible from week 0. 6 tanks across two systems.
    ("B01", 31, 40_000, 3_600.0), ("B01", 33, 40_000, 3_600.0),
    ("B01", 35, 40_000, 3_600.0), ("B01", 41, 40_000, 3_600.0),
    ("B01", 43, 40_000, 3_600.0), ("B01", 45, 40_000, 3_600.0),
    # B02 — mature, the second harvestable batch, so a week needs BOTH tanks
    # of a 6N pair and the sisters engage.
    ("B02", 51, 40_000, 3_400.0), ("B02", 53, 40_000, 3_400.0),
    ("B02", 55, 40_000, 3_400.0), ("B02", 32, 40_000, 3_400.0),
    ("B02", 34, 40_000, 3_400.0), ("B02", 36, 40_000, 3_400.0),
    # B03 — mid grow-out, never harvest-eligible in 12 weeks: the batch that
    # must still be fed and placed around.
    ("B03", 42, 50_000, 2_000.0), ("B03", 44, 50_000, 2_000.0),
    ("B03", 46, 50_000, 2_000.0), ("B03", 52, 50_000, 2_000.0),
    ("B03", 54, 50_000, 2_000.0), ("B03", 56, 50_000, 2_000.0),
    # B04 — entry tier, below the 1 kg lock at start; crosses it during the
    # horizon, so the lock and the forward migration both fire.
    ("B04", 11, 60_000, 900.0), ("B04", 13, 60_000, 900.0),
    ("B04", 15, 60_000, 900.0),
    # B05 — already purging in one complete 6N pair (61 main + 67 sister), so
    # the depuration hold is under test from week 0 rather than week 3.
    ("B05", 61, 20_000, 4_000.0), ("B05", 67, 20_000, 4_000.0),
]
# 23 of 39 OG tanks occupied; OG1S / OG2N / OG2S / OG6S left free so the
# week-4 TranOG arrival has somewhere to land that nothing else competes for.
# Totals: 1,000,000 fish, 2,602,000 kg — inside the 3,600,000 kg facility cap,
# and 520,000 fish are above min_harvest_weight_g, comfortably more than the
# 12 x 30,000 = 360,000 the sales floor needs over the horizon.
# Densities at start: 40,000 x 3.6 kg / 1720 m3 = 83.7; 50,000 x 2.0 = 58.1;
# 60,000 x 0.9 = 31.4; 20,000 x 4.0 = 46.5 — all under the 95 cap with room
# for growth, so nothing is illegal on day one.

# --------------------------------------------------------------------------
# Batches. The four already in the facility carry a tran_og_date in the PAST,
# which is how forecast/biology.project_all_batches knows to take them from the
# ProductionReport instead of re-projecting them as incoming lifecycles.
# B06 and B07 arrive DURING the horizon and are the only projected entries.
# --------------------------------------------------------------------------
def _dt(y, m, d) -> datetime:
    return datetime(y, m, d)


BATCHES: list[BatchInput] = [
    # id     input        count    tran_sf      tran_og      og_count  og_wt  cv
    BatchInput("B01", _dt(2024, 11, 4), 1_400_000, _dt(2025, 2, 3), _dt(2025, 5, 5),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "mature grow-out"),
    BatchInput("B02", _dt(2024, 12, 30), 1_400_000, _dt(2025, 3, 31), _dt(2025, 6, 30),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "mature grow-out"),
    BatchInput("B03", _dt(2025, 4, 28), 1_400_000, _dt(2025, 7, 28), _dt(2025, 10, 27),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "mid grow-out"),
    BatchInput("B04", _dt(2025, 8, 25), 1_400_000, _dt(2025, 11, 24), _dt(2026, 2, 23),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "entry tier"),
    BatchInput("B05", _dt(2024, 9, 2), 1_400_000, _dt(2024, 12, 2), _dt(2025, 3, 3),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "in 6N depuration"),
    # --- incoming during the horizon ---
    # The FW leg runs input -> tran_sf (~3 months to start-feeding) -> tran_og
    # (~9 months total). A shorter leg cannot reach the 370 g transfer target
    # and the FW auto-calibration then clamps at auto_calibrate_fw_max, which
    # shows up as a large residual and makes the baseline look broken.
    BatchInput("B06", _dt(2025, 9, 22), 1_400_000, _dt(2025, 12, 22), _dt(2026, 6, 22),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "TranOG in week 4"),
    BatchInput("B07", _dt(2025, 11, 17), 1_400_000, _dt(2026, 2, 17), _dt(2026, 8, 17),
               1_260_000, 370.0, 16.0, "FCR_118_Quick", 1.0, 1.0, "TranOG past horizon"),
]


def _control():
    """Start from the live control so the fixture is representative, then pin
    the handful of values the baseline's numbers actually hinge on."""
    c = load_control(_ROOT / "config")
    c.scenario_name = "Reference fixture"
    c.forecast_start = FORECAST_START
    c.horizon_weeks = HORIZON_WEEKS
    # Round, legible operating limits.
    #
    # max_biomass_kg is set just ABOVE the fixture's 2,602,000 kg opening
    # stock, because the harvest controller works off a BAND around the cap
    # (+/- facility_biomass_deviation_pct). A facility that opens far below
    # its band is not "safe", it is idle: the controller harvests only the
    # contract floor and waits to grow into the band. An earlier draft with a
    # 3,600,000 cap produced 2 harvest events in 12 weeks for exactly this
    # reason. Opening AT the band is what a real Bluehouse week looks like and
    # is what makes the baseline exercise normal operation.
    c.max_biomass_kg = 2_650_000.0        # facility ceiling, ~2% over opening
    c.max_feed_per_day_kg = 27_500.0      # deliberately NOT binding
    c.max_harvest_per_week = 55_000.0     # processing limit
    c.min_harvest_per_week = 30_000.0     # sales contract floor
    c.min_harvest_weight_g = 3_300.0      # F01/F02/F05 eligible; F03/F04 not
    c.default_hog_yield = 0.81
    c.handling_mortality_pct = 0.01
    c.density_target_pct = 0.80
    c.tran_og_default_tanks = 3
    # 6N stays in PURGE mode for the whole horizon: production starts well
    # past the 12-week window, so the fixture tests the depuration pipeline
    # and not the transition machinery.
    c.sixn_growth = False
    c.sixn_production_start = date(2028, 1, 1)
    return c


def _limits(facility):
    """One binding per-system cap, everything else slack.

    OG3N holds 3 tanks of B01 = 3 x 40,000 x 3.6 kg = 432,000 kg. Its physical
    ceiling is 3 x 1720 x 95 = 490,200 kg, so the cap is set to 400,000 — just
    under the opening load (1.08x). That is enough to make the realized
    rebalancer act in week 0 without being so far over that the week is
    infeasible and the baseline becomes a study of failure handling. Every
    other system sits at its physical ceiling, so exactly one biomass cap
    binds and any breach in the baseline is unambiguous.
    """
    og_systems = sorted({t.system_id for t in facility.tanks if t.type == "OG"})
    defaults: dict[tuple[str, str], float] = {}
    for sysid in og_systems:
        tanks = [t for t in facility.tanks if t.system_id == sysid]
        physical = sum(t.volume_m3 * t.max_density_kg_m3 for t in tanks)
        feed = sum(t.max_feed_kg_day for t in tanks)
        defaults[(sysid, "biomass")] = 400_000.0 if sysid == "OG3N" else physical
        defaults[(sysid, "feed_per_day")] = feed
    return (FacilityLimits(overrides={}),
            SystemLimits(caps={}, defaults=defaults, mode_defaults={},
                         sixn_growth=False,
                         sixn_production_start=date(2028, 1, 1)))


def _write_production_report(path: Path) -> None:
    """One row per (batch, tank), grouped under a batch row, under the closing
    banner — the exact shape forecast/production_report.py parses."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProductionReport"

    ws.cell(1, 6, "Stock")
    for col, label in [(6, "Opening Count"), (7, "Closing Count"),
                       (8, "Opening Biomass [kg]"), (9, "Closing Biomass [kg]"),
                       (10, "Opening Avg weight"), (11, "Closing Avg weight")]:
        ws.cell(2, col, label)

    tot_n = sum(s[2] for s in STOCK)
    tot_kg = sum(s[2] * s[3] / 1000.0 for s in STOCK)
    r = 3
    # Column A carries the closing banner as a STRING in m/d/yyyy — an actual
    # date cell is ignored by the parser.
    ws.cell(r, 1, f"Closing Month: {PR_CLOSING.month}/{PR_CLOSING.day}/{PR_CLOSING.year}")
    ws.cell(r, 7, tot_n); ws.cell(r, 9, round(tot_kg, 1))
    ws.cell(r, 11, round(tot_kg * 1000.0 / tot_n, 1))
    r += 1
    ws.cell(r, 2, "Site: Reference"); ws.cell(r, 7, tot_n)
    ws.cell(r, 9, round(tot_kg, 1)); ws.cell(r, 11, round(tot_kg * 1000.0 / tot_n, 1))
    r += 1

    for batch in sorted({s[0] for s in STOCK}):
        rows = [s for s in STOCK if s[0] == batch]
        bn = sum(x[2] for x in rows)
        bkg = sum(x[2] * x[3] / 1000.0 for x in rows)
        # The batch row MUST precede its unit rows — a Unit row with no
        # current batch is dropped silently.
        ws.cell(r, 3, f"Fish group name: {batch}")
        ws.cell(r, 7, bn); ws.cell(r, 9, round(bkg, 1))
        ws.cell(r, 11, round(bkg * 1000.0 / bn, 1))
        r += 1
        for _b, tank, n, wt in rows:
            ws.cell(r, 4, f"Unit: {tank}")
            ws.cell(r, 6, n)                      # opening (decorative)
            ws.cell(r, 7, n)                      # closing count — must be NUMERIC
            ws.cell(r, 8, round(n * wt / 1000.0, 1))
            ws.cell(r, 9, round(n * wt / 1000.0, 1))   # closing biomass kg
            ws.cell(r, 10, wt)
            ws.cell(r, 11, wt)                    # closing avg weight g — authoritative for OG
            r += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    facility = load_facility_config(_ROOT / "config")   # REAL geometry
    tables = load_biology_tables(_ROOT / "config")      # REAL biology
    control = _control()
    fac_lim, sys_lim = _limits(facility)

    cfg_dir = OUT / "config"
    scn_dir = OUT / "scenario"
    cfg_dir.mkdir(parents=True, exist_ok=True)
    scn_dir.mkdir(parents=True, exist_ok=True)

    dump_config(cfg_dir, control=control, tables=tables, facility=facility)
    dump_scenario(scn_dir, batches=BATCHES,
                  facility_limits=fac_lim, system_limits=sys_lim)
    _write_production_report(OUT / "production_report.xlsx")

    og = [t for t in facility.tanks if t.type == "OG"]
    print(f"reference fixture written to {OUT}")
    print(f"  facility : {len(facility.tanks)} tanks "
          f"({len(og)} OG across {len({t.system_id for t in og})} systems, real geometry)")
    print(f"  biology  : {len(tables.sgr_size_g)} SGR rows, "
          f"FCR models {sorted(tables.fcr_by_model)}")
    print(f"  batches  : {len(BATCHES)}")
    print(f"  PR       : {len(STOCK)} tank rows, "
          f"{sum(s[2] for s in STOCK):,} fish, "
          f"{sum(s[2]*s[3]/1000 for s in STOCK):,.0f} kg, closing {PR_CLOSING}")
    print(f"  horizon  : {HORIZON_WEEKS} weeks from {FORECAST_START}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
