"""Weekly/MonthlyReport `Peak_Density` — the column that used to be a literal 0.

`_ledger_value_cells` wrote the constant `0` into the density column of EVERY
row of both reports. A column of zeros does not read as "no data", it reads as
"density is fine" — the single most consequential thing this project's reports
can say wrongly, since per-tank density over-cap is the standing failure mode.

It is now the WORST tank the batch occupied in the period (max over its
BatchLocations rows), blank when the batch held no tank at all. These pin that
the number is real, that it is a max and not a mean, that it rolls up to months
as a max, and that the blank case stays blank.
"""
from __future__ import annotations

from datetime import date

import openpyxl

from forecast import excel_io
from forecast.placement import BatchLocationRow

_MON = date(2026, 8, 3)          # 2026-W32
_MON2 = date(2026, 8, 10)        # 2026-W33
_DENS_COL = _LEDGER_IDX = excel_io._LEDGER_COLS.index("Peak_Density (kg/m³)")


def _loc(week, wkstart, batch, tank, count, wt_g, bio, dens):
    return BatchLocationRow(
        week_label=week, week_start=wkstart, batch_id=batch, tank_id=tank,
        location_id=f"OG3-{tank}", system_id="OG3", count=count,
        avg_wt_g=wt_g, biomass_kg=bio, density_kg_m3=dens)


_LOCS = [
    # B1, W32: three tanks — one of them well over the 95 cap.
    _loc("2026-W32", _MON, "B1", 40, 10_000, 4000.0, 40_000.0, 80.0),
    _loc("2026-W32", _MON, "B1", 41, 10_000, 4000.0, 40_000.0, 142.5),
    _loc("2026-W32", _MON, "B1", 42, 10_000, 4000.0, 40_000.0, 60.0),
    # B1, W33: back inside the cap.
    _loc("2026-W33", _MON2, "B1", 40, 15_000, 4200.0, 63_000.0, 88.0),
    _loc("2026-W33", _MON2, "B1", 41, 15_000, 4200.0, 63_000.0, 91.0),
]


def _rows():
    return excel_io._build_batch_week_ledger(_LOCS, [], [])


def _sheet_rows(ws):
    """(header row, data rows) — the ledger block starts at the header line."""
    all_rows = [r for r in ws.iter_rows(values_only=True)]
    hi = next(i for i, r in enumerate(all_rows)
              if r and r[0] == "Scenario")
    return all_rows[hi], [r for r in all_rows[hi + 1:] if r and r[0] is not None]


class TestPeakDensityIsReal:
    def test_the_column_is_no_longer_a_constant_zero(self):
        """Negative control: on the parent commit every value cell here is 0."""
        vals = [d["peak_density"] for d in _rows()]
        assert any(v for v in vals), "Peak_Density is still all zero/blank"

    def test_it_is_the_worst_tank_not_the_mean(self):
        """A mean (94.2 here) sits comfortably under the 95 cap while one tank
        runs at 142.5. Reporting the mean would hide exactly the breach the
        column exists to expose."""
        by_week = {(d["batch"], d["week"]): d["peak_density"] for d in _rows()}
        assert by_week[("B1", "2026-W32")] == 142.5
        assert by_week[("B1", "2026-W33")] == 91.0

    def test_a_batch_with_no_tank_rows_reads_blank_not_zero(self):
        """No tank that week (a freshwater week carried by the projection) must
        be an EMPTY cell. 0 would say "density is fine"."""
        from forecast.models import BatchWeekState
        s = BatchWeekState(
            batch_id="B9", week_label="2026-W32", week_start=_MON,
            days_since_input=7, week_from_input=1, count=5000.0,
            avg_weight_g=5.0, biomass_kg=25.0, feed_kg_day=1.0,
            feed_kg_week=7.0, sgr_pct_day=3.0, fcr=0.9, stage="FW",
            feed_type="FW", mortality_pct_weekly=0.1)
        rows = excel_io._build_batch_week_ledger(_LOCS, [], [s])
        fw = next(d for d in rows if d["batch"] == "B9")
        assert fw["peak_density"] is None
        assert excel_io._ledger_value_cells(fw)[_DENS_COL] is None


class TestItReachesBothSheets:
    def test_weekly_report_writes_the_peak_and_labels_it(self):
        wb = openpyxl.Workbook()
        excel_io.write_weekly_report(wb, _LOCS, [], [])
        ws = wb["WeeklyReport"]
        header, data = _sheet_rows(ws)
        assert header[4 + _DENS_COL] == "Peak_Density (kg/m³)"
        w32 = next(r for r in data if r[1] == "2026-W32")
        assert w32[4 + _DENS_COL] == 142.5
        # The legend must be IN the sheet — a peak is only interpretable if the
        # reader knows a row spans several tanks.
        blurb = "\n".join(str(r[0]) for r in ws.iter_rows(values_only=True)
                          if r and r[0] is not None)
        assert "WORST tank" in blurb and "not a mean" in blurb

    def test_monthly_rolls_the_peak_up_as_a_max_not_a_prorated_flow(self):
        wb = openpyxl.Workbook()
        excel_io.write_monthly_report(wb, _LOCS, [], [])
        ws = wb["MonthlyReport"]
        header, data = _sheet_rows(ws)
        assert header[3 + _DENS_COL] == "Peak_Density (kg/m³)"
        aug = next(r for r in data if r[1] == "2026-08")
        # max(142.5, 91.0) — NOT a sum (233.5) and NOT a day-weighted average.
        assert aug[3 + _DENS_COL] == 142.5
