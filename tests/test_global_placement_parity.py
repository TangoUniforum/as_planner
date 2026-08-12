"""GLOBAL-method constraint parity — negative controls for the placement repair.

The Global engines must respect every constraint the controller family does;
otherwise a compare board is not a comparison. This suite guards the repairs
made 2026-08-11, each one a constraint that was ABSENT rather than merely
mis-tuned. Policy is the project's: a check exists to DETECT a defect, and a
check that cannot fire is itself a defect — so every constraint here gets a
control proving it BINDS plus a positive control proving it stays quiet on
clean input.

Measured on the operator's 7.29 PR + scenario/manual_events/2026-07-31.yaml:
  * 33-vs-36: `production_tanks_per_system` / `production_systems_for_week`
    were CORRECT but had ZERO call sites — `plan_l3` used the mode-blind
    `n_tanks_per_system`, so production-era weeks were placed into a facility
    3 tanks smaller than the real one. Cost: B65+B66 = 1,140,000 fish dropped.
  * never-drop: a batch L1 seeded but the tank pick could not place simply
    VANISHED from BatchLocations while the batch-level ReconciliationReport
    still called it standing and conserving (continuity audits are blind to a
    batch that was never placed). 570,000 fish at 2028-W49, with 13 of 39
    tanks empty and 400,000 kg of headroom.
  * STARVE-is-a-state: the 6N stage stamp keyed off the TANK ID, so once the
    mains became production grow-out their fish would have been stamped
    off-feed depuration — hiding them from every density/welfare metric (all
    of which exclude purge rows) and corrupting the depuration-hold audit.
"""
import os
import sys
from datetime import datetime

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from forecast import excel_io, window_weeks                       # noqa: E402
from forecast.global_planner_l3_poc import (                      # noqa: E402
    n_tanks_per_system,
    production_systems_for_week,
    production_tanks_per_system,
)
from forecast.models import ControlParams, FacilityConfig, TankConfig  # noqa: E402
from forecast.sixn import SIXN_MAIN_TANKS, SIXN_SISTER_TANKS      # noqa: E402


# --------------------------------------------------------------------------- #
# A facility shaped like the real one: 11 production systems x 3 tanks, plus
# OG6N's 3 mains + 3 sisters.
# --------------------------------------------------------------------------- #
def _facility():
    tanks = []
    for grp, sysbase in enumerate(["OG1", "OG2", "OG3", "OG4", "OG5"], start=1):
        for i in range(6):
            tid = grp * 10 + 1 + i
            sysid = sysbase + ("N" if i % 2 == 0 else "S")
            tanks.append(TankConfig(f"{sysid}-{tid}", sysid, tid, 1720.0, 95.0,
                                    3000.0, "OG"))
    for tid in sorted(set(SIXN_MAIN_TANKS) | set(SIXN_SISTER_TANKS)):
        tanks.append(TankConfig(f"OG6N-{tid}", "OG6N", tid, 1720.0, 95.0,
                                3000.0, "OG"))
    # OG6S: the 7th grow-out system (3 tanks), as on the real facility.
    for i, tid in enumerate((66, 68, 70)):
        tanks.append(TankConfig(f"OG6S-{tid}", "OG6S", tid, 1720.0, 95.0,
                                3000.0, "OG"))
    return FacilityConfig(tanks=tanks)


def _control(prod_start="2028-01-01"):
    c = ControlParams(
        forecast_start=datetime(2026, 8, 3), horizon_weeks=130,
        scenario_name="t", max_feed_per_day_kg=34000.0, max_biomass_kg=3.8e6,
        max_harvest_per_week=55000.0, min_harvest_weight_g=3500.0,
        min_harvest_per_week=30000.0, min_tank_control=7000.0,
        default_hog_yield=0.81, facility_biomass_deviation_pct=0.005,
        handling_mortality_pct=0.01, sixn_growth=False)
    c.sixn_production_start = datetime.fromisoformat(prod_start)
    return c


PURGE_WK, PROD_WK = "2027-W20", "2028-W20"


class TestModeAwareTankCount:
    """33 tanks in purge era, 36 in production era — and OG6N's 3 SISTERS are
    never production in either mode (they are harvest staging)."""

    def test_purge_week_has_no_production_6n_tanks(self):
        n = production_tanks_per_system(_facility(), _control(), PURGE_WK)
        assert n.get("OG6N", 0) == 0
        assert sum(n.values()) == 33
        assert "OG6N" not in production_systems_for_week(
            _facility(), _control(), PURGE_WK)

    def test_production_week_adds_the_three_6n_mains(self):
        n = production_tanks_per_system(_facility(), _control(), PROD_WK)
        assert n.get("OG6N", 0) == 3, "the 3 MAINS become grow-out production"
        assert sum(n.values()) == 36
        assert "OG6N" in production_systems_for_week(
            _facility(), _control(), PROD_WK)

    def test_sisters_are_never_production_in_either_mode(self):
        """The 3 sisters exist so two batches can stage for harvest the same
        week without co-mingling; counting them as grow-out would invent
        capacity that physically is not there."""
        f = _facility()
        for wk in (PURGE_WK, PROD_WK):
            n = production_tanks_per_system(f, _control(), wk)
            assert n.get("OG6N", 0) <= len(SIXN_MAIN_TANKS)

    def test_the_mode_blind_count_really_is_different(self):
        """NEGATIVE CONTROL for the whole fix: if the mode-blind helper agreed
        with the mode-aware one, wiring it would be a no-op and this suite
        would be proving nothing. It does not — it over-counts by the 3
        sisters and is flat across the mode boundary."""
        f = _facility()
        raw = n_tanks_per_system(f)
        assert raw.get("OG6N", 0) == 6            # mains + sisters, no mode
        assert (production_tanks_per_system(f, _control(), PURGE_WK)
                != production_tanks_per_system(f, _control(), PROD_WK))


class TestPlanL3UsesTheModeAwareCount:
    """The helpers above were correct for months while having ZERO call sites.
    Correctness is worthless unwired, so guard the WIRING, not just the math."""

    def test_plan_l3_calls_production_tanks_per_system(self, monkeypatch):
        """Spy: plan_l3 must consult the mode-aware count. Patched to raise a
        sentinel, so reaching it proves it is on the placement path."""
        import forecast.global_planner_l3_poc as l3

        class _Reached(Exception):
            pass

        def _spy(*a, **k):
            raise _Reached()

        monkeypatch.setattr(l3, "production_tanks_per_system", _spy)

        class _L1:
            batch_standing = [
                l3.BatchStandingRow(week=0, week_label=PROD_WK, batch_id="B1",
                                    count=100000.0, biomass_kg=100000.0,
                                    avg_wt_g=1000.0, feed_kg_day=500.0),
            ]

        with pytest.raises(_Reached):
            l3.plan_l3(_L1(), _control(), _facility(), None)

    def test_the_spy_would_not_fire_without_the_wiring(self, monkeypatch):
        """Positive control for the spy itself: patching the OTHER (mode-blind)
        helper must NOT raise, so the test above is detecting the specific
        call it claims to."""
        import forecast.global_planner_l3_poc as l3

        class _Reached(Exception):
            pass

        monkeypatch.setattr(l3, "n_tanks_per_system",
                            lambda *a, **k: (_ for _ in ()).throw(_Reached()))

        class _L1:
            batch_standing = [
                l3.BatchStandingRow(week=0, week_label=PROD_WK, batch_id="B1",
                                    count=100000.0, biomass_kg=100000.0,
                                    avg_wt_g=1000.0, feed_kg_day=500.0),
            ]
        try:
            l3.plan_l3(_L1(), _control(), _facility(), None)
        except _Reached:
            pytest.fail("plan_l3 still routes its tank capacity through the "
                        "mode-blind n_tanks_per_system")
        except Exception:
            pass          # any other failure is fine; we only assert the route


class TestEntryTierIsNotAClosedBox:
    """R2 forward relief — the whole of the remaining density breach.

    Measured on the operator's 7.29 PR: 74 of 75 tank-weeks over the 95 kg/m3
    cap were sub-1 kg fish crammed into OG1/OG2 (up to 187.4 kg/m3) while ~8
    grow-out tanks sat free AND spread-eligible in every one of those weeks.
    The cause was neither cap headroom nor tank availability (both hypotheses
    were tested and refuted) but a tier lock in the Global code that is
    STRICTER than the operator's own rule module: tiers.R2 allows an entry-tier
    cohort to move forward to any OG3/4/5/6 tank AT ANY WEIGHT.
    """

    def test_the_rule_module_permits_forward_moves_at_any_weight(self):
        """The authority. If this ever changes, the relief below is illegal and
        must change with it — which is why it is asserted, not assumed."""
        from forecast.tiers import move_allowed
        for wt in (200.0, 500.0, 999.0, 1500.0):
            ok, why = move_allowed("OG1N", "OG3N", wt)
            assert ok, f"R2 forward move refused at {wt} g: {why}"

    def test_backward_is_still_forbidden_at_every_weight(self):
        """NEGATIVE CONTROL: R2 must not be read as "movement is free". R4 is
        untouched — nothing may come back into the entry tier."""
        from forecast.tiers import move_allowed
        for wt in (200.0, 999.0, 5000.0):
            ok, _ = move_allowed("OG3N", "OG1N", wt)
            assert not ok, f"backward move wrongly allowed at {wt} g"

    def test_intra_entry_move_still_locked_above_1kg(self):
        """R3 likewise untouched."""
        from forecast.tiers import move_allowed
        assert move_allowed("OG1N", "OG2N", 500.0)[0]
        assert not move_allowed("OG1N", "OG2N", 1500.0)[0]

    def test_l3_offers_a_nursery_batch_the_forward_systems(self):
        """WIRING: L3 and the tank pick must AGREE that a nursery batch may sit
        in grow-out. When only the pick allowed it, L3 pulled the batch back the
        next week and the relief was re-emitted as a BACKWARD move."""
        import inspect
        from forecast import global_planner_l3_poc as l3
        src = inspect.getsource(l3.plan_l3)
        assert "if d.tier == TIER_NURSERY:" in src
        assert "fwd = [s for s in GROWOUT_SYSTEMS if s in sys_set]" in src

    def test_the_pick_agrees_and_still_refuses_to_go_backward(self):
        import inspect
        from forecast import global_tank_pick_poc as tp
        src = inspect.getsource(tp.pick_tanks)
        assert "(nurs_sys + grow_sys)" in src          # forward relief exists
        assert "not is_entry(tank_sys.get(t," in src           # R4 guard exists

    def test_r4_is_enforced_monotonically_on_the_main_placement_pass(self):
        """Once ANY of a batch's fish leave the entry tier, no part of it may be
        sent back. L3 plans in system COUNTS and can legally hand a batch back to
        a nursery system a week later; realizing that emits a grow-out -> entry
        move, which R4 forbids at any weight. Measured: this guard cut emitted
        topology violations from 272 to 251."""
        import inspect
        from forecast import global_tank_pick_poc as tp
        src = inspect.getsource(tp.pick_tanks)
        assert "if (is_entry(system)" in src


class TestTransferTopologyIsJudgedAndSurfaced:
    """R1-R7 conformance of the EMITTED transfer stream.

    The controller family emits ZERO topology violations; Global emitted 208
    even before this repair series and no gate had ever measured it. A plan that
    plays by different movement rules is not comparable to one that does not,
    which is the whole point of the compare board.
    """

    def test_the_pick_judges_pairings_with_the_shared_rule_module(self):
        """Both families must be held to identical code — forecast.tiers — so a
        violation means the same thing on either side of the board."""
        import inspect
        from forecast import global_tank_pick_poc as tp
        assert "move_allowed" in inspect.getsource(tp.pick_tanks)

    def test_retention_prefers_the_forward_tier(self):
        """A batch that straddles tiers must keep its FORWARD tanks and release
        entry ones, so any consolidation runs entry -> grow-out (R2, legal at
        any weight) rather than grow-out -> entry (R4, never legal). Measured:
        251 -> 53 emitted violations."""
        import inspect
        from forecast import global_tank_pick_poc as tp
        src = inspect.getsource(tp.pick_tanks)
        assert 'key=lambda t: (is_entry(tank_sys.get(t, "")), t)' in src

    def test_a_breach_that_cannot_be_avoided_is_reported_not_hidden(self):
        """Conservation wins when no legal source exists (the fish must come
        from somewhere), so the move is emitted — but it must be recorded as an
        ERROR, never dropped silently."""
        from forecast.excel_io import write_validation_log
        wb = openpyxl.Workbook()
        write_validation_log(wb, invariant_warnings=[
            "TOPOLOGY VIOLATION - 2026-W34: batch B45 OG6S-64 -> OG1N-11 at "
            "1958 g. R4: backward move OG6S->OG1N"])
        rows = [r for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert rows[0][1] == "ERROR - Topology violation (R1-R7)"

    def test_the_rule_module_is_the_authority_for_both_families(self):
        """NEGATIVE CONTROL: the checker must actually reject the two shapes we
        measured, or 'zero violations' would be meaningless."""
        from forecast.tiers import move_allowed
        assert not move_allowed("OG6S", "OG1N", 1958.0)[0]   # R4 backward
        assert not move_allowed("OG1N", "OG1S", 1500.0)[0]   # R3 intra-entry
        assert move_allowed("OG1N", "OG3N", 500.0)[0]        # R2 forward, legal


class TestCpSatWasOverConstrainedNotOverFull:
    """R6 in the CP-SAT model — the cause of the 103/127 infeasible weeks.

    R6: ">= 1 kg fish MAY remain in entry-tier tanks (stuck-in-place is legal;
    the >= 1 kg overflow in OG1/2 is measured-necessary -- never force-evict)."
    CP-SAT barred them outright, so all heavy biomass had to fit the grow-out
    tanks alone: 106 of 130 weeks then needed MORE grow-out tanks than exist
    while the entry tier sat at 1-3 of its 12. Respecting R6 took infeasible
    weeks 103 -> 36 and solver slack 4,878,644 -> 210,481 kg.

    Proof it was never a time budget: the infeasible count was EXACTLY 103 at
    both a 6 s and a 30 s per-week deterministic budget (802 s vs 2034 s total).
    """

    def test_r6_permits_heavy_fish_to_remain_in_the_entry_tier(self):
        """The rule the model was contradicting. tiers has no prohibition on a
        heavy batch OCCUPYING an entry tank — only on MOVING into one (R4)."""
        from forecast.tiers import move_allowed
        # staying put is not a move at all; moving backward is what is banned
        assert not move_allowed("OG3N", "OG1N", 5000.0)[0]
        # and forward is always fine
        assert move_allowed("OG1N", "OG3N", 5000.0)[0]

    def test_the_model_offers_a_heavy_batch_only_the_entry_tank_it_already_holds(self):
        """R6 without breaking R4: occupancy of a RETAINED entry tank is legal,
        acquiring a new one is not. The eligibility must be conditioned on the
        previous occupant being this same batch."""
        import inspect
        from forecast import global_placement_milp_poc as mp
        src = inspect.getsource(mp.solve_cpsat_perweek)
        assert "prev_tb.get(t) == b" in src
        assert "og_w[t] in nset" in src

    def test_the_solver_is_seeded_with_real_starting_occupancy(self):
        """THE remaining 36 infeasible weeks. R6 lets a >= 1 kg batch KEEP an
        entry tank it already occupies, but `prev_tb` started EMPTY, so in week
        0 no batch occupied anything and the allowance could never engage —
        heavy fish were confined to the 21/24 grow-out tanks, and in 40 weeks
        they need more than that while the entry tier sits at 2-4 of its 12.
        Seeding from the facility's actual handoff state took CP-SAT from 36
        proven-INFEASIBLE weeks to 0, with 0 kg of cap slack."""
        import inspect
        from forecast import global_placement_milp_poc as mp
        sig = inspect.signature(mp.solve_cpsat_perweek)
        assert "initial_tb" in sig.parameters
        src = inspect.getsource(mp.solve_cpsat_perweek)
        assert "prev_tb: dict = dict(initial_tb or {})" in src

    def test_unplaced_weeks_are_reported_by_solver_verdict(self):
        """INFEASIBLE and UNKNOWN mean opposite things — one says a constraint
        forbids every layout (more time is useless), the other says the solver
        found nothing in budget. Collapsing them hid which fix was needed."""
        import inspect
        from forecast import global_placement_milp_poc as mp
        assert "_st_counts" in inspect.getsource(mp.solve_cpsat_perweek)
        from tools.run_global_forecast import cpsat_degrade_warning
        w = cpsat_degrade_warning({"n_weeks": 127, "n_infeasible": 36,
                                   "unplaced_status": {"INFEASIBLE": 36}})
        assert "INFEASIBLE" in w
        assert cpsat_degrade_warning({"n_weeks": 127, "n_infeasible": 0}) == ""

    def test_light_fish_are_still_barred_from_grow_out_entry_rules_intact(self):
        """NEGATIVE CONTROL: the relaxation is for HEAVY fish keeping an entry
        tank only. A sub-1 kg batch must still be offered nursery tanks."""
        from forecast.global_placement_milp_poc import _eligible_tanks
        og = {11: "OG1N", 31: "OG3N"}
        assert _eligible_tanks(500.0, og, {"OG3N"}, {"OG1N"}) == [11]
        assert _eligible_tanks(5000.0, og, {"OG3N"}, {"OG1N"}) == [31]


class TestSolvesAreReproducible:
    """A non-reproducible measurement is a measurement bug (project law).

    `time_limit` is WALL CLOCK, so whichever incumbent branch-and-bound held
    when the clock ran out became the plan: the same PR gave 54 / 55 / 81
    idle-tank weeks purely with CPU contention, which silently poisons every
    A/B built on top of it. A bigger budget is NOT the fix -- measured, 9
    Pass A.2 weeks still bound at 120 s and a run took 21 minutes instead of 3.
    """

    def test_only_a_proved_solve_is_used(self):
        """The rule that makes the output load-independent: a limit-bound
        incumbent is discarded for a reproducible fallback."""
        import inspect
        from forecast import global_planner_l3_poc as l3
        src = inspect.getsource(l3._solve_passA_per_week)
        assert '_r2_ok = getattr(r2, "status", None) == 0' in src
        srcb = inspect.getsource(l3._solve_passB_per_week)
        assert 'if getattr(res, "status", None) != 0:' in srcb

    def test_symmetry_breaking_is_applied_between_interchangeable_systems(self):
        """L3's variables are per-SYSTEM tank counts, so identical tanks are
        already collapsed — the symmetry that matters is between SYSTEMS, which
        on this facility are identical (3 tanks, 400,000 kg, 3,000 kg/day) and
        carry no continuity term in Pass A. Breaking it took unprovable weeks
        from 20-24 to 2 and halved the runtime."""
        import inspect
        from forecast import global_planner_l3_poc as l3
        src = inspect.getsource(l3._solve_passA_per_week)
        assert "SYMMETRY BREAKING" in src
        assert "_classes" in src

    def test_symmetry_breaking_only_removes_permutations(self):
        """SAFETY: the constraint orders interchangeable systems by load. Any
        solution can be relabelled to satisfy it, so the OPTIMUM is unchanged —
        it must never be applied across systems that differ in tier, tank count
        or caps, or it would forbid real solutions."""
        import inspect
        from forecast import global_planner_l3_poc as l3
        src = inspect.getsource(l3._solve_passA_per_week)
        # the class key must include tier, tank count AND caps
        assert "s in NURSERY_SYSTEMS" in src
        assert "n_tanks_w.get(w, {}).get(s, 0)" in src
        assert "_caps_here.get(s)" in src

    def test_symmetry_breaking_is_confined_to_the_stateless_pass(self):
        """The operator's objection: identical systems are only interchangeable
        if they are also in the same STATE, or the "relabelling" is really a
        demand to move fish. It holds in general and not here, because the
        constraint lives ONLY in Pass A, which is stateless — and Pass B, the
        pass that knows physical occupancy, is free to permute back. If anyone
        ever adds the constraint to Pass B, or teaches Pass A about occupancy,
        this guard fails and the soundness argument has to be re-made."""
        import inspect, re
        from forecast import global_planner_l3_poc as l3
        a = inspect.getsource(l3._solve_passA_per_week)
        b = inspect.getsource(l3._solve_passB_per_week)
        assert "SYMMETRY BREAKING" in a
        assert "SYMMETRY BREAKING" not in b, (
            "Pass B knows occupancy — ordering there would force real moves")
        # Pass A must remain stateless: no prior-occupancy inputs in its model.
        code = "\n".join(ln for ln in a.splitlines()
                         if not ln.strip().startswith("#"))
        for token in ("prev_state", "last_sys", "prev_by_batch", "initial_state"):
            assert token not in code, f"Pass A now sees {token}; symmetry class is unsound"
        # Pass B must still carry the stickiness term it permutes back with.
        assert "last_sys" in b

    def test_every_degrade_is_recorded(self):
        from forecast.excel_io import write_validation_log
        wb = openpyxl.Workbook()
        write_validation_log(wb, invariant_warnings=[
            "NON-DETERMINISTIC SOLVE - Pass A.1 hit its wall-clock time limit on week 2027-W04.",
            "PASS A.2 FALLBACK - 2 week(s) could not PROVE the cap-slack refinement.",
            "PASS B FALLBACK - 18 week(s) kept Pass A's layout.",
        ])
        cats = [r[1] for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert cats[0] == "ERROR - Non-reproducible solve"
        assert cats[1].startswith("WARNING - Pass A.2 fallback")
        assert cats[2].startswith("WARNING - Pass B fallback")


class TestUnplacedBatchIsLoud:
    """Fish with L1 standing but no physical tank must be IMPOSSIBLE to miss.
    They previously vanished while conservation still reported them standing."""

    def test_unplaced_batch_files_as_an_error_not_a_hydration_note(self):
        wb = openpyxl.Workbook()
        excel_io.write_validation_log(wb, invariant_warnings=[
            "UNPLACED BATCH - 2028-W49: batch B66 (570,000 kg) has L1 standing "
            "but NO legal free tank in its tier (grow-out); it is absent from "
            "BatchLocations."])
        rows = [r for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert len(rows) == 1
        assert rows[0][1].startswith("ERROR"), \
            f"an unplaced batch filed as {rows[0][1]!r}"
        assert "Unplaced batch" in rows[0][1]

    def test_unplaced_warning_cannot_be_read_as_a_manual_window_week(self):
        """It carries an ISO week label, so it MUST NOT also carry the manual
        markers — a planner week wrongly tagged 'window' is excluded from the
        harvest-compliance gates, hiding breaches in the degraded run."""
        wb = openpyxl.Workbook()
        excel_io.write_validation_log(wb, invariant_warnings=[
            "UNPLACED BATCH - 2028-W49: batch B66 has L1 standing but no tank.",
            "MANUAL EVENT OK - 2026-W31: harvested 21,812 fish",
        ])
        assert window_weeks.manual_window_weeks(wb) == {"2026-W31"}

    def test_the_pick_exposes_its_unplaced_list(self):
        """The pick's result must CARRY the misses so the caller can record
        them; a stdout-only degrade is invisible to the graders."""
        from forecast.global_tank_pick_poc import TankPickResult
        r = TankPickResult(
            batch_locations=[], transfers=[], tranog_events=[],
            harvest_events=[], realized_biology={}, mort_states=[],
            n_transfers=0, n_oversub_rows=0, oversub_weeks=[], n_tank_weeks=0)
        assert r.unplaced_warnings == []       # clean run stays quiet
        r2 = TankPickResult(
            batch_locations=[], transfers=[], tranog_events=[],
            harvest_events=[], realized_biology={}, mort_states=[],
            n_transfers=0, n_oversub_rows=0, oversub_weeks=[], n_tank_weeks=0,
            unplaced_warnings=["UNPLACED BATCH - x"])
        assert r2.unplaced_warnings


class TestPurgeReleaseRespectsTheWeeklyLimit:
    """The weekly processing LIMIT must bind the 6N RELEASE, not only the draw.

    The hold length changes at sixn_production_start (2 purge weeks -> 1 in
    production), so two weeks' draws mature on the same week. Measured on the
    operator's PR: the 2027-W52 draw (hold 2) and the 2028-W01 draw (hold 1)
    both landed on 2028-W02 and released 72,040 fish -- 31% over the 55,000
    limit, 19% over the 60,500 relief ceiling.
    """
    LIMIT = 55_000.0

    def _buf(self):
        # the real collision: two matured cohorts on one week
        return {2: [{"batch_id": "B53", "count": 44_101.0, "biomass_kg": 182_578.0},
                    {"batch_id": "B54", "count": 27_939.0, "biomass_kg": 111_478.0}]}

    def test_the_collision_week_is_capped_at_the_limit(self):
        from forecast.global_planner_poc import release_due_capped
        buf = self._buf()
        out = release_due_capped(buf, 2, self.LIMIT)
        assert sum(c for _, c, _ in out) <= self.LIMIT + 1e-6

    def test_the_excess_is_deferred_not_dropped(self):
        """Conservation is non-negotiable: every fish held back must still be
        in the buffer, and biomass must follow the count on a split cohort."""
        from forecast.global_planner_poc import release_due_capped
        buf = self._buf()
        before_c = sum(e["count"] for v in buf.values() for e in v)
        before_kg = sum(e["biomass_kg"] for v in buf.values() for e in v)
        out = release_due_capped(buf, 2, self.LIMIT)
        after_c = sum(e["count"] for v in buf.values() for e in v)
        after_kg = sum(e["biomass_kg"] for v in buf.values() for e in v)
        assert sum(c for _, c, _ in out) + after_c == pytest.approx(before_c)
        assert sum(k for _, _, k in out) + after_kg == pytest.approx(before_kg)
        assert 3 in buf, "the remainder must wait for the FOLLOWING week"

    def test_a_split_cohort_keeps_its_frozen_mean_weight(self):
        """Held fish are off-feed and frozen at move-in weight; a deferral must
        not silently re-price them."""
        from forecast.global_planner_poc import release_due_capped
        buf = self._buf()
        src = dict(buf[2][1])
        out = release_due_capped(buf, 2, self.LIMIT)
        rel = next((c, k) for b, c, k in out if b == "B54")
        defer = next(e for e in buf[3] if e["batch_id"] == "B54")
        assert rel[1] / rel[0] == pytest.approx(src["biomass_kg"] / src["count"])
        assert (defer["biomass_kg"] / defer["count"]
                == pytest.approx(src["biomass_kg"] / src["count"]))

    def test_an_under_limit_week_is_untouched(self):
        """POSITIVE CONTROL: the cap must not perturb an ordinary week."""
        from forecast.global_planner_poc import release_due_capped
        buf = {2: [{"batch_id": "B1", "count": 30_000.0, "biomass_kg": 120_000.0}]}
        out = release_due_capped(buf, 2, self.LIMIT)
        assert out == [("B1", 30_000.0, 120_000.0)]
        assert 3 not in buf

    def test_no_cap_configured_releases_everything(self):
        """Byte-identical to the pre-fix behaviour when the limit is unset, so
        no existing caller changes."""
        from forecast.global_planner_poc import release_due_capped
        buf = self._buf()
        out = release_due_capped(buf, 2, 0)
        assert sum(c for _, c, _ in out) == pytest.approx(72_040.0)
        assert 3 not in buf

    def test_the_defect_reproduces_without_the_cap(self):
        """NEGATIVE CONTROL: prove the fixture really contains the defect --
        uncapped, this exact week emits the 72,040 the operator saw, 19% over
        the 60,500 relief ceiling. Without this, the cap test proves nothing."""
        from forecast.global_planner_poc import release_due_capped
        uncapped = sum(c for _, c, _ in release_due_capped(self._buf(), 2, 0))
        assert uncapped == pytest.approx(72_040.0)
        assert uncapped > 60_500.0
        capped = sum(c for _, c, _ in release_due_capped(self._buf(), 2, self.LIMIT))
        assert capped <= 60_500.0


class TestPlacementGapIsVisibleAtAll:
    """L1 standing that never reached a tank.

    THE BLIND SPOT: Advisory and SystemLimitsAudit are both WRITTEN FROM
    batch_locations, so checking them against BatchLocations is circular — it
    validates the arithmetic of what was placed and is silent about what was
    left out. (This audit made exactly that mistake first: "Advisory ==
    BatchLocations == SystemLimitsAudit to +/-9 kg" is true and proves nothing
    here.) StandingTrace is L1's own facility standing, so it is the first
    independent witness. On the operator's PR it exposed 18,900,608 kg-weeks of
    unplaced standing that every existing gate reported as clean.
    """

    class _Loc:
        def __init__(self, wl, kg):
            self.week_label, self.biomass_kg = wl, kg

    class _Trace:
        def __init__(self, wl, og, purge=0.0):
            self.week_label = wl
            self.og_biomass_kg, self.purge_biomass_kg = og, purge

    class _Gft:
        def __init__(self, trace, locs):
            self.trace, self.batch_locations = trace, locs

    def test_the_alarm_rings_when_standing_never_reaches_a_tank(self):
        from tools.run_global_forecast import placement_gap_warnings
        gft = self._Gft(
            [self._Trace("2026-W50", 500_000.0)],
            [self._Loc("2026-W50", 165_229.0)])          # 334,771 kg unplaced
        w = placement_gap_warnings(gft)
        assert len(w) == 1 and w[0].startswith("PLACEMENT GAP")
        assert "334,771 kg" in w[0]

    def test_a_fully_placed_plan_is_silent(self):
        """POSITIVE CONTROL — the alarm must not cry wolf on a clean plan."""
        from tools.run_global_forecast import placement_gap_warnings
        gft = self._Gft(
            [self._Trace("2026-W50", 400_000.0, 100_000.0)],
            [self._Loc("2026-W50", 500_000.0)])
        assert placement_gap_warnings(gft) == []

    def test_rounding_dust_is_not_an_alarm(self):
        from tools.run_global_forecast import placement_gap_warnings
        gft = self._Gft([self._Trace("2026-W50", 500_000.0)],
                        [self._Loc("2026-W50", 499_991.0)])   # 9 kg
        assert placement_gap_warnings(gft) == []

    def test_the_purge_pool_counts_as_placeable_standing(self):
        """6N depuration fish occupy real tanks, so they belong on the L1 side
        of the comparison; omitting them would invent a permanent false gap."""
        from tools.run_global_forecast import placement_gap_warnings
        gft = self._Gft([self._Trace("2026-W50", 300_000.0, 200_000.0)],
                        [self._Loc("2026-W50", 500_000.0)])
        assert placement_gap_warnings(gft) == []

    def test_it_files_as_an_error_row(self):
        from tools.run_global_forecast import placement_gap_warnings
        gft = self._Gft([self._Trace("2026-W50", 500_000.0)],
                        [self._Loc("2026-W50", 100_000.0)])
        wb = openpyxl.Workbook()
        excel_io.write_validation_log(
            wb, invariant_warnings=placement_gap_warnings(gft))
        cats = [r[1] for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert cats == ["ERROR - Placement gap (L1 standing unplaced)"]

    def test_missing_inputs_do_not_fake_a_clean_result(self):
        """A workbook without StandingTrace must yield NO claim either way —
        absence of the witness is not evidence of a full placement."""
        from tools.run_global_forecast import placement_gap_warnings
        assert placement_gap_warnings(self._Gft([], [])) == []


class TestStarveIsAStateNotATankId:
    """Once the 6N mains carry production grow-out, stamping them STARVE would
    hide those fish from every density/welfare metric (they all exclude purge
    rows) and corrupt the depuration-hold audit."""

    def test_density_metrics_exclude_starve_rows(self):
        """The reason the stamp matters: a STARVE row is invisible to the
        density peak. If production fish were stamped STARVE, an over-packed
        6N main would read as a clean facility."""
        from forecast.optimize import _is_purge_row
        hdr_si, hdr_sti = 4, 9
        starve = ("2028-W20", None, "B1", 61, "OG6N", 1, 1.0, 1, 500.0, "STARVE")
        grow = ("2028-W20", None, "B1", 61, "OG6N", 1, 1.0, 1, 500.0, "SW")
        assert _is_purge_row(starve, hdr_si, hdr_sti) is True
        assert _is_purge_row(grow, hdr_si, hdr_sti) is False

    def test_stage_is_keyed_off_depuration_not_the_tank_id(self):
        """WIRING guard: the pick must decide STARVE from the set of tanks it
        actually filled with depuration fish, never from `tid in sixn_set`."""
        import inspect
        from forecast import global_tank_pick_poc as tp
        src = inspect.getsource(tp.pick_tanks)
        assert 'stage = "STARVE" if tid in _depurating else ""' in src
        assert 'stage = "STARVE" if tid in sixn_set else ""' not in src
