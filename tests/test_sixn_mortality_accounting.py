"""What the 6N off-feed mortality note actually describes, measured.

The app's Known limits page said:

    "6N off-feed mortality is slightly under-counted per tank — a
     few-fish-per-week approximation that nets out facility-wide and is
     deliberately left (correcting it destabilizes the facility-level balance
     it currently cancels against)."

The review brief asked whether that is an intentional approximation, stale
documentation, or two compensating errors. Measured independently on a real
85-week run -- my arithmetic, the engine's reported rows -- it is the first,
described wrongly:

  * `ReconciliationReport.Count_Delta` is 0 on all 572 rows and the sheet is
    INTERNALLY CONSISTENT: recomputing its identity from its own columns
    reproduces it to under one fish (rounding only). Fish are not being lost.
  * The engine applies NO mortality to fish while off feed in 6N. excel_io
    says so: "STARVE (6N production in-place purge) tanks neither grow nor take
    mortality ... exclude them from the growth + mortality expectation (else
    they read as drift)."
  * `WeeklyReport` nevertheless REPORTS mortality on those batch-weeks -- about
    1,000 fish over the horizon -- that the population never lost. So the two
    sheets disagree, by up to 28 fish, on the same (batch, week), and only in
    the Mortality column: Open, Cull, Harvest, Input and Close match exactly.
  * It does NOT "net out facility-wide": non-6N weeks carry +325 and 6N weeks
    -1,147, leaving -822 across 385 of 1,191 ledger rows. Partial offset, not
    cancellation, and the residual is signed.

So: count conservation is REAL, the modelling choice is deliberate (and two
attempts to change it backfired -- off-feed mortality is load-bearing), and the
defect is a REPORTING one plus a wrong description. This test does not change
the model. It pins the boundary, so that if the discrepancy ever escapes 6N or
grows past what the off-feed rule can explain, it fails here instead of being
absorbed into a mortality term.
"""
import collections
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

WB = os.path.join(
    r"C:\Users\julian.f\AppData\Local\Temp\claude"
    r"\C--Users-julian-f-OneDrive---Atlantic-Sapphire-Production"
    r"-Forecasts-Tool-Python\9713a940-2d28-4a9e-af3d-22a34efbec92\scratchpad",
    "ref_main.xlsm")

pytestmark = pytest.mark.skipif(not os.path.exists(WB),
                                reason="needs a generated output workbook")


@pytest.fixture(scope="module")
def sheets():
    import openpyxl
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)

    def load(name, first):
        ws, hdr, out = wb[name], None, {}
        for r in ws.iter_rows(values_only=True):
            if hdr is None:
                if r and str(r[0]).strip() == first:
                    hdr = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not r or hdr.get("Week") is None or hdr.get("Batch") is None:
                continue
            w, b = r[hdr["Week"]], r[hdr["Batch"]]
            if w and b and str(w).startswith("20"):
                out[(str(b).strip(), str(w).strip())] = (hdr, r)
        return out

    starve = collections.defaultdict(float)
    for r in wb["BatchLocations"].iter_rows(min_row=5, values_only=True):
        if not r or not r[0]:
            continue
        if str(r[9]).strip() == "STARVE":
            starve[(str(r[2]).strip(), str(r[0]).strip())] += float(r[5])
    return (load("ReconciliationReport", "Week"),
            load("WeeklyReport", "Scenario"), starve)


def _g(pair, k):
    hdr, r = pair
    i = hdr.get(k)
    v = r[i] if i is not None and i < len(r) else None
    return float(v) if isinstance(v, (int, float)) else 0.0


def test_count_conservation_is_real_not_a_rounded_headline(sheets):
    """Recompute the reconciliation's identity from its OWN columns. It must
    agree with the Count_Delta it publishes to under one fish -- if a future
    change starts absorbing a real imbalance into the mortality term, the
    published zero and the recomputed identity part company here."""
    rec, _, _ = sheets
    worst = 0.0
    for p in rec.values():
        expected = (_g(p, "Open_Count") - _g(p, "Mortality_Count")
                    - _g(p, "Cull_Count") - _g(p, "Harvest_Count")
                    + _g(p, "Input_Count"))
        worst = max(worst, abs(expected - _g(p, "Actual_Close")))
        assert abs(_g(p, "Count_Delta")) < 0.5
    assert worst <= 1.0 + 1e-9, (
        "the published identity and its own columns disagree by %.2f fish — "
        "more than the whole-fish rounding that explains it" % worst)


def test_the_two_sheets_disagree_only_about_mortality(sheets):
    """Open, Cull, Harvest, Input and Close must match exactly. If a second
    column starts drifting, the off-feed explanation no longer covers it."""
    rec, wkl, _ = sheets
    pairs = [("Open_Count", "Open_Count (fish)"),
             ("Cull_Count", "Cull_Count (fish)"),
             ("Harvest_Count", "Harv_Count (fish)"),
             ("Input_Count", "Input_Count (fish)"),
             ("Actual_Close", "Close_Count (fish)")]
    bad = []
    for k in set(rec) & set(wkl):
        for a, b in pairs:
            if abs(_g(rec[k], a) - _g(wkl[k], b)) > 0.5:
                bad.append((k, a))
    assert not bad, "columns beyond Mortality now disagree: %r" % bad[:5]


def test_there_are_exactly_two_effects_and_they_have_different_shapes(sheets):
    """THE BOUNDARY, corrected by measurement. My first draft asserted the
    discrepancy lives only on off-feed weeks. It does not, and the test caught
    me: 77 ordinary grow-out batch-weeks also differ -- always WeeklyReport
    LOWER, never on a harvest week, 314 fish in total, about 4 per row against
    batches of ~300,000. That is rounding scale (0.007%), consistent with
    whole-fish rounding applied per tank per day and then summed, against a
    rate computed once for the batch.

    So there are two effects, and what this pins is their SHAPE, not a story:
    the off-feed one may be large per row, the grow-out one must stay at
    rounding scale. If an ordinary week ever drifts past that, something other
    than rounding is in play and the off-feed explanation no longer covers the
    ledger."""
    rec, wkl, starve = sheets
    loud = []
    for k in set(rec) & set(wkl):
        if starve.get(k, 0.0) > 0:
            continue
        d = abs(_g(rec[k], "Mortality_Count") - _g(wkl[k], "Mort_Count (fish)"))
        opened = _g(rec[k], "Open_Count")
        if opened > 0 and d > 0.0002 * opened:
            loud.append((k, d, opened))
    assert not loud, (
        "mortality differs on a NON-off-feed week by more than rounding can "
        "explain: %r" % loud[:5])


def test_the_size_of_the_approximation_is_bounded(sheets):
    """Quantified, not hand-waved: the fish WeeklyReport reports as dead in 6N
    that the population never lost cannot exceed the off-feed population times
    a plausible weekly rate. ~1,000 fish over 85 weeks against 3.4M harvested.
    A regression that made it materially larger fails here."""
    rec, wkl, starve = sheets
    diff = sum(abs(_g(wkl[k], "Mort_Count (fish)") - _g(rec[k], "Mortality_Count"))
               for k in set(rec) & set(wkl) if starve.get(k, 0.0) > 0)
    fish_weeks = sum(v for v in starve.values())
    assert fish_weeks > 0
    assert diff < 0.005 * fish_weeks, (
        "the 6N mortality discrepancy is %.0f fish against %.0f off-feed "
        "fish-weeks (%.4f%%) — larger than the off-feed rule can explain"
        % (diff, fish_weeks, diff / fish_weeks * 100))
