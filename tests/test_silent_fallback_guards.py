"""Guards for the 2026-08 silent-fallback audit's two behavioural fixes.

Both are the same defect class: a MISSING measurement defaulted to the
PASSING value, so a read failure looked like a clean result. The fixes make
absence read as "no data" (N/A / INVESTIGATE), never as a pass. Everything
else in that audit batch is log-lines only (no behaviour change) and is not
tested here.
"""
import math
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from forecast import tuning  # noqa: E402


def _wb_with_section_b(tmp_path, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TransferTemplate"
    ws.append(header)
    for r in rows:
        ws.append(r)
    p = tmp_path / "out.xlsx"
    wb.save(p)
    return str(p)


def test_missing_peak_density_column_reads_as_no_data_not_all_clean(tmp_path):
    """DEFECT: with the Peak_Density column absent, every batch parsed as peak
    0.0 and the distribution reported "0 severe, all clean" — a missing
    MEASUREMENT presented as a compliant result. It must read as "no data"."""
    p = _wb_with_section_b(tmp_path, ["Batch", "Tanks"], [["B51", 3], ["B52", 2]])
    peaks, detail = tuning._peaks_and_detail(p)
    assert peaks == [] and detail == []
    # ...and the analysis-layer lens then reports N/A (None), never PASS.
    from forecast.analysis import density_review
    assert density_review(p) is None


def test_present_peak_density_column_still_measures(tmp_path):
    p = _wb_with_section_b(
        tmp_path,
        ["Batch", "Peak_Density_Ratio", "Peak_Wk", "Wks_from_Start"],
        [["B51", 1.42, 10, 4], ["B52", 0.95, 8, 2]])
    peaks, detail = tuning._peaks_and_detail(p)
    assert peaks == [1.42, 0.95]
    assert [d["Batch"] for d in detail] == ["B51"]   # >= DETAIL_RATIO only


def test_scan_audit_drift_missing_facility_row_cannot_read_as_pass():
    """DEFECT: the facility 'Count (fish)' totals initialized to 0.0, so a
    summary row that was missing (or unparseable) DEFAULTED to the passing
    value of the caller's `abs(fac_signed) < 1.0` cleanliness test. NaN fails
    that comparison, so the verdict reads INVESTIGATE."""
    from tools.run_global_forecast import _scan_audit_drift
    wb = openpyxl.Workbook()
    ws = wb.active   # no 'Count (fish)' row at all
    ws.append(["Some", "other", "row"])
    n_tank, n_bio, fac_signed, fac_abs = _scan_audit_drift(ws)
    assert n_tank == 0 and n_bio == 0
    assert math.isnan(fac_signed) and math.isnan(fac_abs)
    assert not (abs(fac_signed) < 1.0)      # the caller's clean test FAILS


def test_scan_audit_drift_parseable_row_unchanged():
    from tools.run_global_forecast import _scan_audit_drift
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Count (fish)", 0.4, 12.0])
    _, _, fac_signed, fac_abs = _scan_audit_drift(ws)
    assert fac_signed == 0.4 and fac_abs == 12.0
    assert abs(fac_signed) < 1.0            # genuinely clean still passes


# --------------------------------------------------------------------------- #
# global-milp: CP-SAT infeasible weeks must not be a silent degrade
#
# DEFECT (2026-08-11 global audit): solve_cpsat_perweek emits `q_by_w[w] = {}`
# for a week it cannot place; those weeks fall through to the tank pick's
# fallback, which applies NO per-tank density cap. On the operator's board leg
# CP-SAT failed 103 of 127 weeks (81%) and the resulting fallback layout —
# peak density 689.9 kg/m3 against a 95 cap — still reached the compare board
# labelled "Global - CP-SAT optimal" with a PASS gate. The failure was printed
# to stdout and recorded NOWHERE the graders or the workbook could see it.
# --------------------------------------------------------------------------- #
def test_cpsat_perweek_info_reports_the_horizon_denominator():
    """n_infeasible is meaningless without its denominator: "103" reads very
    differently from "103 of 127". The solver must self-report both."""
    from forecast.global_placement_milp_poc import solve_cpsat_perweek
    import inspect
    src = inspect.getsource(solve_cpsat_perweek)
    assert '"n_weeks": len(weeks)' in src
    assert '"n_infeasible": n_infeasible' in src


def test_degraded_placement_is_recorded_as_an_error_not_a_note():
    """A fallback-laid-out horizon must read as an ERROR in the ValidationLog.
    Previously any unrecognised warning fell to "WARNING - Hydration", which
    would have filed a total placement failure under a hydration note."""
    from forecast.excel_io import write_validation_log
    wb = openpyxl.Workbook()
    msg = ("PLACEMENT DEGRADED - CP-SAT could not place 103 of 127 week(s) "
           "(81% of the horizon). Those weeks were laid out by the tank-pick "
           "FALLBACK, which enforces no per-tank density cap.")
    write_validation_log(wb, invariant_warnings=[msg])
    rows = [r for r in wb["ValidationLog"].iter_rows(values_only=True)
            if r and r[0] == 1]
    assert len(rows) == 1
    cat, detail = rows[0][1], rows[0][2]
    assert cat.startswith("ERROR"), f"degrade filed as {cat!r}, not an ERROR"
    assert "Placement degraded" in cat
    assert "103 of 127" in detail


def test_degrade_warning_cannot_be_mistaken_for_a_manual_window_week():
    """window_weeks.manual_window_weeks() recovers operator-scripted weeks by
    scanning ValidationLog text. The degrade line must not poison that read —
    a planner week wrongly marked "window" would be EXCLUDED from the harvest
    compliance gates, hiding breaches in the very run that degraded."""
    from forecast import window_weeks
    from forecast.excel_io import write_validation_log
    wb = openpyxl.Workbook()
    write_validation_log(wb, invariant_warnings=[
        "PLACEMENT DEGRADED - CP-SAT could not place 103 of 127 week(s) "
        "(81% of the horizon).",
        "MANUAL EVENT OK - 2026-W31: harvested 21,812 fish",
    ])
    # Only the genuine manual row is recognised; the degrade line adds nothing.
    assert window_weeks.manual_window_weeks(wb) == {"2026-W31"}


def test_a_fully_solved_placement_raises_no_degrade_warning():
    """The clean case must stay silent — a warning on every optimal run would
    train the operator to ignore the one that matters."""
    from tools.run_global_forecast import cpsat_degrade_warning
    assert cpsat_degrade_warning({"n_weeks": 127, "n_infeasible": 0}) == ""
    assert cpsat_degrade_warning({}) == ""
    assert cpsat_degrade_warning(None) == ""


def test_infeasible_weeks_produce_a_degrade_warning_with_both_numbers():
    """The operator's actual board leg: 103 of 127 weeks unplaced."""
    from tools.run_global_forecast import cpsat_degrade_warning
    w = cpsat_degrade_warning({"n_weeks": 127, "n_infeasible": 103})
    assert w.startswith("PLACEMENT DEGRADED")
    assert "103 of 127" in w and "81%" in w
    assert "NOT an optimal placement" in w
    # never mistakable for an operator-scripted window row
    assert "MANUAL EVENT" not in w and "MANUAL WINDOW" not in w
    import re
    assert not re.search(r"\b\d{4}-W\d{2}\b", w)


def test_degrade_warning_survives_a_missing_denominator():
    """An older/partial info dict must still raise the alarm rather than crash
    or silently return "" (absence must not read as success)."""
    from tools.run_global_forecast import cpsat_degrade_warning
    w = cpsat_degrade_warning({"n_infeasible": 5})
    assert w.startswith("PLACEMENT DEGRADED") and "5 of 0" in w


def test_run_global_routes_the_degrade_warning_into_the_validation_log():
    """End-to-end wiring, without paying for a real 40-minute CP-SAT solve:
    the warning the helper produces must reach write_validation_log and land
    as an ERROR row."""
    import inspect
    from tools import run_global_forecast as rgf
    src = inspect.getsource(rgf.run_global)
    assert "cpsat_degrade_warning(_cpsat_info)" in src
    assert "manual_warnings=list(_mw_warns) + _engine_warns" in src
    assert "return q, info" in inspect.getsource(rgf._solve_cpsat_q)

    from forecast.excel_io import write_validation_log
    wb = openpyxl.Workbook()
    write_validation_log(wb, invariant_warnings=[
        rgf.cpsat_degrade_warning({"n_weeks": 127, "n_infeasible": 103})])
    cats = [r[1] for r in wb["ValidationLog"].iter_rows(values_only=True)
            if r and isinstance(r[0], int)]
    assert cats == ["ERROR - Placement degraded (fallback)"]


# =========================================================================== #
# 2026-08 documentation audit — three more categories that could never fire,
# or fired under the wrong name. Same class as everything above: the tool made
# a decision and the operator had no way to read it.
# =========================================================================== #
def _caps_for(*system_ids):
    """A SystemLimits with a default cap for each named system.

    These tests are about the Pass-B fallback WARNING, not about caps — but
    L3 no longer invents a ceiling for a system the operator never
    configured (it raises naming the missing input), so the fixture has to
    state one. The values are arbitrary and non-binding for this demand.
    """
    from forecast.caps import SystemLimits
    return SystemLimits(defaults={
        (s, m): v for s in system_ids
        for m, v in (("biomass", 400_000.0), ("feed_per_day", 3_000.0))})


class TestPassBFallbackIsActuallyProduced:
    """`WARNING - Pass B fallback (no stickiness)` was a ValidationLog category
    with NO producer anywhere in the repo — and the fallback it names is real
    and live: `_solve_passB_per_week` collected `_passB_fallbacks` and then
    threw the list away. A horizon could lose transfer minimisation on most of
    its weeks and the workbook would say nothing."""

    def _run_with_a_solver_that_never_proves(self):
        import numpy as np
        from forecast import global_planner_l3_poc as l3
        from forecast.caps import SystemLimits

        class _Failed:
            status, x, message = 1, None, "iteration limit"

        def _never_proves(*a, **kw):
            return _Failed()

        demand = {
            ("B1", w): l3.TankDemandRow(
                week=w, week_label=f"2026-W{31 + w}", batch_id="B1", tier="grow",
                tanks=1, biomass_kg=50_000.0, feed_kg_day=500.0,
                avg_wt_g=4000.0, per_tank_biomass_kg=50_000.0,
                per_tank_feed_kg_day=500.0)
            for w in (0, 1)
        }
        y_meta = [("B1", "OG3", 0), ("B1", "OG3", 1)]
        xA = np.zeros(len(y_meta) + 3 * 2 + len(y_meta))
        l3.SOLVER_WARNINGS.clear()
        l3._solve_passB_per_week(
            [0, 1], ["OG3"], {("OG3", 0): 0, ("OG3", 1): 1},
            {("OG3", 0): [0], ("OG3", 1): [1]}, y_meta, demand,
            {0: {"OG3": 4}, 1: {"OG3": 4}},
            {0: "2026-W31", 1: "2026-W32"}, _caps_for("OG3"),
            {0: (0.0, 0.0), 1: (0.0, 0.0)}, xA, 1.0, np, _never_proves,
            0.02, False)
        return list(l3.SOLVER_WARNINGS)

    def test_every_fallen_back_week_is_reported(self):
        """NEGATIVE CONTROL: on the parent commit this list is empty."""
        warns = self._run_with_a_solver_that_never_proves()
        hits = [w for w in warns if w.startswith("PASS B FALLBACK")]
        assert len(hits) == 1, f"Pass B fell back on every week and said: {warns}"
        # The denominator matters as much as the count (the 2026-08 lesson):
        # "2" reads very differently from "2 of 2".
        assert "2 of 2 week(s)" in hits[0]
        # And it must name the CONSEQUENCE, not just the event.
        assert "transfer" in hits[0].lower()

    def test_a_clean_pass_b_stays_silent(self):
        """A warning on every run trains the operator to ignore the one that
        matters — the proved path must add nothing."""
        import numpy as np
        from forecast import global_planner_l3_poc as l3
        from forecast.caps import SystemLimits
        l3.SOLVER_WARNINGS.clear()

        class _Ok:
            status, message = 0, ""
            x = np.zeros(1 + 3 * 1)

        l3._solve_passB_per_week(
            [0], ["OG3"], {("OG3", 0): 0}, {("OG3", 0): [0]},
            [("B1", "OG3", 0)],
            {("B1", 0): l3.TankDemandRow(
                week=0, week_label="2026-W31", batch_id="B1", tier="grow",
                tanks=1, biomass_kg=50_000.0, feed_kg_day=500.0,
                avg_wt_g=4000.0, per_tank_biomass_kg=50_000.0,
                per_tank_feed_kg_day=500.0)},
            {0: {"OG3": 4}}, {0: "2026-W31"}, _caps_for("OG3"), {0: (0.0, 0.0)},
            np.zeros(1 + 3 * 1 + 1), 1.0, np, lambda *a, **kw: _Ok(),
            0.02, False)
        assert not [w for w in l3.SOLVER_WARNINGS
                    if w.startswith("PASS B FALLBACK")]

    def test_the_warning_still_files_under_its_category(self):
        from forecast.excel_io import write_validation_log
        warns = self._run_with_a_solver_that_never_proves()
        wb = openpyxl.Workbook()
        write_validation_log(wb, invariant_warnings=warns)
        cats = [r[1] for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert any(str(c).startswith("WARNING - Pass B fallback") for c in cats)


class TestFwCalibrationIsNotFiledAsHydration:
    """`AUTO-FW-CALIB ...` messages are FRESHWATER GROWTH CALIBRATIONS — the
    model rewriting a batch's fw_correction. Every one of them fell through to
    the `WARNING - Hydration` catch-all, telling an operator scanning the log
    that the Production Report had read badly."""

    def _cats(self, msgs):
        from forecast.excel_io import write_validation_log
        wb = openpyxl.Workbook()
        write_validation_log(wb, invariant_warnings=msgs)
        return [str(r[1]) for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]

    def test_an_applied_calibration_is_its_own_info_category(self):
        cat, = self._cats(["AUTO-FW-CALIB B51: fw_correction 1.000 -> 1.180 "
                           "to land pre-cull on 120g at transfer"])
        assert "Hydration" not in cat, "a growth calibration filed as hydration"
        assert "FW growth calibration" in cat and cat.startswith("INFO")

    def test_a_clamped_or_unconverged_solve_stays_a_warning(self):
        """These two mean the transfer target is UNREACHABLE at this growth —
        the model did not do what was asked. That is not an INFO."""
        clamped, diverged = self._cats([
            "AUTO-FW-CALIB B51: landing on 120g needs fw_correction 2.400 — "
            "CLAMPED to 1.500 [0.50,1.50] (target likely unreachable at this growth)",
            "AUTO-FW-CALIB B52: FW correction did not converge; kept configured 1.000",
        ])
        for cat in (clamped, diverged):
            assert cat.startswith("WARNING - FW growth calibration"), cat
            assert "unreachable" in cat


class TestHybridGuideLedgerReachesTheOperator:
    """`HarvestGuide.ledger` recorded every lever the guide REFUSED and every
    week it declined to steer — and nothing anywhere read it. The purge lever
    silently self-disables whenever sixn_level_drains is off, so a run could
    ignore a lever the operator had switched on and never say so."""

    def test_a_refused_lever_is_recorded_at_all(self):
        """The producing decision, pinned so the ledger cannot quietly empty."""
        import inspect
        from forecast import hybrid_guide
        src = inspect.getsource(hybrid_guide.build_harvest_guide)
        assert "purge lever REFUSED" in src
        assert "notes.append" in src

    def test_run_wires_the_ledger_into_the_validation_log(self):
        import inspect
        from forecast import run as frun
        src = inspect.getsource(frun.main)
        assert 'f"HYBRID GUIDE - {m}"' in src, \
            "the guide ledger is still written to nothing"
        assert "+ _guide_notes" in src, \
            "the ledger notes never reach write_validation_log"

    def test_the_notes_land_in_a_readable_category(self):
        from forecast.excel_io import write_validation_log
        from forecast.hybrid_guide import HarvestGuide
        g = HarvestGuide(weeks={}, follow="full", band=0.05,
                         min_harvest=30_000.0, max_harvest=55_000.0)
        g.note("purge lever REFUSED: sixn_level_drains is off, which is the "
               "guard against over-filling one 6N pair")
        wb = openpyxl.Workbook()
        write_validation_log(wb, invariant_warnings=[f"HYBRID GUIDE - {m}"
                                                     for m in g.ledger])
        rows = [r for r in wb["ValidationLog"].iter_rows(values_only=True)
                if r and isinstance(r[0], int)]
        assert len(rows) == 1
        assert rows[0][1] == "INFO - Hybrid guide (L1) decision"
        assert "REFUSED" in str(rows[0][2])

    def test_a_guide_with_nothing_to_report_writes_nothing(self):
        from forecast.hybrid_guide import HarvestGuide
        g = HarvestGuide(weeks={}, follow="full", band=0.05,
                         min_harvest=30_000.0, max_harvest=55_000.0)
        assert g.ledger == []
