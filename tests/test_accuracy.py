"""FORECAST vs ACTUALS — tests for the biology-grading comparison.

House rule (see tests/test_negative_controls.py): a check that cannot fire is
itself a defect. An accuracy report is a MEASUREMENT, and a measurement that
cannot report an error is worse than none at all — it would certify the growth
model as perfect. So every metric here has a matched pair:

  * a POSITIVE control — identical prediction and actuals must read ~0, proving
    the comparison does not manufacture error out of its own plumbing;
  * a NEGATIVE control — a deliberately wrong prediction must show up, with the
    right SIGN and roughly the right MAGNITUDE, proving the alarm rings.

The other thing under test is the distinction the whole design rests on:
batch-level = biology, tank-level = plan adherence. Moving fish to a different
tank than planned must leave the biology score untouched. If that ever breaks,
the report starts blaming the growth model for the operator's decisions.
"""
from __future__ import annotations

import json
from datetime import date, datetime, timedelta

import pytest
from openpyxl import Workbook

from forecast import accuracy


# --------------------------------------------------------------------------- #
# Builders — minimal workbooks in the exact shape the real writers produce
# --------------------------------------------------------------------------- #

def _fc_wb(rows, anchor=date(2026, 7, 1), weeks=4):
    """A forecast output workbook with a BatchLocations sheet.

    `rows` is a callable (week_index, week_start) -> list of
    (batch, tank, system, count, biomass_kg) for that week.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BatchLocations"
    ws.append(["BATCH LOCATIONS"])
    ws.append(["Per-tank batch occupancy from the forecast plan. Auto-generated."])
    ws.append([])
    ws.append(["Week", "Week_Start", "Batch", "Tank", "System",
               "Count (fish)", "AvgWt (kg)", "Biomass (kg)",
               "Density (kg/m3)", "Stage"])
    for i in range(weeks):
        start = anchor + timedelta(days=7 * i)
        for (b, t, sysid, cnt, bio) in rows(i, start):
            ws.append([f"2026-W{27 + i:02d}",
                       datetime(start.year, start.month, start.day),
                       b, t, sysid, round(cnt), round(bio / cnt, 3) if cnt else 0,
                       round(bio), 50.0, "SW"])
    return wb


def _pr_wb(rows, closing=date(2026, 7, 22)):
    """A ProductionReport workbook. `rows` = [(batch, tank, count, biomass_kg)]."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ProductionReport"
    ws.append([None, None, None, None, None, "Stock"])
    ws.append([" ", None, None, None, None, "Opening Count", "Closing Count",
               "Opening Biomass [kg]", "Closing Biomass [kg]",
               "Opening Avg weight", "Closing Avg weight"])
    ws.append([f"Closing Month: {closing.month}/{closing.day}/{closing.year} "])
    ws.append([None, "Site: Homestead "])
    seen = set()
    for (b, t, cnt, bio) in rows:
        if b not in seen:
            seen.add(b)
            ws.append([None, None, f"Fish group name: {b} "])
        ws.append([None, None, None, f"Unit: {t} ", None,
                   0, cnt, 0, bio, 0, (bio / cnt * 1000.0) if cnt else 0])
    return wb


def _save(wb, tmp_path, name):
    p = tmp_path / name
    wb.save(p)
    return p


def _flat(batch_spec):
    """Constant-through-time rows helper."""
    return lambda i, start: list(batch_spec)


# --------------------------------------------------------------------------- #
# POSITIVE CONTROL — identical inputs must read ~0
# --------------------------------------------------------------------------- #

def test_identical_prediction_and_actuals_show_no_error(tmp_path):
    """The plumbing must not manufacture error. Same numbers on both sides,
    graded on a date that IS a week end, must come out flat zero."""
    spec = [("B41", 61, "OG6N", 10_000, 30_000.0),
            ("B42", 32, "OG3N", 20_000, 40_000.0)]
    fc = _save(_fc_wb(_flat(spec)), tmp_path, "fc.xlsx")
    # Week 0 = [7/1, 7/8) so it ENDS 7/8 — grade exactly there.
    pr = _save(_pr_wb([(b, t, c, m) for (b, t, _s, c, m) in spec],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert len(rep.graded) == 2
    for b in rep.graded:
        assert b.wt_err_pct == pytest.approx(0.0, abs=1e-6)
        assert b.count_err_pct == pytest.approx(0.0, abs=1e-6)
        assert b.biomass_err_pct == pytest.approx(0.0, abs=1e-6)
    assert rep.bias["wt_typical_abs_pct"] == pytest.approx(0.0, abs=1e-6)
    assert "No systematic bias" in rep.bias["verdict"]


# --------------------------------------------------------------------------- #
# NEGATIVE CONTROLS — a wrong prediction MUST show, signed and sized
# --------------------------------------------------------------------------- #

def test_a_deliberately_hot_forecast_is_reported_as_hot(tmp_path):
    """THE alarm. The forecast claims fish 20% heavier than they really are;
    the report must say so, with a POSITIVE sign, and must call it systematic
    because every batch misses the same way."""
    # Predicted: 3.0 kg/fish. Actual: 2.5 kg/fish -> +20%.
    fc = _save(_fc_wb(_flat([("B41", 61, "OG6N", 10_000, 30_000.0),
                             ("B42", 32, "OG3N", 10_000, 30_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B41", 61, 10_000, 25_000.0),
                       ("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    for b in rep.graded:
        assert b.wt_err_pct == pytest.approx(20.0, abs=0.01)
    assert rep.bias["wt_median_signed_pct"] == pytest.approx(20.0, abs=0.01)
    assert rep.bias["over_predicted"] == 2
    assert rep.bias["under_predicted"] == 0
    assert "Systematic" in rep.bias["verdict"]
    assert "HOT" in rep.bias["verdict"]
    assert accuracy.headline(rep)["typical_wt_err_pct"] == pytest.approx(20.0, abs=0.01)


def test_a_deliberately_cold_forecast_is_reported_as_cold(tmp_path):
    """The mirror control — the sign must not be hard-coded to 'hot'."""
    fc = _save(_fc_wb(_flat([("B41", 61, "OG6N", 10_000, 20_000.0),
                             ("B42", 32, "OG3N", 10_000, 20_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B41", 61, 10_000, 25_000.0),
                       ("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.bias["wt_median_signed_pct"] == pytest.approx(-20.0, abs=0.01)
    assert rep.bias["under_predicted"] == 2
    assert "COLD" in rep.bias["verdict"]


def test_scatter_both_ways_is_not_called_systematic(tmp_path):
    """A bias verdict must require the signs to AGREE. Equal-and-opposite
    errors are noise; calling them a systematic bias would send the operator
    hunting a model defect that isn't there."""
    fc = _save(_fc_wb(_flat([("B41", 61, "OG6N", 10_000, 30_000.0),
                             ("B42", 32, "OG3N", 10_000, 20_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B41", 61, 10_000, 25_000.0),
                       ("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.bias["over_predicted"] == 1
    assert rep.bias["under_predicted"] == 1
    assert "Systematic" not in rep.bias["verdict"]


def test_a_tiny_one_signed_error_is_not_promoted_to_systematic(tmp_path):
    """Signs agreeing is not enough — 8 of 8 batches at +0.1% is rounding, not
    a finding. Without this the verdict would cry wolf on every report."""
    fc = _save(_fc_wb(_flat([("B41", 61, "OG6N", 10_000, 25_012.5),
                             ("B42", 32, "OG3N", 10_000, 25_012.5)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B41", 61, 10_000, 25_000.0),
                       ("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.bias["one_sided_share_pct"] == 100.0
    assert abs(rep.bias["wt_median_signed_pct"]) < 1.0
    assert "Systematic" not in rep.bias["verdict"]


# --------------------------------------------------------------------------- #
# The load-bearing distinction: biology vs plan adherence
# --------------------------------------------------------------------------- #

def test_fish_moved_to_a_different_tank_is_adherence_not_model_error(tmp_path):
    """THE design invariant. The operator put the fish in tank 33 while the
    plan said tank 32. The growth prediction was perfect. The biology view must
    read zero error; only the tank (adherence) view may show the difference."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 33, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    # Biology: untouched.
    assert len(rep.graded) == 1
    assert rep.graded[0].wt_err_pct == pytest.approx(0.0, abs=1e-6)
    assert rep.graded[0].count_err_pct == pytest.approx(0.0, abs=1e-6)
    # Adherence: the move is visible, and ONLY here.
    assert rep.coverage["tank_adherence_pct"] == 0.0
    assert {t.present for t in rep.tanks} == {"forecast-only", "actual-only"}


def test_a_split_batch_still_grades_clean(tmp_path):
    """Same fish, spread over two tanks instead of one. Batch-level sums over
    tanks, so this must be invisible to the biology score."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 5_000, 12_500.0),
                       ("B42", 33, 5_000, 12_500.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.graded[0].wt_err_pct == pytest.approx(0.0, abs=1e-6)
    assert rep.graded[0].count_err_pct == pytest.approx(0.0, abs=1e-6)


# --------------------------------------------------------------------------- #
# Coverage — what is NOT graded must not be silently averaged in
# --------------------------------------------------------------------------- #

def test_a_batch_in_only_one_source_is_excluded_from_the_averages(tmp_path):
    """A harvested-out batch is absent from the PR. Folding its whole mass in
    as '100% error' would be a fabricated finding: the PR simply cannot see
    fish that were sold."""
    fc = _save(_fc_wb(_flat([("B41", 61, "OG6N", 10_000, 25_000.0),
                             ("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0),
                       ("B43", 34, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.coverage["batches_graded"] == 1
    assert rep.coverage["batches_forecast_only"] == ["B41"]
    assert rep.coverage["batches_actual_only"] == ["B43"]
    assert rep.bias["n"] == 1
    # Facility totals also cover the graded batch ONLY.
    assert rep.facility["act_count"] == 10_000


def test_the_limits_travel_with_every_report(tmp_path):
    """The things this cannot measure must be attached to the numbers, not left
    in a docstring nobody opens."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    blob = " ".join(rep.limits).lower()
    assert "harvest" in blob            # execution is not graded
    assert "freshwater" in blob         # FW is in neither file
    assert rep.limits == accuracy.LIMITS


# --------------------------------------------------------------------------- #
# Date alignment — the measurement's own biggest error source
# --------------------------------------------------------------------------- #

def test_the_prediction_is_read_at_the_exact_closing_date(tmp_path):
    """Weekly snapshots rarely land on a PR closing date, and on this facility
    the weight error moves ~0.8pp per day of gap — enough to swamp the signal.
    A date halfway between two week ends must therefore read halfway between
    the two predictions, not snap to whichever week is nearer."""
    # 1.0 kg at the end of week 0 (7/8), 2.0 kg at the end of week 1 (7/15).
    def rows(i, start):
        return [("B42", 32, "OG3N", 10_000, 10_000.0 * (i + 1))]

    fc = _save(_fc_wb(rows), tmp_path, "fc.xlsx")
    # 7/11 or 7/12 is ~half a week past 7/8; use 7/11 (t = 3/7).
    pr = _save(_pr_wb([("B42", 32, 10_000, 10_000.0)],
                      closing=date(2026, 7, 11)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.basis["method"] == "interpolated"
    assert rep.basis["weeks"] == ["2026-W27", "2026-W28"]
    # Predicted weight interpolates 1000g -> 2000g at t = 3/7.
    expected = 1000.0 + 1000.0 * (3 / 7)
    assert rep.graded[0].pred_wt_g == pytest.approx(expected, rel=1e-6)


def test_a_date_on_a_week_end_needs_no_interpolation(tmp_path):
    def rows(i, start):
        return [("B42", 32, "OG3N", 10_000, 10_000.0 * (i + 1))]

    fc = _save(_fc_wb(rows), tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 10_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.graded[0].pred_wt_g == pytest.approx(1000.0, rel=1e-9)
    assert rep.alignment_offset_days == 0


def test_a_date_past_the_horizon_degrades_to_the_last_snapshot_and_says_so(tmp_path):
    """Outside the horizon there is nothing to interpolate between. It must
    fall back to a single snapshot AND declare the gap, rather than quietly
    grading against a date it never predicted."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)]), weeks=2),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 9, 30)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.basis["method"] == "snapshot"
    assert rep.notes, "an out-of-horizon comparison must warn"
    assert any("horizon" in n for n in rep.notes)


def test_the_files_the_wrong_way_round_are_flagged(tmp_path):
    """Grading a forecast against a PR that predates it is a user error the
    report must name, not silently score."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)]),
                      anchor=date(2026, 8, 1)), tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 1)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert any("BEFORE" in n for n in rep.notes)


def test_alignment_sensitivity_shows_the_neighbouring_weeks(tmp_path):
    """The report must be able to SHOW how much the week choice is worth,
    rather than asserting the alignment is fine."""
    def rows(i, start):
        return [("B42", 32, "OG3N", 10_000, 10_000.0 * (i + 1))]

    fc = _save(_fc_wb(rows), tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 10_000.0)],
                      closing=date(2026, 7, 11)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    s = rep.sensitivity
    assert s["graded"] and s["next"]
    # Adjacent weeks disagree hugely here by construction; the point is that
    # the spread is REPORTED rather than hidden.
    assert s["graded"]["typical_wt_err_pct"] != s["next"]["typical_wt_err_pct"]


# --------------------------------------------------------------------------- #
# Units — the easiest way to be confidently 1000x wrong
# --------------------------------------------------------------------------- #

def test_grams_and_kilograms_are_reconciled(tmp_path):
    """BatchLocations writes AvgWt in KG; the PR reports grams. If that were
    read straight across, a perfect forecast would score a 99.9% error."""
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.graded[0].pred_wt_g == pytest.approx(2500.0, rel=1e-6)
    assert rep.graded[0].act_wt_g == pytest.approx(2500.0, rel=1e-6)


def test_columns_are_found_by_name_not_position(tmp_path):
    """An inserted column must not silently shift the reader onto the wrong
    metric — the failure mode would be a confident wrong answer, not a crash."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BatchLocations"
    ws.append(["BATCH LOCATIONS"])
    ws.append(["sub"])
    ws.append([])
    ws.append(["Week", "Week_Start", "NEW COLUMN", "Batch", "Tank", "System",
               "Count (fish)", "AvgWt (kg)", "Biomass (kg)",
               "Density (kg/m3)", "Stage"])
    ws.append(["2026-W27", datetime(2026, 7, 1), "x", "B42", 32, "OG3N",
               10_000, 2.5, 25_000, 50.0, "SW"])
    fc = _save(wb, tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)],
                      closing=date(2026, 7, 8)), tmp_path, "pr.xlsx")

    rep = accuracy.compare(fc, pr)

    assert rep.graded[0].pred_count == 10_000
    assert rep.graded[0].pred_wt_g == pytest.approx(2500.0, rel=1e-6)


def test_a_workbook_without_batchlocations_is_refused_clearly(tmp_path):
    wb = Workbook()
    wb.active.title = "Something Else"
    fc = _save(wb, tmp_path, "fc.xlsx")
    pr = _save(_pr_wb([("B42", 32, 10_000, 25_000.0)]), tmp_path, "pr.xlsx")

    with pytest.raises(ValueError, match="BatchLocations"):
        accuracy.compare(fc, pr)


def test_a_pr_without_a_closing_date_is_refused_clearly(tmp_path):
    fc = _save(_fc_wb(_flat([("B42", 32, "OG3N", 10_000, 25_000.0)])),
               tmp_path, "fc.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "ProductionReport"
    ws.append([None, None, None, None, None, "Stock"])
    ws.append([None, None, "Fish group name: B42 "])
    ws.append([None, None, None, "Unit: 32 ", None, 0, 10_000, 0, 25_000, 0, 2500])
    pr = _save(wb, tmp_path, "pr.xlsx")

    with pytest.raises(ValueError, match="Closing Month"):
        accuracy.compare(fc, pr)


# --------------------------------------------------------------------------- #
# Calibration history
# --------------------------------------------------------------------------- #

def test_calibration_log_round_trips(tmp_path):
    p = str(tmp_path / "calib.jsonl")
    recs = [accuracy.calibration_record(
        "B37", ts="2026-08-01T10:00:00", configured=1.0, applied=0.774,
        solved=0.774, target_wt_g=340.0, clamped=False, converged=True,
        lo=0.5, hi=1.5, pr_closing=date(2026, 7, 31), source="t")]
    accuracy.append_calibration_log(recs, p)
    back = accuracy.read_calibration_log(p)

    assert len(back) == 1
    assert back[0]["batch"] == "B37"
    assert back[0]["applied"] == 0.774
    assert back[0]["pr_closing"] == "2026-07-31"


def test_a_missing_log_reads_empty_and_an_unwritable_path_never_raises(tmp_path):
    """Logging is a diagnostic. A diagnostic that can take the pipeline down is
    worse than no diagnostic — same contract as optimize/adoption history."""
    assert accuracy.read_calibration_log(str(tmp_path / "nope.jsonl")) == []
    accuracy.append_calibration_log(
        [{"batch": "B1"}], str(tmp_path / "no" / "such" / "dir" / "x.jsonl"))


def test_a_corrupt_line_does_not_lose_the_rest_of_the_history(tmp_path):
    p = tmp_path / "calib.jsonl"
    p.write_text(json.dumps({"batch": "B1", "applied": 0.8}) + "\n"
                 + "{not json\n"
                 + json.dumps({"batch": "B2", "applied": 0.9}) + "\n",
                 encoding="utf-8")

    back = accuracy.read_calibration_log(str(p))

    assert [r["batch"] for r in back] == ["B1", "B2"]


def test_drift_flags_a_correction_the_model_keeps_needing(tmp_path):
    """THE point of keeping the history: a correction applied every month is a
    standing model error, not a one-off. It has to be distinguishable from a
    batch that only ever needed one nudge."""
    recs = []
    for i in range(6):
        recs.append(accuracy.calibration_record(
            "B49", ts=f"2026-0{i + 1}-01T00:00:00", configured=1.0,
            applied=0.82, solved=0.82, target_wt_g=370.0, clamped=False,
            converged=True))
    recs.append(accuracy.calibration_record(
        "B50", ts="2026-06-01T00:00:00", configured=1.0, applied=1.0,
        solved=1.0, target_wt_g=370.0, clamped=False, converged=True))

    drift = {d["batch"]: d for d in accuracy.calibration_drift(recs)}

    assert drift["B49"]["runs"] == 6
    assert drift["B49"]["median_applied"] == 0.82
    assert drift["B49"]["gap"] == pytest.approx(-0.18)
    assert drift["B49"]["persistent"] is True
    # B50 needed nothing and must NOT be flagged.
    assert drift["B50"]["persistent"] is False


def test_a_single_run_is_not_yet_a_standing_error(tmp_path):
    """One month is an observation, not a pattern. Flagging it would train the
    operator to ignore the flag."""
    recs = [accuracy.calibration_record(
        "B49", ts="2026-06-01T00:00:00", configured=1.0, applied=0.82,
        solved=0.82, target_wt_g=370.0, clamped=False, converged=True)]

    drift = accuracy.calibration_drift(recs)

    assert drift[0]["persistent"] is False


def test_clamped_and_unconverged_runs_are_counted(tmp_path):
    """'Target likely unreachable' is the strongest signal in the calibration
    data — it must survive into the history rather than only into a log line."""
    recs = [
        accuracy.calibration_record(
            "B49", ts="2026-06-01T00:00:00", configured=1.0, applied=0.5,
            solved=0.31, target_wt_g=370.0, clamped=True, converged=True),
        accuracy.calibration_record(
            "B49", ts="2026-07-01T00:00:00", configured=1.0, applied=1.0,
            solved=None, target_wt_g=370.0, clamped=False, converged=False),
    ]

    d = accuracy.calibration_drift(recs)[0]

    assert d["clamped_runs"] == 1
    assert d["not_converged_runs"] == 1


class TestSgrBackSolve:
    """The seawater analogue of the FW correction solver.

    FW self-corrects against the plan's own TranOG target; seawater has no
    target, so a growth error can only be seen against ACTUALS. Before this,
    `sgr_correction` was a knob with no feedback at all.
    """

    @staticmethod
    def _b(anchor, pred, act, confounded=False):
        from forecast.accuracy import BatchAccuracy
        return BatchAccuracy(batch_id="B1", anchor_wt_g=anchor, pred_wt_g=pred,
                             act_wt_g=act, act_count=1000.0, pred_count=1000.0,
                             exec_confounded=confounded)

    def test_scale_is_the_ratio_of_log_growth(self):
        import math
        b = self._b(1000.0, 1200.0, 1150.0)
        assert b.sgr_scale == pytest.approx(
            math.log(1150 / 1000) / math.log(1200 / 1000))
        assert b.sgr_scale < 1.0          # model grew faster than reality

    def test_a_perfect_forecast_scales_by_one(self):
        assert self._b(1000.0, 1200.0, 1200.0).sgr_scale == pytest.approx(1.0)

    def test_model_too_slow_scales_above_one(self):
        assert self._b(1000.0, 1100.0, 1200.0).sgr_scale > 1.0

    def test_no_growth_or_no_anchor_is_unsolvable(self):
        # A flat or shrinking batch makes the log ratio meaningless — better to
        # return nothing than a number that looks like a recalibration.
        assert self._b(1000.0, 1000.0, 1100.0).sgr_scale is None   # no predicted growth
        assert self._b(1000.0, 1200.0, 900.0).sgr_scale is None    # actual shrank
        assert self._b(0.0, 1200.0, 1150.0).sgr_scale is None      # not in the water yet

    def test_harvested_batches_are_excluded_not_just_flagged(self):
        """A partial harvest takes the BIGGEST fish, so survivors' mean weight
        falls for reasons unrelated to growth. Including one would read as the
        model being wildly too fast."""
        from forecast.accuracy import summarize_sgr_recalibration
        clean = self._b(1000.0, 1200.0, 1190.0)
        harvested = self._b(1000.0, 1200.0, 1010.0, confounded=True)
        out = summarize_sgr_recalibration([clean, harvested])
        assert out["n"] == 1
        assert out["excluded_exec_confounded"] == 1
        # the summary rounds to 4 dp for reporting
        assert out["scale"] == pytest.approx(clean.sgr_scale, abs=1e-4)

    def test_aggregate_weights_by_growth_not_by_batch(self):
        """Summed log-growth, so a barely-grown batch cannot dominate. A median
        of per-batch ratios would let it, which is how a 3% weight error over
        three weeks reads as a 40% rate error."""
        from forecast.accuracy import summarize_sgr_recalibration
        big = self._b(1000.0, 2000.0, 1980.0)     # lots of growth, near-perfect
        tiny = self._b(1000.0, 1010.0, 1005.0)    # negligible growth, ratio ~0.5
        out = summarize_sgr_recalibration([big, tiny])
        assert out["scale"] > 0.9, "the barely-grown batch must not dominate"

    def test_clean_tracking_says_so(self):
        from forecast.accuracy import summarize_sgr_recalibration
        out = summarize_sgr_recalibration([self._b(1000.0, 1200.0, 1199.0)])
        assert "no recalibration indicated" in out["verdict"]
