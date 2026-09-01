"""Streamlit UI for the AS Production Forecast tool.

Local-only single-operator workflow:
  1. Upload an input workbook (Forecast.xlsm).
  2. Click "Run forecast" → pipeline runs in-memory; input is never mutated.
  3. Review summary KPIs + Advisory + tank-occupancy heatmap.
  4. Download the output workbook with all populated tabs.

Run with:  streamlit run app.py
"""
from __future__ import annotations

import io
import os
import sys
import uuid
import tempfile
import time
import traceback
from collections import defaultdict
from contextlib import redirect_stdout
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook

# Local imports — adjust path so this file works whether streamlit is
# launched from the project root or elsewhere.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from forecast.run import main as run_pipeline  # noqa: E402
from forecast import optimize  # noqa: E402
from forecast import methods as _methods  # noqa: E402
from forecast import levers as _levers  # noqa: E402

# The ONE method list. Board roster, run-mode label and the engine dispatch all
# read this, so adding a method is a single register() call in forecast/methods.
_METHODS = _methods.REGISTRY
# The method ▶ Run forecast uses until the board picks another. The hybrid, not
# the plain controller: once zero-harvest weeks were actually counted (34ecbaf)
# the controller turned out to leave an empty week on 5 of 6 real PRs, which
# breaks the hard steady-harvest contract rule. See forecast/methods.py.
_DEFAULT_METHOD = "controller-hybrid"
# Operator-facing runtime hints + which methods the board runs unprompted.
_TYPICAL = {"controller": "~30 s", "controller-hybrid": "~40 s",
            "controller-lns": "~30 s", "global-lp": "~4 min",
            "global-milp": "~30 min+"}
# Behind the opt-in checkbox (slow, and benchmarks only -- see the label).
# MUST stay a subset of _BOARD_ORDER: when the roster dropped to the three
# controller arms on 2026-08-27 (84d3e90) this set stopped intersecting it,
# so the comprehensions below returned the same three arms whether the box
# was ticked or not -- a control that silently did nothing, while five
# places of UI text promised it worked.
_BOARD_OPTIONAL = {"global-lp", "global-milp"}
# Pseudo-method: run the controller pipeline on the given config EXACTLY as-is,
# with NO registry pins layered on top. Optimize's sweep measures variants that
# way (variant knobs onto the live config, nothing else), so its verification
# runs must too — passing "controller" here would pin hybrid_follow off and
# verify a DIFFERENT engine than every variant the sweep just scored.
_AS_CONFIGURED = _methods.Method(
    key="as-configured", label="As configured", family="Controller",
    blurb="", engine="controller")


def _method_obj(key):
    """Method object for a stored `_chosen_method`, INCLUDING the as-configured
    pseudo-method. `_METHODS` is the registry and does not contain
    "as-configured", so a plain `_METHODS.get(key) or _METHODS[_DEFAULT_METHOD]`
    silently renames it "Controller — hybrid" — the run would use the right
    engine while the caption named a different one."""
    if key == _AS_CONFIGURED.key:
        return _AS_CONFIGURED
    return _METHODS.get(key) or _METHODS[_DEFAULT_METHOD]

# App-managed config (Phase 1) + scenario (Phase 2) live here. In PR-only
# mode the app reads these instead of pulling everything from the upload;
# the uploaded workbook then supplies only the ProductionReport.
_ROOT = Path(__file__).resolve().parent
CONFIG_DIR = _ROOT / "config"
SCENARIO_DIR = _ROOT / "scenario"


def _config_ready() -> bool:
    return (CONFIG_DIR / "control.yaml").exists()


def _scenario_ready() -> bool:
    return (SCENARIO_DIR / "batches.yaml").exists()


# A file EXISTING is not the same as a file LOADING. `_config_ready` /
# `_scenario_ready` only test existence, so every render path that reads an
# operator-editable file has to survive that file being unreadable.
_READ_FIX_HINT = ("Fix the file on disk, or re-import it from "
                  "**Configure → Template & import**.")


def _read_or_explain(loader, what: str, hint: str = _READ_FIX_HINT):
    """Run a config/scenario loader; NAME a failure instead of blanking out.

    Streamlit renders an exception escaping a rerun as a BLANK PAGE with no
    text at all. That is the worst failure this tool has: it tells the
    operator nothing, and in the sidebar — which runs in EVERY mode, before
    the mode dispatch — it also removes the only route back to the editor
    that would repair the file. A defaults-plus-overrides migration, a hand
    edit, a value that is suddenly required: all of them land here.

    So every read of an operator-editable file on a render path goes through
    this, and a bad file costs one named message naming the file, the
    exception and the way to fix it — never the whole app.

    Returns `(True, value)` or `(False, None)`; callers render a stub and
    return on False. The project standard: a failure must be VISIBLE and
    NAMED, never silent.
    """
    try:
        return True, loader()
    except Exception as e:  # noqa: BLE001 — any loader failure, named not hidden
        st.error(f"**{what}** could not be read — {type(e).__name__}: {e}\n\n"
                 f"{hint}")
        return False, None


def _cpu_workers() -> int:
    """Parallel-work budget from the sidebar's "Computer power" percent: that
    share of this machine's logical CPUs, at least 1. One number feeds both
    kinds of heavy work — CP-SAT search threads and Optimize sweep processes."""
    pct = int(st.session_state.get("cpu_pct", 40))
    return max(1, round((os.cpu_count() or 2) * pct / 100.0))


def _active_config_summary(cd: dict) -> list[tuple]:
    """Plain-language (label, value, what-it-does) for the settings that actually
    shape a run — so the operator can always see WHAT the app is doing. Reads a
    control dict (from the live config, or a run's RunConfig)."""
    def g(k, default=None):
        return cd.get(k, default)
    lv, hl = bool(g("rebalance_level")), bool(g("harvest_level_load"))
    pm = str(g("placement_method", "greedy"))
    dt = g("density_target_pct", 0) or 0
    # Feed leveling and the rebalancer budget are reported by forecast.levers,
    # which knows that leveling SHARES the rebalancer budget and steers nothing
    # when it is 0. This panel used to read `rebalance_level` alone and print
    # "ON" over a budget of 0 — stating something false about the plan the
    # operator was about to run.
    _lv = {s.key: s for s in _levers.effective_levers(cd)}

    def _lever_row(key, fallback_label):
        st = _lv.get(key)
        if st is None:
            return (fallback_label, "—", "")
        if st.status == _levers.ACTIVE:
            val = "ON" if st.raw is True else st.raw
        elif st.status == _levers.OFF:
            val = "OFF" if st.raw is False else f"OFF ({st.raw})"
        else:
            val = f"{'ON' if st.raw is True else st.raw} — {st.status.replace('_', ' ').lower()}"
        return (st.label, val, st.reason)

    rows = [
        _lever_row("rebalance_level", "Feed leveling"),
        ("Harvest smoother", "ON" if hl else "OFF",
         "holds the weekly harvest cap hard + pre-harvests → flat harvest"
         if hl else "reactive → harvest can be lumpy"),
        ("TranOG tanks / arrival", g("tran_og_default_tanks"),
         "more tanks = lower per-system feed but a tighter facility (spikier harvest)"),
        ("Biomass setpoint band", f"{(g('facility_biomass_deviation_pct') or 0) * 100:.2f}%",
         "how close to the FW-inclusive cap the harvest controller runs (the only soft "
         "margin; tighter = higher utilisation)"),
        ("Density target", f"{dt * 100:.0f}%",
         "packs each tank to this % of its density cap (lower = more headroom, "
         "fewer hot spots)"),
        _lever_row("rebalance_balance_budget", "Rebalancer budget"),
        ("Placement engine", pm.upper(),
         "greedy heuristic + rebalancer (default)" if pm == "greedy"
         else "LP-guided LNS optimal-layout refinement"),
        ("FW auto-calibration", "ON" if bool(g("auto_calibrate_fw")) else "OFF",
         "FW growth is auto-adjusted so each batch hits its pre-cull transfer target "
         f"(clamped {g('auto_calibrate_fw_min', 0.5):.2f}–{g('auto_calibrate_fw_max', 1.5):.2f}) "
         "— a planning assumption, not a guarantee"
         if bool(g("auto_calibrate_fw"))
         else "FW growth uses each batch's configured FW_Correction (residuals shown in "
              "Diagnostics)"),
        ("Harvest guide (hybrid)", str(g("hybrid_follow", "off")).upper(),
         f"a whole-horizon harvest plan is computed first and the weekly "
         f"controller must track it within ±"
         f"{(g('hybrid_follow_band') or 0) * 100:.0f}% — it harvests LESS in "
         f"fat weeks so those fish still exist for lean ones (the fix for "
         f"empty harvest weeks; costs a higher biomass peak)"
         if str(g("hybrid_follow", "off")) == "full" else
         "reactive week-by-week harvest only — no long-horizon guide "
         "(measured to leave empty harvest weeks)"),
        ("Caps", f"biomass {g('max_biomass_kg', 0):,.0f} kg · "
                 f"feed {g('max_feed_per_day_kg', 0):,.0f} kg/day",
         "the facility limits the plan must hold under (both counted "
         "FW-inclusive)"),
        ("Harvest floor / limit",
         f"{g('min_harvest_per_week', 0):,.0f} – "
         f"{g('max_harvest_per_week', 0):,.0f} fish/wk"
         + (f" (relief ceiling {g('max_harvest_per_week', 0) * (1 + (g('harvest_relief_pct') or 0)):,.0f})"
            if (g("harvest_relief_pct") or 0) > 0 else ""),
         "every week must ship at least the floor (contract); the limit is "
         "the plant's weekly capacity — a cap, never a target"),
        ("Handling budget", f"{g('max_transfers_per_week', 0):,.0f} moves/wk",
         "deferrable quality moves stop here; essential moves are never "
         "blocked (Controller engines only — Global ignores this)"),
    ]
    # Surface the opt-in knobs only when they're engaged (off by default).
    if (g("min_transfer_count") or 0) > 0:
        rows.append(("Min transfer size", f"{g('min_transfer_count'):,.0f} fish",
            "rebalancer won't split a smaller sub-group out of a tank"))
    if (g("cap_repair_budget") or 0) > 0:
        rows.append(_lever_row("cap_repair_budget", "End-of-week cap repair"))
    # Anything that reads as set but is NOT shaping the plan gets its own rows,
    # so "why did my change do nothing?" is answered on the page rather than in
    # placement.py.
    for st in _levers.not_steering(cd):
        if st.key in ("rebalance_level", "rebalance_balance_budget",
                      "cap_repair_budget"):
            continue          # already rendered above
        rows.append((f"⚠ {st.label}",
                     f"{'ON' if st.raw is True else st.raw} — "
                     f"{st.status.replace('_', ' ').lower()}",
                     st.reason))
    return rows


def _render_active_config(cd: dict, title: str):
    """Render the active-config summary as a collapsible panel (informs without
    clutter)."""
    import streamlit as _st
    with _st.expander(title, expanded=False):
        if not cd:
            _st.caption("No config loaded yet.")
            return
        for label, value, effect in _active_config_summary(cd):
            _st.markdown(f"**{label}: {value}** — {effect}")


def _ingest_pr(uploaded):
    """Parse + validate the uploaded ProductionReport (cached by file identity).

    The PR is the anchor: forecast_start is DERIVED from its closing date
    (+1 day). Returns dict(ok, forecast_start, closing, n_og, n_fw, errors,
    warnings). `ok` is False (locks downstream actions) on any hard error.
    """
    # Key on CONTENT, not (name, size): re-uploading an edited PR under the same
    # name at the same size served the previous parse — stale forecast_start,
    # stale counts and stale validation warnings, silently, for the whole
    # session. Everything downstream anchors on forecast_start, so this is not a
    # cosmetic staleness. The hydration cache below already hashes content; this
    # was the one PR cache that did not. md5 of a few MB is ~milliseconds.
    import hashlib
    key = hashlib.md5(uploaded.getvalue()).hexdigest()
    if st.session_state.get("_pr_key") == key:
        return st.session_state["_pr"]
    from datetime import datetime as _dt, timedelta as _td
    from forecast.excel_io import load_workbook
    from forecast.production_report import read_production_report
    res = {"ok": False, "forecast_start": None, "closing": None,
           "n_og": 0, "n_fw": 0, "errors": [], "warnings": []}
    try:
        wd = Path(tempfile.mkdtemp(prefix="as_pr_"))
        p = wd / uploaded.name
        p.write_bytes(uploaded.getvalue())
        wb = load_workbook(p)
        # Tolerant on the NAME (any spelling of "production report"), strict on
        # the CONTENT — the checks below still require a closing date and tank
        # rows. A workbook in the right format was being rejected over a space
        # or a capital letter.
        from forecast.production_report import find_pr_sheet, pr_sheet_name
        _pr_ws = find_pr_sheet(wb)
        if _pr_ws is None:
            res["errors"].append(
                "No ProductionReport sheet in the workbook (any spelling of "
                "'production report' is accepted). Sheets found: "
                + ", ".join(wb.sheetnames))
        else:
            _nm = pr_sheet_name(wb)
            if _nm and _nm != "ProductionReport":
                res["warnings"].append(
                    f"Using sheet '{_nm}' as the ProductionReport.")
            closing, og, fw = read_production_report(wb)
            res["closing"], res["n_og"], res["n_fw"] = closing, len(og), len(fw)
            if closing is None:
                res["errors"].append(
                    "ProductionReport has no parseable 'Closing Month' date.")
            else:
                res["forecast_start"] = _dt(closing.year, closing.month,
                                            closing.day) + _td(days=1)
            if not og and not fw:
                res["errors"].append(
                    "ProductionReport has no tank rows (no OG/FW records).")
            if (og or fw) and _config_ready() and _scenario_ready():
                try:
                    from forecast.config_io import load_facility_config
                    from forecast.scenario_io import load_batches
                    fac_ids = {t.tank_id for t in load_facility_config(CONFIG_DIR).tanks}
                    batch_ids = {b.batch_id for b in load_batches(SCENARIO_DIR)}
                    pr_b = {r.batch_id for r in og} | {r.batch_id for r in fw}
                    miss_b = sorted(pr_b - batch_ids)
                    if miss_b:
                        res["warnings"].append(
                            f"PR batches not in config Batches {miss_b} — "
                            f"their fish would be dropped.")
                    unk_t = sorted({r.tank_id for r in og} - fac_ids)
                    if unk_t:
                        res["warnings"].append(
                            f"PR tank ids not in Facility config: {unk_t}.")
                except Exception as e:  # noqa: BLE001
                    # The cross-check DEGRADING is fine (config may be mid-
                    # edit); degrading SILENTLY is not — this check is what
                    # warns about fish that would be dropped.
                    res["warnings"].append(
                        f"Config cross-check skipped ({type(e).__name__}: {e})"
                        f" — PR-vs-config batch/tank mismatches were NOT "
                        f"checked.")
        wb.close()
        res["ok"] = not res["errors"]
    except Exception as e:  # noqa: BLE001
        res["errors"].append(f"Failed to read the workbook: {e}")
    st.session_state["_pr_key"] = key
    st.session_state["_pr"] = res
    return res


# ============================================================
# Mother ship — in-app config + scenario editor
# ============================================================
# Edit the app's stable config (biology/facility/control) and scenario
# (forward batches + limits) directly, saved back to the YAML the engine
# runs from. This is what makes the models + control points live in the
# app instead of a workbook. Round-trip converters are kept pure (no
# Streamlit) so they can be unit-tested.

def _records(df):
    """DataFrame -> list[dict] with NATIVE python types (data_editor returns
    numpy scalars / NaN which yaml.safe_dump can't serialize). JSON round-trip
    coerces numpy->native and NaN->None."""
    import json
    return json.loads(df.to_json(orient="records"))


def _preserved_facility_limits(fl_cur, shown_weeks):
    """Facility limit records for weeks the grid is NOT showing.

    The limits editor renders only the current forecast horizon but Save
    REPLACES limits.yaml, so anything without a column here would be deleted.
    See _edit_limits. Extracted so this is testable without Streamlit.
    """
    shown = set(shown_weeks)
    return [{"week": wk, "metric": m, "value": v}
            for (wk, m), v in fl_cur.items() if wk not in shown]


def _preserved_system_limits(sl_cur, shown_weeks):
    """System limit records for weeks the grid is NOT showing (see above)."""
    shown = set(shown_weeks)
    return [{"week": wk, "system": s, "metric": m, "value": v}
            for (wk, s, m), v in sl_cur.items() if wk not in shown]


# ---- System capacity DEFAULTS grid (the common case) ----------------------
#
# A capacity is a fact about the facility, so the editor for it is one row
# per system — not the 130-column week grid, which is technically editable
# and humanly useless. These four helpers are the pure grid<->model mapping,
# kept out of the Streamlit body so they are testable headlessly.

def _system_defaults_records(defaults, systems, metrics):
    """One record per system: {"system": s, <metric>: value or None}."""
    return [{"system": s, **{m: defaults.get((s, m)) for m in metrics}}
            for s in systems]


def _system_defaults_from_records(records, metrics):
    """Grid rows -> {(system, metric): value}. Blank cell = no default."""
    out = {}
    for r in records or []:
        s = (r.get("system") or "").strip()
        if not s:
            continue
        for m in metrics:
            v = r.get(m)
            if v is None or v == "":
                continue
            out[(s, m)] = float(v)
    return out


def _mode_only_cells(defaults, mode_defaults):
    """"SYSTEM METRIC" for every cap stated ONLY as a mode default.

    Those cells render BLANK in the system-capacities grid while the system is
    genuinely capped (today: OG6N biomass, which exists only as purge /
    production rows). "Blank = no cap" is therefore false for exactly them, so
    the caption names them — read from the data, never typed, so it stays true
    if the operator fills one in or gives another system a mode row.
    """
    return sorted({f"{s} {m}" for (s, _mode, m) in (mode_defaults or {})
                   if (s, m) not in (defaults or {})})


def _mode_default_records(mode_defaults):
    """{(system, mode, metric): v} -> sorted list of editable rows."""
    return [{"system": s, "mode": mode, "metric": m, "value": v}
            for (s, mode, m), v in sorted(mode_defaults.items())]


def _mode_defaults_from_records(records):
    """Editable rows -> {(system, mode, metric): value}.

    Skips rows that are not fully specified: a mode default needs all four
    fields to mean anything, and a half-filled row is an accident (the
    dynamic editor appends a blank row the moment you click +).
    """
    out = {}
    for r in records or []:
        s = (r.get("system") or "").strip()
        mode = (r.get("mode") or "").strip()
        m = (r.get("metric") or "").strip()
        v = r.get("value")
        if not (s and mode and m) or v is None or v == "":
            continue
        out[(s, mode, m)] = float(v)
    return out


def _result_rid(r):
    """Stable identity of a result dict, for binding derived data to the run it
    came from. Results stored before `_rid` existed fall back to something
    run-unique."""
    r = r or {}
    return r.get("_rid") or f"{r.get('output_path')}|{r.get('elapsed')}"


def _blank(v):
    return v is None or (isinstance(v, str) and not v.strip())


def _clean_rows(records, key_field, what):
    """Drop never-filled rows from a `num_rows="dynamic"` grid, and refuse
    half-filled ones.

    Streamlit's dynamic editor appends an empty row as soon as you click +, and
    it is returned whether or not you type in it. Saved verbatim those become
    junk config: a batch literally named "None", or a facility tank with a null
    tank_id that bricks the NEXT run deep inside precalc. An all-blank row is
    unambiguously an accident, so drop it silently. A row with data but no
    identifier is NOT — the operator typed something and would lose it, so say
    so instead of guessing.
    """
    out = []
    for i, r in enumerate(records, start=1):
        if all(_blank(v) for v in r.values()):
            continue
        if _blank(r.get(key_field)):
            raise ValueError(
                f"row {i} has no {key_field} — every {what} needs one. "
                f"Fill it in, or clear the row to discard it."
            )
        out.append(r)
    return out


def _persist(key, loader):
    """Load a DataFrame into session_state ONCE and return it.

    Streamlit reruns the whole script on every click; loading from YAML each
    rerun would discard in-progress edits. Holding the base in session_state
    (constant) + a keyed data_editor (key=`<key>_w`) keeps edits across reruns
    and Run/Configure mode switches until the user explicitly Saves or Reloads.
    """
    if key not in st.session_state:
        st.session_state[key] = loader()
    return st.session_state[key]


def _reset_keys(*keys):
    """Drop a working df + its data_editor widget state so it reloads fresh."""
    for k in keys:
        st.session_state.pop(k, None)
        st.session_state.pop(k + "_w", None)


def _clear_all_editor_state():
    """Drop every editor's cached working copy so they reload from disk.

    Used after an import (which rewrites the YAML out from under the open
    editors) so the tabs reflect the freshly-imported config, not stale
    session_state from before the import.
    """
    _reset_keys("bio_growth", "bio_mort", "bio_feed", "bio_cull",
                "fac_df", "batch_df", "flim_wide", "slim_wide")
    for k in ("bio_models", "_lim_weeks", "_tmpl_bytes", "_tmpl_fp"):
        st.session_state.pop(k, None)


def _biology_to_frames(tables):
    """BiologyTables -> (growth_df, mort_df, feed_df, cull_df, model_keys)."""
    n = len(tables.sgr_size_g)
    models = sorted(tables.fcr_by_model.keys())
    growth = []
    for i in range(n):
        row = {"size_g": tables.sgr_size_g[i],
               "SGR_FW": tables.sgr_fw_pct_day[i],
               "SGR_SW": tables.sgr_sw_pct_day[i]}
        for m in models:
            col = tables.fcr_by_model.get(m, [])
            row[f"FCR_{m}"] = col[i] if i < len(col) else None
        growth.append(row)
    growth_df = pd.DataFrame(growth)
    mort_df = pd.DataFrame({"week_from_input": tables.mortality_week_from_input,
                            "mortality_pct": tables.mortality_pct_weekly})
    feed_df = pd.DataFrame([{"max_size_g": mx, "feed_name": nm}
                            for mx, nm in tables.feed_types])
    cull_df = pd.DataFrame([{"days_since_input": d, "cull_pct": p}
                            for d, p in tables.culling])
    return growth_df, mort_df, feed_df, cull_df, models


def _frames_to_biology(growth_df, mort_df, feed_df, cull_df, models):
    """Inverse of _biology_to_frames -> BiologyTables."""
    from forecast.models import BiologyTables
    g = growth_df.dropna(subset=["size_g"])
    sizes = [float(x) for x in g["size_g"]]

    def _col(df, name):
        return [None if pd.isna(x) else float(x) for x in df[name]]

    md = mort_df.dropna(subset=["week_from_input"])
    fd = feed_df.dropna(subset=["max_size_g"]) if len(feed_df) else feed_df
    cd = cull_df.dropna(subset=["days_since_input"]) if len(cull_df) else cull_df
    return BiologyTables(
        sgr_size_g=sizes,
        sgr_fw_pct_day=_col(g, "SGR_FW"),
        sgr_sw_pct_day=_col(g, "SGR_SW"),
        fcr_size_g=list(sizes),
        fcr_by_model={
            m: [float("nan") if pd.isna(x) else float(x)
                for x in g[f"FCR_{m}"]]
            for m in models if f"FCR_{m}" in g.columns
        },
        mortality_week_from_input=[int(x) for x in md["week_from_input"]],
        mortality_pct_weekly=[0.0 if pd.isna(x) else float(x)
                              for x in md["mortality_pct"]],
        feed_types=[(float(mx), str(nm))
                    for mx, nm in zip(fd["max_size_g"], fd["feed_name"])],
        culling=[(int(d), float(p))
                 for d, p in zip(cd["days_since_input"], cd["cull_pct"])],
    )


# Per-parameter explanations shown as the hover-"?" tooltip on each Control
# widget. Keep in sync with docs/USER_GUIDE.md §3.2 and ControlParams in
# forecast/models.py (the authoritative descriptions).
# Every knob the Control editor renders gets a tooltip here, written for an
# operator who has never seen the codebase: (1) what it is in plain language,
# (2) what raising/lowering (or toggling) it does to the plan, (3) the unit and
# the current validated setting where one exists. Knobs whose values were
# CHOSEN BY MEASUREMENT carry a closing caution to tune via Analyze, not by
# hand.
_VALIDATED = (" Validated setting — chosen by measurement; prefer tuning it "
              "via Analyze rather than editing by hand.")
_CONTROL_HELP = {
    "forecast_start":
        "The forecast's week 1. Computed automatically from the uploaded "
        "ProductionReport (its closing date + 1 day) every run — the value "
        "stored here is only a placeholder and is ignored.",
    "horizon_weeks":
        "How far into the future the forecast plans. Longer shows more of the "
        "plan but every run takes longer. Unit: weeks (130 is about 2.5 "
        "years).",
    "scenario_name":
        "A free-text name for this planning scenario. It appears in report "
        "headers and the run's settings snapshot. Changes nothing in the plan "
        "itself.",
    "max_feed_per_day_kg":
        "The most feed the whole facility may deliver in one day. The planner "
        "keeps every week under this; when feed (not space) is the bottleneck "
        "it harvests or moves fish earlier to make feed room. Unit: kg/day. "
        "Per-week overrides live on the Limits tab.",
    "max_biomass_kg":
        "The most fish, by total weight, the whole facility may hold at once "
        "— freshwater, seawater grow-out AND fish waiting in depuration all "
        "count. The harvest controller rides just under this line: raising it "
        "lets fish grow bigger before harvest; lowering it forces earlier "
        "harvests. Unit: kg. Per-week overrides live on the Limits tab.",
    "max_harvest_per_week":
        "THE weekly processing limit: the most fish the plant takes in a "
        "normal week. It is a CONSTRAINT, not a goal — the harvest is decided "
        "by what the fish need (biomass, density, the weekly floor, "
        "contracts) and then capped here; no pass ever sizes harvest UP to "
        "reach it. With the harvest smoother on (the shipped default) this is "
        "a hard weekly ceiling and the 6N drain will hold a purge tank back a "
        "rotation rather than exceed it; with the smoother OFF it only clamps "
        "the 6N fill and the weekly target, so weeks can run over. Unit: "
        "fish/week.",
    "harvest_relief_pct":
        "The pressure-relief band used to JUDGE a plan: how far past the "
        "weekly processing limit an exceptional week may go, as a fraction of "
        "the limit. It derives an absolute ceiling = limit x (1 + this). "
        "IMPORTANT: this is a grading threshold, not an allowance handed to "
        "the planner — no engine reads it. The planner's own weekly ceiling "
        "is the processing limit itself; weeks land in the relief band when a "
        "whole 6N pair had to drain or a force-empty overdrew, and that "
        "overage is borrowed back from the following week. What this knob "
        "decides is how such weeks are SCORED: the checklist flags 1-3 relief "
        "weeks amber, and more than 3 — or any week past the derived ceiling "
        "— red, meaning the plan should ramp its harvests up earlier. Unit: "
        "fraction of the limit.",
    "min_harvest_weight_g":
        "The 3.5 kg sales gate: a fish must weigh at least this (live weight) "
        "before it may be harvested. A business constant, not a tuning knob. "
        "Unit: grams.",
    "min_harvest_per_week":
        "The weekly harvest floor from the sales contracts: EVERY week must "
        "ship at least this many fish — never an empty week. This knob is a "
        "floor only; the holding-back-for-lean-weeks behaviour comes from the "
        "harvest guide (hybrid_follow) further down. Unit: fish/week.",
    "min_tank_control":
        "The 'no dribbles' rule: if a harvest or transfer would leave a tank "
        "holding fewer fish than this, the tank is emptied completely instead "
        "— tiny leftover groups waste a whole tank. Unit: fish.",
    "min_transfer_count":
        "The smallest partial group the rebalancer may move OUT of a tank — "
        "moves smaller than this cost handling for little relief. Raising it "
        "= fewer, larger transfers but slightly more crowding left "
        "unrelieved; 0 = no floor. Whole-tank moves are unaffected. Unit: "
        "fish.",
    "max_transfers_per_week":
        "The weekly HANDLING BUDGET: the most tank-to-tank transfer moves the "
        "crew should perform in one week. When a week's essential moves (6N "
        "purge fills, making room for an arriving batch, following the plan) "
        "have used the budget, the deferrable quality passes (even-out, load "
        "balancing, variable-quantity trims, remnant clean-up) wait for a "
        "calmer week — the split pass is NOT budget-gated. Essential moves "
        "are never blocked; a week they alone exceed the budget shows "
        "amber/red on the checklist instead. 0 switches the budget off. NOTE: "
        "only the Controller engines read this — the Global engines ignore it "
        "entirely. Unit: moves/week.",
    "default_hog_yield":
        "Converts live (gross) weight to sold weight — HOG means head-off, "
        "gutted. Sold kg = live kg × this. Used wherever harvest tonnage or "
        "revenue is reported. Unit: ratio. Per-week overrides live on the "
        "Limits tab.",
    "facility_biomass_deviation_pct":
        "The comfort band under the facility's EFFECTIVE ceiling — how close "
        "to the line the harvest controller aims to ride. The ceiling is the "
        "LOWER of the biomass cap and the biomass at which feed hits its cap, "
        "so on a feed-bound week this is a band under the feed limit. It is "
        "also the tolerance both caps are judged against. UNIT TRAP: this is "
        "a RAW FRACTION — 0.005 = 0.5% (the OPPOSITE convention from Handling "
        "mortality below, which is a percent). Smaller = closer to the cap "
        "(more production, more over-cap risk); larger = harvest earlier, "
        "leaving more slack."
        + _VALIDATED,
    "handling_mortality_pct":
        "Fish lost to handling every time a group is transferred between "
        "tanks. UNIT TRAP: this value is a PERCENT that is divided by 100 "
        "before use — 0.01 means 0.01% (1 fish in 10,000) and 1 means 1%. "
        "That is the OPPOSITE convention from the biomass band above, which "
        "is a raw fraction, so a value of 0.01 here means 0.01% lost per "
        "transfer.",
    "sixn_growth":
        "MASTER SWITCH for the 6N system. Off (unchecked) = normal operation: "
        "6N is the depuration station (fish sit off-feed there before "
        "harvest) until the production start date below, then becomes grow- "
        "out. On = 6N is an ordinary grow-out system for the WHOLE forecast "
        "and there is NO depuration model at all (the date below is ignored). "
        "Leave off unless you truly mean to remove depuration — switching it "
        "on by accident has produced misleading harvest craters before.",
    "sixn_production_start":
        "The date the 6N system stops being the depuration station and "
        "becomes ordinary grow-out (the facility's planned mode change). "
        "Ignored when 'Run 6N as grow-out' is on. Format: YYYY-MM-DD.",
    "sixn_transition_weeks":
        "Rest/empty weeks for the 6N tanks at the switchover date above, "
        "before grow-out fish move in. Unit: weeks.",
    "tran_og_default_tanks":
        "How many tanks a new seawater arrival is spread across on entry — "
        "the strongest feed-vs-harvest lever. More tanks = feed spread "
        "thinner (fewer feed-cap problems) but more of the facility occupied "
        "(bigger make-room harvests); fewer tanks = the reverse. Values below "
        "2 are raised to 2. Unit: tanks.",
    "global_buffer_pct":
        "Headroom above each per-system feed/biomass cap before the planner "
        "treats that system as over-loaded — it is both the rebalancer's "
        "trigger and the audit's tolerance, so it changes the plan, not just "
        "the report. ONE-SIDED: it only allows a system to read over its cap; "
        "being under a cap is never a problem. UNIT: raw fraction — 0.05 = "
        "5%. Separate from the facility biomass band above.",
    "chronic_pressure_frac":
        "An entry-tier (OG1/2) tank sitting at or above this share of its "
        "density cap for 'Chronic pressure weeks' running is treated as "
        "STRUCTURALLY short of tanks — it gets another tank once, instead of "
        "being shaved every week forever. Keep it CLEAR of 'Density relief "
        "target': when the two were equal, relieved tanks landed exactly on "
        "the trigger and flipped in and out of 'chronic' on rounding noise. "
        "Unit: fraction 0-1.",
    "chronic_pressure_weeks":
        "How many consecutive weeks a tank must sit above 'Chronic pressure "
        "level' before it counts as chronic. Shorter reacts sooner and spends "
        "more handling; longer waits for a clearer signal. Unit: weeks.",
    "chronic_relief_pct":
        "A chronic tank is emptied down to this share of its cap — deeper than "
        "the ordinary relief target, because trimming a tank that has sat at "
        "91% for a month back to 90% moves almost nothing and it is chronic "
        "again next week. Unit: fraction 0-1.",
    "chronic_max_frees_per_week":
        "Cap on how many tanks the ANTICIPATORY pass may free per week. "
        "Consolidation and 6N harvest staging draw on the same weekly transfer "
        "budget, so an unbounded sweep starves harvest and misses the sales "
        "floor. Tanks ALREADY over cap are urgent and ignore this cap. "
        "0 stops the anticipatory tank-FREEING only — chronic tanks are still "
        "detected, and still shed to the deeper 'Chronic relief target' by the "
        "weekly even-out pass. Unit: tanks per week.",
    "density_relief_pct":
        "When an over-cap OG1/2 tank is relieved, fish are moved out until it "
        "sits at this share of its cap. NOT 1.0: relieving to exactly the cap "
        "leaves no margin and one week of growth puts the tank straight back "
        "over — which is what made the same tanks breach every week for "
        "months. Unit: fraction 0-1.",
    "consolidation_fill_pct":
        "When a batch's grow-out tanks are consolidated to free one, the "
        "keepers are filled to this share of their cap. 0.80 rather than 0.90 "
        "for the same reason as above — a tank packed to its cap re-breaches "
        "within a week of growth. Unit: fraction 0-1.",
    "global_assume_primed_6n":
        "Applies to the GLOBAL methods AND to the Controller's long-horizon "
        "harvest guide (any 'Harvest guide (hybrid)' setting other than off), "
        "which runs the same L1 planner as a pre-pass and reads this flag "
        "there. OFF (default) models the REAL 6N handover: the "
        "planner starts its depuration pipeline from the fish actually in 6N "
        "on the ProductionReport, so the first couple of weeks harvest less "
        "while the pipeline fills — a real startup ramp. ON assumes 6N is "
        "already full at steady state, which gives a smooth harvest from week "
        "1 but is not physically realisable: there are only six 6N tanks and "
        "the report has already filled them, so those fish end up harvested "
        "straight out of production tanks. Leave OFF unless you specifically "
        "want the idealised envelope.",
    "grade_efficiency":
        "How cleanly a real grader separates sizes, 0-1. A grader does NOT cut "
        "the population neatly at the threshold: fish near the cut line go both "
        "ways, so the two graded populations OVERLAP — the 'big' side keeps "
        "some small fish and vice versa. 1.0 models a perfect grader (the two "
        "sides as far apart as the maths allows); 0.85 (the default, matching "
        "the VBA) narrows that separation by 15%. Lower = more overlap. Total "
        "biomass is unchanged at any setting — only how it splits between the "
        "two legs. 0 means 'off' and behaves as perfect. Unit: fraction 0-1.",
    "starvation_period_days":
        "Only for 6N grow-out mode (after the production start date): "
        "harvest-bound fish stop being fed IN PLACE for this many days before "
        "harvest, since there is no separate depuration tank anymore. Unit: "
        "days.",
    "harvest_grade_to_min":
        "HISTORICAL — this switch no longer controls anything. The behaviour "
        "it used to gate (on a lean depuration week that would miss the "
        "weekly harvest floor, size-sort a near-market tank and send just "
        "enough of its BIGGEST fish to 6N to reach the floor) now runs "
        "UNCONDITIONALLY, because leaving it off produced empty harvest weeks "
        "and that breaks the steady-harvest contract. Flipping this box "
        "changes only a line in the run summary. Kept so older configs still "
        "load. Nothing to tune: no Analyze/Optimize search touches it, and no "
        "setting of it changes a plan.",
    "sixn_level_drains":
        "Levels the flow through depuration: caps how full one 6N pair may "
        "get (at the weekly harvest limit) so weekly fills don't pile into a "
        "single pair and starve the following weeks. On = steadier weekly "
        "harvests (measured: biggest weekly drain 110k -> 68k fish, more "
        "weeks meeting the floor). Only affects depuration mode — but note "
        "that turning it OFF also disables the harvest guide's 6N staging "
        "lever, whatever that checkbox says. Validated by measurement — and "
        "held OUT of every Analyze/Optimize search as a safety guard, so this "
        "one you do set by hand.",
    "density_target_pct":
        "How full to pack each tank when placing fish, as a fraction of that "
        "tank's density cap. 0.90 = fill to 90%, leaving 10% headroom for "
        "growth between weekly checks. Higher = fewer tanks used but more "
        "crowding risk; lower = gentler but needs more tanks. UNIT: raw "
        "fraction."
        + _VALIDATED,
    "density_welfare_threshold_kg_m3":
        "The fish-welfare crowding line, BELOW each tank's hard density cap "
        "(Configure → Facility): "
        "fish reared above it count as 'crowded' in the quality reports (Run "
        "KPI, Compare 'Best welfare', Optimize 'Product quality'). "
        "Reporting/scoring only — it never changes the plan. Unit: kg/m3.",
    "rebalance_varqty_budget":
        "Fine-trim moves per week: shave a precise number of fish off an "
        "over-full system into one with room. More moves = slightly less "
        "over-crowding at the cost of extra handling; 0 = off. Unit: "
        "moves/week.",
    "rebalance_split_budget":
        "Fan-out moves per week: split one over-crowded tank's fish across "
        "several free tanks. Unit: moves/week.",
    "rebalance_balance_budget":
        "The main rebalancer's weekly move budget: relieve tanks and systems "
        "over their caps into destinations with room, weighing crowding, feed "
        "and biomass together. Load leveling (below) shares this budget. "
        "Unit: moves/week.",
    "sixn_overdue_drain_weeks":
        "Rescue fish stuck in purge. A 6N tank whose fish would not fit in "
        "the week's remaining processing budget is held for the next "
        "rotation - but a nearly-full tank never fits, so it can be held "
        "forever (measured: 53,006 fish held 58 weeks off feed; 5 of 8 test "
        "months trap fish this way). Set to N and a tank that has purged "
        "longer than N weeks drains anyway, using the exceptional relief "
        "band. 0 = off, the shipped behaviour. Unit: weeks.",
    "sixn_drain_largest_first":
        "6N drain order. On = a purge pair empties its BIGGEST tank first. Off "
        "= tank-number order, which can spend the week's processing limit on a "
        "small tank and leave its large partner held - every rotation, "
        "indefinitely (measured: 53,006 fish held 58 weeks off feed). Off is "
        "the shipped behaviour.",
    "rebalance_headroom_days":
        "Forward headroom for the rebalancer. It scores destinations on the "
        "load it can see NOW, but the per-system audit measures the END of the "
        "week, a full week of growth later. Set to 7 and it projects each tank "
        "forward at its own growth rate before checking headroom, so a move "
        "cannot fill a system that is about to be over cap. 0 = off, unchanged "
        "behaviour. Unit: days.",
    "rebalance_level":
        "Load leveling. On = each week, spread load off the hottest grow-out "
        "system onto the coolest, leveling feed, biomass and crowding "
        "together — the fix for per-system feed spikes (measured: feed-over- "
        "cap system-weeks 312 → 25). Off = the old crowding-only behavior. "
        "Uses the main rebalancer's move budget."
        + _VALIDATED,
    "cap_repair_budget":
        "End-of-week cap repair. Every other rebalancing pass runs BEFORE the "
        "week's growth is applied, but the numbers you see are measured AFTER "
        "it — so a system left just under its cap grows straight back over it "
        "with nothing left to catch it. This pass runs last, on the state that "
        "is actually reported, and moves the least it can out of any system "
        "still over its feed or biomass cap into the coolest system that can "
        "legally take it. It never breaks a transfer rule and never exceeds "
        "the weekly handling budget. 0 = off. Unit: moves/week. "
        "SHIPPED OFF, and it is NOT a recommendation to switch on. What it "
        "buys is per-system balance, and that part is robust: measured across "
        "8 starting states at budget 8, over-cap system-weeks fell 1,223 → 724 "
        "and every state improved. What it costs is the harvest floor, and "
        "that part is HIGH-VARIANCE — it swings with the ProductionReport, not "
        "with the setting. On the 7.29 PR it made the worst harvest week "
        "BETTER (19,630 → 23,235 fish) and worst density better (116.8 → "
        "102.2); on the 8.13 PR the same budget COLLAPSED the worst week "
        "(23,259 → 4,578) and added a week above the relief ceiling, which is "
        "why it was switched back off. So: if per-system utilisation is your "
        "binding problem, try 8 (15 measured identical — the leftover handling "
        "budget binds first), then check the worst harvest week and the relief "
        "ceiling on YOUR PR in Analyze before keeping it. Never adopt it on "
        "the system-balance numbers alone.",
    "harvest_setpoint_lookahead_weeks":
        "DOES NOTHING (inactive). Superseded by the newer harvest logic — no "
        "part of the plan reads this value anymore; it remains only so older "
        "saved configs still load. Editing it has no effect. (It is still "
        "written into the run's config snapshot, so ignore it there too.)",
    "harvest_level_load":
        "The harvest smoother. On = the weekly processing max is held as a "
        "hard ceiling and fish are pre-harvested a little early, so weekly "
        "harvest is flat instead of dump-then-nothing (measured: weeks over "
        "the weekly limit 15 → 10, steadier weekly totals). Off = the old reactive "
        "behavior. Travels together with Load leveling above, which otherwise "
        "makes harvest spikier."
        + _VALIDATED,
    "harvest_smooth_lookahead_weeks":
        "The harvest smoother's look-ahead: how many weeks of soon-to-be- "
        "ready fish it spreads early harvesting over. Bigger = flatter and "
        "earlier; smaller = closer to reactive. Only used when the harvest "
        "smoother is on. Unit: weeks.",
    "harvest_level_target":
        "Optional FLOOR on how much the harvest smoother pre-harvests each "
        "week — it can only push harvest up, never pin it to a flat number, "
        "and it is still capped by the weekly processing limit. Blank = "
        "computed automatically from how fast the fish are actually growing "
        "(recommended). Only used when the harvest smoother is on. Unit: "
        "fish/week.",
    "placement_method":
        "Which engine assigns fish to physical tanks. 'greedy' (the "
        "production engine) places week by week. 'lns' runs greedy first, "
        "then a refinement pass moves groups off the most crowded systems "
        "onto cooler ones; every move is audit-checked (no fish lost, "
        "strictly less crowding) with greedy as the fallback, so it can't "
        "make a run worse — it just takes longer, and only helps when free "
        "tanks exist.",
    "lns_max_moves":
        "How many refinement moves the 'lns' placement engine may make per "
        "run. Only used when the placement engine is 'lns'. Higher = chases "
        "more crowding but runs slower. Unit: moves.",
    "auto_calibrate_fw":
        "Auto-tune freshwater growth so each batch lands exactly on its "
        "planned seawater-entry weight. On = the forecast ASSUMES the growth "
        "needed to hit target (a planning assumption, not a guarantee the "
        "fish grow that fast); each batch's tuned value is clamped to the "
        "min/max below and flagged if it hits a clamp. Off = use each batch's "
        "hand-set FW correction.",
    "auto_calibrate_fw_min":
        "Lower clamp on the auto-tuned freshwater growth multiplier. A batch "
        "that would need LESS growth than this allows is capped here and "
        "flagged in the log. Only used when auto-calibrate is on. Unit: "
        "multiplier.",
    "auto_calibrate_fw_max":
        "Upper clamp on the auto-tuned freshwater growth multiplier. A batch "
        "that would need MORE growth than this allows is capped here and "
        "flagged as likely unreachable. Only used when auto-calibrate is on. "
        "Unit: multiplier.",
    "hybrid_follow":
        "The long-horizon harvest guide (the 'hybrid'). 'full': before "
        "planning, the app computes a whole-horizon harvest "
        "plan and the weekly controller aims at it as a target band — it "
        "is told to harvest LESS in fat weeks so those fish are still there "
        "for the lean ones, the one thing a week-by-week planner cannot see "
        "for itself. IT STEERS ONLY THROUGH THE TWO LEVERS BELOW ('guide "
        "lever: 6N staging (purge)' and 'guide lever: harvest cap "
        "(production)'): config/control.yaml ships both OFF, so on the PLAIN "
        "controller the guide is computed and then IGNORED. The "
        "'Controller — hybrid' METHOD pins both levers ON (and "
        "hybrid_follow='full'), and a method's pins are written OVER "
        "control.yaml for its own run — so on THAT arm the guide really does "
        "steer. The byte-identical measurement across 21 PRs (2026-08-21) "
        "predates those pins (2026-08-27) and no longer describes it. The purge lever is refused outright whenever "
        "'Level 6N purge drains' is off (it is, here). Switching the levers "
        "on does work: on the real workbook weeks under the harvest floor "
        "fall 20 → 16 with both, 20 → 14 with the production lever alone, "
        "with no empty harvest weeks. Measured on 6 real PRs on 2026-08-03 "
        "WITH THE LEVERS ON, before four 2026-08-20/21 changes and not "
        "reproduced since: totally empty harvest weeks 6 → 0; the cost is a "
        "higher biomass peak (the held-back fish are still "
        "in the water). 'off' = old reactive-only planning (leaves empty "
        "weeks). 'floor' = only lifts short weeks (a no-op in practice)."
        + _VALIDATED,
    "hybrid_follow_band":
        "How tightly the weekly controller must follow the long-horizon "
        "harvest guide, as ± a fraction of the guide's weekly number. Tighter "
        "(smaller) = steadier harvest and better-protected lean weeks; looser "
        "= more freedom to chase the current week. UNIT: raw fraction — 0.05 "
        "= ±5%. The shipped 0.05 beat 0.10 in a 90-cell paired sweep."
        + _VALIDATED,
    "hybrid_guide_min_frac":
        "Housekeeping for the harvest guide: a guide week below this fraction "
        "of the weekly harvest floor is ignored rather than followed down to "
        "nothing — such weeks are structural gaps in the long-horizon plan "
        "(its start-up and tail weeks), not real advice. UNIT: raw fraction "
        "of the harvest floor. Default: 0.25. Ask before changing.",
    "hybrid_guide_smooth_weeks":
        "Optional averaging of the harvest guide's weekly numbers over this "
        "many weeks before the controller follows them. 0 or 1 = follow the "
        "raw curve — recommended; smoothing measured worse (it blunts exactly "
        "the fat-week/lean-week signal the guide exists to carry). Unit: "
        "weeks. Default: 0.",
    "hybrid_purge_lever":
        "Lets the harvest guide steer how much is staged into 6N depuration "
        "each week. NOT a diagnostic switch — it is one of the two gates that "
        "decide whether the guide does anything at all: with BOTH levers off "
        "the guide is computed and then ignored, which made the "
        "'Controller — hybrid' method byte-identical to the plain controller. "
        "Default is ON in code, but this install ships it OFF — check the box "
        "below rather than assuming. Forced off whenever 'Level 6N purge "
        "drains' is off (that is a safety guard against over-filling one 6N "
        "pair, and the guide is not allowed to remove it). NOT TUNABLE: since "
        "2026-08-27 this lever is arm IDENTITY (methods.UNTUNABLE_KNOBS) and "
        "register() refuses any search space containing it, so no tuned run "
        "will settle it for you. What sets it is the METHOD you run — "
        "'Controller — hybrid' pins it ON; the plain controller leaves it at "
        "the value below.",
    "hybrid_production_lever":
        "Lets the harvest guide steer the weekly harvest ceiling and off-feed "
        "entry once 6N is in grow-out mode. NOT a diagnostic switch — see the "
        "purge lever above: with both off the guide steers nothing. Default is "
        "ON in code, but this install ships it OFF. Measured on the real "
        "workbook (2026-08-21): turning THIS lever on alone took weeks under "
        "the contract floor from 20 to 14, better than turning both on (16). "
        "NOT TUNABLE: like the purge lever it is arm IDENTITY "
        "(methods.UNTUNABLE_KNOBS), so a tuned run cannot settle it — "
        "'Controller — hybrid' pins it ON, the plain controller does not.",
}

# Friendly display labels for the Control editor (the raw field name stays the key).
_CONTROL_LABEL = {
    "forecast_start": "Forecast start (derived from PR)",
    "horizon_weeks": "Horizon (weeks)",
    "scenario_name": "Scenario name",
    "max_feed_per_day_kg": "Max feed / day (kg)",
    "max_biomass_kg": "Max facility biomass (kg)",
    "max_harvest_per_week": "Harvest limit / week (fish)",
    "harvest_relief_pct": "Harvest relief band (fraction of limit)",
    "min_harvest_per_week": "Min harvest / week (fish)",
    "min_harvest_weight_g": "Min harvest weight (g)",
    "min_tank_control": "Force-empty floor (fish)",
    "min_transfer_count": "Min transfer size (fish)",
    "max_transfers_per_week": "Handling budget (moves / week)",
    "default_hog_yield": "Default HOG yield",
    "facility_biomass_deviation_pct": "Biomass setpoint band (R24)",
    "handling_mortality_pct": "Handling mortality (per transfer)",
    "sixn_growth": "Run 6N as grow-out",
    "sixn_production_start": "6N production start date",
    "sixn_transition_weeks": "6N transition fallow (weeks)",
    "sixn_level_drains": "Level 6N purge drains (on/off)",
    "tran_og_default_tanks": "TranOG default tanks",
    "global_buffer_pct": "System-cap buffer (R29)",
    "starvation_period_days": "In-place purge length (days)",
    "grade_efficiency": "Grader efficiency (0-1)",
    "global_assume_primed_6n": "6N handover: assume primed at start",
    "chronic_pressure_frac": "Chronic pressure level (fraction of cap)",
    "chronic_pressure_weeks": "Chronic pressure weeks",
    "chronic_relief_pct": "Chronic relief target (fraction of cap)",
    "chronic_max_frees_per_week": "Anticipatory frees per week",
    "density_relief_pct": "Density relief target (fraction of cap)",
    "consolidation_fill_pct": "Consolidation fill (fraction of cap)",
    "density_target_pct": "Density target (fraction of cap)",
    "density_welfare_threshold_kg_m3": "Welfare density line (kg/m³)",
    "rebalance_balance_budget": "Rebalancer moves / week",
    "rebalance_headroom_days": "Rebalancer forward headroom (days)",
    "sixn_drain_largest_first": "6N: drain biggest tank first",
    "sixn_overdue_drain_weeks": "6N: force-drain after (weeks)",
    "rebalance_split_budget": "Split-pass moves / week",
    "rebalance_varqty_budget": "Variable-qty moves / week",
    "rebalance_level": "Load leveling (on/off)",
    "cap_repair_budget": "End-of-week cap repair / week",
    "harvest_setpoint_lookahead_weeks": "Setpoint lookahead (INACTIVE)",
    "harvest_level_load": "Harvest smoother (on/off)",
    "harvest_smooth_lookahead_weeks": "Harvest smoother window K",
    "harvest_level_target": "Harvest level target (fish/wk)",
    "harvest_grade_to_min": "Grade-harvest to the floor (INACTIVE)",
    "placement_method": "Placement engine",
    "lns_max_moves": "LNS move budget",
    "auto_calibrate_fw": "Auto-calibrate FW to transfer target",
    "auto_calibrate_fw_min": "  ↳ FW correction clamp — min",
    "auto_calibrate_fw_max": "  ↳ FW correction clamp — max",
    "hybrid_follow": "Harvest guide (hybrid): off / floor / full",
    "hybrid_follow_band": "  ↳ guide follow band (± fraction)",
    "hybrid_guide_min_frac": "  ↳ guide week drop threshold",
    "hybrid_guide_smooth_weeks": "  ↳ guide smoothing (weeks)",
    "hybrid_purge_lever": "  ↳ guide lever: 6N staging (purge)",
    "hybrid_production_lever": "  ↳ guide lever: harvest cap (production)",
}


def _harvest_limit(default: float = 55_000.0) -> float:
    """The live weekly processing limit (max_harvest_per_week), for readouts
    that would otherwise hardcode 55k and lie the moment it is retuned. Never
    raises — a readout must render even with no config seeded."""
    try:
        from forecast.config_io import load_control
        return float(load_control(CONFIG_DIR).max_harvest_per_week or default)
    except Exception:  # noqa: BLE001 — a label must never break the page
        return default


def _ctl_fmt(v) -> str:
    """One knob value, rendered the way an operator reads it. Pure."""
    if v is None or v == "":
        return "blank (auto)"
    if isinstance(v, bool):
        return "on" if v else "off"
    if isinstance(v, float):
        # 3,800,000 not 3800000.0; 0.005 not 0.00500
        return f"{v:,.0f}" if abs(v) >= 1000 and float(v).is_integer() else f"{v:g}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


_NO_VALUE = object()


def _ctl_help(k: str, v=_NO_VALUE) -> str:
    """Tooltip for one Control knob: the static explanation PLUS this install's
    live value.

    The value is appended at RENDER time on purpose. The prose used to assert
    it ("Current setting: 0.85") and silently rotted on every retune — the
    2026-08-12 sweep caught density_target_pct claiming 0.85 against a config
    holding 0.9, and the welfare line claiming 80 against 85. A tooltip may
    describe what a knob DOES; only the config may say what it is SET to.

    Pure (no Streamlit, no disk) so the contract is testable headlessly."""
    base = _CONTROL_HELP.get(k) or ""
    if v is _NO_VALUE:
        return base
    return f"{base}\n\nCurrently set to: {_ctl_fmt(v)}."


def _ctl_label(k: str) -> str:
    """Friendly Control-editor label for a knob (falls back to a prettified name)."""
    return _CONTROL_LABEL.get(k, k.replace("_", " ").capitalize())


# Column tooltips for the tabular editors (shown on the column header in the grid).
_FACILITY_HELP = {
    "location_id": "Human-readable tank name (free text), e.g. 'OG3N-1'. Used in "
        "every report and the manual window grid.",
    "system_id": "The system (water loop) this tank belongs to, e.g. OG3N. Tanks "
        "in one system share the per-system feed and biomass caps set on the "
        "Limits tab.",
    "tank_id": "Unique tank number — the engine's identifier for this tank. Must "
        "not repeat.",
    "volume_m3": "Tank water volume. Sets the tank's biomass cap together with "
        "the density cap: cap (kg) = volume × max density. Unit: m³.",
    "max_density_kg_m3": "The most fish weight allowed per m³ of this tank — the "
        "hard crowding cap. Tank biomass cap (kg) = volume × this. Unit: kg/m³. "
        "Grow-out tanks are 95.",
    "max_feed_kg_day": "The most feed this tank can deliver in one day. For "
        "grow-out fish, feed is usually the binding limit before space is. "
        "Unit: kg/day. Grow-out tanks are 1,000.",
    "type": "What the tank is for: FW = freshwater stages (eggs to smolt), "
        "OG = seawater grow-out (including the 6N depuration tanks).",
}
_BATCH_HELP = {
    "batch_id": "Unique batch (cohort) label, e.g. B53. Must not repeat.",
    "input_date": "The day the fry are stocked into freshwater. Format: "
        "YYYY-MM-DD.",
    "input_count": "How many fry are stocked on the input date. Unit: fish.",
    "tran_sf_date": "The day the batch moves from first-feeding to the "
        "smolt stage within freshwater. Format: YYYY-MM-DD.",
    "tran_og_date": "The day the batch enters seawater (the TranOG transfer). "
        "From here on the forecast tracks it tank by tank. Format: YYYY-MM-DD.",
    "tran_og_count": "Planned number of fish entering seawater — fish above "
        "this count at transfer are culled to hit it. Unit: fish.",
    "tran_og_avg_wt_g": "Planned average fish weight at seawater entry (the "
        "pre-cull target the freshwater phase aims for). Unit: grams.",
    "tran_og_cv": "How spread-out the fish sizes are at entry (coefficient of "
        "variation). Drives the big/small grading split at transfer. Unit: %. "
        "Typical: 16.",
    "fcr_model": "Which feed-conversion curve this batch eats by, e.g. "
        "FCR_116_Quick = 1.16 kg feed per kg growth. The curves themselves "
        "live on the Biology tab.",
    "fw_correction": "Freshwater growth multiplier for this batch: 1.0 = grow "
        "exactly by the freshwater growth table, 1.1 = 10% faster, 0.9 = 10% "
        "slower. Ignored (auto-tuned) when 'Auto-calibrate FW' is on in "
        "Control.",
    "sgr_correction": "Seawater growth multiplier for this batch: 1.0 = grow "
        "exactly by the seawater growth table; above/below = faster/slower. "
        "Calibrate against how the cohort is actually performing.",
    "notes": "Free-text notes — reporting only, changes nothing.",
}

# Biology-tab grids: one help per column, keyed per grid (growth FCR columns
# are dynamic — one per model — and get a generated tooltip).
_BIO_GROWTH_HELP = {
    "size_g": "Fish size this row applies to. The tables form curves — the "
        "engine interpolates between neighboring sizes. Unit: grams.",
    "SGR_FW": "Freshwater growth rate at this size: % of body weight gained "
        "per day (SGR). Unit: %/day.",
    "SGR_SW": "Seawater growth rate at this size: % of body weight gained per "
        "day (SGR). Unit: %/day.",
}
_BIO_MORT_HELP = {
    "week_from_input": "Batch age, counted in weeks since freshwater stocking.",
    "mortality_pct": "Fish lost during that week of age, as a % of the batch. "
        "Unit: %/week.",
}
_BIO_FEED_HELP = {
    "max_size_g": "Fish up to this size eat this feed type; the next row takes "
        "over above it. Unit: grams.",
    "feed_name": "Feed product name — used to break the feed forecast out by "
        "type. Reporting only.",
}
_BIO_CULL_HELP = {
    "days_since_input": "Batch age, in days since freshwater stocking, at "
        "which this planned cull happens.",
    "cull_pct": "Share of the batch removed at that age, taken from the "
        "SMALLEST fish (a bottom cull). Unit: %.",
}


# Knob GROUPS for the Control editor. 56 knobs rendered as one flat list in
# dataclass order is a wall an operator has to read end-to-end to find anything,
# and it puts a sales contract (min_harvest_per_week) next to a solver detail
# (lns_max_moves) with nothing to say they are different kinds of thing.
#
# Ordered by WHAT THE OPERATOR OWNS, hardest commitment first:
#   1. what the facility IS and the business REQUIRES  — mostly UNTUNABLE_KNOBS
#   2. how the planner behaves                          — the real levers
#   3. engine internals                                 — rarely touched
#
# A key not listed here still renders, in "Everything else" — the editor's
# invariant is that EVERY key lands in `new`, so grouping may never be a filter.
# tests/test_app_control_editor.py holds that line.
_CONTROL_GROUPS = [
    ("🏭 The facility and the contract",
     "What the site IS and what the business REQUIRES. Most of these are "
     "operator inputs no automatic search may touch — the tuner may change how "
     "the model plans, never what the facility is or what you have promised.",
     ["max_feed_per_day_kg", "max_biomass_kg", "max_harvest_per_week",
      "harvest_relief_pct", "min_harvest_per_week", "min_harvest_weight_g",
      "min_tank_control", "max_transfers_per_week", "min_transfer_count",
      "tran_og_default_tanks", "density_target_pct",
      "density_welfare_threshold_kg_m3", "default_hog_yield",
      "grade_efficiency", "handling_mortality_pct"]),
    ("📅 Horizon and scenario",
     "When the plan starts, how far it runs, and the 6N production-mode "
     "switchover.",
     ["forecast_start", "horizon_weeks", "scenario_name",
      "starvation_period_days", "sixn_growth", "sixn_production_start",
      "sixn_transition_weeks"]),
    ("🎣 Harvest control",
     "How the weekly harvest is paced against the biomass cap and the contract "
     "floor.",
     ["facility_biomass_deviation_pct", "harvest_level_load",
      "harvest_smooth_lookahead_weeks", "harvest_level_target",
      "harvest_grade_to_min", "harvest_setpoint_lookahead_weeks"]),
    ("⚖️ Rebalancer and repair",
     "The passes that move fish to relieve crowding and per-system load. "
     "MEASURED 2026-08-30/31 across 8 starting states: most of these do less "
     "than they look like they do — check the Active configuration panel, "
     "which reports what each one is ACTUALLY doing on your config.",
     ["rebalance_balance_budget", "rebalance_level", "rebalance_headroom_days",
      "rebalance_split_budget", "rebalance_varqty_budget", "cap_repair_budget",
      "density_relief_pct", "consolidation_fill_pct",
      "chronic_pressure_frac", "chronic_pressure_weeks", "chronic_relief_pct",
      "chronic_max_frees_per_week"]),
    ("🧊 6N depuration",
     "The purge pipeline: how fills are sized and how tanks drain. Two of "
     "these were measured and rejected as defaults in August 2026 — see the "
     "tooltips.",
     ["sixn_level_drains", "sixn_drain_largest_first",
      "sixn_overdue_drain_weeks"]),
    ("⚙️ Engine internals",
     "Which planner runs and how hard it works. Rarely worth touching by hand — "
     "Decide tunes these for you.",
     ["placement_method", "lns_max_moves", "hybrid_follow", "hybrid_follow_band",
      "hybrid_guide_min_frac", "hybrid_guide_smooth_weeks", "hybrid_purge_lever",
      "hybrid_production_lever", "global_buffer_pct", "global_assume_primed_6n",
      "auto_calibrate_fw", "auto_calibrate_fw_min", "auto_calibrate_fw_max"]),
]


def _edit_control():
    from forecast.config_io import (
        load_control, load_biology_tables, load_facility_config,
        control_to_dict, control_from_dict, dump_config,
    )
    st.caption("Caps defaults, horizon, and planner knobs. "
               "`forecast_start` is derived from the ProductionReport at "
               "run time — the stored value is an ignored seed.")
    _pr = _ingest_pr(uploaded) if uploaded is not None else None
    _derived = (_pr["forecast_start"].date()
                if (_pr and _pr["ok"] and _pr["forecast_start"]) else None)
    _ok, d = _read_or_explain(lambda: control_to_dict(load_control(CONFIG_DIR)),
                              "config/control.yaml")
    if not _ok:
        return
    # Knobs no engine reads. They cannot simply be deleted: `dump_config`
    # rewrites control.yaml from the ControlParams dataclass, which emits every
    # field, so a removed knob reappears in the file after any Save — and the
    # field itself has to stay for configs predating the redesign to load. They
    # are therefore kept, labelled, and (here) moved OUT of the working set so
    # nobody tunes one expecting an effect. Derived from the label convention
    # rather than a second hardcoded list, so marking a future knob INACTIVE
    # moves it automatically.
    _inactive = {k for k in d if "(INACTIVE)" in _ctl_label(k)}
    with st.form("control_form"):
        new = {}

        def _knob(k, v):
            """Render one knob. EVERY key in `d` must land in `new` — a key
            that is skipped is silently dropped from the saved config."""
            if isinstance(v, bool):
                new[k] = st.checkbox(_ctl_label(k), value=v,
                                     help=_ctl_help(k, v))
            elif isinstance(v, int):
                new[k] = int(st.number_input(_ctl_label(k), value=int(v), step=1,
                                             help=_ctl_help(k, v)))
            elif isinstance(v, float):
                new[k] = float(st.number_input(_ctl_label(k), value=float(v),
                                               format="%.5f",
                                               help=_ctl_help(k, v)))
            else:
                new[k] = st.text_input(_ctl_label(k), value="" if v is None else str(v),
                                       help=_ctl_help(k, v)) or None

        # Render group by group. `_seen` guarantees every key lands in `new`
        # exactly once: a key listed in no group falls through to "Everything
        # else", and a key listed twice is rendered once. Grouping is a LAYOUT,
        # never a filter — a skipped key is silently dropped from the config.
        _seen = set()
        _grouped = []
        for _title, _blurb, _keys in _CONTROL_GROUPS:
            _mine = [k for k in _keys if k in d and k not in _inactive
                     and k not in _seen]
            _seen.update(_mine)
            if _mine:
                _grouped.append((_title, _blurb, _mine))
        _rest = [k for k in d if k not in _seen and k not in _inactive]
        if _rest:
            _grouped.append(("📦 Everything else",
                             "Not yet grouped — still saved and still read.",
                             _rest))

        def _render_key(k):
            v = d[k]
            if k == "forecast_start":
                if _derived is not None:
                    st.success(f"forecast_start = **{_derived}** — derived from "
                               f"the uploaded ProductionReport (the stored seed "
                               f"is ignored).")
                else:
                    st.info("forecast_start is derived from the ProductionReport "
                            "at run time — upload a PR to see it. The stored "
                            "value is an ignored seed.")
                new[k] = v
            else:
                _knob(k, v)

        for _i, (_title, _blurb, _keys) in enumerate(_grouped):
            st.markdown(f"**{_title}**")
            st.caption(_blurb)
            for k in _keys:
                _render_key(k)
            if _i < len(_grouped) - 1:
                st.divider()
        if _inactive:
            with st.expander(f"⚪ Inactive settings ({len(_inactive)}) — kept only "
                             f"so older configs load", expanded=False):
                st.caption(
                    "No engine reads these. They stay in the file because saving "
                    "rewrites every field of the config, and older configs must "
                    "still load — editing one has no effect on any plan.")
                for k in _inactive:
                    _knob(k, d[k])
        if st.form_submit_button("💾 Save Control"):
            # control_from_dict coerces to the declared types and raises on a
            # value that cannot be one (e.g. text typed into a knob that is
            # currently null, so it rendered as a text box). Catch it here: the
            # alternative is a saved string that fails much later in arithmetic.
            # dump_config needs all three files, so an unreadable biology.yaml /
            # facility.yaml fails the SAVE too — inside the try, so it reports
            # rather than blanking the page mid-edit.
            try:
                _ctl = control_from_dict(new)
                dump_config(CONFIG_DIR, control=_ctl,
                            tables=load_biology_tables(CONFIG_DIR),
                            facility=load_facility_config(CONFIG_DIR))
            except Exception as e:  # noqa: BLE001
                st.error(f"Not saved — {type(e).__name__}: {e}")
            else:
                st.success("Saved config/control.yaml")


def _edit_biology():
    from forecast.config_io import (
        load_control, load_biology_tables, load_facility_config, dump_config,
    )
    st.caption("Growth (SGR FW/SW), FCR curves, mortality, feed types, and "
               "culling. Edits persist until you Save or Reload.")
    if "bio_models" not in st.session_state:
        # _biology_to_frames indexes the SGR columns positionally, so a ragged
        # curve (a size row added without its SGR values) raises here too.
        _ok, _frames = _read_or_explain(
            lambda: _biology_to_frames(load_biology_tables(CONFIG_DIR)),
            "config/biology.yaml")
        if not _ok:
            return
        g, m, f, c, models = _frames
        st.session_state.update({"bio_growth": g, "bio_mort": m, "bio_feed": f,
                                 "bio_cull": c, "bio_models": models})
    models = st.session_state["bio_models"]
    st.markdown("**Growth + FCR** (by fish size, grams)")
    # Column tooltips: the FCR columns are dynamic (one per model), so their
    # help is generated per model name.
    _g_cfg = {c: st.column_config.Column(help=h)
              for c, h in _BIO_GROWTH_HELP.items()}
    for _m in models:
        _g_cfg[f"FCR_{_m}"] = st.column_config.Column(
            help=f"Feed-conversion curve '{_m}': kg of feed per kg of growth "
                 f"at this fish size. A batch uses the one curve its "
                 f"'fcr_model' names on the Batches tab.")
    g2 = st.data_editor(st.session_state["bio_growth"], num_rows="dynamic",
                        hide_index=True, use_container_width=True,
                        key="bio_growth_w", column_config=_g_cfg)
    cols = st.columns(3)
    with cols[0]:
        st.markdown("**Mortality** (% / wk)")
        m2 = st.data_editor(st.session_state["bio_mort"], num_rows="dynamic",
                            hide_index=True, key="bio_mort_w",
                            column_config={c: st.column_config.Column(help=h)
                                           for c, h in _BIO_MORT_HELP.items()})
    with cols[1]:
        st.markdown("**Feed types**")
        f2 = st.data_editor(st.session_state["bio_feed"], num_rows="dynamic",
                            hide_index=True, key="bio_feed_w",
                            column_config={c: st.column_config.Column(help=h)
                                           for c, h in _BIO_FEED_HELP.items()})
    with cols[2]:
        st.markdown("**Culling**")
        c2 = st.data_editor(st.session_state["bio_cull"], num_rows="dynamic",
                            hide_index=True, key="bio_cull_w",
                            column_config={c: st.column_config.Column(help=h)
                                           for c, h in _BIO_CULL_HELP.items()})
    b1, b2, _ = st.columns([1, 1, 3])
    if b1.button("💾 Save Biology", key="save_bio"):
        try:
            tables2 = _frames_to_biology(g2, m2, f2, c2, models)
            dump_config(CONFIG_DIR, control=load_control(CONFIG_DIR),
                        tables=tables2, facility=load_facility_config(CONFIG_DIR))
            _reset_keys("bio_growth", "bio_mort", "bio_feed", "bio_cull")
            st.session_state.pop("bio_models", None)
            st.success("Saved config/biology.yaml")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Save failed: {e}")
    if b2.button("↻ Reload", key="reload_bio"):
        _reset_keys("bio_growth", "bio_mort", "bio_feed", "bio_cull")
        st.session_state.pop("bio_models", None)
        st.rerun()


def _edit_facility():
    from forecast.config_io import (
        load_control, load_biology_tables, load_facility_config,
        facility_to_dict, facility_from_dict, dump_config,
    )
    st.caption("Tank definitions: system, stage, volume, density/feed caps, type.")
    _ok, _tanks = _read_or_explain(
        lambda: facility_to_dict(load_facility_config(CONFIG_DIR))["tanks"],
        "config/facility.yaml")
    if not _ok:
        return
    base = _persist("fac_df", lambda: pd.DataFrame(_tanks))
    edited = st.data_editor(base, num_rows="dynamic", hide_index=True,
                            use_container_width=True, key="fac_df_w",
                            column_config={c: st.column_config.Column(help=h)
                                           for c, h in _FACILITY_HELP.items()})
    b1, b2, _ = st.columns([1, 1, 3])
    if b1.button("💾 Save Facility", key="save_fac"):
        try:
            fac2 = facility_from_dict(
                {"tanks": _clean_rows(_records(edited), "tank_id", "tank")})
            dump_config(CONFIG_DIR, control=load_control(CONFIG_DIR),
                        tables=load_biology_tables(CONFIG_DIR), facility=fac2)
            _reset_keys("fac_df")
            st.success("Saved config/facility.yaml")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Save failed: {e}")
    if b2.button("↻ Reload", key="reload_fac"):
        _reset_keys("fac_df")
        st.rerun()


def _edit_batches():
    from forecast.config_io import load_control
    from forecast.scenario_io import (
        load_batches, load_limits, batches_to_list, batches_from_list,
        dump_scenario,
    )
    st.caption("Forward batch schedule + metadata. In-flight state comes from "
               "the ProductionReport; this is the planning/metadata layer.")
    _ok, _rows = _read_or_explain(
        lambda: batches_to_list(load_batches(SCENARIO_DIR)),
        "scenario/batches.yaml")
    if not _ok:
        return
    base = _persist("batch_df", lambda: pd.DataFrame(_rows))
    edited = st.data_editor(base, num_rows="dynamic", hide_index=True,
                            use_container_width=True, key="batch_df_w",
                            column_config={c: st.column_config.Column(help=h)
                                           for c, h in _BATCH_HELP.items()})
    b1, b2, _ = st.columns([1, 1, 3])
    if b1.button("💾 Save Batches", key="save_batch"):
        try:
            batches2 = batches_from_list(
                _clean_rows(_records(edited), "batch_id", "batch"))
            # BOUND to Control: capacities may carry mode-specific
            # defaults (6N purge vs production), and resolving one
            # unbound raises. Cost of getting this wrong is a blank
            # page, so it is bound at every call site, not most.
            fl, sl = load_limits(SCENARIO_DIR, load_control(CONFIG_DIR))
            dump_scenario(SCENARIO_DIR, batches=batches2,
                          facility_limits=fl, system_limits=sl)
            _reset_keys("batch_df")
            st.success("Saved scenario/batches.yaml")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Save failed: {e}")
    if b2.button("↻ Reload", key="reload_batch"):
        _reset_keys("batch_df")
        st.rerun()


# ============================================================
# Manual override window editor (Run mode) — script week-by-week
# operations the forecast EXECUTES before the planner takes over.
# ============================================================

_MANUAL_COLS = ["week", "type", "batch", "from_tank", "to_tanks", "count",
                "mode", "notes"]


def _hydrate_state_from_upload(uploaded):
    """Hydrate a FacilityState from the uploaded PR, so the editor can populate
    tank/batch context + dry-run validate events. Returns (state, fw, ctx).

    Cached by PR content hash AND config fingerprint: `ctx` embeds control /
    tables / batches read from disk below, so keying on the PR alone left the
    manual window validating against pre-edit config for the rest of the
    session after any in-app config save."""
    import hashlib
    import io
    data = uploaded.getvalue()
    ck = ("_hydrated_" + hashlib.md5(data).hexdigest() + "_"
          + _config_fingerprint())
    if ck in st.session_state:
        return st.session_state[ck]
    from openpyxl import load_workbook
    from datetime import datetime as _dt, timedelta as _td
    from forecast.config_io import load_config
    from forecast.scenario_io import load_batches
    from forecast.production_report import read_production_report, hydrate_facility_state
    from forecast.state import FacilityState
    wb = load_workbook(io.BytesIO(data), data_only=True,
                       keep_vba=str(uploaded.name).lower().endswith(".xlsm"))
    control, tables, facility = load_config(CONFIG_DIR)
    batches = load_batches(SCENARIO_DIR)
    pc, og, fw = read_production_report(wb)
    fs = _dt(pc.year, pc.month, pc.day) + _td(days=1)
    # Mirror the real run: the PR closing date is the forecast start (run.py
    # DetectForecastStart). The faithful manual-event validation needs it on
    # control so its FW projection anchors to the same week the run will.
    control.forecast_start = fs
    state = FacilityState.from_facility_config(facility, today=fs.date())
    hydrate_facility_state(state, og, batches)
    # Run-time context so the editor can validate events FAITHFULLY (per-week
    # biology + the FW projection for fw_to_og), matching forecast/run.py.
    ctx = {
        "batch_by_id": {b.batch_id: b for b in batches},
        "tables": tables,
        "forecast_start": fs.date(),
        "control": control,
        "pr_closing": pc,
        "fw_records": fw,
    }
    # Keep only the current hydration: each entry is a deep FacilityState, and
    # the key now varies with config too, so an un-evicted cache would grow one
    # entry per (upload × config save) for the life of the session.
    for k in [k for k in st.session_state
              if isinstance(k, str) and k.startswith("_hydrated_") and k != ck]:
        st.session_state.pop(k, None)
    st.session_state[ck] = (state, fw, ctx)
    return state, fw, ctx


def _dest_token(d):
    """One `to_tanks` token: `tank`, `tank:count`, `tank@size`, `tank:count@size`."""
    tok = str(d.tank) if d.count is None else f"{d.tank}:{int(d.count)}"
    return f"{tok}@{d.size_class}" if d.size_class else tok


def _parse_dest_token(tok):
    """Inverse of _dest_token -> (tank, count|None, size_class|None)."""
    size = None
    if "@" in tok:
        tok, size = tok.split("@", 1)
        size = size.strip().lower() or None
    if ":" in tok:
        _t, _c = tok.split(":", 1)
        return int(float(_t)), float(_c), size
    return int(float(tok)), None, size


def _manual_events_to_df_rows(events):
    rows = []
    for e in events:
        # Encode per-dest counts as "tank:count" so explicit / UNEQUAL counts
        # round-trip losslessly; a bare "tank" means None (split the `count`
        # column evenly across the bare tanks at run time). "@big"/"@small"
        # carries fw_to_og size routing — WITHOUT it a round-trip through this
        # grid silently collapsed the big/small split, so "Apply to window"
        # quietly planned a different transfer than the one that was saved.
        to_tanks = ",".join(_dest_token(d) for d in e.destinations)
        count = e.count if e.type not in ("og_transfer", "og_to_6n") else None
        # `mode` matters only for graded_harvest ("stage" = 6N purge staging —
        # the 6N-pickup default; "harvest" = drained to processing in the
        # scripted week) — round-trip the EFFECTIVE timing explicitly, or
        # applying the grid could silently flip an in-week harvest back to the
        # staged default.
        from forecast.manual_events import is_staged_graded
        mode = ""
        if e.type == "graded_harvest":
            mode = "stage" if is_staged_graded(e) else "harvest"
        rows.append({"week": e.week, "type": e.type, "batch": e.batch or "",
                     "from_tank": e.from_tank, "to_tanks": to_tanks,
                     "count": count, "mode": mode, "notes": e.notes})
    return rows


def _rows_to_manual_events(rows):
    """Flat editor rows -> list[ManualEvent]. `count` = total to move (split
    evenly across to_tanks) for transfer/6N; = target for fw_to_og; = amount for
    harvest."""
    from forecast.manual_events import ManualEvent, ManualDest
    out = []
    for r in rows:
        typ = str(r.get("type") or "").strip()
        if not typ:
            continue
        week = int(r.get("week") or 1)
        batch = str(r.get("batch")).strip() if r.get("batch") else None
        ft = r.get("from_tank")
        from_tank = int(ft) if ft not in (None, "") else None
        cnt = r.get("count")
        count = float(cnt) if cnt not in (None, "") else None
        # parse to_tanks: "tank" (bare), "tank:count" (explicit per-dest), either
        # optionally suffixed "@big"/"@small" for fw_to_og size routing. Bare
        # tanks share the `count` column evenly.
        specs = []
        for tok in str(r.get("to_tanks") or "").replace(" ", "").split(","):
            if not tok:
                continue
            specs.append(_parse_dest_token(tok))
        notes = str(r.get("notes") or "")
        if typ in ("og_transfer", "og_to_6n"):
            bare = [t for t, c, _s in specs if c is None]
            per_bare = (count / len(bare)) if (bare and count is not None) else None
            dests = [ManualDest(tank=t, count=(c if c is not None else per_bare),
                                size_class=s)
                     for t, c, s in specs]
            out.append(ManualEvent(type=typ, week=week, from_tank=from_tank,
                                   destinations=dests, batch=batch, notes=notes))
        elif typ == "harvest":
            out.append(ManualEvent(type=typ, week=week, from_tank=from_tank,
                                   count=count, batch=batch, notes=notes))
        elif typ == "fw_to_og":
            # size_class is the whole point here: it routes the big half of the
            # FW cohort to one tank and the small half to another.
            dests = [ManualDest(tank=t, size_class=s) for t, c, s in specs]
            out.append(ManualEvent(type=typ, week=week, batch=batch, count=count,
                                   destinations=dests, notes=notes))
        elif typ == "graded_harvest":
            # from_tank = source, count = biggest-N to harvest, to_tanks =
            # pickup[,retention] (retention defaults to the source). mode
            # "harvest" = drain the pickup in the scripted week; blank/"stage"
            # on a 6N pickup = purge staging (the default; harvested later).
            _mode = str(r.get("mode") or "").strip().lower()
            dests = [ManualDest(tank=t, size_class=s) for t, c, s in specs]
            out.append(ManualEvent(type=typ, week=week, from_tank=from_tank,
                                   count=count, destinations=dests,
                                   batch=batch, notes=notes,
                                   mode=(_mode or "transfer")))
        else:
            out.append(ManualEvent(type=typ, week=week, notes=notes))
    return out


# ---- Shared working set: one in-memory list[ManualEvent] both the visual
# editor and the Advanced raw grid mutate. Seeded once from the YAML; a Save
# dumps it back. (Single source of truth avoids the two surfaces clobbering each
# other's unsaved edits at the YAML boundary.)

def _pr_closing():
    """The uploaded PR's closing date — the PR's business identity, and the
    key manual-event files are scoped by. None when no valid PR is loaded."""
    return pr["closing"] if (pr is not None and pr.get("ok")) else None


def _mw_events():
    """The working set — PR-SCOPED: seeded from THIS PR's event file, and
    RESEEDED whenever the uploaded PR changes (events are statements about
    one PR's starting reality; carrying them across PRs was the bug the
    operator caught 2026-08-07)."""
    from forecast.manual_events import load_manual_events
    cur = st.session_state.get("_pr_key")
    if ("mw_events" not in st.session_state
            or st.session_state.get("_mw_events_pr", "±") != cur):
        # Guarded: this runs as soon as a PR is uploaded, so an event file the
        # loader chokes on used to blank Run mode outright.
        _ok, _evs = _read_or_explain(
            lambda: load_manual_events(SCENARIO_DIR, pr_closing=_pr_closing()),
            "scenario/manual_events/ (this PR's operations)",
            hint="The editor below is showing an EMPTY window — **do not save "
                 "it over the file** until the error above is fixed, or the "
                 "stored operations are lost.")
        st.session_state["mw_events"] = _evs if _ok else []
        st.session_state["_mw_events_pr"] = cur
        _mw_bump_grid()
    return st.session_state["mw_events"]


def _mw_bump_grid():
    """Bump the raw-grid remount nonce so it re-seeds from the working set after
    the visual editor changes it (a data_editor keeps widget state by key)."""
    st.session_state["mw_grid_nonce"] = st.session_state.get("mw_grid_nonce", 0) + 1


def _mw_set(events):
    st.session_state["mw_events"] = list(events)
    _mw_bump_grid()


def _mw_add(ev):
    _mw_events().append(ev)
    _mw_bump_grid()


def _mw_tanks(state):
    """All tanks in heatmap order (system, then tank id)."""
    return sorted(state.tanks_by_id.values(),
                  key=lambda t: (t.system_id or "", t.tank_id))


def _mw_loc(state, tid):
    t = state.tanks_by_id.get(int(tid))
    return t.location_id if t else f"#{tid}"


def _mw_sig(events, extra=""):
    """Cheap stable signature of the working set (+ context) for caching the
    heavy biology projection / validation across idle reruns.

    Includes the config fingerprint: the projection and the reject-at-entry
    validation both run against control/tables/batches, so an in-app config
    save has to invalidate them — otherwise the editor keeps judging events
    against pre-edit knobs while ▶ Run forecast reads the fresh YAML."""
    import hashlib
    import json
    from forecast.manual_events import manual_events_to_list
    payload = json.dumps({"e": manual_events_to_list(events), "x": extra,
                          "pr": st.session_state.get("_mw_pr_key", ""),
                          "cfg": _config_fingerprint()},
                         sort_keys=True, default=str)
    return hashlib.md5(payload.encode()).hexdigest()


def _mw_project(state, ctx, events, n_weeks, view="open"):
    """Project the facility through `n_weeks` of the override window (operator
    events + full biology) on a COPY of the hydrated state — the SAME engine the
    real run uses (forecast.manual_window.advance_facility_window) — and return
    (rows, labels, moves) for `view`:

      view="open"  → start-of-week, pre-event, pre-biology snapshot (what's in
                     the tank WHEN you click to act on it).
      view="close" → end-of-week, post-event, post-biology snapshot (what holds
                     fish and what's empty at week's end; a tank you harvest or
                     move shows empty here).

    `moves` is the view-independent list of tank→tank relocations that happened
    in the window (OG→OG splits and OG→6N sends), one dict per Transfer:
    {"week": label, "src": tank_id, "dests": [tank_id,...], "batch", "count"}.
    The grid uses it to light up BOTH ends of a move in the week it fires — the
    end that actually holds the fish in this view solid, the counterpart end as
    a ghost arrow — so a relocation is legible without diffing open vs close.

    Both snapshots + the moves come from ONE projection and are cached together
    by (PR, events, n_weeks), so toggling the view or clicking around never
    recomputes biology."""
    import copy
    from datetime import timedelta
    from forecast.manual_window import advance_facility_window
    from forecast.time_grid import forecast_week_labels
    # "proj3" (bump on every cached-schema change) so a stale cache from an
    # older schema can never satisfy this {sig,open,close,labels,moves} read.
    sig = _mw_sig(events, extra=f"proj4:{n_weeks}")  # view is NOT in the sig:
    cache = st.session_state.get("_mw_proj_cache")    # one projection feeds both
    if not (cache and cache.get("sig") == sig):
        labels = forecast_week_labels(ctx["forecast_start"], n_weeks)
        moves: list[dict] = []
        err = None
        try:
            sc = copy.deepcopy(state)
            win = advance_facility_window(
                sc, ctx["batch_by_id"], ctx["tables"], ctx["forecast_start"],
                n_weeks, events=events, control=ctx["control"],
                pr_closing=ctx["pr_closing"], fw_records=ctx["fw_records"])
            # Fall back to the closing snapshot for older engines without an
            # opening_locations, and vice-versa, so neither view blanks out.
            open_rows = win.get("opening_locations") or win["batch_locations"]
            close_rows = win.get("batch_locations") or open_rows
            # Each Transfer is dated at its week's start (forecast_start + i*7),
            # the same arithmetic the window loop uses — map it back to a label.
            label_by_date = {ctx["forecast_start"] + timedelta(days=7 * i): lbl
                             for i, lbl in enumerate(labels)}
            for tr in (win.get("transfer_events") or []):
                lbl = label_by_date.get(getattr(tr, "event_date", None))
                if lbl is None:
                    continue
                if hasattr(tr, "pickup_tank_id"):
                    # GradedHarvest carries pickup/retention tanks, not
                    # .destinations — without this branch a scripted grading
                    # draws NO arrows and the fish just "appear" in the grid.
                    dests = [tr.pickup_tank_id]
                    cnt = float(tr.pickup_count)
                    if tr.retention_tank_id != tr.source_tank_id:
                        dests.append(tr.retention_tank_id)
                        cnt += float(tr.retention_count)
                    moves.append({"week": lbl, "src": tr.source_tank_id,
                                  "dests": dests, "batch": tr.batch_id,
                                  "count": cnt})
                    continue
                dests = [a.tank_id for a in getattr(tr, "destinations", []) or []]
                if not dests:
                    continue
                moves.append({"week": lbl, "src": tr.source_tank_id,
                              "dests": dests, "batch": tr.batch_id,
                              "count": getattr(tr, "count_transferred", 0.0)})
        except Exception as e:  # noqa: BLE001 — a bad event must not blank the view
            # Record WHY. Empty rows are indistinguishable from "facility is
            # fine" downstream, so a silent failure here reports an all-clear.
            open_rows, close_rows = [], []
            err = f"{type(e).__name__}: {e}"
        cache = {"sig": sig, "open": open_rows, "close": close_rows,
                 "labels": labels, "moves": moves, "error": err}
        st.session_state["_mw_proj_cache"] = cache
    rows = cache["close"] if view == "close" else cache["open"]
    return rows, cache["labels"], cache.get("moves", [])


def _mw_proj_error():
    """Why the last _mw_project failed, or None.

    A failed projection caches EMPTY rows, and empty rows read downstream as
    "nothing in the facility" — which renders as a clean bill of health rather
    than a failure. Callers must check this before showing any all-within-limits
    verdict, or a crashed projection silently reports the safest possible answer.
    """
    return (st.session_state.get("_mw_proj_cache") or {}).get("error")


def _mw_dark_handoff(state, ctx, events):
    """Handoff-continuity lint for the scripted window: which handoff-era weeks
    have NO harvestable 6N fish under the depuration hold?

    Pure arithmetic on the hydrated start state + the scripted events (the
    detector is forecast.manual_window.dark_handoff_weeks — the same handoff
    semantics both engines now honor); no biology, no extra runs. Returns
    (dark_week_labels, window_n, hold_weeks, stage_lo, stage_hi) or None when
    the handoff is covered. stage_lo..stage_hi is the window-week range where a
    scripted Send-to-6N / Graded-to-6N would arrive in time to cover the dark
    week(s).
    """
    from forecast.global_planner_poc import _PURGE_HOLD_WEEKS as _hold
    from forecast.manual_window import dark_handoff_weeks
    from forecast.sixn import SIXN_ALL_TANKS
    from forecast.time_grid import forecast_week_labels
    if not events:
        return None
    try:
        sixn_start = {t.tank_id: t.count for t in state.tanks_by_id.values()
                      if t.tank_id in SIXN_ALL_TANKS and not t.is_empty}
        n = max(int(e.week or 1) for e in events)
        dark = dark_handoff_weeks(sixn_start, events,
                                  window_weeks=n, hold_weeks=_hold)
        if not dark:
            return None
        labels = forecast_week_labels(ctx["forecast_start"], n + _hold)
        dark_labels = [labels[w - 1] for w in dark if 0 < w <= len(labels)]
        # A staging at window week k releases at k + hold: to cover the first
        # dark week the move must land by (first_dark - hold); anything later
        # than (last_dark - hold) arrives too late for every dark week.
        stage_lo = max(1, min(dark) - _hold)
        stage_hi = min(n, max(dark) - _hold)
        stage_hi = max(stage_hi, stage_lo)
        return dark_labels, n, _hold, stage_lo, stage_hi
    except Exception as e:  # noqa: BLE001 — a lint must never break the editor
        # ...but a crashed lint must not read as "handoff covered": None is
        # also the all-clear return, so say the check didn't run.
        st.caption(f"⚠ dark-handoff lint unavailable ({type(e).__name__}: {e}) "
                   f"— handoff coverage NOT checked.")
        return None


def _mw_validate(state, ctx, events):
    """{event_index(1-based): [messages]} for every infeasible event, faithful to
    the run (forecast.manual_events.validate_manual_events). Cached by the working
    set so the Save gate + per-row status don't re-run biology every rerun."""
    from forecast.manual_events import validate_manual_events
    if not events:
        return {}
    sig = _mw_sig(events, extra="val")
    cache = st.session_state.get("_mw_val_cache")
    if cache and cache.get("sig") == sig:
        return cache["bad"]
    try:
        res = validate_manual_events(state, events, **ctx)
        bad = {i: msgs for i, ok, msgs in res if not ok}
    except Exception as e:  # noqa: BLE001
        bad = {-1: [f"validation unavailable ({type(e).__name__}: {e})"]}
    st.session_state["_mw_val_cache"] = {"sig": sig, "bad": bad}
    return bad


def _mw_fw_avail(ctx, window_labels):
    """{batch_id: {week_label: (count, avg_wt_g, cv)}} for every in-flight FW
    cohort still in freshwater somewhere in the window — the candidates a manual
    FW→OG intake can pull from (projected exactly like the run's _build_fw_lookup,
    but over ALL FW batches, not just ones already referenced by an event)."""
    from collections import defaultdict
    from forecast.biology import project_in_flight_fw_batch
    fw_records = ctx.get("fw_records") or []
    if not fw_records or ctx.get("control") is None:
        return {}
    agg = defaultdict(lambda: {"count": 0.0, "biomass_kg": 0.0})
    for r in fw_records:
        agg[r.batch_id]["count"] += r.closing_count
        agg[r.batch_id]["biomass_kg"] += r.closing_biomass_kg
    win = set(window_labels)
    out: dict[str, dict] = {}
    _skipped = []
    for bid, a in agg.items():
        b_meta = ctx["batch_by_id"].get(bid)
        if a["count"] <= 0 or b_meta is None:
            continue
        avg_wt = a["biomass_kg"] * 1000.0 / a["count"]
        try:
            states, _, _ = project_in_flight_fw_batch(
                b_meta, ctx["tables"], ctx["control"], a["count"], avg_wt,
                ctx["pr_closing"])
        except Exception:  # noqa: BLE001
            _skipped.append(bid)
            continue
        cv = b_meta.tran_og_cv or 16.0
        wk = {s.week_label: (s.close_count, s.close_avg_weight_g, cv)
              for s in states if s.stage == "FW" and s.week_label in win}
        if wk:
            out[bid] = wk
    if _skipped:
        st.caption(f"⚠ FW cohort(s) {', '.join(sorted(_skipped))} could not be "
                   f"projected — excluded from the FW-intake picker.")
    return out


def _mw_fw_load(ctx, window_labels):
    """{week_label: {"open_bio": kg, "close_bio": kg, "feed": kg/day}} summed over
    every in-flight FW cohort still in freshwater in the window — the standing FW
    LOAD to fold into the system rollup as an "FW" row + into the facility total.

    Same projection as _mw_fw_avail, but keeps biomass + feed. The feed figure is
    the projection's own stage-correct daily feed (FW-stage SGR/FCR while the fish
    are in freshwater) — NOT realized_feed_kg_day, which assumes seawater. FW fish
    are fed in the freshwater area, not from OG feed capacity, so the rollup shows
    this row neutral (uncapped) and only rolls it into the neutral TOTAL."""
    from collections import defaultdict
    from forecast.biology import project_in_flight_fw_batch
    fw_records = ctx.get("fw_records") or []
    if not fw_records or ctx.get("control") is None:
        return {}
    agg = defaultdict(lambda: {"count": 0.0, "biomass_kg": 0.0})
    for r in fw_records:
        agg[r.batch_id]["count"] += r.closing_count
        agg[r.batch_id]["biomass_kg"] += r.closing_biomass_kg
    win = set(window_labels)
    out: dict[str, dict] = defaultdict(
        lambda: {"open_bio": 0.0, "close_bio": 0.0, "feed": 0.0})
    _skipped = []
    for bid, a in agg.items():
        b_meta = ctx["batch_by_id"].get(bid)
        if a["count"] <= 0 or b_meta is None:
            continue
        avg_wt = a["biomass_kg"] * 1000.0 / a["count"]
        try:
            states, _, _ = project_in_flight_fw_batch(
                b_meta, ctx["tables"], ctx["control"], a["count"], avg_wt,
                ctx["pr_closing"])
        except Exception:  # noqa: BLE001
            _skipped.append(bid)
            continue
        for s in states:
            if s.stage != "FW" or s.week_label not in win:
                continue
            rec = out[s.week_label]
            rec["open_bio"] += s.open_biomass_kg or s.biomass_kg
            rec["close_bio"] += s.close_biomass_kg or s.biomass_kg
            rec["feed"] += s.feed_kg_day
    if _skipped:
        st.caption(f"⚠ FW cohort(s) {', '.join(sorted(_skipped))} could not be "
                   f"projected — the FW row and facility totals UNDERSTATE "
                   f"their load.")
    return dict(out)


def _mw_grid(state, rows, labels, color_by, batch_filter=None, moves=None):
    """Colour-styled DataFrame of the projected facility (index = tank, columns =
    weeks, cell text = "batch · avg-weight · density") for a CLICKABLE st.dataframe.
    The per-cell weight + density let you read grow-out state at a glance to decide
    moves without clicking every tank. color_by 'fill'
    shades by density-vs-cap (green→red); 'batch' gives each batch its own colour.
    `batch_filter` (a set of batch ids, or None) restricts the rows to only the
    tanks that hold one of those batches in some displayed week — so the operator
    can focus on a few cohorts instead of the whole facility; batch COLOURS stay
    consistent with the unfiltered view.

    `moves` (from _mw_project) lights up BOTH ends of a relocation in the week it
    fires: the tank that holds the fish in this snapshot gets a solid cell with a
    trailing arrow (⇢ leaving / ⇠ arrived), and the counterpart tank — empty in
    this snapshot — gets a faint GHOST cell naming where the fish went / came
    from. So a move reads at a glance instead of by diffing open vs close.

    Returns (styler, ylabels, tank_by_y). Unlike a plotly heatmap, a single click
    on a dataframe row reliably emits a Streamlit selection."""
    from collections import defaultdict
    from forecast.sixn import SIXN_ALL_TANKS
    idx = {(r.tank_id, r.week_label): r for r in rows}
    loc_by_tank = {t.tank_id: t.location_id for t in state.tanks_by_id.values()}

    def _uniq(seq):
        seen, out = set(), []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return ",".join(out)

    # Per-(tank, week) move markers: out_to = where this tank's fish LEAVE to;
    # in_from = where a tank's fish ARRIVE from. Keyed by location label so the
    # arrow text reads OG1N/6N-63 (what the operator sees), not a raw tank id.
    out_to: dict = defaultdict(list)
    in_from: dict = defaultdict(list)
    for m in (moves or []):
        wk, src = m["week"], m["src"]
        out_to[(src, wk)].extend(loc_by_tank.get(d, f"#{d}") for d in m["dests"])
        for d in m["dests"]:
            in_from[(d, wk)].append(loc_by_tank.get(src, f"#{src}"))

    tanks = _mw_tanks(state)
    if batch_filter:
        keep = {r.tank_id for r in rows
                if r.count > 0 and r.batch_id in batch_filter}
        # keep both ends of a filtered batch's moves, even where a snapshot shows
        # the tank empty (the ghost end) — else half the move would vanish.
        for m in (moves or []):
            if m["batch"] in batch_filter:
                keep.add(m["src"])
                keep.update(m["dests"])
        tanks = [t for t in tanks if t.tank_id in keep]
    ubatches = sorted({r.batch_id for r in rows if r.count > 0})
    _pal = px.colors.qualitative.Light24
    bcolor = {b: _pal[i % len(_pal)] for i, b in enumerate(ubatches)}
    # Per-batch BOLD FONT colour so a cohort is trackable across tanks even in Fill
    # mode (where the background encodes density, not batch). Dark24 is a dark
    # palette -> readable on the light/green/amber fills; on the dark over-cap reds
    # we lighten the same hue so it stays legible without losing batch identity.
    _fpal = px.colors.qualitative.Dark24
    fcolor = {b: _fpal[i % len(_fpal)] for i, b in enumerate(ubatches)}

    def _rgb(h):
        if h.startswith("rgb"):
            return tuple(int(x) for x in h[h.find("(") + 1:h.find(")")].split(","))
        h = h.lstrip("#")
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))

    def _lum(h):
        r, g, b = _rgb(h)
        return (0.299 * r + 0.587 * g + 0.114 * b) / 255.0

    def _toward_white(h, t):
        r, g, b = _rgb(h)
        return "#%02x%02x%02x" % (round(r + (255 - r) * t),
                                  round(g + (255 - g) * t),
                                  round(b + (255 - b) * t))

    def _fill_hex(frac):
        if frac <= 0:
            return "#f0f0f0"
        if frac < 0.8:
            return "#a8d5a8"
        if frac < 1.0:
            return "#f5d49a"
        if frac < 1.15:
            return "#e8615e"
        return "#7a0d0b"

    ylabels, tank_by_y = [], {}
    text_grid, css_grid = [], []
    for t in tanks:
        is6n = t.tank_id in SIXN_ALL_TANKS
        yl = f"{t.location_id}" + ("  ⛔6N" if is6n else "")
        ylabels.append(yl)
        tank_by_y[yl] = t.tank_id
        cap = t.max_density_kg_m3 or 0.0
        trow, crow = [], []
        for wk in labels:
            r = idx.get((t.tank_id, wk))
            outs = out_to.get((t.tank_id, wk))
            ins = in_from.get((t.tank_id, wk))
            if r is not None and r.count > 0:
                # Solid occupancy cell. If this tank is a move end THIS week,
                # trail an arrow: ⇢ when the fish here are leaving (open view of
                # a source), ⇠ when they arrived (close view of a destination).
                arrow = (f"  ⇢{_uniq(outs)}" if outs
                         else f"  ⇠{_uniq(ins)}" if ins else "")
                trow.append(
                    f"{r.batch_id} · {r.avg_wt_g / 1000:.2f}kg · "
                    f"{r.density_kg_m3:.0f}{arrow}")
                if color_by == "batch":
                    # background already encodes the batch; keep a dark bold font.
                    bg, fg = bcolor.get(r.batch_id, "#cccccc"), "#1f1f1f"
                else:
                    bg = _fill_hex((r.density_kg_m3 / cap) if cap > 0 else 0.0)
                    base = fcolor.get(r.batch_id, "#1f1f1f")
                    fg = _toward_white(base, 0.7) if _lum(bg) < 0.6 else base
                crow.append(f"background-color:{bg};color:{fg};font-weight:700")
            elif outs or ins:
                # GHOST counterpart: this end is empty in THIS snapshot, but a
                # move touched it this week — show where the fish went / came
                # from so both ends of the relocation are visible in one column.
                trow.append(f"⇢ {_uniq(outs)}" if outs else f"⇠ {_uniq(ins)}")
                crow.append("background-color:#e3e7f0;color:#3a4a7a;"
                            "font-style:italic;font-weight:600")
            else:
                trow.append("")
                crow.append("background-color:#f0f0f0;color:#1f1f1f")
        text_grid.append(trow)
        css_grid.append(crow)
    df = pd.DataFrame(text_grid, index=ylabels, columns=labels)
    css_df = pd.DataFrame(css_grid, index=ylabels, columns=labels)
    styler = df.style.apply(lambda _: css_df, axis=None)
    return styler, ylabels, tank_by_y


def _mw_system_rollup(state, rows, labels, tables, batch_by_id, ctx=None,
                      view="open"):
    """Render, under the grid, per-SYSTEM **biomass (tonnes)** and **feed
    (kg/day)** tables (systems = rows, weeks = columns). Cells are coloured
    green→red by fraction of the system's tank capacity (biomass = Σ volume ×
    max_density; feed = Σ max_feed/day) so system-level capacity pressure — which
    the per-tank grid can't show, especially FEED — is visible at a glance.

    Follows the grid's Week open / Week close `view`. When `ctx` is supplied, a
    neutral **FW (freshwater)** row folds the standing freshwater cohorts into
    each table + the facility TOTAL (see _mw_fw_load); STARVE (6N) tanks feed 0."""
    from collections import defaultdict
    from forecast.biology import realized_feed_kg_day
    sys_bio_cap, sys_feed_cap = defaultdict(float), defaultdict(float)
    for t in state.tanks_by_id.values():
        if t.type != "OG":
            continue
        sys_bio_cap[t.system_id] += (t.volume_m3 or 0.0) * (t.max_density_kg_m3 or 0.0)
        sys_feed_cap[t.system_id] += (t.max_feed_kg_day_cap or 0.0)
    systems = sorted(sys_bio_cap)
    bio, feed = defaultdict(float), defaultdict(float)
    for r in rows:
        if r.count <= 0:
            continue
        bio[(r.system_id, r.week_label)] += r.biomass_kg
        if getattr(r, "stage", "") != "STARVE":
            feed[(r.system_id, r.week_label)] += realized_feed_kg_day(
                r.avg_wt_g, r.biomass_kg, batch_by_id.get(r.batch_id), tables)

    # Standing FW load as a neutral row folded into the TOTAL (biomass matches the
    # view's open/close snapshot; feed is the FW-stage projected daily feed).
    fw = _mw_fw_load(ctx, labels) if ctx else {}
    fw_bio = {wk: v["open_bio" if view == "open" else "close_bio"]
              for wk, v in fw.items()}
    fw_feed = {wk: v["feed"] for wk, v in fw.items()}
    FW_LABEL = "FW (freshwater)"

    def _fill(frac):
        if frac <= 0:
            return "#f0f0f0"
        if frac < 0.8:
            return "#a8d5a8"
        if frac < 1.0:
            return "#f5d49a"
        if frac < 1.15:
            return "#e8615e"
        return "#7a0d0b"

    # Feed is coloured by ABSOLUTE kg/day/system load (operator setting), not by the
    # per-system cap fraction: green below FEED_AMBER, amber between, red at/above
    # FEED_RED. Change the two numbers to retune the thresholds.
    FEED_AMBER, FEED_RED = 2600.0, 2800.0

    def _fill_feed_abs(v):
        if v <= 0:
            return "#f0f0f0"
        if v < FEED_AMBER:
            return "#a8d5a8"
        if v < FEED_RED:
            return "#f5d49a"
        return "#e8615e"

    def _table(agg, cap, scale, extra=None, color_of=None):
        # extra = {wk: raw_value} for a neutral (uncapped) FW row folded into the
        # facility TOTAL — FW fish aren't in an OG system, so no cap fraction.
        # color_of(value, system_cap) -> css colour; default = fraction-of-cap fill.
        if color_of is None:
            color_of = lambda v, caps: _fill((v / caps) if caps else 0.0)  # noqa: E731
        txt, css, idx = [], [], []
        for s in systems:
            idx.append(s)
            trow, crow = [], []
            for wk in labels:
                v = agg.get((s, wk), 0.0)
                trow.append(f"{v * scale:,.0f}")
                crow.append(f"background-color:{color_of(v, cap.get(s, 0.0))};"
                            f"color:#1f1f1f")
            txt.append(trow)
            css.append(crow)
        if extra:
            idx.append(FW_LABEL)
            trow, crow = [], []
            for wk in labels:
                trow.append(f"{extra.get(wk, 0.0) * scale:,.0f}")
                crow.append("background-color:#eef1e8;color:#3a4a2a;"
                            "font-style:italic")
            txt.append(trow)
            css.append(crow)
        idx.append("TOTAL")           # facility total (neutral, not cap-coloured)
        trow, crow = [], []
        for wk in labels:
            tot = sum(agg.get((s, wk), 0.0) for s in systems)
            if extra:
                tot += extra.get(wk, 0.0)
            trow.append(f"{tot * scale:,.0f}")
            crow.append("background-color:#e8eaf0;color:#1f1f1f;font-weight:700")
        txt.append(trow)
        css.append(crow)
        d = pd.DataFrame(txt, index=idx, columns=labels)
        c = pd.DataFrame(css, index=idx, columns=labels)
        return d.style.apply(lambda _: c, axis=None)

    _w = "open" if view == "open" else "close"
    _nrows = len(systems) + 1 + (1 if fw else 0)
    _h = min(560, 44 + 33 * _nrows)
    _fwnote = (f" The **{FW_LABEL}** row is standing freshwater cohorts — fed in "
               "the FW area (no OG cap), shown neutral and folded only into the "
               "TOTAL." if fw else "")
    # Render the matrices in the SAME left 3/5 column the tank grid uses (the
    # st.columns([3, 2]) split above the grid) and let them fill it, so the
    # biomass/feed tables line up under the grid at the same width.
    _mcol, _ = st.columns([3, 2], gap="medium")
    with _mcol:
        st.caption(f"Biomass colour = fraction of each OG system's tank capacity "
                   f"(green roomy, amber near cap, red over).{_fwnote} Same "
                   f"week-{_w} state as the grid.")
        st.markdown(f"**{_w.capitalize()} biomass — tonnes / system / week**")
        st.dataframe(_table(bio, sys_bio_cap, 0.001, fw_bio if fw else None),
                     use_container_width=True, height=_h)
        st.markdown(f"**{_w.capitalize()} feed — kg/day / system / week** "
                    f"(6N depuration eats 0)")
        st.caption(f"Feed colour = absolute load: green < {FEED_AMBER:,.0f} · amber "
                   f"{FEED_AMBER:,.0f}–{FEED_RED:,.0f} · red ≥ {FEED_RED:,.0f} "
                   f"kg/day/system.")
        st.dataframe(_table(feed, sys_feed_cap, 1.0, fw_feed if fw else None,
                            color_of=lambda v, caps: _fill_feed_abs(v)),
                     use_container_width=True, height=_h)


def _mw_recommendations(state, rows, labels, ctx, view="open"):
    """Rank the projected window's cap breaches (most out of bounds FIRST) and
    suggest a relief action for each — harvest the heaviest tank in the offending
    system when it's at harvest weight, else move it to the system with the most
    feed headroom; a per-tank density breach recommends splitting that tank.

    Covers per-system FEED, per-system BIOMASS, per-tank DENSITY and FACILITY
    biomass, all against the SAME caps the System-rollup shows (tank-derived
    system caps + the Control facility cap). Reads the current `rows`, so as the
    operator scripts harvests/moves the breaches shrink. Returns
    (collapsed_worst_first, weeks_with_breaches, {week: breaches_worst_first}); each
    breach is {frac, week, tank_id, msg, action}, `week` the ISO week-of-year label
    (the project's canonical id — forecast.time_grid). The by-week map + week list
    drive the panel's week picker; `collapsed` is the 'all weeks' default (each
    distinct breach shown once, at its worst week)."""
    from collections import defaultdict
    from forecast.biology import realized_feed_kg_day
    from forecast.sixn import SIXN_ALL_TANKS
    from forecast.sixn import is_purge_mode as _is_purge_mode
    from forecast.tiers import effective_density_cap as _eff_cap
    control, tables = ctx.get("control"), ctx.get("tables")
    batch_by_id = ctx.get("batch_by_id") or {}
    min_hg = float(getattr(control, "min_harvest_weight_g", 0) or 0)
    fac_bio_cap = float(getattr(control, "max_biomass_kg", 0) or 0)

    sys_bio_cap, sys_feed_cap, tank_cap = defaultdict(float), defaultdict(float), {}
    # 6N systems are depuration (off-feed, harvest-staging) — never a valid relief
    # destination and not grow-out feed/biomass constraints, so exclude them.
    sixn_systems = {t.system_id for t in state.tanks_by_id.values()
                    if t.tank_id in SIXN_ALL_TANKS}
    for t in state.tanks_by_id.values():
        if t.type != "OG":
            continue
        sys_bio_cap[t.system_id] += (t.volume_m3 or 0.0) * (t.max_density_kg_m3 or 0.0)
        sys_feed_cap[t.system_id] += (t.max_feed_kg_day_cap or 0.0)
        tank_cap[t.tank_id] = t.max_density_kg_m3 or 0.0

    sys_bio, sys_feed, fac_bio = (defaultdict(float), defaultdict(float),
                                  defaultdict(float))
    rows_by_sw, rows_by_w = defaultdict(list), defaultdict(list)
    for r in rows:
        if r.count <= 0:
            continue
        sys_bio[(r.system_id, r.week_label)] += r.biomass_kg
        fac_bio[r.week_label] += r.biomass_kg
        rows_by_sw[(r.system_id, r.week_label)].append(r)
        rows_by_w[r.week_label].append(r)
        if getattr(r, "stage", "") != "STARVE":
            sys_feed[(r.system_id, r.week_label)] += realized_feed_kg_day(
                r.avg_wt_g, r.biomass_kg, batch_by_id.get(r.batch_id), tables)

    def _loc(tid):
        t = state.tanks_by_id.get(tid)
        return t.location_id if t else f"#{tid}"

    def _heaviest(rws):
        cand = [x for x in rws if x.count > 0 and x.tank_id not in SIXN_ALL_TANKS]
        return max(cand, key=lambda x: x.avg_wt_g) if cand else None

    def _roomiest_feed(wk, exclude):
        best, best_head = None, 0.0
        for s, cap in sys_feed_cap.items():
            if s == exclude or s in sixn_systems:  # never relocate INTO 6N
                continue
            head = cap - sys_feed.get((s, wk), 0.0)
            if head > best_head:
                best, best_head = s, head
        return best, best_head

    def _shed_action(rws, sysid, wk):
        """Harvest the heaviest ready tank, else move it to a roomier system."""
        tank = _heaviest(rws)
        if tank is None:
            return "no non-6N tank here to relieve — check 6N / FW inflow", None
        loc = _loc(tank.tank_id)
        if tank.avg_wt_g >= min_hg > 0:
            return (f"**Harvest {loc}** ({tank.batch_id} @ "
                    f"{tank.avg_wt_g / 1000:.2f} kg)"), tank.tank_id
        dest, head = _roomiest_feed(wk, sysid)
        if dest and head > 0:
            return (f"**Move {loc} → {dest}** (feed room ~{head:,.0f} kg/day) — "
                    f"too light to harvest ({tank.avg_wt_g / 1000:.2f} kg)"), tank.tank_id
        return (f"**Move {loc} off {sysid}** — no system has feed headroom, "
                f"consider 6N"), tank.tank_id

    breaches = []
    for (s, wk), used in sys_feed.items():
        cap = sys_feed_cap.get(s, 0.0)
        if s not in sixn_systems and cap > 0 and used > cap:
            breaches.append(("feed", s, wk, used, cap))
    for (s, wk), used in sys_bio.items():
        cap = sys_bio_cap.get(s, 0.0)
        if s not in sixn_systems and cap > 0 and used > cap:
            breaches.append(("sysbio", s, wk, used, cap))
    for r in rows:
        if r.count <= 0:
            continue
        # R8 (tiers.effective_density_cap) -- the SAME rule as the engine and
        # run.py's audit, so a tank called legal there is never called a
        # breach here. This replaced `r.tank_id in SIXN_ALL_TANKS`, a
        # membership test that got the ADVICE wrong, not merely the number: a
        # harvest-prep tank outside 6N is deliberately consolidated and
        # deliberately dense, and flagging it made this panel recommend
        # SPLITTING it -- undoing the consolidation on purpose. The same test
        # also hid real breaches once 6N runs as a PRODUCTION system.
        cap = _eff_cap(tank_cap.get(r.tank_id, 0.0), r.system_id,
                       getattr(r, "stage", ""),
                       _is_purge_mode(control, r.week_start))
        if cap > 0 and cap != float("inf") and r.density_kg_m3 > cap:
            breaches.append(("dens", r.tank_id, r.week_label,
                             r.density_kg_m3, cap))
    # FACILITY biomass must be FW-INCLUSIVE to match the cap it is judged against.
    # `rows` are TANK rows, so they cover OG + 6N only; the freshwater cohorts are
    # tracked separately and were simply missing here. The engine's setpoint counts
    # FW (see the dual-limit setpoint), so leaving it out understated facility
    # biomass and showed headroom that does not exist — the same OG-only-vs-
    # FW-inclusive mismatch already fixed in the results biomass charts.
    _fw_load = _mw_fw_load(ctx, labels) if ctx else {}
    # Match the FW snapshot to the view `rows` came from — mixing open OG rows
    # with close FW biomass would disagree with the rollup TOTAL beside it.
    _fw_key = "open_bio" if view == "open" else "close_bio"
    for wk, used in fac_bio.items():
        used += float((_fw_load.get(wk) or {}).get(_fw_key, 0.0) or 0.0)
        if fac_bio_cap > 0 and used > fac_bio_cap:
            breaches.append(("facbio", None, wk, used, fac_bio_cap))

    out = []
    for kind, where, wk, used, cap in breaches:
        frac = used / cap if cap else 0.0
        if kind == "feed":
            act, tid = _shed_action(rows_by_sw.get((where, wk), []), where, wk)
            msg = (f"**{where}** feed {used:,.0f} / {cap:,.0f} kg/day "
                   f"({frac * 100:.0f}%)")
        elif kind == "sysbio":
            act, tid = _shed_action(rows_by_sw.get((where, wk), []), where, wk)
            msg = (f"**{where}** biomass {used / 1000:.1f} / {cap / 1000:.1f} t "
                   f"({frac * 100:.0f}%)")
        elif kind == "dens":
            tid = where
            act = f"**Split {_loc(where)}** — move part to an empty / roomier tank"
            msg = (f"**{_loc(where)}** density {used:.0f} / {cap:.0f} kg/m³ "
                   f"({frac * 100:.0f}%)")
        else:  # facbio — only harvest reduces the facility total
            tank = _heaviest(rows_by_w.get(wk, []))
            tid = tank.tank_id if tank else None
            act = (f"**Harvest {_loc(tid)}** (heaviest ready, {tank.batch_id} @ "
                   f"{tank.avg_wt_g / 1000:.2f} kg)"
                   if tank and tank.avg_wt_g >= min_hg > 0
                   else "**Harvest facility-wide** — total over cap")
            msg = (f"**Facility** biomass {used / 1000:.1f} / {cap / 1000:.1f} t "
                   f"({frac * 100:.0f}%)")
        out.append({"frac": frac, "week": wk, "tank_id": tid,
                    "msg": msg, "action": act,
                    "key": (kind, where if where is not None else "FAC")})
    # Per-week view (everything out of bounds in a chosen week) + the ordered set
    # of weeks that have ANY breach, for the panel's week picker.
    by_week: dict = {}
    for o in out:
        by_week.setdefault(o["week"], []).append(o)
    for _w in by_week:
        by_week[_w].sort(key=lambda x: -x["frac"])
    breach_weeks = sorted(by_week, key=lambda w: labels.index(w) if w in labels else 10**6)
    # Collapse a breach that recurs across weeks to its WORST week, so the default
    # 'all weeks' top-N shows distinct problems instead of one tank every week.
    worst = {}
    for o in out:
        if o["key"] not in worst or o["frac"] > worst[o["key"]]["frac"]:
            worst[o["key"]] = o
    collapsed = sorted(worst.values(), key=lambda x: -x["frac"])
    return collapsed, breach_weeks, by_week


# ---- The contextual action panel (opens on a tank click) ----

def _mw_split_dests(picks, total, whole):
    """Build ManualDest list mirroring the run's even-split semantics: whole tank
    -> count=None dests (engine splits the whole source); a partial total ->
    explicit per-dest counts (the UI does the division, like the raw grid)."""
    from forecast.manual_events import ManualDest
    if whole or not total:
        return [ManualDest(tank=int(d)) for d in picks]
    per = float(total) / len(picks)
    return [ManualDest(tank=int(d), count=per) for d in picks]


def _mw_cut_weights(avg_wt_g, cv_pct, count, k):
    """(big_avg_g, small_avg_g) for a top-`k`-by-size cut — the SAME split the run
    applies (forecast.biology.upper_truncated_split at the count-implied cutoff),
    so the panel's live readout matches what will actually be moved. Returns
    (None, None) for a degenerate cut (k<=0, k>=count, or no weight)."""
    from statistics import NormalDist
    from forecast.biology import upper_truncated_split
    if not k or k <= 0 or k >= count or avg_wt_g <= 0:
        return None, None
    cv = cv_pct or 16.0
    sigma = avg_wt_g * (cv / 100.0)
    if sigma <= 0:
        return avg_wt_g, avg_wt_g
    z = NormalDist().inv_cdf(1.0 - k / count)
    # Same imperfect grader the engine applies (control.grade_efficiency).
    # Omitting it took upper_truncated_split's 1.0 default -- a PERFECT cut --
    # so this preview promised a cleaner split than the run would deliver.
    _ge = 1.0
    try:
        from forecast.config_io import load_control as _lc
        _ge = float(getattr(_lc(CONFIG_DIR), "grade_efficiency", 1.0) or 1.0)
    except Exception:  # noqa: BLE001
        pass
    return upper_truncated_split(avg_wt_g, cv, avg_wt_g + sigma * z,
                                 grade_efficiency=_ge)


def _mw_occ_at(rows, wlabel):
    """{tank_id: (batch_id, density_kg_m3)} for tanks occupied at `wlabel`."""
    return {x.tank_id: (x.batch_id, x.density_kg_m3)
            for x in rows if x.week_label == wlabel and x.count > 0}


def _mw_dest_fmt(state, occ):
    """format_func for a destination picker: 'LOC · BATCH · DENSITY' when the tank
    holds fish at the selected week (so you see the current batch + density before
    picking), else 'LOC · empty'. `occ` = _mw_occ_at() map."""
    def _f(tid):
        loc = _mw_loc(state, tid)
        o = occ.get(tid)
        return f"{loc} · empty" if o is None else f"{loc} · {o[0]} · {o[1]:.0f} kg/m³"
    return _f


def _mw_action_panel(state, ctx, rows, labels, sel, date_for):
    from forecast.sixn import SIXN_ALL_TANKS
    from forecast.manual_events import ManualDest, ManualEvent
    from forecast.tiers import is_entry, move_allowed
    tid, wlabel, wk = sel
    r = next((x for x in rows if x.tank_id == tid and x.week_label == wlabel), None)
    occupied = r is not None and r.count > 0
    loc = _mw_loc(state, tid)
    dt = date_for.get(wlabel)
    ds = f" · {dt.strftime('%b %d')}" if dt else ""
    head = st.columns([8, 1])
    head[0].markdown(f"#### ▶ {loc} — {wlabel}{ds}")
    if head[1].button("✕", key="mw_close_sel", help="Close this panel"):
        st.session_state.pop("mw_sel", None)
        # bump the grid remount nonce so the selected row clears too
        st.session_state["mw_grid_nonce2"] = \
            st.session_state.get("mw_grid_nonce2", 0) + 1
        st.rerun()
    if occupied:
        st.caption(f"Projected here: batch **{r.batch_id}** · {r.count:,.0f} fish "
                   f"@ {r.avg_wt_g / 1000:.2f} kg · {r.density_kg_m3:.0f} kg/m³")
    else:
        st.caption("Projected **empty** at this week — pick it as a destination "
                   "in a Move or an FW→OG intake.")
        return

    other_og = [t for t in _mw_tanks(state)
                if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                and t.tank_id != tid]
    # Current occupancy at this week + a shared picker format (LOC · batch · density,
    # or LOC · empty) so EVERY destination dropdown shows what's in each tank.
    occ_map = _mw_occ_at(rows, wlabel)
    dfmt = _mw_dest_fmt(state, occ_map)
    # Scope input keys to THIS tank+week so a previous selection's destinations /
    # counts don't linger when you click a different cell.
    sfx = f"{tid}_{wk}"
    # Tier rules (R5): the clicked tank's SYSTEM decides whether harvest / 6N
    # staging is possible at all; its projected avg weight gates entry moves.
    _src_sys = state.tanks_by_id[tid].system_id
    _entry_src = is_entry(_src_sys)
    act = st.radio("What do you want to do here?",
                   ["Harvest", "Graded → 6N", "Move (OG→OG)",
                    "Send to 6N depuration"],
                   horizontal=True, key="mw_act")

    if act == "Harvest":
        if _entry_src:
            st.warning(f"Fish can't be harvested from {loc} ({_src_sys}) — "
                       f"OG1/2 is the entry tier (rule R5): move them forward "
                       f"to OG3-6 first, then harvest from there.")
            return
        whole = st.checkbox("Harvest the whole tank", value=True, key=f"mw_h_whole_{sfx}")
        cnt = None
        if not whole:
            cnt = st.number_input("Fish to harvest", min_value=0.0,
                                  value=float(r.count), step=1000.0, key=f"mw_h_cnt_{sfx}")
        if st.button(f"➕ Add harvest in {wlabel}", key="mw_h_add", type="primary"):
            _mw_add(ManualEvent(type="harvest", week=wk, from_tank=tid,
                                count=(None if whole else cnt)))
            st.rerun()

    elif act == "Graded → 6N":
        if _entry_src:
            st.warning(f"Fish can't be graded to 6N from {loc} ({_src_sys}) — "
                       f"OG1/2 is the entry tier (rule R5): no harvest and no "
                       f"6N staging from entry-tier tanks; move them forward "
                       f"first.")
            return
        st.caption("Grade the tank by size — it **empties**: the **biggest N "
                   "fish** route through a 6N tank and the **smaller remainder "
                   "moves to an OG tank** to keep growing. Conserves count + "
                   "biomass exactly.")
        from forecast.manual_events import MODE_HARVEST, MODE_STAGE
        _timing = st.radio(
            "When are the graded fish harvested?",
            ["Purge first — stage in 6N off-feed, harvested after the "
             "~2-week hold",
             "Harvest them this week (via the 6N staging tank)"],
            key=f"mw_g6_timing_{sfx}",
            help="Purge-first parks the biggest N in the 6N tank frozen "
                 "off-feed; script a later harvest of that tank, or the "
                 "planner takes it after the depuration hold. "
                 "Harvest-this-week drains them to processing in the scripted "
                 "week (the 6N tank is just the staging route — it ends the "
                 "week empty). Either way the run logs a MANUAL EVENT OK line "
                 "saying exactly what happened.")
        _g6_staged = _timing.startswith("Purge")
        # Destinations = EMPTY tanks OR tanks already holding THIS batch (top-up),
        # roomiest-first; each option shows its current batch + density.
        def _dest_opts(tank_ids):
            opts = [t for t in tank_ids
                    if t not in occ_map or occ_map[t][0] == r.batch_id]
            return sorted(opts, key=lambda t: occ_map.get(t, (None, 0.0))[1])
        dest_6n = _dest_opts(sorted(SIXN_ALL_TANKS))   # mains 61/63/65 + sisters 67/69/71
        dest_og = _dest_opts([t.tank_id for t in other_og])
        n_big = st.number_input(
            "Fish to send to 6N (the biggest N)", min_value=0.0,
            max_value=float(r.count), value=float(int(r.count // 2)), step=1000.0,
            key=f"mw_g6_cnt_{sfx}",
            help="The N largest fish are graded out at their (higher) mean weight; "
                 "the rest stay at their (lower) mean.")
        # Live cut-weight readout — the SAME split the run applies. CV travels
        # with the BATCH through every event, so read the hydrated tank's cv
        # only while it still holds this batch; if the window moved fish here,
        # fall back to any hydrated tank of the batch, then the batch's PR cv —
        # the hydration-time cv of the clicked tank would be a stale 0.
        _cvt = state.tanks_by_id.get(tid)
        if _cvt is not None and _cvt.batch_id == r.batch_id and _cvt.cv_pct:
            _cv = _cvt.cv_pct
        else:
            _cv = next((t.cv_pct for t in state.tanks_by_id.values()
                        if t.batch_id == r.batch_id and t.cv_pct), 0.0)
            if not _cv:
                _bm = (ctx.get("batch_by_id") or {}).get(r.batch_id)
                _cv = float(getattr(_bm, "tran_og_cv", 0.0) or 0.0)
        _big, _small = _mw_cut_weights(r.avg_wt_g, _cv, r.count, n_big)
        if _big is not None:
            st.caption(f"↳ biggest **{n_big:,.0f} ≈ {_big / 1000:.2f} kg** → 6N · "
                       f"smaller {r.count - n_big:,.0f} ≈ "
                       f"{_small / 1000:.2f} kg → OG")
        dest6n = st.selectbox(
            "6N depuration tank — biggest fish (· batch · density)",
            options=dest_6n, format_func=dfmt, key=f"mw_g6_dest_{sfx}",
            help="Mains 61/63/65 + sisters 67/69/71 — empty or same-batch; each "
                 "shows its current batch + density (a same-pair main holding a "
                 "different batch = a mixed harvest, so watch the batch column).",
        ) if dest_6n else None
        # Grading empties the source, so the smaller remainder is graded OUT too and
        # moves to an OG tank (empty, or one already holding this batch). Required.
        ret_tank = st.selectbox(
            "Send the smaller fish to — OG tank (· batch · density)",
            options=dest_og, format_func=dfmt, key=f"mw_g6_ret_{sfx}",
            help="Empty or same-batch OG tanks. Grading empties the source; the "
                 "smaller fish move here.",
        ) if dest_og else None
        if not dest_6n:
            st.caption("⚠ No empty / same-batch 6N tank available this week.")
        if not dest_og:
            st.caption("⚠ No empty / same-batch OG tank for the smaller fish.")
        dests = ([ManualDest(tank=int(dest6n)), ManualDest(tank=int(ret_tank))]
                 if dest6n is not None and ret_tank is not None else [])
        if st.button(f"➕ Add graded 6N move in {wlabel}", key="mw_g6_add",
                     type="primary",
                     disabled=(dest6n is None or ret_tank is None
                               or not n_big or n_big <= 0)):
            _mw_add(ManualEvent(type="graded_harvest", week=wk, from_tank=tid,
                                count=n_big, destinations=dests,
                                mode=(MODE_STAGE if _g6_staged else MODE_HARVEST)))
            st.rerun()

    elif act == "Move (OG→OG)":
        # Regular OG grow-out tanks only (the 6N depuration system is reached via
        # Send-to-6N / Graded->6N, not a plain grow-out move). Offer EMPTY tanks or
        # ones already holding this batch, each showing current density, roomiest-first.
        # Tier rules R2-R4 (tiers.move_allowed, judged on the tank's projected avg
        # weight): non-entry sources may never move back into OG1/2; entry sources
        # >= 1 kg may only move forward (OG3-6); entry sources < 1 kg may do both.
        move_dests = sorted(
            (t.tank_id for t in other_og
             if (t.tank_id not in occ_map or occ_map[t.tank_id][0] == r.batch_id)
             and move_allowed(_src_sys, t.system_id, r.avg_wt_g)[0]),
            key=lambda t: occ_map.get(t, (None, 0.0))[1])
        if _entry_src and r.avg_wt_g >= 1000.0:
            st.caption("↳ ≥ 1 kg in the entry tier: forward moves (OG3-6) only "
                       "(rule R3 — the intra-OG1/2 equipment limit).")
        elif not _entry_src:
            st.caption("↳ Grow-out fish never move back into OG1/2 (rule R4).")
        picks = st.multiselect(
            "Destination grow-out tank(s) — empty or same batch (· batch · density)",
            options=move_dests, format_func=dfmt, key=f"mw_m_dest_{sfx}")
        whole = st.checkbox("Move the whole tank (split evenly)", value=True,
                            key=f"mw_m_whole_{sfx}")
        total = None
        if not whole:
            total = st.number_input("Total fish to move (split evenly across dests)",
                                    min_value=0.0, value=float(r.count), step=1000.0,
                                    key=f"mw_m_total_{sfx}")
        if st.button(f"➕ Add move in {wlabel}", key="mw_m_add", type="primary",
                     disabled=not picks):
            _mw_add(ManualEvent(type="og_transfer", week=wk, from_tank=tid,
                                destinations=_mw_split_dests(picks, total, whole),
                                count=(None if whole else total)))
            st.rerun()

    else:  # Send to 6N depuration
        if _entry_src:
            st.warning(f"Fish can't be staged to 6N from {loc} ({_src_sys}) — "
                       f"OG1/2 is the entry tier (rule R5): move them forward "
                       f"to OG3-6 first.")
            return
        picks = st.multiselect(
            "6N depuration tank(s) — mains + sisters (· batch · density)",
            options=sorted(SIXN_ALL_TANKS), format_func=dfmt, key=f"mw_6_dest_{sfx}")
        whole = st.checkbox("Move the whole tank (split evenly)", value=True,
                            key=f"mw_6_whole_{sfx}")
        total = None
        if not whole:
            total = st.number_input("Total fish to send (split evenly)",
                                    min_value=0.0, value=float(r.count), step=1000.0,
                                    key=f"mw_6_total_{sfx}")
        if st.button(f"➕ Add 6N move in {wlabel}", key="mw_6_add", type="primary",
                     disabled=not picks):
            _mw_add(ManualEvent(type="og_to_6n", week=wk, from_tank=tid,
                                destinations=_mw_split_dests(picks, total, whole),
                                count=(None if whole else total)))
            st.rerun()


def _mw_fw_split_preview(ctx, bid, fw_count, fw_avg_wt_g, fw_cv, target, event_date):
    """(big_n, big_avg_g, small_n, small_avg_g) for the entry grade of a FW→OG
    intake — replays the run's handling-mortality + reconcile-to-target bottom
    cull + median size split (manual_events._apply_fw_to_og / biology), so the
    picker preview matches exactly what will be placed. None on a bad projection."""
    from forecast.biology import _apply_bottom_cull, compute_size_class_split
    control = ctx.get("control")
    hf = (control.handling_mortality_pct / 100.0) if control is not None else 0.0
    cnt = fw_count * (1.0 - hf)
    wt = fw_avg_wt_g
    if cnt <= 0 or wt <= 0:
        return None
    if target and cnt > target:
        cnt, wt, _cn, _cb = _apply_bottom_cull(cnt, wt, fw_cv, 1.0 - target / cnt)
    try:
        split = compute_size_class_split(
            batch_id=bid, tran_og_date=event_date,
            post_cull_count=cnt, post_cull_avg_wt_g=wt, cv_pct=fw_cv)
    except Exception:  # noqa: BLE001
        return None
    return (split.big_class_count, split.big_class_avg_wt_g,
            split.small_class_count, split.small_class_avg_wt_g)


def _mw_fw_intake(state, ctx, rows, labels, date_for):
    """FW→OG intake — a freshwater cohort isn't a tank yet, so it gets its own
    picker: choose the cohort, the week, a target count (engine culls down to
    it), and — because the cohort is graded into a bigger + smaller class on
    entry — separate destination tanks for each grade. Rendered into the
    caller's container (no inner expander — the editor already lives in one)."""
    from forecast.sixn import SIXN_ALL_TANKS
    from forecast.manual_events import ManualEvent, ManualDest
    avail = _mw_fw_avail(ctx, labels)
    if not avail:
        st.caption("No in-flight freshwater cohorts are still in freshwater "
                   "during this window.")
        return
    # Seed both pickers from durable copies: adding/deleting a grid event calls
    # st.rerun() BEFORE this section renders, so Streamlit drops the widget-
    # backed keys on that interrupted pass and the selection silently snapped
    # back to the first cohort (same cleanup mechanism as the cpsat_depth fix).
    _opts = sorted(avail)
    _saved_bid = st.session_state.get("_mw_fw_batch_saved")
    bid = st.selectbox("Freshwater cohort", options=_opts, key="mw_fw_batch",
                       index=(_opts.index(_saved_bid)
                              if _saved_bid in _opts else 0))
    st.session_state["_mw_fw_batch_saved"] = bid
    wk_labels = [w for w in labels if w in avail.get(bid, {})]
    if not wk_labels:
        st.caption("This cohort has already crossed to seawater in this window.")
        return
    _saved_wk = st.session_state.get("_mw_fw_week_saved")
    wlabel = st.selectbox(
        "Week to bring it in", options=wk_labels,
        format_func=lambda w: f"{w}"
        + (f" · {date_for[w].strftime('%b %d')}" if date_for.get(w) else ""),
        key="mw_fw_week",
        index=(wk_labels.index(_saved_wk) if _saved_wk in wk_labels else 0))
    st.session_state["_mw_fw_week_saved"] = wlabel
    cnt, _wt, _cv = avail[bid][wlabel]
    wk = labels.index(wlabel) + 1
    # Scope input keys to THIS cohort+week — a fixed key would keep the previous
    # cohort's target count / tank picks alive (Streamlit ignores value= once a
    # key has state), silently scripting e.g. a huge bottom-cull on the new one.
    sfx = f"{bid}_{wlabel}"

    # Planned vs. current: the batch's originally-scheduled TranOG (from the PR /
    # scenario) next to what you're actually picking, so you can see how far off
    # the plan — in timing and in size — this intake is.
    from forecast.time_grid import iso_week_label, _as_date
    b_meta = ctx["batch_by_id"].get(bid)
    # tran_og_date may be a datetime; normalise to date so it subtracts cleanly
    # against the (date) week-starts below.
    _pd = getattr(b_meta, "tran_og_date", None) if b_meta else None
    p_date = _as_date(_pd) if _pd else None
    p_wt = getattr(b_meta, "tran_og_avg_wt_g", None) if b_meta else None
    p_week = iso_week_label(p_date) if p_date else "—"
    p_wt_s = f"{p_wt / 1000:.2f} kg" if p_wt else "—"
    st.markdown(
        "| | Transfer week | Avg weight |\n"
        "|---|---|---|\n"
        f"| **Planned** (PR) | {p_week} | {p_wt_s} |\n"
        f"| **This intake** | {wlabel} | {_wt / 1000:.2f} kg |")
    # One-line read-out of the deltas that matter operationally.
    notes = []
    cur_date = date_for.get(wlabel)
    if p_date and cur_date:
        dwk = round((cur_date - p_date).days / 7)
        notes.append("same week as planned" if dwk == 0 else
                     f"{abs(dwk)} wk {'earlier' if dwk < 0 else 'later'} than planned")
    if p_wt:
        dwt = (_wt - p_wt) / 1000.0
        notes.append("on planned weight" if abs(dwt) < 0.005 else
                     f"{abs(dwt):.2f} kg {'lighter' if dwt < 0 else 'heavier'} "
                     f"than planned")
    if notes:
        st.caption("↳ " + " · ".join(notes))
    # The target is judged AFTER handling mortality (validate_manual_events:
    # avail = fw_count * (1 - handling/100)) — defaulting to the raw FW count
    # would pre-fill a target the validator itself rejects as infeasible.
    _hf = float(getattr(ctx.get("control"), "handling_mortality_pct", 0.0)
                or 0.0) / 100.0
    avail_sw = cnt * (1.0 - _hf)
    st.caption(f"Projected freshwater state: ~{cnt:,.0f} fish at {wlabel}; "
               f"after handling mortality **~{avail_sw:,.0f} can enter "
               f"seawater**. Target is the count entering seawater (the "
               f"engine culls down to it).")
    target = st.number_input("Target fish entering seawater", min_value=0.0,
                             value=float(avail_sw), step=1000.0,
                             key=f"mw_fw_target_{sfx}")

    # Live entry-grade preview — the two classes after handling + cull, so the
    # operator sees the counts/weights they're placing before picking tanks.
    prev = _mw_fw_split_preview(ctx, bid, cnt, _wt, _cv, target, date_for.get(wlabel))
    if prev:
        big_n, big_avg, small_n, small_avg = prev
        st.caption(f"Entry grade → **bigger {big_n:,.0f} ≈ {big_avg / 1000:.2f} kg** · "
                   f"**smaller {small_n:,.0f} ≈ {small_avg / 1000:.2f} kg**")

    occ = {x.tank_id for x in rows if x.week_label == wlabel and x.count > 0}
    # R1: FW arrivals enter ONLY the entry tier (OG1/2) — the pool offers
    # empty entry-tier tanks, never OG3+.
    from forecast.tiers import is_entry as _is_entry
    empty_og = [t.tank_id for t in _mw_tanks(state)
                if t.type == "OG" and t.tank_id not in SIXN_ALL_TANKS
                and _is_entry(t.system_id)
                and t.tank_id not in occ]
    if not empty_og:
        st.caption("⚠ No empty entry-tier (OG1/2) tank at this week — FW "
                   "arrivals may only enter OG1/2 (rule R1); free entry tanks "
                   "first (move their fish forward).")
    dfmt = _mw_dest_fmt(state, _mw_occ_at(rows, wlabel))
    big_picks = st.multiselect(
        "Tank(s) for the BIGGER grade", options=empty_og,
        format_func=dfmt, key=f"mw_fw_big_{sfx}")
    # A tank can't hold both grades — drop the big picks from the small options.
    small_opts = [t for t in empty_og if t not in big_picks]
    small_picks = st.multiselect(
        "Tank(s) for the SMALLER grade", options=small_opts,
        format_func=dfmt, key=f"mw_fw_small_{sfx}")

    need_big = bool(prev and prev[0] > 0)
    need_small = bool(prev and prev[2] > 0)
    gaps = []
    if need_big and not big_picks:
        gaps.append("a tank for the bigger grade")
    if need_small and not small_picks:
        gaps.append("a tank for the smaller grade")
    if gaps:
        st.caption("⚠ Still need " + " and ".join(gaps) + ".")
    dests = ([ManualDest(tank=int(t), size_class="big") for t in big_picks]
             + [ManualDest(tank=int(t), size_class="small") for t in small_picks])
    if st.button(f"➕ Add FW→OG intake in {wlabel}", key="mw_fw_add",
                 type="primary", disabled=bool(gaps) or not dests):
        _mw_add(ManualEvent(type="fw_to_og", week=wk, batch=bid, count=target,
                            destinations=dests))
        st.rerun()


# ---- Readback timeline + save bar ----

def _mw_iso_week(week, forecast_start):
    """ISO week-of-year label (e.g. '2026-W08') for a 1-based override-window week —
    the project's canonical week identifier (forecast.time_grid), matching the grid
    columns, the feed matrix and the co-pilot. Falls back to 'Wk N' when
    forecast_start is unknown (non-hydrated PR)."""
    from forecast.time_grid import week_label
    if forecast_start is None or not week:
        return f"Wk {week}"
    try:
        return week_label(int(week) - 1, forecast_start)
    except Exception:  # noqa: BLE001
        return f"Wk {week}"


def _mw_move_amount(ev):
    """How much a transfer/6N event actually moves, for the timeline label.

    These event types carry their counts PER DESTINATION, not on ev.count — so
    reading ev.count alone always found None and every partial move was
    described as "whole tank". That misreads what the plan will do, and it is
    exactly what a grid round-trip produces (each dest gets an explicit count).
    Only claim a number when EVERY destination has one; a mix of explicit and
    bare destinations is genuinely ambiguous (the bare ones split the remainder).
    """
    ds = ev.destinations or []
    if ev.count is not None:
        return f"{ev.count:,.0f}"
    if ds and all(d.count is not None for d in ds):
        return f"{sum(d.count for d in ds):,.0f}"
    return "whole tank"


def _mw_event_summary(state, ev, forecast_start=None):
    loc = lambda t: _mw_loc(state, t)  # noqa: E731
    dests = ", ".join(loc(d.tank) for d in ev.destinations) or "—"
    wk = _mw_iso_week(ev.week, forecast_start)   # week-of-year, like the matrix
    if ev.type == "harvest":
        amt = f"{ev.count:,.0f} fish" if ev.count is not None else "the whole tank"
        return f"{wk}: **Harvest** {amt} from {loc(ev.from_tank)}"
    if ev.type == "og_transfer":
        return (f"{wk}: **Move** {_mw_move_amount(ev)} from "
                f"{loc(ev.from_tank)} → {dests}")
    if ev.type == "og_to_6n":
        return (f"{wk}: **Send to 6N** {_mw_move_amount(ev)} from "
                f"{loc(ev.from_tank)} → {dests}")
    if ev.type == "fw_to_og":
        tgt = f"target {ev.count:,.0f}" if ev.count else "all available"
        big = ", ".join(loc(d.tank) for d in ev.destinations
                        if (d.size_class or "").lower() == "big")
        small = ", ".join(loc(d.tank) for d in ev.destinations
                          if (d.size_class or "").lower() == "small")
        if big or small:
            return (f"{wk}: **FW→OG** {ev.batch} → bigger {big or '—'} · "
                    f"smaller {small or '—'} ({tgt})")
        return f"{wk}: **FW→OG** {ev.batch} → {dests} ({tgt})"
    if ev.type == "graded_harvest":
        from forecast.manual_events import is_staged_graded
        from forecast.sixn import SIXN_ALL_TANKS
        amt = f"{ev.count:,.0f}" if ev.count else "?"
        pk_id = ev.destinations[0].tank if ev.destinations else None
        pk = loc(pk_id) if pk_id is not None else "—"
        ret = (loc(ev.destinations[1].tank) if len(ev.destinations) >= 2
               else "source")
        if pk_id in SIXN_ALL_TANKS and is_staged_graded(ev):
            return (f"{wk}: **Graded → 6N (purge)** biggest {amt} from "
                    f"{loc(ev.from_tank)} → 6N {pk} off-feed (harvested "
                    f"later), retain smaller in {ret}")
        return (f"{wk}: **Graded harvest** biggest {amt} from "
                f"{loc(ev.from_tank)} harvested this week (via {pk}), "
                f"retain smaller in {ret}")
    return f"{wk}: {ev.type}"


def _mw_timeline(state, events, bad, forecast_start=None):
    st.markdown("**Scripted operations — this window**")
    if not events:
        st.caption("None yet. Click a tank in the grid above (or use the FW→OG "
                   "intake) to add your first operation.")
        return
    for i, ev in enumerate(events, 1):
        c1, c2 = st.columns([10, 1])
        problems = bad.get(i)
        with c1:
            if problems:
                st.markdown(f"❌ {_mw_event_summary(state, ev, forecast_start)}")
                st.caption("&nbsp;&nbsp;&nbsp;↳ " + "; ".join(problems))
            else:
                st.markdown(f"✅ {_mw_event_summary(state, ev, forecast_start)}")
            if ev.notes:
                st.caption(f"&nbsp;&nbsp;&nbsp;_{ev.notes}_")
        if c2.button("🗑", key=f"mw_del_{i}", help="Delete this operation"):
            _mw_events().pop(i - 1)
            _mw_bump_grid()
            st.rerun()


def _mw_raw_grid(state):
    """Power-user fallback: the same five event types as a flat table. Seeds from
    and writes back to the shared working set (not the YAML directly)."""
    st.caption(
        "The same five event types as a raw table — for bulk edits or unequal "
        "per-tank splits the click flow doesn't cover. **Apply to window** pushes "
        "these rows into the visual editor + timeline above.")
    st.caption(
        "**og_transfer**: from_tank → to_tanks (count split evenly) · "
        "**harvest**: from_tank, count · **graded_harvest**: from_tank, "
        "count=biggest-N, to_tanks=pickup[,retention]; a 6N pickup defaults to "
        "mode `stage` (purge, harvested later) — set mode `harvest` to drain "
        "it in the scripted week · "
        "**og_to_6n**: from_tank → 6N to_tanks · "
        "**fw_to_og**: batch + count=target → to_tanks. "
        "to_tanks: comma-separated; `tank:count` for an explicit per-tank "
        "amount; `tank@big` / `tank@small` to route an fw_to_og size split "
        "(combinable: `45:1000@small`).")
    nonce = st.session_state.get("mw_grid_nonce", 0)
    base = pd.DataFrame(_manual_events_to_df_rows(_mw_events()), columns=_MANUAL_COLS)
    edited = st.data_editor(
        base, num_rows="dynamic", hide_index=True, use_container_width=True,
        key=f"mw_grid_{nonce}",
        column_config={
            "week": st.column_config.NumberColumn("Week", min_value=1, step=1,
                help="1-based forecast week the event fires in"),
            "type": st.column_config.SelectboxColumn("Type",
                options=["og_transfer", "harvest", "graded_harvest",
                         "og_to_6n", "fw_to_og"]),
            "batch": st.column_config.TextColumn("Batch", help="FW batch (fw_to_og)"),
            "from_tank": st.column_config.NumberColumn("From tank", step=1),
            "to_tanks": st.column_config.TextColumn("To tanks",
                help="comma-separated tank IDs; tank:count for an explicit amount"),
            "count": st.column_config.NumberColumn("Count / target", step=1000),
            "mode": st.column_config.SelectboxColumn("Mode",
                options=["", "stage", "harvest"],
                help="graded_harvest only. What decides the outcome is the "
                     "PICKUP tank plus this field. Pickup is a 6N tank + mode "
                     "blank or 'stage' = park the biggest N there to purge, "
                     "harvested later (the default). Pickup is a 6N tank + "
                     "mode 'harvest' = drain them to processing in the "
                     "scripted week. Pickup is an ordinary OG tank = harvest "
                     "now, and mode 'stage' is refused."),
            "notes": st.column_config.TextColumn("Notes"),
        })
    if st.button("Apply to window", key="mw_grid_apply"):
        try:
            _mw_set(_rows_to_manual_events(_records(edited)))
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Couldn't parse the grid: {e}")


def _mw_save_bar(events, bad):
    from forecast.manual_events import dump_manual_events, load_manual_events
    n = len(events)
    # -1 is _mw_validate's "validation itself crashed" sentinel — it has no
    # timeline row, so without this branch the bar says "fix the ❌ rows above"
    # while every row shows ✅ and the actual exception is displayed nowhere.
    _sentinel = (bad or {}).get(-1)
    _n_bad = len([k for k in (bad or {}) if k != -1])
    if n and not bad:
        st.success(f"All {n} operation(s) feasible against the uploaded PR.")
    elif _sentinel:
        st.error("Couldn't validate the window — " + "; ".join(_sentinel)
                 + ". Saving is disabled until validation runs; check the "
                   "operations in the raw table below, or ↻ Reload from file.")
        if _n_bad:
            st.warning(f"{_n_bad} operation(s) infeasible — fix the ❌ rows "
                       f"above before saving.")
    elif bad:
        st.warning(f"{len(bad)} operation(s) infeasible — fix the ❌ rows above "
                   f"before saving.")
    c1, c2, c3, _ = st.columns([1, 1, 1, 2])
    if c1.button("💾 Save window", key="mw_save", disabled=bool(bad),
                 help="Reject-at-entry: disabled while any operation is infeasible."):
        try:
            dump_manual_events(SCENARIO_DIR, events, pr_closing=_pr_closing())
            st.success(f"Saved {n} operation(s) for THIS PR (closing "
                       f"{_pr_closing()}) — scenario/manual_events/. "
                       f"Click ▶ Run forecast.")
        except Exception as e:  # noqa: BLE001
            st.error(f"Save failed: {e}")
    if c2.button("↻ Reload from file", key="mw_reload"):
        _ok_r, _evs_r = _read_or_explain(
            lambda: load_manual_events(SCENARIO_DIR, pr_closing=_pr_closing()),
            "scenario/manual_events/ (this PR's operations)",
            hint="Nothing was reloaded — the window on screen is unchanged.")
        if _ok_r:
            _mw_set(_evs_r)
            st.rerun()
    if c3.button("🧹 Clear window", key="mw_clear", disabled=not n):
        _mw_set([])
        st.rerun()


def _mw_copilot(uploaded, events, forecast_start=None, bad=None):
    """Human-in-the-loop co-pilot (Sequential). Runs the planners forward ONCE
    from the scripted window (respect mode — your transfers are fixed) and lets
    you browse the recommended moves for the next several weeks. Only the NEXT
    week is approvable: ticked moves append as that week's operations, extending
    the window by one week — then run again for the week after. The look-ahead
    weeks are projections from the current plan and refresh once the nearer weeks
    are scripted. Engine is forecast.copilot (UI-free); this is just the shell.

    `bad` is the save bar's infeasibility map. Both buttons here write
    scenario/manual_events.yaml — the same file ▶ Run forecast reads — so they
    honour the same reject-at-entry gate the Save button does."""
    import shutil
    import tempfile
    from forecast.copilot import propose_upcoming, to_manual_events
    from forecast.manual_events import dump_manual_events
    _n = max((e.week or 1) for e in events) if events else 0
    _LOOKAHEAD = 6
    _blocked = bool(bad)
    if _blocked:
        st.warning(f"{len(bad)} operation(s) in the window are infeasible — fix the "
                   "❌ rows above before running the co-pilot. It has to save your "
                   "window to disk first, and that is the file the forecast reads.")
    # A stale proposal is worse than none: it was computed against a different
    # working set / PR, so its week number and tank picks may no longer line up.
    _cp_sig = _mw_sig(events, extra="copilot1")
    if st.session_state.get("mw_cp_sig") != _cp_sig:
        st.session_state.pop("mw_cp_props", None)
    # Approve reruns immediately, so its confirmation has to survive the rerun.
    _flash = st.session_state.pop("mw_cp_flash", None)
    if _flash:
        st.success(_flash)
    st.caption(
        "Runs the controller + global optimiser forward from your window — your "
        "transfers stay fixed — and recommends the **next** week plus a few weeks "
        "of **look-ahead**. Harvest + 6N staging come from the validated "
        "controller (pre-ticked); the OG↔OG transfer plan comes from the global "
        "optimiser (ranked, opt-in). Tick what you want on the **next** week, "
        "approve, and it's added as that week's ops — then run again for the week "
        "after. Look-ahead weeks are view-only projections. **~20-30 s per run.**")
    if events:
        from forecast.manual_events import is_staged_graded as _isg
        _tr = sum(1 for e in events if e.type == "og_transfer")
        _hv = sum(1 for e in events if e.type == "harvest"
                  or (e.type == "graded_harvest" and not _isg(e)))
        _s6 = sum(1 for e in events if e.type == "og_to_6n"
                  or (e.type == "graded_harvest" and _isg(e)))
        _fw = sum(1 for e in events if e.type == "fw_to_og")
        st.caption(
            f"✓ Building on your **{len(events)} scripted operation(s)** through "
            f"{_mw_iso_week(_n, forecast_start)}: {_tr} transfer · {_hv} harvest · "
            f"{_s6} into 6N · {_fw} FW→OG. Both engines run your full manual window "
            f"first, so every recommendation is computed on top of these.")
    _next_iso = _mw_iso_week(_n + 1, forecast_start)
    if st.button(f"🤖 Recommend from {_next_iso}", key="mw_cp_run", type="primary",
                 disabled=_blocked,
                 help="Disabled while any operation is infeasible — the co-pilot "
                      "saves your window to disk before it runs."
                      if _blocked else None):
        try:
            dump_manual_events(SCENARIO_DIR, events, pr_closing=_pr_closing())
        except Exception as e:  # noqa: BLE001
            st.error(f"Couldn't save your events first: {e}")
            return
        _wd = tempfile.mkdtemp(prefix="as_copilot_")   # per-run dir: no cross-session clobber
        tmp = Path(_wd) / "copilot_pr.xlsm"
        tmp.write_bytes(uploaded.getvalue())
        with st.spinner(f"Running the controller + global optimiser from "
                        f"{_next_iso}… (~20-30 s)"):
            try:
                st.session_state["mw_cp_props"] = propose_upcoming(
                    str(tmp), str(CONFIG_DIR), str(SCENARIO_DIR), n_weeks=_LOOKAHEAD)
                # Stamp the working set these proposals were computed against —
                # any later edit (or a new PR) invalidates them on the next render.
                st.session_state["mw_cp_sig"] = _cp_sig
                st.session_state["mw_cp_nonce"] = \
                    st.session_state.get("mw_cp_nonce", 0) + 1
            except Exception as e:  # noqa: BLE001
                st.session_state.pop("mw_cp_props", None)
                st.error(f"Co-pilot failed: {type(e).__name__}: {e}")
                return
            finally:
                shutil.rmtree(_wd, ignore_errors=True)

    props = st.session_state.get("mw_cp_props")
    if not props:
        return
    handoff = props[0]

    # Week picker — browse any upcoming week; approval stays on the handoff week
    # (Sequential model: you approve one contiguous week at a time).
    def _wk_label(i):
        p = props[i]
        return (f"{p.week_label}  ·  week {p.window_week}  ·  "
                f"{'next — approvable' if i == 0 else 'look-ahead (view only)'}")
    sel = st.selectbox("Show recommendations for week", range(len(props)),
                       format_func=_wk_label, key="mw_cp_week_sel")
    prop = props[sel]
    is_handoff = (sel == 0)

    for _w in prop.warnings:
        st.caption(f"⚠ {_w}")
    if prop.is_empty():
        st.info(f"The planners recommend no moves for {prop.week_label}.")
        return

    if is_handoff:
        st.markdown(f"**Recommended for {prop.week_label}** — ticked moves become "
                    f"that week's operations:")
    else:
        st.info(f"👁 **Look-ahead: {prop.week_label}.** "
                f"A projection from the current plan — approve advances one week at "
                f"a time, so approve {handoff.week_label} first. These numbers "
                f"refresh once the "
                f"nearer weeks are scripted.")

    _nonce = st.session_state.get("mw_cp_nonce", 0)
    picks: list = []

    def _row(m, i, label, default):
        # Handoff: a tickable checkbox (approvable). Look-ahead: a read-only line.
        if not is_handoff:
            st.markdown(f"- {label}")
            return
        key = f"mw_cp_pick_{_nonce}_{i}"
        c1, c2 = st.columns([1, 16])
        c1.checkbox("pick", value=default, key=key, label_visibility="collapsed")
        c2.markdown(label)
        picks.append((m, key))

    i = 0
    if prop.harvest_recs:
        st.markdown("**① Harvest — controller (contract / caps)**")
        for m in prop.harvest_recs:
            _row(m, i, f"Harvest **{m.from_loc}** · {m.batch} · "
                       f"{m.count:,.0f} fish @ {m.avg_wt_kg:.2f} kg", True)
            i += 1
    if prop.sixn_recs:
        st.markdown("**② Stage into 6N for harvest — controller**")
        for m in prop.sixn_recs:
            _lbl = (f"**{m.from_loc} → {m.to_loc}** · {m.batch} · "
                    f"{m.count:,.0f} fish")
            if m.kind == "grade_to_6n":     # graded pickup — say what approval does
                _lbl += f" · _{m.note}_"
            _row(m, i, _lbl, True)
            i += 1
    for opt in prop.transfer_options:
        mv = sorted(opt.moves, key=lambda x: -x.count)
        st.markdown(f"**③ OG↔OG transfers — {opt.label}** · _{opt.why}_ "
                    f"— {len(mv)} moves (the optimiser's full transition; "
                    f"tick the ones you want)")
        for m in mv:
            _row(m, i, f"**{m.from_loc} → {m.to_loc}** · {m.batch} · "
                       f"{m.count:,.0f} fish · _{m.note}_", False)
            i += 1

    if not is_handoff:
        st.caption(f"Viewing a look-ahead week — switch the picker back to "
                   f"**{handoff.week_label}** to approve.")
        return

    if st.button(f"✓ Approve ticked → add to {prop.week_label}",
                 key="mw_cp_approve", type="primary", disabled=_blocked,
                 help="Disabled while any operation is infeasible."
                      if _blocked else None):
        chosen = [m for m, key in picks if st.session_state.get(key)]
        if not chosen:
            st.warning("Nothing ticked — tick at least one move to approve.")
        else:
            for ev in to_manual_events(chosen, prop.window_week):
                _mw_add(ev)
            try:
                dump_manual_events(SCENARIO_DIR, _mw_events(), pr_closing=_pr_closing())
            except Exception as e:  # noqa: BLE001
                # Never claim success on a failed write: scenario/ is OneDrive-
                # synced and the dump can lose a lock race (see yaml_atomic).
                # The ops are in the working set, so Save window can retry.
                st.session_state.pop("mw_cp_props", None)
                st.error(f"Added {len(chosen)} operation(s) to the window, but "
                         f"couldn't write scenario/manual_events.yaml: {e}. "
                         f"They are NOT on disk yet — use 💾 Save window to retry.")
                return
            st.session_state.pop("mw_cp_props", None)
            # The window grew, so the copilot signature moved with it; re-stamp
            # it or the next render would drop the (already consumed) proposal.
            st.session_state["mw_cp_sig"] = _mw_sig(_mw_events(), extra="copilot1")
            st.session_state["mw_cp_flash"] = (
                f"Added {len(chosen)} operation(s) to {prop.week_label}. "
                f"Run me again for the next week "
                f"({_mw_iso_week(prop.window_week + 1, forecast_start)}).")
            st.rerun()


def _manual_window_editor(uploaded):
    """Run-mode editor: SEE the projected facility week by week, click a tank to
    act on it in context (harvest / move / 6N / FW→OG), validated against the
    uploaded PR, saved to scenario/manual_events.yaml (which the run reads). The
    flat grid lives on behind an Advanced expander. No Excel sheets involved."""
    from forecast.time_grid import week_start as _week_start
    # The section-toggle widgets (rollup / FW intake / advanced) render BELOW the
    # clickable grid + action panel. When an action-panel button (add harvest/
    # move/6N, close ✕) or a timeline delete calls st.rerun(), the script aborts
    # BEFORE those toggles are re-instantiated that run — and Streamlit drops the
    # state of any keyed widget it didn't render, snapping the toggles back to
    # off (their content vanishes while the switch still looks on). Re-touching
    # the keys here, at the top (which always runs), keeps their state across
    # such a rerun. (Verified against a headless Streamlit repro.)
    for _tk in ("mw_rollup_toggle", "mw_fw_toggle", "mw_adv_toggle",
                "mw_copilot_toggle"):
        if _tk in st.session_state:
            st.session_state[_tk] = st.session_state[_tk]
    with st.expander("🗓 Starting setup — manual override window (optional)",
                     expanded=False):
        st.caption(
            "See the facility projected forward and **click a tank to act on "
            "it** — harvest it, move/split it, send it to 6N, size-grade it "
            "into 6N, or bring a freshwater cohort into OG. The forecast "
            "EXECUTES your operations with full biology (growth/mortality/"
            "feed), records them in the reports, then the planner takes over "
            "after your last scripted week. **A window week runs your script "
            "and nothing else** — no planner logic at all, so a week you give "
            "no harvest harvests nothing (you get warned, not blocked). Every "
            "operation writes a `MANUAL EVENT OK` or `MANUAL EVENT REFUSED` "
            "line into the output workbook's ValidationLog, so nothing ever "
            "happens — or fails to happen — silently. Leave the window empty "
            "to let the planner do everything.")

        try:
            state, _fw, ctx = _hydrate_state_from_upload(uploaded)
            import hashlib
            st.session_state["_mw_pr_key"] = hashlib.md5(uploaded.getvalue()).hexdigest()
            hydrated = True
        except Exception as e:  # noqa: BLE001
            st.warning(f"Couldn't read the facility from this PR "
                       f"({type(e).__name__}: {e}) — the visual view needs a "
                       f"hydratable PR. Use the raw grid below.")
            state, ctx, hydrated = None, None, False

        events = _mw_events()
        bad = _mw_validate(state, ctx, events) if hydrated else {}

        # Handoff lint (WARNING, not a block — the operator may intend it): a
        # window that drains 6N without restaging leaves the planner nothing
        # releasable at handoff under the depuration hold — a dark week the
        # engines will now faithfully show instead of papering over.
        if hydrated:
            _dh = _mw_dark_handoff(state, ctx, events)
            if _dh:
                _dl, _dn, _dhold, _slo, _shi = _dh
                _wk_txt = ", ".join(f"**{w}**" for w in _dl)
                _rng = (f"week {_slo}" if _slo == _shi
                        else f"weeks {_slo}-{_shi}")
                st.warning(
                    f"⚠ **Your window drains 6N without restaging** — "
                    f"{'week ' if len(_dl) == 1 else 'weeks '}{_wk_txt} will "
                    f"have **no harvestable fish** under the {_dhold}-week "
                    f"depuration hold (fish must sit off-feed in 6N for "
                    f"{_dhold} weeks before harvest, so the planner can't "
                    f"release anything it hasn't staged in time). Script "
                    f"**Send-to-6N** / **Graded-to-6N** moves in {_rng}, or "
                    f"accept a dark week.")

        if hydrated:
            horizon = int(getattr(ctx["control"], "horizon_weeks", 52) or 52)
            max_ev = max((e.week or 1) for e in events) if events else 0
            cap_view = min(max(1, horizon - 1), 26)
            default_view = min(max(8, max_ev), cap_view)
            if cap_view <= 1:
                view = 1
            else:
                view = st.slider(
                    "Weeks to project / act in", 1, cap_view,
                    min(max(default_view, 1), cap_view),
                    help="How far ahead to project the facility and let you act. "
                         "The saved window length stays implicit — it runs through "
                         "your last scripted operation, then the planner takes over.")
            n_weeks = max(view, max_ev, 1)
            _vmode = st.radio(
                "Show tank state at", ["Week open", "Week close"],
                horizontal=True, key="mw_view_at",
                help="Week open = start-of-week, before that week's growth AND "
                     "before your scripted events run — what's in the tank when "
                     "you click to act on it. Week close = end-of-week, after "
                     "growth and after your events run — so you can see what "
                     "holds fish and what's empty at week's end (a tank you "
                     "harvest or move shows empty here, at the week you act).")
            _view = "close" if _vmode.startswith("Week close") else "open"
            rows, labels, moves = _mw_project(
                state, ctx, events, n_weeks, view=_view)
            date_for = {lbl: _week_start(i, ctx["forecast_start"])
                        for i, lbl in enumerate(labels)}

            _cmode = st.radio(
                "Colour cells by", ["Fill (density)", "Batch"], horizontal=True,
                key="mw_color_by",
                help="Fill = how full each tank is vs its cap (green→red). Batch = "
                     "a distinct colour per batch, to see which tanks hold which "
                     "fish and how a batch moves across the weeks.")
            _cb = "batch" if _cmode.startswith("Batch") else "fill"
            _when = (
                "**week-open** (start of the week, before that week's growth and "
                "before your scripted events) — what's in the tank when you act"
                if _view == "open" else
                "**week-close** (end of the week, after growth and after your "
                "scripted events) — what holds fish and what's empty at week's "
                "end; a tank you harvest or move shows empty here")
            st.caption(
                ("**Each batch has its own background colour** (grey = empty). "
                 if _cb == "batch" else
                 "**Background = how full each tank is vs its cap** — grey empty, "
                 "green roomy, amber near cap, red over — and the **batch id is bold "
                 "in its own colour** so you can follow a cohort across tanks. ")
                + "Columns are weeks, rows are tanks (⛔6N = depuration), and each cell "
                  "shows **batch · avg weight · density** at " + _when + ". "
                  "A **move lights up both ends in its week**: the tank that holds "
                  "the fish in this view is solid with an arrow (**⇢** leaving / "
                  "**⇠** arrived), and the counterpart tank shows a faint **ghost "
                  "arrow** naming where the fish went / came from. "
                  "**Click a tank's cell at the week you want** to act on it. "
                  "The grid redraws as you script.")

            # Optional batch filter — show only the tanks holding the selected
            # cohort(s) instead of the whole facility (empty = show all).
            _all_batches = sorted({r.batch_id for r in rows if r.count > 0})
            _sel_batches = st.multiselect(
                "Filter to batches (empty = whole facility)", options=_all_batches,
                default=[], key="mw_batch_filter",
                help="Show only the tanks that hold the selected batch(es) in some "
                     "displayed week. Batch colours stay the same as the full view.")
            _bf = set(_sel_batches) or None

            # Clickable facility grid (left) + contextual action panel (right), side
            # by side so a cell click shows the options right next to the grid instead
            # of below a tall, internally-scrolling table.
            styler, ylabels, tank_by_y = _mw_grid(
                state, rows, labels, color_by=_cb, batch_filter=_bf, moves=moves)
            if _bf and not ylabels:
                st.caption("No tanks hold the selected batch(es) in this window.")
            gnonce = st.session_state.get("mw_grid_nonce2", 0)
            grid_col, panel_col = st.columns([3, 2], gap="medium")
            with grid_col:
                # single-cell selection (Streamlit >=1.49) reliably emits the clicked
                # (row, column), which plotly-heatmap clicks do not. _mw_grid lays the
                # frame out rows=tanks / columns=week-labels, so one click picks BOTH
                # the tank AND the week — no separate week control needed.
                gev = st.dataframe(
                    styler, use_container_width=True,
                    # Full height: show EVERY tank row without a vertical
                    # scrollbar so the whole facility is visible at once (tall
                    # for large facilities, by design — weeks still scroll
                    # horizontally). +2px avoids a residual scrollbar from
                    # row-height rounding.
                    height=46 + 35 * len(ylabels),
                    on_select="rerun", selection_mode="single-cell",
                    key=f"mw_grid_sel_{gnonce}")
                try:
                    _cells = list(gev["selection"]["cells"])
                except Exception:  # noqa: BLE001
                    _cells = list(
                        getattr(getattr(gev, "selection", None), "cells", []) or [])
                if _cells:
                    _rowpos, _wlabel = _cells[0]
                    if 0 <= _rowpos < len(ylabels) and _wlabel in labels:
                        st.session_state["mw_sel"] = (
                            tank_by_y[ylabels[_rowpos]], _wlabel,
                            labels.index(_wlabel) + 1)
            with panel_col:
                # Recommendations — what's most out of bounds this window and the
                # relief action, ranked worst-first. Reads the current projection,
                # so it updates as you script events. ▶ jumps to that tank.
                _collapsed, _breach_weeks, _by_week = _mw_recommendations(
                    state, rows, labels, ctx, view=_view)
                with st.container(border=True):
                    st.markdown("**⚠ Most out of bounds — recommended actions**")
                    _proj_err = _mw_proj_error()
                    if _proj_err:
                        # The projection failed, so `rows` is empty and NOTHING can
                        # look out of bounds. Never let that read as an all-clear.
                        st.error(
                            f"Projection failed — **limits could not be checked**. "
                            f"Treat this as unknown, not as within-limits.\n\n"
                            f"`{_proj_err}`"
                        )
                    elif not _breach_weeks:
                        # Say WHAT was checked. The 6N depuration tanks and
                        # systems are excluded from all three per-tank/per-
                        # system checks here (harvest-size fish held dense and
                        # off-feed is expected) — but 6N's own biomass cap IS
                        # enforced on a real run's SystemLimitsAudit sheet, so
                        # a bare "everything is within limits" would promise
                        # more than this panel looked at.
                        st.caption(
                            "✓ Every grow-out tank and system, and the "
                            "whole-facility biomass total, are within limits "
                            "across this window. Tanks whose fish are off "
                            "feed for harvest carry NO density cap (rule R8) "
                            "— 6N while it purges, and ANY tank starving in "
                            "place — so they are not flagged here (dense, "
                            "off-feed fish awaiting harvest is normal). 6N's "
                            "system biomass cap is still "
                            "checked on the SystemLimitsAudit sheet after a "
                            "run.")
                    else:
                        # Week picker: 'All weeks' collapses each distinct breach to
                        # its worst week; a specific week shows everything wrong then.
                        _opts = ["All weeks (worst first)", *_breach_weeks]
                        if st.session_state.get("mw_rec_week") not in _opts:
                            st.session_state["mw_rec_week"] = _opts[0]
                        _sel = st.selectbox(
                            "Show recommendations for week", _opts, key="mw_rec_week",
                            help="'All weeks' lists each distinct breach at its worst "
                                 "week; pick a week to see everything out of bounds then.")
                        _recs = (_collapsed if _sel.startswith("All weeks")
                                 else _by_week.get(_sel, []))
                        st.caption("Ranked by how far over cap. **▶** jumps to the "
                                   "tank so you can act on it.")
                        for _i, _rc in enumerate(_recs[:6]):
                            _sev = "🔴" if _rc["frac"] >= 1.15 else "🟠"
                            _ca, _cb = st.columns([9, 1])
                            _ca.markdown(
                                f"{_sev} **{_rc['week']}** · {_rc['msg']}  \n"
                                f"↳ {_rc['action']}")
                            if (_rc["tank_id"] is not None
                                    and _rc["week"] in labels
                                    and _cb.button("▶", key=f"mw_rec_{_i}",
                                                   help="Select this tank")):
                                st.session_state["mw_sel"] = (
                                    _rc["tank_id"], _rc["week"],
                                    labels.index(_rc["week"]) + 1)
                                st.rerun()
                        if len(_recs) > 6:
                            st.caption(f"… and {len(_recs) - 6} more breach(es).")

                sel = st.session_state.get("mw_sel")
                if sel and sel[1] in labels:
                    with st.container(border=True):
                        _mw_action_panel(state, ctx, rows, labels, sel, date_for)
                else:
                    st.info("👆 Click a tank's cell in the grid to harvest it, move / "
                            "split it, or send it to 6N — the options appear here.")

            _roll_when = "open" if _view == "open" else "close"
            if st.toggle(f"📊 System rollup — {_roll_when} biomass + feed/day "
                         f"per week", key="mw_rollup_toggle"):
                _mw_system_rollup(state, rows, labels, ctx["tables"],
                                  ctx["batch_by_id"], ctx=ctx, view=_view)

            if st.toggle("🐟 FW→OG intake — bring a freshwater cohort into OG",
                         key="mw_fw_toggle"):
                with st.container(border=True):
                    _mw_fw_intake(state, ctx, rows, labels, date_for)

            if st.toggle("🤖 Co-pilot — let the forecast propose the next week",
                         key="mw_copilot_toggle"):
                with st.container(border=True):
                    _mw_copilot(uploaded, events,
                                ctx.get("forecast_start") if ctx else None, bad)

            st.divider()
            _mw_timeline(state, events, bad, ctx.get("forecast_start") if ctx else None)

            if st.toggle("⚙ Advanced — raw event grid (power users)",
                         key="mw_adv_toggle"):
                with st.container(border=True):
                    _mw_raw_grid(state)
        else:
            _mw_raw_grid(None)

        st.divider()
        _mw_save_bar(events, bad)


def _og_systems_app():
    """OG system ids from facility config, or the standard 12 default."""
    try:
        from forecast.config_io import load_facility_config
        s = sorted({t.system_id for t in load_facility_config(CONFIG_DIR).tanks
                    if t.type == "OG" and t.system_id})
        if s:
            return s
    except Exception as e:  # noqa: BLE001
        st.caption(f"⚠ facility config unreadable ({type(e).__name__}) — "
                   f"showing the standard 12 OG systems.")
    return ["OG1N", "OG1S", "OG2N", "OG2S", "OG3N", "OG3S",
            "OG4N", "OG4S", "OG5N", "OG5S", "OG6N", "OG6S"]


def _limit_week_cols(fl_cur, sl_cur):
    """Week columns for the limits grid: the forecast horizon (PR start +
    Control horizon) when a PR is uploaded, else the weeks already present."""
    pr = _ingest_pr(uploaded) if uploaded is not None else None
    if pr and pr["ok"] and _config_ready():
        try:
            from forecast.config_io import load_control
            from forecast.time_grid import forecast_week_labels
            return forecast_week_labels(pr["forecast_start"],
                                        int(load_control(CONFIG_DIR).horizon_weeks))
        except Exception:  # noqa: BLE001
            pass
    return sorted({k[0] for k in fl_cur} | {k[0] for k in sl_cur})


def _sixn_prod_start_str():
    """The live `sixn_production_start`, formatted, or a plain-language note.

    Read at RENDER time. The mode split shown in this editor is decided by
    this date, so the tooltip must quote the value the config actually
    holds rather than a number baked into prose that rots on the next edit
    (app.py help-text contract, `_ctl_help`).
    """
    try:
        from forecast.config_io import load_control
        c = load_control(CONFIG_DIR)
        if getattr(c, "sixn_growth", False):
            return ("not applicable — 'Run 6N as grow-out' is ON, so every "
                    "week is production mode")
        d = getattr(c, "sixn_production_start", None)
        if d is None:
            return "not set — every week counts as purge mode"
        return (d.date().isoformat() if hasattr(d, "date") else str(d))
    except Exception:  # noqa: BLE001
        return "unreadable (check config/control.yaml)"


def _edit_limits():
    from forecast.config_io import load_control
    from forecast.scenario_io import (
        load_batches, load_limits, facility_limits_to_list,
        system_limits_to_list, facility_limits_from_list, dump_scenario,
    )
    from forecast.caps import (METRIC_BIOMASS, METRIC_FEED_DAY, METRIC_MAX_HARVEST,
                               METRIC_MIN_HARVEST, METRIC_HOG_YIELD, METRIC_SGR_OG,
                               SYSTEM_MODES, SystemLimits)
    st.caption(
        "Capacity limits. A capacity is a fact about the facility, so it is "
        "stated **once** per system below and applies to every week of every "
        "horizon. A single unusual week is an *exception* — see the expander "
        "at the bottom. Resolution order, highest first: per-week exception → "
        "system + mode default → system default → no cap at all."
    )
    # Guarded: this is the tab the operator opens to REPAIR limits.yaml, so it
    # must survive BOTH inputs being unreadable — limits.yaml itself, and the
    # control.yaml whose 6N production-start date binds the mode-specific caps.
    _ok, _limits = _read_or_explain(
        lambda: load_limits(SCENARIO_DIR, load_control(CONFIG_DIR)),
        "scenario/limits.yaml (bound to config/control.yaml)")
    if not _ok:
        return
    fl, sl = _limits
    fl_cur = {(r["week"], r["metric"]): r["value"] for r in facility_limits_to_list(fl)}
    sl_cur = {(r["week"], r["system"], r["metric"]): r["value"]
              for r in system_limits_to_list(sl)}
    fl_metrics = [METRIC_BIOMASS, METRIC_FEED_DAY, METRIC_MAX_HARVEST,
                  METRIC_MIN_HARVEST, METRIC_HOG_YIELD, METRIC_SGR_OG]
    sl_metrics = [METRIC_BIOMASS, METRIC_FEED_DAY]
    systems = _og_systems_app()
    # Weeks drive the EXCEPTION grid only; the defaults editor needs none, so
    # an empty horizon is no longer a dead end (it used to return early and
    # leave the operator with nothing editable at all).
    weeks = _limit_week_cols(fl_cur, sl_cur)

    if "sysdef_grid" not in st.session_state:
        st.session_state["sysdef_grid"] = pd.DataFrame(
            _system_defaults_records(sl.defaults, systems, sl_metrics)
        ).astype({m: "float64" for m in sl_metrics})
        st.session_state["modedef_grid"] = pd.DataFrame(
            _mode_default_records(sl.mode_defaults)
            or [{"system": "", "mode": "", "metric": "", "value": None}])
        st.session_state["flim_wide"] = pd.DataFrame(
            [{"metric": m, **{wk: fl_cur.get((wk, m)) for wk in weeks}}
             for m in fl_metrics]).astype({wk: "float64" for wk in weeks})
        st.session_state["slim_wide"] = pd.DataFrame(
            [{"system": s, "metric": m,
              **{wk: sl_cur.get((wk, s, m)) for wk in weeks}}
             for s in systems for m in sl_metrics]
        ).astype({wk: "float64" for wk in weeks})
        st.session_state["_lim_weeks"] = weeks
    weeks = st.session_state["_lim_weeks"]

    # ---------------- System capacities (the common case) ----------------
    # A blank cell here does NOT always mean "no cap": a mode-specific row
    # below can supply one, and today exactly that is true of OG6N's biomass.
    # Name the affected cells at RENDER time rather than asserting a config
    # value in prose (app.py help-text contract, `_ctl_help`).
    _mode_only = _mode_only_cells(sl.defaults, sl.mode_defaults)
    st.markdown("**System capacities**")
    st.caption(
        "One row per system — change a capacity in one cell. This is the "
        "system's STANDING capacity, applied in every week. A blank cell means "
        "no standing capacity is set, which leaves that system with **no cap "
        "at all** for that metric — unless a mode-specific row below supplies "
        "one."
        + (f" Blank here but covered by mode: **{', '.join(_mode_only)}**."
           if _mode_only else ""))
    sysdef_cfg = {
        "system": st.column_config.Column(
            pinned=True, disabled=True,
            help="The grow-out system (water loop). Its tanks share these "
                 "caps together — a tank-level density cap is separate "
                 "(Facility config)."),
        METRIC_BIOMASS: st.column_config.NumberColumn(
            "biomass (kg)", format="%.0f",
            help="Most standing fish weight this system may hold, in "
                 "kilograms, in any week. Raising it lets the planner "
                 "concentrate more fish here before it must move or harvest "
                 "them; lowering it pushes load onto the other systems and, "
                 "once every system is tight, forces earlier harvest. "
                 "Unit: kg."),
        METRIC_FEED_DAY: st.column_config.NumberColumn(
            "feed/day (kg/day)", format="%.0f",
            help="Most feed this system may deliver per day, in kilograms "
                 "per day — the physical limit of its feed line. Binds "
                 "before biomass does on the grow-out systems. Unit: kg/day."),
    }
    sysdef_df = st.data_editor(st.session_state["sysdef_grid"], hide_index=True,
                               column_config=sysdef_cfg, key="sysdef_grid_w",
                               use_container_width=True)

    # ---------------- Mode-specific capacities ----------------
    _psd = _sixn_prod_start_str()
    st.markdown("**Mode-specific capacities**")
    st.caption(
        f"For a system whose capacity depends on what it is being USED for. "
        f"6N is the one that does: it holds more while it is the depuration "
        f"station than it does once its 3 mains become grow-out. A row here "
        f"overrides that system's plain capacity above, for the weeks in that "
        f"mode. Which weeks are which is derived from the 6N production start "
        f"date in Control — currently **{_psd}** — so this split can never "
        f"drift away from that date."
    )
    modedef_cfg = {
        "system": st.column_config.SelectboxColumn(
            options=systems, required=False,
            help="The system this mode-specific capacity belongs to."),
        "mode": st.column_config.SelectboxColumn(
            options=list(SYSTEM_MODES), required=False,
            help="purge = the system is running depuration (pre-transition). "
                 "production = it is ordinary grow-out (from the 6N "
                 "production start date). Weeks are assigned to a mode by "
                 "that date, not by hand."),
        "metric": st.column_config.SelectboxColumn(
            options=sl_metrics, required=False,
            help="biomass = kg of standing fish; feed_per_day = kg of feed "
                 "per day."),
        "value": st.column_config.NumberColumn(
            format="%.0f",
            help="The cap for this system while it is in this mode. Unit "
                 "follows the metric: kg for biomass, kg/day for "
                 "feed_per_day."),
    }
    modedef_df = st.data_editor(st.session_state["modedef_grid"],
                                hide_index=True, num_rows="dynamic",
                                column_config=modedef_cfg, key="modedef_grid_w",
                                use_container_width=True)

    # ---------------- Facility limits ----------------
    # Metric names must be the TOKENS the `metric` column actually shows
    # (forecast.caps.METRIC_*), or the tooltip explains labels that appear
    # nowhere on the grid.
    _metric_help = (
        "Which cap this row sets, for the week in each column: "
        "`biomass` = most standing fish weight (kg); "
        "`feed_per_day` = most feed per day (kg/day); "
        "`max_harvest_per_week` / `min_harvest_per_week` = that week's harvest "
        "ceiling / floor (fish); "
        "`hog_yield` = live-to-sold weight ratio for that week; "
        "`sgr_correction_og` = growth factor for the OG (seawater) tanks that "
        "week — 1.0 (or blank) is the modelled growth, 0.90 means you expect "
        "only 90% of it. It multiplies ON TOP of the growth curve and each "
        "batch's own sgr_correction, and feed follows it (feed = biomass x "
        "SGR x FCR), so a 90% week eats 90% and grows 90%. Freshwater is "
        "unaffected. "
        "The per-week system grid carries the first two only — a harvest or "
        "yield cap is a whole-facility number.")
    wk_cfg = {wk: st.column_config.NumberColumn(
        width="small",
        help=f"The cap value for ISO week {wk} (unit = whatever the row's "
             f"metric uses). Blank facility cell = fall back to the Control "
             f"default; blank system cell = use the system capacity above.")
        for wk in weeks}
    with st.expander(f"Facility limits — per-week overrides "
                     f"({len(fl_cur)} set)", expanded=False):
        st.caption("Whole-facility caps for one week. Blank = use the Control "
                   "default for that metric.")
        if weeks:
            fdf = st.data_editor(
                st.session_state["flim_wide"], hide_index=True,
                key="flim_wide_w",
                column_config={"metric": st.column_config.Column(
                    pinned=True, disabled=True, help=_metric_help), **wk_cfg})
        else:
            fdf = st.session_state["flim_wide"]
            st.info("No weeks to show — upload a ProductionReport to set the "
                    "horizon.")

    # ---------------- Per-week system exceptions ----------------
    with st.expander(f"Per-week system exceptions — advanced "
                     f"({len(sl_cur)} set)", expanded=False):
        st.caption(
            "Only for a genuinely unusual week — a shutdown, a trial, "
            "maintenance. A value here overrides that system's capacity above "
            "for that ONE week. Leave blank (the normal case) and the system "
            "capacity applies. Weeks run across the top; the label columns "
            "stay frozen as you scroll.")
        if weeks:
            sdf = st.data_editor(
                st.session_state["slim_wide"], hide_index=True, height=400,
                key="slim_wide_w",
                column_config={
                    "system": st.column_config.Column(
                        pinned=True, disabled=True,
                        help="The grow-out system this exception applies to."),
                    "metric": st.column_config.Column(
                        pinned=True, disabled=True, help=_metric_help),
                    **wk_cfg})
        else:
            sdf = st.session_state["slim_wide"]
            st.info("No weeks to show — upload a ProductionReport to set the "
                    "horizon.")
        _hidden = ({k[0] for k in fl_cur} | {k[0] for k in sl_cur}) - set(weeks)
        if _hidden:
            st.caption(
                f"ℹ️ {len(_hidden)} week(s) in `limits.yaml` fall outside the "
                f"current forecast horizon and are not shown here "
                f"({min(_hidden)} … {max(_hidden)}). They are **kept** on "
                f"save, not deleted — edit them by loading a PR whose horizon "
                f"covers them.")

    b1, b2, _ = st.columns([1, 1, 3])
    if b1.button("💾 Save Limits", key="save_lim"):
        try:
            # Save REPLACES limits.yaml wholesale, but the exception grid only
            # shows the current forecast horizon (_limit_week_cols). Any week
            # stored in the file OUTSIDE that horizon has no column here, so
            # rebuilding purely from the grid would silently DELETE it — e.g.
            # every earlier week after uploading a PR that starts later. Carry
            # those through untouched; the operator never saw them and cannot
            # have edited them. (Defaults have no week axis, so they are not
            # exposed to this hazard at all — which is the point of them.)
            fl_recs = _preserved_facility_limits(fl_cur, weeks)
            sl_recs = _preserved_system_limits(sl_cur, weeks)
            fl_recs += [{"week": wk, "metric": r["metric"], "value": float(r[wk])}
                        for r in _records(fdf) for wk in weeks
                        if r.get(wk) not in (None, "")]
            sl_recs += [{"week": wk, "system": r["system"], "metric": r["metric"],
                         "value": float(r[wk])}
                        for r in _records(sdf) for wk in weeks
                        if r.get(wk) not in (None, "")]
            fl_recs.sort(key=lambda r: (r["week"], r["metric"]))
            sl_recs.sort(key=lambda r: (r["week"], r["system"], r["metric"]))
            new_sl = SystemLimits(
                caps={(r["week"], r["system"], r["metric"]): float(r["value"])
                      for r in sl_recs},
                defaults=_system_defaults_from_records(_records(sysdef_df),
                                                       sl_metrics),
                mode_defaults=_mode_defaults_from_records(_records(modedef_df)),
            )
            dump_scenario(SCENARIO_DIR, batches=load_batches(SCENARIO_DIR),
                          facility_limits=facility_limits_from_list(fl_recs),
                          system_limits=new_sl)
            _reset_keys("flim_wide", "slim_wide", "sysdef_grid", "modedef_grid")
            st.session_state.pop("_lim_weeks", None)
            st.success("Saved scenario/limits.yaml")
            st.rerun()
        except Exception as e:  # noqa: BLE001
            st.error(f"Save failed: {e}")
    if b2.button("↻ Reload", key="reload_lim"):
        _reset_keys("flim_wide", "slim_wide", "sysdef_grid", "modedef_grid")
        st.session_state.pop("_lim_weeks", None)
        st.rerun()


def _config_template_bytes(horizon_weeks=None, forecast_start=None) -> bytes:
    from forecast.config_template import write_config_template
    wd = Path(tempfile.mkdtemp(prefix="as_tmpl_"))
    out = wd / "config_template.xlsx"
    write_config_template(
        out,
        config_dir=CONFIG_DIR if _config_ready() else None,
        scenario_dir=SCENARIO_DIR if _scenario_ready() else None,
        horizon_weeks=horizon_weeks, forecast_start=forecast_start,
    )
    return out.read_bytes()


def _current_horizon_start():
    """Default horizon + start for the template, from current control if any."""
    from datetime import date
    h, s = 52, date.today()
    if _config_ready():
        try:
            from forecast.config_io import load_control
            c = load_control(CONFIG_DIR)
            h = int(c.horizon_weeks)
            fs = c.forecast_start
            s = fs.date() if hasattr(fs, "date") else (fs or s)
        except Exception as e:  # noqa: BLE001
            st.caption(f"⚠ Control unreadable ({type(e).__name__}) — template "
                       f"defaults to {h} weeks from today.")
    return h, s


# Analysis-layer files are SCORING overlays, not engine inputs — editing a
# harvest target or a price band changes no run's output, so they must not
# invalidate cached board legs / sweep results (which would force hours of
# re-runs to change a number the checklist re-judges instantly).
_NON_ENGINE_CONFIG = {"targets.yaml", "economics.yaml", "analysis_defaults.yaml"}


def _config_fingerprint() -> str:
    """Hash of config/ + scenario/ ENGINE-INPUT file CONTENT — changes
    whenever any config that affects a run is saved, so cached templates,
    board legs and sweep results invalidate. Content, not name+mtime: the
    2026-08-10 stale-board incident was a scenario edit (a W33 manual
    harvest) the mtime proxy did not register, so every disk-cached stock
    engine leg replayed the pre-edit scenario and poisoned the stock-vs-tuned
    comparison. The scan stays RECURSIVE (per-PR manual-event files live in
    scenario/manual_events/<closing>.yaml and are engine inputs);
    analysis-overlay files are excluded (see _NON_ENGINE_CONFIG)."""
    from forecast import analysis as _ana
    return _ana.dirs_fingerprint((CONFIG_DIR, SCENARIO_DIR),
                                 exclude=_NON_ENGINE_CONFIG)


def _engine_fingerprint() -> str:
    """Hash of the ENGINE SOURCE that produces a run (forecast/ + tools/).

    Cached engine results are identified by inputs AND code. Without this, the
    2026-08-12 stale-board incident: the Global engines were rebuilt while
    config/ and scenario/ stayed byte-identical, so _config_fingerprint never
    moved and four of five board legs replayed Tuesday's pre-repair plans as
    though they were today's. Only .py content is hashed (never __pycache__)."""
    from forecast import analysis as _ana
    return _ana.code_fingerprint((_ROOT / "forecast", _ROOT / "tools"))


def _sweep_inputs_sig() -> str:
    """Identity of the inputs a sweep ran against (PR content + config/scenario
    state) — stored beside Tune/Optimize/Frontier results so a recommendation
    computed on different inputs is flagged instead of presented as current.
    Folds in the metric-semantics version too: identical inputs measured under
    changed counter rules (e.g. the manual-window exclusion) are DIFFERENT
    measurements, so cached variants graded under old rules age out instead of
    silently defeating a metrics fix."""
    import hashlib
    from forecast import optimize as _opt_sig
    return hashlib.md5(
        f"{st.session_state.get('_pr_key', '')}|{_config_fingerprint()}"
        f"|{_engine_fingerprint()}|{_opt_sig.METRICS_SCHEMA}"
        .encode()).hexdigest()


class _WriteThroughCache(dict):
    """dict that persists itself to the analysis result cache on every write —
    so each finished knob-search variant survives a mid-search crash (the
    2026-08-06 pickling incident cost a whole 15-min phase; with this, a
    re-search reuses every variant that had already finished)."""

    def __init__(self, name: str, data: dict):
        super().__init__(data)
        self._name = name

    def __setitem__(self, key, value):
        super().__setitem__(key, value)
        try:
            from forecast import analysis as _ana
            _ana.cache_save(self._name, dict(self))
        except Exception:  # noqa: BLE001 — durability is best-effort
            pass


def _variant_cache(method_key: str = "") -> "_WriteThroughCache":
    """The knob-search variant cache for the CURRENT inputs (PR + config) —
    keyed by the sweep input signature, so a config/PR change simply starts
    an empty cache while the old one ages out of the store.

    `method_key` further keys the cache per METHOD for the tuned tournament:
    the same knob dict means a DIFFERENT run under a different method's pins
    only when the pins differ (they always do today), but per-method names
    also keep each method's search reuse legible and independently evictable.
    Empty = the live-config search (Optimize / quick Analyze), unchanged."""
    from forecast import analysis as _ana
    name = f"optvar_{_sweep_inputs_sig()[:20]}"
    if method_key:
        name += f"_{method_key}"
    data = _ana.cache_load_all(prefix=name).get(name) or {}
    return _WriteThroughCache(name, data)


def _warn_if_sweep_stale(sig_key: str, what: str) -> None:
    """Compare a stored sweep's input signature to the live inputs and warn —
    the Compare board has this staleness check; the sweeps were missing it, so
    a knob set validated on another PR/config could be saved as if current."""
    stored = st.session_state.get(sig_key)
    if stored is not None and stored != _sweep_inputs_sig():
        st.warning(f"⚠ These {what} were computed on a **different PR or "
                   f"config** than what's loaded now — re-run before trusting "
                   f"or saving anything from them.")


def _config_io_section():
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Template — build config offline**")
        pr = _ingest_pr(uploaded) if uploaded is not None else None
        if not (pr and pr["ok"]):
            st.info("⬆ Upload a valid **ProductionReport** (sidebar) to enable "
                    "the template — its week labels come from the PR's closing "
                    "date. The forecast length comes from Control → horizon.")
        else:
            h = _current_horizon_start()[0]
            fs = pr["forecast_start"]
            # Invalidate a cached template if config changed since it was built,
            # so Download always reflects the latest saved config.
            _fp = _config_fingerprint()
            if st.session_state.get("_tmpl_fp") != _fp:
                st.session_state.pop("_tmpl_bytes", None)
            st.caption(f"Template covers **{h} weeks from {fs.date()}** — start "
                       f"derived from the PR, horizon from Control (edit on the "
                       f"Control tab). Every per-week limit slot is laid out to "
                       f"fill in; current values pre-filled.")
            if st.button("Build template", use_container_width=True):
                st.session_state["_tmpl_bytes"] = _config_template_bytes(int(h), fs)
                st.session_state["_tmpl_fp"] = _fp
            if "_tmpl_bytes" in st.session_state:
                st.download_button(
                    "⬇ Download config template (.xlsx)",
                    data=st.session_state["_tmpl_bytes"],
                    file_name="config_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.caption("↑ Click **Build template** to (re)generate from the "
                           "current saved config.")
    with c2:
        st.markdown("**Import — load config from a file**")
        st.caption("A filled config template, or a saved forecast workbook "
                   "(RunConfig snapshot). Overwrites current config/ + scenario/. "
                   "The analysis overlays are deliberately NOT carried in a "
                   "workbook and are never overwritten by an import: "
                   "`analysis_defaults.yaml` (your promoted Quick-run default), "
                   "`targets.yaml` and `economics.yaml` (the scoring yardstick). "
                   "They steer how a run is judged, not what it computes — a "
                   "workbook must not be able to move them.")
        imp = st.file_uploader("Config template or saved workbook",
                               type=["xlsx", "xlsm"], key="cfg_import")
        if st.button("📥 Import config", disabled=imp is None, key="do_import",
                     use_container_width=True):
            try:
                from forecast.config_template import (
                    import_config_template, is_config_template,
                )
                from forecast.config_snapshot import (
                    import_config_snapshot, read_config_snapshot,
                    describe_run_config_sheet,
                )
                wd = Path(tempfile.mkdtemp(prefix="as_import_"))
                p = wd / imp.name
                p.write_bytes(imp.getvalue())
                wb = load_workbook(p, keep_vba=(p.suffix.lower() == ".xlsm"))
                why = ""
                if is_config_template(wb):
                    restored = import_config_template(wb, CONFIG_DIR, SCENARIO_DIR)
                    src = "config template"
                elif read_config_snapshot(wb):
                    restored = import_config_snapshot(wb, CONFIG_DIR, SCENARIO_DIR)
                    src = "RunConfig snapshot"
                else:
                    # Say WHICH RunConfig sheet this is. A Global method-stamp
                    # workbook has one — reporting a flat "not found" for it
                    # sent the operator hunting a file-format problem that
                    # never existed.
                    restored, src = [], None
                    why = describe_run_config_sheet(wb)
                wb.close()
                if not restored:
                    st.error("Nothing to import from that file. "
                             + (why or "It is neither a config template nor a "
                                       "RunConfig snapshot."))
                else:
                    _clear_all_editor_state()  # refresh open editors from disk
                    st.success(f"Imported {len(restored)} file(s) from {src}: "
                               f"{', '.join(restored)}")
                    st.rerun()
            except Exception as e:  # noqa: BLE001
                st.error(f"Import failed: {e}")


def _edit_targets_prices():
    """Harvest targets + price bands — the ANALYSIS overlay. These score and
    judge plans (Analyze mode's checklist + revenue) but change no engine
    output, so saving here never invalidates cached runs."""
    from forecast import analysis as _ana

    st.markdown("**Harvest targets** — monthly / yearly harvest the plan "
                "should deliver. Judged with a tolerance and **penalized, "
                "never disqualifying**: Analyze flags shortfalls and prefers "
                "plans that meet them, but a miss doesn't hide a plan.")
    _ok, t = _read_or_explain(lambda: _ana.load_targets(CONFIG_DIR),
                              "config/targets.yaml")
    if not _ok:
        return
    t = t or {"basis": "hog", "tolerance_pct": 5.0, "monthly": {}, "yearly": {}}
    c1, c2 = st.columns(2)
    basis = c1.radio("Target basis", ["hog", "gross"], horizontal=True,
                     index=0 if t["basis"] == "hog" else 1, key="tgt_basis",
                     help="Which weight your targets are written in: hog = "
                          "head-off gutted (sold) kg; gross = live kg out of "
                          "the water. Pick the one your sales plan uses.")
    tol = c2.number_input("Tolerance (%)", min_value=0.0, max_value=50.0,
                          value=float(t["tolerance_pct"]), step=1.0,
                          key="tgt_tol",
                          help="Grace margin when judging a target: landing "
                               "within this % under it counts as CLOSE (a "
                               "soft warning) instead of MISSED. Unit: %. "
                               "Default 5.")
    mdf = st.data_editor(
        pd.DataFrame([{"Month": k, "Target_kg": v}
                      for k, v in sorted(t["monthly"].items())]
                     or [{"Month": "", "Target_kg": None}]),
        num_rows="dynamic", hide_index=True, use_container_width=True,
        key="tgt_monthly",
        column_config={
            "Month": st.column_config.TextColumn(
                "Month (YYYY-MM)",
                help="Calendar month this target applies to, e.g. 2026-11."),
            "Target_kg": st.column_config.NumberColumn(
                "Target (kg)", min_value=0.0, step=1000.0,
                help="Harvest the plan should deliver that month, in the "
                     "chosen basis (hog/gross). Unit: kg."),
        })
    ydf = st.data_editor(
        pd.DataFrame([{"Year": k, "Target_kg": v}
                      for k, v in sorted(t["yearly"].items())]
                     or [{"Year": "", "Target_kg": None}]),
        num_rows="dynamic", hide_index=True, use_container_width=True,
        key="tgt_yearly",
        column_config={
            "Year": st.column_config.TextColumn(
                "Year (YYYY)",
                help="Calendar year this target applies to, e.g. 2027."),
            "Target_kg": st.column_config.NumberColumn(
                "Target (kg)", min_value=0.0, step=10000.0,
                help="Harvest the plan should deliver that year, in the "
                     "chosen basis (hog/gross). Unit: kg."),
        })

    st.divider()
    st.markdown("**Price per fish size** — turns harvest into revenue on the "
                "Analyze board. Each harvest event is priced by its average "
                "fish weight; harvest falling in **no band is reported as "
                "unpriced** (a loud gap, never an invented price).")
    _ok, e = _read_or_explain(lambda: _ana.load_economics(CONFIG_DIR),
                              "config/economics.yaml")
    if not _ok:
        return
    e = e or {"currency": "USD", "basis": "hog", "model_cv_pct": 18.0,
              "price_bands": []}
    c3, c4, c5 = st.columns(3)
    cur = c3.text_input("Currency", value=e["currency"], key="eco_cur",
                        help="Currency label shown on revenue figures (e.g. "
                             "USD). Display only — no conversion happens.")
    ebasis = c4.radio("Price basis", ["hog", "gross"], horizontal=True,
                      index=0 if e["basis"] == "hog" else 1, key="eco_basis",
                      help="Which weight the price bands and revenue are "
                           "written in: hog = head-off gutted (sold) kg; "
                           "gross = live kg. Match your sales price list.")
    mcv = c5.number_input(
        "Sales model CV (%)", min_value=0.0, max_value=60.0,
        value=float(e.get("model_cv_pct", 18.0)), step=1.0, key="eco_cv",
        help="Harvest weight-distribution CV: each event's kg is spread "
             "across the bands with a size-biased lognormal around its "
             "average weight (your Excel LOGNORM method). Re-tune against "
             "historical harvest results. Per-month price overrides live "
             "in config/economics.yaml under each band's `monthly:` key.")
    bdf = st.data_editor(
        pd.DataFrame(e["price_bands"]
                     or [{"min_kg": None, "max_kg": None, "price_per_kg": None}]),
        num_rows="dynamic", hide_index=True, use_container_width=True,
        key="eco_bands",
        column_config={
            "min_kg": st.column_config.NumberColumn(
                "Min fish wt (kg, incl.)", min_value=0.0, step=0.25,
                help="Smallest fish weight this price band covers (included). "
                     "In the chosen basis (hog/gross). Unit: kg per fish."),
            "max_kg": st.column_config.NumberColumn(
                "Max fish wt (kg, excl.)", min_value=0.0, step=0.25,
                help="Fish weight where this band ends (excluded — the next "
                     "band takes over exactly here, so bands never overlap). "
                     "Unit: kg per fish."),
            "price_per_kg": st.column_config.NumberColumn(
                "Price / kg", min_value=0.0, step=0.1,
                help="Default sales price for fish in this size band, per kg "
                     "in the chosen basis. Per-month price overrides live in "
                     "config/economics.yaml under each band's 'monthly:' key."),
        })

    if st.button("💾 Save targets & prices", key="tgt_save", type="primary"):
        import re as _re
        monthly, yearly, errs = {}, {}, []
        for rec in _records(mdf):
            mth, kg = str(rec.get("Month") or "").strip(), rec.get("Target_kg")
            if not mth and kg in (None, ""):
                continue
            if not _re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])", mth):
                errs.append(f"bad month '{mth}' (want YYYY-MM)")
            elif kg in (None, ""):
                errs.append(f"month {mth} has no target")
            else:
                monthly[mth] = float(kg)
        for rec in _records(ydf):
            yr, kg = str(rec.get("Year") or "").strip(), rec.get("Target_kg")
            if not yr and kg in (None, ""):
                continue
            if not _re.fullmatch(r"20\d{2}", yr):
                errs.append(f"bad year '{yr}' (want YYYY)")
            elif kg in (None, ""):
                errs.append(f"year {yr} has no target")
            else:
                yearly[yr] = float(kg)
        bands = []
        # The grid edits only default prices — carry each band's per-month
        # overrides through the save (matched by weight range), or a UI save
        # would silently strip the monthly price ladder from the YAML.
        _old_monthly = {(b["min_kg"], b["max_kg"]): b.get("monthly") or {}
                        for b in (e.get("price_bands") or [])}
        for rec in _records(bdf):
            lo, hi, p = (rec.get("min_kg"), rec.get("max_kg"),
                         rec.get("price_per_kg"))
            if lo in (None, "") and hi in (None, "") and p in (None, ""):
                continue
            if None in (lo, hi, p) or "" in (lo, hi, p):
                errs.append("a price band is half-filled")
            elif float(hi) <= float(lo):
                errs.append(f"band {lo}–{hi}: max must exceed min")
            else:
                bands.append({"min_kg": float(lo), "max_kg": float(hi),
                              "price_per_kg": float(p),
                              "monthly": _old_monthly.get(
                                  (float(lo), float(hi)), {})})
        if errs:
            for x in errs:
                st.error(x)
        else:
            _ana.save_targets(str(CONFIG_DIR), {
                "basis": basis, "tolerance_pct": float(tol),
                "monthly": monthly, "yearly": yearly})
            _ana.save_economics(str(CONFIG_DIR), {
                "currency": cur or "USD", "basis": ebasis,
                "model_cv_pct": float(mcv), "price_bands": bands})
            st.success(f"Saved — {len(monthly)} monthly + {len(yearly)} yearly "
                       f"target(s), {len(bands)} price band(s). Analyze "
                       f"re-judges instantly; no runs are invalidated.")


def _config_editor():
    st.header("⚙️ Configure — models & control")
    st.caption("Build the forecast config here — saved to `config/` + `scenario/` "
               "YAML (the source of truth). Or populate it offline via the "
               "template/import below, separately from the PR.")
    ready = _config_ready() and _scenario_ready()
    with st.expander("📦 Template & import", expanded=not ready):
        _config_io_section()
    if not ready:
        st.warning("No config yet — download the template above, fill it in, and "
                   "import; or seed from a workbook in Run mode.")
        return

    tabs = st.tabs(["Control", "Biology models", "Facility (tanks)",
                    "Batches", "Limits", "Targets & prices"])
    with tabs[0]:
        _edit_control()
    with tabs[1]:
        _edit_biology()
    with tabs[2]:
        _edit_facility()
    with tabs[3]:
        _edit_batches()
    with tabs[4]:
        _edit_limits()
    with tabs[5]:
        _edit_targets_prices()


# ============================================================
# How it works — the plain-language rulebook (operator request)
# ============================================================

def _hv_error_model():
    """The measured forecast-error model, or None when there isn't one.

    Absence is the normal state of a fresh clone (no backtest has been run), and
    the caller must then show plan numbers with NO band rather than inventing
    one. Cached per session: it is a small static JSON.
    """
    key = "_hv_error_model_cache"
    if key not in st.session_state:
        from forecast.error_bands import load_error_model
        st.session_state[key] = load_error_model(_ROOT)
    return st.session_state[key]


def _hv_month_bands(months, values, model):
    """Per-month (low_t, high_t), or None for months that must not be banded.

    Horizon is counted from the FIRST harvest month in the plan, which is the
    forecast's own start — the same clock the error model was measured on. A
    month outside the quotable window, or one the model rates too thin, returns
    None so the chart draws no whisker there at all.
    """
    from forecast.error_bands import apply_band, band_for_horizon, months_between
    import datetime as _d
    out = []
    if not months:
        return out
    def _first(m):
        return _d.date(int(str(m)[:4]), int(str(m)[5:7]), 1)
    start = _first(min(months))
    for m, v in zip(months, values):
        try:
            h = months_between(start, _first(m))
        except (ValueError, TypeError):
            out.append(None)
            continue
        out.append(apply_band(float(v), band_for_horizon(model, h)))
    return out


def _hv_band_caption(model, bands) -> str:
    """One honest sentence under the chart. Says what the whiskers mean, what
    they do NOT cover, and why some months have none."""
    from forecast.error_bands import QUOTABLE_MAX_MONTHS, describe
    if not model:
        return ("Bars are the planned tonnage. " + describe(None) +
                " Run `tools/backtest.py` then `tools/error_model.py` to "
                "measure one from your own Production Reports.")
    n_banded = sum(1 for b in bands if b)
    n_plain = len(bands) - n_banded
    msg = ("Whiskers show where past forecasts of THIS facility actually "
           "landed at the same distance ahead — not a modelling assumption. "
           + describe(model))
    if n_plain:
        msg += (f" {n_plain} later month(s) carry no whisker: beyond "
                f"{QUOTABLE_MAX_MONTHS} months the measurement is still "
                f"distorted by harvest timing, so a band there would look "
                f"more certain than it is.")
    msg += (" The band covers BIOLOGY (how fast fish grow). It does NOT cover "
            "execution: harvest COUNT is your plan, so harvesting a different "
            "number of fish moves tonnage for reasons no growth model can see.")
    return msg


def _hiw_knobs():
    """Live Control values for the How-it-works page, with dataclass fallbacks
    — so the page shows the numbers THIS install actually runs with and never
    goes stale when a knob is retuned. Never raises (page must render with no
    config seeded)."""
    vals = {"min_hv": 30_000.0, "max_hv": 55_000.0, "relief_pct": 0.10,
            "min_wt": 3_500.0, "bio_cap": 3_800_000.0, "feed_cap": 34_000.0,
            "moves": 15, "min_tank": 7_000.0, "band": 0.005,
            "prod_start": "2028-01-01", "n_entry_tanks": 2,
            # The harvest guide's follow band, and the 6N tanks' own density
            # cap — both quoted in the prose below, so both are read, never
            # typed (the same contract `_ctl_help` follows).
            "guide_band": 0.05, "sixn_density": 95.0}
    try:
        from forecast.config_io import load_control
        c = load_control(CONFIG_DIR)
        vals.update({
            "min_hv": float(c.min_harvest_per_week or vals["min_hv"]),
            "relief_pct": float(getattr(c, "harvest_relief_pct",
                                        vals["relief_pct"]) or 0.0),
            "max_hv": float(c.max_harvest_per_week or vals["max_hv"]),
            "min_wt": float(c.min_harvest_weight_g or vals["min_wt"]),
            "bio_cap": float(c.max_biomass_kg or vals["bio_cap"]),
            "feed_cap": float(c.max_feed_per_day_kg or vals["feed_cap"]),
            "moves": int(getattr(c, "max_transfers_per_week", 0)
                         or vals["moves"]),
            "min_tank": float(c.min_tank_control or vals["min_tank"]),
            "band": float(c.facility_biomass_deviation_pct or vals["band"]),
            "prod_start": (c.sixn_production_start.date().isoformat()
                           if getattr(c, "sixn_production_start", None)
                           else vals["prod_start"]),
            "n_entry_tanks": int(c.tran_og_default_tanks
                                 or vals["n_entry_tanks"]),
            "guide_band": float(getattr(c, "hybrid_follow_band", 0.0)
                                or vals["guide_band"]),
        })
    except Exception:  # noqa: BLE001 — static page must render regardless
        pass
    try:
        # The 6N fill cap is a per-TANK facility input, not a Control knob.
        from forecast.config_io import load_facility_config
        from forecast.sixn import SIXN_ALL_TANKS
        _d = [t.max_density_kg_m3 for t in load_facility_config(CONFIG_DIR).tanks
              if t.tank_id in SIXN_ALL_TANKS and t.max_density_kg_m3]
        if _d:
            vals["sixn_density"] = float(max(_d))
    except Exception:  # noqa: BLE001 — static page must render regardless
        pass
    # The never-exceed ceiling is DERIVED, never stored: limit × (1 + relief).
    vals["ceiling_hv"] = vals["max_hv"] * (1.0 + vals["relief_pct"])
    return vals


def _how_it_works():
    """The plain-language rulebook: every pipeline layer — what it decides,
    what it is forbidden to do, which checks bind it — plus the honest list of
    known limits. Static text (no runs); the numbers are read live from the
    current Control config so the page tracks retuning."""
    k = _hiw_knobs()
    st.header("📖 How it works — the rules, layer by layer")
    st.caption(
        "What the forecast actually does, in plain language. Each layer below "
        "says **what it decides**, **what it may never do**, and **which "
        "checks bind it** — plus an honest list of known limits at the end. "
        "The numbers shown are read from your current Control config, so this "
        "page stays true when a knob is retuned. Nothing here runs anything.")

    with st.expander("🧭 The big picture — one paragraph", expanded=True):
        st.markdown(f"""
You upload a **ProductionReport** (the facility's true state today). The tool
grows every cohort forward week by week — freshwater, seawater grow-out, then
the **6N depuration station** — and decides each week's **transfers** and
**harvests** so that: every week ships at least **{k['min_hv']:,.0f} fish**
(the contract floor — never an empty week), the facility never exceeds
**{k['bio_cap']:,.0f} kg** of standing fish or **{k['feed_cap']:,.0f} kg/day**
of feed, fish move only along the physically legal routes (the tier rules
R1-R8 below), and every fish is accounted for from stocking to harvest. Layers
that *decide* are separated from checks that *audit* — a plan is never trusted
because the planner says so, only because the independent audits reconcile it.""")

    with st.expander("1 · Inputs & the starting state"):
        st.markdown(f"""
**What it decides.** The uploaded ProductionReport is hydrated into the week-0
facility: which batch sits in which tank, at what count and weight. The
forecast start date is **derived** from the PR's closing date (+1 day) — the
config value is only a seed. Models and knobs come from Configure
(`config/` + `scenario/` YAML): biology curves, the facility's tanks, the
batch schedule, per-week cap overrides.

**What it may never do.** The PR is *state only* — it carries no instructions.
Nothing else is read from the workbook, and the source file is never written
back. A batch in the PR can never be silently ignored: every in-horizon batch
must reach the facility or the run fails its input-conservation gate.

**What binds it.** Hydration validation (unknown tanks/batches are refused
loudly), the derived-start contract, and the input-conservation audit
(layer 9).""")

    with st.expander("2 · The manual override window (optional operator prefix)"):
        st.markdown("""
**What it decides.** Nothing — *you* do. For weeks 1..N you script the exact
operations and the engine executes ONLY those, plus real biology (growth,
mortality, feed). A window week runs **no planner logic at all**: if you
script no harvest that week, that week harvests nothing. The planner takes
over at week N+1 (the handoff) from exactly the state your scripted weeks
produced.

**The five things you can script:**

| Type | What it does |
|---|---|
| **harvest** | Take fish out of a grow-out or 6N tank. No count = the whole tank. |
| **og_transfer** | Move / split fish between grow-out tanks. |
| **og_to_6n** | Send fish into a 6N tank to purge — they go off-feed immediately. |
| **graded_harvest** | Size-sort a tank: the biggest N go to a pickup tank, the rest stay growing. |
| **fw_to_og** | Bring a freshwater cohort into the entry tier, culling down to your target count. |

**Graded → 6N: staging or harvest?** When the pickup tank is a **6N** tank the
default is **staging** — the big fish sit off-feed and are harvested *later*,
after the 2-week hold (either you script that harvest, or the planner takes
them). Set the event's **mode** to `harvest` to drain them to processing in
the scripted week instead; the 6N tank is then only the route and ends the
week empty. (A pickup tank that is *not* a 6N tank always means harvest now,
and `mode: stage` there is refused.)

**Every scripted event says what happened.** There are no silent no-ops: each
event writes either a `MANUAL EVENT OK` line naming exactly what it did, or a
`MANUAL EVENT REFUSED` line naming the reason and confirming the fish stayed
put — both into the ValidationLog sheet of the output workbook. Read that
sheet after a windowed run.

**What it may never do.** The window may not invent operations you didn't
script — and neither may the planner *pretend* you did: no engine may assume
fish were staged into depuration during your window weeks (that bug made
engine comparisons dishonest and is fixed + pinned by tests). Scripted events
that break the tier rules are hard-blocked in the editor; events conserve
fish exactly (a refused move leaves the source untouched).

**What binds it.** Per-event validation against the hydrated PR, the tier
rules R1-R8, and the **dark-handoff lint**: if your window drains 6N without
restaging, the editor warns you *which* handoff weeks will have nothing
harvestable under the 2-week depuration hold — a warning, not a block,
because you may intend a dark week.""")

    with st.expander("3 · The freshwater phase (egg → smolt → transfer)"):
        st.markdown("""
**What it decides.** Each batch's growth from stocking to its seawater
transfer date, on the freshwater growth/mortality/cull tables, calibrated
per batch (auto-calibrated to land each batch on its planned entry weight
when 'Auto-calibrate FW' is on — a planning assumption, not a guarantee).

**What it may never do.** Freshwater fish are **never harvested** and never
placed by the planner — their trajectory is a given. But they are never
invisible either: FW biomass and feed **count against the facility caps**,
so the seawater side must make room for a known FW peak (it pre-positions
harvest up to 8 weeks ahead of one).

**What binds it.** The closed FW mass-balance audit (first FW count ==
seawater entry + FW mortality + FW culls, per batch) and the FW→SW
reconciliation signal (realized entry vs plan, flagged over 5% — your cue to
recalibrate a batch's correction).""")

    with st.expander("4 · Seawater entry — TranOG arrivals (rule R1)"):
        st.markdown(f"""
**What it decides.** When a batch's transfer date arrives, its fish enter the
**entry tier only** — systems OG1N/OG1S/OG2N/OG2S (rule **R1**) — spread
across at least **{k['n_entry_tanks']}** tanks, split big/small by the batch's
size spread. If no entry tank is free, the planner **makes room**: it frees a
tank by moving its fish forward (or into 6N to purge), never by dropping the
arrival. Room-making is *anticipatory*: empty tanks near an arrival (≤3 weeks
out) are **reserved** so a rebalancing pass can't consume them first, with a
6-week lookahead budgeting the rest.

**What it may never do.** Arrivals may not enter grow-out or 6N directly; an
arrival may never be silently dropped (a whole class of lost-fish bugs was
closed by this gate); and freed remnant tanks may not strand an arrival —
sub-minimum leftovers are folded forward first.

**What binds it.** The input-conservation gate (0 dropped batches), the tier
rules, and the remnant floor (layer 5).""")

    with st.expander("5 · Grow-out placement, the tier rules R1-R8, and the "
                     "handling budget"):
        st.markdown(f"""
**What it decides.** Week by week, which tank each group occupies: forward
moves along the conveyor (entry tier → OG3-6 → 6N → harvest), plus
quality passes that level crowding, feed and biomass across systems
(splitting hot tanks, evening out load).

**The movement rules — physical law for the planner:**

| Rule | Plain meaning |
|---|---|
| **R1** | New seawater fish enter only the entry tier (OG1/2). |
| **R2** | Entry-tier fish may move forward to any OG3-6 tank at any weight. |
| **R3** | *Within* the entry tier, moves are legal only while the source tank averages **< 1 kg** (equipment limit). At/over 1 kg: forward only. |
| **R4** | **Never backward** — a grow-out tank may never send fish to the entry tier. |
| **R5** | **No harvest and no 6N staging from entry-tier tanks** — fish route forward first. |
| **R6** | Fish ≥ 1 kg *may stay* in an entry tank (stuck-in-place is legal and measured-necessary — never force-evicted). |
| **R7** | **6N is one-way**: fish moved into depuration leave only by harvest, never by transfer. |
| **R8** | **Fish preparing for harvest carry no density cap** — judged on stage (`STARVE`), so it covers both 6N depuration and, after the 6N production switch, in-place starvation in an ordinary grow-out tank. They are off feed, not growing, and gone within the hold, so the feed-loading reason for the cap does not apply. It is also what makes harvest-prep consolidation legal: the whole group fits in one tank and the rest go back to the rotation. |

**Moving fish costs fish.** Every tank-to-tank deposit is charged
`handling_mortality_pct` (2026-08-21) — rebalances, consolidation, relief
moves, 6N move-ins, the TranOG entry. The source tank is drained by the full
amount, the destination keeps the rest, and the difference is booked as
mortality so the tank audit still balances to zero. The handling budget below
is therefore not only a crew-capacity limit: a plan that shuffles more fish
also kills more of them.

**What it may never do.** Exceed **{k['moves']} transfer moves per week** —
the handling budget. Deferrable quality passes simply stop and wait for a
calmer week; *essential* moves (6N rotation fills, arrival make-room) are
never blocked, but a week they alone push past the budget is **reported on
the handling gate**, never hidden. It may never mix two batches in one tank,
never leave a "remnant" — every partial draw takes everything or leaves at
least **{k['min_tank']:,.0f} fish** (a tank below that isn't worth
operating), and never plan a move the tier rules refuse.

**What binds it.** The tier rules above (enforced in the event layer — an
illegal move is refused, not patched), the per-tank continuity audit (every
tank-week must balance to zero drift), the per-system limits audit, and the
handling gate.""")

    with st.expander("6 · The harvest controller — floor, limit, relief"):
        st.markdown(f"""
**What it decides.** How many fish to harvest each week, aiming to ride just
under the facility's **effective ceiling** — the lower of the biomass cap
({k['bio_cap']:,.0f} kg) and the biomass at which feed hits its cap
({k['feed_cap']:,.0f} kg/day) — with one soft margin
(±{k['band'] * 100:.1f}%). Below the band it harvests only the floor and lets
the facility **fill up**; at the band it ramps harvest to hold the line.
Harvest is **demand-driven**: the week takes what holding the caps requires,
never a quota.

**The three harvest numbers:**
* **Floor {k['min_hv']:,.0f} fish/week** — the sales contract. Every week
  must ship at least this; a totally empty week is a hard business-rule
  breach, ranked above every other quality measure.
* **Limit {k['max_hv']:,.0f} fish/week** — the processing plant's weekly
  capacity. A **constraint, not a level**: no pass ever sizes harvest *up*
  to reach it — demand-driven harvest is simply capped here, and the
  long-horizon guide ramps harvests **earlier** rather than let demand pile
  into one week.
* **Relief ceiling {k['ceiling_hv']:,.0f} fish/week** — derived, never
  stored: limit × (1 + {k['relief_pct'] * 100:.0f}% relief). A **pressure
  valve for exceptional weeks only** — a whole 6N tank that must drain, a
  make-room dump that saves an arrival. Nothing may ever exceed it.

Relief is *allowed, but not routine*: the checklist gate reads **PASS** at 0
relief weeks, **WARN** at 1-3 ("pressure relief used — acceptable if
exceptional"), **FAIL** beyond 3 or on any week above the relief ceiling —
and the failure message says what to do about it: **ramp harvests up
earlier**. A plan that pins to relief every week isn't using a buffer, it's
hiding a restructuring problem.

When the **harvest guide** is on (Configure → Control, `hybrid_follow` — on by
default), a whole-horizon harvest envelope (the Global engine's long view) is
computed first and fed to the weekly controller as a **target band**
(±{k['guide_band'] * 100:.0f}%): it tells the controller to harvest *less* in
fat weeks so those fish still exist for lean ones — the one thing a week-by-week
planner cannot see. Switch the guide off and this paragraph does not apply:
that is the plain reactive controller, which leaves empty harvest weeks. The
smoother additionally spreads harvest early so weeks are flat, not
dump-then-nothing.

**What it may never do.** Harvest a fish under **{k['min_wt']:,.0f} g** (the
sales gate); harvest from entry-tier tanks (R5); harvest a production tank
directly while 6N is in depuration mode — *all* harvest flows through 6N
(layer 7); plan a week above the {k['max_hv']:,.0f} limit except as
exceptional relief; or breach the {k['ceiling_hv']:,.0f} relief ceiling,
ever.

**What binds it.** The steady-harvest gate (no near-empty week past the
startup handoff), the floor gate, the **relief gate** (PASS 0 / WARN 1-3 /
FAIL >3 relief weeks or any ceiling breach), the biomass/feed cap checks,
and — when trial corrections are evaluated — a **lexicographic rule**: a
candidate that adds an empty harvest week never wins, whatever else it
improves.""")

    with st.expander("7 · 6N depuration — the two-week hold and the rotation"):
        st.markdown(f"""
**What it decides.** Before harvest, fish sit **off-feed for 2 weeks** in the
6N station to purge (a product requirement, not a scheduling nicety). 6N runs
a **3-pair fallow rotation** (pairs 61/67, 63/69, 65/71, fixed order
61→63→65): two pairs purge while one rests; each week the front pair is
harvested and the resting pair refills from the oldest mature fish.

**The 6N-specific rules:**
* **One batch, one tank** — a purge tank has **no density cap at all**
  (operator rule, 2026-08-21): the fish are off feed, not growing, and gone
  within the hold, so what bounds the tank is the *harvest schedule*, not
  kg/m³. A whole batch therefore fills ONE tank however dense it gets.
  The sister (67/69/71) is **not overflow capacity** — it exists so a
  *second, different* batch needing harvest the same week is not mixed into
  an occupied tank, because mixing destroys per-batch count fidelity at
  harvest. Spending a sister on one batch's overflow burns the slot that
  separation needs.
* **The hold is real** — a tank may **not** be drained on the rotation right
  after its fill (that leak shipped fish with 1 week of purge instead of 2;
  fixed and audited). Fish hydrated from the PR already mid-purge are the one
  exemption — their residency clock predates the forecast.
* **R7 one-way** — once in depuration, fish leave only by harvest.
* **R8 no density cap on harvest-prep** — a purging tank is bounded by the
  harvest schedule, not by kg/m³.
* Make-room routes through 6N too: a tank freed for an arrival sends its
  fish to purge, staging them for harvest — never harvested in place.

After **{k['prod_start']}** the station switches to production mode: the
**main** tanks (61/63/65) become ordinary grow-out and every cap applies to
them again — including their configured
**{k['sixn_density']:.0f} kg/m³** density cap, which does NOT bind while they
are purging — while the three **sisters** (67/69/71) are not production capacity
at all. Harvest-bound fish instead go off feed *in place* in their own
grow-out tank for the starvation period — and while they do, that tank is
density-exempt under R8 exactly as a purge tank is, and a batch's in-place
purge tanks are consolidated into one to hand the rest back to the rotation.

**What it may never do.** Skip or shorten the hold, mix batches within one
tank, or transfer fish back out of depuration. (Stocking a purge tank "past
its density cap" is **not** on this list any more — there is no such cap
while fish are preparing for harvest.)

**What binds it.** The depuration-hold audit (every run reports what fraction
of harvest left 6N early — must be the PR-hydrated exemptions only), R7
enforcement in the event layer, and the pair rotation's own continuity in the
tank audit. Density does **not** bind here: rule **R8**
(`tiers.density_exempt`) exempts any tank whose fish are preparing for
harvest — both 6N depuration and, after the production switch, in-place
starvation in an ordinary grow-out tank.""")

    with st.expander("8 · The engines — two families, and how far you can "
                     "trust each"):
        st.markdown(f"""
**What it decides.** Which planning *method* produced your plan. Every method
reads the same inputs and every method's output goes through the same audits —
but they do **not** all enforce the same rules while planning. That difference
is the single most important thing on this page, so it is spelled out below
rather than smoothed over.

**The Controller family — runnable operating plans.**

* **Controller — hybrid (L1-guided harvest)** — *the default.* The validated
  week-by-week controller, guided by a whole-horizon harvest envelope. It is
  the only method measured to harvest something every single week.
* **Controller — reactive greedy** — the same controller with the guide
  switched off. Simpler to reason about; it leaves empty harvest weeks on 5 of
  6 real ProductionReports.
* **Controller — greedy + LNS** — adds an audited relocation pass. It only
  changes anything when there are free tanks to relocate into, so on a
  capacity-bound facility it usually matches the plain controller exactly.

All three enforce the tier rules R1-R8 *while planning* (an illegal move is
refused, not logged), respect the **{k['moves']}-move handling budget** by
deferring their optional quality passes, and route all harvest through 6N.

**The Global family — whole-horizon benchmarks, not drop-in operating plans.**

* **Global — lexicographic LP** — plans the whole horizon up front: harvest
  envelope → per-batch share per system → tank placement.
* **Global — CP-SAT optimal** — the same front end, but the grow-out tank
  layout is re-solved week by week with a constraint solver. Slowest by far;
  tightest, most evenly balanced layouts. Its advantage is **not foresight** —
  it plans one week at a time, seeded by last week's occupancy, exactly as
  myopic as the controller — but an explicit min-max *balance* term the
  controller has no equivalent of. Same-week tank swaps are only *softly*
  penalised in that objective, so it buys them freely: expect a
  transfer-heavy plan.

Both Global methods conserve fish exactly and pass the tank-continuity audit —
that part is real. **What they do not do:**

* They **never read the handling budget.** No Global code looks at
  `max_transfers_per_week`; they minimise moves in their objective but nothing
  caps a week. Expect weeks well above {k['moves']} moves.
* They enforce only **part** of the tier rulebook. R2/R3/R4 are checked when a
  transfer is paired up, and R6 is respected by the CP-SAT layout. **R1, R5 and
  R7 are not checked at all** while planning. When no legal source exists for a
  needed move, the Global pick **emits the move anyway** and writes a
  `TOPOLOGY VIOLATION` row to the ValidationLog — the controller would have
  refused it. Read that sheet before treating a Global plan as executable.
* Their planning pass **decomposes the horizon into independent weekly
  problems**, which is why week-to-week topology can break in the first place.
* Only **CP-SAT** enforces a real per-tank density cap (each tank's own
  kg/m³ × volume, from your facility config). The **LP** arm sizes tanks off a
  single facility-wide number — the *smallest* OG tank's legal mass — and where
  a batch cannot get enough tanks it packs it denser and flags the row.
  Nothing rejects an over-cap tank on the LP path.
* If CP-SAT cannot solve a week, that week falls through to an
  **unconstrained** placement (no density test) and the run writes
  `PLACEMENT DEGRADED` to the ValidationLog naming how many weeks it hit.
  On the current facility + PR this is 0 weeks, but it is a real path, not a
  removed one — so check for that row rather than assuming.

Use Global to ask *"how good could a plan be if nothing constrained handling
or topology?"* Use a Controller method for a plan the crew can actually run.

**What they may never do.** Diverge on the starting state: the manual window,
the PR hydration and the biology are shared — after a scripted window, *no*
engine may assume unscripted pre-start staging, so comparisons are honest
(a dark handoff week is dark in every engine, and the editor warned you
about it first).

**What binds them.** Conservation and tank continuity bind every method
equally. The rest is graded, not enforced: Compare & Choose shows four
hard-rule badges (conserves · fully placed · no empty week · under cap), and
Analyze's checklist adds the density and handling-budget gates, which are
**flagged, never disqualifying**, plus the **6N one-way rule (R7), which is
hard** — any outbound depuration transfer FAILs it and drops the plan to the
bottom of the ranking. That is why the Global family left the default roster: it
consumed most of an 8h35m tournament to produce arms that hard-fail R7. A plan
can still top a lens and be unrunnable for a SOFT reason; the badges and the
ValidationLog are how you tell.""")

    with st.expander("9 · The checks that bind everything (the audit net)"):
        st.markdown("""
Independent invariants, each catching a *different* failure (the hard lesson:
"all tests green" once coexisted with a silent 17% production loss):

1. **In-facility continuity (0 drift)** — every tank-week balances exactly.
2. **Input conservation, both ends** — every batch reaches the facility (0
   dropped) and none harvests more than it stocked (0 over-produced).
3. **Facility-level distributed loss** — catches many small same-sign leaks
   that per-row tolerances would each forgive.
4. **FW → seawater reconciliation** — realized entry vs plan per batch (a
   calibration signal, not a lost-fish gate).
5. **Graded-harvest + HOG consistency** — every event type accounted; sold
   tonnage consistent across sheets.
6. **Closed freshwater mass-balance** — the FW phase can't leak fish either.
7. **Steady-harvest contract** — no near-empty week past the startup handoff.

**On top of the audits, Analyze grades every plan on twelve gates, in this
order.** **Three** are hard — conservation, never-an-empty-week, and the 6N
one-way rule (R7) — and a hard FAIL sinks a plan no matter how well it scores on
everything else. The other eight are flagged and penalised, never
disqualifying:

| # | Gate | Hard? | Reads |
|---|---|---|---|
| 1 | Conservation (no fish created or lost) | **HARD** | 0 dropped **and** 0 over-produced, or FAIL |
| 2 | Never an empty harvest week | **HARD** | 0 empty weeks, or FAIL |
| 3 | Weekly contract floor (min harvest/week) | soft | PASS if every planner week clears `min_harvest_per_week`, else WARN with the count + the worst week |
| 4 | Facility biomass cap | soft | PASS ≤100% of cap · WARN ≤110% · FAIL above |
| 5 | Converges: red -> green -> stays green | soft | judges only **avoidable** red weeks · PASS if it never goes red, if every red week was forced, or if it settles and holds · WARN if it settles only after an avoidable relapse, or holds on <0.5% headroom · FAIL if the final week is still over **and** the plan owns it |
| 6 | Per-system feed capacity | soft | PASS if every system stays within its own feed cap · WARN on breaches within 10% · FAIL above 1.10× or on >25% of system-weeks — the feed system cannot deliver it |
| 7 | Weekly processing limit + relief | soft | PASS 0 relief weeks · WARN 1-3 · FAIL >3 **or** any week past the relief ceiling |
| 8 | Harvest targets (monthly/yearly) | soft | never worse than WARN — targets are penalised, never disqualifying |
| 9 | Per-batch density quality | soft | PASS if no batch peaks ≥1.3× its tank cap, else WARN |
| 10 | 6N one-way commitment (R7) | **HARD** | PASS if nothing left a depuration tank except by harvest · **FAIL** on any outbound transfer, which disqualifies the plan |
| 11 | Weekly handling budget | soft | PASS every week within budget · WARN any week over ~80% · FAIL any week over |
| 12 | Fish stuck in 6N purge | soft | PASS if every 6N tank drains within its rotation · WARN past 5 weeks in purge · **FAIL** past 8 — those fish are never harvested and would not survive. Conservation cannot see it: nothing is lost, they simply stand at horizon end |

**Gate 5 judges the plan; gate 4 judges what you inherited.** Peak biomass is
mostly a property of the *starting state* — hand every engine a Production
Report that opens at 101% of the cap and every engine peaks in week 0, so gate
4 cannot tell them apart. Gate 5 asks the operator's question instead: it
starts red, so **how fast does it work down to green, and can it hold there?**
It reads each week against **that week's** resolved limit (the cap moves —
3.80M in one week, 3.65M in the next), and it separates *touching* green from
*settling* into it, so a plan that dips under the cap in September and bulges
back to 107% in December reads as the relapse it is rather than "green by week
5". An inherited red start is not held against a plan; **reaching green and
giving it back is.**

**And it only charges what the plan could control.** Fish reach the plant only
through the 6N purge, and only fish at or above the minimum harvest weight may
be staged. When one cohort has passed through and the next has not yet grown
into the window there is *no mature inventory*: harvest cannot rise, biomass
climbs, and no engine under any knobs could do otherwise. The same holds for a
week already staging at the weekly processing limit — there is no legal move
left. Those weeks are **forced**: reported, never charged, as are
operator-scripted window weeks. A red week counts against the plan only when the
weeks that could still have fed it could between them have supplied the fish
needed to **erase the excess** — each offering the smaller of its remaining
plant capacity and the mature fish the pick could actually see. Maturity is read
from the snapshot that pick could SEE (the prior week), never the week-end one
that counts fish which grew into the window afterwards.

The yardstick is the excess itself, never an engine's own target. Asking "did it
get what it asked for?" would judge the controller on a number it records while
Global, which records none, kept being judged on plant capacity — the same
asymmetry that let Global's OG-only biomass flatter it for months. On the
8.23.26 PR this separates a controller plan with **1 avoidable** red week from a
Global plan with **14**, where the raw counts (15 and 30) largely measure the
same maturity trough twice.

**Gate 3 is the contract, gate 2 is only its degenerate case.** "Never an
empty week" catches a week that harvests *literally nothing*; the rule you
actually signed is a weekly **floor**. A plan can pass gate 2 and still miss
the floor nine times. Gate 3 is deliberately soft — near full utilisation
every plan misses it sometimes, and a gate that always FAILs teaches you to
ignore it — so use it to **compare** candidates, not to accept or reject one.

**A knob search cannot sell a rule to buy a better score.** The emphasis score
is not a safe place to keep a constraint, and twice it was measured failing to
hold one — so two guards now sit ABOVE the score, applied in this order and
independent of which emphasis you pick:

1. **The relief ceiling.** The weekly processing limit is represented in the
   objective by exactly one term, and the shipped *Product quality* preset
   weights that term **0** — so under it nothing, objective or gate, could see
   a breach. Across 717 measured plans its winners planned 82-83k-fish weeks on
   4 of 8 starting states: ~50% over the limit, ~37% over the ceiling the
   config itself calls never legal. A winner must now breach it in **zero**
   weeks.
2. **The contract floor.** The score has no floor term at all (its harvest
   components are a variability CV and an over-the-limit count; measured
   correlation between the worst week and the score: −0.03). On the 7.29 PR the
   search once chose settings that cut the plain controller's worst week
   20,526 → 16,185 fish. A winner's worst harvest week must now be **at least
   as good as that method's own un-tuned run**.

The ceiling goes first because it is the harder rule: a week over the ceiling
cannot be executed at all, while a lean week is a shortfall. Neither guard ever
empties the field — if nothing clears one it **stands down**, and the tool says
which one did, so the trade is yours to judge rather than the tool's to hide.
The same two rules guard **Adopt** and **Promote** on this page, which are the
other two ways a plan reaches your config: there they do not exclude, they
require you to acknowledge the finding by name, and it is saved alongside.

**Which weeks each gate judges.** Gates 2, 3 and 5 judge the **planner's**
weeks only: weeks you scripted yourself in the manual override window are
excluded from those counts, and the verdict says how many were excluded. Your
scripted weeks are policed instead by the `MANUAL WINDOW` lints in the
ValidationLog. Every other gate — including conservation — judges the **whole
horizon**, scripted weeks included.

Ranking order when Analyze picks a winner: hard FAILs → soft FAILs → total
warnings → target shortfall → the emphasis score. No emphasis or weighting can
lift a plan above one that beats it on an earlier tier.""")

    with st.expander("⚠ 10 · Known limits — the honest list"):
        st.markdown(f"""
* **"0 drift" proves bookkeeping, not biology.** The audits reconcile the
  plan against the *same* growth/feed curves that produced it — an internally
  consistent but biologically wrong model reconciles to itself. The check
  against reality is field data (the FW reconciliation signal is your first
  hint a correction needs retuning).
* **The hybrid's biomass peak is the price of steady harvest.** Holding fish
  back for lean weeks means they're still in the water: peak biomass runs a
  few percent over the cap (measured ~102.6 → ~107.1% switching the guide
  on). Every knob that shaves that peak was measured to put empty harvest
  weeks back — the spike *is* the reserve.
* **Pre-{k['prod_start'][:4]} weekly harvest is lumpy by design.** The 6N
  pair rotation quantizes weekly drains; leveling fills troughs but the pair
  granularity remains until production mode.
* **Fish ≥ 1 kg standing in entry tanks is necessary, not a bug (R6).**
  OG3-6 feed capacity binds before space does, so some overflow stays behind
  — measured on real PRs; forcing it forward breaks feed caps.
* **A week can exceed the {k['moves']}-move handling budget.** Essential
  moves (rotation fills, arrival make-room) are never blocked; such a week
  is flagged on the handling gate rather than silently truncated. Note the
  budget is a *Controller* mechanism — the Global engines never read it at
  all (layer 8).
* **The Global engines are benchmarks, not runnable plans.** They do not
  enforce R1, R5 or R7 while planning, they ignore the handling budget, and
  the LP variant has no per-tank density constraint. They conserve fish and
  balance the facility beautifully; that is a different question from "can
  the crew execute this". Layer 8 has the full list — read it before
  adopting a Global plan.
* **Three gates can actually sink a plan.** Conservation,
  never-an-empty-week and the 6N one-way commitment (R7) are the hard
  ones. THREE more can never even reach FAIL by design — the weekly
  contract floor, harvest targets and per-batch density stop at WARN.
  (The floor still binds, just not here: a tuned tournament will not
  promote a variant that starves its lean weeks.) The remaining five
  (biomass cap, convergence, per-system feed, processing limit, handling
  budget) *can* read FAIL but are soft: they rank a plan down, they do
  not disqualify it.
  So a plan can be recommended with a red handling gate — the checklist
  shows it, and it is your call, not the tool's.
* **Severe per-batch density clusters are a stocking problem.** When a
  cohort's tanks collide mid-grow-out, no planner knob fixes it — the remedy
  is stocking fewer fish (the stocking frontier in Analyze quantifies the
  trade).
* **6N off-feed mortality is slightly under-counted per tank** — a
  few-fish-per-week approximation that nets out facility-wide and is
  deliberately left (correcting it destabilizes the facility-level balance
  it currently cancels against).
* **PR-hydrated purge fish can show short residency at horizon start** — a
  measurement artifact (their clock predates the forecast), exempted from
  the hold audit on purpose.""")

    st.caption(
        "Sources: the tier rulebook (forecast/tiers.py), the Control config "
        "and its tooltips (Configure → Control), and docs/USER_GUIDE.md — "
        "which adds the full tuning guidance this page deliberately leaves "
        "out.")


# ============================================================
# Accuracy — forecast vs actuals
# ============================================================

def _acc_fmt(v, suffix="%", nd=1):
    """None-safe number formatting: a missing measurement prints as '—', never
    as 0.0 (a zero error and an unmeasurable one are opposite findings)."""
    if v is None:
        return "—"
    return f"{v:,.{nd}f}{suffix}"


def _forecast_vs_actuals():
    """Grade a past forecast against the ProductionReport that followed it.

    The only mode that measures the BIOLOGY. Everything else in this app
    measures the plan; 700 tests prove the bookkeeping. This answers the
    question none of them can: how far off is the growth model.
    """
    from forecast import accuracy as _acc

    st.header("🎯 Accuracy — forecast vs actuals")
    st.caption(
        "Grades a forecast you ran **earlier** against the ProductionReport "
        "that came **after** it. Both files are read only — nothing is run, "
        "nothing is saved, no config is touched, and this can never change a "
        "plan. It measures the **growth model**, which no other mode does. "
        "What it CANNOT see: harvest execution (a PR shows what is in the "
        "water, so fish already sold are simply absent), freshwater (it is not "
        "in either file), and any batch that appears in only one of the two — "
        "those are listed as coverage instead of being averaged in.")

    st.subheader("1. The forecast being graded")
    prev = st.file_uploader(
        "A forecast output workbook you produced earlier (.xlsm / .xlsx)",
        type=["xlsm", "xlsx"],
        help="Any workbook this tool wrote — it needs the BatchLocations "
             "sheet. Pick one anchored BEFORE the ProductionReport below; the "
             "gap between the two is the horizon being graded.",
        key="acc_prev_fc",
    )

    st.subheader("2. The actuals")
    _pr_ok = pr is not None and pr["ok"]
    alt = st.file_uploader(
        "Optional — grade against a different ProductionReport",
        type=["xlsm", "xlsx"],
        help="Leave empty to use the ProductionReport already loaded in the "
             "sidebar, which is the normal case: today's PR IS the actuals.",
        key="acc_alt_pr",
    )
    if alt is not None:
        st.caption(f"Using the uploaded **{alt.name}** as actuals.")
    elif _pr_ok:
        st.caption(f"Using the sidebar ProductionReport — closing "
                   f"**{pr['closing']:%Y-%m-%d}**.")
    else:
        st.info("Upload a **ProductionReport** in the sidebar (or above) to "
                "supply the actuals.")

    if prev is None:
        st.info("Upload a previous forecast workbook above to grade it. "
                "Nothing else in the app is affected until you do.")
        _acc_calibration_section()
        return
    if alt is None and not _pr_ok:
        _acc_calibration_section()
        return

    # Recompute only when either input changes — a report computed from other
    # files must never be shown as if it described these ones.
    import hashlib
    act_bytes = alt.getvalue() if alt is not None else uploaded.getvalue()
    sig = hashlib.md5(prev.getvalue()).hexdigest() + "|" + \
        hashlib.md5(act_bytes).hexdigest()
    cached = st.session_state.get("_acc_report")
    if not cached or cached.get("sig") != sig:
        try:
            rep = _acc.compare(io.BytesIO(prev.getvalue()),
                               io.BytesIO(act_bytes))
        except Exception as exc:  # noqa: BLE001 — a bad pair must not kill the page
            st.error(f"Could not grade these two files: {exc}")
            _acc_calibration_section()
            return
        st.session_state["_acc_report"] = {"sig": sig, "rep": rep}
        cached = st.session_state["_acc_report"]
    rep = cached["rep"]

    _acc_render(rep)
    _acc_calibration_section()


def _acc_render(rep):
    """The report itself. Leads with the number that answers 'how much should
    I trust this forecast?' — typical and worst batch weight error, over the
    elapsed time they accumulated across."""
    from forecast import accuracy as _acc
    h = _acc.headline(rep)

    st.divider()
    st.subheader("How much should I trust this forecast?")
    if h["batches_graded"] == 0:
        st.warning(
            "No batch appears in **both** files, so there is nothing to "
            "grade. That usually means the two files are from different "
            "periods — check the dates below.")
    else:
        c = st.columns(4)
        c[0].metric(
            "Typical weight error",
            _acc_fmt(h["typical_wt_err_pct"]),
            help="Median across graded batches of |predicted − actual| mean "
                 "weight, as a % of actual. THE headline: the error you "
                 "should expect on a typical batch over the elapsed time "
                 "shown. Read it next to that horizon — an error means "
                 "nothing without the time it accumulated over.")
        c[1].metric(
            "Worst batch",
            _acc_fmt(h["worst_wt_err_pct"]),
            help="The largest single-batch weight error. A plan difference "
                 "smaller than this is inside the model's own noise for at "
                 "least one batch.")
        c[2].metric(
            "Signed median", _acc_fmt(h["signed_median_pct"]),
            help="The same errors WITH their sign. Positive = the model "
                 "predicted heavier fish than reality (runs hot). This is the "
                 "number that matters most: a one-signed error can be "
                 "corrected at source, while scatter both ways is noise the "
                 "monthly re-anchor already absorbs.")
        c[3].metric(
            "Elapsed", f"{rep.elapsed_weeks:g} wk",
            help="Time between the forecast's anchor and the actuals' closing "
                 "date — the horizon these errors accumulated across.")

        v = rep.bias.get("verdict", "")
        if "Systematic" in v:
            st.warning(f"**{v}**")
        else:
            st.info(f"**{v}**")

    st.caption(
        f"Forecast anchored **{rep.forecast_anchor:%Y-%m-%d}** → actuals "
        f"closing **{rep.actual_closing:%Y-%m-%d}** "
        f"({rep.elapsed_days} days). Prediction read "
        + ("by interpolating between weeks "
           f"**{'** and **'.join(rep.basis.get('weeks', []))}** so it lands on "
           "the exact closing date."
           if rep.basis.get("method") == "interpolated"
           else f"from the single nearest weekly snapshot "
                f"(**{rep.aligned_week}**, "
                f"{rep.basis.get('offset_days', 0):+d} days off)."))

    for n in rep.notes:
        st.warning(f"⚠ {n}")

    # ---- batch level: the biology ------------------------------------------
    st.divider()
    st.subheader("Per batch — the biology")
    st.caption(
        "**This is the honest model-error view.** Fish are summed per batch "
        "over whatever tanks they ended up in, so placing fish differently "
        "from the plan does not show up here as a bad prediction. **Weight** "
        "is the growth-model score. **Count** and **biomass** also move with "
        "harvest, culling, grading and transfers, so they mix model error "
        "with execution — do not read them as pure model error.")
    graded = rep.graded
    if graded:
        rows = [{
            "Batch": b.batch_id,
            "Pred wt (g)": round(b.pred_wt_g),
            "Actual wt (g)": round(b.act_wt_g),
            "Weight err %": round(b.wt_err_pct, 2) if b.wt_err_pct is not None else None,
            "Pred count": round(b.pred_count),
            "Actual count": round(b.act_count),
            "Count err %": round(b.count_err_pct, 2) if b.count_err_pct is not None else None,
            "Pred biomass (kg)": round(b.pred_biomass_kg),
            "Actual biomass (kg)": round(b.act_biomass_kg),
            "Biomass err %": round(b.biomass_err_pct, 2) if b.biomass_err_pct is not None else None,
        } for b in graded]
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)

        fig = px.bar(
            pd.DataFrame([{"Batch": b.batch_id, "Weight error %": b.wt_err_pct}
                          for b in graded if b.wt_err_pct is not None]),
            x="Batch", y="Weight error %",
            title="Weight error by batch (positive = model predicted heavier "
                  "than reality)")
        fig.add_hline(y=0, line_width=1)
        # use_container_width, not width=: st.plotly_chart has no `width`
        # parameter, so it would be swallowed into **kwargs and silently do
        # nothing. This matches the convention used by every other chart here.
        st.plotly_chart(fig, use_container_width=True)

        f = rep.facility
        st.markdown("**Facility totals** (graded batches only)")
        fc = st.columns(3)
        fc[0].metric("Mean weight", _acc_fmt(f["wt_err_pct"]),
                     help="Facility mean weight: predicted vs actual, as a %. "
                          "Signed.")
        fc[1].metric("Total count", _acc_fmt(f["count_err_pct"]),
                     help="Includes harvest timing, not just model error.")
        fc[2].metric("Total biomass", _acc_fmt(f["biomass_err_pct"]),
                     help="Count error and weight error combined.")

    cov = rep.coverage
    if cov.get("batches_forecast_only") or cov.get("batches_actual_only"):
        st.caption(
            f"**Not graded** — in one file only: forecast-only "
            f"{cov.get('batches_forecast_only') or '—'}, actuals-only "
            f"{cov.get('batches_actual_only') or '—'}. A forecast-only batch "
            f"is usually one that was harvested out; an actuals-only batch is "
            f"one that entered after this forecast was made. Neither is "
            f"evidence about the growth model, so neither is averaged in.")

    # ---- alignment sensitivity ---------------------------------------------
    s = rep.sensitivity or {}
    if s.get("graded"):
        with st.expander("📅 How much does the date alignment matter?"):
            st.caption(
                "The forecast only produces a value once a week, so grading "
                "against the wrong week charges the calendar gap to the "
                "model. The headline above avoids that by reading the exact "
                "closing date. This shows what the neighbouring weekly "
                "snapshots would have said against the same actuals — the "
                "spread is how much the question 'which week?' is worth.")
            srows = [{"Forecast week": d["week"],
                      "Week ends": f"{d['week_end']:%Y-%m-%d}",
                      "Typical weight error %": round(d["typical_wt_err_pct"], 2)}
                     for k in ("previous", "graded", "next")
                     for d in [s.get(k)] if d]
            st.dataframe(pd.DataFrame(srows), width="stretch", hide_index=True)

    # ---- tank level: adherence, NOT model error ----------------------------
    st.divider()
    st.subheader("Per tank — plan adherence (NOT model error)")
    st.caption(
        "**Read this as a different question entirely.** It asks whether the "
        "fish ended up where the plan put them. A mismatch is an operator "
        "decision — or a plan you improved on — and is **not** evidence that "
        "the growth model is wrong. Unlike the biology view above, this is a "
        f"single weekly snapshot (**{rep.aligned_week}**, "
        f"{rep.alignment_offset_days:+d} days from the closing date), because "
        "tank occupancy is discrete and cannot be interpolated.")
    adh = cov.get("tank_adherence_pct")
    st.metric("Tank cells matching the plan", _acc_fmt(adh),
              help="Share of (batch, tank) cells present in BOTH the plan and "
                   "the actuals. Low is not bad in itself — it means the "
                   "facility was run differently from this particular plan.")
    if rep.tanks:
        with st.expander("Per-(batch, tank) detail"):
            st.dataframe(pd.DataFrame([{
                "Batch": t.batch_id, "Tank": t.tank_id, "Where": t.present,
                "Pred count": round(t.pred_count),
                "Actual count": round(t.act_count),
                "Diff": round(t.count_err),
            } for t in rep.tanks]), width="stretch", hide_index=True)

    st.divider()
    with st.expander("⚠️ What this measurement cannot tell you", expanded=False):
        for lim in rep.limits:
            st.markdown(f"- {lim}")


def _acc_calibration_section():
    """Freshwater calibration drift — the model error the tool already
    measures on every run and used to throw away."""
    from forecast import accuracy as _acc

    st.divider()
    st.subheader("🧪 Freshwater calibration history")
    st.caption(
        "Every run with FW auto-calibration on back-solves each freshwater "
        "batch's `fw_correction` — the factor that makes the model land on "
        "the transfer target. A value of 0.77 means the model grew that batch "
        "**23% faster than reality**. Those rewrites used to scroll past in "
        "the run log and were never kept, so a correction needed every month "
        "for six months looked exactly like a one-off. They are now appended "
        "to `fw_calibration_history.jsonl` beside the optimize and adoption "
        "logs. A batch flagged **persistent** below has needed the same "
        "correction repeatedly — that is a standing model error to fix in the "
        "biology config, not to re-correct every month. This is freshwater "
        "only; it says nothing about seawater growth.")
    recs = _acc.read_calibration_log(str(_ROOT / _acc.DEFAULT_CALIB_LOG))
    if not recs:
        st.info("No calibration history yet — it starts filling on your next "
                "run with FW auto-calibration on.")
        return
    drift = _acc.calibration_drift(recs)
    st.caption(f"{len(recs):,} rewrites recorded across "
               f"{len({r.get('ts') for r in recs})} run(s).")
    st.dataframe(pd.DataFrame([{
        "Batch": d["batch"], "Runs": d["runs"],
        "Configured": d["median_configured"],
        "Applied (median)": d["median_applied"],
        "Gap": d["gap"], "Spread": d["spread"],
        "Clamped runs": d["clamped_runs"],
        "Persistent?": "⚠ yes" if d["persistent"] else "",
    } for d in drift]), width="stretch", hide_index=True)
    persistent = [d["batch"] for d in drift if d["persistent"]]
    if persistent:
        st.warning(
            f"⚠ Standing model error on {', '.join(persistent)} — the applied "
            f"correction has sat away from the configured value across "
            f"several runs. Fix it in the biology config rather than letting "
            f"auto-calibration re-discover it every month.")


# ============================================================
# Page setup
# ============================================================

st.set_page_config(
    page_title="AS Forecast — NEXT (v2)",
    page_icon="🐟",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("AS Production Forecast — NEXT (v2)")
st.caption("Development build on branch `v2/roster-and-emphasis`. "
           "Your production app is the one under OneDrive; this one is safe to break.")
st.caption(
    "Upload an input workbook, run the planner, review the results, "
    "and download a new workbook with all tabs populated. "
    "The input workbook is never modified."
)


# ============================================================
# Sidebar — upload + run controls
# ============================================================

with st.sidebar:
    # "Computer power" — how much of this machine the heavy runs may use. Stored
    # as a PERCENT of logical CPUs (operator-friendly); _cpu_workers() translates
    # it into CP-SAT search threads / sweep worker processes at the call sites.
    st.slider(
        "Computer power", min_value=10, max_value=100, value=40, step=10,
        format="%d%%", key="cpu_pct",
        help="How much of this computer the heavy runs may use — the Global "
             "optimal (CP-SAT) solver (also inside Compare & Choose) and the "
             "Optimize sweeps. Higher = faster runs, but other applications "
             "may feel slower (and Optimize sweeps use more memory) while a "
             "run is going. A plain controller Run forecast is "
             "sequential and unaffected.",
    )
    st.caption(f"Heavy runs may use up to **{_cpu_workers()}** of "
               f"{os.cpu_count() or 2} processor cores.")
    st.divider()
    # If ▶ Run forecast was clicked from another mode, jump to Run forecast HERE —
    # before the radio is instantiated, so setting its session_state value is allowed
    # (Streamlit forbids mutating a widget's key after it renders). The pending run is
    # then honored by the run handler below.
    if st.session_state.pop("_goto_run_mode", False):
        st.session_state["app_mode"] = "Run forecast"
    # Listed in the ORDER OF OPERATIONS an operator actually works in:
    # set up -> run -> analyse -> compare -> hand-steer -> read the rulebook.
    # Each option carries a one-line caption saying what it is FOR, so the
    # choice is legible without opening a tooltip.
    # If a stored selection references the retired Tune mode, reset it BEFORE
    # the radio instantiates (a stored value outside the options raises).
    if str(st.session_state.get("app_mode", "")).startswith("Tune"):
        st.session_state["app_mode"] = "Run forecast"
    # Analyze / Compare & Choose / Optimize merged into Decide (2026-08-31).
    # A stored value outside the options raises, so migrate BEFORE the radio
    # instantiates — same reason as the Tune migration above. Each retired name
    # lands on the tab that carries its capabilities, so a returning operator
    # arrives where they left off rather than at a generic front page.
    _MERGED_INTO_DECIDE = {"Analyze": 0, "Compare": 1, "Optimize": 2}
    _stored = str(st.session_state.get("app_mode", ""))
    for _old, _tab in _MERGED_INTO_DECIDE.items():
        if _stored.startswith(_old):
            st.session_state["app_mode"] = "Decide (which plan should I run?)"
            st.session_state.setdefault("_decide_tab", _tab)
            break
    # The list reads as the order of operations, but the app must still LAND on
    # the everyday mode, not on Configure (option 0). Seed the stored value on
    # first render only — after that the operator's own selection wins.
    st.session_state.setdefault("app_mode", "Run forecast")
    app_mode = st.radio(
        "Mode",
        ["Configure (models & control)", "Run forecast",
         "Decide (which plan should I run?)",
         "Accuracy (forecast vs actuals)", "How it works (the rules)"],
        captions=[
            "Set up once — biology curves, tanks, batches, per-week limits, "
            "control knobs, harvest targets and prices.",
            "The everyday step — run your chosen plan on today's "
            "ProductionReport and download the workbook.",
            "“Which plan should I run?” — the monthly lever check, the "
            "engine board, and knob tuning, in one place.",
            "“How much should I trust this?” — grades a past forecast against "
            "the PR that followed it.",
            "The rulebook — what each layer decides, what it may never do, "
            "and the honest list of known limits.",
        ],
        help="Listed in the order you normally work: Configure once, then Run "
             "forecast every day. Analyze answers 'which plan?' end to end; "
             "Compare & Choose and Optimize are the two halves of that "
             "question you can steer by hand (Compare & Choose changes the "
             "ENGINE, Optimize changes the KNOBS of one engine). How it works "
             "is the plain-language rulebook — read it once before trusting "
             "or challenging a plan.",
        key="app_mode",
    )
    with st.expander("ℹ️ Which mode? — the order of operations"):
        st.markdown(
            "**The workflow, in order:** set up **Configure** once (models, "
            "facility, targets, prices) → run **Analyze** to pick and tune the "
            "best plan and **⭐ Promote** it → then **Run forecast** (or "
            "Analyze's ⚡ Quick run) is the everyday step. **Compare & Choose** "
            "and **Optimize** are the two halves of the Analyze decision, "
            "available on their own when you want to steer one by hand.\n\n"
            "- **Configure (models & control)** — hand-edit the biology curves, "
            "the tank list, the batch schedule, the per-week limits, every "
            "control knob, and your harvest targets + price bands. Saved to "
            "`config/` + `scenario/` YAML, which is the source of truth for "
            "every run. Every field has a tooltip explaining what it does, its "
            "unit, and what it trades off.\n"
            "- **Run forecast** — runs the pipeline with your **current** "
            "Control knobs and your **currently picked method**, and produces "
            "the plan + reports. This is the everyday mode. *\"Run with tuned "
            "knobs\"* just means a normal Run **after** Analyze/Optimize has "
            "saved better knobs into your config.\n"
            "- **Analyze (find my best plan)** — the one-flow version of the "
            "whole decision: every engine + a knob search + the hard-rule "
            "checklist (conservation and never-an-empty-week are the two HARD "
            "rules; caps, handling, your harvest targets, revenue and the "
            "per-batch density lens are scored but never disqualifying) → one "
            "recommendation card with **Adopt** (use it now) and **Promote** "
            "(make it the Quick-run default). The stocking-for-quality "
            "frontier lives here too. Finished runs are cached to disk and "
            "shared with Compare & Choose, so nothing runs twice.\n"
            "- **Compare & Choose (all methods)** — runs the *different "
            "engines* (Controller plain / hybrid / +LNS, Global LP, Global "
            "CP-SAT) on one PR, grades them on several lenses (fewest moves, "
            "steadiest harvest, between/within-system balance, density, "
            "welfare, footprint) with hard-rule badges, and lets you pick "
            "which whole plan becomes the report. **This is where the "
            "planning method is chosen** — ▶ Run forecast re-runs whatever you "
            "picked here. Unlike Optimize (same engine, different knobs), this "
            "compares *engines*.\n"
            "- **Optimize (multi-objective)** — sweeps the **controller "
            "family's** knobs against several goals at once (flat biomass, "
            "feed, handling, cap compliance) on a *selectable* weighted "
            "objective, ranks the settings, and can apply the best. It finds "
            "knob *combinations* a one-knob-at-a-time sweep can't. It does not "
            "tune the Global engines — those have no tunable knobs (see "
            "Analyze's run-budget table).\n"
            "- **Accuracy (forecast vs actuals)** — the only mode that grades "
            "the *biology* rather than the plan. Upload a forecast workbook "
            "you produced earlier plus the ProductionReport that came after "
            "it, and it reports how far each batch's predicted weight and "
            "count missed reality, and whether the misses lean one way. Costs "
            "nothing to run and changes nothing: it reads two files. It cannot "
            "grade harvest execution or freshwater — see the mode's own "
            "limits list.\n"
            "- **How it works (the rules)** — the plain-language rulebook: "
            "every pipeline layer (inputs → manual window → freshwater → entry "
            "→ placement → harvest → depuration → audits), what each decides, "
            "what it may never do, which checks bind it, and the honest "
            "known-limits list. Read it once before trusting or challenging a "
            "plan.\n\n"
            "*(The old Tune mode retired — its density distribution + severe-batch "
            "readout is now a checklist gate + drill-in on the Analyze board, and "
            "the stocking frontier moved there with it. The headless density sweep "
            "is still available via `tools/tune_sweep.py`.)*"
        )
    st.divider()

    st.header("ProductionReport")
    # Keyed by a nonce so "Clear PR & last run" can reset the widget (bumping
    # the nonce gives the uploader a fresh key → it renders empty).
    uploaded = st.file_uploader(
        "ProductionReport workbook (.xlsx / .xlsm)",
        type=["xlsm", "xlsx"],
        help="Only the ProductionReport sheet is read. The forecast start is "
             "derived from its closing date. Models/limits come from config.",
        key=f"pr_upload_{st.session_state.get('_pr_nonce', 0)}",
    )

    pr = _ingest_pr(uploaded) if uploaded is not None else None
    if pr is not None:
        if pr["ok"]:
            st.success(
                f"PR ✓ — forecast start **{pr['forecast_start'].date()}** "
                f"(closing {pr['closing']}) · {pr['n_og']} OG + {pr['n_fw']} FW rows"
            )
        else:
            for e in pr["errors"]:
                st.error(e)
        for w in pr.get("warnings", []):
            st.warning(w)

    _cfg_ok = _config_ready() and _scenario_ready()
    _pr_ok = pr is not None and pr["ok"]
    if not _cfg_ok:
        st.info("No config yet — set it up in **Configure**.")

    st.header("Run")
    # The planning method is chosen ONCE, on the Compare & Choose board, where
    # you can see every method graded side by side. Run forecast just re-runs
    # whichever plan you picked — no second, blind choice on the main screen.
    _chosen = st.session_state.get("_chosen_method", _DEFAULT_METHOD)
    _chosen_m = _method_obj(_chosen)
    st.caption(f"Method: **{_chosen_m.label}**")
    if _chosen == _DEFAULT_METHOD:
        st.caption("The default. It is the only method measured to harvest "
                   "something every single week — the plain Controller has an "
                   "empty week on 5 of 6 real PRs. Compare & Choose runs every "
                   "method and lets you pick a different one.")
    else:
        st.caption("Picked on the Compare & Choose board. Pick another there, "
                   "or re-select the hybrid to go back to the default.")
    if _cfg_ok:
        # Guarded because this runs in EVERY mode, before the mode dispatch: an
        # unreadable control.yaml here used to blank the whole app, including
        # the Configure tab that would fix it.
        from forecast.config_io import load_control, control_to_dict
        _ok_cd, _cd = _read_or_explain(
            lambda: control_to_dict(load_control(CONFIG_DIR)),
            "config/control.yaml")
        if _ok_cd:
            _render_active_config(
                _cd, "ℹ️ Active configuration — what this run will do")
    run_clicked = st.button(
        "▶ Run forecast",
        type="primary",
        disabled=(not _pr_ok or not _cfg_ok),
        use_container_width=True,
        help=None if (_pr_ok and _cfg_ok)
        else "Upload a valid ProductionReport and set up config first.",
    )

    # The ▶ Run forecast button lives in the sidebar in EVERY mode, but the run
    # results render only in Run forecast mode — so clicking it from Configure/
    # Optimize used to silently do nothing (the main panel st.stop()s before the run
    # handler). Jump to Run forecast and run there instead.
    if run_clicked and app_mode != "Run forecast":
        st.session_state["_goto_run_mode"] = True
        st.session_state["_pending_run"] = True
        st.rerun()

    if "result" in st.session_state and st.session_state.result.get("ok"):
        r = st.session_state.result
        st.success(
            f"Last run: {r['elapsed']:.1f}s, {r['violations']} viols"
        )
        st.download_button(
            label="⬇ Download output workbook",
            data=r["output_bytes"],
            file_name=r["output_name"],
            mime=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                  if str(r["output_name"]).lower().endswith(".xlsx")
                  else "application/vnd.ms-excel.sheet.macroenabled.12"),
            use_container_width=True,
        )
        # Fallback: show where the file lives on disk in case the
        # download button doesn't work (browser quirks, file size, etc.)
        if r.get("output_path"):
            st.caption(f"Also saved at:\n`{r['output_path']}`")
    elif "result" in st.session_state and not st.session_state.result.get("ok"):
        st.error("Last run failed — see error in main panel.")
        if st.session_state.result.get("output_path"):
            st.caption(
                f"Partial output (if any):\n"
                f"`{st.session_state.result['output_path']}`"
            )
    else:
        st.caption(
            "Upload a workbook, click ▶ Run forecast. The download "
            "button appears here after the run completes."
        )

    # ---- Clear / start fresh (config + scenario are kept) ----
    if uploaded is not None or "result" in st.session_state:
        st.divider()
        if st.button("🗑 Clear PR & last run", use_container_width=True,
                     help="Remove the uploaded ProductionReport and the last run "
                          "result, to start a fresh forecast. Your config + "
                          "scenario (models, limits, batches) are kept."):
            for _k in ("result", "_pr", "_pr_key"):
                st.session_state.pop(_k, None)
            # Bump the nonce so the file_uploader widget resets to empty.
            st.session_state["_pr_nonce"] = st.session_state.get("_pr_nonce", 0) + 1
            st.rerun()


# ============================================================
# Pipeline runner
# ============================================================

class _TeeIO(io.StringIO):
    """Captures stdout AND forwards each completed line to `on_line`.

    The pipeline narrates its stages to stdout; capturing it into a plain
    StringIO meant the operator saw nothing until the run returned — a static
    spinner for anything up to half an hour. This tees the stream so the UI can
    show the current stage live, while `getvalue()` stays byte-identical to the
    old capture (super().write runs first and unconditionally, whatever the
    callback does). A callback that raises disables itself rather than
    spamming; callbacks must only touch `st` APIs — printing from one would
    re-enter this writer under redirect_stdout and pollute the transcript.
    """

    def __init__(self, on_line=None):
        super().__init__()
        self._on_line = on_line
        self._buf = ""

    def write(self, s: str) -> int:
        n = super().write(s)            # transcript first: never lose output
        if self._on_line and s:
            self._buf += s
            while "\n" in self._buf:
                line, self._buf = self._buf.split("\n", 1)
                line = line.strip()
                if not line:
                    continue
                try:
                    self._on_line(line)
                except Exception:  # noqa: BLE001 — narration is best-effort
                    self._on_line = None
                    break
        return n


def _run_with_workbook_bytes(
    input_bytes: bytes,
    input_name: str,
    config_dir: str | None = None,
    scenario_dir: str | None = None,
    method: str = "controller",
    cpsat_time: float = 300.0,
    cpsat_workers: int | None = None,
    cpsat_det_time: float | None = None,
    on_line=None,
) -> dict:
    """Run the pipeline against `input_bytes` in a temp directory.

    `cpsat_workers` = CP-SAT search threads for the global-optimal method
    (None -> the engine default of 8); callers pass _cpu_workers() so the
    sidebar's "Computer power" percent governs it.

    `on_line` receives each stage line the pipeline prints, as it prints it, so
    the caller can narrate progress. The captured stdout is unaffected.

    When config_dir/scenario_dir are given (PR-only mode), the stable
    config + scenario load from YAML and the uploaded workbook supplies
    only the ProductionReport. Returns a dict with metrics + the output
    workbook bytes + parsed data needed for visualization.
    """
    work_dir = Path(tempfile.mkdtemp(prefix="as_forecast_"))
    in_path = work_dir / input_name
    _m = (_AS_CONFIGURED if method == "as-configured"
          else _METHODS.get(method) or _METHODS["controller"])
    _is_global_engine = (_m.engine == "global")
    _is_optimal = bool(_m.engine_kwargs.get("optimal"))
    # The global method emits a fresh .xlsx (no VBA to carry); the controller
    # keeps the uploaded macro workbook's suffix.
    if _is_global_engine and _is_optimal:
        out_name = Path(input_name).stem + "_planned_OPTIMAL.xlsx"
    elif _is_global_engine:
        out_name = Path(input_name).stem + "_planned_GLOBAL.xlsx"
    else:
        # Match the output extension to the workbook's MACRO STATE, not the input
        # name's suffix: the controller keeps VBA on load, so a macro-enabled input
        # yields a macro-enabled output that MUST be .xlsm — Excel refuses a macro
        # workbook wearing a .xlsx extension. (run.py also backstops this on save.)
        from forecast.excel_io import is_macro_enabled_workbook
        _suf = ".xlsm" if is_macro_enabled_workbook(input_bytes) else ".xlsx"
        out_name = Path(input_name).stem + "_planned" + _suf
    out_path = work_dir / out_name
    in_path.write_bytes(input_bytes)

    # A method's control-knob overrides (LNS placement, hybrid follow mode, ...)
    # are applied via a throwaway config COPY so the user's control.yaml is
    # never touched — same contract as forecast.methods.run_method.
    run_config_dir = config_dir
    if _m.overrides and config_dir:
        import shutil
        import yaml as _yaml
        _tmpcfg = work_dir / f"config_{_m.key.replace('-', '_')}"
        shutil.copytree(config_dir, _tmpcfg)
        _cy = _tmpcfg / "control.yaml"
        _d = _yaml.safe_load(_cy.read_text()) or {}
        _d.update(_m.overrides)
        _cy.write_text(_yaml.safe_dump(_d, sort_keys=False))
        run_config_dir = str(_tmpcfg)

    # Run the pipeline, capturing console output for display.
    t0 = time.time()
    captured = _TeeIO(on_line)
    try:
        with redirect_stdout(captured):
            if _is_global_engine:
                from tools.run_global_forecast import run_global
                rc = run_global(in_path, out_path, run_config_dir, scenario_dir,
                                optimal=_is_optimal,
                                cpsat_time=cpsat_time,
                                cpsat_workers=(cpsat_workers or 8),
                                cpsat_det_time=(cpsat_det_time or 30.0))
            else:
                rc = run_pipeline(input_path=in_path, output_path=out_path,
                                  config_dir=run_config_dir, scenario_dir=scenario_dir)
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc(),
            "stdout": captured.getvalue(),
            "elapsed": time.time() - t0,
        }
    elapsed = time.time() - t0

    # run.py coerces the output extension to match the workbook's macro state (.xlsm
    # vs .xlsx, so Excel can open it), which may differ from the name the app chose —
    # locate the actual saved file (same stem, any extension) before reading it.
    if not out_path.exists():
        _alt = next(iter(work_dir.glob(out_path.stem + ".*")), None)
        if _alt is not None:
            out_path = _alt

    if rc != 0 or not out_path.exists():
        return {
            "ok": False,
            "error": f"Pipeline returned rc={rc} or no output produced",
            "stdout": captured.getvalue(),
            "elapsed": elapsed,
            "output_path": str(out_path) if out_path.exists() else None,
        }

    # Read parsed outputs for visualization.
    output_bytes = out_path.read_bytes()
    parsed = _parse_output_workbook(out_path)
    # Capture the EFFECTIVE config this run used, so the result can always show
    # what produced it. MUST read run_config_dir, not config_dir: with method
    # overrides it is the throwaway copy they were applied to, while config_dir
    # still holds the user's un-overridden values. Reading config_dir made an LNS
    # run report placement_method=greedy, and would now report the pinned
    # controller arms as hybrid_follow=full (the base config's value) — i.e. the
    # panel would describe a different plan than the one on screen. It falls back
    # to config_dir automatically: run_config_dir IS config_dir when no overrides.
    config_used = {}
    if run_config_dir:
        try:
            from forecast.config_io import load_control, control_to_dict
            config_used = control_to_dict(load_control(run_config_dir))
        except Exception as e:  # noqa: BLE001
            # Config WAS used — only the read-back for display failed. An
            # empty dict makes the "Configuration this run used" panel vanish
            # as if no config existed; say what actually happened.
            print(f"WARN: could not read back the effective run config "
                  f"({type(e).__name__}: {e}) — the config panel will be "
                  f"empty for this run")
            config_used = {}
    parsed.update({
        "ok": True,
        "elapsed": elapsed,
        "stdout": captured.getvalue(),
        "output_bytes": output_bytes,
        "output_name": out_path.name,
        "output_path": str(out_path),
        "config_used": config_used,
        # Identity of this run, so the results view can memoize its derived
        # frames and only rebuild them when a DIFFERENT run is displayed.
        "_rid": uuid.uuid4().hex,
        # Provenance: WHEN the engine actually ran — travels with the result
        # through the board store / disk cache so a replayed leg can say
        # "cached run of <this time>" instead of passing as current.
        "run_ts": datetime.now().isoformat(timespec="seconds"),
    })
    # Session-scoped freshness: anything in this set ran in THIS session;
    # everything else on screen is a cache replay (see _res_is_fresh).
    try:
        st.session_state.setdefault("_fresh_rids", set()).add(parsed["_rid"])
    except Exception:  # noqa: BLE001 — headless callers have no session
        pass
    return parsed


def _parse_output_workbook(path: Path) -> dict:
    """Extract data from the saved workbook for the UI's visualization."""
    wb = load_workbook(path, keep_vba=str(path).lower().endswith(".xlsm"),
                       data_only=False)

    # Per-tank density caps from facility config (a control input) — never a
    # hardcoded literal. Each tank's over-cap is judged against ITS OWN cap
    # (nursery 30–65, grow-out 95), so the Violations count stays correct if
    # the cap is retuned or if lower-cap tanks ever enter the placement.
    tank_caps, sys_cap_biomass = {}, {}
    growout_cap = 95.0
    try:
        from forecast.config_io import load_facility_config
        _fac = load_facility_config(CONFIG_DIR)
        for t in _fac.tanks:
            tank_caps[t.tank_id] = t.max_density_kg_m3
            sys_cap_biomass[t.system_id] = (
                sys_cap_biomass.get(t.system_id, 0.0)
                + t.volume_m3 * t.max_density_kg_m3
            )
        # Representative grow-out density cap for facility-wide reference lines:
        # the OG production tanks, excluding the OG6N depuration pool.
        _grow = [t.max_density_kg_m3 for t in _fac.tanks
                 if t.system_id.startswith("OG") and t.system_id != "OG6N"]
        if _grow:
            growout_cap = float(max(_grow))
    except Exception as e:  # noqa: BLE001
        # Degrading to the 95 default is fine; doing it silently undoes the
        # per-tank-cap fix without a word — land the warning in the run log.
        print(f"WARN: facility config unreadable ({type(e).__name__}: {e}) — "
              f"density KPI/heatmap judged against the {growout_cap:g} kg/m³ "
              f"default cap")

    # R8 needs the purge/production boundary, which lives in Control. If
    # Control is unreadable, fall back to PURGE: that reproduces the previous
    # behaviour (6N never flagged) instead of inventing breaches out of a
    # missing config. The facility-config warning above already tells the
    # operator when the density judgement is degraded.
    from forecast.tiers import effective_density_cap as _eff_cap_kpi
    _ctl_kpi = None
    try:
        from forecast.config_io import load_control
        _ctl_kpi = load_control(CONFIG_DIR)
    except Exception as e:  # noqa: BLE001
        print(f"WARN: control config unreadable ({type(e).__name__}: {e}) — "
              f"6N density judged as PURGE (exempt) for the whole horizon")

    def _purge_kpi(ctl, when):
        if ctl is None or when is None:
            return True
        try:
            from forecast.sixn import is_purge_mode
            return is_purge_mode(ctl, when)
        except Exception:  # noqa: BLE001
            return True
    # Density violations from BatchLocations (header at row 4).
    violations = []
    bl_rows = []
    if "BatchLocations" in wb.sheetnames:
        ws = wb["BatchLocations"]
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i < 5 or not row or row[0] is None:
                continue
            wk, ws_d, bid, tid, sys_id, count, avg_wt, biomass, density = row[:9]
            # Stage (col 10) drives R8: a STARVE tank is exempt wherever
            # it sits, so the alert needs it, not just the system name.
            stage = row[9] if len(row) > 9 else ""
            bl_rows.append({
                "Week": wk, "Batch": bid, "Tank": tid, "System": sys_id,
                "Count": count, "AvgWt_kg": avg_wt, "Biomass_kg": biomass,
                "Density_kg_m3": density,
            })
            # Density alert judged through R8 (tiers.effective_density_cap) --
            # ONE rule, shared with the engine, run.py's audit and the
            # advisory panel. This parse feeds EVERY pipeline's KPI
            # (controller + global), so a wrong rule here mis-scores all of
            # them at once. It replaced `sys_id != "OG6N"`, which both hid
            # real breaches once 6N runs as a production system and counted
            # in-place harvest-prep tanks outside 6N as violations.
            _cap = _eff_cap_kpi(tank_caps.get(tid, growout_cap), sys_id or "",
                                stage, _purge_kpi(_ctl_kpi, ws_d))
            if (isinstance(density, (int, float)) and _cap > 0
                    and _cap != float("inf") and density > _cap):
                violations.append(density)

    # BiologyProjection — per (batch, week) explicit mortality % + cull
    # events. BatchLocations only shows END-OF-WEEK count (mortality
    # implicitly applied); this sheet has the per-week mortality fraction
    # and cull counts/biomass so we can chart "weekly losses" at scale.
    bio_rows = []
    if "BiologyProjection" in wb.sheetnames:
        ws = wb["BiologyProjection"]
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if header is None and row and row[0] == "Batch":
                header = list(row)
                idx = {h: j for j, h in enumerate(header)}
                continue
            if header is None or not row or row[0] is None:
                continue
            def g(name, default=None):
                j = idx.get(name)
                if j is None or j >= len(row):
                    return default
                return row[j]
            bid = g("Batch")
            wk = g("Week")
            stage = g("Stage")
            count = g("Count")
            mort_pct_wk = g("Mortality_pct_wk", 0) or 0
            cull_count = g("Cull_Count", 0) or 0
            if not isinstance(count, (int, float)):
                continue
            # Weekly mortality count = current count * mort_pct/100
            # (approximation; the geometric daily compound is close).
            mort_count = float(count) * float(mort_pct_wk) / 100.0
            bio_rows.append({
                "Batch": bid, "Week": wk, "Stage": stage,
                "Count": float(count),
                "Mortality_count_wk": mort_count,
                "Cull_count_wk": float(cull_count),
            })

    # Harvest events from HarvestPlan. Single-table format (title rows 1-3,
    # header row 4, data rows from row 5):
    #   Week | Batch | Tank | Count (fish) | Gross_AvgWt (kg) | Gross_Biomass (kg)
    #   | HOG_Yield | HOG_AvgWt | HOG_Biomass
    # Data rows are identified by an ISO week label in col 0 ("2026-W23").
    harvest_kg = 0.0
    harvest_count = 0
    harvest_events = []
    if "HarvestPlan" in wb.sheetnames:
        ws = wb["HarvestPlan"]
        for row in ws.iter_rows(values_only=True):
            if not row or not isinstance(row[0], str) or "-W" not in row[0]:
                continue
            if len(row) < 6:
                continue
            cnt = row[3]
            gross_kg = row[5]
            gross_avg_kg = row[4]
            hog_kg = row[8] if len(row) > 8 and isinstance(row[8], (int, float)) else None
            hog_avg_kg = row[7] if len(row) > 7 and isinstance(row[7], (int, float)) else None
            if isinstance(cnt, (int, float)) and isinstance(gross_kg, (int, float)):
                harvest_count += cnt
                harvest_kg += gross_kg
                harvest_events.append({
                    "Week": row[0], "Batch": row[1], "Tank": row[2],
                    "Count": cnt, "Gross_kg": gross_kg,
                    "Avg_wt_kg": gross_avg_kg,
                    "HOG_kg": hog_kg if hog_kg is not None else 0.0,
                    "HOG_avg_kg": hog_avg_kg,
                })

    # FACILITY-WIDE weekly biomass + feed, straight from the Advisory sheet.
    # This is the FW-INCLUSIVE basis the engine's cap actually governs (audit
    # H2/M3): FW/EGG fish are real facility biomass but live in FW tanks, so
    # they never appear in BatchLocations. Charts that summed tank rows and
    # drew the facility cap line beside them understated the load by the FW
    # share — which peaks near 7% of cap, so a chart reading 98% was really
    # ~105%. Read the number the cap is enforced against instead.
    facility_weekly = []
    if "Advisory" in wb.sheetnames:
        ws = wb["Advisory"]
        hdr = None
        for r in ws.iter_rows(values_only=True):
            if hdr is None:
                if r and r[0] == "Week" and any(
                        str(c).startswith("Total_Biomass") for c in r if c):
                    hdr = {str(c): i for i, c in enumerate(r) if c}
                continue
            if not r or not r[0] or not str(r[0]).startswith("20"):
                continue

            def _num(key, _row=r, _h=hdr):
                k = next((c for c in _h if c.startswith(key)), None)
                v = _row[_h[k]] if k is not None else None
                return float(v) if isinstance(v, (int, float)) else None

            facility_weekly.append({
                "Week": str(r[0]),
                "Total_Biomass_kg": _num("Total_Biomass"),
                "Biomass_Limit_kg": _num("Biomass_Limit"),
                "Total_Feed_kg_day": _num("Total_Feed"),
                "Feed_Limit_kg_day": _num("Feed_Limit"),
            })

    # Validation warnings now live in ValidationLog ("# | Category | Detail",
    # data rows have a numeric '#'); the Advisory sheet is the per-week capacity
    # table. Build the issues-by-category summary + detail list from ValidationLog.
    advisory_entries = []
    advisory_summary = []
    if "ValidationLog" in wb.sheetnames:
        ws = wb["ValidationLog"]
        cat_counts: dict[str, int] = defaultdict(int)
        for r in ws.iter_rows(values_only=True):
            if not r or not isinstance(r[0], (int, float)):
                continue  # title/header rows have no numeric '#'
            cat = str(r[1]) if len(r) > 1 and r[1] else ""
            det = str(r[2]) if len(r) > 2 and r[2] else ""
            advisory_entries.append({"#": int(r[0]), "Category": cat, "Detail": det})
            cat_counts[cat] += 1
        advisory_summary = [{"Category": c, "Count": n}
                            for c, n in sorted(cat_counts.items(), key=lambda x: -x[1])]

    # Yearly summary (facility-wide per-year rollup) for the app's yearly trends.
    yearly = []
    if "YearlySummary" in wb.sheetnames:
        ws = wb["YearlySummary"]
        hdr = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Year":
                hdr = [str(c) for c in row if c is not None]
                continue
            if hdr and row and isinstance(row[0], (int, float)):
                yearly.append({hdr[i]: row[i] for i in range(min(len(hdr), len(row)))})

    # TransferTemplate: Section A = the canonical production-flow template (the
    # seawater journey every batch follows), Section B = per-batch plan summary.
    # Parse both for the Plan tab.
    plan_summary = []
    flow_template = []
    if "TransferTemplate" in wb.sheetnames:
        ws = wb["TransferTemplate"]
        hdr = None            # Section B (per-batch) header
        flow_hdr = None       # Section A (production-flow) header
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Stage":          # Section A header row
                flow_hdr = [str(c) if c is not None else "" for c in row]
                continue
            if row and row[0] == "Batch":          # Section B header row
                hdr = [str(c) for c in row if c is not None]
                continue
            if (flow_hdr and row and isinstance(row[0], str)
                    and row[0][:1].isdigit()):     # Section A stage rows ("1. …")
                flow_template.append({flow_hdr[i]: row[i]
                                      for i in range(min(len(flow_hdr), len(row)))
                                      if flow_hdr[i]})
            if (hdr and row and isinstance(row[0], str)
                    and row[0].startswith("B") and len(row[0]) > 1 and row[0][1].isdigit()):
                plan_summary.append({hdr[i]: row[i] for i in range(min(len(hdr), len(row)))})

    # Control status (R8-R16).
    status = {}
    if "Control" in wb.sheetnames:
        ws = wb["Control"]
        labels = {
            8: "status", 9: "timestamp", 10: "scenario",
            11: "forecast_start", 12: "horizon", 13: "batches",
            14: "og_tanks", 15: "elapsed", 16: "warnings",
        }
        for r, k in labels.items():
            status[k] = ws.cell(row=r, column=2).value

    from forecast.optimize import _density_quality, WELFARE_DENSITY_KG_M3
    _wl = WELFARE_DENSITY_KG_M3
    _wl_note = None
    try:                                    # operator's welfare line from Configure
        from forecast.config_io import load_control
        # `or default`: an unset/zero line means "the default 80", the SAME
        # resolution every other surface (board, optimizer, frontier) uses —
        # a 0 here would mark ALL biomass crowded on this KPI only.
        _wl = float(load_control(CONFIG_DIR).density_welfare_threshold_kg_m3
                    or WELFARE_DENSITY_KG_M3)
    except Exception as _e:  # noqa: BLE001
        # Was `pass`: the KPI then judged density against a line the operator
        # never set, with nothing on screen to say so. A fallback is fine; a
        # SILENT one is not (its sibling read a few lines up already WARNs).
        _wl_note = (f"welfare density line unreadable "
                    f"({type(_e).__name__}) — judged against the "
                    f"{WELFARE_DENSITY_KG_M3:.0f} kg/m³ default")
    _q_mean, _q_fw, _q_frac = _density_quality(wb, _wl)
    return {
        "violations": len(violations),
        "worst_density": max(violations, default=0.0),
        "growout_density_cap": growout_cap,
        "mean_rearing_density": _q_mean,
        "crowded_biomass_fraction": _q_frac,
        "crowded_fish_weeks": _q_fw,
        "welfare_density": _wl,
        "welfare_density_note": _wl_note,
        "system_biomass_cap": sys_cap_biomass,
        "harvest_kg": harvest_kg,
        "harvest_count": harvest_count,
        "batch_locations": bl_rows,
        "harvest_events": harvest_events,
        "biology_projection": bio_rows,
        "advisory_summary": advisory_summary,
        "advisory_entries": advisory_entries,
        "facility_weekly": facility_weekly,
        "control_status": status,
        "yearly": yearly,
        "plan_summary": plan_summary,
        "flow_template": flow_template,
    }


# ============================================================
# Stocking-for-quality frontier — called from the Analyze board
# ============================================================

def _stocking_frontier_section():
    """Stocking-for-quality frontier: sweep a stocking CUT across FUTURE batches
    and plot the quality-vs-volume trade — the lever that works when the density
    knobs can't (a tank-full facility). Engine = forecast.stocking_frontier."""
    import plotly.graph_objects as pgo
    st.divider()
    st.subheader("🌿 Stocking-for-quality frontier")
    st.caption(
        "When the facility is **tank-full** the density knobs above can't lower "
        "density — the real quality lever is stocking **fewer fish**. This sweeps a "
        "stocking cut across your **future** batches (fish already in the facility "
        "are fixed) and shows the trade: fewer fish rear **gentler** (lower "
        "experienced density) but yield **less harvest**. Each point runs the full "
        "forecast (~20s); your config/scenario/PR are never touched.")
    _fd = st.radio("Frontier depth", ["Quick (0, 10%)",
                                      "Full (0, 5, 10, 15, 20%)"],
                   horizontal=True, key="frontier_depth")
    reductions = ((0.0, 0.10) if _fd.startswith("Quick")
                  else (0.0, 0.05, 0.10, 0.15, 0.20))
    if st.button(f"▶ Run stocking frontier ({len(reductions)} runs, "
                 f"~{len(reductions) * 20 // 60 + 1} min)", key="frontier_go"):
        from forecast.config_io import load_control
        from forecast.stocking_frontier import stocking_frontier
        _wl = 80.0
        try:
            # `or 80`: unset/zero resolves to the default, same as every surface.
            _wl = float(load_control(CONFIG_DIR).density_welfare_threshold_kg_m3
                        or 80.0)
        except Exception:  # noqa: BLE001
            pass
        import shutil
        _wd = tempfile.mkdtemp(prefix="as_frontier_")   # per-run dir: no cross-session clobber
        _tmp = Path(_wd) / "frontier_pr.xlsm"
        _tmp.write_bytes(uploaded.getvalue())
        try:
            with st.spinner(f"Running {len(reductions)} forecasts…"):
                st.session_state["frontier_pts"] = stocking_frontier(
                    str(_tmp), str(CONFIG_DIR), str(SCENARIO_DIR),
                    reductions=reductions, welfare_density=_wl)
                st.session_state["_frontier_sig"] = _sweep_inputs_sig()
        finally:
            shutil.rmtree(_wd, ignore_errors=True)

    pts = st.session_state.get("frontier_pts")
    if not pts:
        return
    _warn_if_sweep_stale("_frontier_sig", "frontier points")
    for p in pts:
        if p.error:
            st.caption(f"⚠ {p.reduction * 100:.0f}% cut failed: {p.error}")
    ok = [p for p in pts if not p.error and p.conserves]
    if not ok:
        st.warning("No valid frontier points (all failed or broke conservation).")
        return
    df = pd.DataFrame([{
        "Stocking cut": f"{p.reduction * 100:.0f}%",
        "Future batches cut": p.scaled_batches,
        "Harvest (t)": round(p.harvest_t),
        "Reared density (kg/m³)": round(p.mean_rearing_density, 1),
        "% crowded": round(p.crowded_biomass_fraction * 100, 1),
        "Worst density": round(p.worst_density),
    } for p in ok])
    st.dataframe(df, hide_index=True, use_container_width=True)
    fig = pgo.Figure()
    fig.add_trace(pgo.Scatter(
        x=[p.harvest_t for p in ok], y=[p.mean_rearing_density for p in ok],
        mode="lines+markers+text",
        text=[f"{p.reduction * 100:.0f}%" for p in ok], textposition="top center",
        line=dict(color="#2e7d32")))
    fig.update_layout(
        height=380, title="Quality vs volume — each point is a stocking cut",
        xaxis_title="Harvest volume (t) — more is better →",
        yaxis_title="Reared density (kg/m³) — lower is gentler ↓")
    st.plotly_chart(fig, use_container_width=True)
    base = next((p for p in ok if p.reduction == 0), None)
    if base is None:
        # The 0% run failed/broke conservation — the smallest surviving cut
        # stands in, and the captions must SAY so: silently labelling deltas
        # "vs 0%" against a non-zero base misstates every number.
        base = ok[0]
        st.caption(f"⚠ The 0% (no-cut) baseline failed — deltas below are vs "
                   f"the **{base.reduction * 100:.0f}% cut**, not vs today's "
                   f"stocking.")
    for p in ok:
        if p is not base and p.reduction > base.reduction:
            dq = base.mean_rearing_density - p.mean_rearing_density
            dv = base.harvest_t - p.harvest_t
            st.caption(
                f"**{p.reduction * 100:.0f}% fewer future fish** → reared "
                f"**{dq:+.1f} kg/m³ gentler** "
                f"({base.crowded_biomass_fraction * 100:.0f}% → "
                f"{p.crowded_biomass_fraction * 100:.0f}% crowded), for "
                f"**{dv:,.0f} t less harvest**.")


# ============================================================
# Optimizer helpers
# ============================================================

def _opt_winner(results, rec):
    """The variant the recommendation actually chose.

    Resolve by OVERRIDES, not by label: coordinate_descent names each candidate
    for the single knob it changed that step, so the same label recurs across
    rounds carrying different accumulated overrides. A by-label lookup returns
    the earliest match — a round-1 partial set — which then gets run and (with
    auto-save on) written to control.yaml in place of the winning combination.
    Falls back to the old behaviour for a recommendation without overrides.
    """
    ov = dict(getattr(rec, "overrides", None) or {})
    for v in results:                      # exact winner: label AND knobs
        if v.label == rec.best_label and dict(v.overrides) == ov:
            return v
    for v in results:                      # knobs alone still identify it
        if dict(v.overrides) == ov:
            return v
    return next((v for v in results if v.label == rec.best_label), results[0])


def _opt_table(results) -> pd.DataFrame:
    rows = []
    for v in results:
        m = v.metrics
        failed = bool(getattr(v, "failed", None))
        rows.append({
            "Variant": v.label,
            "Score": None if failed else round(v.score, 3),
            "Sys_over-cap": None if failed else round(m.system_overshoot, 3),
            "Density_over-cap": None if failed else round(m.density_overshoot, 3),
            "Biomass_overshoot": None if failed else round(m.biomass_overshoot, 3),
            "Biomass_var": None if failed else round(m.biomass_var, 3),
            "Util_gap": None if failed else round(m.biomass_util_gap, 3),
            "Harvest_var": None if failed else round(m.harvest_var, 3),
            "Harvest_overshoot": None if failed else round(m.harvest_overshoot, 3),
            "Feed_load": None if failed else round(m.feed_load),
            "Transfers/fish": None if failed else round(m.transfers_per_fish, 2),
            "Wks_over_limit": None if failed else m.weeks_over_harvest_cap,
            "Conservation": (f"INFEASIBLE — {v.failed[:90]}" if failed
                             else "OK" if v.conservation_ok
                             else f"FAIL ({v.dropped}/{v.overprod})"),
        })
    return pd.DataFrame(rows)


# System -> conveyor tier, for the per-batch journey ("how it got there").
_BATCH_TIER = {
    "OG1N": "Nursery (OG1/2)", "OG1S": "Nursery (OG1/2)",
    "OG2N": "Nursery (OG1/2)", "OG2S": "Nursery (OG1/2)",
    "OG3N": "Grow-out OG3", "OG3S": "Grow-out OG3",
    "OG4N": "Grow-out OG4", "OG4S": "Grow-out OG4",
    "OG5N": "Grow-out OG5", "OG5S": "Grow-out OG5",
    "OG6S": "Finishing OG6", "OG6N": "Finishing/depuration OG6N",
}
_BATCH_TIER_ORDER = ["Nursery (OG1/2)", "Grow-out OG3", "Grow-out OG4",
                     "Grow-out OG5", "Finishing OG6", "Finishing/depuration OG6N"]


def _derive_batch_plans(bl_df, he_df):
    """Per-batch journey from BatchLocations (+ HarvestPlan): a summary header plus
    the milestone timeline (the tier transitions each batch makes, when, at what
    weight/tanks, through to harvest). Pure derivation from data already in the
    output workbook — the 'where each batch is + how it got there' traceability."""
    plans = []
    if bl_df is None or bl_df.empty:
        return plans
    bl = bl_df.copy()
    bl["Tier"] = bl["System"].map(lambda s: _BATCH_TIER.get(str(s), str(s)))
    bl["Week"] = bl["Week"].astype(str)
    hv = {}
    if he_df is not None and not he_df.empty:
        for b, g in he_df.groupby("Batch"):
            wks = sorted(str(w) for w in g["Week"])
            hv[str(b)] = {"first": wks[0], "last": wks[-1],
                          "hog_t": float(g["HOG_kg"].sum()) / 1000.0,
                          "avg_wt": float(pd.to_numeric(g["Avg_wt_kg"], errors="coerce").mean())}
    for b, g in bl.groupby("Batch"):
        g = g.sort_values("Week")
        weeks = list(dict.fromkeys(g["Week"]))
        peak_tanks = int(g.groupby("Week")["Tank"].nunique().max())
        milestones, seen = [], set()
        for wk in weeks:
            gw = g[g["Week"] == wk]
            here = set(gw["Tier"])
            for tier in _BATCH_TIER_ORDER:
                if tier in here and tier not in seen:
                    seen.add(tier)
                    sub = gw[gw["Tier"] == tier]
                    avgwt = pd.to_numeric(sub["AvgWt_kg"], errors="coerce").mean()
                    if not milestones:   # first appearance: real entry vs in-flight
                        ev = ("Seawater entry (TranOG)"
                              if pd.notna(avgwt) and avgwt < 0.6
                              else "In-flight at forecast start")
                    else:
                        ev = f"→ {tier}"
                    milestones.append({
                        "Week": wk, "Event": ev,
                        "Systems": ", ".join(sorted(set(str(s) for s in sub["System"]))),
                        "AvgWt (kg)": round(float(avgwt), 2) if pd.notna(avgwt) else None,
                        "Tanks": int(sub["Tank"].nunique()),
                    })
        h = hv.get(str(b))
        if h:
            milestones.append({"Week": f"{h['first']}–{h['last']}", "Event": "Harvest",
                               "Systems": "→ harvest",
                               "AvgWt (kg)": round(h["avg_wt"], 2), "Tanks": None})
        plans.append({"Batch": str(b), "SW_entry": weeks[0] if weeks else "—",
                      "Peak_tanks": peak_tanks,
                      "Harvest_window": (f"{h['first']}–{h['last']}" if h else "—"),
                      "HOG_t": round(h["hog_t"], 0) if h else 0.0,
                      "milestones": milestones})
    plans.sort(key=lambda p: p["SW_entry"])
    return plans


def _harvest_mode_label(config_dir) -> str:
    """Short description of the harvest controller mode in a config dir, so the
    Results view can always say WHICH run is on screen (keeping the correct data)."""
    import yaml
    try:
        c = yaml.safe_load(open(Path(config_dir) / "control.yaml"))
    except Exception:  # noqa: BLE001
        return "current app config"
    if c.get("harvest_level_load"):
        return f"level-load ON (K={c.get('harvest_smooth_lookahead_weeks')})"
    return "level-load OFF (default controller)"


def _rv_memo(name: str, rid: str, build):
    """Per-run memo for the results view.

    st.tabs renders ALL tab bodies on every rerun, so without this every widget
    tick anywhere in the app (including the manual-window editor above the
    results) rebuilt the pivots, groupbys and derived tables and re-opened the
    output workbook from disk. Holds exactly one run's artifacts: a different
    `rid` clears the lot, so stale data can't survive a new run.

    Values are returned BY REFERENCE — callers must .copy() before mutating,
    which the tab code already does throughout.
    """
    cache = st.session_state.setdefault("_rv_cache", {})
    if cache.get("_rid") != rid:
        cache.clear()
        cache["_rid"] = rid
    if name not in cache:
        cache[name] = build()
    return cache[name]


def _system_feed_audit(out_path):
    """REALIZED per-system feed rows + PER-SYSTEM caps, from the output
    workbook's SystemLimitsAudit sheet.

    Returns (rows, {system: cap_kg_day}). The cap used to be a single scalar
    overwritten on every row, so it ended up holding whatever the LAST audit row
    happened to say — one cap line drawn for every system, wrong for all but one
    of them whenever the per-system caps differ (which is the normal case: OG1/2
    and OG3-6 have very different feed capacity).

    Returns ([], {}) when the file is gone: the output lives in a temp dir that
    the OS can sweep mid-session, and this read previously took the whole
    results view down with it."""
    from pathlib import Path as _P
    if not out_path or not _P(out_path).exists():
        return [], {}
    rows, caps = [], {}
    try:
        _wb = load_workbook(out_path, data_only=True, read_only=True)
        try:
            if "SystemLimitsAudit" in _wb.sheetnames:
                _hdr = None
                for _row in _wb["SystemLimitsAudit"].iter_rows(values_only=True):
                    if _hdr is None:
                        if _row and _row[0] == "Week":
                            _hdr = {h: j for j, h in enumerate(_row)}
                        continue
                    if not _row or _row[0] is None:
                        continue
                    _sys = _row[_hdr.get("System", 1)]
                    _fd = _row[_hdr.get("Feed_kg_day", 5)]
                    _fc = _row[_hdr.get("Feed_cap", 6)]
                    if _sys is None or _fd is None:
                        continue
                    rows.append({"System": str(_sys),
                                 "Week": str(_row[_hdr.get("Week", 0)]),
                                 "Feed_kg_day": float(_fd)})
                    if _fc:
                        caps[str(_sys)] = float(_fc)
        finally:
            _wb.close()   # was outside any guard: an exception mid-scan leaked it
    except Exception:  # noqa: BLE001 — an unreadable audit just hides the chart
        return [], {}
    return rows, caps


def _transfer_plan_rows(out_path) -> list[dict]:
    """TransferPlan sheet -> list of dicts, generically by its header row —
    every engine writes the same single-table layout (Week | Batch | Type |
    From_Tank | To_Tank | Count | Avg_Weight). Missing sheet/file -> []."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(out_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — file gone -> empty, caller says so
        return []
    try:
        if "TransferPlan" not in wb.sheetnames:
            return []
        hdr = None
        out = []
        for row in wb["TransferPlan"].iter_rows(values_only=True):
            if hdr is None:
                if (row and str(row[0]).strip() == "Week"
                        and any(str(c).strip() == "Batch" for c in row if c)):
                    hdr = [str(c).strip() if c else "" for c in row]
                continue
            if not row or not str(row[0]).startswith("20"):
                continue
            out.append({hdr[i]: row[i]
                        for i in range(min(len(hdr), len(row))) if hdr[i]})
        return out
    finally:
        wb.close()


def _daily_harvest_table(he_df):
    """Per-day (Mon–Fri) breakout of the WEEK's total harvest for the Harvest tab.

    All tanks harvesting in the same ISO week are COMBINED into one block: their
    count + biomass are summed and split evenly across the five operating days,
    with blended average weights (total biomass ÷ total fish), a **Total** row,
    and a blank row before the next week. The Tank/Batch columns list every tank
    and batch that contributed. Returns (DataFrame, {total-row positions},
    {blank-row positions}) so the caller can shade the totals."""
    import datetime as _dt
    cols = ["Week", "Date", "Tank", "Batch", "Count", "Live kg",
            "Avg live (kg)", "HOG kg", "Avg HOG (kg)"]
    rows: list[dict] = []
    total_pos, blank_pos = set(), set()
    if he_df.empty or "Week" not in he_df.columns:
        return pd.DataFrame(rows, columns=cols), total_pos, blank_pos
    df = he_df.copy()
    df["Week"] = df["Week"].astype(str)

    def _num(s):
        return float(pd.to_numeric(s, errors="coerce").fillna(0).sum())

    for wk in sorted(df["Week"].unique()):
        sub = df[df["Week"] == wk]
        try:
            y, w = int(wk[:4]), int(wk[6:8])
            days = [_dt.date.fromisocalendar(y, w, 1) + _dt.timedelta(days=i)
                    for i in range(5)]
        except Exception:  # noqa: BLE001 — a non-week label just gets skipped
            continue
        cnt = _num(sub["Count"])
        gross = _num(sub["Gross_kg"]) if "Gross_kg" in sub else 0.0
        hog = _num(sub["HOG_kg"]) if "HOG_kg" in sub else 0.0
        live_avg = gross / cnt if cnt else 0.0     # blended live kg/fish
        hog_avg = hog / cnt if cnt else 0.0         # blended HOG kg/fish
        tanks = (", ".join(str(int(t)) if float(t).is_integer() else str(t)
                           for t in sorted(sub["Tank"].dropna().unique()))
                 if "Tank" in sub else "")
        bats = (", ".join(sorted(sub["Batch"].dropna().astype(str).unique()))
                if "Batch" in sub else "")
        n = len(days)
        for d in days:
            rows.append({
                "Week": wk, "Date": d.strftime("%Y-%m-%d"),
                "Tank": tanks, "Batch": bats,
                "Count": f"{round(cnt / n):,}", "Live kg": f"{round(gross / n):,}",
                "Avg live (kg)": f"{live_avg:.2f}",
                "HOG kg": f"{round(hog / n):,}", "Avg HOG (kg)": f"{hog_avg:.2f}"})
        total_pos.add(len(rows))
        rows.append({
            "Week": wk, "Date": "Total", "Tank": tanks, "Batch": bats,
            "Count": f"{round(cnt):,}", "Live kg": f"{round(gross):,}",
            "Avg live (kg)": f"{live_avg:.2f}", "HOG kg": f"{round(hog):,}",
            "Avg HOG (kg)": f"{hog_avg:.2f}"})
        blank_pos.add(len(rows))
        rows.append({c: "" for c in cols})
    return pd.DataFrame(rows, columns=cols), total_pos, blank_pos


def _quick_viz(r):
    """Inline visualization of a forecast result (used in the Optimize tab so the
    optimized run can be visualized without switching modes). Charts the harvest-
    flattening + facility biomass straight from the parsed result."""
    he = pd.DataFrame(r.get("harvest_events") or [])
    bl = pd.DataFrame(r.get("batch_locations") or [])
    if not he.empty and "Week" in he and "Count" in he:
        hw = (he.groupby("Week", as_index=False)["Count"].sum()
                .sort_values("Week"))
        fig = px.bar(hw, x="Week", y="Count", title="Harvest — fish per week")
        _hv_cap = float((r.get("config_used") or {}).get("max_harvest_per_week")
                        or 55000)
        fig.add_hline(y=_hv_cap, line_dash="dot",
                      annotation_text=f"{_hv_cap / 1000:,.0f}k cap")
        fig.update_layout(height=340, xaxis_title="", yaxis_title="fish")
        st.plotly_chart(fig, use_container_width=True)
    # WHOLE-FACILITY biomass (OG + 6N + freshwater) — the basis the cap is
    # enforced on. Summing tank rows omits the FW phase entirely (~7% of cap at
    # peak), which drew a line the plan appeared to sit comfortably under while
    # the engine was at or over it.
    _fw = pd.DataFrame(r.get("facility_weekly") or [])
    if not _fw.empty and _fw["Total_Biomass_kg"].notna().any():
        _fw = _fw.dropna(subset=["Total_Biomass_kg"]).sort_values("Week")
        _fw["Biomass_t"] = _fw["Total_Biomass_kg"] / 1000.0
        fig2 = px.line(_fw, x="Week", y="Biomass_t",
                       title="Facility biomass per week (t) — whole facility, "
                             "incl. freshwater")
        _lim = _fw["Biomass_Limit_kg"]
        if _lim.dropna().nunique() > 1:
            # The cap is a per-week limit (FacilityLimits overrides) — a single
            # hline at week 1's value misreads any ramp across the horizon.
            fig2.add_scatter(x=_fw["Week"], y=_lim / 1000.0, mode="lines",
                             line={"dash": "dot", "color": "red"},
                             name="cap (per-week)")
        else:
            _cap_t = (float(_lim.dropna().iloc[0]) if _lim.notna().any() else
                      float((r.get("config_used") or {}).get("max_biomass_kg")
                            or 3_800_000)) / 1000.0
            fig2.add_hline(y=_cap_t, line_dash="dot",
                           annotation_text=f"{_cap_t / 1000:.2f}M cap")
        fig2.update_layout(height=340, xaxis_title="", yaxis_title="tonnes")
        st.plotly_chart(fig2, use_container_width=True)


def _optimizer():
    st.header("🧭 Optimize — multi-objective")
    st.caption(
        "Sweeps the controller knobs and ranks variants on a **selectable** "
        "objective. The goal is to **walk the line**: hold biomass and harvest "
        "close to their limits AND flat (no lumps, no breaches), with feed and "
        "handling minimized — gated on conservation. The transfer/density trade "
        "is real (relieving density adds transfers), so you pick the emphasis; "
        "nothing is auto-decided. Conservation-failing variants are rejected. "
        "The winner is **not** whatever scores best: three emphasis-independent "
        "guards run first — never an empty harvest week, never a week above the "
        "relief ceiling, and never a worse leanest week than the baseline "
        "config. They exist because the score itself is blind to those "
        "(the floor has no term at all, and the relief band is protected by a "
        "single weight that one shipped preset sets to 0). If a guard excludes "
        "the top-scoring variant — or has to stand down because nothing cleared "
        "it — the recommendation below says so by name, in amber."
    )

    _cfg_ok = _config_ready() and _scenario_ready()
    _pr_ok = pr is not None and pr["ok"]
    if not _cfg_ok:
        st.info("No config yet — set it up in **Configure** first.")
        return
    if not _pr_ok:
        st.info("Upload a valid **ProductionReport** in the sidebar first.")
        return

    from forecast.config_io import load_control, control_to_dict
    _ok_bc, _base_cd = _read_or_explain(
        lambda: control_to_dict(load_control(CONFIG_DIR)), "config/control.yaml")
    if not _ok_bc:
        return
    _render_active_config(
        _base_cd, "ℹ️ Base configuration — the search tunes knobs ON TOP of this")

    # Optimize tunes the controller-family pipeline (the live config's engine).
    # If the plan picked on the Compare board is a GLOBAL engine, knobs found
    # here were never measured on it — say so instead of letting a save look
    # like it was validated for the chosen plan.
    _ch = st.session_state.get("_chosen_method", _DEFAULT_METHOD)
    _chm = _METHODS.get(_ch)
    if _chm is not None and _chm.engine == "global":
        st.warning(
            f"Your picked plan is **{_chm.label}** (a Global engine). Optimize "
            f"sweeps and validates the **controller-family** engine, so a "
            f"recommendation saved here was not measured on your picked plan — "
            f"re-run **Compare & Choose** after saving to see its effect there.")

    _hist = optimize.read_run_log(n=15)
    if _hist:
        with st.expander(f"📜 Recent auto-optimize runs ({len(_hist)}) — settings used + results",
                         expanded=False):
            _rows = []
            for h in reversed(_hist):   # newest first
                mt = h.get("metrics", {}) or {}
                kb = h.get("winning_knobs") or {}
                _rows.append({
                    "When": h.get("ts", ""),
                    # This is the SEARCH method (grid/deep), not a planning
                    # engine — label it so it can't be misread as one.
                    "Search": h.get("method", ""),
                    "Emphasis": h.get("emphasis", ""),
                    "Winning knobs": ", ".join(f"{k}={v}" for k, v in kb.items()) or "(baseline)",
                    "Hot spot": mt.get("system_peak"),
                    "Feed": mt.get("feed_load"),
                    "Wks>limit": mt.get("weeks_over_harvest_cap"),
                    "Saved": "✓" if h.get("saved_to_config") else "",
                    "Dropped": h.get("dropped"),
                })
            st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True)
            st.caption("Each Auto-optimize run is logged to `optimize_history.jsonl` — "
                       "the settings used and what it produced, kept across sessions.")

    emphasis = st.radio("Objective emphasis", list(optimize.EMPHASIS_PRESETS.keys()),
                        horizontal=True,
                        help="What 'best' should mean when ranking the tried "
                             "settings — each preset weighs the soft goals "
                             "differently (steady biomass, less feed strain, "
                             "fewer transfers, gentler crowding). Hard rules "
                             "(contracts, caps, conservation) always come "
                             "first. Re-scoring is instant — change this after "
                             "a sweep without re-running.")
    custom = None
    with st.expander("Advanced: custom weights"):
        st.caption("Override the preset (all 'less is better'). 0 drops a component.")
        base = optimize.EMPHASIS_PRESETS[emphasis]
        cols = st.columns(4)
        custom = {}
        for i, comp in enumerate(optimize.COMPONENTS):
            with cols[i % 4]:
                custom[comp] = st.number_input(comp, min_value=0.0, max_value=10.0,
                                               value=float(base.get(comp, 0.0)), step=0.5,
                                               key=f"w_{comp}")
        if st.checkbox("Use custom weights", key="use_custom_w"):
            pass
        else:
            custom = None

    method = st.radio(
        "Search method",
        ["Quick grid", "Full grid", "Deep search (finds combos)",
         "Grid + Deep (best of both)"],
        horizontal=True,
        help="TWO search algorithms, offered as FOUR choices. GRID (Quick = 2 configs, "
             "Full = 23) enumerates a hand-picked list — fast and broad, but mostly one "
             "knob at a time, so it misses COMBINATIONS. DEEP SEARCH is a coordinate "
             "descent that tunes one knob at a time and FINDS combinations (~15–30 runs). "
             "GRID + DEEP runs the full grid, then deep-searches FROM the grid's best and "
             "keeps the global best of both — grid explores, descent exploits (most "
             "thorough, ~30–45 runs; what Auto-optimize uses). The emphasis above guides "
             "deep/combined.")
    combined = method.startswith("Grid +")
    deep = method.startswith("Deep")
    if combined:
        st.write(f"**Grid + Deep** — the full grid, then coordinate descent from its "
                 f"best, guided by the **{emphasis}** emphasis (~30–45 runs). Returns "
                 f"the best of both methods. Config never modified.")
    elif deep:
        st.write(f"**Deep search** — greedy local search guided by the **{emphasis}** "
                 f"emphasis (~15–30 runs). Finds knob COMBINATIONS. Config never modified.")
    else:
        grid = optimize.opt_grid_for(method == "Quick grid")
        n = len(grid)
        st.write(f"**{method}** — runs the forecast **{n} times** "
                 f"(~{n * 90 // 60}–{max(1, n * 100 // 60)} min). Config never modified.")
    _c1, _c2 = st.columns(2)
    _run_opt = _c1.button("▶ Run optimization", type="primary", use_container_width=True)
    _auto_opt = _c2.button(
        "🤖 Auto-optimize & run", use_container_width=True,
        help="One click: find the best config (this method + emphasis), then run the "
             "FULL forecast with it and load it into the tabs. The winning knobs are "
             "validated TOGETHER, so it's safe to apply them as a set.")
    _auto_save = st.checkbox(
        "When auto-optimizing, save the winning knobs to config", value=True,
        key="auto_save_cfg",
        help="Writes the best knobs into config/control.yaml so future normal runs use "
             "them. Uncheck to just produce the optimized run without changing config.")
    if _run_opt or _auto_opt:
        work = Path(tempfile.mkdtemp(prefix="as_opt_in_"))
        in_path = work / (uploaded.name or "input.xlsm")
        in_path.write_bytes(uploaded.getvalue())
        bar = st.progress(0.0, text="Starting…")
        _prog = lambda i, m, label: bar.progress(  # noqa: E731
            min(0.98, i / m) if m else min(0.95, i / 40.0),
            text=f"[{i}{'/' + str(m) if m else ''}] {label} …")
        _w = optimize.weights_for(emphasis, custom)
        _vc = _variant_cache()
        try:
            if combined:
                results = optimize.deep_search_combined(
                    str(in_path), str(CONFIG_DIR), str(SCENARIO_DIR),
                    emphasis=emphasis, weights=_w, progress=_prog,
                    max_workers=_cpu_workers(), variant_cache=_vc)
            elif deep:
                results = optimize.coordinate_descent(
                    str(in_path), str(CONFIG_DIR), str(SCENARIO_DIR),
                    emphasis=emphasis, weights=_w, progress=_prog,
                    max_workers=_cpu_workers(), variant_cache=_vc)
            else:
                results = optimize.sweep(
                    str(in_path), str(CONFIG_DIR), str(SCENARIO_DIR), grid=grid,
                    progress=_prog, max_workers=_cpu_workers(), variant_cache=_vc)
        except Exception as e:  # noqa: BLE001
            bar.empty()
            st.error(f"Optimization failed: {e}")
            st.code(traceback.format_exc())
            return
        bar.progress(1.0, text="Done")
        st.session_state["_opt_results"] = results
        st.session_state["_opt_sig"] = _sweep_inputs_sig()
        if _auto_opt:
            # AUTO: pick the validated best, run the FULL forecast with it, load it
            # into the viz tabs, and (optionally) persist the winning knobs.
            _rec0 = optimize.recommend(results, emphasis=emphasis, weights=_w)
            _best0 = _opt_winner(results, _rec0)
            _knobs = optimize.overrides_yaml(_best0.overrides).replace("\n", " · ") or "baseline"
            with st.spinner(f"Auto-optimize — running the full forecast with {_knobs} …"):
                try:
                    _tmpcfg = optimize.config_dir_with_overrides(str(CONFIG_DIR), _best0.overrides)
                    # "as-configured": the SAME engine the sweep measured (the
                    # live config + winning knobs, no method pins) — the default
                    # "controller" method pins hybrid_follow off and would load
                    # a different engine's plan into the tabs than was scored.
                    _res = _run_with_workbook_bytes(
                        uploaded.getvalue(), uploaded.name,
                        config_dir=_tmpcfg, scenario_dir=str(SCENARIO_DIR),
                        method="as-configured")
                except Exception as e:  # noqa: BLE001
                    st.error(f"Auto-optimize run failed: {e}")
                    _res = {"ok": False}
            if _res.get("ok"):
                _res["_run_label"] = "Auto-optimized — " + _knobs
                st.session_state.result = _res
                _saved = bool(_auto_save and _best0.overrides)
                if _saved:
                    optimize.save_overrides_to_config(str(CONFIG_DIR), _best0.overrides)
                    _clear_all_editor_state()
                    # The save moved the config fingerprint — refresh the sig
                    # so only EXTERNAL changes flag the results stale.
                    st.session_state["_opt_sig"] = _sweep_inputs_sig()
                # Log this run (settings + results) to optimize_history.jsonl so
                # there's a durable record of what was run and what it produced.
                from datetime import datetime as _dt
                optimize.append_run_log(optimize.make_run_record(
                    _best0, method, emphasis,
                    ts=_dt.now().isoformat(timespec="seconds"),
                    saved=_saved, source="auto-optimize (app)"))
                st.success(
                    f"🤖 Auto-optimized → **{_best0.label}** ({_knobs})"
                    + (" · **saved to config**" if _saved else " · config unchanged")
                    + ". Loaded into the **Run forecast** tabs. *(logged to history)*")
                # Auto-optimize can WRITE to config, so never let a guard
                # decision pass silently here.
                for _gn in (getattr(_rec0, "guard_notes", None) or []):
                    st.warning(_gn)

    results = st.session_state.get("_opt_results")
    if not results:
        return
    _warn_if_sweep_stale("_opt_sig", "optimization results")

    # Some variants may be INFEASIBLE on this PR (the engine refused to plan them —
    # e.g. a TranOG arrival with no free tanks). They're excluded from selection but
    # kept in the table; surface how many so the operator sees the search wasn't clean.
    _infeasible = [v for v in results if getattr(v, "failed", None)]
    if _infeasible:
        _lbls = ", ".join(v.label for v in _infeasible[:6])
        st.warning(
            f"⚠ **{len(_infeasible)} of {len(results)} variants were infeasible** on "
            f"this ProductionReport (excluded from selection): {_lbls}"
            f"{' …' if len(_infeasible) > 6 else ''}. This means the tighter/fuller "
            f"settings can't physically place every TranOG arrival here — a "
            f"capacity limit (add 6N depuration, raise the biomass cap, or re-time "
            f"the TranOG schedule), not a tuning problem. See the table for each "
            f"reason.")

    # Re-score instantly against the currently selected emphasis (no re-run).
    rec = optimize.recommend(results, emphasis=emphasis,
                             weights=optimize.weights_for(emphasis, custom))
    _no_feasible = rec.best_label == "(none)"
    if _no_feasible:
        st.error(
            f"**No feasible variant** on this ProductionReport — {rec.text} Every "
            "setting hit the engine's capacity guard (each reason is in the table "
            "below). This is a capacity limit, not a tuning problem: add 6N "
            "depuration capacity, raise the facility biomass cap, or re-time the "
            "TranOG arrival schedule, then re-run.")
    elif rec.is_capacity_bound:
        st.warning(f"**Capacity-bound:** {rec.text}")
    elif getattr(rec, "guard_notes", None):
        # A winner-eligibility guard excluded the best-scoring candidate, or
        # stood down because nothing cleared it. Never green — the operator
        # must read WHY before applying or saving these knobs.
        st.warning(f"**Recommendation:** {rec.text}")
    else:
        st.success(f"**Recommendation:** {rec.text}")

    # ---- Feed the recommendation back into the forecast ----
    # When nothing is feasible, `best` is an infeasible variant — the Apply panel's
    # Save button is hidden (no overrides) and its Run button surfaces the same
    # capacity error, so it's survivable; the error above + the table make the
    # situation clear.
    best = _opt_winner(results, rec)
    with st.container(border=True):
        st.markdown(f"**Apply & verify — `{best.label}`**")
        st.caption("These are control-knob overrides — the same knobs a normal run "
                   "reads. Paste them into Configure → Control to keep them, or run a "
                   "full forecast now to verify the optimized result end-to-end.")
        st.code(optimize.overrides_yaml(best.overrides) or "# baseline", language="yaml")
        _bcol1, _bcol2 = st.columns(2)
        with _bcol2:
            if best.overrides and st.button("💾 Save these knobs to my config",
                                            key="opt_save", use_container_width=True):
                optimize.save_overrides_to_config(str(CONFIG_DIR), best.overrides)
                _clear_all_editor_state()
                # The save moved the config fingerprint — refresh the sig so
                # only EXTERNAL changes flag the results stale.
                st.session_state["_opt_sig"] = _sweep_inputs_sig()
                st.success("Saved to config — **Run forecast** and future runs now "
                           "use these knobs (no longer baseline).")
        _run_clicked_opt = _bcol1.button("▶ Run full forecast with these knobs",
                                         key="opt_apply_run", use_container_width=True)
        if _run_clicked_opt:
            with st.spinner("Running the full pipeline with the recommended knobs…"):
                try:
                    # Run via the SAME path as Run-forecast mode, against a temp
                    # config with the knobs applied — so the result populates the
                    # full visualization tabs AND the download, not just metrics.
                    tmpcfg = optimize.config_dir_with_overrides(str(CONFIG_DIR), best.overrides)
                    # "as-configured" — match the sweep's engine exactly (see
                    # the auto-optimize call above for why).
                    result = _run_with_workbook_bytes(
                        uploaded.getvalue(), uploaded.name,
                        config_dir=tmpcfg, scenario_dir=str(SCENARIO_DIR),
                        method="as-configured")
                    if result.get("ok"):
                        result["_run_label"] = ("Optimized — "
                            + optimize.overrides_yaml(best.overrides).replace("\n", " · "))
                        st.session_state.result = result  # lights up Run-forecast tabs
                        # welfare_density: without it this one path scored
                        # crowding against the 80.0 module default while the
                        # rest of the app uses the configured welfare line.
                        m, dropped, overprod = optimize.metrics_from_workbook(
                            result["output_path"],
                            optimize._harvest_cap(str(CONFIG_DIR), best.overrides),
                            welfare_density=optimize._welfare_density(
                                str(CONFIG_DIR), best.overrides))
                        st.session_state["_opt_run"] = {
                            "dropped": dropped, "overprod": overprod,
                            "cv": m.harvest_var, "over": m.weeks_over_harvest_cap,
                            # Bind these metrics to the run they describe. Without
                            # it the panel survives a later ordinary Run-forecast,
                            # pairing THESE numbers with THAT workbook and still
                            # offering it as "Forecast_optimized.xlsm".
                            "rid": _result_rid(result),
                        }
                    else:
                        st.error(f"Run failed: {result.get('error', 'unknown')}")
                except Exception as e:  # noqa: BLE001
                    st.error(f"Run failed: {e}")
                    st.code(traceback.format_exc())
        run_out = st.session_state.get("_opt_run")
        _cur = st.session_state.get("result") or {}
        # Only render while the loaded result IS the run these metrics came from.
        # A later plain Run-forecast replaces `result` but leaves `_opt_run`
        # behind, and the panel would then describe a workbook it never measured.
        if run_out and _cur.get("ok") and run_out.get("rid") == _result_rid(_cur):
            r = st.session_state.result
            ok = run_out["dropped"] == 0 and run_out["overprod"] == 0
            cc1, cc2, cc3 = st.columns(3)
            cc1.metric("Conservation", "PASS ✓" if ok else "FAIL ✗",
                       help=f"{run_out['dropped']} dropped / {run_out['overprod']} over-produced")
            cc2.metric("Harvest CV", f"{run_out['cv']:.3f}",
                       help="How lumpy the weekly harvest is (coefficient of "
                            "variation). Lower = steadier week-to-week; 0 = a "
                            "perfectly flat harvest.")
            _hvlim = _harvest_limit()
            cc3.metric(f"Weeks over {_hvlim:,.0f}", run_out["over"],
                       help=f"Number of weeks whose harvest exceeds the "
                            f"{_hvlim:,.0f}-fish weekly processing limit "
                            f"(max_harvest_per_week) — pressure-relief weeks, "
                            f"acceptable only as rare exceptions. Weeks you "
                            f"scripted yourself in the manual override window "
                            f"are not counted.")
            st.download_button(
                "⬇ Download optimized forecast workbook",
                data=r["output_bytes"], file_name="Forecast_optimized.xlsm",
                mime="application/vnd.ms-excel.sheet.macroenabled.12",
                use_container_width=True)
            # Visualize inline — no mode-switch needed (the harvest chart is the
            # level-loading result you want to see).
            st.markdown("**Visualize this run**")
            _quick_viz(r)
            st.caption("For the full interactive tabs (Per-Batch, Yearly, Plan, "
                       "occupancy heatmap), switch **Mode → Run forecast** — this "
                       "run is already loaded there. Or open the downloaded "
                       "workbook in Excel for every report sheet.")

    df = _opt_table(results).sort_values("Score")

    def _hl(row):
        if row["Variant"] == rec.best_label:
            return ["background-color: #d7f0d7"] * len(row)
        if "FAIL" in str(row["Conservation"]):
            return ["color: #999"] * len(row)
        return [""] * len(row)

    st.dataframe(df.style.apply(_hl, axis=1), use_container_width=True, hide_index=True)

    # ---- Pareto trade-off map: the feed <-> harvest tension across variants ----
    st.subheader("Trade-off map — feed/biomass vs harvest")
    st.caption(
        "Every knob setting plotted by its two competing cap pressures: per-system "
        f"feed/biomass over-cap (x) vs weeks over the {_harvest_limit():,.0f}-fish "
        "weekly processing limit (y). "
        "**Lower-left is best** — both caps held. The lower-left envelope is the "
        "Pareto frontier; your operating point is a choice along it. E.g. "
        "`dev=0.005 (tight)` runs the facility closer to the cap than "
        "`dev=0.02 (loose)`, trading headroom for harvest — so you can SEE the "
        "trade instead of discovering it after a run. (The old worked example "
        "named `tran_og=3`; seawater-entry spread is an operator INPUT, not a "
        "search row, so it can never appear on this plot.)"
    )
    pdf = df.copy()
    pdf["Kind"] = [
        "Recommended" if v == rec.best_label
        else ("Rejected" if "FAIL" in str(c) else "Variant")
        for v, c in zip(pdf["Variant"], pdf["Conservation"])
    ]
    fig = px.scatter(
        pdf, x="Sys_over-cap", y="Wks_over_limit", text="Variant", color="Kind",
        color_discrete_map={"Recommended": "#2e7d32", "Rejected": "#bbbbbb",
                            "Variant": "#1f77b4"},
        title="Operating-point trade-off (lower-left = both caps held)",
    )
    fig.update_traces(textposition="top center", marker=dict(size=11))
    fig.update_layout(height=430,
                      xaxis_title="Per-system feed/biomass over-cap (fraction)",
                      yaxis_title=f"Weeks over the {_harvest_limit():,.0f}-fish weekly limit")
    st.plotly_chart(fig, use_container_width=True)

    best = _opt_winner(results, rec)
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Component scores — recommended")
        comp_df = pd.DataFrame(
            [{"Component": c, "Normalized": round(best.norm.get(c, 0.0), 3)}
             for c in optimize.COMPONENTS])
        fig = px.bar(comp_df, x="Normalized", y="Component", orientation="h",
                     title=f"{best.label} (lower = better)")
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        st.subheader("Transfers by type — recommended")
        tt = best.metrics.transfers_by_type
        st.caption("TranOG is the FW→SW arrival (From_Tank reads 'FW'). "
                   "TRANSFER rows are the real "
                   "tank-to-tank moves — structural progression PLUS whatever "
                   "the rebalancer's budgeted relief passes add — and they "
                   "alone spend the weekly handling budget. Grade legs "
                   "re-split a tank's own fish and are not moves between "
                   "tanks.")
        st.dataframe(pd.DataFrame([{"Type": k, "Fish moved": round(v)} for k, v in tt.items()]),
                     use_container_width=True, hide_index=True)

    # Per-system biomass peak/CV for the recommended variant.
    ps = best.metrics.per_system
    if ps:
        st.subheader("Per-system biomass — recommended")
        psd = pd.DataFrame([
            {"System": s, "Mean_kg": round(d["mean"]), "Peak_kg": round(d["peak"]),
             "CV": round(d["cv"], 3),
             "Peak/cap": round(d["peak"] / d["cap"], 3) if d.get("cap") else None}
            for s, d in sorted(ps.items())])
        st.dataframe(psd, use_container_width=True, hide_index=True)


# ============================================================
# Compare & Choose board — run the methods, grade, pick the plan
# ============================================================

# Grading lenses: (label, getter(res)->value [lower is better], one-line blurb).
_BOARD_LENSES = [
    ("Fewest fish moves", lambda r: r["_score"]["metrics"].transfers_per_fish,
     "least handling / stress"),
    ("Steadiest harvest", lambda r: r["_score"]["metrics"].harvest_var,
     "flattest weekly harvest"),
    ("Most balanced across systems",
     lambda r: r["_score"]["metrics"].between_system.get("bio_cv_mean"),
     "even load system-to-system"),
    ("Most even within systems",
     lambda r: r["_score"]["metrics"].within_system.get("bio_cv_mean"),
     "even load tank-to-tank"),
    ("Tightest density", lambda r: r["_score"]["metrics"].density_peak,
     "most density headroom"),
    ("Best welfare / product quality",
     lambda r: r["_score"]["metrics"].crowded_biomass_fraction,
     "least product reared above the welfare density line"),
    ("Smallest tank footprint", lambda r: r["_score"]["metrics"].tank_footprint_mean,
     "fewest grow-out tanks used"),
    ("Fastest run", lambda r: r.get("elapsed"), "least wall time"),
]


def _board_score(out_path):
    """Metrics + HARD-GATE status for one method's output workbook. Gates are
    pass/fail badges shown on every method; a method that fails Conserves or
    Fully placed is excluded from winning a lens (it lost / never reared fish —
    see _board_lens_pool), the rest are warning flags the operator weighs.
    Reuses the compare lobby's authoritative verdicts."""
    import yaml as _yaml
    from forecast import optimize as _opt
    from tools.run_compare import _conservation_verdict, _harvest_extras
    with open(CONFIG_DIR / "control.yaml") as _f:
        _cfg = _yaml.safe_load(_f) or {}
    hv_cap = float(_cfg.get("max_harvest_per_week", 55000) or 55000)
    _relief = _cfg.get("harvest_relief_pct", 0.10)
    _relief = float(_relief) if _relief is not None else 0.0
    hv_ceiling = hv_cap * (1.0 + _relief) if _relief > 0 else None
    min_hv = float(_cfg.get("min_harvest_per_week", 0) or 0)
    welfare = float(_cfg.get("density_welfare_threshold_kg_m3", 80) or 80)
    mv_cap = int(float(_cfg.get("max_transfers_per_week", 15) or 0)) or None
    m, _dropped, _overprod = _opt.metrics_from_workbook(out_path, hv_cap,
                                                        welfare_density=welfare,
                                                        relief_ceiling=hv_ceiling,
                                                        move_cap=mv_cap,
                                                        min_harvest=min_hv or None)
    verdict = _conservation_verdict(out_path)
    harv = _harvest_extras(out_path, min_hv)
    # "No empty week": the HARD contract rule is "never a NEAR-EMPTY week". A week a
    # little under the floor (a pinned startup week, or 15 fish short of rounding) is
    # not a breach; a crater (e.g. 377 fish) is. Flag only weeks below a quarter of
    # the floor so the badge isolates real craters from benign sub-floor weeks.
    near_empty = 0.25 * min_hv
    min_wk = harv.get("min_week", 0) or 0
    # "Under cap": a method riding its DESIGNED deviation band (~0.5% crest) is at the
    # cap, not over it; only a material overshoot fails. Tolerance = the band + margin.
    dev = float(_cfg.get("facility_biomass_deviation_pct", 0.005) or 0.005)
    cap_tol = 1.0 + max(dev, 0.005) + 0.01
    under_cap = (m.overall_peak_biomass <= m.biomass_cap * cap_tol) if m.biomass_cap else True
    gates = {
        "Conserves": verdict["gate"] != "FAIL",
        "Fully placed": verdict.get("unplaced_batches", 0) == 0,
        "No empty week": (min_wk >= near_empty) if min_hv else True,
        "Under cap": under_cap,
    }
    # R7 lens for the analysis checklist: depuration-era outbound 6N moves.
    try:
        from forecast import analysis as _r7ana
        sixn_out = _r7ana.sixn_outbound_transfers(
            out_path, str(_cfg.get("sixn_production_start") or ""))
    except Exception:                                            # noqa: BLE001
        sixn_out = None
    # The schema stamp versions the grading SEMANTICS (which weeks count,
    # what a gate judges) independently of the engine inputs — see
    # _ensure_board_score / analysis.drop_stale_grades.
    return {"metrics": m, "verdict": verdict, "harvest": harv, "gates": gates,
            "sixn_outbound_purge": sixn_out, "move_cap": mv_cap,
            "schema": _opt.METRICS_SCHEMA}


def _board_badges(gates):
    return "  ".join(f"{'✅' if ok else '⚠️'} {name}" for name, ok in gates.items())


# --------------------------------------------------------------------------- #
# Provenance — every displayed result says where it came from
# --------------------------------------------------------------------------- #
def _fmt_ts_minutes(iso_ts) -> str:
    """'2026-08-10T11:08:03' -> '2026-08-10 11:08'. Junk/None -> ''. Pure."""
    s = str(iso_ts or "")
    if len(s) >= 16 and s[4:5] == "-" and s[7:8] == "-":
        return s[:16].replace("T", " ")
    return ""


def _provenance_line(res, sig: str = "", fresh=None) -> str:
    """One compact caption line saying WHERE a displayed result came from —
    the label that would have made the 2026-08-10 stale-board replay operator-
    visible without pickle-spelunking. Covers the four provenance axes:

      * fresh-run vs cache-replay (`fresh` True/False; None = origin unknown,
        claim nothing rather than guess),
      * the engine-run wall time (res['run_ts'], stamped by
        _run_with_workbook_bytes; older cached legs predate the stamp),
      * the grading-rules version (_score['schema'] = METRICS_SCHEMA) — and,
        when the grade was REDONE from the cached workbook after a rules bump
        (the drop_stale_grades path), it says so with the re-grade time,
      * a short inputs-signature prefix, enough to eyeball "same inputs?"
        across cards.

    Pure (dict in -> str out) so it's testable headlessly."""
    res = res or {}
    ts = _fmt_ts_minutes(res.get("run_ts"))
    if fresh is True:
        head = "● fresh run" + (f" {ts}" if ts else " this session")
    elif fresh is False:
        head = "⟲ cached run" + (f" of {ts}" if ts else " (time not recorded)")
    else:
        head = (f"run {ts}" if ts else "run time not recorded")
    parts = [head]
    schema = (res.get("_score") or {}).get("schema")
    if res.get("_regraded"):
        gts = _fmt_ts_minutes(res.get("_graded_ts"))
        parts.append("re-graded under current rules"
                     + (f" {gts}" if gts else "")
                     + (f" ({schema})" if schema else ""))
    elif schema:
        parts.append(f"graded {schema}")
    if sig:
        parts.append(f"inputs {str(sig)[:8]}")
    return " · ".join(parts)


def _res_is_fresh(res):
    """True = this result's engine ran in THIS browser session; False = it was
    replayed from the cache (disk hydration / an earlier session); None =
    unknown (no rid, or no session runtime). Session-scoped on purpose: the
    operator's question is "did anything just run, or am I looking at a
    replay?"."""
    try:
        rid = (res or {}).get("_rid")
        if not rid:
            return None
        return rid in st.session_state.get("_fresh_rids", set())
    except Exception:  # noqa: BLE001 — headless import has no session
        return None


def _board_lens_pool(scored: dict) -> dict:
    """The methods allowed to WIN a grading lens: must conserve AND be fully
    placed. A PARTIAL plan (fish dropped for lack of space) would otherwise win
    quality lenses on the fish it never reared — unplaced fish can't be crowded,
    moved, or crammed, so every per-fish/per-biomass metric flatters it. Falls
    back to the whole board when nothing passes (cards still render, badges show
    why)."""
    eligible = {k: v for k, v in scored.items()
                if v["_score"]["gates"]["Conserves"]
                and v["_score"]["gates"]["Fully placed"]}
    return eligible or scored


# FULL_ROSTER, not DEFAULT_ROSTER: the Global arms must be PRESENT here for
# the opt-in filter to have anything to add. They are excluded by default
# via _BOARD_OPTIONAL, so an unticked board is still the three controller
# arms -- identical to before, but now the checkbox actually reaches them.
_BOARD_ORDER = tuple(_methods.FULL_ROSTER)


def _board_method_sig(mkey: str, pr_md5: str) -> str:
    """Identity of one board leg's ENGINE inputs, so a finished method can be
    reused instead of re-run. "board3" is a format tag — bump it when the
    stored result shape or the key composition changes (board2→board3 = the
    content-based _config_fingerprint; every mtime-keyed leg ages out once).
    The CP-SAT knobs enter only the method they affect, so moving the Computer
    power slider doesn't needlessly invalidate the fast methods.

    Deliberately EXCLUDES METRICS_SCHEMA: this sig identifies what the ENGINE
    produced (PR + config/scenario + method), while grading is stamped with
    its own schema inside _score and self-invalidates (_ensure_board_score).
    The two axes are independent — a metrics-code bump re-grades the cached
    workbook, it must not force 30-minute engine re-runs."""
    import hashlib
    parts = ["board4", pr_md5, _config_fingerprint(), _engine_fingerprint(), mkey]
    if (_METHODS.get(mkey) or _METHODS["controller"]).engine_kwargs.get("optimal"):
        parts += [f"cpsat{_cpsat_det_time()}", str(_cpu_workers())]
    return hashlib.md5("|".join(parts).encode()).hexdigest()


def _cpsat_det_time() -> float:
    """CP-SAT's per-week DETERMINISTIC work budget — the criterion that actually
    stops each solve. (The wall-clock limit is only a safety cap for a
    pathological week, so tuning it changes nothing.) Higher = tighter layout,
    longer solve.

    Reads the durable copy when the Compare-board slider isn't rendered —
    Streamlit deletes widget-backed keys on any rerun that doesn't draw the
    widget, so without the fallback every other mode (including ▶ Run forecast
    re-running a picked CP-SAT plan) silently reverted to 30.0."""
    return float(st.session_state.get(
        "cpsat_depth", st.session_state.get("_cpsat_depth_saved", 30.0)))


def _restore_output_path(res: dict, tag: str) -> None:
    """A cached result's output_path points into a temp dir the OS may have
    cleaned — regenerate the workbook from the cached bytes so drill-ins and
    re-grading keep working after any amount of downtime."""
    p = res.get("output_path")
    if not p or Path(p).exists() or not res.get("output_bytes"):
        return
    from forecast import analysis as _ana
    d = _ana._default_cache_dir() / "workbooks"
    try:
        d.mkdir(parents=True, exist_ok=True)
        newp = d / f"{tag}__{Path(p).name}"
        newp.write_bytes(res["output_bytes"])
        res["output_path"] = str(newp)
    except OSError:
        pass  # grading degrades gracefully (gates from the stored _score)


def _res_for_disk(res: dict) -> dict:
    """A pickle-safe copy: the _score's Metrics dataclass becomes a plain
    dict. Class instances are tied to a module generation — after a source
    hot-reload they can't be pickled at all ("not the same object as
    forecast.optimize.Metrics", 2026-08-07) — plain data always can."""
    import dataclasses
    sc = res.get("_score")
    if not sc or not dataclasses.is_dataclass(sc.get("metrics")):
        return res
    out = dict(res)
    out["_score"] = {**sc, "metrics": None,
                     "_metrics_plain": dataclasses.asdict(sc["metrics"])}
    return out


def _res_from_disk(res: dict) -> dict:
    """Rebuild the Metrics instance from the CURRENT class. If the stored
    fields no longer fit (schema drift), drop _score — _ensure_board_score
    re-grades from the workbook on demand, so the worst case is a re-grade,
    never a crash or a wrong verdict."""
    import dataclasses
    sc = res.get("_score")
    if not (sc and sc.get("metrics") is None and sc.get("_metrics_plain")):
        return res
    from forecast import optimize as _o
    out = dict(res)
    try:
        fields = {f.name for f in dataclasses.fields(_o.Metrics)}
        m = _o.Metrics(**{k: v for k, v in sc["_metrics_plain"].items()
                          if k in fields})
        out["_score"] = {k: v for k, v in sc.items() if k != "_metrics_plain"}
        out["_score"]["metrics"] = m
    except Exception:  # noqa: BLE001
        out.pop("_score", None)
    return out


def _board_store() -> dict:
    """The per-method finished-run store, hydrated from the DISK cache once
    per session — so a page reload, a frozen tab, or a browser restart never
    loses a finished leg (a CP-SAT leg is 30 minutes of compute). Staleness
    is unchanged: every entry carries its sig and is checked at use."""
    store = st.session_state.setdefault("_board_store", {})
    if not st.session_state.get("_board_cache_hydrated"):
        from forecast import analysis as _ana
        for name, obj in _ana.cache_load_all(prefix="board_").items():
            mkey = name[len("board_"):]
            if obj.get("res"):
                obj = {**obj, "res": _res_from_disk(obj["res"])}
            if store.setdefault(mkey, obj) is obj and obj.get("res"):
                _restore_output_path(obj["res"], mkey)
        st.session_state["_board_cache_hydrated"] = True
    return store


def _board_persist(mkey: str) -> None:
    """Write one method's finished entry through to the disk cache (as plain
    data — see _res_for_disk). Failures are non-fatal (the session copy still
    works) but surfaced."""
    from forecast import analysis as _ana
    entry = st.session_state.get("_board_store", {}).get(mkey)
    if not entry:
        return
    try:
        payload = entry
        if entry.get("res"):
            payload = {**entry, "res": _res_for_disk(entry["res"])}
        _ana.cache_save(f"board_{mkey}", payload)
    except Exception as e:  # noqa: BLE001
        st.caption(f"⚠ couldn't disk-cache {mkey}: {e} (session copy kept)")


def _ensure_board_score(res: dict, label: str) -> None:
    """Grade a finished run, once per METRICS_SCHEMA. The grading is three full
    workbook reads, so it lives inside the stored result — and a run whose
    grading failed transiently can be re-graded on reuse without re-running the
    solve. A grade computed under an OLDER schema (or before schema-stamping
    existed) self-invalidates here: the engine output is reused, the judgement
    is redone from the cached workbook — the 2026-08-10 stale board replayed
    pre-fix zero_weeks verdicts precisely because a stored _score was trusted
    forever."""
    from forecast import analysis as _anacache
    from forecast import optimize as _opt
    _was_stale = _anacache.drop_stale_grades(res, _opt.METRICS_SCHEMA)
    if not (res.get("ok") and res.get("output_path")) or res.get("_score"):
        return
    with st.spinner(f"Grading {label}…"):
        try:
            res["_score"] = _board_score(res["output_path"])
            # Provenance: when this judgement was made, and whether it REPLACED
            # a grade computed under older rules (the drop_stale_grades path) —
            # the card then says "re-graded under current rules", so a reused
            # engine run never passes its verdict off as contemporaneous.
            res["_graded_ts"] = datetime.now().isoformat(timespec="seconds")
            if _was_stale:
                res["_regraded"] = True
            res.pop("_score_err", None)
        except Exception as e:  # noqa: BLE001
            res["_score"] = None
            res["_score_err"] = str(e)


def _compare_and_choose():
    st.header("⚖️ Compare & Choose — run the methods, pick the plan")
    st.caption(
        "Runs the planning methods on your PR, grades them on several lenses, and "
        "lets **you** pick which plan becomes the report — and which method ▶ Run "
        "forecast uses from then on. Each plan is internally consistent (0-drift, "
        "tank continuity) — you choose a whole plan, not a splice. Four badges "
        "(conserves · fully placed · no empty week · under cap) ride on every "
        "method, so a low-transfer plan can't hide a lost batch or a harvest "
        "crater. They are a floor, not the whole rulebook: they say nothing "
        "about the handling budget, the tier rules or the depuration hold, and "
        "the empty-week badge catches craters rather than the weekly contract "
        "floor itself. Read the legend below before picking a **Global** "
        "method, and Analyze's checklist for the rest.")

    _cfg_ok = _config_ready() and _scenario_ready()
    _pr_ok = pr is not None and pr["ok"]
    if not _cfg_ok:
        st.info("No config yet — set it up in **Configure** first.")
        return
    if not _pr_ok:
        st.info("Upload a valid **ProductionReport** in the sidebar first.")
        return

    include_milp = st.checkbox(
        "Include the Global engines (global-lp ~4 min, CP-SAT ~30 min+) — "
        "BENCHMARKS, not runnable plans: both plan the horizon as independent "
        "weekly problems, never read the handling budget, and currently "
        "hard-fail the 6N one-way rule (R7), which disqualifies them",
        value=False, key="board_milp")
    _always = [k for k in _BOARD_ORDER if k not in _BOARD_OPTIONAL]
    st.caption(
        ", ".join(f"{_METHODS[k].label} ({_TYPICAL.get(k, '?')})" for k in _always)
        + " always run. The CP-SAT leg gives EVERY week of your horizon its own "
        "solver budget, so its total runtime scales with the horizon and can run "
        "well past the estimate — uncheck it for a fast compare and add it later, "
        "since finished methods are reused. On a capacity-bound config (facility "
        "full at peak) **Controller + LNS usually matches plain Controller** — "
        "LNS only diverges when there's tank slack to relocate into.")
    if include_milp:
        # value= seeds from the durable copy so leaving Compare mode (which
        # drops the widget key) doesn't snap the depth back to Balanced — that
        # both re-ran picked plans at the wrong budget and falsely marked
        # finished CP-SAT legs stale (the board sig embeds _cpsat_det_time()).
        st.select_slider(
            "CP-SAT solve depth", options=[8.0, 30.0, 60.0],
            value=st.session_state.get("_cpsat_depth_saved", 30.0),
            format_func=lambda v: {8.0: "Quick", 30.0: "Balanced",
                                   60.0: "Thorough"}[v],
            key="cpsat_depth",
            help="Deterministic work budget per week — the criterion that "
                 "actually stops each solve (the wall-clock limit is only a "
                 "safety cap). Quick trades layout tightness for a much shorter "
                 "run; Balanced is the validated default.")
        st.session_state["_cpsat_depth_saved"] = float(
            st.session_state["cpsat_depth"])

    roster = [(k, _METHODS[k].label) for k in _BOARD_ORDER
              if k not in _BOARD_OPTIONAL or include_milp]

    _b1, _b2 = st.columns([3, 2])
    run_all = _b1.button("▶ Run all methods & compare", type="primary",
                         use_container_width=True,
                         help="Reuses any method already finished for these exact "
                              "inputs — only missing or stale methods run.")
    rerun_all = _b2.button("↻ Re-run all from scratch", use_container_width=True,
                           help="Discards finished results and runs every method again.")
    st.caption("Each method is saved the moment it finishes, so an interrupted "
               "compare keeps the legs that completed — click ▶ again to finish "
               "the rest.")

    if run_all or rerun_all:
        import hashlib
        from datetime import datetime as _dtn
        from forecast import analysis as _anacache
        store = _board_store()
        if rerun_all:
            store.clear()
        st.session_state["_board_roster"] = roster
        pr_md5 = hashlib.md5(uploaded.getvalue()).hexdigest()
        n = len(roster)
        bar = st.progress(0.0, text="Starting…")
        for i, (mkey, mlabel) in enumerate(roster):
            msig = _board_method_sig(mkey, pr_md5)
            done = store.get(mkey)
            # A leg whose stored sig doesn't match the CURRENT inputs (or an
            # old-format leg with no sig at all) is absent — re-run, never replay.
            if (_anacache.board_leg_current(done, msig) and done["res"].get("ok")
                    and done["res"].get("output_path")):
                _ensure_board_score(done["res"], mlabel)
                _board_persist(mkey)   # capture a freshly-added _score too
                _rts = _fmt_ts_minutes(done["res"].get("run_ts"))
                bar.progress((i + 1) / n,
                             text=f"{mlabel}: reusing finished result"
                                  + (f" from {_rts}" if _rts else ""))
                continue
            # The bar can only move between methods — the engine call below
            # blocks the script — so say so, and give a clock to judge against.
            bar.progress(i / n, text=(
                f"{i}/{n} finished · running {mlabel} (typically "
                f"{_TYPICAL.get(mkey, '?')}, started {_dtn.now():%H:%M}) — "
                f"this bar next moves when the method finishes"))
            with st.status(f"Running {mlabel}…", expanded=False) as _ms:
                res = _run_with_workbook_bytes(
                    uploaded.getvalue(), uploaded.name,
                    config_dir=str(CONFIG_DIR), scenario_dir=str(SCENARIO_DIR),
                    method=mkey, cpsat_time=300.0,
                    cpsat_det_time=_cpsat_det_time(),
                    cpsat_workers=_cpu_workers(),
                    on_line=lambda ln, _s=_ms, _l=mlabel: _s.update(
                        label=f"{_l} — {ln[:100]}"))
                _ms.update(
                    label=(f"{mlabel} — done in {res.get('elapsed', 0):,.0f}s"
                           if res.get("ok") else f"{mlabel} — failed"),
                    state="complete" if res.get("ok") else "error")
            res["_label"] = mlabel
            _ensure_board_score(res, mlabel)
            store[mkey] = {"sig": msig, "res": res}   # persists NOW, per method
            _board_persist(mkey)                       # ...and to DISK
            bar.progress((i + 1) / n,
                         text=f"✓ {mlabel} done in {res.get('elapsed', 0):,.0f}s "
                              f"({i + 1}/{n})")
        bar.progress(1.0, text=f"All {n} method(s) complete")

    store = _board_store()
    results = {k: store[k]["res"] for k in _BOARD_ORDER if k in store}

    # Results outlive the inputs that produced them: a config save, a scenario
    # edit or a new PR doesn't clear the board. A leg whose stored sig doesn't
    # match the CURRENT inputs is treated as ABSENT — never replayed onto the
    # board (2026-08-10: stale stock legs shown next to fresh tuned runs
    # poisoned the comparison). ▶ re-runs exactly those.
    import hashlib as _hl
    from forecast import analysis as _anacache
    _now_pr = _hl.md5(uploaded.getvalue()).hexdigest()
    _stale = {k for k in results
              if not _anacache.board_leg_current(store[k],
                                                 _board_method_sig(k, _now_pr))}
    if _stale:
        st.warning(f"Inputs changed since {len(_stale)} of these result(s) were "
                   f"computed ({', '.join(results[k].get('_label', k) for k in _stale)}"
                   f") — treating them as not run. **▶ Run all methods & "
                   f"compare** re-runs just those.")
        results = {k: v for k, v in results.items() if k not in _stale}
    if not results:
        return

    _planned = st.session_state.get("_board_roster") or []
    _missing = [lbl for k, lbl in _planned if k not in results]
    if _missing:
        st.warning(f"Partial compare — {len(results)} of {len(_planned)} methods "
                   f"finished. Missing: {', '.join(_missing)}. Click **▶ Run all "
                   f"methods & compare** to run only those.")

    # Grades must match the CURRENT metrics semantics even when nothing re-ran:
    # a valid leg hydrated from disk may carry a _score computed under an older
    # METRICS_SCHEMA — re-grade it from the cached workbook (engine reused) and
    # write the refreshed grade back through to disk.
    for k, v in results.items():
        _sc0 = v.get("_score")
        _ensure_board_score(v, v.get("_label", k))
        if v.get("_score") is not _sc0:
            _board_persist(k)

    scored = {k: v for k, v in results.items() if v.get("ok") and v.get("_score")}
    for k, v in results.items():
        if k not in scored:
            st.error(f"**{v.get('_label', k)}** failed: "
                     f"{v.get('error') or v.get('_score_err') or 'no output produced'}")
    if not scored:
        return

    # ---- Grading-lens cards: who wins each (conservation-passers only) ----
    st.subheader("Grading lenses — who wins each")
    # Values the legend quotes are read from the SAME config the badges were
    # computed against, at render time — the welfare line in this legend was
    # written as "~80" and the config had been retuned to 85.
    _lg_welfare, _lg_floor, _lg_dev = 80.0, 0.0, 0.005
    try:
        from forecast.config_io import load_control as _lc
        _lcc = _lc(CONFIG_DIR)
        _lg_welfare = float(_lcc.density_welfare_threshold_kg_m3 or 80.0)
        _lg_floor = float(_lcc.min_harvest_per_week or 0.0)
        _lg_dev = float(_lcc.facility_biomass_deviation_pct or 0.005)
    except Exception:  # noqa: BLE001 — a legend must never break the page
        pass
    with st.expander("ℹ️ What do the badges, metrics and lenses mean?"):
        st.markdown(
            "**Hard-gate badges** (✅ pass · ⚠️ flag) — the non-negotiables every "
            "plan is judged on first:\n"
            "- **Conserves** — no fish lost or created; mass balance ties out "
            "(0 drift). ⚠️ = the plan lost fish, which disqualifies it from winning "
            "any lens.\n"
            "- **Fully placed** — every batch got tanks; none dropped for lack of "
            "space. ⚠️ also disqualifies from winning a lens: unplaced fish can't "
            "be crowded or moved, so every quality metric flatters the plan.\n"
            "- **No empty week** — no *crater* week. It flags a week only below "
            f"a QUARTER of the weekly contract floor"
            + (f" (under {0.25 * _lg_floor:,.0f} of {_lg_floor:,.0f} fish)"
               if _lg_floor else "")
            + ", which is deliberate: near full utilisation a plan often lands "
            "a little under the floor, and a badge that reddened for that would "
            "redden for everything. So ✅ here does **not** mean the floor was "
            "met every week — that is the contract-floor gate in **Analyze**, "
            "and it is the number to compare candidates on.\n"
            "- **Under cap** — facility biomass stays within its cap plus its "
            f"deviation band (±{_lg_dev * 100:.1f}%) plus a further 1% "
            "measurement margin; ⚠️ = an overshoot past all of that.\n\n"
            "**Per-method metrics** (lower is better on all of these):\n"
            "- **peak % cap** — the single busiest *system-week's* biomass/feed "
            "load vs that system's cap. 100% = right at the cap; over 100% = a "
            "system runs hot that week.\n"
            "- **moves/fish** — fish handled (TranOG arrivals + tank-to-tank "
            "transfers + grading legs) ÷ fish stocked. Lower = less "
            "handling, stress and labour.\n"
            "- **density** — the worst per-tank density (kg/m³) among "
            "tanks still REARING fish. Harvest-prep tanks are excluded, "
            "judged on STAGE (`STARVE`, rule R8), so it covers 6N "
            "depuration AND in-place starvation in an ordinary grow-out "
            "tank. Compare it to that tank's own cap in Configure → "
            "Facility (grow-out tanks ship at 95). Lower = more headroom.\n"
            "- **between-sys CV** — how *evenly* biomass is spread **system-to-"
            "system**. 0 = perfectly balanced; higher = some systems packed while "
            "others sit light.\n"
            "- **within-sys CV** — the same, but **tank-to-tank inside** each "
            "system.\n"
            "- **reared … kg/m³ (…% crowded)** — the **product-quality** view: the "
            "biomass-weighted average density your fish were *reared at*, and the "
            "fraction of grow-out biomass that spent time **above the welfare line** "
            f"(currently **{_lg_welfare:.0f} kg/m³**, set in Configure → Control, "
            "below the hard cap). Lower = gentler rearing = better "
            "welfare / flesh quality — but usually means fewer fish / more tanks.\n\n"
            "**Grading lenses** — each card names the method that's best on one "
            "axis (fewest moves, steadiest harvest, most balanced, tightest "
            "density, best welfare, smallest footprint, fastest). Only methods "
            "that pass **Conserves** and **Fully placed** are eligible to win a "
            "lens. No method wins them all — the board shows the trade-offs so "
            "**you** pick the plan that fits your priority, then press **Use "
            "this plan** (which also becomes the method ▶ Run forecast uses "
            "from then on).\n\n"
            "⚠️ **These four badges are not the whole rulebook.** They do not "
            "check the handling budget, the tier rules R1-R8, or the "
            "depuration hold. That matters most for the two **Global** "
            "methods: they plan the horizon as independent weekly problems, "
            "never read the handling budget, and do not enforce R1, R5 or R7 "
            "while planning — where no legal move exists they emit the move "
            "anyway and log a `TOPOLOGY VIOLATION` row. A Global plan can "
            "therefore win several lenses and still not be executable. Before "
            "adopting one, open its workbook's **ValidationLog** and look for "
            "`TOPOLOGY VIOLATION`, `DEPURATION HOLD` and `PLACEMENT DEGRADED` "
            "rows, and check the handling-budget gate over in **Analyze**. "
            "The Controller methods enforce all of it while planning, so they "
            "are the ones to reach for when you want a plan the crew runs "
            "rather than a benchmark to measure against.")
    pool = _board_lens_pool(scored)
    cols = st.columns(2)
    for i, (label, getter, blurb) in enumerate(_BOARD_LENSES):
        vals = {}
        for k, v in pool.items():
            try:
                x = getter(v)
            except Exception:  # noqa: BLE001
                x = None
            if isinstance(x, (int, float)):
                vals[k] = x
        if not vals:
            continue
        win_k = min(vals, key=vals.get)
        win = scored[win_k]
        with cols[i % 2].container(border=True):
            st.markdown(f"**{label}** — *{blurb}*")
            st.markdown(f"→ **{win['_label']}**  ·  `{vals[win_k]:,.3f}`")
            st.caption(_board_badges(win["_score"]["gates"]))
            st.caption(_provenance_line(
                win, sig=(store.get(win_k) or {}).get("sig", ""),
                fresh=_res_is_fresh(win)))

    # ---- Per-method summary + pick ----
    st.subheader("Pick the plan for your report")
    for k, v in scored.items():
        m = v["_score"]["metrics"]
        with st.container(border=True):
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(f"**{v['_label']}**  ·  {v.get('elapsed', 0):.0f}s")
                st.caption(_board_badges(v["_score"]["gates"]))
                st.caption(_provenance_line(
                    v, sig=(store.get(k) or {}).get("sig", ""),
                    fresh=_res_is_fresh(v)))
                st.caption(
                    f"peak {m.overall_peak_biomass / (m.biomass_cap or 1) * 100:.0f}% cap"
                    f"  ·  {m.transfers_per_fish:.2f} moves/fish"
                    f"  ·  density {m.density_peak:.0f}"
                    f"  ·  between-sys CV {m.between_system.get('bio_cv_mean', 0):.3f}"
                    f"  ·  within-sys CV {m.within_system.get('bio_cv_mean', 0):.3f}"
                    f"  ·  reared {m.mean_rearing_density:.0f} kg/m³ "
                    f"({m.crowded_biomass_fraction * 100:.0f}% crowded)")
            with c2:
                if st.button("Use this plan", key=f"board_pick_{k}",
                             use_container_width=True):
                    # Shallow copy: `v` IS the board's stored entry, so writing
                    # _run_label straight onto it would relabel the board card
                    # too (and any later annotation of the active result would
                    # leak back into the store). The big payloads stay shared.
                    st.session_state.result = {
                        **v, "_run_label": f"Compare & Choose — {v['_label']}"}
                    # This is where the planning method is chosen — ▶ Run
                    # forecast re-runs THIS method from now on.
                    st.session_state["_chosen_method"] = k
                    st.session_state["_goto_run_mode"] = True
                    st.rerun()


# ============================================================
# Analyze — ONE flow: engines → knobs → checklist → a single card
# ============================================================

def _ana_grade(res, targets, econ):
    """Analysis-layer grading for one finished run: gate checklist + target
    review + revenue. Harvest rows are cached on the result dict (keyed by
    rid) so target/price edits re-judge instantly without re-reading the
    workbook — the overlay files are deliberately outside the config
    fingerprint for the same reason."""
    from forecast import analysis as _ana
    from forecast import optimize as _opt
    _schema = _opt.METRICS_SCHEMA
    # Re-validate the stored grade first: a _score/_ana_rows computed under an
    # older METRICS_SCHEMA is dropped and re-derived from the cached workbook
    # (engine output reused, judgement redone) — see analysis.drop_stale_grades.
    _ensure_board_score(res, res.get("_label") or "this run")
    rid = _result_rid(res)
    cached = res.get("_ana_rows")
    if not cached or cached.get("rid") != rid or cached.get("schema") != _schema:
        _rows_err = None
        try:
            rows = (_ana.harvest_rows(res["output_path"])
                    if res.get("output_path") else [])
        except Exception as e:  # noqa: BLE001 — grading must not kill the board
            rows = []
            _rows_err = f"{type(e).__name__}: {e}"
        cached = {"rid": rid, "schema": _schema, "rows": rows, "err": _rows_err}
        res["_ana_rows"] = cached
    if cached.get("err"):
        # Empty-because-unreadable must not display as empty-because-no-harvest:
        # without this the targets gate says "no harvest targets configured"
        # and revenue silently disappears.
        st.caption(f"⚠ {res.get('_label', 'this run')}: HarvestPlan unreadable "
                   f"({cached['err']}) — targets/revenue unavailable for it.")
    rows = cached["rows"]
    tr = None
    if targets:
        monthly, yearly = _ana.harvest_by_period(
            rows, basis=targets.get("basis", "hog"))
        tr = _ana.review_targets(monthly, yearly, targets)
    rev = _ana.revenue_for(rows, econ) if (econ and rows) else None
    dcached = res.get("_ana_density")
    if not dcached or dcached.get("rid") != rid or dcached.get("schema") != _schema:
        dcached = {"rid": rid, "schema": _schema,
                   "review": (_ana.density_review(res["output_path"])
                              if res.get("output_path") else None)}
        res["_ana_density"] = dcached
    dr = dcached["review"]
    # Same rid+schema caching as the density lens: reading Advisory is cheap,
    # but the board re-renders on every widget interaction.
    ccached = res.get("_ana_convergence")
    if not ccached or ccached.get("rid") != rid or ccached.get("schema") != _schema:
        ccached = {"rid": rid, "schema": _schema,
                   "review": (_ana.convergence_review(res["output_path"])
                              if res.get("output_path") else None)}
        res["_ana_convergence"] = ccached
    cvr = ccached["review"]
    sfcached = res.get("_ana_sysfeed")
    if not sfcached or sfcached.get("rid") != rid or sfcached.get("schema") != _schema:
        sfcached = {"rid": rid, "schema": _schema,
                    "review": (_ana.system_feed_review(res["output_path"])
                               if res.get("output_path") else None)}
        res["_ana_sysfeed"] = sfcached
    sfr = sfcached["review"]
    # Fish stuck in 6N purge — same rid+schema caching as the lenses above.
    tpcached = res.get("_ana_trapped")
    if not tpcached or tpcached.get("rid") != rid or tpcached.get("schema") != _schema:
        tpcached = {"rid": rid, "schema": _schema,
                    "review": (_ana.sixn_trapped_review(res["output_path"])
                               if res.get("output_path") else None)}
        res["_ana_trapped"] = tpcached
    tpr = tpcached["review"]
    sc = res.get("_score") or {}
    m = sc.get("metrics")
    v = sc.get("verdict") or {}
    h = sc.get("harvest") or {}
    peak_pct = None
    if m is not None and getattr(m, "biomass_cap", 0):
        peak_pct = m.overall_peak_biomass / m.biomass_cap * 100.0
    ctx = {
        "dropped": v.get("dropped", 0), "overprod": v.get("overprod", 0),
        "zero_weeks": h.get("zero_weeks"),
        "zero_weeks_excluded": h.get("window_weeks_excluded"),
        # The CONTRACT FLOOR, beside the degenerate empty-week rule. Computed
        # since forever by _harvest_extras and written to the RunComparison
        # sheet — but until 2026-08-12 read by no gate and no score component.
        "weeks_below_floor": h.get("weeks_below_min"),
        "min_week": h.get("min_week"),
        "min_harvest": h.get("min_harvest"),
        "weeks_over_harvest_cap": (m.weeks_over_harvest_cap
                                   if m is not None else None),
        "weeks_over_relief_ceiling": (
            getattr(m, "weeks_over_relief_ceiling", None)
            if m is not None else None),
        "sixn_outbound_purge": sc.get("sixn_outbound_purge"),
        "weeks_moves_over_cap": (getattr(m, "weeks_moves_over_cap", None)
                                 if m is not None else None),
        "weeks_moves_warn": (getattr(m, "weeks_moves_warn", None)
                             if m is not None else None),
        "moves_week_max": (getattr(m, "moves_week_max", None)
                           if m is not None else None),
        # The budget the counts above were measured against, so the gate text
        # names the operator's real number instead of a hardcoded 15.
        "move_cap": sc.get("move_cap"),
        "peak_pct_of_cap": peak_pct,
        "targets_review": tr,
        "density_review": dr,
        "convergence": cvr,
        "system_feed": sfr,
        "sixn_trapped": tpr,
    }
    return {"gates": _ana.evaluate_gates(ctx), "targets_review": tr,
            "revenue": rev, "metrics": m, "density_review": dr}


_ANA_ICON = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌", "N/A": "◽"}


def _ana_checklist(gates):
    for g in gates:
        hard = " **(hard rule)**" if g["hard"] else ""
        st.markdown(f"{_ANA_ICON.get(g['status'], '◽')} **{g['label']}**{hard}"
                    f" — {g['detail']}")


def _adoption_refusal(cand, acknowledged: bool):
    """The message to show when an Adopt/Promote write must be REFUSED, or
    None to let it through. Pure — the button handlers call exactly this, so
    the refusal rule is testable without a Streamlit runtime.

    Analyze ranks candidates by gate-failure counts; it does not filter on
    them, and the relief-ceiling gate is SOFT. So the card's winner — and any
    candidate the promote picker can reach — may carry a breach that both
    other doors to config (`tournament.pick_winner`, `optimize.recommend`)
    would refuse. The operator may still overrule that; they may not do it by
    accident."""
    from forecast import analysis as _ana
    br = (cand or {}).get("breaches") or []
    if not _ana.adoption_blocked(br, acknowledged):
        return None
    return ("**Nothing was saved.** This plan fails "
            f"{len(br)} of the adoption checks: " + "; ".join(br)
            + ". Tick the acknowledgement box to save it anyway — what you "
              "accepted is recorded with it.")


def _adoption_gate(cand, sig: str, slot: str, box=None) -> bool:
    """Render the breach panel + the explicit acknowledgement for one adoption
    candidate; return whether a write is confirmed.

    SHAPE, decided deliberately: the button never disappears. Hiding it would
    take a judgement call away from the operator on the one surface that is
    theirs — they can see the plan, its checklist and its provenance, and a
    hard-gate FAIL on a Global leg may be a known modelling gap they mean to
    accept. Leaving it unguarded, though, is how a relief-ceiling breach walks
    into control.yaml unremarked. So: adopt anything you can justify, never
    anything you did not notice.

    The acknowledgement key carries the input signature AND the candidate
    label, so a tick never survives into a different plan or a re-analysis on
    different inputs (state outliving its inputs is a defect class this file
    has paid for twice)."""
    box = box or st
    br = (cand or {}).get("breaches") or []
    if not br:
        return True
    import hashlib as _hl
    box.error(
        f"⛔ **{cand['label']} fails {len(br)} adoption check(s)** — these are "
        f"the same winner-eligibility rules the tuned tournament and Optimize "
        f"apply before they will crown a plan, so neither of them would pick "
        f"this one:\n\n"
        + "\n".join(f"- {b}" for b in br)
        + "\n\nThe list is wider than the checklist above on purpose. It "
          "covers the two HARD gates (conservation, never an empty week), the "
          "**relief ceiling** — whose checklist gate is SOFT, so a breach only "
          "ranks a plan down and would otherwise be adopted in silence — and "
          "the contract floor against this method's own un-tuned run. An entry "
          "reading *never measured* is not a breach but an UNKNOWN, and an "
          "unknown is never read as a pass.\n\n"
          "You can still adopt or promote it on your own judgement; what you "
          "accepted is then recorded with what you save.")
    _k = _hl.md5(f"{sig}|{slot}|{cand['label']}".encode()).hexdigest()[:10]
    return bool(box.checkbox(
        f"I have read the {len(br)} finding(s) above and accept them for "
        f"**{cand['label']}**", key=f"ana_ack_{_k}"))


def _monthly_lever_check(uploaded, targets):
    """The monthly two-way check: run a SHORT set of declared lever legs on THIS
    month's PR and say which holds the contract best — or that nothing beats what
    you already have.

    This is not a search. Four levers were measured across eight starting states
    in August 2026 and all four rejected as defaults, because each helps some
    starting states and hurts others. There is no better default to find; there
    is a cheap per-month answer. See forecast/monthly_check.py for the rules and
    docs/LEVELING_TRADE_2026-08-30.md for the measurements.

    It never writes config. The operator applies a winner by hand in Configure,
    which keeps the decision — and the audit trail — theirs.
    """
    from forecast import monthly_check as _mc
    st.subheader("🗓️ Monthly lever check")
    st.caption(
        "Runs your config against a couple of declared alternatives on THIS "
        "PR and ranks them on the CONSTRAINTS — hard gates, then the weekly "
        "contract floor, then per-system feed, then handling. Score is shown "
        "but never decides. No setting is right for every month: measured "
        "across eight starting states, each lever helped some and hurt others, "
        "so this is a per-month question with a 2-minute answer."
    )
    with st.expander("What it runs, and why each leg is a candidate"):
        for leg in _mc.LEGS:
            ov = ", ".join(f"{k}={v}" for k, v in leg["overrides"].items()) or "—"
            st.markdown(f"**{leg['name']}** (`{ov}`) — {leg['why']}")
        st.caption(_mc.noise_caveat())

    _key = "_monthly_check_res"
    if st.button("Run the monthly check (~2 min)", key="_btn_monthly_check"):
        work = Path(tempfile.mkdtemp(prefix="as_mcheck_"))
        in_path = work / (uploaded.name or "input.xlsm")
        in_path.write_bytes(uploaded.getvalue())
        try:
            from tools.measure_leveling import measure as _measure
        except Exception as e:                                # noqa: BLE001
            st.error(f"Could not load the measurement harness: {e}")
            return
        rows = []
        bar = st.progress(0.0, text="Starting…")
        for i, leg in enumerate(_mc.LEGS):
            bar.progress(i / len(_mc.LEGS), text=f"Running {leg['name']}…")
            try:
                rows.extend(_measure(
                    str(in_path), [leg], str(CONFIG_DIR), str(SCENARIO_DIR),
                    str(work), f"mc{i}", targets))
            except Exception as e:                            # noqa: BLE001
                rows.append({"name": leg["name"], "error": f"{type(e).__name__}: {e}"})
        bar.progress(1.0, text="Done")
        st.session_state[_key] = rows

    rows = st.session_state.get(_key)
    if not rows:
        return
    v = _mc.decide(rows)
    if v.keep_current:
        st.success(
            f"**Keep your current settings.** {v.reason}")
    else:
        st.info(f"**{v.winner}** looks better on this PR — {v.reason}.")
        _ov = next((l["overrides"] for l in _mc.LEGS if l["name"] == v.winner), {})
        st.caption(
            "Nothing has been changed. To use it, set these in **Configure → "
            "Control** yourself, then re-run: "
            + ", ".join(f"`{k} = {val}`" for k, val in _ov.items()))
    st.dataframe(_mc.summary_rows(rows), width="stretch", hide_index=True)
    for note in v.notes:
        st.caption(f"• {note}")
    for name, why in v.disqualified:
        st.warning(f"**{name}** — {why}")
    st.caption(_mc.noise_caveat())


def _decide():
    """One entry point for "which plan should I run?".

    Analyze, Compare & Choose and Optimize answered three views of ONE question,
    and an operator had to know which mode fitted before they could ask it. Two
    of them ran the same engine legs over the same roster writing the same cache
    (app.py Analyze phase 1 vs the Compare run loop), so the split cost a
    decision without buying anything.

    This is deliberately a WRAPPER, not a rewrite. The three functions are
    unchanged and rendered as tabs, so all thirteen capabilities the mode audit
    found living only in Compare and Optimize survive BY CONSTRUCTION — the
    between-system and within-system CV lenses, the tank-footprint and
    fastest-run lenses, the raw density peak, the per-method metric readout,
    setting the standing engine to a NON-winning method, picking a plan while
    writing nothing to config, the force re-run, the ~100-second engine-only
    path, the partial-roster warning, the CP-SAT solve-depth control (which
    Analyze itself reads), and naming a failed engine leg with its error.

    The repo's documented failure mode is retiring a mode and quietly losing
    what only it could do (USER_GUIDE 13.1). A wrapper cannot do that; a deeper
    integration could, and is a separate decision made with the tests in
    tests/test_decide_mode.py holding the line.

    Tab order is the order the question is actually answered: which LEVERS this
    month (cheap, 2 minutes, and most months the answer is "keep what you
    have"), then which ENGINE, then which KNOBS.
    """
    st.header("🧭 Decide — which plan should I run?")
    st.caption(
        "Three views of one question. **Recommend** runs every engine, tunes "
        "the knobs and grades every candidate against the checklist — start "
        "here. **Compare engines** puts the planners side by side on eight "
        "lenses. **Tune knobs** sweeps one engine's settings on an objective "
        "you choose. Finished runs are shared between them; nothing runs twice."
    )
    _tabs = st.tabs(["Recommend a plan", "Compare engines", "Tune knobs"])
    with _tabs[0]:
        _analyze()
    with _tabs[1]:
        _compare_and_choose()
    with _tabs[2]:
        _optimizer()


def _analyze():
    import hashlib
    from datetime import datetime as _dtn
    from forecast import analysis as _ana
    # No header here: _decide() already titled the page, and a second one made
    # the tab read as a separate mode inside a mode. The caption also named
    # Compare & Choose and Optimize as "modes" after they became tabs.
    st.caption(
        "Run every planning engine, tune the knobs on top, judge every "
        "candidate on the **hard-rule checklist**, and recommend ONE plan. "
        "**Most months this tab is the whole answer** — the other two are here "
        "to steer a phase by hand if you disagree with it.")

    _cfg_ok = _config_ready() and _scenario_ready()
    _pr_ok = pr is not None and pr["ok"]
    if not _cfg_ok:
        st.info("No config yet — set it up in **Configure** first.")
        return
    if not _pr_ok:
        st.info("Upload a valid **ProductionReport** in the sidebar first.")
        return

    _ok_t, targets = _read_or_explain(lambda: _ana.load_targets(CONFIG_DIR),
                                      "config/targets.yaml")
    _ok_e, econ = _read_or_explain(lambda: _ana.load_economics(CONFIG_DIR),
                                   "config/economics.yaml")
    if not (_ok_t and _ok_e):
        return

    # The cheap, decisive question first: which LEVERS for this month? It is a
    # 2-minute check, it answers what an operator actually asks monthly, and it
    # is independent of the engine tournament below.
    _monthly_lever_check(uploaded, targets)
    st.divider()
    _t_bits = []
    _t_bits.append(f"🎯 {len((targets or {}).get('monthly', {}))} monthly + "
                   f"{len((targets or {}).get('yearly', {}))} yearly target(s)"
                   if targets else "🎯 no harvest targets set")
    _t_bits.append(f"💰 {len(econ['price_bands'])} price band(s) ({econ['currency']})"
                   if econ else "💰 no prices set")
    st.caption(" · ".join(_t_bits) + " — edit in **Configure → Targets & prices**; "
               "edits re-judge existing results instantly (no re-runs).")

    # ---- ⚡ Quick run — the operator-promoted default ----
    promoted = _ana.load_promoted_default(CONFIG_DIR)
    if promoted:
        with st.container(border=True):
            _pk = ", ".join(f"{k}={v}" for k, v in
                            (promoted.get("overrides") or {}).items()) or "no knob overrides"
            st.markdown(f"**⚡ Promoted default** — `{promoted['method']}` · {_pk} · "
                        f"promoted {promoted.get('promoted_ts', '?')}"
                        + (f" · _{promoted['note']}_" if promoted.get("note") else ""))
            # A default promoted OVER a known breach must say so every time it
            # is offered — the acknowledgement happened once, in a session that
            # is long gone; whoever presses ⚡ next may not be the same reader.
            _pev = promoted.get("evidence")
            _pev = _pev if isinstance(_pev, dict) else {}
            if _pev.get("accepted_with_breach"):
                st.warning(
                    "⚠ This default was promoted WITH accepted rule breach(es): "
                    + "; ".join(str(b) for b in (_pev.get("breaches") or []))
                    + " — a quick run replays exactly that plan.")
            qc1, qc2 = st.columns([1, 3])
            if qc1.button("⚡ Quick run this default", key="ana_quick",
                          help="One run + the checklist — minutes, not the full "
                               "analysis. Uses the promoted method + knobs; "
                               "changes nothing."):
                with st.status("Quick run — promoted default…", expanded=False) as _qs:
                    _qcfg = optimize.config_dir_with_overrides(
                        str(CONFIG_DIR), promoted.get("overrides") or {})
                    _qm = (promoted["method"] if promoted["method"] in _METHODS
                           else "as-configured")
                    qres = _run_with_workbook_bytes(
                        uploaded.getvalue(), uploaded.name, config_dir=_qcfg,
                        scenario_dir=str(SCENARIO_DIR), method=_qm,
                        cpsat_time=300.0, cpsat_det_time=_cpsat_det_time(),
                        cpsat_workers=_cpu_workers(),
                        on_line=lambda ln, _s=_qs: _s.update(label=ln[:100]))
                    _qs.update(state="complete" if qres.get("ok") else "error")
                if qres.get("ok"):
                    qres["_label"] = "Quick run — promoted default"
                    _ensure_board_score(qres, qres["_label"])
                    st.session_state["_ana_quick"] = {
                        "res": qres, "sig": _sweep_inputs_sig()}
                else:
                    st.error(f"Quick run failed: {qres.get('error', 'unknown')}")
            qr = st.session_state.get("_ana_quick")
            if qr and qr["res"].get("ok"):
                if qr.get("sig") != _sweep_inputs_sig():
                    st.warning("This quick run predates a PR/config change — "
                               "re-run it.")
                _qg = _ana_grade(qr["res"], targets, econ)
                st.caption(_provenance_line(qr["res"], sig=qr.get("sig", ""),
                                            fresh=_res_is_fresh(qr["res"])))
                _ana_checklist(_qg["gates"])
                if _qg["revenue"]:
                    _rv = _qg["revenue"]
                    st.caption(f"Revenue ≈ **{_rv['total']:,.0f} {_rv['currency']}**"
                               + (f" · ⚠ {_rv['unpriced_kg']:,.0f} kg unpriced "
                                  f"(outside every band)" if _rv["unpriced_kg"] else ""))
                if qc2.button("Load this run into the Run-forecast tabs",
                              key="ana_quick_load"):
                    st.session_state.result = {
                        **qr["res"], "_run_label": qr["res"]["_label"]}
                    st.session_state["_goto_run_mode"] = True
                    st.rerun()

    # ---- Full analysis ----
    st.subheader("Full analysis")
    emphasis = st.selectbox(
        "What should 'best' mean? (the emphasis for the knob search + final score)",
        list(optimize.EMPHASIS_PRESETS.keys()), key="ana_emph",
        help="Hard rules always come first regardless of emphasis; this weights "
             "the soft objectives (flat biomass, feed, handling, density).")
    include_milp = st.checkbox(
        "Include the Global engines (adds ~4 min + ~30 min; finished legs are "
        "reused) — benchmarks only: they hard-fail R7, so they rank last",
        value=False, key="ana_milp")
    depth = st.radio(
        "Analysis depth",
        ["Quick tournament — engines at stock + knob search on the live config",
         "Tuned tournament — find each method's best knobs, then compare the "
         "TUNED methods"],
        key="ana_depth",
        help="QUICK — run every engine once exactly as your config stands, then "
             "do ONE knob search on the live config. Fast, but it compares one "
             "tuned engine against the rest at stock, so a method can lose "
             "merely for not having been tuned.\n\n"
             "TUNED — the fair version: EVERY method gets its own knob search, "
             "restricted to the knobs that method actually reads, and the board "
             "then compares the methods at their best. A method that already "
             "fails a hard rule at stock gets a cheap one-knob probe first; if "
             "no probed knob clears the failure it is marked GATE-BOUND and the "
             "full search is skipped (honestly — nothing there to find).\n\n"
             "Note on the two Global methods: they have NO tunable knobs at "
             "all, so they compete at stock under either depth. That is "
             "deliberate — the only knobs their code path reads were measured "
             "to break Global's own conservation proof when overridden, so the "
             "registry refuses to put them in a search space. It also means a "
             "Global method that fails a hard rule reads GATE-BOUND with no "
             "probe: there is simply no knob to try.")
    tuned_mode = depth.startswith("Tuned")
    _n_eng = len([k for k in _BOARD_ORDER if k not in _BOARD_OPTIONAL or include_milp])
    if tuned_mode:
        # ---- Honest budget: what pressing go can cost, and what's already paid ----
        from forecast import tournament as _tour
        _brows = []
        for _bk in _BOARD_ORDER:
            if _bk in _BOARD_OPTIONAL and not include_milp:
                continue
            _bm = _METHODS[_bk]
            _bb = _tour.estimate_budget(_bm)
            _bvc = _variant_cache(_bk)
            _bcached = _tour.cached_count(_bvc, _tour.search_grid(_bm))
            _brows.append({
                "Method": _bm.label,
                "Engine run": 1,
                "Probe (only if a hard gate fails)": _bb["probe_if_gate_fails"],
                "Grid": _bb["grid"],
                "Descent (≤)": _bb["descent_max"],
                "Verify": _bb["verify"],
                "Already cached": _bcached,
                "Worst case": _bb["max_total"],
            })
        with st.expander("💰 Run budget — engine runs per method (before you "
                         "press go)", expanded=False):
            st.dataframe(pd.DataFrame(_brows), hide_index=True,
                         use_container_width=True)
            st.caption("Counts are FULL forecast runs (~30–40 s each for the "
                       "controller family, parallelized). 'Probe' only happens "
                       "if the method fails a hard rule at stock; 'Descent' is "
                       "an upper bound — it stops early when a round finds no "
                       "improvement, so the real cost is usually well under "
                       "the worst case. 'Already cached' = grid rows this "
                       "PR+config already measured (keyed on file CONTENT, not "
                       "timestamps), so re-running the tournament is cheap. "
                       "Methods showing 0 grid and 0 descent have no tunable "
                       "knobs and compete at stock — that is the two Global "
                       "methods.")
    go = st.button(
        f"▶ Run {'TUNED tournament' if tuned_mode else 'full analysis'} "
        f"({_n_eng} engines + knob search + checklist)",
        type="primary", key="ana_go")

    if go:
        pr_md5 = hashlib.md5(uploaded.getvalue()).hexdigest()
        roster = [(k, _METHODS[k].label) for k in _BOARD_ORDER
                  if k not in _BOARD_OPTIONAL or include_milp]
        store = _board_store()
        n_phases = (2 * len(roster) + 1) if tuned_mode else (len(roster) + 2)
        bar = st.progress(0.0, text="Phase 1/3 — engine round…")
        # Phase 1: every engine once on the CURRENT config (board legs reused
        # both ways — a leg run here shows up finished on Compare & Choose).
        for i, (mkey, mlabel) in enumerate(roster):
            msig = _board_method_sig(mkey, pr_md5)
            done = store.get(mkey)
            # Sig mismatch / old-format leg = absent: re-run, never replay.
            if (_ana.board_leg_current(done, msig) and done["res"].get("ok")
                    and done["res"].get("output_path")):
                _ensure_board_score(done["res"], mlabel)
                _board_persist(mkey)   # capture a freshly-added _score too
                _rts = _fmt_ts_minutes(done["res"].get("run_ts"))
                bar.progress((i + 1) / n_phases,
                             text=f"{mlabel} — reused ✓"
                                  + (f" (run of {_rts})" if _rts else ""))
                continue
            bar.progress(i / n_phases, text=f"Phase 1/3 — running {mlabel} "
                         f"(typically {_TYPICAL.get(mkey, '?')})…")
            with st.status(f"Running {mlabel}…", expanded=False) as _ms:
                res = _run_with_workbook_bytes(
                    uploaded.getvalue(), uploaded.name,
                    config_dir=str(CONFIG_DIR), scenario_dir=str(SCENARIO_DIR),
                    method=mkey, cpsat_time=300.0,
                    cpsat_det_time=_cpsat_det_time(),
                    cpsat_workers=_cpu_workers(),
                    on_line=lambda ln, _s=_ms, _l=mlabel: _s.update(
                        label=f"{_l} — {ln[:100]}"))
                _ms.update(state="complete" if res.get("ok") else "error")
            res["_label"] = mlabel
            _ensure_board_score(res, mlabel)
            store[mkey] = {"sig": msig, "res": res}
            _board_persist(mkey)
        work = Path(tempfile.mkdtemp(prefix="as_ana_"))
        in_path = work / (uploaded.name or "input.xlsm")
        in_path.write_bytes(uploaded.getvalue())
        _w = optimize.weights_for(emphasis)

        if tuned_mode:
            # Phase 2 (tuned): EVERY method gets its own knob search on its
            # own space — gate-passers the full search, gate-failers the cheap
            # probe (gate-bound when no knob fixes them), knobless methods
            # compete at stock. Phase 3: verify each winner on ITS OWN engine.
            from forecast import tournament as _tour
            tuned_methods = {}
            for j, (mkey, mlabel) in enumerate(roster):
                done = store.get(mkey)
                if not (done and done["res"].get("ok")):
                    tuned_methods[mkey] = {"status": "run-failed",
                                           "overrides": None}
                    continue
                _g = _ana_grade(done["res"], targets, econ)
                _fails = _tour.hard_gate_fails(_g["gates"])
                m = _METHODS[mkey]
                _vc = _variant_cache(mkey)
                _pre = set(_vc.keys())
                bar.progress((len(roster) + j) / n_phases,
                             text=f"Phase 2/3 — tuning {mlabel}…")
                try:
                    tr = _tour.tune_method(
                        m, str(in_path), str(CONFIG_DIR), str(SCENARIO_DIR),
                        emphasis=emphasis, weights=_w,
                        stock_hard_fails=_fails,
                        progress=lambda i, n, label, _j=j, _l=mlabel:
                            bar.progress(
                                min((len(roster) + _j + 0.95) / n_phases,
                                    (len(roster) + _j
                                     + (i / n if n else 0.9)) / n_phases),
                                text=f"Phase 2/3 — tuning {_l} "
                                     f"[{i}{'/' + str(n) if n else ''}] "
                                     f"{label}…"),
                        max_workers=_cpu_workers(), variant_cache=_vc,
                        # Arms the contract-floor no-regression guard: a tuned
                        # winner may never harvest LESS in the leanest week
                        # than this method's own stock run
                        # (forecast.tournament.floor_eligible).
                        stock_min_week=((done["res"].get("_score") or {})
                                        .get("harvest") or {}).get("min_week"))
                except Exception as e:  # noqa: BLE001 — one method must not kill the board
                    st.error(f"Knob search for {mlabel} failed: {e}")
                    st.code(traceback.format_exc())
                    tuned_methods[mkey] = {"status": "search-error",
                                           "overrides": None}
                    continue
                entry = {
                    "status": tr["status"],
                    "overrides": (dict(tr["winner_overrides"])
                                  if tr["winner_overrides"] else None),
                    "stock_hard_fails": list(_fails),
                    # A guard that STOOD DOWN means the winner breaks the rule
                    # it protects. The headless tournament already printed
                    # this; the app dropped it, so the summary showed a tuned
                    # winner with no hint that a rule had to be waived.
                    "ceiling_guard": tr.get("ceiling_guard"),
                    "floor_guard": tr.get("floor_guard"),
                    "n_variants": len(tr["variants"]),
                    "n_cache_reused": sum(
                        1 for v in tr["variants"]
                        if optimize._overrides_key(v.overrides) in _pre),
                    "res": None,
                }
                if tr["status"] == "tuned":
                    winner = dict(tr["winner_overrides"])
                    if winner == dict(m.overrides):
                        # The search's best IS the stock config — the engine
                        # leg already on the board is the tuned candidate.
                        entry["status"] = "stock-best"
                    else:
                        bar.progress((len(roster) + j + 0.95) / n_phases,
                                     text=f"Phase 3/3 — verifying {mlabel} "
                                          f"(tuned)…")
                        _tcfg = optimize.config_dir_with_overrides(
                            str(CONFIG_DIR), winner)
                        vres = _run_with_workbook_bytes(
                            uploaded.getvalue(), uploaded.name,
                            config_dir=_tcfg,
                            scenario_dir=str(SCENARIO_DIR), method=mkey,
                            cpsat_time=300.0,
                            cpsat_det_time=_cpsat_det_time(),
                            cpsat_workers=_cpu_workers())
                        if vres.get("ok"):
                            vres["_label"] = _tour.tuned_label(
                                mlabel, winner, m.overrides)
                            _ensure_board_score(vres, vres["_label"])
                            entry["res"] = vres
                        else:
                            entry["status"] = "verify-failed"
                            st.warning(
                                f"{mlabel}: the tuned winner's verification "
                                f"run failed — competing at stock only. "
                                f"({vres.get('error', 'unknown')})")
                tuned_methods[mkey] = entry
            bar.progress(1.0, text="Tuned tournament complete")
            st.session_state["_ana"] = {
                "sig": _sweep_inputs_sig(), "emphasis": emphasis,
                "made": _dtn.now().isoformat(timespec="seconds"),
                "engine_keys": [k for k, _ in roster],
                "mode": "tuned", "tuned": None,
                "tuned_methods": tuned_methods,
            }
        else:
            # Phase 2 (quick): knob search on the live-config engine (Grid +
            # Deep — what Auto-optimize uses), then verify the winner on the
            # SAME engine.
            bar.progress(len(roster) / n_phases,
                         text="Phase 2/3 — knob search (Grid + Deep)…")
            try:
                opt_results = optimize.deep_search_combined(
                    str(in_path), str(CONFIG_DIR), str(SCENARIO_DIR),
                    emphasis=emphasis, weights=_w,
                    progress=lambda i, m, label: bar.progress(
                        min((len(roster) + 0.9) / n_phases,
                            (len(roster) + (i / m if m else 0.9)) / n_phases),
                        text=f"Phase 2/3 — knob search [{i}"
                             f"{'/' + str(m) if m else ''}] {label}…"),
                    max_workers=_cpu_workers(), variant_cache=_variant_cache())
                _rec = optimize.recommend(opt_results, emphasis=emphasis,
                                          weights=_w)
                _best = _opt_winner(opt_results, _rec)
            except Exception as e:  # noqa: BLE001
                st.error(f"Knob search failed: {e}")
                st.code(traceback.format_exc())
                return
            # The tuned candidate can be Adopted (and its knobs saved), so say
            # out loud when a guard excluded a better-scoring one. Outside the
            # try: a render problem here is not a "knob search failed".
            for _gn in (getattr(_rec, "guard_notes", None) or []):
                st.warning(_gn)
            bar.progress((n_phases - 1) / n_phases,
                         text="Phase 3/3 — verifying the tuned winner…")
            tuned_res = None
            _knob_str = ", ".join(f"{k}={v}" for k, v in
                                  (_best.overrides or {}).items())
            if _best.overrides:
                _tcfg = optimize.config_dir_with_overrides(str(CONFIG_DIR),
                                                           _best.overrides)
                tuned_res = _run_with_workbook_bytes(
                    uploaded.getvalue(), uploaded.name, config_dir=_tcfg,
                    scenario_dir=str(SCENARIO_DIR), method="as-configured")
                if tuned_res.get("ok"):
                    tuned_res["_label"] = f"Tuned config — {_knob_str}"
                    _ensure_board_score(tuned_res, tuned_res["_label"])
                else:
                    st.warning("The tuned winner's verification run failed — "
                               "recommending among the engine round only. "
                               f"({tuned_res.get('error', 'unknown')})")
                    tuned_res = None
            bar.progress(1.0, text="Analysis complete")
            st.session_state["_ana"] = {
                "sig": _sweep_inputs_sig(), "emphasis": emphasis,
                "made": _dtn.now().isoformat(timespec="seconds"),
                "engine_keys": [k for k, _ in roster],
                "mode": "quick",
                "tuned": ({"overrides": dict(_best.overrides), "res": tuned_res}
                          if tuned_res else None),
                "tuned_methods": None,
            }
        # The card survives reloads/frozen tabs too — same disk cache as the
        # engine legs (the 2026-08-06 tab freeze made this non-optional).
        # Both depths persist here: quick's "tuned" res and the tuned
        # tournament's per-method verification runs become plain data.
        try:
            _payload = dict(st.session_state["_ana"])
            if _payload.get("tuned") and _payload["tuned"].get("res"):
                _payload["tuned"] = {**_payload["tuned"],
                                     "res": _res_for_disk(_payload["tuned"]["res"])}
            if _payload.get("tuned_methods"):
                _payload["tuned_methods"] = {
                    k: ({**e, "res": _res_for_disk(e["res"])}
                        if e.get("res") else dict(e))
                    for k, e in _payload["tuned_methods"].items()}
            _ana.cache_save("ana_summary", _payload)
        except Exception as e:  # noqa: BLE001
            st.caption(f"⚠ couldn't disk-cache the analysis: {e}")

    ana = st.session_state.get("_ana")
    if ana is None:
        # New session (reload, new tab, browser restart): restore the last
        # finished analysis from disk — sig-checked below like any other.
        ana = _ana.cache_load_all(prefix="ana_summary").get("ana_summary")
        if ana is not None:
            if ana.get("tuned") and ana["tuned"].get("res"):
                ana["tuned"] = {**ana["tuned"],
                                "res": _res_from_disk(ana["tuned"]["res"])}
                _restore_output_path(ana["tuned"]["res"], "tuned")
            for _tk, _te in (ana.get("tuned_methods") or {}).items():
                if _te.get("res"):
                    _te["res"] = _res_from_disk(_te["res"])
                    _restore_output_path(_te["res"], f"tuned_{_tk}")
            st.session_state["_ana"] = ana
    if not ana:
        st.caption("No analysis yet — click ▶ above. Roughly "
                   f"{'1½–2 h' if include_milp else '45–75 min'} hands-off; "
                   "finished engine legs are reused across re-runs and shared "
                   "with Compare & Choose.")
        return
    if ana["sig"] != _sweep_inputs_sig():
        st.warning("⚠ This analysis was computed on a **different PR or "
                   "config** than what's loaded now — re-run it before "
                   "adopting anything.")

    # ---- Build + grade candidates (targets/prices re-judged live) ----
    store = _board_store()
    cands = []
    # Engine legs join the board ONLY if their stored sig matches the current
    # inputs — a leg run on a pre-edit scenario next to fresh tuned runs is
    # exactly the 2026-08-10 stale-cache poisoning. Mismatched legs are
    # treated as absent (▶ re-runs them), and a reused leg's grade is
    # re-validated against the current METRICS_SCHEMA before it's trusted.
    _pr_now = hashlib.md5(uploaded.getvalue()).hexdigest()
    _skipped_stale = []
    for k in ana["engine_keys"]:
        done = store.get(k)
        if not _ana.board_leg_current(done, _board_method_sig(k, _pr_now)):
            if done:
                _lbl = (done["res"].get("_label", k)
                        if isinstance(done.get("res"), dict) else k)
                _skipped_stale.append(_lbl)
            continue
        _ensure_board_score(done["res"], done["res"].get("_label", k))
        if done["res"].get("ok") and done["res"].get("_score"):
            cands.append({"key": k, "label": done["res"].get("_label", k),
                          "overrides": {}, "res": done["res"],
                          "prov": _provenance_line(
                              done["res"], sig=done.get("sig", ""),
                              fresh=_res_is_fresh(done["res"]))})
        elif done["res"].get("ok"):
            # An engine leg that RAN but couldn't be graded must not vanish
            # without a word (Compare & Choose st.error()s the same state).
            st.warning(f"{done['res'].get('_label', k)}: grading failed "
                       f"({done['res'].get('_score_err', 'unknown')}) — left "
                       f"off the candidate board.")
    if _skipped_stale:
        st.warning(f"{len(_skipped_stale)} stock engine leg(s) were computed on "
                   f"different inputs (PR/config/scenario) and are left OFF the "
                   f"board: {', '.join(_skipped_stale)}. Press ▶ above to re-run "
                   f"just those.")
    if ana.get("tuned") and ana["tuned"]["res"].get("ok"):
        cands.append({"key": "_tuned",
                      "label": ana["tuned"]["res"].get("_label", "Tuned config"),
                      "overrides": ana["tuned"]["overrides"],
                      "res": ana["tuned"]["res"],
                      "prov": _provenance_line(
                          ana["tuned"]["res"], sig=ana.get("sig", ""),
                          fresh=_res_is_fresh(ana["tuned"]["res"]))})
    # Tuned-tournament candidates: each method's tuned winner joins the board
    # as "METHOD (tuned: knobs)". key = the real method key, so Adopt/Promote
    # store a replayable method + overrides pair (Quick run replays BOTH).
    for _tk, _te in (ana.get("tuned_methods") or {}).items():
        _tr = _te.get("res")
        if _tr and _tr.get("ok"):
            # Re-grades from the cached workbook when the stored _score
            # predates the current METRICS_SCHEMA (or didn't survive disk).
            _ensure_board_score(_tr, _tr.get("_label", f"{_tk} (tuned)"))
        if _tr and _tr.get("ok") and _tr.get("_score"):
            cands.append({"key": _tk,
                          "label": _tr.get("_label", f"{_tk} (tuned)"),
                          "overrides": _te.get("overrides") or {},
                          "res": _tr,
                          "prov": _provenance_line(
                              _tr, sig=ana.get("sig", ""),
                              fresh=_res_is_fresh(_tr))})
    if not cands:
        st.error("No graded candidates survived — check the engine round above.")
        return
    for c in cands:
        g = _ana_grade(c["res"], targets, econ)
        c.update(gates=g["gates"], targets_review=g["targets_review"],
                 revenue=g["revenue"], metrics=g["metrics"],
                 density_review=g["density_review"], score=None)
    # Comparable emphasis score across candidates (same scorer as Optimize).
    _w = optimize.weights_for(ana.get("emphasis", "balanced"))
    _variants = [optimize.OptVariant(label=c["label"], overrides=c["overrides"],
                                     metrics=c["metrics"], dropped=0, overprod=0)
                 for c in cands if c["metrics"] is not None]
    if _variants:
        optimize.score_variants(_variants, _w)
        _by_label = {v.label: v.score for v in _variants}
        for c in cands:
            c["score"] = _by_label.get(c["label"])
    # Winner-ELIGIBILITY, the same rules the other two doors to config apply.
    # rank_key below RANKS on gate failures; it never filters on them, and the
    # relief-ceiling gate is soft — so this is what stands between a
    # rule-breaking plan and control.yaml / analysis_defaults.yaml. Computed
    # for EVERY candidate (the promote picker can reach any row), not just the
    # card's winner.
    for c in cands:
        c["breaches"] = _ana.adoption_breaches(
            c, _ana.stock_reference_min_week(c, cands))

    from forecast.analysis import rank_key as _rank_key
    ranked = sorted(cands, key=_rank_key)
    winner, runner = ranked[0], (ranked[1] if len(ranked) > 1 else None)

    # ---- The card ----
    st.divider()
    st.subheader("🏆 Recommended plan")
    st.caption(
        f"Analysis of {ana['made']} · emphasis **{ana.get('emphasis')}** · "
        "pick order: hard-rule FAILs → soft-rule FAILs → total warnings → "
        "target shortfall → emphasis score. Only conservation and "
        "never-an-empty-week are hard; everything else ranks a plan down "
        "without disqualifying it, so **read the checklist before adopting** "
        "— a recommended plan can still carry a red handling or R7 gate. "
        "Adopt/Promote apply the same winner-eligibility rules as the tuned "
        "tournament and Optimize (hard gates, the relief ceiling, the contract "
        "floor): a plan that breaks one can still be saved, but only after you "
        "acknowledge the breach by name, and the breach is saved with it.")
    with st.container(border=True):
        st.markdown(f"### {winner['label']}")
        if winner.get("prov"):
            st.caption(winner["prov"])
        if winner["overrides"]:
            st.code(optimize.overrides_yaml(winner["overrides"]), language="yaml")
        _ana_checklist(winner["gates"])
        if winner["revenue"]:
            _rv = winner["revenue"]
            st.markdown(f"**Revenue ≈ {_rv['total']:,.0f} {_rv['currency']}**"
                        + (f" · ⚠ {_rv['unpriced_kg']:,.0f} kg unpriced"
                           if _rv["unpriced_kg"] else ""))
        if runner is not None:
            st.caption(f"Runner-up: **{runner['label']}** — kept below for "
                       "drill-in; the full table shows every candidate.")
        # The last door to config. One acknowledgement covers both buttons —
        # they save the same plan, so they carry the same breach.
        _card_ok = _adoption_gate(winner, ana.get("sig", ""), "card")

        def _log_adoption(cand, action, method, ok_msg):
            """Durable record of what was accepted, breach included, so a later
            reader isn't left with only the knobs."""
            from datetime import datetime as _dtn3
            _ana.append_adoption_log(
                _ana.adoption_record(
                    cand, ts=_dtn3.now().isoformat(timespec="seconds"),
                    action=action, method=method,
                    overrides=cand.get("overrides"),
                    breaches=cand.get("breaches"),
                    source=f"Analyze ({ana.get('emphasis')}) on {uploaded.name}"),
                str(_ROOT / _ana.DEFAULT_ADOPTION_LOG))
            if cand.get("breaches"):
                st.warning(f"{ok_msg} **with {len(cand['breaches'])} accepted "
                           f"rule breach(es)** — recorded in "
                           f"`{_ana.DEFAULT_ADOPTION_LOG}`.")

        a1, a2 = st.columns(2)
        if a1.button("✅ Adopt this plan", type="primary", key="ana_adopt",
                     help="Saves the winning knobs (if any) to config, makes "
                          "this the method ▶ Run forecast uses, and loads the "
                          "run into the tabs."):
            _refusal = _adoption_refusal(winner, _card_ok)
            if _refusal:
                st.error(_refusal)
            else:
                # The pinned method must be the engine that PRODUCED this
                # plan, or the next Run forecast reproduces a DIFFERENT one.
                # The quick-depth tuned winner is keyed "_tuned" and was
                # verified with method="as-configured" (the live config, no
                # registry pins). Falling back to _DEFAULT_METHOD here pinned
                # controller-hybrid, whose arm pins BOTH hybrid levers ON while
                # the graded plan ran with them OFF (config ships them false) —
                # so the adopted plan silently would not reproduce.
                _m = (winner["key"] if winner["key"] in _METHODS
                      else (_AS_CONFIGURED.key if winner["key"] == "_tuned"
                            else _DEFAULT_METHOD))
                if winner["overrides"]:
                    optimize.save_overrides_to_config(str(CONFIG_DIR),
                                                      winner["overrides"])
                    _clear_all_editor_state()
                # Logged AFTER the write, so a failed write never leaves a
                # record of an adoption that did not happen.
                _log_adoption(winner, "adopt", _m, "Adopted")
                st.session_state["_chosen_method"] = _m
                st.session_state.result = {
                    **winner["res"],
                    "_run_label": (f"Analyze — {winner['label']}"
                                   + (" ⚠ adopted with an accepted rule breach"
                                      if winner.get("breaches") else ""))}
                st.session_state["_goto_run_mode"] = True
                st.rerun()

        def _promote(cand, note_prefix):
            from datetime import datetime as _dtn2
            _gsum = {g["key"]: g["status"] for g in cand["gates"]}
            _br = list(cand.get("breaches") or [])
            # Same rule as Adopt: store the engine that PRODUCED the plan. The
            # quick-depth tuned candidate is keyed "_tuned" and was run
            # as-configured; coercing it to _DEFAULT_METHOD stored
            # controller-hybrid, whose arm pins both hybrid levers ON while the
            # promoted plan ran with them OFF — so ⚡ Quick run replayed a
            # different plan than the one promoted.
            _m = (cand["key"] if cand["key"] in _METHODS
                  else (_AS_CONFIGURED.key if cand["key"] == "_tuned"
                        else _DEFAULT_METHOD))
            _ana.save_promoted_default(
                str(CONFIG_DIR),
                method=_m,
                overrides=cand["overrides"],
                promoted_ts=_dtn2.now().isoformat(timespec="seconds"),
                note=(("⚠ ACCEPTED WITH A KNOWN RULE BREACH — " if _br else "")
                      + f"{note_prefix} on {uploaded.name}"),
                evidence={"gates": _gsum, "score": cand.get("score"),
                          "emphasis": ana.get("emphasis"),
                          # Travels with the promoted default so ⚡ Quick run —
                          # which replays exactly this — can say the plan
                          # carries a flagged rule.
                          "breaches": _br,
                          "accepted_with_breach": bool(_br)})
            _log_adoption(cand, "promote", _m, "Promoted")
            st.success(f"Promoted **{cand['label']}** — the ⚡ Quick run card "
                       "at the top now uses this plan.")

        if a2.button("⭐ Promote as Quick-run default", key="ana_promote",
                     help="Stores method + knobs in "
                          "config/analysis_defaults.yaml, next to the rest of "
                          "your config. It is NOT written into an output "
                          "workbook, so it cannot be lost to a run — but note "
                          "that also means importing config from a workbook "
                          "will not restore it. Promoting changes nothing "
                          "about the current run; it only sets what the ⚡ "
                          "Quick run card at the top of this page will "
                          "re-run. Manual by design: the tool never changes "
                          "its own defaults."):
            _refusal = _adoption_refusal(winner, _card_ok)
            if _refusal:
                st.error(_refusal)
            else:
                _promote(winner, "won analysis")

    # ---- Tuned-tournament summary: what each method's search concluded ----
    if ana.get("tuned_methods"):
        _TM_STATUS = {
            "tuned": "✅ tuned — winner verified on its own engine",
            "stock-best": "✅ stock config already best — the engine leg IS "
                          "the tuned candidate",
            "stock-only": "◽ no tunable knobs — competes at stock (its knobs "
                          "would break its own conservation proof)",
            "gate-bound": "⛔ gate-bound — fails a hard rule and no knob can "
                          "fix it (either no probed knob cleared it, or the "
                          "method has no tunable knobs to probe)",
            "search-failed": "❌ no search variant conserved",
            "verify-failed": "❌ tuned winner's verification run failed",
            "run-failed": "❌ engine run failed",
            "search-error": "❌ knob search crashed",
        }
        with st.expander("🔬 Tuned tournament — per-method search summary",
                         expanded=True):
            _tmrows = []
            for _tk in ana["engine_keys"]:
                _te = (ana.get("tuned_methods") or {}).get(_tk)
                if not _te:
                    continue
                _pins = (_METHODS[_tk].overrides if _tk in _METHODS else {})
                _chosen = {k: v for k, v in (_te.get("overrides") or {}).items()
                           if _pins.get(k) != v}
                _waived = [n for n, _g in
                           (("relief ceiling", _te.get("ceiling_guard")),
                            ("contract floor", _te.get("floor_guard")))
                           if _g == "stood-down"]
                _tmrows.append({
                    "Method": (_METHODS[_tk].label if _tk in _METHODS else _tk),
                    "Outcome": _TM_STATUS.get(_te["status"], _te["status"]),
                    "Winning knobs": (", ".join(f"{k}={v}" for k, v in
                                                sorted(_chosen.items()))
                                      or "—"),
                    "Hard fails at stock": ", ".join(
                        _te.get("stock_hard_fails") or []) or "none",
                    "Guard stood down": ("⚠ " + ", ".join(_waived)
                                         if _waived else "—"),
                    "Variants run": _te.get("n_variants", 0),
                    "From cache": _te.get("n_cache_reused", 0),
                })
            st.dataframe(pd.DataFrame(_tmrows), hide_index=True,
                         use_container_width=True)
            st.caption(
                "Winning knobs exclude the method's own pinned overrides (its "
                "identity). **Guard stood down** = no candidate in that "
                "method's search cleared that rule, so the guard was dropped "
                "rather than return nothing — the winner on that row therefore "
                "**breaks it**: 'relief ceiling' means it plans a week the "
                "plant cannot take, 'contract floor' means its leanest week is "
                "worse than not tuning at all. Treat those rows as findings to "
                "judge, not results to adopt. '—' means every guard held (or "
                "no search ran). 'From cache' = search runs reused from earlier "
                "searches on the same PR + config — re-running the tournament "
                "is cheap.")

    # ---- Full candidate table ----
    st.subheader("All candidates")
    _rows = []
    for c in ranked:
        tr = c.get("targets_review") or {}
        _rows.append({
            "Candidate": c["label"],
            "Gates": " ".join(_ANA_ICON.get(g["status"], "◽")
                              for g in c["gates"]),
            "Hard fails": sum(1 for g in c["gates"]
                              if g["hard"] and g["status"] == "FAIL"),
            "Targets met": (f"{tr.get('met', 0)}/{tr.get('judged', 0)}"
                            if tr.get("judged") else "—"),
            "Shortfall (t)": (round(tr.get("total_shortfall_kg", 0) / 1000.0, 1)
                              if tr.get("judged") else None),
            "Revenue": (f"{c['revenue']['total']:,.0f}" if c.get("revenue")
                        else "—"),
            "Score": (round(c["score"], 3) if c.get("score") is not None
                      else None),
            "Provenance": c.get("prov", ""),
        })
    st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True)
    st.caption("Gate icons in checklist order: " +
               " · ".join(g["label"] for g in ranked[0]["gates"]) +
               ". Scores are the emphasis-weighted objective (lower is "
               "better), comparable across candidates. Provenance: ● = engine "
               "ran this session, ⟲ = replayed from the result cache; "
               "'re-graded under current rules' = the engine output was reused "
               "but its verdict was recomputed after a grading-rules update; "
               "'inputs' = signature prefix of the PR + config the run saw.")
    # Promote ANY candidate, not only the card's winner — the first real
    # analysis promoted the runner-up (the tuned winner was refuted cross-PR),
    # which needed a by-hand YAML write. Now it's a picker.
    _pc1, _pc2 = st.columns([3, 1])
    _pick_lbl = _pc1.selectbox(
        "Promote a different candidate as the Quick-run default",
        [c["label"] for c in ranked], key="ana_promote_pick",
        help="The card's ⭐ promotes the winner; this promotes whichever "
             "candidate YOUR judgment picks (e.g. after cross-PR evidence).")
    # This picker reaches EVERY row, including the ones ranked last precisely
    # because they fail a rule — so it needs the same door as the card.
    _pick_cand = next((c for c in ranked if c["label"] == _pick_lbl), None)
    _pick_ok = (_adoption_gate(_pick_cand, ana.get("sig", ""), "picker", _pc1)
                if _pick_cand is not None else True)
    if _pc2.button("⭐ Promote selected", key="ana_promote_any",
                   use_container_width=True):
        _refusal = (_adoption_refusal(_pick_cand, _pick_ok)
                    if _pick_cand is not None else None)
        if _pick_cand is None:
            st.error("That candidate is no longer on the board — re-pick.")
        elif _refusal:
            st.error(_refusal)
        else:
            _promote(_pick_cand, "operator pick from the candidates table")
    with st.expander("🎯 Target detail — every period, every candidate"):
        for c in ranked:
            tr = c.get("targets_review")
            if not tr or not tr.get("rows"):
                continue
            st.markdown(f"**{c['label']}**")
            st.dataframe(pd.DataFrame([
                {"Period": r["period"], "Target (t)": r["target_kg"] / 1000.0,
                 "Planned (t)": round(r["actual_kg"] / 1000.0, 1),
                 "% of target": (round(r["pct"], 1) if r["pct"] is not None
                                 else None),
                 "Status": r["status"]}
                for r in tr["rows"]]), hide_index=True,
                use_container_width=True)
    with st.expander("📊 Density quality — distribution + severe batches "
                     "(the old Tune readout, per candidate)"):
        st.caption("Each number is a batch's PEAK density as a multiple of its "
                   "tank's cap (1.0× = exactly at cap). Read the "
                   "DISTRIBUTION, not the raw over-cap count: 1.0–1.1× is "
                   "normal at full utilisation. Only **severe (≥1.3×)** "
                   "batches matter, and they are what the gate counts — the "
                   "table below lists everything from **1.2×** up so you can "
                   "see what is approaching severe. If the severe ones cluster "
                   "in time and peak mid-grow-out it's a **stocking/capacity** "
                   "problem (use the stocking frontier below), not a knob to "
                   "re-tune.")
        for c in ranked:
            dr = c.get("density_review")
            if not dr:
                continue
            st.markdown(f"**{c['label']}** — {dr['n']} batches · worst "
                        f"{dr['worst']:.2f}× · median {dr['median']:.2f}×")
            st.dataframe(pd.DataFrame([dr["buckets"]]), hide_index=True,
                         use_container_width=True)
            if dr["severe_rows"]:
                st.dataframe(pd.DataFrame(dr["severe_rows"]), hide_index=True,
                             use_container_width=True)
            else:
                st.caption("No batch over 1.2× cap in this plan.")

    # The stocking-for-quality frontier — the REMEDY for what the density lens
    # diagnoses (severe cluster = stock fewer fish, no knob fixes it). Moved
    # here from the retired Tune mode; runs on demand with its own button.
    _stocking_frontier_section()


if app_mode.startswith("Decide"):
    _decide()
    st.stop()

if app_mode.startswith("Analyze"):      # legacy deep-link, kept working
    _analyze()
    st.stop()

if app_mode.startswith("Accuracy"):
    _forecast_vs_actuals()
    st.stop()

if app_mode.startswith("Compare"):
    _compare_and_choose()
    st.stop()

if app_mode.startswith("Configure"):
    _config_editor()
    st.stop()

if app_mode.startswith("How it works"):
    _how_it_works()
    st.stop()


if app_mode.startswith("Optimize"):
    _optimizer()
    st.stop()


# ---- Run mode: manual override window editor (above the run results) ----
if uploaded is not None:
    _manual_window_editor(uploaded)


# ============================================================
# Run the pipeline when the button is clicked
# ============================================================

if (run_clicked or st.session_state.pop("_pending_run", False)) and uploaded is not None:
    _method = st.session_state.get("_chosen_method", _DEFAULT_METHOD)
    _mobj = _method_obj(_method)
    _spin = (f"Running {_mobj.label} — typically {_TYPICAL.get(_method, '?')}...")
    with st.status(_spin, expanded=True) as status:
        st.write("Config + scenario from the app; ProductionReport from upload...")

        def _narrate(line, _s=status):
            """Show each pipeline stage as it happens: the newest line as the
            status label, the whole sequence in the body below it."""
            _s.update(label=line[:110])
            st.write(line)

        result = _run_with_workbook_bytes(
            uploaded.getvalue(), uploaded.name,
            config_dir=str(CONFIG_DIR), scenario_dir=str(SCENARIO_DIR),
            method=_method, cpsat_time=300.0,
            cpsat_det_time=_cpsat_det_time(),
            cpsat_workers=_cpu_workers(),
            on_line=_narrate,
        )
        if result["ok"]:
            st.write(
                f"✓ Pipeline complete in {result['elapsed']:.1f}s — "
                f"{result['violations']} violations, "
                f"worst {result['worst_density']:.1f} kg/m³"
            )
            status.update(label="✓ Forecast complete", state="complete")
            result["_run_label"] = (
                _mobj.label if _method != "controller"
                else f"Controller — {_harvest_mode_label(CONFIG_DIR)}")
            st.session_state.result = result
        else:
            st.error(f"Pipeline failed: {result.get('error', 'unknown')}")
            if result.get("traceback"):
                st.code(result["traceback"])
            if result.get("stdout"):
                with st.expander("Console output (stdout)", expanded=False):
                    st.code(result["stdout"], language="text")
            status.update(label="✗ Pipeline failed", state="error")
            # Store the failure result so the sidebar can show diagnostics.
            st.session_state.result = result


# ============================================================
# Results view
# ============================================================

if "result" in st.session_state and st.session_state.result.get("ok"):
    r = st.session_state.result
    # Cache key for every derived frame below.
    _rid = _result_rid(r)

    # Provenance — always show WHICH run is on screen (keep the correct data),
    # and WHERE it came from (fresh vs cache-replay, run time, grading rules).
    st.caption(f"📋 Showing: **{r.get('_run_label', 'forecast run')}** · "
               + _provenance_line(r, fresh=_res_is_fresh(r)))
    if r.get("config_used"):
        _render_active_config(r["config_used"],
                              "ℹ️ Configuration this run used")

    # ---- KPIs + prominent download button ----
    top_kpi, top_dl = st.columns([3, 1])
    with top_kpi:
        st.subheader("Summary")
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Violations", r["violations"],
                  help="Tank-weeks where realized density exceeds that tank's "
                       "own cap (per-tank, from facility config). Fish "
                       "PREPARING FOR HARVEST are excluded — judged on stage "
                       "(STARVE, rule R8), so it covers both 6N depuration and "
                       "in-place starvation in an ordinary grow-out tank after "
                       "the 6N production switch. Those tanks have no density "
                       "cap at all: the fish are off feed, not growing, and "
                       "gone within the hold.")
        k2.metric("Worst density", f"{r['worst_density']:.1f} kg/m³",
                  help="Worst density among the tanks COUNTED ABOVE — i.e. the "
                       "deepest breach, not the highest density in the plan. "
                       "Harvest-prep tanks legitimately run far denser and are "
                       "excluded; 0.0 here means no tank breached its cap, not "
                       "that no tank was dense.")
        _wl = r.get("welfare_density", 80)
        k3.metric("Reared density",
                  f"{r.get('mean_rearing_density', 0):.0f} kg/m³",
                  help=f"Product-quality view: the biomass-weighted average density "
                       f"your fish were REARED at over grow-out (lower = gentler = "
                       f"better welfare / flesh quality). The delta is the "
                       f"share of grow-out biomass-WEEKS spent above the "
                       f"{_wl:.0f} kg/m³ welfare line — time-weighted "
                       f"exposure, NOT the share of biomass that was ever "
                       f"crowded.",
                  delta=f"{r.get('crowded_biomass_fraction', 0) * 100:.0f}% crowded",
                  delta_color="inverse")
        if r.get("welfare_density_note"):
            k3.caption(f"⚠ {r['welfare_density_note']}")
        k4.metric("Total harvest", f"{r['harvest_kg']/1000:,.1f} t",
                  help="Sum of all harvest events across the horizon")
        k5.metric("Run time", f"{r['elapsed']:.1f}s")
    with top_dl:
        st.subheader("Output")
        st.download_button(
            label="⬇ Download workbook",
            data=r["output_bytes"],
            file_name=r["output_name"],
            mime="application/vnd.ms-excel.sheet.macroenabled.12",
            use_container_width=True,
            type="primary",
        )
        if r.get("output_path"):
            st.caption(f"Saved to:\n`{r['output_path']}`")

    # ---- Tabs ----
    bl_df = _rv_memo("bl_df", _rid, lambda: (
        pd.DataFrame(r["batch_locations"]) if r["batch_locations"] else pd.DataFrame()))
    he_df = _rv_memo("he_df", _rid, lambda: (
        pd.DataFrame(r["harvest_events"]) if r["harvest_events"] else pd.DataFrame()))
    bio_df = _rv_memo("bio_df", _rid,
                      lambda: pd.DataFrame(r.get("biology_projection", [])))

    tab_over, tab_batch, tab_period, tab_harvest, tab_yearly, tab_plan = st.tabs([
        "Overview",
        "Per-Batch",
        "Period Summary",
        "Harvest",
        "Yearly",
        "Plan",
    ])

    # ============================================================
    # Tab 1: Overview — Advisory + occupancy heatmap
    # ============================================================
    with tab_over:
        st.subheader("Advisory")
        if r["advisory_summary"]:
            col_table, col_detail = st.columns([1, 2])
            with col_table:
                st.caption("Issues by category")
                st.dataframe(
                    pd.DataFrame(r["advisory_summary"]),
                    hide_index=True, use_container_width=True,
                )
            with col_detail:
                st.caption("Details (expand each category)")
                entries = r["advisory_entries"]
                by_cat: dict[str, list] = defaultdict(list)
                for e in entries:
                    by_cat[e["Category"]].append(e["Detail"])
                for cat in sorted(by_cat, key=lambda c: -len(by_cat[c])):
                    with st.expander(f"{cat} ({len(by_cat[cat])})"):
                        for d in by_cat[cat][:50]:
                            st.text(d)
                        if len(by_cat[cat]) > 50:
                            st.caption(f"… and {len(by_cat[cat]) - 50} more")
        else:
            st.info("No advisory entries — clean run.")

        st.subheader("Tank occupancy over time")
        if not bl_df.empty:
            def _build_heatmap():
                """Pivots + hover matrix. Widget-independent, and the hover loop
                is O(tanks x weeks) scalar lookups — the single most expensive
                thing the results view used to redo on every click."""
                df = bl_df.copy()
                df["TankLabel"] = df.apply(
                    lambda r: f"{r['System']}-{r['Tank']}", axis=1)
                tank_order = sorted(
                    df["TankLabel"].unique(),
                    key=lambda t: (t.split("-")[0],
                                   int(t.split("-")[1]) if t.split("-")[1].isdigit() else 0),
                )
                weeks = sorted(df["Week"].dropna().unique())
                density_pivot = df.pivot_table(
                    index="TankLabel", columns="Week",
                    values="Density_kg_m3", aggfunc="first",
                ).reindex(index=tank_order, columns=weeks)
                batch_pivot = df.pivot_table(
                    index="TankLabel", columns="Week",
                    values="Batch", aggfunc="first",
                ).reindex(index=tank_order, columns=weeks)
                _dv = pd.to_numeric(df["Density_kg_m3"], errors="coerce")
                vmax = max(130.0, float(_dv.max()) if _dv.notna().any() else 130.0)
                customdata = []
                for tl in tank_order:
                    row_cd = []
                    for wk in weeks:
                        bid = batch_pivot.loc[tl, wk] if wk in batch_pivot.columns else None
                        d = density_pivot.loc[tl, wk] if wk in density_pivot.columns else None
                        row_cd.append([str(bid) if bid else "—",
                                       f"{d:.1f}" if isinstance(d, (int, float)) else "—"])
                    customdata.append(row_cd)
                return density_pivot, batch_pivot, tank_order, weeks, vmax, customdata

            (density_pivot, batch_pivot, tank_order, weeks, vmax,
             customdata) = _rv_memo("ov_heatmap", _rid, _build_heatmap)
            # Severity-honest scale: span the TRUE worst density (clamping at 130
            # hid 3.8x spikes as ordinary red). Hard color break at the cap;
            # over-cap tanks deepen toward dark red as they get worse.
            CAP = float(r.get("growout_density_cap") or 95.0)
            c_lo, c_cap = 80.0 / vmax, CAP / vmax
            colorscale = sorted({
                0.0: "#f0f0f0",          # empty / ~zero
                min(c_lo, c_cap * 0.5): "#a8d5a8",   # green, under target
                max(0.0, c_cap - 1e-3): "#f5d49a",   # amber, just under cap
                c_cap: "#e8615e",        # red AT the cap
                1.0: "#7a0d0b",          # dark red at the worst observed
            }.items())
            # 6N rows are depuration/purge — empty cells there are NOT free
            # growout capacity. Mark them so they aren't read as availability.
            disp_y = [f"{t}  ⛔purge" if t.startswith("OG6N") else t
                      for t in tank_order]
            fig = px.imshow(
                density_pivot.values,
                x=weeks, y=disp_y,
                color_continuous_scale=[(p, c) for p, c in colorscale],
                range_color=[0, vmax],
                labels=dict(x="Week", y="Tank", color="Density (kg/m³)"),
                aspect="auto",
            )
            fig.update_traces(
                customdata=customdata,
                hovertemplate=(
                    "Tank: %{y}<br>Week: %{x}<br>Batch: %{customdata[0]}"
                    "<br>Density: %{customdata[1]} kg/m³<extra></extra>"
                ),
            )
            fig.update_layout(height=600, margin=dict(l=80, r=20, t=20, b=40))
            st.plotly_chart(fig, use_container_width=True)
            st.caption(
                f"Color = per-tank density. Green = under target, amber = "
                f"approaching the {CAP:.0f} kg/m³ cap, red = over cap, deepening to "
                f"dark red at the worst observed ({vmax:.0f} kg/m³) — so a 3.8x "
                f"spike no longer looks like a mild one. Rows tagged ⛔purge are "
                f"OG6N depuration tanks: empty cells there are NOT free growout "
                f"capacity. Hover any cell for the batch and exact density."
            )

            # ---- Per-system biomass + feed over time ----
            st.subheader("Per-system biomass + feed")
            sys_bio = (
                bl_df.assign(Biomass_kg=bl_df["Biomass_kg"].fillna(0))
                .groupby(["System", "Week"]).agg(
                    Biomass_kg=("Biomass_kg", "sum"),
                ).reset_index().sort_values(["Week", "System"])
            )

            c1, c2 = st.columns(2)
            with c1:
                fig = px.line(
                    sys_bio, x="Week", y="Biomass_kg", color="System",
                    markers=True,
                    title="Per-system biomass (kg) over time",
                )
                fig.update_layout(height=380, yaxis_title="kg",
                                  legend=dict(title="System"))
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                # Per-system biomass as % of typical density-cap capacity
                # (avg tank volume × density cap × tank count per system).
                # Approximation: pull tank count per system from data.
                # Exact per-system density-cap capacity from facility config:
                # Σ(tank volume × tank max_density), not a flat 95×1720 guess.
                sys_cap_biomass = r.get("system_biomass_cap") or {}
                sys_bio_pct = sys_bio.copy()
                sys_bio_pct["Cap_kg"] = sys_bio_pct["System"].map(
                    lambda s: sys_cap_biomass.get(s, 0.0)
                )
                sys_bio_pct["Pct_of_cap"] = (
                    sys_bio_pct["Biomass_kg"] / sys_bio_pct["Cap_kg"] * 100
                ).where(sys_bio_pct["Cap_kg"] > 0, 0)
                fig = px.line(
                    sys_bio_pct, x="Week", y="Pct_of_cap", color="System",
                    markers=True,
                    title="Per-system utilization (% of density-cap capacity)",
                )
                fig.add_hline(y=100, line_dash="dash", line_color="red",
                              annotation_text="100% (cap)")
                fig.add_hline(y=85, line_dash="dot", line_color="orange",
                              annotation_text="85% target")
                fig.update_layout(height=380, yaxis_title="% of cap",
                                  legend=dict(title="System"))
                st.plotly_chart(fig, use_container_width=True)

            st.caption(
                "Left: absolute biomass per system, summed across that "
                "system's tanks. Right: same as % of the system's "
                "density-cap capacity (Σ tank volume × tank max_density, "
                "from facility config). "
                "Watch for systems pinned at 100% while others sit idle "
                "— that's the operational signal of imbalance."
            )

            # Per-system feed/day — REALIZED, read straight from the
            # SystemLimitsAudit sheet (the per-(system, week) feed the caps are
            # actually checked against). NOT BiologyProjection.Feed_kg_day, which
            # is the raw UNHARVESTED projection (fish growing along the curve,
            # ignoring harvest + caps) and spikes far above any realized
            # per-system feed (e.g. >10,000 vs a realized ~3,980 / 3,000 cap) —
            # the same projection-vs-realized gap fixed at the report layer in
            # 63dd6cc. This chart must mirror the realized plan, so it reads the
            # audit directly (also exact, not a biomass-share approximation).
            sys_feed_rows, feed_caps = _rv_memo(
                "sysfeed", _rid, lambda: _system_feed_audit(r.get("output_path")))
            if sys_feed_rows:
                sys_feed = pd.DataFrame(sys_feed_rows).sort_values(["Week", "System"])
                fig = px.line(
                    sys_feed, x="Week", y="Feed_kg_day", color="System",
                    markers=True,
                    title="Per-system feed (kg/day) over time — REALIZED",
                )
                # One dashed line per DISTINCT cap, labelled with the systems it
                # applies to. Systems sharing a cap collapse to a single line so
                # the chart is not striped with duplicates.
                _by_cap = {}
                for _s, _c in (feed_caps or {}).items():
                    _by_cap.setdefault(round(float(_c), 6), []).append(_s)
                for _c, _syss in sorted(_by_cap.items()):
                    _lbl = (", ".join(sorted(_syss)) if len(_syss) <= 3
                            else f"{len(_syss)} systems")
                    fig.add_hline(y=_c, line_dash="dash", line_color="red",
                                  opacity=0.55,
                                  annotation_text=f"{_lbl}: {_c:,.0f} kg/day cap",
                                  annotation_font_size=10)
                fig.update_layout(height=380, yaxis_title="kg/day",
                                  legend=dict(title="System"))
                st.plotly_chart(fig, use_container_width=True)
                st.caption(
                    "REALIZED per-system feed from SystemLimitsAudit — the actual "
                    "fed plan (after harvest + FIFO), the exact series the feed "
                    "caps are checked against. NOT the unharvested biology "
                    "projection (which ignores harvest and spikes well past the "
                    "cap). Each dashed line is the cap for the system(s) named "
                    "on it; as shipped EVERY OG system carries the same "
                    "3,000 kg/day cap, so they collapse to one line. Lines riding "
                    "just under their own cap = leveled correctly; brief "
                    "crossings are the residual over-cap weeks."
                )

    # ============================================================
    # Tab 2: Per-Batch lifecycle
    # ============================================================
    with tab_batch:
        st.subheader("Per-batch trajectories")
        if bl_df.empty:
            st.info("No batch data to display.")
        else:
            # Aggregate per (Batch, Week): sum count + biomass.
            def _build_batch_agg():
                df = bl_df.copy()
                df["Biomass_kg"] = df["Biomass_kg"].fillna(0)
                df["Count"] = df["Count"].fillna(0)
                agg = df.groupby(["Batch", "Week"]).agg(
                    Count=("Count", "sum"),
                    Biomass_kg=("Biomass_kg", "sum"),
                    MaxDensity=("Density_kg_m3", "max"),
                    MeanDensity=("Density_kg_m3", "mean"),
                    Tanks=("Tank", "nunique"),
                ).reset_index()
                agg["AvgWt_kg"] = (agg["Biomass_kg"] / agg["Count"]).where(
                    agg["Count"] > 0, 0)
                return agg

            agg = _rv_memo("pb_agg", _rid, _build_batch_agg)
            batches = sorted(agg["Batch"].dropna().unique())
            default = ["B46", "B47"] if all(b in batches for b in ("B46", "B47")) else batches[:2]
            all_weeks = sorted(agg["Week"].dropna().unique())
            # Keyed so the selection survives reruns triggered elsewhere in the
            # app; dropped when a new run's options no longer contain it (a
            # select_slider raises on a value outside its options).
            if ("pb_batches" in st.session_state
                    and not set(st.session_state["pb_batches"]) <= set(batches)):
                del st.session_state["pb_batches"]
            if ("pb_period" in st.session_state
                    and not all(w in all_weeks for w in st.session_state["pb_period"])):
                del st.session_state["pb_period"]

            ctrl_l, ctrl_r = st.columns([2, 3])
            with ctrl_l:
                picked = st.multiselect(
                    "Batches", batches, default=default, key="pb_batches",
                    help="Pick one or more batches to compare trajectories.",
                )
            with ctrl_r:
                if len(all_weeks) >= 2:
                    wk_lo, wk_hi = st.select_slider(
                        "Period",
                        options=all_weeks,
                        value=(all_weeks[0], all_weeks[-1]),
                        key="pb_period",
                        help="Slide endpoints to zoom in on a specific window.",
                    )
                else:
                    # NB: parenthesised so the empty case yields BOTH values —
                    # unparenthesised, `all_weeks[0]` ran before the guard.
                    wk_lo, wk_hi = ((all_weeks[0], all_weeks[-1]) if all_weeks
                                    else (None, None))

            if not picked:
                st.info("Select at least one batch.")
            else:
                view = agg[
                    (agg["Batch"].isin(picked))
                    & (agg["Week"] >= wk_lo)
                    & (agg["Week"] <= wk_hi)
                ].sort_values(["Batch", "Week"])
                c1, c2 = st.columns(2)
                with c1:
                    fig = px.line(
                        view, x="Week", y="AvgWt_kg", color="Batch",
                        markers=True, title="Average weight per fish (kg)",
                    )
                    fig.update_layout(height=350, yaxis_title="kg/fish")
                    st.plotly_chart(fig, use_container_width=True)
                with c2:
                    fig = px.line(
                        view, x="Week", y="Biomass_kg", color="Batch",
                        markers=True, title="Total batch biomass (kg)",
                    )
                    fig.update_layout(height=350, yaxis_title="kg")
                    st.plotly_chart(fig, use_container_width=True)
                c3, c4 = st.columns(2)
                with c3:
                    fig = px.line(
                        view, x="Week", y="MaxDensity", color="Batch",
                        markers=True, title="Max per-tank density (kg/m³)",
                    )
                    _dcap = float(r.get("growout_density_cap") or 95.0)
                    fig.add_hline(y=_dcap, line_dash="dash", line_color="red",
                                  annotation_text="cap")
                    # R31 `density_target_pct` is what the planner actually
                    # aims at (precalc sizing, placement sizing, the Phase D
                    # grade trigger). Read it from the config rather than
                    # hard-coding 85%, which was never this target.
                    _dt = 0.90
                    try:
                        from forecast.config_io import load_control as _lc
                        _dt = float(getattr(_lc(CONFIG_DIR),
                                            "density_target_pct", 0.90) or 0.90)
                    except Exception:  # noqa: BLE001
                        pass
                    fig.add_hline(y=_dcap * _dt, line_dash="dot",
                                  line_color="orange",
                                  annotation_text=f"{_dt * 100:.0f}% target")
                    fig.update_layout(height=350, yaxis_title="kg/m³")
                    st.plotly_chart(fig, use_container_width=True)
                with c4:
                    fig = px.line(
                        view, x="Week", y="Count", color="Batch",
                        markers=True, title="Fish count (end of week)",
                    )
                    fig.update_layout(height=350, yaxis_title="fish")
                    st.plotly_chart(fig, use_container_width=True)

                # Explicit weekly losses: mortality + cull events per
                # week, plotted at their own scale so they don't disappear
                # next to the much larger total fish count.
                if not bio_df.empty:
                    bv = bio_df[
                        (bio_df["Batch"].isin(picked))
                        & (bio_df["Week"] >= wk_lo)
                        & (bio_df["Week"] <= wk_hi)
                    ].copy()
                    bv["Mortality + Cull (fish)"] = (
                        bv["Mortality_count_wk"] + bv["Cull_count_wk"]
                    )
                    fig = px.line(
                        bv, x="Week", y="Mortality + Cull (fish)",
                        color="Batch", markers=True,
                        title="Weekly losses (mortality + scheduled culls)",
                    )
                    fig.update_layout(height=300, yaxis_title="fish lost / week")
                    st.plotly_chart(fig, use_container_width=True)
                    st.caption(
                        "Mortality from the per-week mortality table, plus "
                        "any cull events at scheduled DSI thresholds. Plotted "
                        "separately because per-week losses (~50–200 fish) "
                        "are tiny next to the 200k+ batch total — they "
                        "disappear in the Count chart above."
                    )

                with st.expander("Raw weekly table"):
                    st.dataframe(view, hide_index=True, use_container_width=True)

    # ============================================================
    # Tab 3: Period Summary — facility-wide weekly metrics
    # ============================================================
    with tab_period:
        st.subheader("Facility-wide weekly summary")
        if bl_df.empty:
            st.info("No batch data to display.")
        else:
            def _build_period():
                df = bl_df.copy()
                df["Biomass_kg"] = df["Biomass_kg"].fillna(0)
                wk_facility = df.groupby("Week").agg(
                    FacilityBiomass_kg=("Biomass_kg", "sum"),
                    ActiveTanks=("Tank", "nunique"),
                    ActiveBatches=("Batch", "nunique"),
                    MeanDensity=("Density_kg_m3", "mean"),
                ).reset_index().sort_values("Week")
                # Merge harvest per week (kg + count)
                if not he_df.empty:
                    hw = he_df.groupby("Week").agg(
                        HarvestKg=("Gross_kg", "sum"),
                        HarvestCount=("Count", "sum"),
                    ).reset_index()
                    wk_facility = wk_facility.merge(hw, on="Week", how="left")
                    wk_facility[["HarvestKg", "HarvestCount"]] = (
                        wk_facility[["HarvestKg", "HarvestCount"]].fillna(0)
                    )
                return wk_facility

            wk_facility = _rv_memo("period", _rid, _build_period)

            c1, c2 = st.columns(2)
            with c1:
                # WHOLE-FACILITY basis (OG + 6N + freshwater) — what the cap
                # actually governs. The tank-row sum below omits the FW phase
                # (~7% of cap at peak), so plotting it against this cap line
                # showed headroom that did not exist.
                _fwk = pd.DataFrame(r.get("facility_weekly") or [])
                _has_fac = (not _fwk.empty
                            and _fwk["Total_Biomass_kg"].notna().any())
                if _has_fac:
                    _fwk = _fwk.dropna(subset=["Total_Biomass_kg"]).sort_values("Week")
                    fig = px.line(_fwk, x="Week", y="Total_Biomass_kg",
                                  markers=True,
                                  title="Facility biomass (kg) — whole facility, "
                                        "incl. freshwater")
                    _lim = _fwk["Biomass_Limit_kg"]
                    if _lim.dropna().nunique() > 1:
                        # Per-week cap (FacilityLimits ramp) — one hline at
                        # week 1's value would misread the whole horizon.
                        fig.add_scatter(x=_fwk["Week"], y=_lim, mode="lines",
                                        line={"dash": "dash", "color": "red"},
                                        name="Max Biomass cap (per-week)")
                        _cap_kg = None
                    else:
                        _cap_kg = (float(_lim.dropna().iloc[0])
                                   if _lim.notna().any() else
                                   float((r.get("config_used") or {})
                                         .get("max_biomass_kg") or 3_800_000))
                else:
                    fig = px.line(wk_facility, x="Week", y="FacilityBiomass_kg",
                                  markers=True,
                                  title="Facility biomass (kg) — TANKS ONLY "
                                        "(freshwater not included)")
                    _cap_kg = float((r.get("config_used") or {})
                                    .get("max_biomass_kg") or 3_800_000)
                if _cap_kg is not None:
                    fig.add_hline(
                        y=_cap_kg, line_dash="dash", line_color="red",
                        annotation_text=f"Max Biomass cap ({_cap_kg / 1000:,.0f} t)")
                fig.update_layout(height=350, yaxis_title="kg")
                st.plotly_chart(fig, use_container_width=True)
                if not _has_fac:
                    st.caption("⚠ Advisory sheet not found in this workbook — "
                               "showing tank biomass only, which sits below the "
                               "cap line by the freshwater share (~7% at peak).")
            with c2:
                if "HarvestKg" in wk_facility.columns:
                    fig = px.bar(
                        wk_facility, x="Week", y="HarvestKg",
                        title="Weekly harvest (kg, gross)",
                    )
                    fig.update_layout(height=350, yaxis_title="kg")
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No harvest events.")

            c3, c4 = st.columns(2)
            with c3:
                fig = px.line(
                    wk_facility, x="Week", y="ActiveBatches",
                    markers=True, title="Active batches per week",
                )
                fig.update_layout(height=300, yaxis_title="count")
                st.plotly_chart(fig, use_container_width=True)
            with c4:
                fig = px.line(
                    wk_facility, x="Week", y="MeanDensity",
                    markers=True,
                    title="Mean per-tank density across facility (kg/m³)",
                )
                fig.add_hline(y=float(r.get("growout_density_cap") or 95.0),
                              line_dash="dash", line_color="red",
                              annotation_text="cap")
                fig.update_layout(height=300, yaxis_title="kg/m³")
                st.plotly_chart(fig, use_container_width=True)

            with st.expander("Raw weekly table"):
                st.dataframe(wk_facility, hide_index=True, use_container_width=True)

    # ============================================================
    # Tab 4: Harvest Overview
    # ============================================================
    with tab_harvest:
        st.subheader("Harvest plan overview")
        if he_df.empty:
            st.info("No harvest events.")
        else:
            tot_kg = he_df["Gross_kg"].sum()
            tot_count = he_df["Count"].sum()
            avg_kg = tot_kg / tot_count if tot_count else 0
            k1, k2, k3 = st.columns(3)
            k1.metric("Total harvest", f"{tot_kg/1000:,.1f} t",
                      help="Total gross (live) biomass harvested across the whole "
                           "forecast horizon.")
            k2.metric("Total fish", f"{tot_count:,.0f}",
                      help="Total number of fish harvested across the horizon.")
            k3.metric("Avg weight at harvest", f"{avg_kg:.2f} kg",
                      help="Harvest-weighted average LIVE weight per fish "
                           "(total kg ÷ total fish).")

            c1, c2 = st.columns(2)
            with c1:
                fig = px.bar(
                    he_df, x="Week", y="Gross_kg", color="Batch",
                    title="Harvest kg per week, stacked by batch",
                )
                fig.update_layout(height=400, yaxis_title="kg")
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig = px.bar(
                    he_df, x="Week", y="Count", color="Batch",
                    title="Harvest count per week, stacked by batch",
                )
                fig.update_layout(height=400, yaxis_title="fish")
                st.plotly_chart(fig, use_container_width=True)

            # Avg harvest weight per week (weighted)
            hw = he_df.groupby("Week").agg(
                Count=("Count", "sum"),
                Gross_kg=("Gross_kg", "sum"),
            ).reset_index()
            hw["AvgWt_kg"] = (hw["Gross_kg"] / hw["Count"]).where(hw["Count"] > 0, 0)
            fig = px.line(
                hw, x="Week", y="AvgWt_kg", markers=True,
                title="Average harvest weight per week (kg/fish)",
            )
            _minw = float((r.get("config_used") or {}).get("min_harvest_weight_g")
                          or 3500) / 1000.0
            fig.add_hline(y=_minw, line_dash="dot", line_color="orange",
                          annotation_text=f"Min harvest weight ({_minw:.1f} kg)")
            fig.update_layout(height=300, yaxis_title="kg/fish")
            st.plotly_chart(fig, use_container_width=True)

            # Monthly rollup (sales planning): HOG tonnes + count per calendar
            # month, derived from each event's ISO week.
            import datetime as _dt

            def _wk_to_month(wk):
                try:
                    y, w = int(str(wk)[:4]), int(str(wk)[6:8])
                    return _dt.date.fromisocalendar(y, w, 1).strftime("%Y-%m")
                except Exception:
                    return None

            hm = he_df.copy()
            hm["Month"] = hm["Week"].map(_wk_to_month)
            hm = hm.dropna(subset=["Month"])
            if not hm.empty:
                mo = hm.groupby("Month").agg(
                    HOG_t=("HOG_kg", lambda s: s.sum() / 1000.0),
                    Count=("Count", "sum"),
                ).reset_index()
                st.markdown("**Monthly harvest (sales planning)**")

                # CONFIDENCE BAND on the tonnage, from measured history. The
                # whiskers are not a model assumption: they are where past
                # forecasts of this facility actually landed at the same
                # horizon (tools/backtest.py -> tools/error_model.py). Months
                # with no sound measurement get NO whisker rather than a
                # made-up one — see forecast/error_bands.py.
                _emodel = _hv_error_model()
                _bands = _hv_month_bands(mo["Month"].tolist(),
                                         mo["HOG_t"].tolist(), _emodel)
                st.caption(_hv_band_caption(_emodel, _bands))

                cc1, cc2 = st.columns(2)
                with cc1:
                    fig = px.bar(mo, x="Month", y="HOG_t", text_auto=".0f",
                                 title="HOG tonnes harvested per month")
                    if any(b is not None for b in _bands):
                        fig.update_traces(error_y=dict(
                            type="data", symmetric=False,
                            array=[(b[1] - v) if b else 0.0
                                   for b, v in zip(_bands, mo["HOG_t"])],
                            arrayminus=[(v - b[0]) if b else 0.0
                                        for b, v in zip(_bands, mo["HOG_t"])],
                            color="#444", thickness=1.4, width=6))
                    fig.update_layout(height=320, yaxis_title="t HOG", xaxis_title="")
                    st.plotly_chart(fig, use_container_width=True)
                with cc2:
                    fig = px.bar(mo, x="Month", y="Count", text_auto=".2s",
                                 title="Fish harvested per month")
                    fig.update_layout(height=320, yaxis_title="fish", xaxis_title="")
                    st.plotly_chart(fig, use_container_width=True)

            # Daily harvest schedule — each weekly tank-harvest split Mon–Fri,
            # with a per-week Total row and a blank line between weeks.
            st.markdown("**Daily harvest schedule (Mon–Fri)**")
            st.caption(
                "**All tanks harvesting in a week are combined**, then split "
                "evenly across the five operating days (Mon–Fri), with a **Total** "
                "row per week and a blank line between weeks. Tank/Batch list every "
                "tank + batch that contributed; average weights are blended (total "
                "biomass ÷ total fish). Same as the Excel 'Daily Harvest Schedule' "
                "sheet.")
            _dh, _totpos, _blankpos = _rv_memo(
                "harvest_daily", _rid, lambda: _daily_harvest_table(he_df))
            if _dh.empty:
                st.info("No datable harvest events for a daily breakout.")
            else:
                def _hl_totals(row):
                    if row.name in _totpos:
                        return ["background-color:#e8eaf0;font-weight:700"] * len(row)
                    return [""] * len(row)
                st.dataframe(_dh.style.apply(_hl_totals, axis=1),
                             hide_index=True, use_container_width=True,
                             height=min(760, 44 + 35 * len(_dh)))

            with st.expander("Raw harvest events"):
                st.dataframe(he_df, hide_index=True, use_container_width=True)

    # ============================================================
    # Tab 5: Yearly — facility-wide per-year trends
    # ============================================================
    with tab_yearly:
        st.subheader("Yearly trends (facility-wide)")
        yr = pd.DataFrame(r.get("yearly", []))
        if yr.empty:
            st.info("No yearly summary in this workbook.")
        else:
            st.caption("Per-calendar-year rollup. Partial first/last years reflect "
                       "the forecast horizon, not full calendar years.")
            st.dataframe(yr, hide_index=True, use_container_width=True)
            yr["Year"] = yr["Year"].astype(str)
            charts = [
                ("Harvest_HOG (t)", "Harvest (HOG tonnes) per year"),
                ("Feed (t)", "Feed (tonnes) per year"),
                ("Peak_Biomass (t)", "Peak facility biomass (tonnes) per year"),
                ("Harvest_Count (fish)", "Harvest count (fish) per year"),
            ]
            cols = st.columns(2)
            for i, (col, title) in enumerate(charts):
                if col not in yr.columns:
                    continue
                fig = px.bar(yr, x="Year", y=col, title=title, text_auto=True)
                fig.update_layout(height=320, xaxis_title="")
                cols[i % 2].plotly_chart(fig, use_container_width=True)

    # ============================================================
    # Tab 6: Plan — per-batch plan summary + density risk
    # ============================================================
    with tab_plan:
        # Production-flow template (TransferTemplate Section A) — the canonical
        # journey every batch follows through the seawater conveyor.
        ft = pd.DataFrame(r.get("flow_template", []))
        if not ft.empty:
            st.subheader("Production flow — the canonical batch journey")
            st.caption(
                "Every batch follows this seawater journey; only the timing "
                "shifts with stocking date, the shape is fixed: FW → OG1/2 "
                "nursery → the 1 kg move-lock → grow-out fan-out across systems "
                "→ finishing/depuration in the top systems → harvest drain."
            )
            st.dataframe(ft, hide_index=True, use_container_width=True)
            st.divider()

        st.subheader("Batch plan summary")
        pf = pd.DataFrame(r.get("plan_summary", []))
        if pf.empty:
            st.info("No TransferTemplate plan summary in this workbook.")
        else:
            st.caption("When each batch enters seawater, its tank footprint, density "
                       "risk, and harvest window. Rows flagged OVER CAP peak above the "
                       "density cap at some point in their life.")
            status_col = next((c for c in pf.columns if c.startswith("Density_Status")), None)
            peak_col = next((c for c in pf.columns if c.startswith("Peak_Density")), None)
            n_over = int((pf[status_col] == "OVER CAP").sum()) if status_col else 0
            c1, c2, c3 = st.columns(3)
            c1.metric("Batches", len(pf),
                      help="Number of production batches in this plan.")
            c2.metric("Density risk (OVER CAP)", n_over,
                      help="How many batches peak ABOVE the density cap at some "
                           "point in grow-out — the cohorts to watch for crowding.")
            if peak_col:
                worst = pd.to_numeric(pf[peak_col], errors="coerce").max()
                c3.metric("Worst peak density", f"{worst:.2f}× cap" if pd.notna(worst) else "—",
                          help="The single highest peak density any batch reaches, "
                               "as a multiple of its tank's cap (1.0× = right at "
                               "cap, 1.3× = 30% over).")

            def _hl(row):
                over = status_col and row.get(status_col) == "OVER CAP"
                return ['background-color: #fde2e2' if over else '' for _ in row]
            st.dataframe(pf.style.apply(_hl, axis=1), hide_index=True, use_container_width=True)

            if peak_col and status_col:
                pf2 = pf.copy()
                pf2["_peak"] = pd.to_numeric(pf2[peak_col], errors="coerce")
                pf2 = pf2.dropna(subset=["_peak"])
                if not pf2.empty:
                    fig = px.bar(pf2, x="Batch", y="_peak", color=status_col,
                                 color_discrete_map={"OVER CAP": "#e45756", "OK": "#54a24b"},
                                 title="Peak density per batch (× cap)")
                    fig.add_hline(y=1.0, line_dash="dot", line_color="red",
                                  annotation_text="density cap")
                    fig.update_layout(height=360, yaxis_title="× cap", xaxis_title="")
                    st.plotly_chart(fig, use_container_width=True)

        # ---- Per-batch plan: where each batch is + how it got there ----
        st.divider()
        st.subheader("Per-batch plan — journey + milestones")
        bplans = _rv_memo("bplans", _rid,
                          lambda: _derive_batch_plans(bl_df, he_df))
        if not bplans:
            st.info("No batch-location data to build per-batch plans.")
        else:
            st.caption("Each batch's planned journey through the conveyor — a summary "
                       "header (entry, peak tanks, harvest window, HOG) plus the "
                       "milestone timeline (when it enters each tier, at what weight, "
                       "through to harvest). Pick a batch to review; download all to share.")
            hdr_df = pd.DataFrame([{k: p[k] for k in
                                    ("Batch", "SW_entry", "Peak_tanks", "Harvest_window", "HOG_t")}
                                   for p in bplans])
            st.dataframe(hdr_df, hide_index=True, use_container_width=True)
            # A previous run's pick may not exist in this run's plans (different
            # PR, or a Global-LP plan without per-tank rows) — Streamlit passes
            # the stale value through verbatim, so guard it or next() raises.
            _bp_opts = [p["Batch"] for p in bplans]
            if st.session_state.get("batchplan_pick") not in _bp_opts:
                st.session_state.pop("batchplan_pick", None)
            pick = st.selectbox("Batch", _bp_opts, key="batchplan_pick")
            bp = next((p for p in bplans if p["Batch"] == pick), bplans[0])
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("SW entry", bp["SW_entry"],
                      help="The FIRST week this batch appears in seawater — "
                           "its FW→OG / TranOG transfer, or the forecast "
                           "start week for a batch already in seawater "
                           "(marked 'In-flight at forecast start' below).")
            m2.metric("Peak tanks", bp["Peak_tanks"],
                      help="The most grow-out tanks this batch occupies at once — "
                           "its peak facility footprint.")
            m3.metric("Harvest window", bp["Harvest_window"],
                      help="The span of weeks over which this batch is harvested "
                           "out.")
            m4.metric("HOG (t)", f"{bp['HOG_t']:.0f}",
                      help="Total HOG (head-off, gutted) tonnes this batch yields over its "
                           "harvest window.")
            st.dataframe(pd.DataFrame(bp["milestones"]), hide_index=True,
                         use_container_width=True)
            # Flat export (one row per batch-milestone) for sharing/review.
            _csv = _rv_memo("bplans_csv", _rid, lambda: pd.DataFrame(
                [{"Batch": p["Batch"], **m} for p in bplans for m in p["milestones"]]
            ).to_csv(index=False).encode())
            st.download_button(
                "⬇ Download all batch plans (CSV)",
                data=_csv,
                file_name="batch_plans.csv", mime="text/csv")

    # ---- Transfer plan — the week-by-week move list. Parsed LAZILY from
    # the output workbook (every engine writes a TransferPlan sheet), so it
    # also works for results picked/restored before this view existed, and
    # shows even when per-batch plans are unavailable (Global LP).
    with tab_plan:
        st.divider()
        st.subheader("🚚 Transfer plan — every planned move, week by week")
        st.caption(
            "One **Transfer** row = one real tank-to-tank move (same-week "
            "duplicate legs are merged), and those are the rows the weekly "
            "handling budget counts. The other types are not crew moves in "
            "the same sense: **TranOG** writes one row per destination tank "
            "of a seawater arrival (From_Tank reads `FW`), and **Grade** "
            "writes a pickup row plus a retention row for one size-sort. "
            "Filter by Type to see just the tank-to-tank moves.")
        tp_df = _rv_memo("tp_df", _rid, lambda: pd.DataFrame(
            _transfer_plan_rows(r["output_path"])
            if r.get("output_path") else []))
        if tp_df.empty:
            st.info("No transfers in this plan (or the output workbook is no "
                    "longer on disk — re-run to regenerate it).")
        else:
            _wk_col = tp_df.columns[0]
            f1, f2, f3 = st.columns(3)
            _wks = f1.multiselect(
                "Week(s)", sorted(tp_df[_wk_col].astype(str).unique()),
                default=[], key="tp_weeks",
                help="Empty = all weeks.")
            _typ_col = next((c for c in tp_df.columns
                             if str(c).startswith("Type")), None)
            _typs = (f2.multiselect("Type(s)",
                                    sorted(tp_df[_typ_col].astype(str).unique()),
                                    default=[], key="tp_types")
                     if _typ_col else [])
            _bat_col = next((c for c in tp_df.columns
                             if str(c).startswith("Batch")), None)
            _bats = (f3.multiselect("Batch(es)",
                                    sorted(tp_df[_bat_col].astype(str).unique()),
                                    default=[], key="tp_batches")
                     if _bat_col else [])
            view = tp_df
            if _wks:
                view = view[view[_wk_col].astype(str).isin(_wks)]
            if _typs and _typ_col:
                view = view[view[_typ_col].astype(str).isin(_typs)]
            if _bats and _bat_col:
                view = view[view[_bat_col].astype(str).isin(_bats)]
            c1, c2 = st.columns(2)
            c1.metric("Rows shown", len(view),
                      help=f"{len(tp_df)} rows in the plan. This counts "
                           f"ROWS, not crew moves — filter Type to "
                           f"'Transfer' for the tank-to-tank moves the "
                           f"handling budget counts.")
            c2.metric("Weeks with rows",
                      view[_wk_col].astype(str).nunique())
            st.dataframe(view, hide_index=True, use_container_width=True,
                         height=min(520, 46 + 35 * min(len(view), 13)))
            st.download_button(
                "⬇ Download transfer plan (CSV)",
                data=_rv_memo("tp_csv", _rid,
                              lambda: tp_df.to_csv(index=False).encode()),
                file_name="transfer_plan.csv", mime="text/csv")

    # ---- Run log (collapsed) ----
    with st.expander("Run log (console output)"):
        st.code(r["stdout"], language="text")
else:
    st.info(
        "Upload a workbook in the sidebar and click ▶ Run forecast to "
        "begin. The input workbook is never modified — output is written "
        "to a new file you can download."
    )
